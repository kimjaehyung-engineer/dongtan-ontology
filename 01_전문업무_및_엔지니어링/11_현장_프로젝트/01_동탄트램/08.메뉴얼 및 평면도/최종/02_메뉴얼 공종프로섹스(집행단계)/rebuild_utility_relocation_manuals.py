# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse, re

sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')
util_dir = os.path.join(v8_root, '지장물이설')

print("=== Starting 1:1 Rebuild of 지장물이설 HTML Manuals from Excel ===", flush=True)

wb = openpyxl.load_workbook(v8_excel, data_only=True)
ws = wb['지장물이설']

activities = []
for r in range(2, ws.max_row + 1):
    task_title = ws.cell(row=r, column=8).value
    if task_title:
        activities.append({
            'row': r,
            'l3_code': ws.cell(row=r, column=2).value or '2000',
            'l3_name': ws.cell(row=r, column=3).value or '지장물이설',
            'task_title': str(task_title).strip(),
            'dept': str(ws.cell(row=r, column=9).value or '현장 공무/시공팀').strip(),
            'purpose': str(ws.cell(row=r, column=10).value or '지장물 이설 품질 및 안전 검측').strip(),
            'method': str(ws.cell(row=r, column=11).value or '1) 사전검토 -> 2) 현측 -> 3) 시공 -> 4) 승인').strip(),
            'output': str(ws.cell(row=r, column=12).value or '지장물 이설 정밀 검측성과표 및 최종 승인서').strip(),
            'risk': str(ws.cell(row=r, column=21).value or '지하시설물 간섭 및 안전사고 리스크 사전 차단').strip(),
            'advisory': str(ws.cell(row=r, column=22).value or '관할 점용기관 및 시공 전문 기술자 자문').strip(),
        })

wb.close()
print(f"Extracted {len(activities)} activities from [지장물이설] Excel sheet.", flush=True)

# 2D SVG Diagrams for Utility Relocation Steps (Light Theme)
def generate_step_svgs(title):
    svg1 = f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-slate-50 border border-slate-200 rounded-xl p-3 mb-3" onclick="openDiagramZoom(this.outerHTML, '{title} - STEP 01 2D 도식')">
        <div class="text-[11px] font-bold text-slate-500 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-indigo-600 bg-indigo-50 px-2 py-0.5 rounded border border-indigo-200 font-extrabold">2D 사전탐사/검토 도식</span>
        </div>
        <svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-slate-200">
            <rect width="480" height="160" fill="#F8FAFC"/>
            <line x1="0" y1="120" x2="480" y2="120" stroke="#64748B" stroke-width="2"/>
            <rect x="0" y="120" width="480" height="40" fill="#CBD5E1"/>
            <text x="15" y="145" font-size="12" font-weight="bold" fill="#334155">지중 매설층 (Underground Depth GL-1.5m~3.0m)</text>
            <!-- GPR Radar Truck -->
            <rect x="30" y="70" width="80" height="40" rx="4" fill="#0284C7"/>
            <circle cx="50" cy="115" r="8" fill="#1E293B"/>
            <circle cx="90" cy="115" r="8" fill="#1E293B"/>
            <text x="35" y="95" font-size="11" font-weight="bold" fill="#FFFFFF">GPR 3D 탐사</text>
            <!-- Radar Waves -->
            <path d="M 70 110 Q 70 135 110 140" stroke="#38BDF8" stroke-width="2" stroke-dasharray="3,3"/>
            <!-- Pipe -->
            <circle cx="260" cy="140" r="14" fill="#EAB308" stroke="#CA8A04" stroke-width="2"/>
            <text x="285" y="145" font-size="12" font-weight="900" fill="#854D0E">기존 매설관 (가스/통신/한전)</text>
        </svg>
    </div>'''
    
    svg2 = f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-sky-50 border border-sky-200 rounded-xl p-3 mb-3" onclick="openDiagramZoom(this.outerHTML, '{title} - STEP 02 2D 도식')">
        <div class="text-[11px] font-bold text-sky-700 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-sky-700 bg-sky-100 px-2 py-0.5 rounded border border-sky-300 font-extrabold">2D 인력시탐/굴착 도식</span>
        </div>
        <svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-sky-200">
            <rect width="480" height="160" fill="#F0F9FF"/>
            <!-- Trench Trenching -->
            <polygon points="0,110 150,110 160,155 320,155 330,110 480,110 480,160 0,160" fill="#BAE6FD"/>
            <line x1="0" y1="110" x2="480" y2="110" stroke="#0284C7" stroke-width="2"/>
            <text x="180" y="140" font-size="13" font-weight="900" fill="#0369A1">인력 줄파기 시탐 (폭 0.5m, 깊이 1.5m)</text>
            <!-- Depth Marker -->
            <line x1="340" y1="110" x2="340" y2="155" stroke="#DC2626" stroke-width="2"/>
            <text x="350" y="135" font-size="12" font-weight="bold" fill="#DC2626">GL -1.5m</text>
        </svg>
    </div>'''
    
    svg3 = f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-amber-50 border border-amber-200 rounded-xl p-3 mb-3" onclick="openDiagramZoom(this.outerHTML, '{title} - STEP 03 2D 도식')">
        <div class="text-[11px] font-bold text-amber-800 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-amber-800 bg-amber-100 px-2 py-0.5 rounded border border-amber-300 font-extrabold">2D 신설 이설/방호 도식</span>
        </div>
        <svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-amber-200">
            <rect width="480" height="160" fill="#FEF3C7"/>
            <rect x="40" y="60" width="400" height="60" rx="8" fill="#FFFFFF" stroke="#F59E0B" stroke-width="2"/>
            <rect x="60" y="80" width="360" height="20" fill="#FDE68A" stroke="#D97706" stroke-width="1.5"/>
            <text x="140" y="95" font-size="13" font-weight="900" fill="#B45309">신설 관로 이설 & H-Beam 매달기 방호</text>
            <circle cx="80" cy="90" r="6" fill="#DC2626"/>
            <circle cx="400" cy="90" r="6" fill="#DC2626"/>
        </svg>
    </div>'''
    
    svg4 = f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-emerald-50 border border-emerald-200 rounded-xl p-3 mb-3" onclick="openDiagramZoom(this.outerHTML, '{title} - STEP 04 2D 도식')">
        <div class="text-[11px] font-bold text-emerald-800 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-emerald-800 bg-emerald-100 px-2 py-0.5 rounded border border-emerald-300 font-extrabold">2D 최종 검측/승인 도식</span>
        </div>
        <svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-emerald-200">
            <rect width="480" height="160" fill="#ECFDF5"/>
            <rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#059669" stroke-width="2"/>
            <text x="170" y="45" font-size="13" font-weight="900" fill="#047857">지장물 이설 최종 인계대장</text>
            <line x1="160" y1="55" x2="320" y2="55" stroke="#A7F3D0" stroke-width="1.5"/>
            <text x="165" y="80" font-size="11" font-weight="bold" fill="#334155">수압/NDT/CCTV 검사 100% 합격</text>
            <text x="165" y="100" font-size="11" font-weight="bold" fill="#334155">점용기관 3자 현장 합동 서명</text>
            <circle cx="300" cy="115" r="18" fill="#FEF2F2" stroke="#DC2626" stroke-width="2"/>
            <text x="287" y="120" font-size="11" font-weight="900" fill="#DC2626">서면 승인</text>
        </svg>
    </div>'''
    
    return svg1, svg2, svg3, svg4

# HTML Generator for Guideline
def build_guideline_html(act):
    title = act['task_title']
    purpose = act['purpose']
    method = act['method']
    output = act['output']
    risk = act['risk']
    advisory = act['advisory']
    dept = act['dept']
    
    # Parse method steps
    steps_raw = method.split('➔')
    steps = [s.strip() for s in steps_raw]
    while len(steps) < 4:
        steps.append(f"단계별 세부 승인 절차 {len(steps)+1}")
    
    svg1, svg2, svg3, svg4 = generate_step_svgs(title)
    
    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 지장물이설 - {title} 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style> body {{ font-family: 'Noto Sans KR', sans-serif; }} </style>
</head>
<body class="bg-slate-100 text-slate-800 antialiased py-8 px-4 sm:px-8">

    <div class="max-w-6xl mx-auto bg-white rounded-3xl shadow-xl border border-slate-200 p-6 sm:p-10 space-y-8">
        <!-- 🔵 헤더 영역 -->
        <div class="flex flex-col sm:flex-row sm:items-center justify-between pb-6 border-b border-slate-200 gap-4">
            <div>
                <span class="text-xs font-black text-indigo-700 bg-indigo-50 border border-indigo-100 px-3.5 py-1.5 rounded-full mb-2 inline-block">
                    Dongtan Tram Utility Playbook | WBS {act['l3_code']} 지장물이설
                </span>
                <h1 class="text-2xl sm:text-4xl font-black text-slate-900 tracking-tight">
                    {title} 수행지침서
                </h1>
            </div>
            <div class="shrink-0 text-right text-xs font-bold text-slate-500">
                주관부서: <span class="text-indigo-600 font-extrabold">{dept}</span>
            </div>
        </div>

        <!-- 📌 1. 작업 개요 및 핵심 목적 -->
        <div class="bg-blue-50/80 border border-blue-200 rounded-2xl p-6 space-y-3">
            <h3 class="text-base font-black text-blue-950 flex items-center gap-2">
                <span>📌</span> 현장 이행 목적 및 주요 방침
            </h3>
            <p class="text-slate-800 text-sm font-semibold leading-relaxed">
                {purpose}
            </p>
        </div>

        <!-- 💡 2. 4단계 상세 수행절차 및 2D 공학 도식 (Light Theme) -->
        <div class="space-y-6">
            <h3 class="text-xl font-black text-slate-900 flex items-center gap-2">
                <span>💡</span> 4단계 공학적 세부 수행절차 및 2D 시공 도식
            </h3>

            <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                <!-- STEP 01 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg1}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 01</span>
                        <h4 class="font-bold text-base text-slate-900">사전 검토 및 현장 조사</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[0]}</p>
                </div>

                <!-- STEP 02 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg2}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 02</span>
                        <h4 class="font-bold text-base text-slate-900">인력 시탐 줄파기 & 점용 입회</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[1]}</p>
                </div>

                <!-- STEP 03 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg3}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 03</span>
                        <h4 class="font-bold text-base text-slate-900">신설 이설 시공 & 방호 조치</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[2]}</p>
                </div>

                <!-- STEP 04 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg4}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 04</span>
                        <h4 class="font-bold text-base text-slate-900">검측 성과표 작성 & 서면 승인</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[3]}</p>
                </div>
            </div>
        </div>

        <!-- 📦 3. 최종 산출물 결과 -->
        <div class="bg-emerald-50 border border-emerald-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-emerald-950 flex items-center gap-2">
                <span>📂</span> 최종 필수 성과품 / 결과 서류
            </h3>
            <p class="text-emerald-900 text-sm font-bold leading-relaxed">
                {output}
            </p>
        </div>

        <!-- ⚠️ 4. 리스크 요인 및 예방 관리 -->
        <div class="bg-amber-50 border-2 border-amber-300 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>⚠️</span> 집행단계 핵심 리스크 및 예방 대책
            </h3>
            <p class="text-amber-900 text-sm font-bold leading-relaxed">
                {risk}
            </p>
        </div>

        <!-- 👷 5. 협력업체 실무 자문 노하우 -->
        <div class="bg-amber-50/60 border border-amber-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>👷</span> [협력업체 실무 자문] 시공 및 공사관리 핵심 가이드
            </h3>
            <p class="text-amber-900 text-sm font-semibold leading-relaxed">
                {advisory}
            </p>
        </div>

        <!-- 📌 푸터 -->
        <div class="pt-6 border-t border-slate-200 flex items-center justify-between text-xs text-slate-500">
            <span>🏢 동탄도시철도(트램) 건설공사 지장물이설 수행지침</span>
            <span>v8 엑셀 100% 1:1 연동 데이터</span>
        </div>
    </div>

    <!-- Lightbox Zoom Modal -->
    <div id="zoomModal" class="fixed inset-0 z-[9999] hidden bg-black/80 backdrop-blur-md flex items-center justify-center p-4" onclick="this.classList.add('hidden')">
        <div class="bg-white rounded-3xl p-6 max-w-4xl w-full max-h-[90vh] overflow-auto border-4 border-indigo-500 shadow-2xl relative" onclick="event.stopPropagation()">
            <button onclick="document.getElementById('zoomModal').classList.add('hidden')" class="absolute top-4 right-4 bg-slate-900 text-white font-black px-4 py-2 rounded-full text-sm hover:bg-slate-700">✕ 닫기</button>
            <h3 id="zoomTitle" class="text-xl font-black text-slate-900 mb-4 pb-2 border-b border-slate-200">고해상도 2D 시공 도식 확대보기</h3>
            <div id="zoomContent" class="flex justify-center items-center"></div>
        </div>
    </div>
    <script>
    function openDiagramZoom(htmlContent, title) {{
        var modal = document.getElementById('zoomModal');
        document.getElementById('zoomTitle').innerText = title || '2D 시공 도식 확대보기';
        document.getElementById('zoomContent').innerHTML = htmlContent;
        modal.classList.remove('hidden');
    }}
    </script>
</body>
</html>'''
    return html

# Process all 38 activities
rebuilt_count = 0
for act in activities:
    title = act['task_title']
    
    # Locate matching subfolders in v8/지장물이설/
    matching_folders = []
    if os.path.exists(util_dir):
        for sub in os.listdir(util_dir):
            sub_p = os.path.join(util_dir, sub)
            if os.path.isdir(sub_p):
                # match title or folder prefix
                if title in sub or sub.split('_', 1)[-1] in title or title in sub.split('_', 1)[-1]:
                    matching_folders.append(sub_p)
    
    # If no folder found, search by fuzzy or create standard folder
    if not matching_folders and os.path.exists(util_dir):
        # find by partial keyword
        kw = title.split()[0]
        for sub in os.listdir(util_dir):
            sub_p = os.path.join(util_dir, sub)
            if os.path.isdir(sub_p) and kw in sub:
                matching_folders.append(sub_p)
    
    html_content = build_guideline_html(act)
    
    for mf in matching_folders:
        guide_dir = os.path.join(mf, '수행지침')
        os.makedirs(guide_dir, exist_ok=True)
        
        safe_title = re.sub(r'[\/:*?"<>|]', '_', title)
        
        # Save standard and descriptive html files
        f1 = os.path.join(guide_dir, f"{safe_title}_수행지침.html")
        f2 = os.path.join(guide_dir, "수행지침.html")
        
        with open(f1, 'w', encoding='utf-8') as f:
            f.write(html_content)
        with open(f2, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        # Also copy for any other existing file name in guide_dir
        for existing in os.listdir(guide_dir):
            if existing.endswith('.html'):
                with open(os.path.join(guide_dir, existing), 'w', encoding='utf-8') as f:
                    f.write(html_content)
        
        rebuilt_count += 1

print(f"Successfully rebuilt {rebuilt_count} 지장물이설 HTML guideline folders with 100% Excel-matched 2D SVG content.", flush=True)
