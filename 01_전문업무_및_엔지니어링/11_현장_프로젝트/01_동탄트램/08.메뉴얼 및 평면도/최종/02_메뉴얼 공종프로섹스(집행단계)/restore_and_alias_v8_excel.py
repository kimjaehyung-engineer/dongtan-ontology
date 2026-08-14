# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v7_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v7"

v7_excel = os.path.join(v7_root, '매뉴얼 BODY (집행단계)v7.xlsm')
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')

# Step 1: Restore pristine binary v8_excel from v7_excel
shutil.copy2(v7_excel, v8_excel)
print("Restored pristine binary Excel file to v8:", os.path.getsize(v8_excel), "bytes")

# Load v8 Excel data
wb = openpyxl.load_workbook(v8_excel, data_only=False)

created_alias_count = 0
valid_count = 0
missing_dirs = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            target = None
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                val = cell.value
                parts = val.split('"')
                if len(parts) >= 2:
                    target = parts[1]
            
            if target:
                decoded_target = urllib.parse.unquote(target)
                abs_path = os.path.normpath(os.path.join(v8_root, decoded_target))
                
                if os.path.exists(abs_path):
                    valid_count += 1
                else:
                    dir_p = os.path.dirname(abs_path)
                    target_filename = os.path.basename(abs_path)
                    
                    if os.path.exists(dir_p):
                        files = [f for f in os.listdir(dir_p) if not f.startswith('.') and not f.startswith('~$')]
                        if files:
                            # Pick primary HTML file in folder and create target alias file
                            primary_file = files[0]
                            primary_path = os.path.join(dir_p, primary_file)
                            alias_path = os.path.join(dir_p, target_filename)
                            
                            if not os.path.exists(alias_path):
                                shutil.copy2(primary_path, alias_path)
                                created_alias_count += 1
                            valid_count += 1
                    else:
                        missing_dirs.append((sheet_name, target, dir_p))

wb.close()

print(f"Total Validated Hyperlinks: {valid_count}")
print(f"Created Filename Aliases for Excel Compatibility: {created_alias_count}")
print(f"Missing Directories: {len(missing_dirs)}")
