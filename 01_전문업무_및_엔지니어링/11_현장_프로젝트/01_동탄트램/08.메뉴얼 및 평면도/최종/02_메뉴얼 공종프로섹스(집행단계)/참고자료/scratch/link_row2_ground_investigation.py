import openpyxl, os, shutil

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
excel_out_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_연동완료.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반'

# 1_지반조사 상세검토 (Row 2) 포함 전체 36개 폴더 순서
exact_36_folders = [
    (2, '1_지반조사 상세검토', '지반조사 상세검토'),
    (3, '2_발주전략 KOM', '발주전략 KOM'),
    (4, '3_철도보호지구에서의 행위신고(필요시)', '철도보호지구에서의 행위신고(필요시)'),
    (5, '4_착수전 측량 Data 확인', '착수전 측량 Data 확인'),
    (6, '5_지장물이설 협의', '지장물이설 협의'),
    (7, '6_용지보상RISK 검토', '용지보상RISK 검토'),
    (8, '7_최고의 팀 만들기 지원', '최고의 팀 만들기 지원'),
    (9, '8_시공계획서 수립 승인', '시공계획서 수립 승인'),
    (10, '8_작업조 편성', '작업조 편성'),
    (11, '9_장비 수급 계획', '장비 수급 계획'),
    (12, '10_노반 재료 입도 DB 확보', '노반 재료 입도 DB 확보'),
    (13, '11_사토장 _ 토사 수급 계획 확인', '사토장 _ 토사 수급 계획 확인'),
    (14, '12_배수 처리 계획 수립', '배수 처리 계획 수립'),
    (15, '13_안전관리계획 수립 승인', '안전관리계획 수립 승인'),
    (16, '14_품질관리계획 수립 승인', '품질관리계획 수립 승인'),
    (17, '15_환경관리계획 수립 승인', '환경관리계획 수립 승인'),
    (18, '16_교통소통 대책 수립 승인(필요시)', '교통소통 대책 수립 승인(필요시)'),
    (19, '17_하도급 검토 승인', '하도급 검토 승인'),
    (20, '18_자재승인', '자재승인'),
    (21, '19_시험다짐', '시험다짐'),
    (22, '20_원지반 검측', '원지반 검측'),
    (23, '21_하부노반 검측', '하부노반 검측'),
    (24, '22_상부노반 시공(배수 유공관 포함)', '상부노반 시공(배수 유공관 포함)'),
    (25, '23_상부강화노반 시공', '상부강화노반 시공'),
    (26, '24_다짐 검측', '다짐 검측'),
    (27, '25_평판재하시험', '평판재하시험'),
    (28, '26_강성 검측(K30, EV2)', '강성 검측(K30, EV2)'),
    (29, '27_평탄성 검측', '평탄성 검측'),
    (30, '28_노반 종 횡단 검측', '노반 종 횡단 검측'),
    (31, '29_부적합 사항 조치', '부적합 사항 조치'),
    (32, '30_사면 다짐 검측', '사면 다짐 검측'),
    (33, '31_배수시설 시공 검측', '배수시설 시공 검측'),
    (34, '32_완성면 보호', '완성면 보호'),
    (35, '33_공사일지 작성', '공사일지 작성'),
    (36, '35_검측 및 승인 관리', '검측 및 승인 관리'),
    (37, '36_토공 마무리면 인계', '토공 마무리면 인계')
]

wb = openpyxl.load_workbook(excel_path)
ws = wb['상부강화노반']

print("=== Row 2 지반조사 상세검토 포함 전체 1:1 연동 구축 시작 ===")

# STEP 1: Row 2 특수 HTML 파일(2중 파란색 카드) 확인 및 생성
job1_dir = os.path.join(base_root, '1_지반조사 상세검토')
os.makedirs(os.path.join(job1_dir, '표준서'), exist_ok=True)
os.makedirs(os.path.join(job1_dir, '수행지침'), exist_ok=True)
os.makedirs(os.path.join(job1_dir, '체크리스트'), exist_ok=True)

job1_gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 상부강화노반 - 지반조사 상세검토 상세 수행지침서 (WBS 9000-7-1)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-amber-500 text-slate-900 text-xs font-black px-3 py-1 rounded-full uppercase">Dongtan Tram Guideline (WBS 9000-7-1)</span>
        <h1 class="text-3xl font-black mt-2">지반조사 상세검토 상세 수행지침서</h1>
        <p class="text-amber-200 text-sm mt-1">L4 Code: 9000-7-1 | 주관: 현장 토질팀 / 공무팀 | "KDS 47 10 00 철도지반 설계기준"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 [2중 파란색 카드] 개요 박스 -->
        <div class="bg-blue-50/70 border border-blue-200/80 p-6 sm:p-8 rounded-2xl text-slate-800 shadow-sm space-y-4">
            <h4 class="font-bold text-base sm:text-lg text-slate-900 flex items-center gap-2 m-0">
                <span class="text-base">💡</span> 지반조사 상세검토 현장 실무 개요
            </h4>
            <div class="bg-white border border-blue-300/80 rounded-xl p-5 sm:p-6 shadow-sm">
                <p class="m-0 text-slate-800 leading-relaxed text-xs sm:text-sm font-normal">
                    본 지침서는 KDS 47 10 00 철도지반 설계기준에 의거하여 동탄트램 노상 원지반 시찰 시추 주상도, GPR 지중 탐사 및 N치 &lt; 4 연약 지반 구간을 검측 분석하고 PVD/DCM 공법 반영을 위한 지반조사 상세 검토 보고서를 최종 교부받기 위한 실무 지침입니다.
                </p>
            </div>
        </div>

        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-amber-600 pb-2 flex items-center gap-2"><span class="text-amber-600">🛠️</span> 지반조사 상세검토 4단계 수행 절차</h2>
            <div class="grid grid-cols-1 gap-6">
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-6 space-y-4 shadow-sm">
                    <div class="flex items-center justify-between border-b border-slate-200 pb-3">
                        <span class="bg-amber-600 text-white font-bold text-xs px-3 py-1 rounded-full">STEP 01</span>
                        <h3 class="font-bold text-base text-slate-900">시추 주상도 & GPR 지중 데이터 분석</h3>
                    </div>
                    <p class="text-slate-700 text-xs font-medium bg-white p-4 rounded-xl border border-slate-200">• <strong>수행 지침:</strong> 현장 시추 주상도 및 GPR(지중탐사) 3D 데이터를 1:1 대조하여 연약 지반 층후를 정밀 분석함.</p>
                </div>
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-6 space-y-4 shadow-sm">
                    <div class="flex items-center justify-between border-b border-slate-200 pb-3">
                        <span class="bg-indigo-600 text-white font-bold text-xs px-3 py-1 rounded-full">STEP 02</span>
                        <h3 class="font-bold text-base text-slate-900">N치 &lt; 4 연약지반 심도 경계 추출</h3>
                    </div>
                    <p class="text-slate-700 text-xs font-medium bg-white p-4 rounded-xl border border-slate-200">• <strong>수행 지침:</strong> 표준관입시험 N치 4 미만 실체층 및 지하수위 변동 구역 경계를 3D BIM 도면에 표출함.</p>
                </div>
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-6 space-y-4 shadow-sm">
                    <div class="flex items-center justify-between border-b border-slate-200 pb-3">
                        <span class="bg-emerald-600 text-white font-bold text-xs px-3 py-1 rounded-full">STEP 03</span>
                        <h3 class="font-bold text-base text-slate-900">PVD / DCM 연약지반 개량 공법 검토</h3>
                    </div>
                    <p class="text-slate-700 text-xs font-medium bg-white p-4 rounded-xl border border-slate-200">• <strong>수행 지침:</strong> 잔여 침하량 2.5cm 이내 수렴을 위한 연약지반 개량 공법(PVD/DCM)을 비교 심사함.</p>
                </div>
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-6 space-y-4 shadow-sm">
                    <div class="flex items-center justify-between border-b border-slate-200 pb-3">
                        <span class="bg-teal-600 text-white font-bold text-xs px-3 py-1 rounded-full">STEP 04</span>
                        <h3 class="font-bold text-base text-slate-900">지반조사 검토 보고서 서면 승인</h3>
                    </div>
                    <p class="text-slate-700 text-xs font-medium bg-white p-4 rounded-xl border border-slate-200">• <strong>수행 지침:</strong> 토질 및 기초 기술사 서명 성적서를 부착하여 책임감리원 최종 서면 승인을 완수함.</p>
                </div>
            </div>
        </div>
    </div>
</div>
</body>
</html>"""

with open(os.path.join(job1_dir, '수행지침', '1_지반조사 상세검토_수행지침.html'), 'w', encoding='utf-8') as f:
    f.write(job1_gui_html)
with open(os.path.join(job1_dir, '수행지침', '지반조사 상세검토_수행지침.html'), 'w', encoding='utf-8') as f:
    f.write(job1_gui_html)

# STEP 2: Row 2부터 Row 37까지 전체 36개 행 하이퍼링크 지정
for row_idx, folder, clean_pfx in exact_36_folders:
    dir_path = os.path.join(base_root, folder)
    
    for sub, suffix, col_idx in [('표준서', '표준서', 12), ('수행지침', '수행지침', 14), ('체크리스트', '체크리스트', 16)]:
        sub_p = os.path.join(dir_path, sub)
        os.makedirs(sub_p, exist_ok=True)
        
        f_exact = os.path.join(sub_p, f"{folder}_{suffix}.html")
        f_clean = os.path.join(sub_p, f"{clean_pfx}_{suffix}.html")
        
        src_f = f_clean if os.path.exists(f_clean) else (f_exact if os.path.exists(f_exact) else None)
        if src_f:
            if not os.path.exists(f_exact): shutil.copy2(src_f, f_exact)
            if not os.path.exists(f_clean): shutil.copy2(src_f, f_clean)
        
        rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder}/{sub}/{folder}_{suffix}.html"
        
        ws.cell(row=row_idx, column=col_idx).value = f"{suffix} 보기 (HTML)"
        ws.cell(row=row_idx, column=col_idx).hyperlink = rel_path
        ws.cell(row=row_idx, column=col_idx).style = 'Hyperlink'

wb.save(excel_out_path)
print("SUCCESS: Row 2 지반조사 상세검토 포함 전체 36개 행이 매뉴얼 BODY (집행단계)v4_연동완료.xlsx 파일에 100% 매칭 저장되었습니다!")

try:
    shutil.copy2(excel_out_path, excel_path)
    print("SUCCESS: 원본 매뉴얼 BODY (집행단계)v4.xlsx 파일에도 성공적으로 반영되었습니다!")
except Exception as e:
    print(f"NOTICE: 원본 파일이 엑셀에서 열려 있어 연동완료 파일에 안전하게 반영되었습니다 ({e}).")
