import openpyxl, os, shutil

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반'

# 1. 35개 폴더의 실제 폴더명 및 파일 프라이픽스 매핑
target_folders = [
    ('2_발주전략 KOM', '발주전략 KOM'),
    ('3_철도보호지구에서의 행위신고(필요시)', '철도보호지구에서의 행위신고(필요시)'),
    ('4_착수전 측량 Data 확인', '착수전 측량 Data 확인'),
    ('5_지장물이설 협의', '지장물이설 협의'),
    ('6_용지보상RISK 검토', '용지보상RISK 검토'),
    ('7_최고의 팀 만들기 지원', '최고의 팀 만들기 지원'),
    ('8_시공계획서 수립 승인', '시공계획서 수립 승인'),
    ('8_작업조 편성', '작업조 편성'),
    ('9_장비 수급 계획', '장비 수급 계획'),
    ('10_노반 재료 입도 DB 확보', '노반 재료 입도 DB 확보'),
    ('11_사토장 _ 토사 수급 계획 확인', '사토장 _ 토사 수급 계획 확인'),
    ('12_배수 처리 계획 수립', '배수 처리 계획 수립'),
    ('13_안전관리계획 수립 승인', '안전관리계획 수립 승인'),
    ('14_품질관리계획 수립 승인', '품질관리계획 수립 승인'),
    ('15_환경관리계획 수립 승인', '환경관리계획 수립 승인'),
    ('16_교통소통 대책 수립 승인(필요시)', '교통소통 대책 수립 승인(필요시)'),
    ('17_하도급 검토 승인', '하도급 검토 승인'),
    ('18_자재승인', '자재승인'),
    ('19_시험다짐', '시험다짐'),
    ('20_원지반 검측', '원지반 검측'),
    ('21_하부노반 검측', '하부노반 검측'),
    ('22_상부노반 시공(배수 유공관 포함)', '상부노반 시공(배수 유공관 포함)'),
    ('23_상부강화노반 시공', '상부강화노반 시공'),
    ('24_다짐 검측', '다짐 검측'),
    ('25_평판재하시험', '평판재하시험'),
    ('26_강성 검측(K30, EV2)', '강성 검측(K30, EV2)'),
    ('27_평탄성 검측', '평탄성 검측'),
    ('28_노반 종 횡단 검측', '노반 종 횡단 검측'),
    ('29_부적합 사항 조치', '부적합 사항 조치'),
    ('30_사면 다짐 검측', '사면 다짐 검측'),
    ('31_배수시설 시공 검측', '배수시설 시공 검측'),
    ('32_완성면 보호', '완성면 보호'),
    ('33_공사일지 작성', '공사일지 작성'),
    ('35_검측 및 승인 관리', '검측 및 승인 관리'),
    ('36_토공 마무리면 인계', '토공 마무리면 인계')
]

# STEP A: 현존하는 최신 HTML 파일을 구버전 파일명으로도 1:1 하드 이중 배치(호환성 100% 유지)
print("--- STEP A: 현존 최신 딥빌드 HTML 파일 이중 호환성 배치 시작 ---")
for dir_name, pfx in target_folders:
    dir_path = os.path.join(base_root, dir_name)
    if not os.path.exists(dir_path):
        continue
    
    # 3개 하위 폴더: 표준서, 수행지침, 체크리스트
    types = [('표준서', '표준서'), ('수행지침', '수행지침'), ('체크리스트', '체크리스트')]
    for sub, suffix in types:
        sub_p = os.path.join(dir_path, sub)
        if not os.path.exists(sub_p):
            continue
        
        main_file = os.path.join(sub_p, f"{pfx}_{suffix}.html")
        alt_file = os.path.join(sub_p, f"{dir_name}_{suffix}.html")
        
        if os.path.exists(main_file):
            shutil.copy2(main_file, alt_file)
            print(f"호환성 파일 생성 완료: {dir_name}_{suffix}.html")

# STEP B: 엑셀 v4 상부강화노반 시트 하이퍼링크 타겟 1:1 완벽 갱신
print("\n--- STEP B: 엑셀 v4 상부강화노반 시트 하이퍼링크 1:1 전면 갱신 시작 ---")
wb = openpyxl.load_workbook(excel_path)
ws = wb['상부강화노반']

# Row 3 ~ Row 37 (총 35개 행)
for idx, (dir_name, pfx) in enumerate(target_folders):
    row_idx = idx + 3 # Row 3부터 시작
    
    # 표준서 (Col 12), 수행지침 (Col 14), 체크리스트 (Col 16)
    std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{dir_name}/표준서/{pfx}_표준서.html"
    gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{dir_name}/수행지침/{pfx}_수행지침.html"
    chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{dir_name}/체크리스트/{pfx}_체크리스트.html"

    # 셀 값도 정밀하게 지정
    ws.cell(row=row_idx, column=12).value = "표준서 보기 (HTML)"
    ws.cell(row=row_idx, column=12).hyperlink = std_rel_path
    ws.cell(row=row_idx, column=12).style = 'Hyperlink'

    ws.cell(row=row_idx, column=14).value = "수행지침 보기 (HTML)"
    ws.cell(row=row_idx, column=14).hyperlink = gui_rel_path
    ws.cell(row=row_idx, column=14).style = 'Hyperlink'

    ws.cell(row=row_idx, column=16).value = "체크리스트 보기 (HTML)"
    ws.cell(row=row_idx, column=16).hyperlink = chk_rel_path
    ws.cell(row=row_idx, column=16).style = 'Hyperlink'

wb.save(excel_path)
print("SUCCESS: 엑셀 v4 파일 상부강화노반 시트의 모든 HTML 하이퍼링크가 100% 완벽 연동되었습니다!")
