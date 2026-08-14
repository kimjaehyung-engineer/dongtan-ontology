import openpyxl

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=False)

sheet_name = None
for s in wb.sheetnames:
    if '상부강화노반' in s or '강화노반' in s:
        sheet_name = s
        break

ws = wb[sheet_name]

print("=== Row 1 ~ 3 전체 컬럼 구조 확인 ===")
for r in range(1, 4):
    print(f"--- Row {r} ---")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        val = cell.value
        link = cell.hyperlink.target if cell.hyperlink else None
        if val or link:
            print(f"  Col {c:2d}: val={val} | link={link}")
