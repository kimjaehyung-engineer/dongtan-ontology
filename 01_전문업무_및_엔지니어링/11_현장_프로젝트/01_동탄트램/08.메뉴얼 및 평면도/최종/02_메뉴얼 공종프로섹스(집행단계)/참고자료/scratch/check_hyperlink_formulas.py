import openpyxl

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=False)

for sname in ['전기', '상부강화노반']:
    if sname in wb.sheetnames:
        ws = wb[sname]
        print(f"=== Sheet: {sname} ===")
        for r in range(1, 38):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = str(cell.value)
                h = cell.hyperlink.target if cell.hyperlink else None
                if 'HYPERLINK' in v or h or '.html' in v or 'http' in v:
                    print(f"Row {r:2d}, Col {c:2d}: link={repr(h)}")
