import openpyxl, os

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)'

wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['지장물이설']

print("=== '지장물이설' 시트 Rows 및 헤더 검사 ===")
print(f"최대 행 수: {ws.max_row}, 최대 열 수: {ws.max_column}")

for r in range(1, ws.max_row + 1):
    wbs = str(ws.cell(row=r, column=4).value or '').strip()
    act_id = str(ws.cell(row=r, column=5).value or '').strip()
    act_name = str(ws.cell(row=r, column=6).value or '').strip()
    col8_val = str(ws.cell(row=r, column=8).value or '').strip()
    col12_val = str(ws.cell(row=r, column=12).value or '').strip()
    
    if wbs or act_name or r <= 3:
        print(f"Row {r:2d} | Col 4(WBS): {wbs} | Col 5(ID): {act_id} | Col 6(Task): {act_name}")
        print(f"       | Col 8(표준서요약): {col8_val[:30]}")

# 지장물이설 관련 첨부 폴더 탐색
jijang_folders = []
for d in os.listdir(base_root):
    if '지장물' in d:
        jijang_folders.append(d)

print(f"\n매뉴얼BODY(집행단계-첨부폴더) 내 지장물 관련 폴더: {jijang_folders}")
