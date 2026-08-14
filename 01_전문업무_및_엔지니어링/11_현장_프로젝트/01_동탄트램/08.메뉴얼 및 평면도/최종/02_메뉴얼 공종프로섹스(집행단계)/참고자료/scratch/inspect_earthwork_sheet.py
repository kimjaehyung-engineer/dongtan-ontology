import openpyxl, os

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)'

wb = openpyxl.load_workbook(excel_path, data_only=False)

target_sheet = None
for s in wb.sheetnames:
    if '사전토공' in s or '토공사' in s or '토공' in s:
        target_sheet = s
        break

print(f"찾은 시트 명칭: {target_sheet}")

ws = wb[target_sheet]
print(f"최대 행 수: {ws.max_row}, 최대 열 수: {ws.max_column}")

for r in range(1, ws.max_row + 1):
    wbs = str(ws.cell(row=r, column=4).value or '').strip()
    act_id = str(ws.cell(row=r, column=5).value or '').strip()
    act_name = str(ws.cell(row=r, column=6).value or '').strip()
    col8_val = str(ws.cell(row=r, column=8).value or '').strip()
    
    if wbs or act_name or r <= 3:
        print(f"Row {r:2d} | WBS: {wbs} | ID: {act_id} | Task: {act_name}")

# 첨부 폴더 탐색
earth_folders = []
for d in os.listdir(base_root):
    if '토공' in d or '사전' in d:
        earth_folders.append(d)

print(f"\n매뉴얼BODY(집행단계-첨부폴더) 내 토공 관련 폴더: {earth_folders}")
