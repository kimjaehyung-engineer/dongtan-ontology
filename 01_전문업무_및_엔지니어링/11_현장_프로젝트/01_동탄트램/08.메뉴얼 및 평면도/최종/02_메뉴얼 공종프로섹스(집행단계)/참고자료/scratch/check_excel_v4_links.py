import openpyxl, os

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=False)

sheet_name = None
for s in wb.sheetnames:
    if '상부강화노반' in s or '강화노반' in s:
        sheet_name = s
        break

print(f"Sheet Name: {sheet_name}")

ws = wb[sheet_name]

# Row 3 ~ Row 37
for r in range(3, 38):
    act_name = ws.cell(row=r, column=4).value
    std_cell = ws.cell(row=r, column=8)
    gui_cell = ws.cell(row=r, column=9)
    chk_cell = ws.cell(row=r, column=10)
    
    std_hyper = std_cell.hyperlink.target if std_cell.hyperlink else "NO_HYPERLINK"
    gui_hyper = gui_cell.hyperlink.target if gui_cell.hyperlink else "NO_HYPERLINK"
    chk_hyper = chk_cell.hyperlink.target if chk_cell.hyperlink else "NO_HYPERLINK"
    
    print(f"Row {r:2d} | Activity: {act_name}")
    print(f"  - Std Cell (Col H): val={std_cell.value} | link={std_hyper}")
    print(f"  - Gui Cell (Col I): val={gui_cell.value} | link={gui_hyper}")
    print(f"  - Chk Cell (Col J): val={chk_cell.value} | link={chk_hyper}")
    if r >= 8:
        break
