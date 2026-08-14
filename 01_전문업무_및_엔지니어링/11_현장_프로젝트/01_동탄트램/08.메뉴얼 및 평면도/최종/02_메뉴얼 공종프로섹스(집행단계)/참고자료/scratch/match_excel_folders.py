import openpyxl, os

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반'

# 실제 디렉토리 목록
actual_folders = sorted([d for d in os.listdir(base_root) if os.path.isdir(os.path.join(base_root, d))])

print(f"=== 실제 상부강화노반 폴더 개수: {len(actual_folders)} ===")

wb = openpyxl.load_workbook(excel_path, data_only=False)
ws = wb['상부강화노반']

print("\n=== Excel v4 '상부강화노반' 시트 Row 3~37 목록 ===")
excel_rows = []
for r in range(3, ws.max_row + 1):
    wbs = ws.cell(row=r, column=4).value
    act_id = ws.cell(row=r, column=5).value
    act_name = ws.cell(row=r, column=6).value
    if wbs or act_name:
        excel_rows.append((r, str(wbs).strip(), str(act_id).strip(), str(act_name).strip()))

print(f"Excel 행 개수: {len(excel_rows)}")
print("\n--- Excel 행과 실제 폴더 매칭 상태 확인 ---")
for r, wbs, act_id, act_name in excel_rows:
    # 매수 폴더 찾기
    matched = [f for f in actual_folders if act_name in f or f.endswith(act_name)]
    print(f"Row {r:2d} | WBS: {wbs} | ID: {act_id} | Name: {act_name}")
    print(f"       -> 매칭된 실제 폴더: {matched}")
