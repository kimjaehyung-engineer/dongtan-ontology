import openpyxl

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=False)

# 다른 연동 잘된 시트 (예: 전기 시트) 조사
elec_sheet = None
for s in wb.sheetnames:
    if '전기' in s:
        elec_sheet = s
        break

print(f"전기 시트 명칭: {elec_sheet}")
ws_elec = wb[elec_sheet]

print("\n--- 전기 시트 Row 3의 Col 12, 14, 16 링크 수식/값 샘플 ---")
for c in [12, 14, 16]:
    cell = ws_elec.cell(row=3, column=c)
    link = cell.hyperlink.target if cell.hyperlink else None
    print(f"Col {c:2d}: val={cell.value} | link={link}")

print("\n--- 상부강화노반 시트 Row 3~5 의 Col 12, 14, 16 실측 ---")
ws_tram = wb['상부강화노반']
for r in range(3, 6):
    print(f"Row {r}:")
    for c in [12, 14, 16]:
        cell = ws_tram.cell(row=r, column=c)
        link = cell.hyperlink.target if cell.hyperlink else None
        print(f"  Col {c:2d}: val={cell.value} | link={link}")
