import openpyxl, os

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'

wb = openpyxl.load_workbook(excel_path, data_only=False)

sheet_name = None
for s in wb.sheetnames:
    if '상부강화노반' in s or '강화노반' in s:
        sheet_name = s
        break

print(f"Sheet Name: {sheet_name}")

ws = wb[sheet_name]

# 헤더 행 및 샘플 행 확인
for col in range(1, 15):
    h_val = ws.cell(row=2, column=col).value
    print(f"Col {col:2d}: {h_val}")

print("\n--- Row 3~5 셀 값 및 수식 검사 ---")
for r in range(3, 6):
    print(f"Row {r}:")
    print(f"  Col 4 (WBS/작업명): {ws.cell(row=r, column=4).value}")
    print(f"  Col 8 (표준서)    : val={ws.cell(row=r, column=8).value}, link={ws.cell(row=r, column=8).hyperlink}")
    print(f"  Col 9 (수행지침)  : val={ws.cell(row=r, column=9).value}, link={ws.cell(row=r, column=9).hyperlink}")
    print(f"  Col 10 (체크리스트): val={ws.cell(row=r, column=10).value}, link={ws.cell(row=r, column=10).hyperlink}")
