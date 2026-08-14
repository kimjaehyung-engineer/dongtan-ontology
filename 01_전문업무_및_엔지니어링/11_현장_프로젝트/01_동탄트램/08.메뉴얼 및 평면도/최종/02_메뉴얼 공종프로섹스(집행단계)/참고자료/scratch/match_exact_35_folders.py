import openpyxl, os, shutil, time

excel_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx'
base_root = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반'

# 스크린샷에 보이는 exact 35개 폴더 순서
exact_35_folders = [
    '2_발주전략 KOM',
    '3_철도보호지구에서의 행위신고(필요시)',
    '4_착수전 측량 Data 확인',
    '5_지장물이설 협의',
    '6_용지보상RISK 검토',
    '7_최고의 팀 만들기 지원',
    '8_시공계획서 수립 승인',
    '8_작업조 편성',
    '9_장비 수급 계획',
    '10_노반 재료 입도 DB 확보',
    '11_사토장 _ 토사 수급 계획 확인',
    '12_배수 처리 계획 수립',
    '13_안전관리계획 수립 승인',
    '14_품질관리계획 수립 승인',
    '15_환경관리계획 수립 승인',
    '16_교통소통 대책 수립 승인(필요시)',
    '17_하도급 검토 승인',
    '18_자재승인',
    '19_시험다짐',
    '20_원지반 검측',
    '21_하부노반 검측',
    '22_상부노반 시공(배수 유공관 포함)',
    '23_상부강화노반 시공',
    '24_다짐 검측',
    '25_평판재하시험',
    '26_강성 검측(K30, EV2)',
    '27_평탄성 검측',
    '28_노반 종 횡단 검측',
    '29_부적합 사항 조치',
    '30_사면 다짐 검측',
    '31_배수시설 시공 검측',
    '32_완성면 보호',
    '33_공사일지 작성',
    '35_검측 및 승인 관리',
    '36_토공 마무리면 인계'
]

wb = openpyxl.load_workbook(excel_path)
ws = wb['상부강화노반']

print("=== 35개 스크린샷 폴더 1:1 하이퍼링크 매칭 시작 ===")

for idx, folder in enumerate(exact_35_folders):
    row_idx = idx + 3 # Row 3부터 37까지
    clean_pfx = folder.split('_', 1)[-1] if '_' in folder else folder
    dir_path = os.path.join(base_root, folder)
    
    # 3개 하위 폴더별 파일 생성 및 엑셀 하이퍼링크 설정
    for sub, suffix, col_idx in [('표준서', '표준서', 12), ('수행지침', '수행지침', 14), ('체크리스트', '체크리스트', 16)]:
        sub_p = os.path.join(dir_path, sub)
        os.makedirs(sub_p, exist_ok=True)
        
        f_exact = os.path.join(sub_p, f"{folder}_{suffix}.html")
        f_clean = os.path.join(sub_p, f"{clean_pfx}_{suffix}.html")
        
        # 소스 복사 (어느 하나라도 있으면 서로 복사하여 듀얼 완비)
        src_f = f_clean if os.path.exists(f_clean) else (f_exact if os.path.exists(f_exact) else None)
        if src_f:
            if not os.path.exists(f_exact): shutil.copy2(src_f, f_exact)
            if not os.path.exists(f_clean): shutil.copy2(src_f, f_clean)
        
        # 엑셀 하이퍼링크 지정 (상대경로)
        rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder}/{sub}/{folder}_{suffix}.html"
        
        ws.cell(row=row_idx, column=col_idx).value = f"{suffix} 보기 (HTML)"
        ws.cell(row=row_idx, column=col_idx).hyperlink = rel_path
        ws.cell(row=row_idx, column=col_idx).style = 'Hyperlink'

# 저장 시도 (파일 열림 락 대비)
saved = False
for attempt in range(5):
    try:
        wb.save(excel_path)
        saved = True
        print("\nSUCCESS: 35개 폴더 하이퍼링크가 엑셀 v4 상부강화노반 시트에 100% 저장되었습니다!")
        break
    except PermissionError:
        print(f"Attempt {attempt+1}: 엑셀 파일이 열려 있어 저장이 지연 중입니다. 2초 후 재시도...")
        time.sleep(2)

if not saved:
    print("\nWARNING: 엑셀 파일이 사용 중이어서 직접 저장이 차단되었습니다. 엑셀을 닫고 다시 실행해주시거나 안내해 드립니다.")
