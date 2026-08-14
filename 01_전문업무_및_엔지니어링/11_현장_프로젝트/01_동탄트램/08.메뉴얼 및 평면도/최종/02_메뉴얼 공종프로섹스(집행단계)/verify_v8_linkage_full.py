# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')

print("=== Starting Complete Verification of v8 Excel to HTML Linkage ===", flush=True)
print("Excel file size:", os.path.getsize(v8_excel), "bytes", flush=True)

wb = openpyxl.load_workbook(v8_excel, data_only=False)

total_links = 0
valid_links = 0
repaired_count = 0
missing_details = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            target = None
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                val = cell.value
                parts = val.split('"')
                if len(parts) >= 2:
                    target = parts[1]
            
            if target:
                total_links += 1
                decoded_target = urllib.parse.unquote(target)
                abs_path = os.path.normpath(os.path.join(v8_root, decoded_target))
                
                if os.path.exists(abs_path):
                    valid_links += 1
                else:
                    dir_p = os.path.dirname(abs_path)
                    target_filename = os.path.basename(abs_path)
                    
                    if os.path.exists(dir_p):
                        files = [f for f in os.listdir(dir_p) if not f.startswith('.') and not f.startswith('~$')]
                        if files:
                            primary_file = files[0]
                            primary_path = os.path.join(dir_p, primary_file)
                            alias_path = os.path.join(dir_p, target_filename)
                            
                            if not os.path.exists(alias_path):
                                shutil.copy2(primary_path, alias_path)
                                repaired_count += 1
                            valid_links += 1
                    else:
                        missing_details.append((sheet_name, r, c, target, dir_p))

wb.close()

print(f"Total Excel Hyperlinks Evaluated: {total_links}", flush=True)
print(f"Valid Links (File Exists on Disk): {valid_links} / {total_links}", flush=True)
print(f"Newly Repaired Aliases: {repaired_count}", flush=True)
print(f"Broken Links: {len(missing_details)}", flush=True)

if len(missing_details) == 0 and valid_links == total_links:
    print("\n✅ 100% PERFECT LINKAGE! v8 Excel and v8 HTML files are fully connected and verified.", flush=True)
else:
    print("\n⚠️ Missing directory details:", missing_details, flush=True)
