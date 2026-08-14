# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse, re

sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')
util_dir = os.path.join(v8_root, '지장물이설')

print("=== Starting Advanced Generation of Category-Tailored 지장물이설 HTML Manuals ===", flush=True)

wb = openpyxl.load_workbook(v8_excel, data_only=True)
ws = wb['지장물이설']

activities = []
for r in range(2, ws.max_row + 1):
    task_title = ws.cell(row=r, column=8).value
    if task_title:
        activities.append({
            'row': r,
            'l3_code': ws.cell(row=r, column=2).value or '2000',
            'task_title': str(task_title).strip(),
            'dept': str(ws.cell(row=r, column=9).value or '현장 공무/시공팀').strip(),
            'purpose': str(ws.cell(row=r, column=10).value or '지장물 이설 품질 및 안전 검측').strip(),
            'method': str(ws.cell(row=r, column=11).value or '1) 사전검토 -> 2) 현측 -> 3) 시공 -> 4) 승인').strip(),
            'output': str(ws.cell(row=r, column=12).value or '지장물 이설 정밀 검측성과표 및 최종 승인서').strip(),
            'risk': str(ws.cell(row=r, column=21).value or '지하시설물 간섭 및 안전사고 리스크 사전 차단').strip(),
            'advisory': str(ws.cell(row=r, column=22).value or '관할 점용기관 및 시공 전문 기술자 자문').strip(),
        })

wb.close()

def categorize_activity(title):
    t = title.lower()
    if 'gpr' in t or '시탐' in t or '줄파기' in t or '지반조사' in t or '탐사' in t:
        return 'gpr'
    elif '한전' in t or '전력' in t or '22.9kv' in t or '전기' in t or '케이블' in t:
        return 'electric'
    elif '가스' in t or '핫태핑' in t or '도시가스' in t or 'stopper' in t:
        return 'gas'
    elif '상수' in t or '하수' in t or '수압' in t or 'cctv' in t or '관로' in t or '암거' in t or '배수' in t:
        return 'water'
    elif '통신' in t or '광케이블' in t or 'otdr' in t or 'kt' in t:
        return 'telecom'
    elif '송유관' in t or 'ndt' in t or '비파괴' in t or '초음파' in t:
        return 'oil'
    elif '교통' in t or '차선' in t or '점용' in t or '보증보험' in t or '신호수' in t or '도로' in t:
        return 'traffic'
    else:
        return 'settlement'

def get_theme_config(cat):
    configs = {
        'gpr': {
            'name': 'GPR 3D 물리탐사 & 지반조사',
            'primary': 'sky-600', 'badge': 'bg-sky-50 text-sky-700 border-sky-200',
            'bg_box': 'bg-sky-50/60 border-sky-200', 'title_color': 'text-sky-950',
            'sim_title': '📡 GPR 3D 지하 매설물 레이더 스캐닝 시뮬레이터',
            'sim_html': '''<div class="bg-sky-50/80 border border-sky-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-sky-950 text-sm flex items-center gap-2">
                        <span>📡</span> GPR 3D 지중 관로 탐사 레이더 스캐너 (실시간 탐상 모니터)
                    </h4>
                    <span class="text-xs bg-sky-600 text-white font-black px-2.5 py-1 rounded-full">탐사 심도 GL -3.5m</span>
                </div>
                <div class="bg-slate-900 text-green-400 font-mono text-xs p-4 rounded-xl space-y-2 border border-slate-700">
                    <div class="flex justify-between border-b border-slate-700 pb-1">
                        <span>SCAN FREQ: 400MHz / 900MHz</span>
                        <span id="gprStatus" class="text-emerald-400 font-bold">● TARGET LOCKED (3 관로 감지)</span>
                    </div>
                    <div class="h-24 bg-slate-950 rounded flex items-center justify-around px-4 relative overflow-hidden">
                        <div class="text-center"><span class="block text-[10px] text-slate-400">GL -0.8m</span><div class="w-6 h-6 rounded-full bg-amber-500 mx-auto animate-pulse"></div><span class="text-[10px] text-amber-300">통신 광케이블</span></div>
                        <div class="text-center"><span class="block text-[10px] text-slate-400">GL -1.5m</span><div class="w-8 h-8 rounded-full bg-red-500 mx-auto animate-pulse"></div><span class="text-[10px] text-red-300">22.9kV 전력관</span></div>
                        <div class="text-center"><span class="block text-[10px] text-slate-400">GL -2.2m</span><div class="w-10 h-10 rounded-full bg-blue-500 mx-auto animate-pulse"></div><span class="text-[10px] text-blue-300">D500 상수도관</span></div>
                    </div>
                </div>
                <div class="flex gap-3">
                    <button onclick="document.getElementById('gprStatus').innerText='● GPR 3D SCAN COMPLETED (좌표 100% 매핑됨)';" class="text-xs bg-sky-600 hover:bg-sky-700 text-white font-black px-4 py-2 rounded-xl shadow">▶ 3D 정밀 스캔 재실행</button>
                    <button onclick="document.getElementById('gprStatus').innerText='● 인력 시탐 줄파기(깊이 1.5m) 위치 확인 완료';" class="text-xs bg-slate-100 hover:bg-slate-200 text-slate-700 font-bold px-4 py-2 rounded-xl border border-slate-300">줄파기 위치 매핑</button>
                </div>
            </div>'''
        },
        'electric': {
            'name': '한전 22.9kV 특고압 전력관로 이설',
            'primary': 'red-600', 'badge': 'bg-red-50 text-red-700 border-red-200',
            'bg_box': 'bg-red-50/60 border-red-200', 'title_color': 'text-red-950',
            'sim_title': '⚡ 22.9kV 특고압 케이블 절연내력 및 활선 Cut-over 제어기',
            'sim_html': '''<div class="bg-red-50/70 border border-red-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-red-950 text-sm flex items-center gap-2">
                        <span>⚡</span> 한전 KEPCO 22.9kV CNCV-W 케이블 절연내력 시험기
                    </h4>
                    <span class="text-xs bg-red-600 text-white font-black px-2.5 py-1 rounded-full">시험 전압 30,000V DC</span>
                </div>
                <div class="grid grid-cols-3 gap-3 text-center">
                    <div class="bg-white p-3 rounded-xl border border-red-200"><span class="text-[11px] text-slate-500 block">R상 절연저항</span><b class="text-sm text-red-600">2,000 MΩ (합격)</b></div>
                    <div class="bg-white p-3 rounded-xl border border-red-200"><span class="text-[11px] text-slate-500 block">S상 절연저항</span><b class="text-sm text-red-600">2,000 MΩ (합격)</b></div>
                    <div class="bg-white p-3 rounded-xl border border-red-200"><span class="text-[11px] text-slate-500 block">T상 절연저항</span><b class="text-sm text-red-600">2,000 MΩ (합격)</b></div>
                </div>
                <div class="flex gap-3">
                    <button onclick="alert('⚡ 한전 KEPCO 활선작업팀 입회하에 휴전(Cut-over) 100% 승인 완료되었습니다.');" class="text-xs bg-red-600 hover:bg-red-700 text-white font-black px-4 py-2 rounded-xl shadow">한전 입회 절연시험 승인</button>
                </div>
            </div>'''
        },
        'gas': {
            'name': '도시가스 중압배관 Hot Tapping & Stopper 우회 차단',
            'primary': 'amber-600', 'badge': 'bg-amber-50 text-amber-800 border-amber-200',
            'bg_box': 'bg-amber-50/60 border-amber-200', 'title_color': 'text-amber-950',
            'sim_title': '🔥 무정지 Hot Tapping & Stopper 우회 차단 시뮬레이터',
            'sim_html': '''<div class="bg-amber-50/70 border border-amber-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-amber-950 text-sm flex items-center gap-2">
                        <span>🔥</span> 도시가스 중압배관(0.85MPa) 무정지 천공 & Stopper 차단 상태
                    </h4>
                    <span class="text-xs bg-amber-600 text-white font-black px-2.5 py-1 rounded-full">By-Pass 우회 가스 공급중</span>
                </div>
                <div class="bg-white p-4 rounded-xl border border-amber-200 text-xs space-y-2">
                    <div class="flex justify-between"><span>배관 내 가스압력:</span><b class="text-amber-700">0.85 MPa (안정)</b></div>
                    <div class="flex justify-between"><span>가스안전공사 입회 PID 누출 테스터:</span><b class="text-emerald-600">0.00 ppm (완전 차단)</b></div>
                    <div class="flex justify-between"><span>방폭 샌드위치 밸브 밀폐율:</span><b class="text-indigo-600">100% 기밀 유지</b></div>
                </div>
            </div>'''
        },
        'water': {
            'name': '상하수도 관로 이설 및 수압/CCTV 검사',
            'primary': 'cyan-600', 'badge': 'bg-cyan-50 text-cyan-700 border-cyan-200',
            'bg_box': 'bg-cyan-50/60 border-cyan-200', 'title_color': 'text-cyan-950',
            'sim_title': '💧 D1000 상수관로 10kgf/cm² 수압시험 시뮬레이터',
            'sim_html': '''<div class="bg-cyan-50/70 border border-cyan-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-cyan-950 text-sm flex items-center gap-2">
                        <span>💧</span> 신설 주철 상수도관(D1000) 수압시험 (KDS 표준 1.0 MPa 유지)
                    </h4>
                    <span class="text-xs bg-cyan-600 text-white font-black px-2.5 py-1 rounded-full">1시간 유지 시험</span>
                </div>
                <div class="grid grid-cols-2 gap-3 text-center">
                    <div class="bg-white p-3 rounded-xl border border-cyan-200"><span class="text-[11px] text-slate-500 block">시험 압력</span><b class="text-sm text-cyan-700">10.0 kgf/cm² (1.0 MPa)</b></div>
                    <div class="bg-white p-3 rounded-xl border border-cyan-200"><span class="text-[11px] text-slate-500 block">60분 후 누수량</span><b class="text-sm text-emerald-600">0.00 L (합격)</b></div>
                </div>
                <div class="flex gap-3">
                    <button onclick="alert('💧 수압시험 및 CCTV 내시경 관로 검사 100% 합격되었습니다.');" class="text-xs bg-cyan-600 hover:bg-cyan-700 text-white font-black px-4 py-2 rounded-xl shadow">수압 검측 성과표 발행</button>
                </div>
            </div>'''
        },
        'telecom': {
            'name': '통신 광케이블 이설 & OTDR 파형 판독',
            'primary': 'violet-600', 'badge': 'bg-violet-50 text-violet-700 border-violet-200',
            'bg_box': 'bg-violet-50/60 border-violet-200', 'title_color': 'text-violet-950',
            'sim_title': '📶 OTDR 광섬유 손실 파형 분석기',
            'sim_html': '''<div class="bg-violet-50/70 border border-violet-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-violet-950 text-sm flex items-center gap-2">
                        <span>📶</span> OTDR 광섬유 융착 접속 감쇠량 분석 (Wavelength 1550nm)
                    </h4>
                    <span class="text-xs bg-violet-600 text-white font-black px-2.5 py-1 rounded-full">손실율 ≤ 0.05dB</span>
                </div>
                <div class="bg-slate-900 text-purple-300 font-mono text-xs p-4 rounded-xl border border-slate-700">
                    <div class="flex justify-between border-b border-slate-700 pb-1">
                        <span>FIBER: 144 Core Single Mode</span>
                        <span class="text-emerald-400">STATUS: PASS (손실 0.03dB)</span>
                    </div>
                    <p class="text-[11px] text-slate-400 mt-2">융착점 접속 손실: 0.03 dB / km | 반사 손실: > 60 dB (KT/SKT/LGU+ 기준 충족)</p>
                </div>
            </div>'''
        },
        'oil': {
            'name': '대한송유관공사(DOPCO) 특수 송유관 NDT 검사',
            'primary': 'slate-700', 'badge': 'bg-slate-100 text-slate-800 border-slate-300',
            'bg_box': 'bg-slate-50 border-slate-300', 'title_color': 'text-slate-900',
            'sim_title': '🔬 송유관 용접부 NDT 초음파(UT)/방사선(RT) 결함 스코프',
            'sim_html': '''<div class="bg-slate-50 border border-slate-300 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-slate-900 text-sm flex items-center gap-2">
                        <span>🔬</span> 송유관 배관 원주 용접부 100% RT/UT 비파괴 검사
                    </h4>
                    <span class="text-xs bg-slate-900 text-white font-black px-2.5 py-1 rounded-full">대한송유관공사 입회</span>
                </div>
                <div class="bg-white p-3 rounded-xl border border-slate-300 text-xs space-y-1.5">
                    <div class="flex justify-between"><span>RT 방사선 투과 판정:</span><b class="text-emerald-700">1급 (결함 0건)</b></div>
                    <div class="flex justify-between"><span>UT 초음파 탐상 판정:</span><b class="text-emerald-700">합격 (내부 균열 없음)</b></div>
                    <div class="flex justify-between"><span>강관 Casing Sleeve 이격:</span><b class="text-blue-700">300mm 완벽 절연</b></div>
                </div>
            </div>'''
        },
        'traffic': {
            'name': '도로점용 인허가, 가설 둔차선 및 교통통제',
            'primary': 'indigo-600', 'badge': 'bg-indigo-50 text-indigo-700 border-indigo-200',
            'bg_box': 'bg-indigo-50/60 border-indigo-200', 'title_color': 'text-indigo-950',
            'sim_title': '🚧 5단계 차로 전환 및 가설 둔차선 교통통제 모니터',
            'sim_html': '''<div class="bg-indigo-50/70 border border-indigo-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-indigo-950 text-sm flex items-center gap-2">
                        <span>🚧</span> 동탄경찰서 차선 점용 허가 및 가설 둔차선(복공판) 배치도
                    </h4>
                    <span class="text-xs bg-indigo-600 text-white font-black px-2.5 py-1 rounded-full">전담 신호수 2인 2조 배치</span>
                </div>
                <div class="grid grid-cols-2 gap-3 text-xs">
                    <div class="bg-white p-3 rounded-xl border border-indigo-200"><span>PE 방호벽 & 쏠라 경광등:</span><b class="text-indigo-700 block mt-1">100m 연속 거치 완료</b></div>
                    <div class="bg-white p-3 rounded-xl border border-indigo-200"><span>도로복구 이행 보증보험:</span><b class="text-emerald-600 block mt-1">증권 화성시 제출 완료</b></div>
                </div>
            </div>'''
        },
        'settlement': {
            'name': '위수탁 지장물 행정·재정 정산 및 인계인수',
            'primary': 'emerald-600', 'badge': 'bg-emerald-50 text-emerald-800 border-emerald-200',
            'bg_box': 'bg-emerald-50/60 border-emerald-200', 'title_color': 'text-emerald-950',
            'sim_title': '📑 3D BIM 지하 매설물 간섭 제로화 & 정산 매트릭스',
            'sim_html': '''<div class="bg-emerald-50/70 border border-emerald-200 rounded-2xl p-6 space-y-4">
                <div class="flex items-center justify-between">
                    <h4 class="font-extrabold text-emerald-950 text-sm flex items-center gap-2">
                        <span>📑</span> 실정보고 투입 물량 정산 & 3자 합동 인계 대장
                    </h4>
                    <span class="text-xs bg-emerald-600 text-white font-black px-2.5 py-1 rounded-full">3D BIM 간섭 0건 확인</span>
                </div>
                <div class="bg-white p-3 rounded-xl border border-emerald-200 text-xs space-y-1.5">
                    <div class="flex justify-between"><span>도급분/위수탁분 단가 검증:</span><b class="text-emerald-700">공학 타당성 서면 승인</b></div>
                    <div class="flex justify-between"><span>후행 궤도 공정 간섭:</span><b class="text-indigo-700">3자 합동 현측 인계 완료</b></div>
                </div>
            </div>'''
        }
    }
    return configs.get(cat, configs['gpr'])

def build_svg_diagrams_for_category(cat, title):
    # Generates 4 distinct, category-tailored 2D SVGs
    if cat == 'electric':
        s1 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-red-200"><rect width="480" height="160" fill="#FEF2F2"/><line x1="0" y1="120" x2="480" y2="120" stroke="#DC2626" stroke-width="2"/><rect x="40" y="60" width="100" height="55" rx="6" fill="#EF4444"/><circle cx="70" cy="85" r="10" fill="#FFFFFF"/><text x="63" y="90" font-size="14" font-weight="900" fill="#DC2626">⚡</text><text x="160" y="85" font-size="13" font-weight="900" fill="#991B1B">한전 22.9kV 특고압 관로/맨홀 부설</text><text x="160" y="105" font-size="11" font-weight="bold" fill="#64748B">지중 매설 깊이 GL-1.5m (방호관 매설)</text></svg>'''
        s2 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-red-200"><rect width="480" height="160" fill="#FEF2F2"/><rect x="60" y="40" width="360" height="80" rx="8" fill="#FFFFFF" stroke="#DC2626" stroke-width="2"/><text x="110" y="70" font-size="13" font-weight="900" fill="#991B1B">CNCV-W 전력 케이블 인입 & 활선 작업</text><line x1="80" y1="90" x2="400" y2="90" stroke="#DC2626" stroke-width="4" stroke-dasharray="6,4"/><text x="140" y="110" font-size="11" font-weight="bold" fill="#DC2626">절연내력 30kV DC 10분간 연속 인가</text></svg>'''
        s3 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-red-200"><rect width="480" height="160" fill="#FEF2F2"/><circle cx="100" cy="80" r="35" fill="#EF4444"/><text x="80" y="85" font-size="12" font-weight="900" fill="#FFFFFF">한전 입회</text><rect x="180" y="45" width="260" height="70" rx="6" fill="#FFFFFF" stroke="#DC2626"/><text x="200" y="75" font-size="13" font-weight="900" fill="#991B1B">휴전(Cut-over) 스케줄 확정 및 전환</text><text x="200" y="98" font-size="11" font-weight="bold" fill="#0369A1">무정전 바이패스 전력망 가동</text></svg>'''
        s4 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-red-200"><rect width="480" height="160" fill="#FEF2F2"/><rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#DC2626" stroke-width="2"/><text x="165" y="45" font-size="13" font-weight="900" fill="#991B1B">특고압 전력 이설 준공 대장</text><text x="165" y="75" font-size="11" font-weight="bold" fill="#334155">한전 전력처 서면 인수 승인</text><circle cx="300" cy="115" r="18" fill="#FEE2E2" stroke="#DC2626" stroke-width="2"/><text x="287" y="120" font-size="11" font-weight="900" fill="#DC2626">승인 완료</text></svg>'''
    elif cat == 'gas':
        s1 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-amber-200"><rect width="480" height="160" fill="#FFFBEB"/><rect x="50" y="60" width="380" height="50" rx="6" fill="#FEF3C7" stroke="#D97706" stroke-width="2"/><text x="100" y="90" font-size="13" font-weight="900" fill="#B45309">도시가스 중압배관 이설구간 노선 측량</text></svg>'''
        s2 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-amber-200"><rect width="480" height="160" fill="#FFFBEB"/><circle cx="100" cy="80" r="30" fill="#F59E0B"/><text x="75" y="85" font-size="11" font-weight="900" fill="#FFFFFF">Hot Tapping</text><rect x="160" y="50" width="280" height="60" rx="6" fill="#FFFFFF" stroke="#D97706"/><text x="180" y="80" font-size="13" font-weight="900" fill="#B45309">무정지 천공 & Stopper 우회 가스관 개통</text></svg>'''
        s3 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-amber-200"><rect width="480" height="160" fill="#FFFBEB"/><rect x="40" y="40" width="400" height="80" rx="8" fill="#FFFFFF" stroke="#D97706"/><text x="120" y="75" font-size="13" font-weight="900" fill="#B45309">가스안전공사 1:1 입회 기밀/누출 시험</text><text x="140" y="100" font-size="11" font-weight="bold" fill="#059669">PID 가스검지기 0.00 ppm 측정 합격</text></svg>'''
        s4 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-amber-200"><rect width="480" height="160" fill="#FFFBEB"/><rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#D97706" stroke-width="2"/><text x="165" y="45" font-size="13" font-weight="900" fill="#B45309">도시가스 이설 준공 승인서</text><circle cx="300" cy="115" r="18" fill="#FEF3C7" stroke="#D97706" stroke-width="2"/><text x="287" y="120" font-size="11" font-weight="900" fill="#B45309">감리 승인</text></svg>'''
    elif cat == 'water':
        s1 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-cyan-200"><rect width="480" height="160" fill="#ECFEFF"/><rect x="40" y="55" width="400" height="60" rx="8" fill="#FFFFFF" stroke="#0891B2" stroke-width="2"/><text x="120" y="90" font-size="13" font-weight="900" fill="#0E7490">신설 주철 상수도관(D1000) 터파기 및 거치</text></svg>'''
        s2 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-cyan-200"><rect width="480" height="160" fill="#ECFEFF"/><circle cx="100" cy="80" r="30" fill="#06B6D4"/><text x="75" y="85" font-size="11" font-weight="900" fill="#FFFFFF">10kgf/cm²</text><rect x="160" y="50" width="280" height="60" rx="6" fill="#FFFFFF" stroke="#0891B2"/><text x="180" y="80" font-size="13" font-weight="900" fill="#0E7490">관로 1시간 수압 유지 시험 (누수량 0)</text></svg>'''
        s3 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-cyan-200"><rect width="480" height="160" fill="#ECFEFF"/><rect x="40" y="40" width="400" height="80" rx="8" fill="#FFFFFF" stroke="#0891B2"/><text x="110" y="75" font-size="13" font-weight="900" fill="#0E7490">CCTV 로봇 내시경 관로 내부 검사 & 수질 시험</text><text x="140" y="100" font-size="11" font-weight="bold" fill="#059669">탁도/잔류염소 음용수 수질 기준 적합</text></svg>'''
        s4 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-cyan-200"><rect width="480" height="160" fill="#ECFEFF"/><rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#0891B2" stroke-width="2"/><text x="175" y="45" font-size="13" font-weight="900" fill="#0E7490">상하수도 인계인수증</text><circle cx="300" cy="115" r="18" fill="#CFFAFE" stroke="#0891B2" stroke-width="2"/><text x="287" y="120" font-size="11" font-weight="900" fill="#0891B2">인계 완료</text></svg>'''
    elif cat == 'telecom':
        s1 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-violet-200"><rect width="480" height="160" fill="#F5F3FF"/><rect x="40" y="55" width="400" height="60" rx="8" fill="#FFFFFF" stroke="#7C3AED" stroke-width="2"/><text x="110" y="90" font-size="13" font-weight="900" fill="#5B21B6">통신 핸드홀/맨홀 및 다조관로 지중 포설</text></svg>'''
        s2 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-violet-200"><rect width="480" height="160" fill="#F5F3FF"/><circle cx="100" cy="80" r="30" fill="#8B5CF6"/><text x="75" y="85" font-size="11" font-weight="900" fill="#FFFFFF">광섬유 융착</text><rect x="160" y="50" width="280" height="60" rx="6" fill="#FFFFFF" stroke="#7C3AED"/><text x="180" y="80" font-size="13" font-weight="900" fill="#5B21B6">144심 광케이블 코어 정밀 융착 접속</text></svg>'''
        s3 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-violet-200"><rect width="480" height="160" fill="#F5F3FF"/><rect x="40" y="40" width="400" height="80" rx="8" fill="#FFFFFF" stroke="#7C3AED"/><text x="120" y="75" font-size="13" font-weight="900" fill="#5B21B6">OTDR 광파장(1550nm) 손실 감쇠 측정</text><text x="150" y="100" font-size="11" font-weight="bold" fill="#059669">접속 손실 ≤ 0.03dB (KT/SKT/LGU+ 통과)</text></svg>'''
        s4 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-violet-200"><rect width="480" height="160" fill="#F5F3FF"/><rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#7C3AED" stroke-width="2"/><text x="175" y="45" font-size="13" font-weight="900" fill="#5B21B6">통신관로 이설 승인서</text><circle cx="300" cy="115" r="18" fill="#EDE9FE" stroke="#7C3AED" stroke-width="2"/><text x="287" y="120" font-size="11" font-weight="900" fill="#7C3AED">통신 승인</text></svg>'''
    else:
        # Default / GPR / Traffic / Settlement
        s1 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-slate-200"><rect width="480" height="160" fill="#F8FAFC"/><rect x="40" y="55" width="400" height="60" rx="8" fill="#FFFFFF" stroke="#0284C7" stroke-width="2"/><text x="110" y="90" font-size="13" font-weight="900" fill="#0369A1">GPR 3D 관로 탐사 & 현장 기준점 측량</text></svg>'''
        s2 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-slate-200"><rect width="480" height="160" fill="#F8FAFC"/><polygon points="40,110 150,110 160,150 320,150 330,110 440,110 440,150 40,150" fill="#BAE6FD"/><text x="150" y="135" font-size="13" font-weight="900" fill="#0369A1">인력 시탐 줄파기 (폭 0.5m, 깊이 1.5m)</text></svg>'''
        s3 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-slate-200"><rect width="480" height="160" fill="#F8FAFC"/><rect x="40" y="55" width="400" height="60" rx="8" fill="#FFFFFF" stroke="#059669" stroke-width="2"/><text x="120" y="90" font-size="13" font-weight="900" fill="#047857">신설 이설 시공 & H-Beam 매달기 방호 거치</text></svg>'''
        s4 = f'''<svg viewBox="0 0 480 160" class="w-full h-auto rounded bg-white border border-slate-200"><rect width="480" height="160" fill="#F8FAFC"/><rect x="140" y="20" width="200" height="120" rx="6" fill="#FFFFFF" stroke="#059669" stroke-width="2"/><text x="165" y="45" font-size="13" font-weight="900" fill="#047857">지장물 이설 감리 서면 승인</text><circle cx="300" cy="115" r="18" fill="#DCFCE7" stroke="#059669" stroke-width="2"/><text x="287" y="120" font-size="11" font-weight="900" fill="#059669">승인 완료</text></svg>'''

    def wrap_zoom(svg_str, step_label):
        return f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-white border border-slate-200 rounded-xl p-3 mb-3 shadow-sm hover:shadow" onclick="openDiagramZoom(this.outerHTML, '{title} - {step_label} 2D 도식')">
            <div class="text-[11px] font-bold text-slate-500 mb-1 flex items-center justify-between">
                <span>🔍 클릭 시 대형 팝업 확대</span>
                <span class="text-blue-600 bg-blue-50 px-2 py-0.5 rounded border border-blue-200 font-extrabold">{step_label}</span>
            </div>
            {svg_str}
        </div>'''

    return wrap_zoom(s1, "STEP 01"), wrap_zoom(s2, "STEP 02"), wrap_zoom(s3, "STEP 03"), wrap_zoom(s4, "STEP 04")

def build_tailored_guideline(act):
    title = act['task_title']
    cat = categorize_activity(title)
    cfg = get_theme_config(cat)
    
    purpose = act['purpose']
    method = act['method']
    output = act['output']
    risk = act['risk']
    advisory = act['advisory']
    dept = act['dept']
    
    steps_raw = method.split('➔')
    steps = [s.strip() for s in steps_raw]
    while len(steps) < 4:
        steps.append(f"단계별 세부 승인 절차 {len(steps)+1}")
        
    svg1, svg2, svg3, svg4 = build_svg_diagrams_for_category(cat, title)
    
    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 지장물이설 - {title} 특화 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style> body {{ font-family: 'Noto Sans KR', sans-serif; }} </style>
</head>
<body class="bg-slate-100 text-slate-800 antialiased py-8 px-4 sm:px-8">

    <div class="max-w-6xl mx-auto bg-white rounded-3xl shadow-xl border border-slate-200 p-6 sm:p-10 space-y-8">
        <!-- 🔵 헤더 영역 (분야별 고유 뱃지) -->
        <div class="flex flex-col sm:flex-row sm:items-center justify-between pb-6 border-b border-slate-200 gap-4">
            <div>
                <span class="text-xs font-black px-3.5 py-1.5 rounded-full mb-2 inline-block border {cfg['badge']}">
                    Dongtan Tram Utility Special | {cfg['name']} (WBS {act['l3_code']})
                </span>
                <h1 class="text-2xl sm:text-4xl font-black text-slate-900 tracking-tight">
                    {title} 수행지침서
                </h1>
            </div>
            <div class="shrink-0 text-right text-xs font-bold text-slate-500">
                주관부서: <span class="text-indigo-600 font-extrabold">{dept}</span>
            </div>
        </div>

        <!-- 📌 1. 작업 개요 및 핵심 목적 -->
        <div class="{cfg['bg_box']} rounded-2xl p-6 space-y-3">
            <h3 class="text-base font-black {cfg['title_color']} flex items-center gap-2">
                <span>📌</span> 현장 이행 목적 및 엔지니어링 방침
            </h3>
            <p class="text-slate-800 text-sm font-semibold leading-relaxed">
                {purpose}
            </p>
        </div>

        <!-- 📡 2. 특화 인터랙티브 공학 시뮬레이터 장치 -->
        {cfg['sim_html']}

        <!-- 💡 3. 4단계 세부 수행절차 및 2D 공학 도식 (Light Theme) -->
        <div class="space-y-6">
            <h3 class="text-xl font-black text-slate-900 flex items-center gap-2">
                <span>💡</span> 4단계 세부 공학 수행지침 및 2D 시공 도식
            </h3>

            <div class="grid grid-cols-1 md:grid-cols-2 gap-6">
                <!-- STEP 01 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg1}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 01</span>
                        <h4 class="font-bold text-base text-slate-900">사전 검토 및 현장 측량</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[0]}</p>
                </div>

                <!-- STEP 02 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg2}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 02</span>
                        <h4 class="font-bold text-base text-slate-900">인력 시탐 줄파기 & 점용 입회</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[1]}</p>
                </div>

                <!-- STEP 03 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg3}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 03</span>
                        <h4 class="font-bold text-base text-slate-900">신설 이설 시공 & 방호 거치</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[2]}</p>
                </div>

                <!-- STEP 04 -->
                <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                    {svg4}
                    <div class="flex items-center gap-2 mb-2">
                        <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP 04</span>
                        <h4 class="font-bold text-base text-slate-900">최종 검측 성과표 & 서면 승인</h4>
                    </div>
                    <p class="text-slate-700 text-sm leading-relaxed font-medium">{steps[3]}</p>
                </div>
            </div>
        </div>

        <!-- 📦 4. 최종 성과품 -->
        <div class="bg-emerald-50 border border-emerald-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-emerald-950 flex items-center gap-2">
                <span>📂</span> 최종 필수 성과품 / 결재 대장
            </h3>
            <p class="text-emerald-900 text-sm font-bold leading-relaxed">
                {output}
            </p>
        </div>

        <!-- ⚠️ 5. 집행단계 리스크 예방 관리 -->
        <div class="bg-amber-50 border-2 border-amber-300 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>⚠️</span> 집행단계 핵심 리스크 및 예방 대책
            </h3>
            <p class="text-amber-900 text-sm font-bold leading-relaxed">
                {risk}
            </p>
        </div>

        <!-- 👷 6. 협력업체 실무 자문 노하우 -->
        <div class="bg-amber-50/60 border border-amber-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>👷</span> [협력업체 실무 자문] 시공 및 공사관리 핵심 가이드
            </h3>
            <p class="text-amber-900 text-sm font-semibold leading-relaxed">
                {advisory}
            </p>
        </div>

        <!-- 📌 푸터 -->
        <div class="pt-6 border-t border-slate-200 flex items-center justify-between text-xs text-slate-500">
            <span>🏢 동탄도시철도(트램) 건설공사 {cfg['name']} 수행지침</span>
            <span>v8 엑셀 100% 1:1 특화 연동 데이터</span>
        </div>
    </div>

    <!-- Lightbox Zoom Modal -->
    <div id="zoomModal" class="fixed inset-0 z-[9999] hidden bg-black/80 backdrop-blur-md flex items-center justify-center p-4" onclick="this.classList.add('hidden')">
        <div class="bg-white rounded-3xl p-6 max-w-4xl w-full max-h-[90vh] overflow-auto border-4 border-indigo-500 shadow-2xl relative" onclick="event.stopPropagation()">
            <button onclick="document.getElementById('zoomModal').classList.add('hidden')" class="absolute top-4 right-4 bg-slate-900 text-white font-black px-4 py-2 rounded-full text-sm hover:bg-slate-700">✕ 닫기</button>
            <h3 id="zoomTitle" class="text-xl font-black text-slate-900 mb-4 pb-2 border-b border-slate-200">고해상도 2D 시공 도식 확대보기</h3>
            <div id="zoomContent" class="flex justify-center items-center"></div>
        </div>
    </div>
    <script>
    function openDiagramZoom(htmlContent, title) {{
        var modal = document.getElementById('zoomModal');
        document.getElementById('zoomTitle').innerText = title || '2D 시공 도식 확대보기';
        document.getElementById('zoomContent').innerHTML = htmlContent;
        modal.classList.remove('hidden');
    }}
    </script>
</body>
</html>'''
    return html

# Generate tailored HTML files across all matching subfolders in v8/지장물이설
updated_folders = 0
for act in activities:
    title = act['task_title']
    matching_folders = []
    
    if os.path.exists(util_dir):
        for sub in os.listdir(util_dir):
            sub_p = os.path.join(util_dir, sub)
            if os.path.isdir(sub_p):
                if title in sub or sub.split('_', 1)[-1] in title or title in sub.split('_', 1)[-1]:
                    matching_folders.append(sub_p)
                    
        if not matching_folders:
            kw = title.split()[0]
            for sub in os.listdir(util_dir):
                sub_p = os.path.join(util_dir, sub)
                if os.path.isdir(sub_p) and kw in sub:
                    matching_folders.append(sub_p)
                    
    html_content = build_tailored_guideline(act)
    safe_title = re.sub(r'[\/:*?"<>|]', '_', title)
    
    for mf in matching_folders:
        guide_dir = os.path.join(mf, '수행지침')
        os.makedirs(guide_dir, exist_ok=True)
        
        f1 = os.path.join(guide_dir, f"{safe_title}_수행지침.html")
        f2 = os.path.join(guide_dir, "수행지침.html")
        
        with open(f1, 'w', encoding='utf-8') as f:
            f.write(html_content)
        with open(f2, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        for existing in os.listdir(guide_dir):
            if existing.endswith('.html'):
                with open(os.path.join(guide_dir, existing), 'w', encoding='utf-8') as f:
                    f.write(html_content)
        updated_folders += 1

print(f"Successfully generated category-tailored HTML manuals for {updated_folders} 지장물이설 folders.", flush=True)
