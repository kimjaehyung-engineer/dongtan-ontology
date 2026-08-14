# -*- coding: utf-8 -*-
import openpyxl, os, sys, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')

wb = openpyxl.load_workbook(v8_excel, data_only=False)

total_links = 0
found_count = 0
missing_list = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            target = None
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                val = cell.value
                parts = val.split('"')
                if len(parts) >= 2:
                    target = parts[1]
            
            if target:
                total_links += 1
                decoded_target = urllib.parse.unquote(target)
                abs_path = os.path.normpath(os.path.join(v8_root, decoded_target))
                if os.path.exists(abs_path):
                    found_count += 1
                else:
                    missing_list.append((sheet_name, r, c, target, decoded_target, abs_path))

print(f"Total Excel Hyperlinks: {total_links}")
print(f"Hyperlinks Valid & File Exists on Disk: {found_count} / {total_links}")
print(f"Broken Hyperlinks: {len(missing_list)}")

print("\nSample Broken Hyperlinks (first 20):")
for item in missing_list[:20]:
    sheet, r, c, raw_t, dec_t, abs_p = item
    print(f"[{sheet}] R{r}C{c} | Raw: {raw_t}")
    print(f"    -> Decoded: {dec_t}")
    dir_p = os.path.dirname(abs_p)
    if os.path.exists(dir_p):
        actual_files = os.listdir(dir_p)
        print(f"    -> Directory exists! Actual files in folder: {actual_files}")
    else:
        print(f"    -> Directory DOES NOT EXIST: {dir_p}")
