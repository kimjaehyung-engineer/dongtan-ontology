# -*- coding: utf-8 -*-
import openpyxl, os, sys, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')

wb = openpyxl.load_workbook(v8_excel, data_only=False)

fixed_count = 0
already_valid = 0
not_found_folders = []

print("=== Repairing Excel Hyperlinks in v8 ===")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            target = None
            is_formula = False
            
            if cell.hyperlink and cell.hyperlink.target:
                target = cell.hyperlink.target
            elif cell.value and isinstance(cell.value, str) and 'HYPERLINK' in cell.value:
                is_formula = True
                val = cell.value
                parts = val.split('"')
                if len(parts) >= 2:
                    target = parts[1]
            
            if target:
                decoded_target = urllib.parse.unquote(target)
                abs_path = os.path.normpath(os.path.join(v8_root, decoded_target))
                
                if os.path.exists(abs_path):
                    already_valid += 1
                else:
                    # Target file missing, find actual directory
                    dir_p = os.path.dirname(abs_path)
                    if os.path.exists(dir_p):
                        files = [f for f in os.listdir(dir_p) if not f.startswith('.') and not f.startswith('~$')]
                        if files:
                            # Pick the single html file in this folder
                            actual_file = files[0]
                            rel_dir = os.path.dirname(decoded_target)
                            new_rel_path = os.path.join(rel_dir, actual_file).replace('\\', '/')
                            
                            # Update cell hyperlink or formula
                            if cell.hyperlink:
                                cell.hyperlink.target = new_rel_path
                            if is_formula:
                                # Rebuild HYPERLINK("new_rel_path", "display_text")
                                parts = cell.value.split('"')
                                parts[1] = new_rel_path
                                cell.value = '"'.join(parts)
                            
                            fixed_count += 1
                        else:
                            not_found_folders.append((sheet_name, r, c, target, dir_p, "No files in dir"))
                    else:
                        not_found_folders.append((sheet_name, r, c, target, dir_p, "Dir missing"))

print(f"Already Valid Hyperlinks: {already_valid}")
print(f"Successfully Fixed Hyperlinks: {fixed_count}")
print(f"Unresolvable Hyperlinks: {len(not_found_folders)}")

if not_found_folders:
    print("\nUnresolvable Details:")
    for item in not_found_folders[:10]:
        print("  ", item)

# Save updated excel workbook
wb.save(v8_excel)
print(f"\nSaved updated Excel file: {v8_excel}")
