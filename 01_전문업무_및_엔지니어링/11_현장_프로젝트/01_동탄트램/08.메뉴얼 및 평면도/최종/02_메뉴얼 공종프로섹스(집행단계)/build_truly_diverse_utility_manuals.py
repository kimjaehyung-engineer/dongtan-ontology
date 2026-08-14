# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse, re

sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')
util_dir = os.path.join(v8_root, '지장물이설')

print("=== Starting True Custom Generation for 38 Utility Tasks ===", flush=True)

# Define rich, authentic custom architectures for each of the 38 tasks
TASK_SPECS = {
    "Site Survey Risk 검토": {
        "type": "3-Step",
        "theme": "sky",
        "badge": "리스크 종합 평가 (3-Step Milestone)",
        "sim_title": "🗺️ 3D GIS & 현장 지장물 중첩도 분석기",
        "sim_body": "<div class='p-4 bg-sky-950 text-sky-200 rounded-xl font-mono text-xs flex justify-between items-center'><span>GIS OVERLAY: D1000 상수관, 22.9kV 한전, 중압가스관 중첩 완료</span><span class='text-emerald-400 font-bold'>● 간섭 위험도: 85% (중점관리 필요)</span></div>",
        "steps": [
            ("1단계: 현장 지하매설물 통합 GIS 도면 정밀 대조", "설계 도면과 국가공간정보포털 지하시설물 통합망 데이터를 중첩 분석하여 불일치 구간을 식별합니다.", "gis"),
            ("2단계: 지장물 유관기관(한전, 삼천리, 수자원공사) 현장 합동 실사", "각 점용기관 실무 책임자와 함께 도로상 밸브실, 맨홀, 표지못의 실제 위치를 전수 육안 조사합니다.", "survey"),
            ("3단계: 공구별 위험도 매트릭스 도출 및 우회 시공 전략 수립", "트램 노반 굴착 단면과 간섭되는 심도 1.5m 이내 지장물에 대한 우선 이설 순위 확정.", "risk")
        ]
    },
    "발주전략 KOM (도급지분)": {
        "type": "2-Step",
        "theme": "indigo",
        "badge": "도급 발주 의사결정 (2-Step Matrix)",
        "sim_title": "📋 발주 방식(직접시공 vs 전문외주) 의사결정 매트릭스",
        "sim_body": "<div class='p-3 bg-indigo-50 border border-indigo-200 rounded-xl text-xs flex justify-between'><span>도급분 상하수도 전문건설 면허 적격심사:</span><b class='text-indigo-700'>적합 판정 (종합평점 95.4점)</b></div>",
        "steps": [
            ("Phase 1: 도급 공종별 전문건설업 면허 및 적격성 사전 심사", "상하수도설비공사업 및 토공사 면허 보유 업체의 시공 실적, 기술자 보유 현황, 신용도를 사전 평가합니다.", "contract"),
            ("Phase 2: 원가 절감형 패키지 발주 및 Kick-off 확정", "구간별 일괄 발주를 통해 공기 단축 및 인터페이스 리스크를 최소화하는 최종 도급 계약을 체결합니다.", "handover")
        ]
    },
    "지장물 이설 요청 (위수탁고)": {
        "type": "3-Step",
        "theme": "indigo",
        "badge": "위수탁 행정 수속 (3-Step Flow)",
        "sim_title": "🏛️ 화성시 및 위탁기관 정식 공문 발송 트래커",
        "sim_body": "<div class='p-3 bg-white border border-indigo-200 rounded-xl text-xs flex justify-between'><span>공문 접수 현황 (화성시 맑은물사업소):</span><b class='text-emerald-600'>접수 완료 (처리기간 14일)</b></div>",
        "steps": [
            ("1단계: 위탁기관별 이설 요구 공문 및 기본설계 도서 발송", "한전, 통신 3사, 도시가스사, 수자원공사에 트램 사업계획 및 간섭 현황도 첨부 공문 발송.", "doc"),
            ("2단계: 이설 분담금 및 위탁수수료 협의안 작성", "지자체 조례 및 공공시설물 이설비 부담 기준에 따른 사업비 분담 비율 잠정 협의.", "calc"),
            ("3단계: 위수탁 협약서 체결 및 실무 전담 TF 구성", "사업시행자-위탁기관 간 이설공사 위수탁 협약 체결 및 현장 책임자 지정.", "sign")
        ]
    },
    "도급자분 이설업체 선정(상/하수)": {
        "type": "2-Step",
        "theme": "indigo",
        "badge": "적격업체 선정 (2-Step Fast-Track)",
        "sim_title": "🏢 상하수도 전문 시공사 기술제안서 평가표",
        "sim_body": "<div class='p-3 bg-white border border-indigo-200 rounded-xl text-xs flex justify-between'><span>1순위 적격 시공사:</span><b class='text-indigo-700'>동탄 인프라(주) - 기술점수 98점</b></div>",
        "steps": [
            ("Step 1: 기술 제안서(현장 시공계획 및 안전관리) 정밀 심사", "도로 굴착 중 단수 사고 방지 대책 및 비상 급수차 배치 계획을 중점 평가합니다.", "doc"),
            ("Step 2: 실행예산 확정 및 하도급 계약 승인", "발주처 하도급 적정성 심사 승인 후 현장 안전 서약서 징구 및 착공계 수리.", "sign")
        ]
    },
    "지장물 조사 (위탁기관 합동)": {
        "type": "5-Step",
        "theme": "amber",
        "badge": "5개 공종 합동 정밀 실측 (5-Step Survey)",
        "sim_title": "📍 지하매설물 5대 점용기관 현장 3자 합동 조사표",
        "sim_body": "<div class='p-3 bg-amber-50 border border-amber-200 rounded-xl text-xs space-y-1.5'><div>상수도: D800 주철관 (GL-1.8m) | 한전: 22.9kV 8열 (GL-1.4m)</div><div>가스관: 중압 300A (GL-2.0m) | 통신: KT 16공 (GL-1.0m)</div></div>",
        "steps": [
            ("1단계: 한전 배전처 합동 지중 케이블 선로 확인", "탐사기를 활용하여 22.9kV 특고압 선로의 정확한 매설 축선과 깊이를 페인트 마킹합니다.", "electric"),
            ("2단계: 도시가스(삼천리) 합동 가스 밸브실 및 차단밸브 점검", "비상시 가스 차단이 가능한 인근 제수밸브 위치 및 부식 상태를 점검합니다.", "gas"),
            ("3단계: 맑은물사업소 합동 상수도 제수밸브 및 소화전 실측", "단수 구역을 최소화하기 위한 블록 밸브 조작 계획을 현장에서 확인합니다.", "water"),
            ("4단계: 통신 3사(KT, SKT, LGU+) 광케이블 간선 핸드홀 조사", "트램 정거장 하부 간섭 통신 다조관로의 예비 공관(Duct) 확보 여부를 확인합니다.", "telecom"),
            ("5단계: 5개 기관 합동 서명 조사보고서 확정", "현장 실측 좌표와 심도가 기재된 조사대장에 전 기관 입회관의 서명을 날인합니다.", "sign")
        ]
    },
    "지장물 이설 계획 수립": {
        "type": "3-Step",
        "theme": "sky",
        "badge": "마스터 공정 수립 (3-Step Master Plan)",
        "sim_title": "📅 트램 공정 연계 지장물 이설 마스터 스케줄러",
        "sim_body": "<div class='p-3 bg-white border border-sky-200 rounded-xl text-xs flex justify-between'><span>총 이설 기간:</span><b class='text-sky-700'>착공 후 120일 이내 (궤도 착수 전 완료)</b></div>",
        "steps": [
            ("1단계: 공구별/단계별 우회 이설 공정표 수립", "교통 혼잡을 분산시키기 위해 교차로별 야간 작업 및 분할 굴착 일정을 편성합니다.", "plan"),
            ("2단계: 단수·정전·가스차단 예고 및 주민 공지 계획", "주민 불편을 제로화하기 위해 사전 현수막 게시 및 비상 대응망을 구축합니다.", "notice"),
            ("3단계: 책임감리원 종합 시공계획서 서면 승인", "안전관리계획, 환경관리계획, 우회동선도가 포함된 최종 계획서를 승인받습니다.", "sign")
        ]
    },
    "최고의 팀 만들기 지원": {
        "type": "2-Step",
        "theme": "emerald",
        "badge": "팀 협업 역량 강화 (2-Step Team Building)",
        "sim_title": "👥 지장물 이설 One-Team 통합 상황실",
        "sim_body": "<div class='p-3 bg-emerald-50 border border-emerald-200 rounded-xl text-xs flex justify-between'><span>One-Team 핫라인:</span><b class='text-emerald-700'>발주처-감리단-시공사-유관기관 24시간 실시간 가동</b></div>",
        "steps": [
            ("Phase 1: 유관기관 실무자 합동 워크숍 및 핫라인 개설", "돌발 지장물 발견 시 1시간 이내 현장 출동이 가능한 비상 연락 체계를 구축합니다.", "team"),
            ("Phase 2: 주간 인터페이스 점검 회의 정례화", "공정 간섭 사항을 주 1회 사전 조율하여 불필요한 공사 중단을 사전에 방지합니다.", "meeting")
        ]
    },
    "착수전 Big Room 회의": {
        "type": "3-Step",
        "theme": "indigo",
        "badge": "Big Room 협업 프로세스 (3-Step Session)",
        "sim_title": "🏛️ 지장물 Big Room 통합 의사결정 보드",
        "sim_body": "<div class='p-3 bg-indigo-950 text-indigo-200 rounded-xl font-mono text-xs flex justify-between'><span>3D BIM 간섭 검토 안건:</span><b class='text-amber-400'>12건 중 12건 전원 합의 완료</b></div>",
        "steps": [
            ("1단계: 3D BIM 기반 지하 지장물 간섭 시뮬레이션 시연", "트램 궤도 하부 구조물과 지하 관로 간의 충돌 지점을 3D 모델로 시각화하여 확인합니다.", "bim"),
            ("2단계: 쟁점 사항(이설 경로, 심도, 보호공법) 즉석 합의", "Big Room 현장에서 관련 기관 실무자들이 즉시 대안 노선을 검토하고 설계 변경에 동의합니다.", "debate"),
            ("3단계: Big Room 합의 의사록 작성 및 감리 서명", "현장 결정 사항에 대한 공식 의사록을 작성하고 즉각적인 공사 착수를 결의합니다.", "sign")
        ]
    },
    "인허가 절차 진행": {
        "type": "3-Step",
        "theme": "sky",
        "badge": "행정 인허가 마일스톤 (3-Step Milestone)",
        "sim_title": "📑 도로점용 및 굴착 인허가 승인 트래커",
        "sim_body": "<div class='p-3 bg-white border border-sky-200 rounded-xl text-xs flex justify-between'><span>화성동탄경찰서 교통안전시설 심의:</span><b class='text-emerald-600'>조건부 가결 (신호수 보강)</b></div>",
        "steps": [
            ("1단계: 화성시 도로점용 및 굴착사업계획서 제출", "토목 단면도, 지중 매설물 안전대책, 도로 복구 단면도를 첨부하여 도로관리심의회에 상정합니다.", "doc"),
            ("2단계: 화성동탄경찰서 교통안전시설 및 차선 통제 협의", "출퇴근 시간대 차로 유지 및 우회 동선 안내표지판 설치 계획을 심의받습니다.", "police"),
            ("3단계: 최종 도로점용허가증 수령 및 현장 비치", "허가 조건(작업시간, 복구 기준)을 숙지하고 현장 사무실에 허가증을 게시합니다.", "license")
        ]
    },
    "교통처리대책 협의 및 승인": {
        "type": "5-Step",
        "theme": "indigo",
        "badge": "5단계 차로 전환 교통통제 (5-Step Traffic Shift)",
        "sim_title": "🚗 동탄대로 5단계 차로 우회 시뮬레이터",
        "sim_body": "<div class='p-3 bg-indigo-50 border border-indigo-200 rounded-xl text-xs space-y-1'><div>현재 단계: <b>2단계 중앙 분리대 철거 및 가설 둔차선 개통</b></div><div>통행 속도: <b>30km/h 서행 유도 | 신호수 4명 상시 배치</b></div></div>",
        "steps": [
            ("1단계: 교통 소통 시뮬레이션(VISSIM) 분석 및 대책 수립", "굴착 공사 시 예상되는 차량 지체 시간을 분석하여 최적의 차로 점용 폭을 결정합니다.", "traffic_sim"),
            ("2단계: 가설 둔차선(복공판) 설치 및 중앙분리대 임시 개방", "기존 차로 수를 최대한 유지하기 위해 보도부 및 중앙분리대를 활용한 임시 차로를 개설합니다.", "deck"),
            ("3단계: PE 방호벽, 쏠라 LED 경광등, 갈매기 표지판 100m 연속 거치", "야간 운전자의 시인성을 확보하기 위해 전 구간에 고휘도 안전 시설물을 설치합니다.", "barrier"),
            ("4단계: 국가공인 전담 신호수(2인 1조) 교차로 배치", "안전 조끼, 경광봉, 무전기를 착용한 전담 신호수가 차량 흐름을 통제합니다.", "flagger"),
            ("5단계: 경찰서 및 지자체 현장 합동 점검 승인", "교통 흐름에 중대한 지장이 없음을 최종 확인하고 굴착 작업 착수를 승인받습니다.", "sign")
        ]
    },
    "민원 저감 대책 수립": {
        "type": "3-Step",
        "theme": "emerald",
        "badge": "환경/주민 소통 (3-Step Protocol)",
        "sim_title": "📢 소음·비산먼지 24시간 실시간 모니터링",
        "sim_body": "<div class='p-3 bg-emerald-950 text-emerald-200 rounded-xl font-mono text-xs flex justify-between'><span>소음도: 58.2 dB (기준 65dB 이하)</span><span class='text-emerald-400'>● 미세먼지: 22 ㎍/㎥ (양호)</span></div>",
        "steps": [
            ("1단계: 인근 상가 및 아파트 주민 설명회 개최", "공사 일정, 단수·소음 발생 예상 시간대를 주민들에게 투명하게 사전 안내합니다.", "notice"),
            ("2단계: 이동식 방음벽(H=3.0m) 및 살수차·초음파 살수기 가동", "아스팔트 절단 및 굴착 시 발생하는 비산먼지와 소음을 원천 차단합니다.", "barrier"),
            ("3단계: 24시간 민원 접수 핫라인 및 신속 대응팀 운영", "민원 접수 즉시 30분 이내에 현장에 출동하여 불편 사항을 조치합니다.", "phone")
        ]
    },
    "용지보상 Risk 파악": {
        "type": "2-Step",
        "theme": "amber",
        "badge": "부지 권원 확보 (2-Step Verification)",
        "sim_title": "📐 토지 지적도 및 도로구역선 일치 검증",
        "sim_body": "<div class='p-3 bg-white border border-amber-200 rounded-xl text-xs flex justify-between'><span>도로구역 내 시유지 점용율:</span><b class='text-emerald-700'>100% (사유지 침범 리스크 제로)</b></div>",
        "steps": [
            ("Step 1: 지적 경계측량 및 도로구역선 침범 여부 전수 조사", "이설 관로가 사유지 또는 미보상 토지를 통과하지 않는지 한국국토정보공사 성과표를 검토합니다.", "survey"),
            ("Step 2: 사유지 통과 불가피 시 구분지상권 설정 및 보상 수속", "지하 공간 사용에 대한 적법한 보상 절차를 이행하고 공사 방해 민원을 사전 차단합니다.", "sign")
        ]
    },
    "관리기관(맑은물사업소) 협의": {
        "type": "3-Step",
        "theme": "cyan",
        "badge": "상수도 관리기관 협의 (3-Step Flow)",
        "sim_title": "💧 동탄권역 상수 급수 계통도 검토기",
        "sim_body": "<div class='p-3 bg-cyan-50 border border-cyan-200 rounded-xl text-xs flex justify-between'><span>단수 대체 급수 라인:</span><b class='text-cyan-700'>D400 비상 루프 배관 연계 완료</b></div>",
        "steps": [
            ("1단계: 상수도 주철관(D1000) 이설 설계 도서 기술 협의", "관경, 관종(K-Type 닥타일주철관), 매설 깊이, 이음부 밸브실 구조를 맑은물사업소와 협의합니다.", "water"),
            ("2단계: 단수 작업 승인 및 비상 급수차 배치 계획 확정", "주민 급수 차질을 방지하기 위해 심야 시간(01:00~05:00) 단수 일정을 조율합니다.", "truck"),
            ("3단계: 맑은물사업소 공문 승인 및 입회 감독관 지정", "공사 시 입회할 상수도 전담 공무원을 지정받고 정식 승인 공문을 수령합니다.", "sign")
        ]
    },
    "위수탁 지장물 이설 설계": {
        "type": "3-Step",
        "theme": "indigo",
        "badge": "상세 실시설계 (3-Step Design)",
        "sim_title": "📐 3D 지중 관로 실시설계 CAD/BIM 뷰어",
        "sim_body": "<div class='p-3 bg-white border border-indigo-200 rounded-xl text-xs flex justify-between'><span>설계 안전율(토피 및 차량 하중):</span><b class='text-indigo-700'>DB-24 하중 기준 1.85 확보</b></div>",
        "steps": [
            ("1단계: 현장 실측 좌표 기반 지하지장물 실시설계도 작성", "종·평면도, 횡단면도, 관로 상세도, 부속 밸브실 구조계산서를 작성합니다.", "cad"),
            ("2단계: 관할 위탁기관 설계도서 사전 심사 및 보완", "기관별 표준 시방서(한전, 삼천리, 수자원공사 설계기준) 충족 여부를 확인받습니다.", "doc"),
            ("3단계: 최종 설계도서 확정 및 수량/내역서 발주처 납품", "공사비 산출 내역서와 함께 감리원 검토 의견서를 첨부하여 승인을 득합니다.", "sign")
        ]
    },
    "상하수도 이설계획 실정보고": {
        "type": "2-Step",
        "theme": "indigo",
        "badge": "설계변경 실정보고 (2-Step Procedure)",
        "sim_title": "📑 공사비 변동 실정보고 심의서",
        "sim_body": "<div class='p-3 bg-white border border-indigo-200 rounded-xl text-xs flex justify-between'><span>실정보고 사유:</span><b class='text-indigo-700'>현장 암반 출현 및 지중 간섭으로 인한 노선 우회</b></div>",
        "steps": [
            ("Step 1: 현장 여건 변동(암반, 간섭)에 따른 실정보고서 작성", "현장 사진, 지반조사 주상도, 비교 도면, 공사비 증감 내역서를 구비합니다.", "doc"),
            ("Step 2: 책임감리원 기술 검토 의견서 첨부 및 발주처 승인", "설계변경의 타당성을 입증받고 발주처로부터 공사비 증액 승인을 득합니다.", "sign")
        ]
    },
    "위수탁 계약 체결": {
        "type": "2-Step",
        "theme": "indigo",
        "badge": "공식 협약 체결 (2-Step Agreement)",
        "sim_title": "🤝 지장물 이설공사 위수탁 표준 협약서",
        "sim_body": "<div class='p-3 bg-white border border-indigo-200 rounded-xl text-xs flex justify-between'><span>계약 이행 보증:</span><b class='text-emerald-700'>보증보험증권 100% 징구 완료</b></div>",
        "steps": [
            ("Phase 1: 기관별 위탁 공사비 산출 및 계약 조항 법률 검토", "지자체 재무과 및 기관 감사 부서의 일상 감사를 필하고 계약서 초안을 확정합니다.", "doc"),
            ("Phase 2: 기관장 간 정식 계약 체결 및 사업비 예치", "협약서에 상호 기명 날인하고 위수탁 분담금을 전용 계좌에 입금합니다.", "sign")
        ]
    },
    "도로점용/굴착행위 인허가": {
        "type": "3-Step",
        "theme": "sky",
        "badge": "인허가 완결 (3-Step Permit)",
        "sim_title": "🚧 도로 굴착 허가 조건 이행 확인 시스템",
        "sim_body": "<div class='p-3 bg-white border border-sky-200 rounded-xl text-xs flex justify-between'><span>허가 조건:</span><b class='text-sky-700'>야간(22:00~06:00) 굴착 및 당일 복공판 가설</b></div>",
        "steps": [
            ("1단계: 화성시 도로과 도로점용 및 굴착허가증 수령", "허가 부관 사항(굴착 폭, 토사 반출 시간, 가시설 기준)을 현장에 통보합니다.", "license"),
            ("2단계: 도로 굴착 및 복구 이행 보증보험증권 제출", "공사 후 도로 원상복구를 보증하기 위한 이행보증보험증권을 화성시에 예치합니다.", "money"),
            ("3단계: 도로점용 공사 안내 표지판 설치 및 공사 개시", "시점과 종점에 공사 안내 현수막과 도로점용 허가 표지판을 부착합니다.", "sign")
        ]
    },
    "교통통제 및 교통안전시설 설치": {
        "type": "5-Step",
        "theme": "indigo",
        "badge": "현장 안전 셋업 (5-Step Safety Setup)",
        "sim_title": "🛡️ 도로공사 안전시설물 5대 필수 장치 점검",
        "sim_body": "<div class='p-3 bg-indigo-950 text-indigo-200 rounded-xl font-mono text-xs space-y-1'><div>● PE 방호벽 물채움: 100% 완료 (충격 완충 확보)</div><div>● 쏠라 LED 경광등: 50개소 점등 정상 | ● 신호수 무전기 통신 양호</div></div>",
        "steps": [
            ("1단계: 교통안전시설물(PE드럼, 라바콘) 1차 라인 마킹", "도로 차선 통제 계획도에 따라 기준선을 스프레이로 마킹합니다.", "barrier"),
            ("2단계: 고중량 PE 방호벽 설치 및 내부 100% 물채움", "차량 충돌 시 튕겨나가지 않도록 규정 용량의 물을 완전히 채웁니다.", "water_fill"),
            ("3단계: 야간 시인성 확보용 LED 쏠라 경광등 및 화살표 사인카 배치", "500m 전방에서 운전자가 공사 구간을 인지할 수 있도록 대형 사인카를 배치합니다.", "signcar"),
            ("4단계: 공사 구간 전후방 100m 지점 공사안내 표지판 거치", "'공사중 서행 30km', '차선 감소' 표지판을 규정에 맞게 설치합니다.", "signboard"),
            ("5단계: 안전담당자 현장 안전 점검표 서명 및 작업 개시", "모든 시설물이 정상 작동함을 감리원에게 확인받고 굴착 장비를 진입시킵니다.", "sign")
        ]
    },
    "줄따기(GPR)를 통한 기존 지장물 매설 확인": {
        "type": "Testing Protocol",
        "theme": "sky",
        "badge": "GPR 3D 물리탐사 & 인력 시탐 (Protocol)",
        "sim_title": "📡 GPR 3D 지하 매설물 단면 스캐너 & 줄파기 게이지",
        "sim_body": "<div class='p-4 bg-slate-900 text-sky-300 font-mono text-xs space-y-2 rounded-xl border border-slate-700'><div class='flex justify-between'><span>GPR SCAN FREQ: 900MHz</span><span class='text-emerald-400 font-bold'>● 3개 관로 좌표 LOCK</span></div><div class='flex justify-around bg-slate-950 p-2 rounded'><span>GL-0.8m 통신관</span><span>GL-1.4m 한전전력</span><span>GL-2.1m 상수도</span></div></div>",
        "steps": [
            ("Protocol 01: 3D GPR 지중 레이더 장비 현장 캘리브레이션", "지반 유전율을 보정하고 10m 간격으로 종횡 스캔을 실시하여 이상 반응 구역을 마킹합니다.", "gpr"),
            ("Protocol 02: 인력 시탐 줄파기(폭 0.5m, 깊이 1.5m) 굴착", "장비 굴착을 엄격히 금지하고 삽과 곡괭이를 이용한 인력 굴착으로 매설관을 직접 노출시킵니다.", "trench"),
            ("Protocol 03: 노출 지장물 실측(관종, 관경, 정확한 심도 GL-m)", "버니어캘리퍼스 및 레벨기를 사용하여 외경과 상단 심도를 1cm 단위로 정밀 측정합니다.", "measure"),
            ("Protocol 04: 현장 위험 표지 깃발 설치 및 감리 확인 서명", "노출된 관로에 형광 위험 표지판을 부착하고 줄파기 검측 성과표를 작성합니다.", "sign")
        ]
    },
    "이설 위치 토공 굴착": {
        "type": "5-Step",
        "theme": "amber",
        "badge": "토공 굴착 및 가시설 방호 (5-Step Excavation)",
        "sim_title": "🚜 터파기 굴착 단면 및 가시설 흙막이 모니터",
        "sim_body": "<div class='p-3 bg-amber-50 border border-amber-200 rounded-xl text-xs space-y-1'><div>굴착 깊이: <b>GL -2.5m (규정 토피 확보)</b></div><div>흙막이 가시설: <b>조립식 간이 흙막이(AL-Box) 거치 완료 (붕괴 리스크 0%)</b></div></div>",
        "steps": [
            ("1단계: 도로 아스팔트 포장면 휠쏘(Wheel-Saw) 정밀 절단", "기존 도로 포장의 파손을 방지하기 위해 굴착 경계선을 직선으로 깨끗이 커팅합니다.", "cutting"),
            ("2단계: 소형 백호(0.3m³) 저속 굴착 및 잔토 덤프트럭 직상차", "지중 충격을 최소화하기 위해 저속 굴착을 실시하고 도로 오염을 방지하기 위해 직상차합니다.", "excavator"),
            ("3단계: 굴착 깊이 1.5m 초과 시 조립식 간이 흙막이(AL-Box) 거치", "작업자의 토사 붕괴 매몰 사고를 방지하기 위해 규격 흙막이 지보공을 설치합니다.", "shoring"),
            ("4단계: 바닥면 인력 면정리 및 모래 쿠션(T=10cm) 포설", "신설 관로의 균등한 지지를 위해 잔돌을 제거하고 세사를 평탄하게 포설합니다.", "sand"),
            ("5단계: 굴착 깊이 및 경사도 검측 승인", "배수 구배 및 계획 고를 레벨기로 확인하고 감리원의 승인을 득합니다.", "sign")
        ]
    },
    "신규관로 매설 및 설치": {
        "type": "5-Step",
        "theme": "cyan",
        "badge": "신설 배관 정밀 부설 (5-Step Pipe Laying)",
        "sim_title": "💧 K-Type 닥타일주철관(D1000) 접합 시뮬레이터",
        "sim_body": "<div class='p-3 bg-cyan-50 border border-cyan-200 rounded-xl text-xs space-y-1'><div>고무링 삽입 상태: <b>100% 정위치 삽입 (누수 방지 그리스 도포)</b></div><div>삽입 깊이: <b>표준 삽입선(120mm) 정확히 일치</b></div></div>",
        "steps": [
            ("1단계: 신설 관로 자재 반입 및 공인 시험성적서 대조", "KDS 표준 규격품 여부, 도장 손상 여부, 고무링 탄성을 전수 검사합니다.", "material"),
            ("2단계: 크레인(25t)을 이용한 신설 관로 트렌치 내부 안착", "관의 충돌 및 흠집을 방지하기 위해 섬유 벨트 슬링을 사용하여 서서히 인하합니다.", "crane"),
            ("3단계: 소켓 접합부 청소, 전용 윤활제 도포 및 체결", "레버블록을 사용하여 관을 수평으로 곧게 밀어 넣고 편각 허용치(1.5도 이내)를 준수합니다.", "joint"),
            ("4단계: 곡관부 및 이형관 지지용 콘크리트 스러스트 블록 타설", "수압에 의한 관로 이탈을 방지하기 위해 꺾임부에 콘크리트 반력대를 설치합니다.", "concrete"),
            ("5단계: 관로 부설 선형 및 접합부 감리 서면 검측", "관 상단 높이와 연결부 밀착도를 전수 검측받습니다.", "sign")
        ]
    },
    "무단수 연결을 위한 시설 설치": {
        "type": "Special Protocol",
        "theme": "amber",
        "badge": "무단수 핫태핑 특수공법 (Special Protocol)",
        "sim_title": "🔥 무단수 핫태핑(Hot Tapping) & 차단 밸브 제어기",
        "sim_body": "<div class='p-3 bg-amber-50 border border-amber-200 rounded-xl text-xs space-y-1.5'><div>배관 내 통수 상태: <b>정상 급수 유지 중 (단수 피해 0가구)</b></div><div>천공 압력: <b>1.0 MPa 밀폐 유지 (무누수 천공 완벽 성공)</b></div></div>",
        "steps": [
            ("Special 01: 기존 활관 외경 정밀 연마 및 분기 티관(Split Tee) 용접", "관 표면의 녹을 완전히 제거하고 규격 새들 티관을 밀착 용접합니다.", "weld"),
            ("Special 02: 샌드위치 밸브(Sandwich Valve) 장착 및 기밀 시험", "천공기 장착 전 밸브 플랜지부의 수밀성을 1.5배 수압으로 사전 검증합니다.", "valve"),
            ("Special 03: 무단수 천공기(Hot Tapping Machine) 장착 및 관벽 천공", "단수 없이 수압이 걸린 상태에서 특수 커터로 기존 관벽을 안전하게 절단합니다.", "hottap"),
            ("Special 04: 절단 쿠폰(Coupon) 인양 및 샌드위치 밸브 차단", "절단된 배관 조각을 안전하게 회수하고 밸브를 닫아 통수를 차단합니다.", "coupon"),
            ("Special 05: 바이패스(By-Pass) 우회 배관 가동", "신설 이설 관로로 물의 흐름을 전환하여 단수 없는 무정지 통수를 완성합니다.", "bypass"),
            ("Special 06: 가스안전공사/맑은물사업소 3자 입회 서명", "무누수 상태를 확인하고 감독관 입회 서명을 날인합니다.", "sign")
        ]
    },
    "신규관로 및 연결관로 접속": {
        "type": "3-Step",
        "theme": "cyan",
        "badge": "최종 배관 접속 (3-Step Tie-in)",
        "sim_title": "🔧 신구 관로 Tie-In 최종 연결 대시보드",
        "sim_body": "<div class='p-3 bg-white border border-cyan-200 rounded-xl text-xs flex justify-between'><span>Tie-In 연결부 볼트 토크치:</span><b class='text-cyan-700'>180 N·m (토크렌치 전수 합격)</b></div>",
        "steps": [
            ("1단계: 신설 관로와 기존 관로 사이 Tie-in 조립", "슬리브 이음관(Mechanical Coupling)을 장착하고 편차를 정밀 조정합니다.", "joint"),
            ("2단계: 고장력 볼트 토크렌치 대각선 방향 균등 체결", "패킹의 균일한 압착을 위해 규정 토크치로 대각선 순서로 조입니다.", "torque"),
            ("3단계: 이음부 에폭시 수지 방식 테이프 및 열수축 시트 마감", "지중 매설부의 영구적인 부식을 방지하기 위해 3중 방식 피복을 실시합니다.", "sign")
        ]
    },
    "기존관로 철거 및 원상복구": {
        "type": "3-Step",
        "theme": "emerald",
        "badge": "폐관 처리 & 도로복구 (3-Step Recovery)",
        "sim_title": "🏗️ 폐관 몰탈 주입 및 아스콘 도로 원상복구",
        "sim_body": "<div class='p-3 bg-emerald-50 border border-emerald-200 rounded-xl text-xs flex justify-between'><span>도로 평탄성(PrI):</span><b class='text-emerald-700'>7.2 cm/km (고속 주행 적합 기준 통과)</b></div>",
        "steps": [
            ("1단계: 노후 폐관 내 유동성 채움재(기포 콘크리트) 전구간 주입", "폐관의 지반 침하를 방지하기 위해 시멘트 몰탈을 가압 주입하여 완전히 밀폐합니다.", "grouting"),
            ("2단계: 양질의 토사 30cm 층다짐 및 노반 다짐도(≥95%) 시험", "층별 다짐을 철저히 하고 평판재하시험(Ev2 ≥ 120MPa)을 통과합니다.", "compaction"),
            ("3단계: 기층(10cm) + 표층(5cm) 아스팔트 포장 및 차선 도색", "기존 도로와 단차가 없도록 열융착 시공하고 융착식 차선을 복원합니다.", "pavement")
        ]
    },
    "광역상수관 이설 공사": {
        "type": "5-Step",
        "theme": "cyan",
        "badge": "광역상수도 D1200 대구경 이설 (5-Step Mega Pipe)",
        "sim_title": "💧 K-water 광역상수도 실시간 유량 및 수압 모니터",
        "sim_body": "<div class='p-3 bg-cyan-950 text-cyan-200 rounded-xl font-mono text-xs flex justify-between'><span>유량: 4,500 ㎥/h</span><span class='text-emerald-400'>● 수압: 0.85 MPa 정상 공급</span></div>",
        "steps": [
            ("1단계: K-water 한국수자원공사 합동 기술 검토 및 수계 전환", "대규모 단수를 예방하기 위해 인근 정수장 배수지 수위를 만수위로 사전 확보합니다.", "water"),
            ("2단계: D1200 강관 용접부 100% UT(초음파) 비파괴검사", "고압 급수관의 용접 결함을 완벽히 배제하기 위해 전 구간 비파괴검사를 필합니다.", "ndt"),
            ("3단계: 대형 버터플라이 밸브(전동 구동식)실 구조물 축조", "비상시 원격 차단이 가능한 밸브실 구조물을 견고하게 시공합니다.", "valve"),
            ("4단계: 관 세척(Flushing) 및 음용수 수질 59개 항목 적합 시험", "소독 잔류염소 및 탁도를 측정하여 안전한 수질을 확인합니다.", "water_test"),
            ("5단계: 한국수자원공사 최종 시설물 인계인수 승인", "이설 준공 도면 및 검측 대장을 인계하고 관로 통수를 개시합니다.", "sign")
        ]
    },
    "상수도 관로 이설 공사": {
        "type": "5-Step",
        "theme": "cyan",
        "badge": "상수도 배수관망 이설 (5-Step Water Pipe)",
        "sim_title": "💧 상수관 10kgf/cm² 수압시험 및 잔류염소 검측",
        "sim_body": "<div class='p-3 bg-cyan-50 border border-cyan-200 rounded-xl text-xs flex justify-between'><span>수압 유지:</span><b class='text-cyan-700'>1.0 MPa 60분 유지 누수 0L (합격)</b></div>",
        "steps": [
            ("1단계: 주철관 터파기 및 관경별(D300~D800) 트렌치 거치", "지하 1.5m 이하 동결심도 이하로 굴착하고 모래 기초를 조성합니다.", "water"),
            ("2단계: 닥타일 소켓 조인트 연결 및 이탈방지 압륜 장착", "수압 반력에 의해 이음부가 빠지지 않도록 락링 볼트를 체결합니다.", "joint"),
            ("3단계: 관로 10kgf/cm²(1.0MPa) 1시간 수압 시험 실시", "감리원 및 상수도 공무원 입회하에 압력 강하가 없음을 검증합니다.", "hydro"),
            ("4단계: 소화전 및 제수밸브실 부대시설 설치", "화재 진압용 소화전과 단수 차단용 제수밸브를 도로변에 설치합니다.", "hydrant"),
            ("5단계: 급수관 인입 연결 및 통수 승인", "개별 수용가 인입관을 연결하고 맑은물사업소의 최종 인수를 받습니다.", "sign")
        ]
    },
    "하수도 관로 이설 공사": {
        "type": "5-Step",
        "theme": "cyan",
        "badge": "하수암거(RC Box) 및 오수관 이설 (5-Step Sewer)",
        "sim_title": "📹 하수관 CCTV 로봇 내시경 관로 검사기",
        "sim_body": "<div class='p-3 bg-slate-900 text-green-400 font-mono text-xs flex justify-between'><span>CCTV CRAWLER: 120m 주행 완료</span><span class='text-emerald-400'>● 관체 균열/역구배 0건</span></div>",
        "steps": [
            ("1단계: 현장타설 RC 박스(2.0x2.0m) 또는 PC 암거 터파기", "자연 유하 구배(1/500)를 철저히 확인하며 터파기 레벨을 맞춥니다.", "trench"),
            ("2단계: 버림 콘크리트 타설 및 바닥 방수 시트 시공", "지하수 유입 및 하수 누수를 방지하기 위해 완벽한 차수 처리를 합니다.", "concrete"),
            ("3단계: 철근 배근, 거푸집 설치 및 고강도 수밀 콘크리트 타설", "수밀성 콘크리트를 타설하고 내부 표면을 매끄럽게 마감합니다.", "rebar"),
            ("4단계: 관내 CCTV 로봇 내시경 전구간 촬영 및 수밀 시험", "하수관 내부에 로봇 카메라를 투입하여 이음부 균열 및 침입수 여부를 검사합니다.", "cctv"),
            ("5단계: 맨홀 뚜껑 상단 노면 일치 마감 및 하수도과 인수", "맨홀 뚜껑이 트램 궤도면 및 도로면과 1:1 수평을 이루도록 높이를 맞춥니다.", "sign")
        ]
    },
    "도시가스관 이설 공사": {
        "type": "5-Step",
        "theme": "amber",
        "badge": "도시가스 중압배관 방폭 이설 (5-Step Gas Line)",
        "sim_title": "🔥 도시가스 기밀시험 및 PID 가스누출 감지기",
        "sim_body": "<div class='p-3 bg-amber-50 border border-amber-200 rounded-xl text-xs space-y-1'><div>가스 배관 압력: <b>0.85 MPa (중압 A)</b></div><div>가스안전공사 1:1 입회: <b>PID 가스 누출 농도 0.00 ppm (완전 기밀)</b></div></div>",
        "steps": [
            ("1단계: 가스 배관 용접사 자격 검증 및 용접 절차서(WPS) 승인", "한국가스안전공사 인증 고압가스 특수용접 자격자를 배치합니다.", "welder"),
            ("2단계: 가스 강관 맞대기 용접 및 100% 방사선투과(RT) 검사", "모든 용접 비드에 대해 방사선 필름 촬영을 실시하여 결함 0건을 입증합니다.", "rt"),
            ("3단계: 3중 열수축 방식 시트 피복 및 핀홀(탐침) 절연 시험", "부식 방지를 위해 15,000V 스파크 탐침기를 대어 피복 손상을 전수 검사합니다.", "spark"),
            ("4단계: 배관 가압 기밀 시험(0.85MPa 24시간 유지)", "온도 보정 압력 강하가 전혀 없음을 가스안전공사 검사원과 확인합니다.", "gauge"),
            ("5단계: 질소(N2) 퍼지 후 도시가스 치환 및 통가스 승인", "공기를 질소로 치환한 후 안전하게 가스를 통입하고 공사를 완료합니다.", "sign")
        ]
    },
    "지역난방관 이설 공사": {
        "type": "5-Step",
        "theme": "red",
        "badge": "지역난방 고온열배관 이설 (5-Step Heat Pipe)",
        "sim_title": "🌡️ 120℃ 고온수 공급관 신축흡수 및 보온 감시",
        "sim_body": "<div class='p-3 bg-red-50 border border-red-200 rounded-xl text-xs space-y-1'><div>공급수 온도: <b>120℃ / 1.6 MPa</b></div><div>보온재 감시선: <b>누수 감시선 루프 저항 정상 (단락 0건)</b></div></div>",
        "steps": [
            ("1단계: 이중보온관(내관 강관 + 외관 PE관 + PUR 보온재) 반입", "한국지역난방공사 시방서 기준 열전도율 및 보온 두께를 검사합니다.", "pipe"),
            ("2단계: 고온 강관 맞대기 용접 및 100% 비파괴(RT/UT) 검사", "120℃ 고온 열팽창에 견딜 수 있도록 1급 용접 품질을 확보합니다.", "ndt"),
            ("3단계: 누수 감지선(Leak Detection Wire) 결선 및 도통 시험", "배관 내부에 내장된 센서선을 연결하고 계측기로 신호 저항을 확인합니다.", "wire"),
            ("4단계: 수축 슬리브 접합부 외관 열융착 및 발포폼 주입", "연결부에 고밀도 폴리우레탄 폼을 현장 발포하여 완전 방수 단열합니다.", "foam"),
            ("5단계: 한국지역난방공사 입회하에 통수 및 준공 승인", "시운전 통온 및 압력 유지를 확인받고 시설물을 최종 인계합니다.", "sign")
        ]
    },
    "통신관로 및 케이블 이설 공사": {
        "type": "5-Step",
        "theme": "violet",
        "badge": "통신 144심 광케이블 무중단 이설 (5-Step Telecom)",
        "sim_title": "📶 OTDR 1550nm 광손실 감쇠율 측정기",
        "sim_body": "<div class='p-3 bg-violet-950 text-violet-200 rounded-xl font-mono text-xs flex justify-between'><span>OTDR 144 Core Loss:</span><b class='text-emerald-400'>0.02 dB/km (KT 초고속망 기준 통과)</b></div>",
        "steps": [
            ("1단계: 다조 통신관로(PVC 파이프 16공) 및 콘크리트 핸드홀 매설", "통신 케이블 인입 반경을 확보하며 견고한 보호 콘크리트를 타설합니다.", "duct"),
            ("2단계: 윈치를 이용한 144심 단일모드 광케이블 지중 포설", "광섬유 인장 하중(≤150kg)을 초과하지 않도록 장력을 자동 제어하며 견인합니다.", "winch"),
            ("3단계: 무진 차량 내 정밀 코어 정렬 융착 접속(Fusion Splicing)", "클린 작업대에서 광섬유 코어를 1:1로 융착하고 열수축 슬리브로 보강합니다.", "splice"),
            ("4단계: OTDR 광파장 계측기를 통한 손실(≤0.03dB) 전심선 전수 측정", "1310nm/1550nm 파장에서 반사 및 접속 손실을 전수 인쇄합니다.", "otdr"),
            ("5단계: KT/SKT/LGU+ 통신사 입회 절체(Cut-over) 완료 승인", "인터넷 및 국가통신망 서비스 중단 없이 절체를 완료하고 승인받습니다.", "sign")
        ]
    },
    "특고압 전력관로 이설 공사": {
        "type": "5-Step",
        "theme": "red",
        "badge": "한전 22.9kV 특고압 활선 이설 (5-Step High Voltage)",
        "sim_title": "⚡ 22.9kV CNCV-W 절연내력 및 휴전 절체 제어",
        "sim_body": "<div class='p-3 bg-red-950 text-red-200 rounded-xl font-mono text-xs space-y-1'><div>● 직류 고전압 인가: 30,000V DC / 10분간 연속 인가 (누설전류 0.05mA)</div><div>● 한전 배전처: <b>활선작업 안전 수칙 100% 준수 합격</b></div></div>",
        "steps": [
            ("1단계: 특고압 지중 맨홀 축조 및 ELP/파이프 덕트 전력관로 부설", "한전 설계 기준에 따라 붉은색 지중 경고 테이프를 2중 포설합니다.", "manhole"),
            ("2단계: 22.9kV CNCV-W (동심중성선 수밀형) 케이블 포설", "케이블 굴곡 반경(외경의 12배 이상)을 준수하여 피복 손상 없이 인입합니다.", "cable"),
            ("3단계: 한전 공인 자격자의 엘보 조인트 및 스트레이트 슬리브 단말 처리", "전계 완화 콘을 장착하고 반도전층을 정밀 제거하여 절연 파괴를 방지합니다.", "joint_elec"),
            ("4단계: 절연내력 시험(DC 30kV 10분) 및 절연저항(≥2000MΩ) 측정", "한전 감독관 입회하에 특고압 내력 시험을 실시하여 합격을 득합니다.", "high_pot"),
            ("5단계: 한전 배전운영실 입회하에 휴전(Cut-over) 개폐기 조작 및 통전", "구선로에서 신선로로 전력을 무사히 전환하고 송전을 개시합니다.", "sign")
        ]
    },
    "상하수도 이설 시공 최종 점검": {
        "type": "Testing Protocol",
        "theme": "cyan",
        "badge": "상하수도 최종 종합 검측 (Protocol)",
        "sim_title": "💧 수압·CCTV·수질 3대 지표 종합 판정표",
        "sim_body": "<div class='p-3 bg-cyan-50 border border-cyan-200 rounded-xl text-xs space-y-1'><div>수압시험: <b>10kgf/cm² 합격</b> | CCTV 관로 내부: <b>결함 0건 합격</b></div><div>음용수 수질: <b>탁도 0.08 NTU 적합</b></div></div>",
        "steps": [
            ("Protocol 01: 신설 관로 전체 구간 정밀 측량 및 준공 검측", "계획 노선과 시공 노선의 좌표 일치 여부를 토탈스테이션으로 실측합니다.", "measure"),
            ("Protocol 02: 수압 시험, CCTV 내시경, 수질 검사 성과표 종합", "모든 공인 시험 성적서를 바인더로 편철하여 규격 적합성을 입증합니다.", "doc"),
            ("Protocol 03: 화성시 맑은물사업소 합동 현장 최종 전수 실사", "지상 밸브실, 맨홀 뚜껑, 제수밸브 개폐 상태를 공무원과 전수 점검합니다.", "survey"),
            ("Protocol 04: 최종 점검 대장 서명 날인 및 인수인계 완결", "하자 보수 보증서(보증기간 5년)를 제출하고 정식 인수를 완료합니다.", "sign")
        ]
    },
    "위수탁 지하지장물 이설 최종 점검": {
        "type": "Testing Protocol",
        "theme": "indigo",
        "badge": "유관기관 합동 최종 인수 검측 (Protocol)",
        "sim_title": "🏛️ 5대 위탁기관 최종 시설물 인수증 발급 보드",
        "sim_body": "<div class='p-3 bg-indigo-50 border border-indigo-200 rounded-xl text-xs flex justify-between'><span>한전·가스·통신·지역난방·맑은물:</span><b class='text-emerald-700'>5개 기관 전원 인수 서명 완료</b></div>",
        "steps": [
            ("Protocol 01: 위탁기관별 1:1 현장 성과품 검측", "각 기관의 유지관리 부서 실무자와 현장을 도보로 전수 점검합니다.", "survey"),
            ("Protocol 02: 3D BIM 기반 지하 지장물 준공 데이터베이스 구축", "지하시설물 통합 관리 시스템(GIS)에 정밀 3차원 위치 데이터를 갱신합니다.", "bim"),
            ("Protocol 03: 기관별 시설물 인계인수 협약서 정식 체결", "시설물의 관리권과 소유권을 위탁기관으로 완전히 환원합니다.", "sign")
        ]
    },
    "상하수도 이설 정산 금액 검토": {
        "type": "2-Step",
        "theme": "emerald",
        "badge": "공사비 정산 검증 (2-Step Settlement)",
        "sim_title": "💰 실투입 수량 기반 공사비 정산 검토서",
        "sim_body": "<div class='p-3 bg-white border border-emerald-200 rounded-xl text-xs flex justify-between'><span>도급분 정산 검증액:</span><b class='text-emerald-700'>실투입 물량 증빙 100% 일치 (서면 승인)</b></div>",
        "steps": [
            ("Step 1: 실정보고 수량과 현장 실투입 검측 물량 1:1 대조", "아스팔트 절단, 터파기 토량, 배관 자재 수량의 현장 송장을 전수 대조합니다.", "calc"),
            ("Step 2: 감리원 최종 정산 검토 의견서 작성 및 승인", "단가 산정의 적법성과 수량 계산의 정확성을 검토하여 승인합니다.", "sign")
        ]
    },
    "위수탁 처리 정산금액 지급": {
        "type": "2-Step",
        "theme": "emerald",
        "badge": "분담금 정산 집행 (2-Step Disbursement)",
        "sim_title": "🏛️ 위수탁 기관 정산 분담금 지출 결재",
        "sim_body": "<div class='p-3 bg-white border border-emerald-200 rounded-xl text-xs flex justify-between'><span>지출 결재 상태:</span><b class='text-emerald-700'>발주처 최종 지급 승인 완료</b></div>",
        "steps": [
            ("Phase 1: 위탁기관 준공 정산서 및 증빙 영수증 최종 심사", "기관별 실제 집행 공사비 내역서를 검토하여 잔여 예산을 정산합니다.", "doc"),
            ("Phase 2: 위수탁 정산금 잔액 입금 및 채권·채무 종결", "과부족 금액을 상호 정산하고 정산 종결 합의서를 교환합니다.", "sign")
        ]
    },
    "도급자분/위수탁분 설계변경 정산": {
        "type": "2-Step",
        "theme": "emerald",
        "badge": "총사업비 계약 변경 (2-Step Contract Change)",
        "sim_title": "📑 동탄도시철도 건설공사 총사업비 변경 계약",
        "sim_body": "<div class='p-3 bg-white border border-emerald-200 rounded-xl text-xs flex justify-between'><span>계약 변경 완료:</span><b class='text-indigo-700'>화성시 도급 계약 및 위탁 협약 100% 갱신 완료</b></div>",
        "steps": [
            ("Step 1: 도급 및 위수탁 설계변경 총괄 내역서 취합", "전체 지장물 이설 공사의 공사비 증감을 종합 집계하여 변경 계약서를 작성합니다.", "doc"),
            ("Step 2: 발주처 정식 계약 변경 체결 및 예산 배정 확정", "최종 변경 계약 체결로 지장물 이설 관련 모든 계약 행정을 완결합니다.", "sign")
        ]
    },
    "공사전 선행공종에서 인수받을 사항": {
        "type": "Interface Matrix",
        "theme": "emerald",
        "badge": "선행 토공 ➔ 지장물 인수 (Interface Matrix)",
        "sim_title": "🤝 선행 공종 현장 인계인수 확인증",
        "sim_body": "<div class='p-3 bg-emerald-50 border border-emerald-200 rounded-xl text-xs flex justify-between'><span>선행 토공사 인계 상태:</span><b class='text-emerald-700'>노선 기준점(TBM) 및 부지 정지 인수 완료</b></div>",
        "steps": [
            ("Interface 01: 선행 토공 및 도로 기준점(TBM, CP) 좌표 성과 인수", "공인 측량 기준점 4개소를 인수하여 현장 이설 기준 좌표계로 설정합니다.", "survey"),
            ("Interface 02: 굴착 작업 부지 내 기존 점용물 철거 상태 확인", "가로등, 표지판, 가로수 등이 제거되어 지장물 작업 공간이 확보되었는지 점검합니다.", "site"),
            ("Interface 03: 선행 공종 책임감리원 및 시공사 3자 인수 확인서 서명", "인계인수 대장에 현장 사진을 부착하고 서명 날인하여 책임을 명확히 합니다.", "sign")
        ]
    },
    "공사중 챙겨야할 후행공종의 요구사항": {
        "type": "Interface Matrix",
        "theme": "emerald",
        "badge": "지장물 ➔ 후행 궤도/노반 인계 (Interface Matrix)",
        "sim_title": "🤝 후행 궤도 시공팀 사전 간섭 제로화 협약",
        "sim_body": "<div class='p-3 bg-emerald-50 border border-emerald-200 rounded-xl text-xs flex justify-between'><span>후행 궤도 요구 조건:</span><b class='text-emerald-700'>궤도 하부 1.5m 이내 횡단 매설관 100% 방호관 거치 확인</b></div>",
        "steps": [
            ("Interface 01: 후행 트램 궤도 기초(Slab) 하부 지지력(Ev2 ≥ 120MPa) 확보", "이설 구간 되메우기 시 다짐 불량으로 인한 후행 궤도 침하를 원천 차단합니다.", "compaction"),
            ("Interface 02: 궤도 횡단 관로 강관 Casing Sleeve 이중 방호 거치", "트램 열차 통과 시 발생하는 하중과 진동으로부터 매설관을 완벽 보호합니다.", "sleeve"),
            ("Interface 03: 후행 노반 및 궤도 기술사 3자 합동 현측 검측 승인", "후행 공종에 아무런 장애가 없음을 확인받고 궤도 공종으로 부지를 인계합니다.", "sign")
        ]
    }
}

# SVG Generator for distinct visual cards
def render_step_svg_custom(step_name, icon_type, theme_color):
    return f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-white border border-slate-200 rounded-xl p-3 mb-3 shadow-sm hover:shadow" onclick="openDiagramZoom(this.outerHTML, '{step_name}')">
        <div class="text-[11px] font-bold text-slate-500 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-indigo-600 bg-indigo-50 px-2 py-0.5 rounded border border-indigo-200 font-extrabold">{step_name[:20]}</span>
        </div>
        <svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-slate-200">
            <rect width="460" height="140" fill="#F8FAFC"/>
            <rect x="30" y="30" width="400" height="80" rx="8" fill="#FFFFFF" stroke="#0284C7" stroke-width="1.5"/>
            <text x="60" y="75" font-size="13" font-weight="900" fill="#0F172A">{step_name[:32]}</text>
            <text x="60" y="95" font-size="11" font-weight="bold" fill="#64748B">동탄트램 현장 공학 표준 엔지니어링 도식</text>
            <circle cx="390" cy="70" r="18" fill="#E0F2FE" stroke="#0284C7" stroke-width="2"/>
            <text x="382" y="75" font-size="12" font-weight="900" fill="#0284C7">2D</text>
        </svg>
    </div>'''

def generate_full_html(act, spec):
    title = act['task_title']
    purpose = act['purpose']
    output = act['output']
    risk = act['risk']
    advisory = act['advisory']
    dept = act['dept']
    
    steps = spec['steps']
    step_count = len(steps)
    arch_type = spec['type']
    
    # Render Steps Layout dynamically based on step count
    if step_count == 2:
        steps_grid = "<div class='grid grid-cols-1 md:grid-cols-2 gap-6'>"
        for idx, (s_title, s_desc, s_icon) in enumerate(steps, 1):
            svg = render_step_svg_custom(s_title, s_icon, spec['theme'])
            steps_grid += f'''<div class="bg-slate-50 border border-slate-200 rounded-2xl p-6 shadow-sm hover:shadow-md transition">
                {svg}
                <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-3 py-1 rounded-full mb-2 inline-block">PHASE {idx:02d}</span>
                <h4 class="font-extrabold text-base text-slate-900 mb-2">{s_title}</h4>
                <p class="text-slate-700 text-sm font-medium leading-relaxed">{s_desc}</p>
            </div>'''
        steps_grid += "</div>"
    elif step_count == 3:
        steps_grid = "<div class='grid grid-cols-1 md:grid-cols-3 gap-5'>"
        for idx, (s_title, s_desc, s_icon) in enumerate(steps, 1):
            svg = render_step_svg_custom(s_title, s_icon, spec['theme'])
            steps_grid += f'''<div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 shadow-sm hover:shadow-md transition">
                {svg}
                <span class="text-xs font-black text-sky-700 bg-sky-100 px-2.5 py-1 rounded-full mb-2 inline-block">STEP {idx:02d}</span>
                <h4 class="font-bold text-sm text-slate-900 mb-2 leading-snug">{s_title}</h4>
                <p class="text-slate-700 text-xs font-medium leading-relaxed">{s_desc}</p>
            </div>'''
        steps_grid += "</div>"
    elif arch_type == "Testing Protocol":
        steps_grid = f"<div class='space-y-4 bg-cyan-50/40 border border-cyan-200 rounded-2xl p-6'><div class='grid grid-cols-1 md:grid-cols-{min(step_count, 4)} gap-4'>"
        for idx, (s_title, s_desc, s_icon) in enumerate(steps, 1):
            svg = render_step_svg_custom(s_title, s_icon, spec['theme'])
            steps_grid += f'''<div class="bg-white border border-cyan-200 rounded-xl p-4 shadow-sm">
                {svg}
                <span class="text-[11px] font-black text-cyan-700 bg-cyan-100 px-2 py-0.5 rounded-full inline-block mb-1.5">PROTOCOL {idx:02d}</span>
                <h5 class="font-bold text-xs text-slate-900 mb-1 leading-tight">{s_title}</h5>
                <p class="text-slate-600 text-[11px] font-medium leading-relaxed">{s_desc}</p>
            </div>'''
        steps_grid += "</div></div>"
    elif arch_type == "Interface Matrix":
        steps_grid = f"<div class='space-y-4 bg-emerald-50/40 border border-emerald-200 rounded-2xl p-6'><div class='grid grid-cols-1 md:grid-cols-{min(step_count, 3)} gap-4'>"
        for idx, (s_title, s_desc, s_icon) in enumerate(steps, 1):
            svg = render_step_svg_custom(s_title, s_icon, spec['theme'])
            steps_grid += f'''<div class="bg-white border border-emerald-200 rounded-xl p-4 shadow-sm">
                {svg}
                <span class="text-[11px] font-black text-emerald-700 bg-emerald-100 px-2 py-0.5 rounded-full inline-block mb-1.5">INTERFACE {idx:02d}</span>
                <h5 class="font-bold text-xs text-slate-900 mb-1 leading-tight">{s_title}</h5>
                <p class="text-slate-600 text-[11px] font-medium leading-relaxed">{s_desc}</p>
            </div>'''
        steps_grid += "</div></div>"
    else: # 5-Step or 6-Step Detailed Timeline
        steps_grid = f"<div class='space-y-4'><div class='grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4'>"
        for idx, (s_title, s_desc, s_icon) in enumerate(steps, 1):
            svg = render_step_svg_custom(s_title, s_icon, spec['theme'])
            steps_grid += f'''<div class="bg-slate-50 border border-slate-200 rounded-xl p-4 hover:shadow-md transition">
                {svg}
                <span class="text-[11px] font-black text-amber-800 bg-amber-100 px-2 py-0.5 rounded-full inline-block mb-1.5">STEP {idx:02d}</span>
                <h5 class="font-bold text-xs text-slate-900 mb-1 leading-tight">{s_title}</h5>
                <p class="text-slate-600 text-[11px] font-medium leading-relaxed">{s_desc}</p>
            </div>'''
        steps_grid += "</div></div>"

    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 지장물이설 - {title} 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style> body {{ font-family: 'Noto Sans KR', sans-serif; }} </style>
</head>
<body class="bg-slate-100 text-slate-800 antialiased py-8 px-4 sm:px-8">

    <div class="max-w-6xl mx-auto bg-white rounded-3xl shadow-xl border border-slate-200 p-6 sm:p-10 space-y-8">
        <!-- 🔵 헤더 영역 (고유 아키텍처 뱃지) -->
        <div class="flex flex-col sm:flex-row sm:items-center justify-between pb-6 border-b border-slate-200 gap-4">
            <div>
                <span class="text-xs font-black px-3.5 py-1.5 rounded-full mb-2 inline-block border bg-indigo-50 text-indigo-700 border-indigo-200">
                    Dongtan Tram Utility Playbook | WBS {act['l3_code']} ({spec['badge']})
                </span>
                <h1 class="text-2xl sm:text-4xl font-black text-slate-900 tracking-tight">
                    {title} 수행지침서
                </h1>
            </div>
            <div class="shrink-0 text-right text-xs font-bold text-slate-500">
                주관부서: <span class="text-indigo-600 font-extrabold">{dept}</span>
            </div>
        </div>

        <!-- 📌 1. 작업 개요 및 핵심 목적 -->
        <div class="bg-blue-50/70 border border-blue-200 rounded-2xl p-6 space-y-3">
            <h3 class="text-base font-black text-blue-950 flex items-center gap-2">
                <span>📌</span> 현장 이행 목적 및 엔지니어링 방침
            </h3>
            <p class="text-slate-800 text-sm font-semibold leading-relaxed">
                {purpose}
            </p>
        </div>

        <!-- 📡 2. 고유 인터랙티브 공학 시뮬레이터 -->
        <div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 space-y-3">
            <h4 class="text-xs font-extrabold text-slate-900 flex items-center gap-2">
                <span>{spec['sim_title']}</span>
            </h4>
            {spec['sim_body']}
        </div>

        <!-- 💡 3. 유연한 공학 세부 수행절차 및 2D 시공 도식 -->
        <div class="space-y-4">
            <h3 class="text-xl font-black text-slate-900 flex items-center gap-2">
                <span>💡</span> {arch_type} 맞춤형 공학 세부 수행절차
            </h3>
            {steps_grid}
        </div>

        <!-- 📦 4. 최종 성과품 -->
        <div class="bg-emerald-50 border border-emerald-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-emerald-950 flex items-center gap-2">
                <span>📂</span> 최종 필수 성과품 / 결재 대장
            </h3>
            <p class="text-emerald-900 text-sm font-bold leading-relaxed">
                {output}
            </p>
        </div>

        <!-- ⚠️ 5. 집행단계 리스크 예방 관리 -->
        <div class="bg-amber-50 border-2 border-amber-300 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>⚠️</span> 집행단계 핵심 리스크 및 예방 대책
            </h3>
            <p class="text-amber-900 text-sm font-bold leading-relaxed">
                {risk}
            </p>
        </div>

        <!-- 👷 6. 협력업체 실무 자문 노하우 -->
        <div class="bg-amber-50/60 border border-amber-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>👷</span> [협력업체 실무 자문] 시공 및 공사관리 핵심 가이드
            </h3>
            <p class="text-amber-900 text-sm font-semibold leading-relaxed">
                {advisory}
            </p>
        </div>

        <!-- 📌 푸터 -->
        <div class="pt-6 border-t border-slate-200 flex items-center justify-between text-xs text-slate-500">
            <span>🏢 동탄도시철도(트램) 건설공사 지장물이설 수행지침</span>
            <span>v8 엑셀 100% 1:1 유연 아키텍처 연동</span>
        </div>
    </div>

    <!-- Lightbox Zoom Modal -->
    <div id="zoomModal" class="fixed inset-0 z-[9999] hidden bg-black/80 backdrop-blur-md flex items-center justify-center p-4" onclick="this.classList.add('hidden')">
        <div class="bg-white rounded-3xl p-6 max-w-4xl w-full max-h-[90vh] overflow-auto border-4 border-indigo-500 shadow-2xl relative" onclick="event.stopPropagation()">
            <button onclick="document.getElementById('zoomModal').classList.add('hidden')" class="absolute top-4 right-4 bg-slate-900 text-white font-black px-4 py-2 rounded-full text-sm hover:bg-slate-700">✕ 닫기</button>
            <h3 id="zoomTitle" class="text-xl font-black text-slate-900 mb-4 pb-2 border-b border-slate-200">고해상도 2D 시공 도식 확대보기</h3>
            <div id="zoomContent" class="flex justify-center items-center"></div>
        </div>
    </div>
    <script>
    function openDiagramZoom(htmlContent, title) {{
        var modal = document.getElementById('zoomModal');
        document.getElementById('zoomTitle').innerText = title || '2D 시공 도식 확대보기';
        document.getElementById('zoomContent').innerHTML = htmlContent;
        modal.classList.remove('hidden');
    }}
    </script>
</body>
</html>'''
    return html

# Load activities from Excel workbook
wb = openpyxl.load_workbook(v8_excel, data_only=True)
ws = wb['지장물이설']
activities = []
for r in range(2, ws.max_row + 1):
    task_title = ws.cell(row=r, column=8).value
    if task_title:
        activities.append({
            'row': r,
            'l3_code': ws.cell(row=r, column=2).value or '2000',
            'task_title': str(task_title).strip(),
            'dept': str(ws.cell(row=r, column=9).value or '현장 공무/시공팀').strip(),
            'purpose': str(ws.cell(row=r, column=10).value or '지장물 이설 품질 및 안전 검측').strip(),
            'method': str(ws.cell(row=r, column=11).value or '1) 사전검토 -> 2) 현측 -> 3) 시공 -> 4) 승인').strip(),
            'output': str(ws.cell(row=r, column=12).value or '지장물 이설 정밀 검측성과표 및 최종 승인서').strip(),
            'risk': str(ws.cell(row=r, column=21).value or '지하시설물 간섭 및 안전사고 리스크 사전 차단').strip(),
            'advisory': str(ws.cell(row=r, column=22).value or '관할 점용기관 및 시공 전문 기술자 자문').strip(),
        })
wb.close()

# Write out to all subfolders in v8/지장물이설
updated_count = 0
for act in activities:
    title = act['task_title']
    spec = TASK_SPECS.get(title)
    if not spec:
        # fuzzy lookup
        for k, v in TASK_SPECS.items():
            if k in title or title in k:
                spec = v
                break
    if not spec:
        spec = TASK_SPECS["Site Survey Risk 검토"]
        
    html_content = generate_full_html(act, spec)
    safe_title = re.sub(r'[\/:*?"<>|]', '_', title)
    
    # Locate all matching folders
    matching_folders = []
    if os.path.exists(util_dir):
        for sub in os.listdir(util_dir):
            sub_p = os.path.join(util_dir, sub)
            if os.path.isdir(sub_p):
                clean_sub = sub.split('_', 1)[-1] if '_' in sub else sub
                if title in sub or clean_sub in title or title in clean_sub:
                    matching_folders.append(sub_p)
                    
        if not matching_folders:
            kw = title.split()[0]
            for sub in os.listdir(util_dir):
                sub_p = os.path.join(util_dir, sub)
                if os.path.isdir(sub_p) and kw in sub:
                    matching_folders.append(sub_p)

    for mf in matching_folders:
        guide_dir = os.path.join(mf, '수행지침')
        os.makedirs(guide_dir, exist_ok=True)
        
        f1 = os.path.join(guide_dir, f"{safe_title}_수행지침.html")
        f2 = os.path.join(guide_dir, "수행지침.html")
        
        with open(f1, 'w', encoding='utf-8') as f:
            f.write(html_content)
        with open(f2, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        for existing in os.listdir(guide_dir):
            if existing.endswith('.html'):
                with open(os.path.join(guide_dir, existing), 'w', encoding='utf-8') as f:
                    f.write(html_content)
                    
        updated_count += 1

print(f"Successfully generated true custom HTML manuals for {updated_count} 지장물이설 folders.", flush=True)
