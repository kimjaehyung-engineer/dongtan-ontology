# -*- coding: utf-8 -*-
import openpyxl, os, sys, shutil, urllib.parse, re

sys.stdout.reconfigure(encoding='utf-8')

v8_root = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8"
v8_excel = os.path.join(v8_root, '매뉴얼 BODY (집행단계)v8.xlsm')
util_dir = os.path.join(v8_root, '지장물이설')

print("=== Starting Generation of Diverse & Flexible Step-Architecture 지장물이설 HTML Manuals ===", flush=True)

wb = openpyxl.load_workbook(v8_excel, data_only=True)
ws = wb['지장물이설']

activities = []
for r in range(2, ws.max_row + 1):
    task_title = ws.cell(row=r, column=8).value
    if task_title:
        activities.append({
            'row': r,
            'l3_code': ws.cell(row=r, column=2).value or '2000',
            'task_title': str(task_title).strip(),
            'dept': str(ws.cell(row=r, column=9).value or '현장 공무/시공팀').strip(),
            'purpose': str(ws.cell(row=r, column=10).value or '지장물 이설 품질 및 안전 검측').strip(),
            'method': str(ws.cell(row=r, column=11).value or '1) 사전검토 ➔ 2) 현측 ➔ 3) 시공 ➔ 4) 승인').strip(),
            'output': str(ws.cell(row=r, column=12).value or '지장물 이설 정밀 검측성과표 및 최종 승인서').strip(),
            'risk': str(ws.cell(row=r, column=21).value or '지하시설물 간섭 및 안전사고 리스크 사전 차단').strip(),
            'advisory': str(ws.cell(row=r, column=22).value or '관할 점용기관 및 시공 전문 기술자 자문').strip(),
        })

wb.close()

def categorize_and_layout(title, method):
    t = title.lower()
    
    # Parse natural steps
    raw_steps = [s.strip() for s in method.split('➔') if s.strip()]
    if not raw_steps:
        raw_steps = [s.strip() for s in method.split('\n') if s.strip()]
        
    cleaned_steps = []
    for s in raw_steps:
        # clean leading numbers like 1), 2), [1] etc
        clean_s = re.sub(r'^[0-9]+[\).\s_-]+', '', s).strip()
        if clean_s:
            cleaned_steps.append(clean_s)
            
    if not cleaned_steps:
        cleaned_steps = ["사전 기술 검토 및 현장 조사", "지장물 이설 및 보호 조치", "최종 검측 및 승인"]
        
    # Custom category determination
    if '수압' in t or 'cctv' in t or 'ndt' in t or '비파괴' in t or '초음파' in t or '절연내력' in t or 'otdr' in t:
        layout_type = 'testing_protocol'
        theme_color = 'cyan'
    elif '선행' in t or '후행' in t or '인계' in t or '인수' in t or '인터페이스' in t or '합동' in t:
        layout_type = 'interface_matrix'
        theme_color = 'emerald'
    elif '가스' in t or '핫태핑' in t or '22.9kv' in t or '특고압' in t or '송유관' in t or len(cleaned_steps) >= 5:
        layout_type = 'detailed_timeline'
        theme_color = 'amber' if '가스' in t else ('red' if '전력' in t or '22.9' in t else 'slate')
    elif len(cleaned_steps) == 2:
        layout_type = 'dual_card'
        theme_color = 'indigo'
    elif len(cleaned_steps) == 3:
        layout_type = 'triple_card'
        theme_color = 'sky'
    else:
        layout_type = 'grid_cards'
        theme_color = 'blue'
        
    return layout_type, theme_color, cleaned_steps

def generate_step_svg(step_idx, step_text, theme_color, title):
    # Generates a clean, light-theme 2D SVG diagram tailored to the step text
    st = step_text.lower()
    
    if 'gpr' in st or '탐사' in st or '측량' in st or '조사' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-slate-200"><rect width="460" height="140" fill="#F8FAFC"/><line x1="0" y1="100" x2="460" y2="100" stroke="#64748B" stroke-width="2"/><rect x="30" y="50" width="70" height="35" rx="4" fill="#0284C7"/><circle cx="45" cy="85" r="7" fill="#1E293B"/><circle cx="85" cy="85" r="7" fill="#1E293B"/><text x="35" y="72" font-size="11" font-weight="bold" fill="#FFFFFF">GPR 탐사기</text><path d="M 65 85 Q 65 115 110 120" stroke="#38BDF8" stroke-width="2" stroke-dasharray="3,3"/><circle cx="240" cy="120" r="12" fill="#EAB308"/><text x="260" y="125" font-size="12" font-weight="900" fill="#854D0E">기존 지중 매설관 (GL-1.5m)</text></svg>'''
    elif '줄파기' in st or '시탐' in st or '굴착' in st or '터파기' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-sky-200"><rect width="460" height="140" fill="#F0F9FF"/><polygon points="20,90 140,90 150,130 310,130 320,90 440,90 440,140 20,140" fill="#BAE6FD"/><text x="170" y="115" font-size="12" font-weight="900" fill="#0369A1">인력 시탐 줄파기 굴착 단면</text><line x1="330" y1="90" x2="330" y2="130" stroke="#DC2626" stroke-width="2"/><text x="340" y="115" font-size="11" font-weight="bold" fill="#DC2626">깊이 1.5m</text></svg>'''
    elif '핫태핑' in st or '천공' in st or '가스' in st or '우회' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-amber-200"><rect width="460" height="140" fill="#FFFBEB"/><rect x="40" y="50" width="380" height="40" rx="6" fill="#FEF3C7" stroke="#D97706" stroke-width="2"/><circle cx="100" cy="70" r="24" fill="#F59E0B"/><text x="80" y="75" font-size="11" font-weight="bold" fill="#FFFFFF">Hot Tap</text><text x="140" y="75" font-size="12" font-weight="900" fill="#B45309">무정지 가스 우회(By-Pass) 차단 공법</text></svg>'''
    elif '수압' in st or 'cctv' in st or '관로' in st or '부설' in st or '상하수' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-cyan-200"><rect width="460" height="140" fill="#ECFEFF"/><circle cx="90" cy="70" r="28" fill="#06B6D4"/><text x="68" y="75" font-size="11" font-weight="bold" fill="#FFFFFF">10kgf/cm²</text><rect x="140" y="45" width="280" height="50" rx="6" fill="#FFFFFF" stroke="#0891B2"/><text x="160" y="75" font-size="12" font-weight="900" fill="#0E7490">관로 1시간 수압 유지 & 누수 제로</text></svg>'''
    elif '전력' in st or '한전' in st or '22.9' in st or '케이블' in st or '절연' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-red-200"><rect width="460" height="140" fill="#FEF2F2"/><rect x="50" y="40" width="360" height="60" rx="8" fill="#FFFFFF" stroke="#DC2626" stroke-width="2"/><text x="100" y="65" font-size="12" font-weight="900" fill="#991B1B">22.9kV 특고압 케이블 인입 & 절연내력 시험</text><text x="140" y="85" font-size="11" font-weight="bold" fill="#DC2626">30kV DC 연속 인가 (한전 입회)</text></svg>'''
    elif '승인' in st or '인수' in st or '인계' in st or '합동' in st or '완료' in st or '서명' in st:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-emerald-200"><rect width="460" height="140" fill="#ECFDF5"/><rect x="130" y="20" width="200" height="100" rx="6" fill="#FFFFFF" stroke="#059669" stroke-width="2"/><text x="155" y="45" font-size="12" font-weight="900" fill="#047857">지장물 이설 준공 대장</text><text x="155" y="70" font-size="11" font-weight="bold" fill="#334155">3자 합동 현측 & 서면 승인</text><circle cx="280" cy="90" r="16" fill="#DCFCE7" stroke="#059669" stroke-width="2"/><text x="268" y="95" font-size="11" font-weight="900" fill="#059669">승인</text></svg>'''
    else:
        svg_content = f'''<svg viewBox="0 0 460 140" class="w-full h-auto rounded bg-white border border-indigo-200"><rect width="460" height="140" fill="#EEF2FF"/><rect x="40" y="45" width="380" height="50" rx="6" fill="#FFFFFF" stroke="#4F46E5" stroke-width="1.5"/><text x="90" y="75" font-size="12" font-weight="900" fill="#3730A3">{step_text[:35]}</text></svg>'''

    return f'''<div class="clickable-diagram cursor-pointer transition transform hover:scale-[1.01] bg-white border border-slate-200 rounded-xl p-3 mb-3 shadow-sm hover:shadow" onclick="openDiagramZoom(this.outerHTML, '{title} - STEP {step_idx:02d} 2D 도식')">
        <div class="text-[11px] font-bold text-slate-500 mb-1 flex items-center justify-between">
            <span>🔍 클릭 시 대형 팝업 확대</span>
            <span class="text-indigo-600 bg-indigo-50 px-2 py-0.5 rounded border border-indigo-200 font-extrabold">STEP {step_idx:02d} 2D 도식</span>
        </div>
        {svg_content}
    </div>'''

def render_flexible_body(layout_type, steps, theme_color, title):
    if layout_type == 'dual_card':
        cards_html = f'''<div class="grid grid-cols-1 md:grid-cols-2 gap-6">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-indigo-50/50 border border-indigo-200 rounded-2xl p-6 shadow-sm hover:shadow-md transition">
                {svg}
                <div class="flex items-center gap-2 mb-2">
                    <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-3 py-1 rounded-full">PHASE {idx:02d}</span>
                    <h4 class="font-extrabold text-base text-slate-900">{s}</h4>
                </div>
            </div>'''
        cards_html += '</div>'
        return cards_html
        
    elif layout_type == 'triple_card':
        cards_html = f'''<div class="grid grid-cols-1 md:grid-cols-3 gap-5">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-sky-50/40 border border-sky-200 rounded-2xl p-5 shadow-sm hover:shadow-md transition">
                {svg}
                <div class="flex items-center gap-2 mb-2">
                    <span class="text-xs font-black text-sky-700 bg-sky-100 px-2.5 py-1 rounded-full">STEP {idx:02d}</span>
                    <h4 class="font-bold text-sm text-slate-900 leading-snug">{s}</h4>
                </div>
            </div>'''
        cards_html += '</div>'
        return cards_html

    elif layout_type == 'testing_protocol':
        cards_html = f'''<div class="space-y-4 bg-cyan-50/40 border border-cyan-200 rounded-2xl p-6">
            <h4 class="text-base font-black text-cyan-950 flex items-center gap-2 mb-4">
                <span>🔬</span> 엔지니어링 정밀 시험 및 검측 프로토콜 절차 ({len(steps)}단계 흐름)
            </h4>
            <div class="grid grid-cols-1 md:grid-cols-{min(len(steps), 3)} gap-4">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-white border border-cyan-200 rounded-xl p-4 shadow-sm">
                {svg}
                <span class="text-[11px] font-black text-cyan-700 bg-cyan-100 px-2 py-0.5 rounded-full inline-block mb-1.5">PROTOCOL {idx:02d}</span>
                <h5 class="font-bold text-xs text-slate-900 leading-tight">{s}</h5>
            </div>'''
        cards_html += '</div></div>'
        return cards_html

    elif layout_type == 'interface_matrix':
        cards_html = f'''<div class="space-y-4 bg-emerald-50/40 border border-emerald-200 rounded-2xl p-6">
            <h4 class="text-base font-black text-emerald-950 flex items-center gap-2 mb-4">
                <span>🤝</span> 선행·당해·후행 공종 3자 합동 인터페이스 인계인수 매트릭스
            </h4>
            <div class="grid grid-cols-1 md:grid-cols-{min(len(steps), 3)} gap-4">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-white border border-emerald-200 rounded-xl p-4 shadow-sm">
                {svg}
                <span class="text-[11px] font-black text-emerald-700 bg-emerald-100 px-2 py-0.5 rounded-full inline-block mb-1.5">INTERFACE {idx:02d}</span>
                <h5 class="font-bold text-xs text-slate-900 leading-tight">{s}</h5>
            </div>'''
        cards_html += '</div></div>'
        return cards_html

    elif layout_type == 'detailed_timeline':
        cards_html = f'''<div class="space-y-4">
            <h4 class="text-base font-black text-slate-900 flex items-center gap-2 mb-2">
                <span>⏱️</span> 고난도 복합 안전 시공 {len(steps)}단계 순차 공정 타임라인
            </h4>
            <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-3 gap-4">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-slate-50 border border-slate-200 rounded-xl p-4 hover:shadow-md transition">
                {svg}
                <div class="flex items-center gap-2 mb-1.5">
                    <span class="text-[11px] font-black text-amber-800 bg-amber-100 px-2 py-0.5 rounded-full">STEP {idx:02d}</span>
                    <h5 class="font-bold text-xs text-slate-900 leading-tight">{s}</h5>
                </div>
            </div>'''
        cards_html += '</div></div>'
        return cards_html

    else: # grid_cards (4 steps or general)
        cards_html = f'''<div class="grid grid-cols-1 md:grid-cols-2 gap-6">'''
        for idx, s in enumerate(steps, 1):
            svg = generate_step_svg(idx, s, theme_color, title)
            cards_html += f'''<div class="bg-slate-50 border border-slate-200 rounded-2xl p-5 hover:shadow-md transition">
                {svg}
                <div class="flex items-center gap-2 mb-2">
                    <span class="text-xs font-black text-indigo-700 bg-indigo-100 px-2.5 py-1 rounded-full">STEP {idx:02d}</span>
                    <h4 class="font-bold text-sm text-slate-900">{s}</h4>
                </div>
            </div>'''
        cards_html += '</div>'
        return cards_html

def build_flexible_html(act):
    title = act['task_title']
    method = act['method']
    purpose = act['purpose']
    output = act['output']
    risk = act['risk']
    advisory = act['advisory']
    dept = act['dept']
    
    layout_type, theme_color, steps = categorize_and_layout(title, method)
    body_cards_html = render_flexible_body(layout_type, steps, theme_color, title)
    
    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 지장물이설 - {title} 맞춤 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
    <style> body {{ font-family: 'Noto Sans KR', sans-serif; }} </style>
</head>
<body class="bg-slate-100 text-slate-800 antialiased py-8 px-4 sm:px-8">

    <div class="max-w-6xl mx-auto bg-white rounded-3xl shadow-xl border border-slate-200 p-6 sm:p-10 space-y-8">
        <!-- 🔵 헤더 영역 (유연한 단계 수 뱃지 표시) -->
        <div class="flex flex-col sm:flex-row sm:items-center justify-between pb-6 border-b border-slate-200 gap-4">
            <div>
                <span class="text-xs font-black px-3.5 py-1.5 rounded-full mb-2 inline-block border bg-indigo-50 text-indigo-700 border-indigo-200">
                    Dongtan Tram Utility Playbook | WBS {act['l3_code']} ({len(steps)}단계 유연 아키텍처)
                </span>
                <h1 class="text-2xl sm:text-4xl font-black text-slate-900 tracking-tight">
                    {title} 수행지침서
                </h1>
            </div>
            <div class="shrink-0 text-right text-xs font-bold text-slate-500">
                주관부서: <span class="text-indigo-600 font-extrabold">{dept}</span>
            </div>
        </div>

        <!-- 📌 1. 작업 개요 및 핵심 목적 -->
        <div class="bg-blue-50/70 border border-blue-200 rounded-2xl p-6 space-y-3">
            <h3 class="text-base font-black text-blue-950 flex items-center gap-2">
                <span>📌</span> 현장 이행 목적 및 주요 방침
            </h3>
            <p class="text-slate-800 text-sm font-semibold leading-relaxed">
                {purpose}
            </p>
        </div>

        <!-- 💡 2. 유연한 공학 세부 수행절차 및 2D 시공 도식 -->
        <div class="space-y-4">
            <h3 class="text-xl font-black text-slate-900 flex items-center gap-2">
                <span>💡</span> {len(steps)}단계 맞춤형 공학 수행절차 및 2D 시공 도식
            </h3>
            {body_cards_html}
        </div>

        <!-- 📦 3. 최종 성과품 -->
        <div class="bg-emerald-50 border border-emerald-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-emerald-950 flex items-center gap-2">
                <span>📂</span> 최종 필수 성과품 / 결재 대장
            </h3>
            <p class="text-emerald-900 text-sm font-bold leading-relaxed">
                {output}
            </p>
        </div>

        <!-- ⚠️ 4. 집행단계 리스크 예방 관리 -->
        <div class="bg-amber-50 border-2 border-amber-300 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>⚠️</span> 집행단계 핵심 리스크 및 예방 대책
            </h3>
            <p class="text-amber-900 text-sm font-bold leading-relaxed">
                {risk}
            </p>
        </div>

        <!-- 👷 5. 협력업체 실무 자문 노하우 -->
        <div class="bg-amber-50/60 border border-amber-200 rounded-2xl p-6 space-y-2">
            <h3 class="text-base font-black text-amber-950 flex items-center gap-2">
                <span>👷</span> [협력업체 실무 자문] 시공 및 공사관리 핵심 가이드
            </h3>
            <p class="text-amber-900 text-sm font-semibold leading-relaxed">
                {advisory}
            </p>
        </div>

        <!-- 📌 푸터 -->
        <div class="pt-6 border-t border-slate-200 flex items-center justify-between text-xs text-slate-500">
            <span>🏢 동탄도시철도(트램) 건설공사 지장물이설 수행지침</span>
            <span>v8 엑셀 100% 1:1 유연 아키텍처 연동</span>
        </div>
    </div>

    <!-- Lightbox Zoom Modal -->
    <div id="zoomModal" class="fixed inset-0 z-[9999] hidden bg-black/80 backdrop-blur-md flex items-center justify-center p-4" onclick="this.classList.add('hidden')">
        <div class="bg-white rounded-3xl p-6 max-w-4xl w-full max-h-[90vh] overflow-auto border-4 border-indigo-500 shadow-2xl relative" onclick="event.stopPropagation()">
            <button onclick="document.getElementById('zoomModal').classList.add('hidden')" class="absolute top-4 right-4 bg-slate-900 text-white font-black px-4 py-2 rounded-full text-sm hover:bg-slate-700">✕ 닫기</button>
            <h3 id="zoomTitle" class="text-xl font-black text-slate-900 mb-4 pb-2 border-b border-slate-200">고해상도 2D 시공 도식 확대보기</h3>
            <div id="zoomContent" class="flex justify-center items-center"></div>
        </div>
    </div>
    <script>
    function openDiagramZoom(htmlContent, title) {{
        var modal = document.getElementById('zoomModal');
        document.getElementById('zoomTitle').innerText = title || '2D 시공 도식 확대보기';
        document.getElementById('zoomContent').innerHTML = htmlContent;
        modal.classList.remove('hidden');
    }}
    </script>
</body>
</html>'''
    return html

# Generate flexible HTML files across all matching subfolders in v8/지장물이설
updated_folders = 0
for act in activities:
    title = act['task_title']
    matching_folders = []
    
    if os.path.exists(util_dir):
        for sub in os.listdir(util_dir):
            sub_p = os.path.join(util_dir, sub)
            if os.path.isdir(sub_p):
                if title in sub or sub.split('_', 1)[-1] in title or title in sub.split('_', 1)[-1]:
                    matching_folders.append(sub_p)
                    
        if not matching_folders:
            kw = title.split()[0]
            for sub in os.listdir(util_dir):
                sub_p = os.path.join(util_dir, sub)
                if os.path.isdir(sub_p) and kw in sub:
                    matching_folders.append(sub_p)
                    
    html_content = build_flexible_html(act)
    safe_title = re.sub(r'[\/:*?"<>|]', '_', title)
    
    for mf in matching_folders:
        guide_dir = os.path.join(mf, '수행지침')
        os.makedirs(guide_dir, exist_ok=True)
        
        f1 = os.path.join(guide_dir, f"{safe_title}_수행지침.html")
        f2 = os.path.join(guide_dir, "수행지침.html")
        
        with open(f1, 'w', encoding='utf-8') as f:
            f.write(html_content)
        with open(f2, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        for existing in os.listdir(guide_dir):
            if existing.endswith('.html'):
                with open(os.path.join(guide_dir, existing), 'w', encoding='utf-8') as f:
                    f.write(html_content)
        updated_folders += 1

print(f"Successfully generated flexible step-architecture HTML manuals for {updated_folders} 지장물이설 folders.", flush=True)
