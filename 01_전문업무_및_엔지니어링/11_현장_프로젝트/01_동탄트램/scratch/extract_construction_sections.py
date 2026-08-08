import sys, json
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

path1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
path2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

sections = []

for tool_label, path in [("1공구", path1), ("2공구", path2)]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    current_section = None

    for row in range(12, ws.max_row + 1):
        col_l = ws[f'L{row}'].value  # 시공구간 분할
        col_b = ws[f'B{row}'].value  # No
        col_c = ws[f'C{row}'].value  # 구간명
        col_g = ws[f'G{row}'].value  # 시점 STA
        col_h = ws[f'H{row}'].value  # 종점 STA
        col_i = ws[f'I{row}'].value  # 연장(구간)
        col_k = ws[f'K{row}'].value  # 구간특성
        col_m = ws[f'M{row}'].value  # 시공구간연장
        col_n = ws[f'N{row}'].value  # 소요공기

        if col_l and str(col_l).strip():
            # New section starts
            if current_section:
                sections.append(current_section)

            section_name = str(col_l).strip().replace('\n', ' ')
            current_section = {
                'tool': tool_label,
                'section': section_name,
                'startSta': float(col_g) if col_g else None,
                'endSta': None,
                'length': float(col_m) if col_m else None,
                'duration': round(float(col_n), 1) if col_n else None,
                'segments': [],
                'landmark': str(col_k).strip() if col_k else None
            }

        if current_section and col_b is not None and col_c:
            seg_info = {
                'no': str(col_b),
                'name': str(col_c).strip(),
                'startSta': float(col_g) if col_g else None,
                'endSta': float(col_h) if col_h else None,
                'length': float(col_i) if col_i else None,
            }
            current_section['segments'].append(seg_info)
            if col_h:
                current_section['endSta'] = float(col_h)

    if current_section:
        sections.append(current_section)

    wb.close()

print(f"Total sections extracted: {len(sections)}")
print()
for i, s in enumerate(sections):
    print(f"[{s['tool']}] {s['section']}")
    print(f"  STA: {s['startSta']} ~ {s['endSta']}")
    print(f"  Length: {s['length']}m, Duration: {s['duration']} months")
    print(f"  Landmark: {s['landmark']}")
    print(f"  Segments: {len(s['segments'])} items")
    for seg in s['segments']:
        print(f"    - [{seg['no']}] {seg['name']} STA {seg['startSta']}~{seg['endSta']} ({seg['length']}m)")
    print()

# Save as JSON for use
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\construction_sections.json', 'w', encoding='utf-8') as f:
    json.dump(sections, f, ensure_ascii=False, indent=2)
print("Saved to construction_sections.json")
