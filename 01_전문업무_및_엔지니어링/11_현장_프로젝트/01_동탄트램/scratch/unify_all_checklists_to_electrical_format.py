# -*- coding: utf-8 -*-
"""
동탄도시철도(트램) 매뉴얼BODY(집행단계-첨부폴더)v8
전 공종(11대 공종 375개 액티비티) 체크리스트 HTML 파일 전기분야 마스터 서식 일괄 변환 엔진
"""

import os
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8")
EXCEL_PATH = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")

# 1. 엑셀에서 전체 11개 시트 메타데이터 로드
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

TASK_DB = {}  # key: (disc_name, task_name) -> dict

for sname in wb.sheetnames:
    if sname == "종합통계대시보드":
        continue
    ws = wb[sname]
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 4).value
        task = ws.cell(r, 7).value
        dday = ws.cell(r, 8).value
        dept = ws.cell(r, 9).value
        purpose = ws.cell(r, 10).value
        method = ws.cell(r, 11).value
        deliverable = ws.cell(r, 12).value

        if task:
            task_clean = str(task).strip()
            TASK_DB[(sname, task_clean)] = {
                "sheet": sname,
                "code": str(code).strip() if code else "",
                "task": task_clean,
                "dday": str(dday).strip() if dday else "",
                "dept": str(dept).strip() if dept else "현장 공사팀",
                "purpose": str(purpose).strip() if purpose else "",
                "method": str(method).strip() if method else "",
                "deliverable": str(deliverable).strip() if deliverable else ""
            }

print(f"엑셀에서 로드된 액티비티 총 수: {len(TASK_DB)}개")

def clean_name(folder_or_file_name):
    name = re.sub(r'^\d+[\._\s]*', '', folder_or_file_name)
    name = name.replace('_표준서', '').replace('_수행지침', '').replace('_체크리스트', '').replace('.html', '')
    return name.strip()

# 12대 질문 생성기
def generate_12_questions(discipline, task_name, task_info):
    method_text = task_info.get("method", "") if task_info else ""
    purpose_text = task_info.get("purpose", "") if task_info else ""
    deliverable = task_info.get("deliverable", "검측보고서") if task_info else "검측보고서"

    # 공종별 특화 키워드
    q_specs = {
        "지반조사": ("지반조사 표준시방서(KCS 11 10 10) 및 시추도면", "시추기, 코어배럴, 표준관입시험기", "시추 심도(경암 확인) 및 TCR(≥85%), RQD", "표준관입시험 N치 타격수(63.5kg, 76cm)"),
        "사전토공사": ("토공사 표준시방서(KCS 11 20 00) 및 굴착계획도면", "백호, 덤프트럭, 다짐 롤러", "굴착선형 및 노상 다짐도(평판재하시험 K30)", "지하매설물 방호 및 가시설(토류판/H-Pile) 변위"),
        "지장물이설": ("지장물 이설 이행합의서 및 유관기관(맑은물/한전/가스) 도면", "관로 융착기, 크레인, 수압시험기", "관로 매설 심도(≥1.2m) 및 보호포/경고테이프", "이설 관로 통수/수압/기밀시험(1.0MPa, 60분)"),
        "상부강화노반": ("도시철도 설계기준(노반편) 및 강화노반 시방서", "모터그레이더, 진동롤러, 콘크리트 피니셔", "상부강화노반 평탄성(±3mm) 및 지지력계수(K30)", "콘크리트 타설 슬래브 철근 배근 간격 및 피복두께"),
        "콘크리트도상": ("철도시설 기술기준 및 콘크리트도상 시방서", "레일 연마기, 트랙게이지, 콘크리트 펌프카", "궤간(1,435mm ±1mm), 수평(±1mm), 면맞춤", "체결장치(텐션클램프) 체결 토크 및 절연패드"),
        "건축": ("건축공사 표준시방서 및 역사 마감 상세도면", "골조 거푸집, 비계, 타일/패널 시공장비", "구조체 수직/수평도(±3mm) 및 층고 기준", "방수층 두께, 단열재 밀착 및 외벽 실란트 씰링"),
        "신호": ("철도신호설비 시방서 및 T-ATP/ATO 설계도서", "신호 계측기, 광파워미터, 무선 통신 시험기", "신호기계실 랙 설치 수평도 및 접지저항(≤1.0Ω)", "LTE-R 무선 AP 패킷 손실률(≤1%) 및 안테나 지향각"),
        "전기": ("KEC(한국전기설비규정) 및 수배전반 제작사양도", "절연저항계(메거), 접지저항계, 토크렌치", "특고압 모선 절연거리 및 접지저항(≤1.0Ω)", "케이블 절연저항(≥5MΩ) 및 단자 체결 토크"),
        "통신": ("철도통신설비 표준시방서 및 광전송망 설계서", "OTDR(광손실측정기), 무선 전계강도 측정기", "광케이블 접속 손실(≤0.1dB/접속) 및 곡률반경", "CCTV/PA/안내방송 음압 레벨 및 영상 전송 지연"),
        "기계": ("기계설비공사 시방서 및 소방시설 기술기준", "배관 융착기, 수압시험기, 풍량풍압계", "배관 수평/수직도 및 지지행거 간격(≤2.0m)", "소방배관 수압시험(1.4MPa 2시간) 및 제연풍량"),
        "철도종합시운전": ("철도안전법 제38조 및 국토부 종합시험운행 지침", "속도측정기, 윤중횡압센서, 진동가속도계", "시험구간 선로 출입통제 및 방호벽 시건장치", "속도별 제동거리(70km/h ≤160m) 및 정위치정차")
    }

    spec = q_specs.get(discipline, ("국토교통부 표준시방서 및 승인도면", "공인 교정 검측장비", "시공 치수 및 허용오차 기준", "단위 성능 및 연동 기능"))

    items = [
        ("사전 도서 대조", f"{task_name} 관련 최신 승인도면 및 {spec[0]}에 따라 사전 검토를 철저히 수행하였는가?"),
        ("자재 검수 및 승인", f"현장 반입 자재가 자재공급원 승인원 및 공인 시험성적서 규격과 100% 일치함을 확인하였는가?"),
        ("선행공정 인계", f"선행 구조물(토목/건축/노반)의 수평도, 마감 상태 및 분계점 인수인계를 정밀 점검하였는가?"),
        ("시공장비 및 안전", f"투입 장비({spec[1]})의 공인 검교정 필증을 확인하고 작업장 안전방호 조치를 완료하였는가?"),
        ("정밀 수치 계측", f"{spec[2]} 허용오차 기준을 100% 충족하도록 정밀 실측 검측을 수행하였는가?"),
        ("핵심 성능 시험", f"{spec[3]} 측정 및 단위 기능시험을 엄격히 실시하여 시방 기준 합격을 입증하였는가?"),
        ("체결 및 조립 상태", f"규정 토크값에 따라 전용 공구로 정밀 체결하고 풀림방지 마킹(Torque Seal)을 완료하였는가?"),
        ("배관/배선/배치 정돈", f"설비 및 배관/배선의 굴절반경 규격 준수, 정돈 상태 및 지지간격의 적정성을 확인하였는가?"),
        ("방화/환경 씰링", f"방화구획 관통부 내화 2시간 이상의 방화충전재 밀폐 및 환경오염 방지 처리를 완료하였는가?"),
        ("인터페이스 상호 확인", f"인접 공종(궤도, 전기, 신호, 통신, 건축, 차량)과의 이격거리 확보 및 간섭 배제 조치를 확인하였는가?"),
        ("사진 채증 완료", f"시공 전, 시공 중, 시공 후의 주요 공정 및 숨은 검측 상태를 다각도에서 촬영 채증하였는가?"),
        ("3자 서명 날인", f"시공사, 품질관리자 및 책임감리원 3자 입회 검측 후 {deliverable}에 서명/날인을 완료하였는가?")
    ]
    return items

def build_checklist_html(discipline, code, task_name, std_file, guide_file, task_info):
    items = generate_12_questions(discipline, task_name, task_info)

    rows_html = []
    for idx, (category, question) in enumerate(items, 1):
        r_html = f"""                <tr class="hover:bg-slate-50 border-b border-slate-100">
                    <td class="p-4 text-center font-bold text-slate-500 text-xs">{idx}</td>
                    <td class="p-4 font-bold text-slate-800 text-xs sm:text-sm">{category}</td>
                    <td class="p-4 font-medium text-slate-700 text-xs sm:text-sm leading-relaxed">{question}</td>
                    <td class="p-4 text-center">
                        <span class="inline-flex gap-2">
                            <label class="inline-flex items-center text-xs font-semibold text-emerald-700 cursor-pointer"><input type="checkbox" checked class="rounded border-slate-300 mr-1 text-emerald-600 focus:ring-emerald-500"> 적합</label>
                            <label class="inline-flex items-center text-xs font-semibold text-rose-700 cursor-pointer"><input type="checkbox" class="rounded border-slate-300 mr-1 text-rose-600 focus:ring-rose-500"> 부적합</label>
                        </span>
                    </td>
                    <td class="p-4 text-center font-medium text-slate-500 text-xs">-</td>
                </tr>"""
        rows_html.append(r_html)

    tbody_content = "\n".join(rows_html)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{task_name} 마스터 체크리스트</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; background-color: #f8fafc; }}
        /* Standardized Bottom Navigation Box */
        .nav-box {{ display: flex; gap: 12px; margin-top: 36px; padding-top: 24px; border-top: 1px solid #e2e8f0; width: 100%; box-sizing: border-box; }}
        .nav-btn {{ flex: 1; text-align: center; padding: 14px 16px; border-radius: 6px; font-size: 13px; font-weight: 700; text-decoration: none; transition: 0.2s; display: inline-block; box-sizing: border-box; }}
        .btn-std {{ background: #0f172a !important; color: #ffffff !important; }}
        .btn-guide {{ background: #0284c7 !important; color: #ffffff !important; }}
        .btn-chk {{ background: #d97706 !important; color: #ffffff !important; }}
        .btn-std:hover {{ background: #1e293b !important; }}
        .btn-guide:hover {{ background: #0369a1 !important; }}
        .btn-chk:hover {{ background: #b45309 !important; }}
    </style>
</head>
<body class="p-6 sm:p-10 text-slate-800">
<div class="max-w-5xl mx-auto space-y-6">

    <!-- 대제목 & WBS 코드 -->
    <div class="flex justify-between items-end border-b-2 border-slate-900 pb-4">
        <div>
            <h1 class="text-3xl font-black text-slate-900 tracking-tight">{task_name} 마스터 체크리스트</h1>
        </div>
        <div class="text-right">
            <span class="text-blue-600 font-bold text-sm">WBS Code {code} | {discipline} 검측대장</span>
        </div>
    </div>

    <!-- 📋 상단 안내 상자 (Notice Box) -->
    <div class="bg-blue-50/80 border border-blue-200 rounded-2xl p-6 shadow-sm space-y-2">
        <h3 class="text-base font-bold text-blue-950 flex items-center gap-2">
            <span>📋</span> {task_name} 12대 정밀 검측대장
        </h3>
        <p class="text-xs text-blue-900 leading-relaxed font-medium">
            본 체크리스트는 국토교통부 표준시방서, 철도설계기준 및 동탄트램 특기시방서 기준을 12개 정밀 점검 항목으로 확장 구성하였으며, 모든 항목의 문장 끝은 예외 없이 질문형 어미(<strong class="text-blue-700">~하였는가?</strong>)로 100% 정형화되었습니다.
        </p>
    </div>

    <!-- 3-Column 마스터 검측 테이블 -->
    <div class="bg-white border border-slate-200 rounded-2xl shadow-xl overflow-hidden">
        <table class="w-full border-collapse">
            <thead>
                <tr class="bg-slate-100 text-slate-700 text-xs sm:text-sm font-bold border-b border-slate-200">
                    <th class="p-4 text-center w-12">NO</th>
                    <th class="p-4 text-left w-36">구분</th>
                    <th class="p-4 text-left">검측 및 확인 세부 항목 (질문형 어미 준수)</th>
                    <th class="p-4 text-center w-36">검측 결과</th>
                    <th class="p-4 text-center w-28">조치 사항</th>
                </tr>
            </thead>
            <tbody>
{tbody_content}
            </tbody>
        </table>
    </div>

    <!-- 점검자 및 감리원 서명란 -->
    <div class="bg-white border border-slate-200 rounded-2xl p-6 shadow-sm flex flex-col sm:flex-row justify-between items-center gap-4 text-xs font-semibold text-slate-700">
        <div>
            <span>📌 종합 판정 결과: </span>
            <span class="text-blue-600 font-bold text-sm ml-2">적합 (PASS)</span>
        </div>
        <div class="flex gap-8">
            <div>점검자(시공책임): <span class="border-b border-slate-400 pb-1 px-4 inline-block ml-1">김시공 (인)</span></div>
            <div>확인자(책임감리): <span class="border-b border-slate-400 pb-1 px-4 inline-block ml-1">이감리 (인)</span></div>
        </div>
    </div>

    <!-- Bottom Navigation Block -->
    <div class="nav-box">
      <a href="../표준서/{std_file}" class="nav-btn btn-std">📄 [연계] {task_name} 표준서 열기 ➔</a>
      <a href="../수행지침/{guide_file}" class="nav-btn btn-guide">📘 [연계] {task_name} 수행지침서 열기 ➔</a>
    </div>
</div>
</body>
</html>
"""
    return html

# 2. 전체 디렉토리 순회 및 체크리스트 파일 일괄 변환
converted_count = 0
skipped_count = 0

for root, dirs, files in os.walk(BASE_DIR):
    subdirs = [d for d in dirs if d in ['표준서', '수행지침', '체크리스트']]
    if len(subdirs) >= 2:
        activity_dir = root
        activity_folder_name = os.path.basename(activity_dir)
        task_name = clean_name(activity_folder_name)

        # 상위 공종 폴더명 파악
        rel_to_base = os.path.relpath(activity_dir, BASE_DIR)
        parts = rel_to_base.split(os.sep)
        disc_folder = parts[0]
        disc_clean = re.sub(r'^\d+[\._\s]*', '', disc_folder)

        # Freeze Policy: WBS 11 시공계획 수립 체크리스트 영구 동결 가드
        if "시공계획 수립" in task_name and "5.콘크리트도상" in disc_folder:
            print(f"  🔒 [FREEZE GUARD] WBS 11 {task_name} 체크리스트 건너뜀 (동결 보호)")
            skipped_count += 1
            continue

        std_dir = os.path.join(activity_dir, '표준서')
        guide_dir = os.path.join(activity_dir, '수행지침')
        chk_dir = os.path.join(activity_dir, '체크리스트')

        if not os.path.exists(chk_dir):
            continue

        std_files = [f for f in os.listdir(std_dir) if f.endswith('.html')] if os.path.exists(std_dir) else []
        guide_files = [f for f in os.listdir(guide_dir) if f.endswith('.html')] if os.path.exists(guide_dir) else []
        chk_files = [f for f in os.listdir(chk_dir) if f.endswith('.html')]

        def pick_main(files_list, keyword):
            if not files_list: return ""
            for f in files_list:
                if keyword in f and not f.startswith('9000-'):
                    return f
            for f in files_list:
                if keyword in f:
                    return f
            return files_list[0]

        main_std = pick_main(std_files, '표준서') or (std_files[0] if std_files else f"{task_name}_표준서.html")
        main_guide = pick_main(guide_files, '수행지침') or (guide_files[0] if guide_files else f"{task_name}_수행지침.html")

        # 엑셀 데이터 검색 (매칭 시도)
        task_info = TASK_DB.get((disc_clean, task_name))
        if not task_info:
            for (ds, ts), val in TASK_DB.items():
                if ts in task_name or task_name in ts:
                    task_info = val
                    break

        code_val = task_info.get("code", "") if task_info else f"9000-{disc_folder[:2]}"
        if not code_val:
            code_val = f"9000-{disc_folder[:2]}"

        for cf in chk_files:
            chk_full_path = os.path.join(chk_dir, cf)
            try:
                new_html = build_checklist_html(
                    discipline=disc_clean,
                    code=code_val,
                    task_name=task_name,
                    std_file=main_std,
                    guide_file=main_guide,
                    task_info=task_info
                )
                with open(chk_full_path, "w", encoding="utf-8") as fp:
                    fp.write(new_html)
                converted_count += 1
            except Exception as e:
                print(f"Error converting {chk_full_path}: {e}")

print(f"\n=======================================================")
print(f"전 공종 체크리스트 변환 완료: {converted_count}개 파일 (동결 보호: {skipped_count}개)")
print(f"=======================================================")
