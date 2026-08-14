# -*- coding: utf-8 -*-
"""
동탄트램 8대 공종 엑셀(v8.xlsx, v8.xlsm) 수식 충돌 완전 해결 및 순수 네이티브 단순 하이퍼링크 재구축
(수식 =HYPERLINK 제거 -> 순수 텍스트 + native hyperlink 객체 1:1 매칭)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import glob
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
XLSX_PATH = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")
XLSM_PATH = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
ATTACH_BASE = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")

SHEET_CONFIG = {
    "사전토공사": {
        "folder": "2.사전토공사",
        "col_std": 15, "col_guide": 17, "col_chk": 19
    },
    "지장물이설": {
        "folder": "3.지장물이설",
        "col_std": 12, "col_guide": 14, "col_chk": 16
    },
    "상부강화노반": {
        "folder": "4.상부강화노반",
        "col_std": 12, "col_guide": 14, "col_chk": 16
    },
    "콘크리트도상": {
        "folder": "5.콘크리트도상",
        "col_std": 12, "col_guide": 14, "col_chk": 16
    },
    "건축": {
        "folder": "6.건축",
        "col_std": 11, "col_guide": 13, "col_chk": 15
    },
    "신호": {
        "folder": "7.신호분야",
        "col_std": 11, "col_guide": 13, "col_chk": 15
    },
    "전기": {
        "folder": "8.전기분야",
        "col_std": 11, "col_guide": 13, "col_chk": 15
    },
    "통신": {
        "folder": "9.통신분야",
        "col_std": 11, "col_guide": 13, "col_chk": 15
    }
}

link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

for target_file in [XLSX_PATH, XLSM_PATH]:
    if not os.path.exists(target_file):
        continue
    
    is_vba = target_file.endswith('.xlsm')
    print(f"\n==================================================")
    print(f"파일 처리 중: {os.path.basename(target_file)}")
    wb = openpyxl.load_workbook(target_file, keep_vba=is_vba)

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        folder_abs = os.path.join(ATTACH_BASE, cfg["folder"])
        
        # 태스크 하위폴더 정렬
        subdirs = [d for d in os.listdir(folder_abs) if os.path.isdir(os.path.join(folder_abs, d))]
        def get_sort_key(name):
            m = re.match(r'^(\d+)', name)
            return int(m.group(1)) if m else 999
        subdirs.sort(key=get_sort_key)

        col_std = cfg["col_std"]
        col_guide = cfg["col_guide"]
        col_chk = cfg["col_chk"]

        # 헤더 명칭 정리
        ws.cell(1, col_std).value = "표준서 파일 (HTML)"
        ws.cell(1, col_guide).value = "수행지침 파일 (HTML)"
        ws.cell(1, col_chk).value = "체크리스트 파일 (HTML)"

        linked_count = 0

        for r in range(2, ws.max_row + 1):
            t_idx = r - 2
            if t_idx < len(subdirs):
                s_dir = subdirs[t_idx]
                d_path = os.path.join(folder_abs, s_dir)
                htmls = glob.glob(os.path.join(d_path, '**', '*.html'), recursive=True)

                std_f = next((f for f in htmls if '표준서' in f), None)
                guide_f = next((f for f in htmls if '수행지침' in f), None)
                chk_f = next((f for f in htmls if '체크리스트' in f), None)

                # 1. 표준서
                cell_std = ws.cell(r, col_std)
                if std_f and os.path.exists(std_f):
                    rel_path = os.path.relpath(std_f, BASE_DIR).replace('\\', '/')
                    cell_std.value = "👉 [클릭] 표준서 열기 📄"
                    cell_std.hyperlink = rel_path
                    cell_std.font = link_font
                    cell_std.alignment = align_center
                else:
                    cell_std.value = "-"
                    cell_std.hyperlink = None

                # 2. 수행지침
                cell_guide = ws.cell(r, col_guide)
                if guide_f and os.path.exists(guide_f):
                    rel_path = os.path.relpath(guide_f, BASE_DIR).replace('\\', '/')
                    cell_guide.value = "👉 [클릭] 수행지침 열기 📄"
                    cell_guide.hyperlink = rel_path
                    cell_guide.font = link_font
                    cell_guide.alignment = align_center
                else:
                    cell_guide.value = "-"
                    cell_guide.hyperlink = None

                # 3. 체크리스트
                cell_chk = ws.cell(r, col_chk)
                if chk_f and os.path.exists(chk_f):
                    rel_path = os.path.relpath(chk_f, BASE_DIR).replace('\\', '/')
                    cell_chk.value = "👉 [클릭] 체크리스트 열기 📄"
                    cell_chk.hyperlink = rel_path
                    cell_chk.font = link_font
                    cell_chk.alignment = align_center
                else:
                    cell_chk.value = "-"
                    cell_chk.hyperlink = None

                linked_count += 1

        print(f"  ✓ [{sheet_name}]: {linked_count}개 행 순수 네이티브 하이퍼링크 재구축 완료")

    wb.save(target_file)
    print(f"✓ {os.path.basename(target_file)} 저장 완료!")

print("\n==================================================")
print("전 공종 수식 충돌 Zero 순수 네이티브 하이퍼링크 주입 완료!")
print("==================================================")
