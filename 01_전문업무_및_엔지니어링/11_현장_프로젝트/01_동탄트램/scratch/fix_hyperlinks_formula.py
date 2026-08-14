# -*- coding: utf-8 -*-
"""
Excel에서 한글 경로 file:/// 하이퍼링크가 "지정한 파일을 열 수 없습니다" 오류 발생 시
-> cell.hyperlink 방식 대신 =HYPERLINK() 수식으로 교체하여 확실히 연동
"""

import os
import re
import openpyxl
from openpyxl.styles import Font

BASE_DIR = r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)"
EXCEL_FILES = [
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v6.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v5.xlsx"),
]

font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")

def convert_hyperlink_to_formula(excel_path):
    if not os.path.exists(excel_path):
        print(f"Skipping: {excel_path}")
        return

    excel_abs = os.path.abspath(excel_path)
    excel_dir = os.path.dirname(excel_abs)

    print(f"\nProcessing: {os.path.basename(excel_path)}")
    wb = openpyxl.load_workbook(excel_path)

    total = 0
    converted = 0

    for sname in wb.sheetnames:
        if sname in ['대시보드', '공정매뉴얼', 'GUIDE']:
            continue
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if not cell.hyperlink:
                    continue
                target = cell.hyperlink.target or ""
                if not target:
                    continue
                total += 1

                # 절대경로 정리
                if target.startswith("file:///"):
                    abs_path = target[8:].replace("/", "\\")
                elif not os.path.isabs(target):
                    abs_path = os.path.normpath(os.path.join(excel_dir, target))
                else:
                    abs_path = os.path.normpath(target)

                if not os.path.exists(abs_path):
                    continue

                # 셀 표시 텍스트
                label = str(cell.value or "👉 [클릭] 열기")
                if "표준서" in label or "표준서" in target:
                    label = "👉 [클릭] 표준서 열기 📄"
                elif "수행지침" in label or "수행지침" in target:
                    label = "👉 [클릭] 수행지침 열기 📄"
                elif "체크리스트" in label or "체크리스트" in target:
                    label = "👉 [클릭] 체크리스트 열기 📄"

                # =HYPERLINK() 수식으로 교체 (한글 경로도 완벽 지원)
                # 역슬래시 이스케이프: 수식 내 큰따옴표는 "" 처리
                formula_path = abs_path.replace('"', '""')
                cell.value = f'=HYPERLINK("{formula_path}","{label}")'
                cell.hyperlink = None  # 기존 cell.hyperlink 제거
                cell.font = font_link
                converted += 1

        print(f"  Sheet [{sname}]: {converted}개 수식 변환")
        converted = 0

    wb.save(excel_path)
    print(f"  [SAVED] {os.path.basename(excel_path)}")

for f in EXCEL_FILES:
    convert_hyperlink_to_formula(f)

print("\n=== DONE: All hyperlinks converted to =HYPERLINK() formula! ===")
