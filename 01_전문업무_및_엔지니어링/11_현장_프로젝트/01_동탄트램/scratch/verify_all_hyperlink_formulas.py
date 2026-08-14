# -*- coding: utf-8 -*-
import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

base = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")

for fname in ["매뉴얼 BODY (집행단계)v8.xlsx", "매뉴얼 BODY (집행단계)v8.xlsm"]:
    fpath = os.path.join(base, fname)
    if not os.path.exists(fpath): continue
    wb = openpyxl.load_workbook(fpath, data_only=False, keep_vba=fname.endswith('.xlsm'))

    total = 0
    broken = 0

    for s in ["사전토공사", "지장물이설", "상부강화노반", "콘크리트도상", "건축", "신호", "전기", "통신"]:
        ws = wb[s]
        sheet_links = 0
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(r, c).value or "")
                if val.startswith("=HYPERLINK("):
                    p_start = val.find('"') + 1
                    p_end = val.find('"', p_start)
                    if p_start > 0 and p_end > p_start:
                        rel_p = val[p_start:p_end]
                        full_p = os.path.join(base, rel_p)
                        sheet_links += 1
                        total += 1
                        if not os.path.exists(full_p):
                            print(f"Broken: [{fname}][{s}] R{r}C{c} -> {rel_p}")
                            broken += 1

    print(f"[{fname}] 총 {total}개 =HYPERLINK() 수식 검증 완료 | 깨진 파일: {broken}개")
