# -*- coding: utf-8 -*-
"""
동탄트램 8대 공종 전용 클린 하이퍼링크 엑셀 생성기
(별도 신규 엑셀 파일: 순번, 액티비티명, 표준서, 수행지침, 체크리스트 5개 열 깔끔 구성)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import glob
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
OUT_XLSX = os.path.join(BASE_DIR, "동탄트램_공종별_매뉴얼_하이퍼링크_목록.xlsx")
ATTACH_BASE = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")

os.system('taskkill /f /im excel.exe 2>nul')

DISCIPLINES = [
    ("상부강화노반", "4.상부강화노반"),
    ("사전토공사", "2.사전토공사"),
    ("지장물이설", "3.지장물이설"),
    ("콘크리트도상", "5.콘크리트도상"),
    ("건축", "6.건축"),
    ("신호", "7.신호분야"),
    ("전기", "8.전기분야"),
    ("통신", "9.통신분야")
]

# 스타일 정의
font_title = Font(name="맑은 고딕", size=14, bold=True, color="1E293B")
font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
font_data = Font(name="맑은 고딕", size=10, color="0F172A")
font_data_bold = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")
font_link_std = Font(name="맑은 고딕", size=10, bold=True, color="047857", underline="single")    # 녹색
font_link_guide = Font(name="맑은 고딕", size=10, bold=True, color="0284C7", underline="single")  # 파랑
font_link_chk = Font(name="맑은 고딕", size=10, bold=True, color="D97706", underline="single")    # 주황

fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # 다크네이비
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # 연한회색

border_thin = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

wb = openpyxl.Workbook()
# Remove default sheet
wb.remove(wb.active)

total_activities = 0
total_links = 0

for sheet_title, folder_name in DISCIPLINES:
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    
    # 1. 시트 타이틀
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"동탄도시철도(트램) {sheet_title} 엔지니어링 매뉴얼 링크 대장"
    title_cell.font = font_title
    title_cell.alignment = align_left
    ws.row_dimensions[1].height = 35

    # 2. 헤더 행
    headers = ["순번", "액티비티명 (Activity)", "표준서 (Standard)", "수행지침 (Guideline)", "체크리스트 (Checklist)"]
    ws.row_dimensions[2].height = 28
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(2, col_idx)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # 3. 폴더 내 태스크 수집
    folder_abs = os.path.join(ATTACH_BASE, folder_name)
    subdirs = [d for d in os.listdir(folder_abs) if os.path.isdir(os.path.join(folder_abs, d))]
    
    def get_sort_key(name):
        m = re.match(r'^(\d+)', name)
        return int(m.group(1)) if m else 999
    subdirs.sort(key=get_sort_key)

    row_num = 3
    for idx, s_dir in enumerate(subdirs, 1):
        total_activities += 1
        ws.row_dimensions[row_num].height = 24
        
        # 태스크명 정제
        m = re.match(r'^\d+[._\s]+(.+)$', s_dir)
        act_name = m.group(1).strip() if m else s_dir.strip()

        d_path = os.path.join(folder_abs, s_dir)
        htmls = glob.glob(os.path.join(d_path, '**', '*.html'), recursive=True)

        std_f = next((f for f in htmls if '표준서' in f), None)
        guide_f = next((f for f in htmls if '수행지침' in f), None)
        chk_f = next((f for f in htmls if '체크리스트' in f), None)

        is_even = (idx % 2 == 0)
        row_fill = fill_zebra if is_even else None

        # Col A: 순번
        cA = ws.cell(row_num, 1, value=idx)
        cA.font = font_data_bold
        cA.alignment = align_center
        cA.border = border_thin
        if row_fill: cA.fill = row_fill

        # Col B: 액티비티명
        cB = ws.cell(row_num, 2, value=act_name)
        cB.font = font_data_bold
        cB.alignment = align_left
        cB.border = border_thin
        if row_fill: cB.fill = row_fill

        # Col C: 표준서
        cC = ws.cell(row_num, 3)
        cC.border = border_thin
        cC.alignment = align_center
        if row_fill: cC.fill = row_fill
        if std_f and os.path.exists(std_f):
            rel_path = os.path.relpath(std_f, BASE_DIR)
            cC.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "📄 표준서 열기")'
            cC.font = font_link_std
            total_links += 1
        else:
            cC.value = "-"
            cC.font = font_data

        # Col D: 수행지침
        cD = ws.cell(row_num, 4)
        cD.border = border_thin
        cD.alignment = align_center
        if row_fill: cD.fill = row_fill
        if guide_f and os.path.exists(guide_f):
            rel_path = os.path.relpath(guide_f, BASE_DIR)
            cD.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "📘 수행지침 열기")'
            cD.font = font_link_guide
            total_links += 1
        else:
            cD.value = "-"
            cD.font = font_data

        # Col E: 체크리스트
        cE = ws.cell(row_num, 5)
        cE.border = border_thin
        cE.alignment = align_center
        if row_fill: cE.fill = row_fill
        if chk_f and os.path.exists(chk_f):
            rel_path = os.path.relpath(chk_f, BASE_DIR)
            cE.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "📋 체크리스트 열기")'
            cE.font = font_link_chk
            total_links += 1
        else:
            cE.value = "-"
            cE.font = font_data

        row_num += 1

    # 열 너비 자동 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    print(f"  ✓ [{sheet_title}]: {len(subdirs)}개 액티비티 시트 생성 완료")

# 저장
wb.save(OUT_XLSX)
print(f"\n=======================================================")
print(f"동탄트램 신규 하이퍼링크 엑셀 파일 생성 완료!")
print(f"저장 위치: {OUT_XLSX}")
print(f"총 공종 수: {len(DISCIPLINES)}개 시트")
print(f"총 액티비티: {total_activities}개")
print(f"총 하이퍼링크: {total_links}개")
print(f"=======================================================")
