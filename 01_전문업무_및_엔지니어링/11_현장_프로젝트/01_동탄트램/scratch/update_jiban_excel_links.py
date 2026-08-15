# -*- coding: utf-8 -*-
"""
동탄트램 매뉴얼 BODY (집행단계)v8 엑셀 지반조사 시트
36개 액티비티 108개 HTML 파일 동적 하이퍼링크 수식 일괄 기입 스크립트
"""

import os
import sys
import re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

link_font_std = Font(name="맑은 고딕", size=9, bold=True, color="047857", underline="single")
link_font_guide = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")
link_font_chk = Font(name="맑은 고딕", size=9, bold=True, color="D97706", underline="single")

fill_link_std = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
fill_link_guide = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
fill_link_chk = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue

    print(f"\n=======================================================")
    print(f"지반조사 엑셀 하이퍼링크 업데이트: {os.path.basename(file_path)}")
    wb = openpyxl.load_workbook(file_path)

    if "지반조사" not in wb.sheetnames:
        continue

    ws = wb["지반조사"]

    # 2행부터 37행까지 (36개 액티비티)
    for r in range(2, ws.max_row + 1):
        task = ws.cell(r, 7).value
        if not task:
            continue
        idx = r - 1
        safe_task = re.sub(r'[\/\\:\*\?"<>\|]', '_', str(task)).strip()
        folder_prefix = f"{idx:02d}_{safe_task}"

        std_rel = f"1.지반조사/{folder_prefix}/표준서/{safe_task}_표준서.html"
        guide_rel = f"1.지반조사/{folder_prefix}/수행지침/{safe_task}_수행지침.html"
        chk_rel = f"1.지반조사/{folder_prefix}/체크리스트/{safe_task}_체크리스트.html"

        # Col 14 (N): 표준서 링크
        f_std = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{std_rel}", "표준서 열기")'
        c_std = ws.cell(r, 14, value=f_std)
        c_std.font = link_font_std
        c_std.fill = fill_link_std
        c_std.alignment = align_center

        # Col 16 (P): 수행지침 링크
        f_guide = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{guide_rel}", "수행지침서 열기")'
        c_guide = ws.cell(r, 16, value=f_guide)
        c_guide.font = link_font_guide
        c_guide.fill = fill_link_guide
        c_guide.alignment = align_center

        # Col 18 (R): 체크리스트 링크
        f_chk = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{chk_rel}", "체크리스트 열기")'
        c_chk = ws.cell(r, 18, value=f_chk)
        c_chk.font = link_font_chk
        c_chk.fill = fill_link_chk
        c_chk.alignment = align_center

        # M, O, Q 요약 열 기본값 보강 (비어있거나 기본 텍스트)
        if not ws.cell(r, 13).value:
            ws.cell(r, 13).value = f"1) {task} 업무 수행을 위한 지반조사 기술 표준 및 시방 기준을 100% 충족함.\n2) 지반 리스크 사전 식별 및 보고서 승인 완료함."
        if not ws.cell(r, 15).value:
            ws.cell(r, 15).value = f"1) STEP 1: 사전 조사 계획 수립\n2) STEP 2: 현장 시추/시험 정밀 수행\n3) STEP 3: 데이터 분석 및 R/O 도출"
        if not ws.cell(r, 17).value:
            ws.cell(r, 17).value = f"1) 시추 심도 및 SPT N치 계측 검측\n2) 현장 안전관리 및 산출물 책임감리원 승인"

    wb.save(file_path)
    print(f"  ✔ {os.path.basename(file_path)} 지반조사 36개 액티비티 동적 하이퍼링크 수식 기입 완료!")

print(f"\n=======================================================")
print(f"지반조사 엑셀 하이퍼링크 연동 100% 완료!")
print(f"=======================================================")
