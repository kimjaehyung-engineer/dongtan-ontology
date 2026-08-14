# -*- coding: utf-8 -*-
"""하이퍼링크 경로 vs 실제 파일 존재 여부 비교 진단"""
import openpyxl, os, sys, re
sys.stdout.reconfigure(encoding="utf-8")

p = r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx"
excel_dir = os.path.dirname(os.path.abspath(p))
wb = openpyxl.load_workbook(p)

broken = 0
ok = 0
broken_list = []

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "")
            if not v.startswith("=HYPERLINK("):
                continue
            # 경로 추출: =HYPERLINK("path","label")
            m = re.search(r'=HYPERLINK\("([^"]+)"', v)
            if not m:
                continue
            path = m.group(1)
            if not os.path.isabs(path):
                path = os.path.join(excel_dir, path)
            if os.path.exists(path):
                ok += 1
            else:
                broken += 1
                broken_list.append((sname, cell.coordinate, path))

print(f"=== 진단 결과: OK={ok}, BROKEN={broken} ===")
if broken > 0:
    print(f"\n--- BROKEN 링크 샘플 (최대 10개) ---")
    for sname, coord, path in broken_list[:10]:
        print(f"  [{sname}] {coord}: {path}")
else:
    print("모든 하이퍼링크가 유효합니다!")
