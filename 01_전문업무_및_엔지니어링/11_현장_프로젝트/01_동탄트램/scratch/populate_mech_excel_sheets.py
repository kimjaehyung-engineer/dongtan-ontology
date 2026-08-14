# -*- coding: utf-8 -*-
"""
동탄트램 기계설비·소방설비 36개 액티비티 엑셀 데이터 완벽 입력 및 하이퍼링크 주입 스크립트
(통신 시트 22개 열 완벽 준용 + 동적 절대경로 수식 100% 바인딩)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from mech_part1 import MECH_TASKS_PART1
from mech_part2 import MECH_TASKS_PART2

ALL_TASKS = MECH_TASKS_PART1 + MECH_TASKS_PART2

# Kill any stuck EXCEL processes
os.system('taskkill /f /im excel.exe 2>nul')

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
ATTACH_DIR = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")
MECH_DIR_NAME = "10.기계분야"

TARGET_FILES = [
    os.path.join(ATTACH_DIR, "매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
]

# 스타일 정의
font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
font_data = Font(name="맑은 고딕", size=9, color="0F172A")
font_data_bold = Font(name="맑은 고딕", size=9, bold=True, color="0F172A")

link_font_std = Font(name="맑은 고딕", size=9, bold=True, color="047857", underline="single")
link_font_guide = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")
link_font_chk = Font(name="맑은 고딕", size=9, bold=True, color="D97706", underline="single")

fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

border_thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

HEADERS = [
    "L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "선행", "후행",
    "작업단위 (Level 4 Task/Activity)", "일정 (D-Day)", "주관", "목적",
    "방법", "산출물(결과)", "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)", "체크리스트 (Checklist) 요약",
    "체크리스트 파일 (HTML)", "비고", "첨부서류 연계 상세 설계기준",
    "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
]

for target_file in TARGET_FILES:
    if not os.path.exists(target_file):
        continue

    is_vba = target_file.endswith('.xlsm')
    print(f"\n==================================================")
    print(f"파일 작업 중: {os.path.basename(target_file)}")
    wb = openpyxl.load_workbook(target_file, keep_vba=is_vba)

    # 기계 시트 가져오기 또는 생성
    if "기계" in wb.sheetnames:
        ws = wb["기계"]
        # 기존 데이터 클리어
        for r in range(1, ws.max_row + 10):
            for c in range(1, ws.max_column + 10):
                ws.cell(r, c).value = None
                ws.cell(r, c).hyperlink = None
    else:
        ws = wb.create_sheet("기계")

    ws.views.sheetView[0].showGridLines = True

    # 1. 헤더 작성
    ws.row_dimensions[1].height = 28
    for c_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # 2. 36개 액티비티 데이터 입력
    for idx, item in enumerate(ALL_TASKS, 1):
        t_idx, folder_name, wbs_code, task_title, subtitle, purpose, kpis, specs, steps, diagram_name, terms = item
        row_num = idx + 1
        ws.row_dimensions[row_num].height = 36
        is_even = (idx % 2 == 0)
        row_fill = fill_zebra if is_even else None

        # 선행/후행
        pred = f"9000-8-{idx-1}" if idx > 1 else "-"
        succ = f"9000-8-{idx+1}" if idx < len(ALL_TASKS) else "-"
        dday = f"D-{120 - idx*3}" if idx <= 20 else f"D+{ (idx-20)*5 }"

        # 요약문 생성
        std_sum = f"1) {task_title} ({wbs_code}) 과업 달성을 위해 KDS 47 철도설계기준 및 NFTC 화재안전기준을 100% 충족한다.\n2) {subtitle}을 엄격히 완수하여 책임감리원 최종 승인을 획득한다."
        guide_sum = f"1) {steps[0][0]}: {steps[0][1]}\n2) {steps[1][0]}: {steps[1][1]}"
        chk_sum = f"1) {specs[0][0]} 시방 및 {specs[0][1]} 기준을 100% 충족하였는가?\n2) {specs[1][0]} 검측 및 불합격 요소를 전면 차단하였는가?"

        # 상대경로 계산 (파일 위치 기준)
        file_dir = os.path.dirname(os.path.abspath(target_file))
        is_inside_attach = (file_dir == ATTACH_DIR)

        if is_inside_attach:
            rel_std = f"{MECH_DIR_NAME}\\{folder_name}\\표준서\\{folder_name}_표준서.html"
            rel_guide = f"{MECH_DIR_NAME}\\{folder_name}\\수행지침\\{folder_name}_수행지침.html"
            rel_chk = f"{MECH_DIR_NAME}\\{folder_name}\\체크리스트\\{folder_name}_체크리스트.html"
        else:
            rel_std = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\표준서\\{folder_name}_표준서.html"
            rel_guide = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\수행지침\\{folder_name}_수행지침.html"
            rel_chk = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\체크리스트\\{folder_name}_체크리스트.html"

        row_values = [
            (1, "9000", font_data_bold, align_center),
            (2, "9000-8", font_data_bold, align_center),
            (3, "기계설비·소방설비", font_data_bold, align_center),
            (4, wbs_code, font_data_bold, align_center),
            (5, pred, font_data, align_center),
            (6, succ, font_data, align_center),
            (7, task_title, font_data_bold, align_left),
            (8, dday, font_data, align_center),
            (9, "기계/소방팀", font_data, align_center),
            (10, purpose, font_data, align_left),
            (11, f"KCS 41 시방 기준 및 {steps[0][0]}, {steps[1][0]} 절차서 준수 시공", font_data, align_left),
            (12, f"{task_title} 시공계획서, 검측요청서 및 공인 시험성적서", font_data, align_left),
            (13, std_sum, font_data, align_left),
            (14, f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_std}", "👉 [클릭] 표준서 열기 📄")', link_font_std, align_center),
            (15, guide_sum, font_data, align_left),
            (16, f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_guide}", "👉 [클릭] 수행지침 열기 📄")', link_font_guide, align_center),
            (17, chk_sum, font_data, align_left),
            (18, f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_chk}", "👉 [클릭] 체크리스트 열기 📄")', link_font_chk, align_center),
            (19, "기계·소방 복합 시방 준수", font_data, align_center),
            (20, "● KDS 47 철도설계기준, KCS 41 건축기계시방 및 NFTC/NFPC 화재안전기술기준 100% 준수", font_data, align_left),
            (21, "1) 👉 [더블클릭] 체크리스트 열기 📄", font_data_bold, align_center),
            (22, "기계설비·소방 전문 감리 및 시공 기술자문단 상시 운영", font_data, align_left)
        ]

        for col_idx, val, f_style, a_style in row_values:
            c = ws.cell(row_num, col_idx)
            c.value = val
            c.font = f_style
            c.alignment = a_style
            c.border = border_thin
            if row_fill: c.fill = row_fill

    # 열 너비 자동 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 28
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 36
    ws.column_dimensions['K'].width = 32
    ws.column_dimensions['L'].width = 28
    ws.column_dimensions['M'].width = 36
    ws.column_dimensions['N'].width = 24
    ws.column_dimensions['O'].width = 36
    ws.column_dimensions['P'].width = 24
    ws.column_dimensions['Q'].width = 36
    ws.column_dimensions['R'].width = 24
    ws.column_dimensions['S'].width = 16
    ws.column_dimensions['T'].width = 32
    ws.column_dimensions['U'].width = 24
    ws.column_dimensions['V'].width = 30

    wb.save(target_file)
    print(f"  ✓ {os.path.basename(target_file)} 기계 시트 36개 액티비티 100% 저장 완료!")

# 또한 전용 하이퍼링크 목록 엑셀 파일에도 기계 시트 추가
NAV_XLSX = os.path.join(BASE_DIR, "동탄트램_공종별_매뉴얼_하이퍼링크_목록.xlsx")
if os.path.exists(NAV_XLSX):
    wb_nav = openpyxl.load_workbook(NAV_XLSX)
    if "기계" in wb_nav.sheetnames:
        wb_nav.remove(wb_nav["기계"])
    ws_nav = wb_nav.create_sheet("기계")
    ws_nav.views.sheetView[0].showGridLines = True

    # 타이틀
    ws_nav.merge_cells("A1:E1")
    ws_nav["A1"].value = "동탄도시철도(트램) 기계설비·소방설비 엔지니어링 매뉴얼 링크 대장"
    ws_nav["A1"].font = Font(name="맑은 고딕", size=14, bold=True, color="1E293B")
    ws_nav["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_nav.row_dimensions[1].height = 35

    # 헤더
    ws_nav.row_dimensions[2].height = 28
    nav_headers = ["순번", "액티비티명 (Activity)", "표준서 (Standard)", "수행지침 (Guideline)", "체크리스트 (Checklist)"]
    for c_idx, h_text in enumerate(nav_headers, 1):
        c = ws_nav.cell(2, c_idx, value=h_text)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border_thin

    for idx, item in enumerate(ALL_TASKS, 1):
        t_idx, folder_name, wbs_code, task_title, subtitle, purpose, kpis, specs, steps, diagram_name, terms = item
        r_num = idx + 2
        ws_nav.row_dimensions[r_num].height = 24
        is_even = (idx % 2 == 0)
        row_fill = fill_zebra if is_even else None

        rel_std = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\표준서\\{folder_name}_표준서.html"
        rel_guide = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\수행지침\\{folder_name}_수행지침.html"
        rel_chk = f"매뉴얼BODY(집행단계-첨부폴더)v8\\{MECH_DIR_NAME}\\{folder_name}\\체크리스트\\{folder_name}_체크리스트.html"

        # A: 순번
        cA = ws_nav.cell(r_num, 1, value=idx)
        cA.font = font_data_bold; cA.alignment = align_center; cA.border = border_thin
        if row_fill: cA.fill = row_fill

        # B: 액티비티명
        cB = ws_nav.cell(r_num, 2, value=task_title)
        cB.font = font_data_bold; cB.alignment = align_left; cB.border = border_thin
        if row_fill: cB.fill = row_fill

        # C: 표준서
        cC = ws_nav.cell(r_num, 3, value=f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_std}", "📄 표준서 열기")')
        cC.font = link_font_std; cC.alignment = align_center; cC.border = border_thin
        if row_fill: cC.fill = row_fill

        # D: 수행지침
        cD = ws_nav.cell(r_num, 4, value=f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_guide}", "📘 수행지침 열기")')
        cD.font = link_font_guide; cD.alignment = align_center; cD.border = border_thin
        if row_fill: cD.fill = row_fill

        # E: 체크리스트
        cE = ws_nav.cell(r_num, 5, value=f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_chk}", "📋 체크리스트 열기")')
        cE.font = link_font_chk; cE.alignment = align_center; cE.border = border_thin
        if row_fill: cE.fill = row_fill

    ws_nav.column_dimensions['A'].width = 8
    ws_nav.column_dimensions['B'].width = 42
    ws_nav.column_dimensions['C'].width = 22
    ws_nav.column_dimensions['D'].width = 22
    ws_nav.column_dimensions['E'].width = 22

    wb_nav.save(NAV_XLSX)
    print(f"  ✓ 동탄트램_공종별_매뉴얼_하이퍼링크_목록.xlsx 내 [기계] 시트 추가 완료!")

print(f"\n==================================================")
print(f"전체 대상 엑셀 파일 기계 시트 36개 액티비티 연동 완료!")
print(f"==================================================")
