# -*- coding: utf-8 -*-
"""
매뉴얼BODY(집행단계-첨부폴더)v8 폴더 내 매뉴얼 BODY (집행단계)v8.xlsx 파일 전 공종 하이퍼링크 완벽 주입 스크립트
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import glob
import re
import openpyxl
from openpyxl.styles import Font, Alignment

# 대상 엑셀 파일 (사용자가 열고 있는 첨부폴더 내부 엑셀)
EXCEL_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8")
TARGET_XLSX = os.path.join(EXCEL_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")

# 혹시 열려있는 엑셀 강제 종료하여 파일 락 해제
os.system('taskkill /f /im excel.exe 2>nul')

print(f"작업 대상 엑셀 파일: {TARGET_XLSX}")
print(f"파일 존재 여부: {os.path.exists(TARGET_XLSX)}")

DISCIPLINE_MAP = {
    "사전토공사": "2.사전토공사",
    "상부강화노반": "4.상부강화노반",
    "지장물이설": "3.지장물이설",
    "콘크리트도상": "5.콘크리트도상",
    "건축": "6.건축",
    "신호": "7.신호분야",
    "전기": "8.전기분야",
    "통신": "9.통신분야"
}

link_font_std = Font(name="맑은 고딕", size=10, bold=True, color="047857", underline="single")
link_font_guide = Font(name="맑은 고딕", size=10, bold=True, color="0284C7", underline="single")
link_font_chk = Font(name="맑은 고딕", size=10, bold=True, color="D97706", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

wb = openpyxl.load_workbook(TARGET_XLSX, keep_vba=False)

total_linked = 0

for sheet_name, folder_name in DISCIPLINE_MAP.items():
    if sheet_name not in wb.sheetnames:
        print(f"시트 없음: {sheet_name}")
        continue
    
    ws = wb[sheet_name]
    folder_abs = os.path.join(EXCEL_DIR, folder_name)
    if not os.path.exists(folder_abs):
        print(f"공종 폴더 없음: {folder_abs}")
        continue

    subdirs = [d for d in os.listdir(folder_abs) if os.path.isdir(os.path.join(folder_abs, d))]
    def get_sort_key(name):
        m = re.match(r'^(\d+)', name)
        return int(m.group(1)) if m else 999
    subdirs.sort(key=get_sort_key)

    # 열 위치 탐색
    col_std = None
    col_guide = None
    col_chk = None

    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(1, c).value or "")
        if "표준서" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
            col_std = c
        elif "수행지침" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
            col_guide = c
        elif "체크리스트" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
            col_chk = c

    # 기본 매핑
    if sheet_name == "사전토공사":
        if not col_std: col_std = 15
        if not col_guide: col_guide = 17
        if not col_chk: col_chk = 19
    else:
        if not col_std: col_std = 14
        if not col_guide: col_guide = 16
        if not col_chk: col_chk = 18

    # 헤더명 표준화
    ws.cell(1, col_std).value = "표준서 파일 (HTML)"
    ws.cell(1, col_guide).value = "수행지침 파일 (HTML)"
    ws.cell(1, col_chk).value = "체크리스트 파일 (HTML)"

    print(f"[{sheet_name}] ({folder_name}) -> 표준서={col_std}열, 수행지침={col_guide}열, 체크리스트={col_chk}열 (하위폴더 {len(subdirs)}개)")

    linked_count = 0
    for r in range(2, ws.max_row + 1):
        t_idx = r - 2
        if t_idx < len(subdirs):
            s_dir = subdirs[t_idx]
            d_path = os.path.join(folder_abs, s_dir)
            htmls = glob.glob(os.path.join(d_path, '**', '*.html'), recursive=True)

            std_f = next((f for f in htmls if '표준서' in f), None)
            guide_f = next((f for f in htmls if '수행지침' in f), None)
            chk_f = next((f for f in htmls if '체크리스트' in f), None)

            # 1. 표준서
            cell_std = ws.cell(r, col_std)
            cell_std.hyperlink = None
            if std_f and os.path.exists(std_f):
                rel_path = os.path.relpath(std_f, EXCEL_DIR)
                cell_std.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 표준서 열기 📄")'
                cell_std.font = link_font_std
                cell_std.alignment = align_center
                total_linked += 1
            else:
                cell_std.value = "-"

            # 2. 수행지침
            cell_guide = ws.cell(r, col_guide)
            cell_guide.hyperlink = None
            if guide_f and os.path.exists(guide_f):
                rel_path = os.path.relpath(guide_f, EXCEL_DIR)
                cell_guide.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 수행지침 열기 📄")'
                cell_guide.font = link_font_guide
                cell_guide.alignment = align_center
                total_linked += 1
            else:
                cell_guide.value = "-"

            # 3. 체크리스트
            cell_chk = ws.cell(r, col_chk)
            cell_chk.hyperlink = None
            if chk_f and os.path.exists(chk_f):
                rel_path = os.path.relpath(chk_f, EXCEL_DIR)
                cell_chk.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 체크리스트 열기 📄")'
                cell_chk.font = link_font_chk
                cell_chk.alignment = align_center
                total_linked += 1
            else:
                cell_chk.value = "-"

            linked_count += 1

    print(f"  ✓ [{sheet_name}]: {linked_count}개 행 링크 주입 완료")

wb.save(TARGET_XLSX)
print(f"\n==================================================")
print(f"첨부폴더 내부 v8.xlsx 하이퍼링크 주입 완료!")
print(f"총 주입된 하이퍼링크 수: {total_linked}개")
print(f"==================================================")
