# -*- coding: utf-8 -*-
import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

excel_dir = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8")
xlsx_path = os.path.join(excel_dir, "매뉴얼 BODY (집행단계)v8.xlsx")
wb = openpyxl.load_workbook(xlsx_path, data_only=False)

total = 0
broken = 0

for s in ["사전토공사", "상부강화노반", "지장물이설", "콘크리트도상", "건축", "신호", "전기", "통신"]:
    ws = wb[s]
    sheet_links = 0
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value or "")
            if val.startswith("=HYPERLINK("):
                p_and = val.find('&"') + 2
                p_end = val.find('"', p_and)
                if p_and > 1 and p_end > p_and:
                    rel_p = val[p_and:p_end]
                    full_p = os.path.join(excel_dir, rel_p)
                    sheet_links += 1
                    total += 1
                    if not os.path.exists(full_p):
                        print(f"Broken: [{s}] R{r}C{c} -> {rel_p}")
                        broken += 1
    print(f"Sheet [{s}]: {sheet_links}개 링크 검증 완료")

print(f"\n✓ 총 {total}개 링크 중 깨진 파일: {broken}개 (0이어야 완벽)")
