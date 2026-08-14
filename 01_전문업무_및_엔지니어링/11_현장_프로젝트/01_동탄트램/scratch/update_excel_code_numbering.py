# -*- coding: utf-8 -*-
"""
동탄트램 매뉴얼 BODY (집행단계)v8 엑셀 파일 L2(A), L3(B), L4(D), 선행(E), 후행(F) 코드 체계 일괄 정비 스크립트

시트 순서에 따른 L3 코드 매핑:
1. 지반조사 -> 9000-1 (L4: 9000-1-1 ~ 9000-1-36)
2. 사전토공사 -> 9000-2 (L4: 9000-2-1 ~ 9000-2-31)
3. 지장물이설 -> 9000-3 (L4: 9000-3-1 ~ 9000-3-38)
4. 상부강화노반 -> 9000-4 (L4: 9000-4-1 ~ 9000-4-36)
5. 콘크리트도상 -> 9000-5 (L4: 9000-5-1 ~ 9000-5-23)
6. 건축 -> 9000-6 (L4: 9000-6-1 ~ 9000-6-50)
7. 신호 -> 9000-7 (L4: 9000-7-1 ~ 9000-7-23)
8. 전기 -> 9000-8 (L4: 9000-8-1 ~ 9000-8-32)
9. 통신 -> 9000-9 (L4: 9000-9-1 ~ 9000-9-32)
10. 기계 -> 9000-10 (L4: 9000-10-1 ~ 9000-10-36)
11. 철도종합시운전 -> 9000-11 (L4: 9000-11-1 ~ 9000-11-28)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment

# Kill any stuck Excel processes
os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

font_bold = Font(name="맑은 고딕", size=9, bold=True, color="0F172A")
font_regular = Font(name="맑은 고딕", size=9, color="0F172A")
align_center = Alignment(horizontal="center", vertical="center")

# 각 시트별 실제 유효 작업(Task) 수 매핑
SHEET_CONFIG = {
    "지반조사": {"idx": 1, "name": "지반조사", "tasks": 36},
    "사전토공사": {"idx": 2, "name": "사전토공사", "tasks": 31},
    "지장물이설": {"idx": 3, "name": "지장물이설", "tasks": 38},
    "상부강화노반": {"idx": 4, "name": "상부강화노반", "tasks": 36},
    "콘크리트도상": {"idx": 5, "name": "콘크리트도상", "tasks": 23},
    "건축": {"idx": 6, "name": "건축", "tasks": 50},
    "신호": {"idx": 7, "name": "신호분야", "tasks": 23},
    "전기": {"idx": 8, "name": "전기분야", "tasks": 32},
    "통신": {"idx": 9, "name": "통신분야", "tasks": 32},
    "기계": {"idx": 10, "name": "기계설비·소방설비", "tasks": 36},
    "철도종합시운전": {"idx": 11, "name": "철도종합시험운행", "tasks": 28}
}

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue
    print(f"\n=======================================================")
    print(f"엑셀 파일 처리 중: {file_path}")
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CONFIG:
            print(f"  [건너뜀] 설정에 없는 시트: {sheet_name}")
            continue

        cfg = SHEET_CONFIG[sheet_name]
        s_num = cfg["idx"]
        l3_name = cfg["name"]
        task_count = cfg["tasks"]
        ws = wb[sheet_name]

        print(f"  ▶ 시트 [{sheet_name}]: L3코드=9000-{s_num}, 대공종명={l3_name}, 총 {task_count}개 작업")

        for k in range(1, task_count + 1):
            r = k + 1  # 2행부터 시작

            l2_code = "9000"
            l3_code = f"9000-{s_num}"
            l4_code = f"9000-{s_num}-{k}"
            pred_code = "-" if k == 1 else f"9000-{s_num}-{k-1}"
            succ_code = "-" if k == task_count else f"9000-{s_num}-{k+1}"

            # Col A: L2 코드
            cA = ws.cell(r, 1, value=l2_code)
            cA.font = font_bold
            cA.alignment = align_center

            # Col B: L3 코드
            cB = ws.cell(r, 2, value=l3_code)
            cB.font = font_bold
            cB.alignment = align_center

            # Col C: L3 대공종명
            cC = ws.cell(r, 3, value=l3_name)
            cC.font = font_bold
            cC.alignment = align_center

            # Col D: L4 코드 (문자열로 강제 저장하여 날짜 파싱 방지)
            cD = ws.cell(r, 4, value=str(l4_code))
            cD.font = font_bold
            cD.alignment = align_center

            # Col E: 선행
            cE = ws.cell(r, 5, value=str(pred_code))
            cE.font = font_regular
            cE.alignment = align_center

            # Col F: 후행
            cF = ws.cell(r, 6, value=str(succ_code))
            cF.font = font_regular
            cF.alignment = align_center

        # task_count 이후 잔여 행의 A~F 값 정리
        for r in range(task_count + 2, ws.max_row + 1):
            for c in range(1, 7):
                ws.cell(r, c).value = None

    wb.save(file_path)
    print(f"  ✔ 저장 완료: {os.path.basename(file_path)}")

print(f"\n=======================================================")
print(f"전체 엑셀 파일 L2, L3, L4, 선행, 후행 코드 번호체계 시트 순서 일괄 정비 완료!")
print(f"=======================================================")
