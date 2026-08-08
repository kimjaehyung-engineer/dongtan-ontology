import sys
import os
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 1공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 2공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'

def extract_intersections(path, tool_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    data = []
    
    for r in range(1, 100):
        no_val = sheet.cell(r, 2).value
        if isinstance(no_val, (int, float)) and no_val > 0:
            section = sheet.cell(r, 3).value
            name = sheet.cell(r, 4).value
            start_sta = sheet.cell(r, 9).value
            end_sta = sheet.cell(r, 10).value
            length = sheet.cell(r, 11).value
            method = sheet.cell(r, 12).value
            
            if tool_name == '1공구':
                stage = sheet.cell(r, 13).value
                avg_len = sheet.cell(r, 14).value
            else: # 2공구
                stage = sheet.cell(r, 14).value
                avg_len = sheet.cell(r, 15).value
                
            # Round values cleanly
            start_val = round(float(start_sta), 1) if isinstance(start_sta, (int, float)) else start_sta
            end_val = round(float(end_sta), 1) if isinstance(end_sta, (int, float)) else end_sta
            len_val = round(float(length), 1) if isinstance(length, (int, float)) else length
            avg_val = round(float(avg_len), 1) if isinstance(avg_len, (int, float)) else avg_len
            
            data.append({
                "tool": tool_name,
                "no": int(no_val),
                "code": str(section or '').strip(),
                "name": str(name or '').strip(),
                "startSta": start_val,
                "endSta": end_val,
                "length": len_val,
                "method": str(method or '').strip(),
                "stage": int(stage) if isinstance(stage, (int, float)) else stage,
                "avgLen": avg_val
            })
    return data

intersections_1 = extract_intersections(file1, "1공구")
intersections_2 = extract_intersections(file2, "2공구")
all_intersections = intersections_1 + intersections_2

print(f"Extracted {len(intersections_1)} from 1공구, {len(intersections_2)} from 2공구. Total: {len(all_intersections)}")

# Save to scratch json
output_json_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\intersections.json'
with open(output_json_path, 'w', encoding='utf-8') as f:
    json.dump(all_intersections, f, ensure_ascii=False, indent=2)

print("Saved to intersections.json successfully.")
