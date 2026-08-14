# -*- coding: utf-8 -*-
"""
사전토공사 시트 O열, Q열, S열에
명확한 표시 텍스트 '👉 [클릭] 표준서 열기 📄', '👉 [클릭] 수행지침 열기 📄', '👉 [클릭] 체크리스트 열기 📄'
와 함께 파란색 밑줄 네이티브 하이퍼링크를 완벽하게 주입합니다.
"""

import os, sys, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import win32com.client

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
V8_ATTACH_DIR = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")
EARTH_ATTACH_DIR = os.path.join(V8_ATTACH_DIR, "2.사전토공사")
TARGET_XLSM = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
TARGET_XLSX = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")

# 1. 사전토공사 HTML 파일 매핑
earth_htmls = []
for root, dirs, files in os.walk(EARTH_ATTACH_DIR):
    for f in files:
        if f.lower().endswith(".html"):
            earth_htmls.append(os.path.join(root, f))

def normalize(s):
    if not s: return ""
    return str(s).strip().replace(" ", "").replace("_", "").replace("-", "").replace("/", "").replace("(", "").replace(")", "").lower()

def find_earth_html(act_name, doc_type, l4_code=""):
    n_act = normalize(act_name)
    for sub in os.listdir(EARTH_ATTACH_DIR):
        sub_p = os.path.join(EARTH_ATTACH_DIR, sub)
        if os.path.isdir(sub_p):
            n_sub = normalize(sub)
            if n_act and (n_act in n_sub or n_sub in n_act or n_act[:4] in n_sub):
                doc_sub = os.path.join(sub_p, doc_type)
                if os.path.exists(doc_sub):
                    for f in os.listdir(doc_sub):
                        if f.lower().endswith(".html") and doc_type in f:
                            return os.path.join(doc_sub, f)
                for f in os.listdir(sub_p):
                    if f.lower().endswith(".html") and doc_type in f:
                        return os.path.join(sub_p, f)

    for p in earth_htmls:
        if doc_type in os.path.basename(p):
            n_p = normalize(p)
            if n_act and (n_act in n_p or n_act[:4] in n_p):
                return p

    for p in earth_htmls:
        if doc_type in os.path.basename(p):
            return p

    return None

# 2. openpyxl로 직접 명확한 텍스트 + 하이퍼링크 객체 주입
wb = openpyxl.load_workbook(TARGET_XLSM, keep_vba=True)
ws = wb["사전토공사"]

font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

for r in range(2, ws.max_row + 1):
    act_name = str(ws.cell(r, 8).value or "").strip()
    l4_code = str(ws.cell(r, 4).value or "").strip()

    std_p = find_earth_html(act_name, "표준서", l4_code)
    guide_p = find_earth_html(act_name, "수행지침", l4_code)
    chk_p = find_earth_html(act_name, "체크리스트", l4_code)

    # O열 (표준서)
    c_o = ws.cell(r, 15)
    c_o.value = "👉 [클릭] 표준서 열기 📄"
    c_o.font = font_link
    c_o.alignment = align_center
    if std_p and os.path.exists(std_p):
        c_o.hyperlink = std_p

    # Q열 (수행지침)
    c_q = ws.cell(r, 17)
    c_q.value = "👉 [클릭] 수행지침 열기 📄"
    c_q.font = font_link
    c_q.alignment = align_center
    if guide_p and os.path.exists(guide_p):
        c_q.hyperlink = guide_p

    # S열 (체크리스트)
    c_s = ws.cell(r, 19)
    c_s.value = "👉 [클릭] 체크리스트 열기 📄"
    c_s.font = font_link
    c_s.alignment = align_center
    if chk_p and os.path.exists(chk_p):
        c_s.hyperlink = chk_p

# 열 너비 명시적 설정
ws.column_dimensions['O'].width = 24
ws.column_dimensions['Q'].width = 24
ws.column_dimensions['S'].width = 24

wb.save(TARGET_XLSM)
print("[1] openpyxl로 v8.xlsm 텍스트 및 하이퍼링크 직접 저장 완료!")

# 3. Excel COM으로 열어서 안전하게 최종 저장
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False
excel.ScreenUpdating = False

try:
    wb_com = excel.Workbooks.Open(TARGET_XLSM)
    wb_com.Save()
    print("[2] Excel COM v8.xlsm 저장 완료!")

    # v8.xlsx로도 저장
    if os.path.exists(TARGET_XLSX):
        os.remove(TARGET_XLSX)
    wb_com.SaveAs(TARGET_XLSX, FileFormat=51)
    wb_com.Close()
    print("[3] Excel COM v8.xlsx 저장 완료!")
finally:
    excel.Quit()

print("=== 완벽 적용 완료 ===")
