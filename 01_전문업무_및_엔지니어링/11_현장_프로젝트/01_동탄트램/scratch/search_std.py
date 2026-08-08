import openpyxl
import os

wb_path = 'c:/Users/sskjh/antigravity/01_전문업무_및_엔지니어링/동탄트램/06.체크리스트/260615 트램_표준_리스크_체크리스트(동탄기본설계_검토반영)V3.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=True)

search_codes = ['STD-0145', 'STD-0146', 'STD-0147', 'STD-0148', 'STD-0149', 'STD-0150', 'STD-0151', 'STD-0152', 'STD-0153', 'STD-0154', 'STD-0155', 'STD-0156']

print("Searching standard checklist for codes...")
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for r_idx in range(1, sheet.max_row + 1):
        for c_idx in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r_idx, column=c_idx).value or '')
            for code in search_codes:
                if code in val:
                    print(f"Found {code} in sheet '{sheet_name}', Row {r_idx}, Col {c_idx}: {val[:100]}")
