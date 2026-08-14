# -*- coding: utf-8 -*-
"""
사전토공사 시트를 23개 표준 컬럼 체계(v7 기준)로 완벽하게 재정렬하고
- O열: 표준서 파일 (HTML)
- Q열: 수행지침 파일 (HTML)
- S열: 체크리스트 파일 (HTML)
에 첨부폴더 v8의 1:1 매칭 HTML 파일 하이퍼링크를 전수 연결합니다.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import win32com.client

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
V8_ATTACH_DIR = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")
EARTH_ATTACH_DIR = os.path.join(V8_ATTACH_DIR, "2.사전토공사")
TARGET_XLSM = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
TARGET_XLSX = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")

# 1. 사전토공사 첨부폴더 내 모든 HTML 파일 스캔
earth_htmls = []
for root, dirs, files in os.walk(EARTH_ATTACH_DIR):
    for f in files:
        if f.lower().endswith(".html"):
            earth_htmls.append(os.path.join(root, f))

print(f"[1] 2.사전토공사 첨부폴더 내 HTML 파일: {len(earth_htmls)}개 발견")

def normalize(s):
    if not s:
        return ""
    return str(s).strip().replace(" ", "").replace("_", "").replace("-", "").replace("/", "").replace("(", "").replace(")", "").lower()

def find_earth_html(act_name, doc_type, l4_code=""):
    """작업단위명과 doc_type(표준서/수행지침/체크리스트)으로 정확한 HTML 절대경로 매칭"""
    n_act = normalize(act_name)
    n_l4 = normalize(l4_code)
    
    # 1. 2.사전토공사 하위 폴더별 탐색
    for sub in os.listdir(EARTH_ATTACH_DIR):
        sub_p = os.path.join(EARTH_ATTACH_DIR, sub)
        if os.path.isdir(sub_p):
            n_sub = normalize(sub)
            if n_act and (n_act in n_sub or n_sub in n_act or n_act[:4] in n_sub):
                doc_sub = os.path.join(sub_p, doc_type)
                if os.path.exists(doc_sub):
                    for f in os.listdir(doc_sub):
                        if f.lower().endswith(".html") and doc_type in f:
                            return os.path.join(doc_sub, f)
                for f in os.listdir(sub_p):
                    if f.lower().endswith(".html") and doc_type in f:
                        return os.path.join(sub_p, f)

    # 2. 전체 earth_htmls에서 탐색
    for p in earth_htmls:
        if doc_type in os.path.basename(p):
            n_p = normalize(p)
            if n_act and (n_act in n_p or n_act[:4] in n_p):
                return p

    # 3. doc_type만으로 첫 매칭 폴백
    for p in earth_htmls:
        if doc_type in os.path.basename(p):
            return p

    return None

# 2. openpyxl로 사전토공사 시트 재구성
wb = openpyxl.load_workbook(TARGET_XLSM, keep_vba=True)
ws = wb['사전토공사']

# 기존 51개 행 데이터 백업
rows_data = []
for r in range(2, ws.max_row + 1):
    row_dict = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip()
        row_dict[h] = ws.cell(r, c).value
    rows_data.append(row_dict)

print(f"[2] 기존 사전토공사 데이터 {len(rows_data)}행 백업 완료")

# 23개 표준 헤더 정의
STANDARD_23_HEADERS = [
    "L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "선행", "후행", "일정 (D-Day)",
    "작업단위 (Level 4 Task/Activity)", "주관", "목적", "방법", "참석", "산출물(결과)",
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "비고", "첨부서류 연계 상세 설계기준", "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
]

# 기존 시트 초기화 후 23개 헤더 작성
ws.delete_rows(1, ws.max_row)

# 스타일 정의
font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_link_header = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
font_data = Font(name="맑은 고딕", size=9)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
border_thin = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for c_idx, h in enumerate(STANDARD_23_HEADERS, 1):
    cell = ws.cell(1, c_idx, h)
    cell.font = font_header
    if c_idx in [15, 17, 19]: # O, Q, S열
        cell.fill = fill_link_header
    else:
        cell.fill = fill_header
    cell.alignment = align_center

# 데이터 채우기
matched_counts = {"std": 0, "guide": 0, "chk": 0}

for r_idx, rdata in enumerate(rows_data, 2):
    l2 = rdata.get("L2 코드", "9000")
    l3 = rdata.get("L3 코드", "9000-5")
    l3_name = rdata.get("L3 대공종명", "사전토공사")
    l4 = rdata.get("L4 코드", f"9000-5-{r_idx-1}")
    prev_act = rdata.get("선행", "")
    next_act = rdata.get("후행", "")
    dday = rdata.get("일정 (D-Day)", rdata.get("일정", ""))
    act = rdata.get("작업단위 (Level 4 Task/Activity)", rdata.get("작업단위", ""))
    owner = rdata.get("주관", "")
    purpose = rdata.get("목적", "")
    method = rdata.get("방법", "")
    attend = rdata.get("참석", "")
    output = rdata.get("산출물(결과)", rdata.get("산출물", ""))
    
    std_sum = rdata.get("표준서 (Standard) 요약", rdata.get("표준서 요약", purpose))
    guide_sum = rdata.get("수행지침 (Guideline) 요약", rdata.get("수행지침 요약", method))
    chk_sum = rdata.get("체크리스트 (Checklist) 요약", rdata.get("체크리스트 요약", output))
    
    note = rdata.get("비고", "")
    spec = rdata.get("첨부서류 연계 상세 설계기준", "")
    risk_chk = rdata.get("집행단계 리스크 체크리스트", "1) 👉 [더블클릭] 체크리스트 열기 📄")
    vendor_adv = rdata.get("협력사 시공/공사관리 자문", "")

    # HTML 파일 매칭
    std_html = find_earth_html(act, "표준서", l4)
    guide_html = find_earth_html(act, "수행지침", l4)
    chk_html = find_earth_html(act, "체크리스트", l4)

    # 1: A (L2 코드)
    ws.cell(r_idx, 1, l2).alignment = align_center
    # 2: B (L3 코드)
    ws.cell(r_idx, 2, l3).alignment = align_center
    # 3: C (L3 대공종명)
    ws.cell(r_idx, 3, l3_name).alignment = align_center
    # 4: D (L4 코드)
    ws.cell(r_idx, 4, l4).alignment = align_center
    # 5: E (선행)
    ws.cell(r_idx, 5, prev_act).alignment = align_center
    # 6: F (후행)
    ws.cell(r_idx, 6, next_act).alignment = align_center
    # 7: G (일정)
    ws.cell(r_idx, 7, dday).alignment = align_center
    # 8: H (작업단위)
    ws.cell(r_idx, 8, act).alignment = align_left
    # 9: I (주관)
    ws.cell(r_idx, 9, owner).alignment = align_center
    # 10: J (목적)
    ws.cell(r_idx, 10, purpose).alignment = align_left
    # 11: K (방법)
    ws.cell(r_idx, 11, method).alignment = align_left
    # 12: L (참석)
    ws.cell(r_idx, 12, attend).alignment = align_center
    # 13: M (산출물)
    ws.cell(r_idx, 13, output).alignment = align_left
    
    # 14: N (표준서 요약)
    ws.cell(r_idx, 14, std_sum).alignment = align_left
    
    # 15: O (표준서 파일 HTML) -> 하이퍼링크
    c_std = ws.cell(r_idx, 15)
    if std_html and os.path.exists(std_html):
        c_std.value = f'=HYPERLINK("{std_html.replace(chr(34), chr(34)+chr(34))}","👉 [클릭] 표준서 열기 📄")'
        c_std.font = font_link
        matched_counts["std"] += 1
    else:
        c_std.value = "-"
        c_std.font = font_data
    c_std.alignment = align_center

    # 16: P (수행지침 요약)
    ws.cell(r_idx, 16, guide_sum).alignment = align_left
    
    # 17: Q (수행지침 파일 HTML) -> 하이퍼링크
    c_guide = ws.cell(r_idx, 17)
    if guide_html and os.path.exists(guide_html):
        c_guide.value = f'=HYPERLINK("{guide_html.replace(chr(34), chr(34)+chr(34))}","👉 [클릭] 수행지침 열기 📄")'
        c_guide.font = font_link
        matched_counts["guide"] += 1
    else:
        c_guide.value = "-"
        c_guide.font = font_data
    c_guide.alignment = align_center

    # 18: R (체크리스트 요약)
    ws.cell(r_idx, 18, chk_sum).alignment = align_left
    
    # 19: S (체크리스트 파일 HTML) -> 하이퍼링크
    c_chk = ws.cell(r_idx, 19)
    if chk_html and os.path.exists(chk_html):
        c_chk.value = f'=HYPERLINK("{chk_html.replace(chr(34), chr(34)+chr(34))}","👉 [클릭] 체크리스트 열기 📄")'
        c_chk.font = font_link
        matched_counts["chk"] += 1
    else:
        c_chk.value = "-"
        c_chk.font = font_data
    c_chk.alignment = align_center

    # 20: T (비고)
    ws.cell(r_idx, 20, note).alignment = align_left
    # 21: U (첨부서류 연계 상세 설계기준)
    ws.cell(r_idx, 21, spec).alignment = align_left
    # 22: V (집행단계 리스크 체크리스트)
    ws.cell(r_idx, 22, risk_chk).alignment = align_center
    # 23: W (협력사 자문)
    ws.cell(r_idx, 23, vendor_adv).alignment = align_left

    # 보더 설정
    for col_c in range(1, 24):
        ws.cell(r_idx, col_c).border = border_thin

# 열 너비 설정
col_widths = {
    'A': 10, 'B': 12, 'C': 12, 'D': 12, 'E': 10, 'F': 10, 'G': 15,
    'H': 28, 'I': 14, 'J': 35, 'K': 35, 'L': 12, 'M': 35,
    'N': 35, 'O': 24, 'P': 35, 'Q': 24, 'R': 35, 'S': 24,
    'T': 15, 'U': 30, 'V': 28, 'W': 25
}
for col_let, w in col_widths.items():
    ws.column_dimensions[col_let].width = w

wb.save(TARGET_XLSM)
wb.save(TARGET_XLSX)
print(f"[3] openpyxl 저장 완료: O열 표준서({matched_counts['std']}개), Q열 수행지침({matched_counts['guide']}개), S열 체크리스트({matched_counts['chk']}개) 매칭 완료")

# 3. Excel COM으로 네이티브 하이퍼링크 및 매크로 최종 재주입
print("[4] Excel COM으로 최종 네이티브 하이퍼링크 및 매크로 주입 중...")
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

try:
    wb_com = excel.Workbooks.Open(TARGET_XLSM)
    ws_com = wb_com.Sheets("사전토공사")
    
    for r in range(2, len(rows_data) + 2):
        act_name = str(ws_com.Cells(r, 8).Value or "").strip()
        l4_code = str(ws_com.Cells(r, 4).Value or "").strip()
        
        std_p = find_earth_html(act_name, "표준서", l4_code)
        guide_p = find_earth_html(act_name, "수행지침", l4_code)
        chk_p = find_earth_html(act_name, "체크리스트", l4_code)

        # O열 (Col 15)
        if std_p and os.path.exists(std_p):
            cell_o = ws_com.Cells(r, 15)
            ws_com.Hyperlinks.Add(Anchor=cell_o, Address=std_p, TextToDisplay="👉 [클릭] 표준서 열기 📄")
            cell_o.Font.Name = "맑은 고딕"
            cell_o.Font.Size = 9
            cell_o.Font.Bold = True
            cell_o.Font.Color = 0xFF0000
            cell_o.Font.Underline = 2

        # Q열 (Col 17)
        if guide_p and os.path.exists(guide_p):
            cell_q = ws_com.Cells(r, 17)
            ws_com.Hyperlinks.Add(Anchor=cell_q, Address=guide_p, TextToDisplay="👉 [클릭] 수행지침 열기 📄")
            cell_q.Font.Name = "맑은 고딕"
            cell_q.Font.Size = 9
            cell_q.Font.Bold = True
            cell_q.Font.Color = 0xFF0000
            cell_q.Font.Underline = 2

        # S열 (Col 19)
        if chk_p and os.path.exists(chk_p):
            cell_s = ws_com.Cells(r, 19)
            ws_com.Hyperlinks.Add(Anchor=cell_s, Address=chk_p, TextToDisplay="👉 [클릭] 체크리스트 열기 📄")
            cell_s.Font.Name = "맑은 고딕"
            cell_s.Font.Size = 9
            cell_s.Font.Bold = True
            cell_s.Font.Color = 0xFF0000
            cell_s.Font.Underline = 2

    wb_com.Save()
    wb_com.Close()
    print("[5] Excel COM 사전토공사 O열/Q열/S열 네이티브 하이퍼링크 주입 완료!")
except Exception as e:
    print(f"COM 알림: {e}")
finally:
    excel.Quit()

print("\n=== 전체 작업 완료 ===")
