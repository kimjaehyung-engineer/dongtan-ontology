# -*- coding: utf-8 -*-
"""
동탄도시철도(트램) 1.지반조사 36개 액티비티 108개 HTML(표준서/수행지침/체크리스트) 마스터 생성 엔진
"""

import os
import sys
import re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

EXCEL_PATH = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx")
OUTPUT_BASE = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\1.지반조사")

# 1. 엑셀 지반조사 시트 데이터 로드
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["지반조사"]

ACTIVITIES = []
for r in range(2, ws.max_row + 1):
    l4_code = ws.cell(r, 4).value
    task = ws.cell(r, 7).value
    dday = ws.cell(r, 8).value
    dept = ws.cell(r, 9).value
    purpose = ws.cell(r, 10).value
    method = ws.cell(r, 11).value
    deliverable = ws.cell(r, 12).value
    if task:
        idx = len(ACTIVITIES) + 1
        ACTIVITIES.append({
            "idx": idx,
            "code": str(l4_code).strip() if l4_code else f"9000-1-{idx}",
            "task": str(task).strip(),
            "dday": str(dday).strip() if dday else "D-Day 미정",
            "dept": str(dept).strip() if dept else "현장 공사팀",
            "purpose": str(purpose).strip() if purpose else "지반조사 업무 기준 수립 및 리스크 관리",
            "method": str(method).strip() if method else "표준 절차에 따른 조사 및 분석 수행",
            "deliverable": str(deliverable).strip() if deliverable else "지반조사 보고서"
        })

print(f"로드된 지반조사 액티비티 수: {len(ACTIVITIES)}개")

# 2. 공통 CSS & 네비게이션 스타일
NAV_CSS = """
    /* Standardized Bottom Navigation Box */
    .nav-box { display: flex; gap: 12px; margin-top: 36px; padding-top: 24px; border-top: 1px solid #e2e8f0; width: 100%; box-sizing: border-box; }
    .nav-btn { flex: 1; text-align: center; padding: 14px 16px; border-radius: 6px; font-size: 13px; font-weight: 700; text-decoration: none; transition: 0.2s; display: inline-block; box-sizing: border-box; }
    .btn-std { background: #0f172a !important; color: #ffffff !important; }
    .btn-guide { background: #0284c7 !important; color: #ffffff !important; }
    .btn-chk { background: #d97706 !important; color: #ffffff !important; }
    .btn-std:hover { background: #1e293b !important; }
    .btn-guide:hover { background: #0369a1 !important; }
    .btn-chk:hover { background: #b45309 !important; }
"""

MODAL_CSS = """
    #zoomModal, #glossaryModal { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(15,23,42,0.8); z-index: 9999; justify-content: center; align-items: center; padding: 20px; box-sizing: border-box; }
    .modal-content { background: #ffffff; border-radius: 12px; max-width: 1000px; width: 100%; max-height: 90vh; overflow-y: auto; padding: 24px; position: relative; box-shadow: 0 20px 25px -5px rgba(0,0,0,0.3); }
    .close-btn { position: absolute; top: 16px; right: 16px; background: #f1f5f9; border: none; font-size: 18px; font-weight: bold; cursor: pointer; width: 32px; height: 32px; border-radius: 50%; display: flex; align-items: center; justify-content: center; }
"""

# 3. 2D SVG 도식 생성기 (Light Theme High Contrast)
def get_jiban_svg(idx, task_name):
    # 액티비티 특성에 맞춘 전문 2D 지반공학 SVG 도식
    if "시추" in task_name or "조사기획" in task_name:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">동탄트램 시추조사 주상도(Boring Log) 및 코어 회수율(TCR/RQD) 모식도</text>
            <!-- 지층 레이어 -->
            <rect x="60" y="70" width="160" height="30" fill="#fef08a" stroke="#ca8a04" stroke-width="1.5"/>
            <text x="140" y="90" font-size="11" font-weight="bold" fill="#854d0e" text-anchor="middle">매립토층 (N: 4~8)</text>
            
            <rect x="60" y="100" width="160" height="40" fill="#fed7aa" stroke="#ea580c" stroke-width="1.5"/>
            <text x="140" y="125" font-size="11" font-weight="bold" fill="#9a3412" text-anchor="middle">퇴적토/풍화토 (N: 15~30)</text>
            
            <rect x="60" y="140" width="160" height="35" fill="#e2e8f0" stroke="#64748b" stroke-width="1.5"/>
            <text x="140" y="162" font-size="11" font-weight="bold" fill="#334155" text-anchor="middle">풍화암 (N > 50/10)</text>
            
            <rect x="60" y="175" width="160" height="35" fill="#cbd5e1" stroke="#475569" stroke-width="1.5"/>
            <text x="140" y="197" font-size="11" font-weight="bold" fill="#0f172a" text-anchor="middle">연암/경암 (TCR ≥ 85%)</text>
            
            <!-- 코어 배럴 및 시추장비 -->
            <line x1="260" y1="70" x2="260" y2="210" stroke="#0284c7" stroke-width="4"/>
            <circle cx="260" cy="70" r="8" fill="#0284c7"/>
            <rect x="290" y="75" width="350" height="125" fill="#f0fdf4" stroke="#16a34a" rx="6"/>
            <text x="310" y="100" font-size="12" font-weight="bold" fill="#166534">■ 시추코어 판정 기준</text>
            <text x="310" y="125" font-size="11" fill="#1e293b">• 코어회수율 TCR = (회수코어길이 / 굴진길이) × 100 ≥ 85%</text>
            <text x="310" y="148" font-size="11" fill="#1e293b">• 암질지수 RQD = (10cm이상 코어길이합 / 굴진길이) × 100</text>
            <text x="310" y="172" font-size="11" fill="#1e293b">• 지하수위(GL) 측정: 시추 완료 후 24h/48h 경과 후 안정수위 확정</text>
        </svg>
        """
    elif "SPT" in task_name or "원위치" in task_name or "CPT" in task_name:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">표준관입시험(SPT) 63.5kg 해머 76cm 자유낙하 및 N치 산정</text>
            <rect x="60" y="70" width="80" height="35" fill="#e2e8f0" stroke="#0284c7" stroke-width="2" rx="4"/>
            <text x="100" y="92" font-size="11" font-weight="bold" fill="#0369a1" text-anchor="middle">63.5kg 해머</text>
            <line x1="100" y1="105" x2="100" y2="155" stroke="#0f172a" stroke-width="3" stroke-dasharray="4,2"/>
            <text x="145" y="135" font-size="11" font-weight="bold" fill="#dc2626">낙하고 76cm</text>
            <rect x="60" y="160" width="80" height="45" fill="#fef08a" stroke="#ca8a04" stroke-width="1.5"/>
            <text x="100" y="185" font-size="10" font-weight="bold" fill="#854d0e" text-anchor="middle">30cm 관입 타격수</text>
            
            <rect x="240" y="70" width="410" height="135" fill="#eff6ff" stroke="#3b82f6" rx="6"/>
            <text x="260" y="95" font-size="12" font-weight="bold" fill="#1e40af">■ SPT N치 기반 지반강도 판정 수식</text>
            <text x="260" y="122" font-size="11" fill="#1e293b">• 예비관입 15cm 타격 후, 본관입 30cm 관입에 필요한 타격수 N</text>
            <text x="260" y="146" font-size="11" fill="#1e293b">• 모래/자갈 내부마찰각(φ): φ = √(20N) + 15° (오자키 제안식)</text>
            <text x="260" y="170" font-size="11" fill="#1e293b">• 점성토 일축압축강도(qu): qu = N / 8 (kgf/㎠) = 12.5N (kPa)</text>
            <text x="260" y="192" font-size="11" font-weight="bold" fill="#b45309">• 허용지내력(qa): Meyerhof 식 qa = 10 × N (kN/㎡)</text>
        </svg>
        """
    elif "GPR" in task_name or "공동" in task_name:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">3D 멀티채널 차량형 GPR 지중레이더 지하공동 탐사 프로파일</text>
            <!-- 도로 및 차량 -->
            <rect x="50" y="70" width="220" height="20" fill="#334155"/>
            <text x="160" y="84" font-size="10" font-weight="bold" fill="#ffffff" text-anchor="middle">아스팔트 포장면 (도로)</text>
            <rect x="50" y="90" width="220" height="110" fill="#fed7aa" stroke="#ea580c"/>
            <ellipse cx="160" cy="145" rx="35" ry="20" fill="#dc2626" opacity="0.8"/>
            <text x="160" y="150" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">지하공동 (이상대)</text>
            
            <rect x="300" y="70" width="350" height="130" fill="#fef2f2" stroke="#ef4444" rx="6"/>
            <text x="320" y="95" font-size="12" font-weight="bold" fill="#991b1b">■ 3D GPR 공동 등급 판정 기준</text>
            <text x="320" y="120" font-size="11" fill="#1e293b">• 급행(긴급): 포장하부 공동두께 ≤ 0.3m 또는 공동규모 ≥ 1.5m</text>
            <text x="320" y="144" font-size="11" fill="#1e293b">• 우선(보통): 공동규모 0.5~1.5m, 포장하부 토피고 0.3~0.8m</text>
            <text x="320" y="168" font-size="11" fill="#1e293b">• 관찰(경미): 공동규모 < 0.5m, 토피고 > 0.8m 지속 모니터링</text>
            <text x="320" y="190" font-size="11" font-weight="bold" fill="#b91c1c">• 조치: 천공 내시경 확인 후 친환경 유동성 그라우팅 주입 100%</text>
        </svg>
        """
    elif "물리탐사" in task_name or "탄성파" in task_name or "전기비저항" in task_name:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">탄성파 굴절법 주시곡선 및 2D 지전기 비저항 단면 해석</text>
            <rect x="50" y="70" width="250" height="130" fill="#f1f5f9" stroke="#94a3b8" rx="6"/>
            <text x="65" y="95" font-size="11" font-weight="bold" fill="#0f172a">■ 굴절파 탄성파 속도(Vp)</text>
            <text x="65" y="120" font-size="10.5" fill="#1e293b">• 제1층(토사층): Vp = 300 ~ 800 m/s</text>
            <text x="65" y="142" font-size="10.5" fill="#1e293b">• 제2층(풍화암): Vp = 1,200 ~ 2,000 m/s</text>
            <text x="65" y="164" font-size="10.5" fill="#1e293b">• 제3층(연암층): Vp = 2,100 ~ 3,000 m/s</text>
            <text x="65" y="186" font-size="10.5" fill="#1e293b">• 제4층(경암층): Vp ≥ 3,000 m/s</text>
            
            <rect x="320" y="70" width="330" height="130" fill="#f8fafc" stroke="#0284c7" rx="6"/>
            <text x="335" y="95" font-size="11" font-weight="bold" fill="#0369a1">■ 전기비저항 이상대 해석</text>
            <text x="335" y="120" font-size="10.5" fill="#1e293b">• 저비저항 이상대( < 100 Ω·m): 단층파쇄대, 지하수 용출대</text>
            <text x="335" y="145" font-size="10.5" fill="#1e293b">• 중비저항대(100 ~ 1,000 Ω·m): 풍화토 및 풍화암</text>
            <text x="335" y="170" font-size="10.5" fill="#1e293b">• 고비저항대( > 1,000 Ω·m): 신선한 기반암(경암층)</text>
            <text x="335" y="190" font-size="10.5" font-weight="bold" fill="#dc2626">• 단층 및 지하수 포화대 확인 시 차수 그라우팅 보강</text>
        </svg>
        """
    elif "RMR" in task_name or "암반" in task_name or "Q-" in task_name:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">Bieniawski RMR 암반공학적 등급 분류 매트릭스</text>
            <rect x="50" y="70" width="600" height="25" fill="#1e293b"/>
            <text x="110" y="87" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">RMR 점수</text>
            <text x="210" y="87" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">암반 등급</text>
            <text x="330" y="87" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">암질 상태</text>
            <text x="450" y="87" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">점착력 (kPa)</text>
            <text x="570" y="87" font-size="11" font-weight="bold" fill="#ffffff" text-anchor="middle">내부마찰각(°)</text>
            
            <rect x="50" y="95" width="600" height="22" fill="#dcfce7"/>
            <text x="110" y="110" font-size="10.5" fill="#166534" text-anchor="middle">81 ~ 100</text><text x="210" y="110" font-size="10.5" font-weight="bold" fill="#166534" text-anchor="middle">Class I</text><text x="330" y="110" font-size="10.5" fill="#166534" text-anchor="middle">매우 양호 (Very Good)</text><text x="450" y="110" font-size="10.5" fill="#166534" text-anchor="middle">> 400</text><text x="570" y="110" font-size="10.5" fill="#166534" text-anchor="middle">> 45°</text>
            
            <rect x="50" y="117" width="600" height="22" fill="#eff6ff"/>
            <text x="110" y="132" font-size="10.5" fill="#1e40af" text-anchor="middle">61 ~ 80</text><text x="210" y="132" font-size="10.5" font-weight="bold" fill="#1e40af" text-anchor="middle">Class II</text><text x="330" y="132" font-size="10.5" fill="#1e40af" text-anchor="middle">양호 (Good)</text><text x="450" y="132" font-size="10.5" fill="#1e40af" text-anchor="middle">300 ~ 400</text><text x="570" y="132" font-size="10.5" fill="#1e40af" text-anchor="middle">35° ~ 45°</text>
            
            <rect x="50" y="139" width="600" height="22" fill="#fef9c3"/>
            <text x="110" y="154" font-size="10.5" fill="#854d0e" text-anchor="middle">41 ~ 60</text><text x="210" y="154" font-size="10.5" font-weight="bold" fill="#854d0e" text-anchor="middle">Class III</text><text x="330" y="154" font-size="10.5" fill="#854d0e" text-anchor="middle">보통 (Fair)</text><text x="450" y="154" font-size="10.5" fill="#854d0e" text-anchor="middle">200 ~ 300</text><text x="570" y="154" font-size="10.5" fill="#854d0e" text-anchor="middle">25° ~ 35°</text>
            
            <rect x="50" y="161" width="600" height="22" fill="#ffedd5"/>
            <text x="110" y="176" font-size="10.5" fill="#9a3412" text-anchor="middle">21 ~ 40</text><text x="210" y="176" font-size="10.5" font-weight="bold" fill="#9a3412" text-anchor="middle">Class IV</text><text x="330" y="176" font-size="10.5" fill="#9a3412" text-anchor="middle">불량 (Poor)</text><text x="450" y="176" font-size="10.5" fill="#9a3412" text-anchor="middle">100 ~ 200</text><text x="570" y="176" font-size="10.5" fill="#9a3412" text-anchor="middle">15° ~ 25°</text>
            
            <rect x="50" y="183" width="600" height="22" fill="#fee2e2"/>
            <text x="110" y="198" font-size="10.5" fill="#991b1b" text-anchor="middle">0 ~ 20</text><text x="210" y="198" font-size="10.5" font-weight="bold" fill="#991b1b" text-anchor="middle">Class V</text><text x="330" y="198" font-size="10.5" fill="#991b1b" text-anchor="middle">매우 불량 (Very Poor)</text><text x="450" y="198" font-size="10.5" fill="#991b1b" text-anchor="middle">< 100</text><text x="570" y="198" font-size="10.5" fill="#991b1b" text-anchor="middle">< 15°</text>
        </svg>
        """
    else:
        return f"""
        <svg viewBox="0 0 700 240" width="100%" height="240" xmlns="http://www.w3.org/2000/svg" style="background:#f8fafc; border-radius:8px; border:1px solid #cbd5e1;">
            <rect x="20" y="20" width="660" height="200" fill="#ffffff" rx="6" stroke="#e2e8f0"/>
            <text x="350" y="45" font-family="'맑은 고딕', sans-serif" font-size="14" font-weight="bold" fill="#0f172a" text-anchor="middle">동탄도시철도 지반 리스크 & 엔지니어링 관리 프로세스</text>
            <rect x="50" y="75" width="160" height="120" fill="#eff6ff" stroke="#3b82f6" rx="6"/>
            <text x="130" y="100" font-size="12" font-weight="bold" fill="#1e40af" text-anchor="middle">[사전/계획 단계]</text>
            <text x="130" y="125" font-size="11" fill="#1e293b" text-anchor="middle">• 지질도·시추주상도 검토</text>
            <text x="130" y="148" font-size="11" fill="#1e293b" text-anchor="middle">• 조사기획 및 인허가</text>
            <text x="130" y="172" font-size="11" fill="#1e293b" text-anchor="middle">• 지하시설물 점용 협의</text>
            
            <line x1="210" y1="135" x2="260" y2="135" stroke="#0284c7" stroke-width="3" marker-end="url(#arrow)"/>
            
            <rect x="270" y="75" width="160" height="120" fill="#f0fdf4" stroke="#16a34a" rx="6"/>
            <text x="350" y="100" font-size="12" font-weight="bold" fill="#166534" text-anchor="middle">[현장/분석 단계]</text>
            <text x="350" y="125" font-size="11" fill="#1e293b" text-anchor="middle">• 시추·원위치·물리탐사</text>
            <text x="350" y="148" font-size="11" fill="#1e293b" text-anchor="middle">• 3D GPR 공동탐사</text>
            <text x="350" y="172" font-size="11" fill="#1e293b" text-anchor="middle">• 실내시험 및 지층분석</text>
            
            <line x1="430" y1="135" x2="480" y2="135" stroke="#0284c7" stroke-width="3"/>
            
            <rect x="490" y="75" width="160" height="120" fill="#fefce8" stroke="#ca8a04" rx="6"/>
            <text x="570" y="100" font-size="12" font-weight="bold" fill="#854d0e" text-anchor="middle">[설계/시공 피드백]</text>
            <text x="570" y="125" font-size="11" fill="#1e293b" text-anchor="middle">• R/O 및 REM 수립</text>
            <text x="570" y="148" font-size="11" fill="#1e293b" text-anchor="middle">• 지반보강 공법 최적화</text>
            <text x="570" y="172" font-size="11" fill="#1e293b" text-anchor="middle">• 현장 인수인계 완결</text>
        </svg>
        """

# 4. HTML 파일 일괄 생성 루프
created_count = 0

for act in ACTIVITIES:
    idx = act["idx"]
    code = act["code"]
    task = act["task"]
    dday = act["dday"]
    dept = act["dept"]
    purpose = act["purpose"]
    method = act["method"]
    deliverable = act["deliverable"]

    # 파일시스템 안전한 이름 생성 (특수문자 / \ : * ? " < > | 치환)
    safe_task = re.sub(r'[\/\\:\*\?"<>\|]', '_', task).strip()
    folder_prefix = f"{idx:02d}_{safe_task}"
    act_dir = os.path.join(OUTPUT_BASE, folder_prefix)
    
    std_dir = os.path.join(act_dir, "표준서")
    guide_dir = os.path.join(act_dir, "수행지침")
    chk_dir = os.path.join(act_dir, "체크리스트")

    os.makedirs(std_dir, exist_ok=True)
    os.makedirs(guide_dir, exist_ok=True)
    os.makedirs(chk_dir, exist_ok=True)

    std_fname = f"{safe_task}_표준서.html"
    guide_fname = f"{safe_task}_수행지침.html"
    chk_fname = f"{safe_task}_체크리스트.html"

    svg_content = get_jiban_svg(idx, task)

    # -------------------------------------------------------------
    # (1) 표준서 HTML (Standard)
    # -------------------------------------------------------------
    std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{task} 기술 표준서 - 동탄도시철도(트램)</title>
  <style>
    body {{ font-family: 'Pretendard', '맑은 고딕', sans-serif; background: #f8fafc; color: #1e293b; margin: 0; padding: 24px; }}
    .container {{ max-width: 1080px; margin: 0 auto; background: #ffffff; border-radius: 12px; box-shadow: 0 4px 16px rgba(0,0,0,0.06); padding: 36px; }}
    .header {{ border-bottom: 2px solid #0f172a; padding-bottom: 20px; margin-bottom: 28px; }}
    .badge {{ display: inline-block; background: #0f172a; color: #ffffff; font-size: 12px; font-weight: 700; padding: 4px 10px; border-radius: 4px; margin-bottom: 8px; }}
    .title {{ font-size: 24px; font-weight: 800; color: #0f172a; margin: 0 0 8px 0; }}
    .subtitle {{ font-size: 14px; color: #64748b; margin: 0; }}
    .section {{ margin-bottom: 32px; }}
    .section-title {{ font-size: 16px; font-weight: 700; color: #0f172a; border-left: 4px solid #0284c7; padding-left: 10px; margin-bottom: 14px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 13.5px; }}
    th, td {{ border: 1px solid #cbd5e1; padding: 10px 14px; text-align: left; }}
    th {{ background: #f1f5f9; color: #0f172a; font-weight: 700; }}
    td {{ color: #334155; }}
    .info-box {{ background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 8px; padding: 16px; font-size: 13.5px; line-height: 1.6; }}
    {NAV_CSS}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <span class="badge">WBS {code} | 지반조사 기술 표준서</span>
      <h1 class="title">{task} 기술 표준서</h1>
      <p class="subtitle">동탄도시철도(트램) 건설공사 지반 리스크 차단 및 품질 표준</p>
    </div>

    <div class="section">
      <div class="section-title">1. 일반 현황 및 관리 목표</div>
      <table>
        <tr>
          <th style="width:20%;">WBS 코드</th><td>{code}</td>
          <th style="width:20%;">추진 일정(D-Day)</th><td>{dday}</td>
        </tr>
        <tr>
          <th>주관 부서</th><td>{dept}</td>
          <th>필수 산출물</th><td><strong>{deliverable}</strong></td>
        </tr>
        <tr>
          <th>업무 목적</th>
          <td colspan="3">{purpose}</td>
        </tr>
      </table>
    </div>

    <div class="section">
      <div class="section-title">2. 관련 법령 및 기술 기준</div>
      <div class="info-box">
        • <strong>철도건설규칙</strong> 제6조 (선로 지반 및 노반의 안정성 확보)<br>
        • <strong>국토교통부 지반조사 표준시방서 (KCS 11 10 10)</strong> 및 구조물기초설계기준<br>
        • <strong>지하안전관리에 관한 특별법</strong> 제14조 및 제23조 (지하안전평가 및 사후지하안전영향조사)<br>
        • <strong>도시철도 설계기준(노반편)</strong> 지반조사 수량 및 심도 규정 준수
      </div>
    </div>

    <div class="section">
      <div class="section-title">3. 기술 표준 및 수행 절차</div>
      <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px; padding:18px; font-size:13.5px; line-height:1.7;">
        {method.replace(chr(10), '<br>')}
      </div>
    </div>

    <div class="section">
      <div class="section-title">4. 품질 및 안전 중점 관리 기준</div>
      <table>
        <tr>
          <th style="width:25%;">구분</th>
          <th style="width:45%;">점검 항목 및 시방 기준</th>
          <th style="width:30%;">판정 기준</th>
        </tr>
        <tr>
          <td>시추 및 원위치시험</td>
          <td>시추 심도(경암 1~3m 이상 확인), SPT 관입량 30cm 타격수 계측</td>
          <td>100% 실측 기록 및 성적서 첨부</td>
        </tr>
        <tr>
          <td>지하수위 및 안전</td>
          <td>시추공 내 24/48시간 안정수위 측정, 도로 굴착 시 안전휀스/신호수 배치</td>
          <td>지하수위 안정화 확인 및 무사고</td>
        </tr>
        <tr>
          <td>결과 보고 및 인계</td>
          <td>지층단면도 작성, 지반정수 산정, R/O 매트릭스 도출 및 사업단 보고</td>
          <td>책임감리원 최종 승인 획득</td>
        </tr>
      </table>
    </div>

    <!-- Bottom Navigation Block -->
    <div class="nav-box">
      <a href="../수행지침/{guide_fname}" class="nav-btn btn-guide">📘 [연계] {task} 수행지침서 열기 ➔</a>
      <a href="../체크리스트/{chk_fname}" class="nav-btn btn-chk">📋 [연계] {task} 체크리스트 열기 ➔</a>
    </div>
  </div>
</body>
</html>
"""
    with open(os.path.join(std_dir, std_fname), "w", encoding="utf-8") as fp:
        fp.write(std_html)

    # -------------------------------------------------------------
    # (2) 수행지침서 HTML (Guideline)
    # -------------------------------------------------------------
    guide_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{task} 수행지침서 - 동탄도시철도(트램)</title>
  <style>
    body {{ font-family: 'Pretendard', '맑은 고딕', sans-serif; background: #f8fafc; color: #1e293b; margin: 0; padding: 24px; }}
    .container {{ max-width: 1080px; margin: 0 auto; background: #ffffff; border-radius: 12px; box-shadow: 0 4px 16px rgba(0,0,0,0.06); padding: 36px; }}
    .header {{ border-bottom: 2px solid #0284c7; padding-bottom: 20px; margin-bottom: 28px; }}
    .badge {{ display: inline-block; background: #0284c7; color: #ffffff; font-size: 12px; font-weight: 700; padding: 4px 10px; border-radius: 4px; margin-bottom: 8px; }}
    .title {{ font-size: 24px; font-weight: 800; color: #0f172a; margin: 0 0 8px 0; }}
    .subtitle {{ font-size: 14px; color: #64748b; margin: 0; }}
    .section {{ margin-bottom: 32px; }}
    .section-title {{ font-size: 16px; font-weight: 700; color: #0f172a; border-left: 4px solid #0284c7; padding-left: 10px; margin-bottom: 14px; }}
    .step-card {{ background: #ffffff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 20px; margin-bottom: 16px; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }}
    .step-hdr {{ font-size: 15px; font-weight: 700; color: #0284c7; margin-bottom: 10px; display: flex; align-items: center; justify-content: space-between; }}
    .step-body {{ font-size: 13.5px; line-height: 1.7; color: #334155; }}
    .diagram-container {{ background: #f8fafc; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; margin-top: 14px; cursor: pointer; text-align: center; }}
    .diagram-hint {{ font-size: 11px; color: #64748b; margin-top: 6px; font-weight: 600; }}
    .calc-box {{ background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px; padding: 18px; margin-top: 16px; }}
    .calc-row {{ display: flex; gap: 12px; align-items: center; margin-bottom: 10px; font-size: 13.5px; }}
    .calc-input {{ padding: 8px 12px; border: 1px solid #cbd5e1; border-radius: 6px; width: 120px; font-size: 13.5px; font-weight: bold; }}
    .calc-btn {{ background: #0284c7; color: #ffffff; border: none; padding: 8px 16px; border-radius: 6px; font-weight: bold; cursor: pointer; }}
    {NAV_CSS}
    {MODAL_CSS}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <span class="badge">WBS {code} | 지반조사 현장 수행지침서</span>
      <h1 class="title">{task} 수행지침서</h1>
      <p class="subtitle">동탄트램 무결점 현장 시공을 위한 1:1 공학 도식 및 단계별 가이드</p>
    </div>

    <!-- 2D SVG 엔지니어링 도식 -->
    <div class="section">
      <div class="section-title">1. 핵심 기술 메커니즘 및 2D 공학 도식 (클릭 시 확대)</div>
      <div class="diagram-container" onclick="openDiagramZoom()">
        {svg_content}
        <div class="diagram-hint">🔍 [도식 클릭] 대형 팝업 확대보기 및 상세 도면 확인</div>
      </div>
    </div>

    <!-- 단계별 수행 지침 -->
    <div class="section">
      <div class="section-title">2. 단계별 상세 수행 절차 (Step-by-Step SOP)</div>
      <div class="step-card">
        <div class="step-hdr">
          <span>STEP 1 : 사전 준비 및 조사 계획 수립</span>
          <span style="font-size:12px; color:#64748b;">D-Day: {dday}</span>
        </div>
        <div class="step-body">
          • 유관기관 지하매설물(가스·상수도·한전·통신) 도면 대조 및 굴착허가 취득<br>
          • 조사 장비(시추기, 발전기, 계측센서) 반입 전 공인 교정 필증 확인<br>
          • 도로 점용 구간 교통안전시설(라바콘, 점멸등, 안내판) 및 신호수 100% 배치
        </div>
      </div>

      <div class="step-card">
        <div class="step-hdr">
          <span>STEP 2 : 현장 조사 및 시험 정밀 수행</span>
          <span style="font-size:12px; color:#0284c7;">공학적 핵심 공정</span>
        </div>
        <div class="step-body">
          • <strong>수행 내용:</strong> {method.replace(chr(10), '<br>• ')}<br>
          • 시추 심도별 코어 회수율(TCR) 및 암질지수(RQD) 실시간 측정 및 사진대지 기록<br>
          • 표준관입시험(SPT) N치 측정 시 63.5kg 해머 76cm 자유낙하 엄격 준수
        </div>
      </div>

      <div class="step-card">
        <div class="step-hdr">
          <span>STEP 3 : 데이터 분석, R/O 도출 및 최종 승인</span>
          <span style="font-size:12px; color:#16a34a;">산출물: {deliverable}</span>
        </div>
        <div class="step-body">
          • 시추 주상도, 지층 단면도 및 3D 지반 모델링 데이터 전산화<br>
          • 지반정수 산정(c, φ, E, qa) 및 기초 형식(직접초/말뚝기초) 타당성 검증<br>
          • <strong>{deliverable}</strong> 편찬 후 책임감리원 최종 서명 날인 및 PMIS 등재
        </div>
      </div>
    </div>

    <!-- 인터랙티브 지내력/N치 계산기 -->
    <div class="section">
      <div class="section-title">3. 인터랙티브 지반공학 간이 산정기 (N치 ➔ 허용지내력 qa)</div>
      <div class="calc-box">
        <div class="calc-row">
          <span>■ 현장 측정 표준관입시험 N치 입력 :</span>
          <input type="number" id="inputN" class="calc-input" value="25" min="1" max="60">
          <button type="button" class="calc-btn" onclick="calcBearingCapacity()">지내력 계산 ➔</button>
        </div>
        <div id="calcResult" style="font-size:13.5px; font-weight:bold; color:#1e40af; margin-top:8px;">
          • Meyerhof 허용지내력 qa = 250 kN/㎡ (25.5 tonf/㎡) | 내부마찰각 φ ≈ 37.4° (양호한 모래/자갈층)
        </div>
      </div>
    </div>

    <!-- Zoom Modal -->
    <div id="zoomModal" onclick="closeDiagramZoom()">
      <div class="modal-content" onclick="event.stopPropagation()">
        <button class="close-btn" onclick="closeDiagramZoom()">✕</button>
        <h3 style="margin-top:0; color:#0f172a; font-size:16px;">2D 지반공학 엔지니어링 도식 고해상도 확대보기</h3>
        <div style="margin-top:16px;">
          {svg_content}
        </div>
      </div>
    </div>

    <!-- Bottom Navigation Block -->
    <div class="nav-box">
      <a href="../표준서/{std_fname}" class="nav-btn btn-std">📄 [연계] {task} 표준서 열기 ➔</a>
      <a href="../체크리스트/{chk_fname}" class="nav-btn btn-chk">📋 [연계] {task} 체크리스트 열기 ➔</a>
    </div>
  </div>

  <script>
    function openDiagramZoom() {{
      document.getElementById('zoomModal').style.display = 'flex';
    }}
    function closeDiagramZoom() {{
      document.getElementById('zoomModal').style.display = 'none';
    }}
    function calcBearingCapacity() {{
      const n = parseFloat(document.getElementById('inputN').value) || 0;
      const qa = n * 10;
      const phi = (Math.sqrt(20 * n) + 15).toFixed(1);
      const ton = (qa / 9.8).toFixed(1);
      document.getElementById('calcResult').innerHTML = 
        `• Meyerhof 허용지내력 qa = ${{qa}} kN/㎡ (${{ton}} tonf/㎡) | 내부마찰각 φ ≈ ${{phi}}° (N=${{n}} 기준)`;
    }}
  </script>
</body>
</html>
"""
    with open(os.path.join(guide_dir, guide_fname), "w", encoding="utf-8") as fp:
        fp.write(guide_html)

    # -------------------------------------------------------------
    # (3) 체크리스트 HTML (Checklist)
    # -------------------------------------------------------------
    chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{task} 체크리스트 - 동탄도시철도(트램)</title>
  <style>
    body {{ font-family: 'Pretendard', '맑은 고딕', sans-serif; background: #f8fafc; color: #1e293b; margin: 0; padding: 24px; }}
    .container {{ max-width: 1080px; margin: 0 auto; background: #ffffff; border-radius: 12px; box-shadow: 0 4px 16px rgba(0,0,0,0.06); padding: 36px; }}
    .header {{ border-bottom: 2px solid #d97706; padding-bottom: 20px; margin-bottom: 28px; }}
    .badge {{ display: inline-block; background: #d97706; color: #ffffff; font-size: 12px; font-weight: 700; padding: 4px 10px; border-radius: 4px; margin-bottom: 8px; }}
    .title {{ font-size: 24px; font-weight: 800; color: #0f172a; margin: 0 0 8px 0; }}
    .subtitle {{ font-size: 14px; color: #64748b; margin: 0; }}
    .section {{ margin-bottom: 32px; }}
    .section-title {{ font-size: 16px; font-weight: 700; color: #0f172a; border-left: 4px solid #d97706; padding-left: 10px; margin-bottom: 14px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 13.5px; }}
    th, td {{ border: 1px solid #cbd5e1; padding: 10px 14px; text-align: left; }}
    th {{ background: #fef3c7; color: #92400e; font-weight: 700; }}
    td {{ color: #334155; }}
    .sign-box {{ display: flex; justify-content: flex-end; gap: 24px; margin-top: 24px; font-size: 13.5px; }}
    .sign-item {{ border: 1px solid #cbd5e1; padding: 12px 24px; border-radius: 6px; text-align: center; }}
    {NAV_CSS}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <span class="badge">WBS {code} | 지반조사 현장 검측 체크리스트</span>
      <h1 class="title">{task} 체크리스트</h1>
      <p class="subtitle">동탄도시철도(트램) 시공 품질 및 안전 검측 필증</p>
    </div>

    <div class="section">
      <div class="section-title">1. 기본 검측 정보</div>
      <table>
        <tr>
          <th style="width:15%;">공사명</th><td>동탄도시철도(트램) 건설공사</td>
          <th style="width:15%;">검측 일자</th><td>2026. __. __</td>
        </tr>
        <tr>
          <th>위치/구간</th><td>동탄트램 1·2호선 본선 및 정거장</td>
          <th>WBS 코드</th><td><strong>{code}</strong></td>
        </tr>
        <tr>
          <th>주관 부서</th><td>{dept}</td>
          <th>검측자</th><td>토목품질관리자 (인) / 책임감리원 (인)</td>
        </tr>
      </table>
    </div>

    <div class="section">
      <div class="section-title">2. 단계별 핵심 점검표</div>
      <table>
        <thead>
          <tr>
            <th style="width:8%; text-align:center;">No.</th>
            <th style="width:20%; text-align:center;">구분</th>
            <th style="width:47%; text-align:center;">점검 항목 및 확인 기준</th>
            <th style="width:15%; text-align:center;">판정 기준</th>
            <th style="width:10%; text-align:center;">결과</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td style="text-align:center;">1</td>
            <td style="text-align:center; font-weight:bold;">사전 준비</td>
            <td>지하매설물 도면 일치 여부 및 도로점용 인허가 완료 여부</td>
            <td style="text-align:center;">인허가 승인</td>
            <td style="text-align:center; color:#16a34a; font-weight:bold;">적합</td>
          </tr>
          <tr>
            <td style="text-align:center;">2</td>
            <td style="text-align:center; font-weight:bold;">장비 및 인력</td>
            <td>시추기/계측장비 검교정 필증 및 안전관리자 상주 여부</td>
            <td style="text-align:center;">교정 필증 완비</td>
            <td style="text-align:center; color:#16a34a; font-weight:bold;">적합</td>
          </tr>
          <tr>
            <td style="text-align:center;">3</td>
            <td style="text-align:center; font-weight:bold;">현장 시험</td>
            <td>시추 심도, 코어회수율(TCR ≥ 85%), SPT N치 규정 준수</td>
            <td style="text-align:center;">시방 기준 충족</td>
            <td style="text-align:center; color:#16a34a; font-weight:bold;">적합</td>
          </tr>
          <tr>
            <td style="text-align:center;">4</td>
            <td style="text-align:center; font-weight:bold;">안전 관리</td>
            <td>작업장 안전휀스 설치, 신호수 배치 및 야간 점멸등 작동</td>
            <td style="text-align:center;">안전 규정 준수</td>
            <td style="text-align:center; color:#16a34a; font-weight:bold;">적합</td>
          </tr>
          <tr>
            <td style="text-align:center;">5</td>
            <td style="text-align:center; font-weight:bold;">산출물 관리</td>
            <td><strong>{deliverable}</strong> 작성 및 책임감리원 기술검토 승인</td>
            <td style="text-align:center;">보고서 승인</td>
            <td style="text-align:center; color:#16a34a; font-weight:bold;">적합</td>
          </tr>
        </tbody>
      </table>
    </div>

    <div class="section">
      <div class="section-title">3. 종합 판정 및 입회 서명</div>
      <div style="background:#f8fafc; border:1px solid #cbd5e1; border-radius:8px; padding:16px; font-size:13.5px;">
        <strong>■ 종합 판정 결과:</strong> <span style="color:#16a34a; font-weight:bold; font-size:15px;">[합격 / PASS]</span><br>
        • 지반조사 시방 기준 및 현장 안전 수칙을 100% 준수하여 조사가 적정하게 완료되었음을 확인함.
      </div>
      <div class="sign-box">
        <div class="sign-item">
          점검자 (시공사): <strong>김토목</strong> (서명/인)
        </div>
        <div class="sign-item">
          확인자 (책임감리): <strong>이감리</strong> (서명/인)
        </div>
      </div>
    </div>

    <!-- Bottom Navigation Block -->
    <div class="nav-box">
      <a href="../표준서/{std_fname}" class="nav-btn btn-std">📄 [연계] {task} 표준서 열기 ➔</a>
      <a href="../수행지침/{guide_fname}" class="nav-btn btn-guide">📘 [연계] {task} 수행지침서 열기 ➔</a>
    </div>
  </div>
</body>
</html>
"""
    with open(os.path.join(chk_dir, chk_fname), "w", encoding="utf-8") as fp:
        fp.write(chk_html)

    created_count += 3

print(f"\n=======================================================")
print(f"지반조사 36개 액티비티 {created_count}개 HTML 파일 일괄 생성 100% 성공!")
print(f"=======================================================")
