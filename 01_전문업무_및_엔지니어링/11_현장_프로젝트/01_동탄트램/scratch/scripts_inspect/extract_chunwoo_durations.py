import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def extract_durations_from_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Header scan in row 16-22
    header_row_16 = [ws.cell(row=16, column=c).value for c in range(1, ws.max_column+1)]
    header_row_21 = [ws.cell(row=21, column=c).value for c in range(1, ws.max_column+1)]

    novan_col, gwedo_col, pojang_col, total_col = None, None, None, None
    for idx, val in enumerate(header_row_21):
        if val == '노반': novan_col = idx + 1
        elif val == '궤도': gwedo_col = idx + 1
        elif val == '포장': pojang_col = idx + 1
        elif val == '전체': total_col = idx + 1

    print(f"Detected columns - 노반: {novan_col}, 궤도: {gwedo_col}, 포장: {pojang_col}, 전체: {total_col}")
    
    sec_data = {}
    for r in range(23, ws.max_row + 1):
        sec_name = ws.cell(row=r, column=4).value # D col (index 4)
        if not sec_name:
            sec_name = ws.cell(row=r, column=3).value # C col (index 3)
        if sec_name and isinstance(sec_name, str) and len(sec_name.strip()) > 1 and not '공사기간' in sec_name and not '샘플' in sec_name:
            s_name = sec_name.strip()
            n_days = ws.cell(row=r, column=novan_col).value if novan_col else 0
            g_days = ws.cell(row=r, column=gwedo_col).value if gwedo_col else 0
            p_days = ws.cell(row=r, column=pojang_col).value if pojang_col else 0
            t_days = ws.cell(row=r, column=total_col).value if total_col else 0
            
            sec_data[s_name] = {
                'novanDays': int(n_days) if isinstance(n_days, (int, float)) else 0,
                'gwedoDays': int(g_days) if isinstance(g_days, (int, float)) else 0,
                'pojangDays': int(p_days) if isinstance(p_days, (int, float)) else 0,
                'totalDays': int(t_days) if isinstance(t_days, (int, float)) else 0
            }
            print(f"  {s_name} => 노반:{sec_data[s_name]['novanDays']}일 | 궤도:{sec_data[s_name]['gwedoDays']}일 | 포장:{sec_data[s_name]['pojangDays']}일 | 합계:{sec_data[s_name]['totalDays']}일")

    return sec_data

print("=== 1공구 천우씨엠 소요일수 추출 ===")
d1 = extract_durations_from_file(file1)

print("\n=== 2공구 천우씨엠 소요일수 추출 ===")
d2 = extract_durations_from_file(file2)

merged_durations = {**d1, **d2}
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\chunwoo_durations.json', 'w', encoding='utf-8') as f:
    json.dump(merged_durations, f, ensure_ascii=False, indent=2)

print(f"\n✓ Extracted total {len(merged_durations)} section durations successfully!")
