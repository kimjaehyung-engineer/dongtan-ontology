import openpyxl, json, os, sys, re

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def inspect_header_totals(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Search all rows for "노반합계" or "노반 합계", "궤도합계" or "궤도 합계", "포장합계" or "포장 합계"
    novan_col, gwedo_col, pojang_col = None, None, None
    
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').replace(' ', '').strip()
            if '노반합계' in val:
                novan_col = c
            elif '궤도합계' in val:
                gwedo_col = c
            elif '포장합계' in val:
                pojang_col = c

    print(f"[{zone_name}] Identified Total Columns - 노반합계: C{novan_col}, 궤도합계: C{gwedo_col}, 포장합계: C{pojang_col}")

    section_totals = {}
    for r in range(1, ws.max_row + 1):
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        c4 = ws.cell(row=r, column=4).value
        
        name = None
        for cand in [c3, c4, c2]:
            if cand and isinstance(cand, str) and ('(본)' in cand or '(기' in cand):
                name = cand.strip()
                break
        
        if name:
            n_val = ws.cell(row=r, column=novan_col).value if novan_col else 0
            g_val = ws.cell(row=r, column=gwedo_col).value if gwedo_col else 0
            p_val = ws.cell(row=r, column=pojang_col).value if pojang_col else 0

            nD = int(round(float(n_val))) if isinstance(n_val, (int, float)) and n_val > 0 else 0
            gD = int(round(float(g_val))) if isinstance(g_val, (int, float)) and g_val > 0 else 0
            pD = int(round(float(p_val))) if isinstance(p_val, (int, float)) and p_val > 0 else 0
            tD = nD + gD + pD

            section_totals[name] = {
                'zone': zone_name,
                'novanDays': nD,
                'gwedoDays': gD,
                'pojangDays': pD,
                'totalDays': tD
            }
            print(f"  [{zone_name}] {name:<30} => 🟠노반합계:{nD}일 | 🟢궤도합계:{gD}일 | 🟦포장합계:{pD}일 | 합계:{tD}일")

    return section_totals

print("=== 1공구 노반합계, 궤도합계, 포장합계 스캔 ===")
d1 = inspect_header_totals(file1, "1공구")

print("\n=== 2공구 노반합계, 궤도합계, 포장합계 스캔 ===")
d2 = inspect_header_totals(file2, "2공구")

all_totals = {**d1, **d2}
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\chunwoo_exact_header_totals.json', 'w', encoding='utf-8') as f:
    json.dump(all_totals, f, ensure_ascii=False, indent=2)

print(f"\n✓ Successfully extracted exact 노반합계, 궤도합계, 포장합계 for {len(all_totals)} sections!")
