import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

wb = openpyxl.load_workbook(file1, data_only=True)
ws = wb.active

for r in range(1, 45):
    row_cells = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != '':
            row_cells.append(f"C{c}:{v}")
    if row_cells:
        print(f"Row {r:2d}: {' | '.join(row_cells[:8])}")
