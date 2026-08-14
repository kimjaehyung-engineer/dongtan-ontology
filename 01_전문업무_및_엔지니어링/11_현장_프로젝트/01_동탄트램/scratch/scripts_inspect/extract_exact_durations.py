import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def extract_exact_durations(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    novan_col, gwedo_col, pojang_col, total_col = None, None, None, None

    # Scan rows 1 to 25 to find header positions
    for r in range(1, 25):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').strip()
            if val == '소요공기 정리':
                # Check next rows for 노반, 궤도, 포장, 전체
                for r2 in range(r, r + 5):
                    for c2 in range(c, c + 15):
                        v2 = str(ws.cell(row=r2, column=c2).value or '').strip()
                        if v2 == '노반': novan_col = c2
                        elif v2 == '궤도': gwedo_col = c2
                        elif v2 == '포장': pojang_col = c2
                        elif v2 == '전체': total_col = c2
                break

    print(f"[{zone_name}] Found columns - 노반: {novan_col}, 궤도: {gwedo_col}, 포장: {pojang_col}, 전체: {total_col}")

    durations = {}
    for r in range(1, ws.max_row + 1):
        sec_name = ws.cell(row=r, column=4).value or ws.cell(row=r, column=3).value
        if sec_name and isinstance(sec_name, str) and ('(본)' in sec_name or '(기' in sec_name):
            s_name = sec_name.strip()
            n_val = ws.cell(row=r, column=novan_col).value if novan_col else 0
            g_val = ws.cell(row=r, column=gwedo_col).value if gwedo_col else 0
            p_val = ws.cell(row=r, column=pojang_col).value if pojang_col else 0
            t_val = ws.cell(row=r, column=total_col).value if total_col else 0

            n_days = int(n_val) if isinstance(n_val, (int, float)) else 0
            g_days = int(g_val) if isinstance(g_val, (int, float)) else 0
            p_days = int(p_val) if isinstance(p_val, (int, float)) else 0
            t_days = int(t_val) if isinstance(t_val, (int, float)) else 0

            durations[s_name] = {
                'zone': zone_name,
                'novanDays': n_days,
                'gwedoDays': g_days,
                'pojangDays': p_days,
                'totalDays': t_days
            }
            print(f"  {s_name} => 🟠노반:{n_days}일 | 🟢궤도:{g_days}일 | 🟦포장:{p_days}일 | ➔ 총합:{t_days}일")

    return durations

d1 = extract_exact_durations(file1, "1공구")
d2 = extract_exact_durations(file2, "2공구")

all_durations = {**d1, **d2}
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\chunwoo_durations.json', 'w', encoding='utf-8') as f:
    json.dump(all_durations, f, ensure_ascii=False, indent=2)

print(f"\n✓ Successfully extracted exact Chunwoo CM durations for {len(all_durations)} sections!")
