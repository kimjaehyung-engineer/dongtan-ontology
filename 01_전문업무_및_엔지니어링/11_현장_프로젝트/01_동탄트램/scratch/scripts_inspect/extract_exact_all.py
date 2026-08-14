import openpyxl, json, os, sys, re

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def extract_all_rows(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Header scan in row 21
    novan_col, gwedo_col, pojang_col, total_col = 95, 96, 97, 98
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=21, column=c).value or '').strip()
        if v == '노반': novan_col = c
        elif v == '궤도': gwedo_col = c
        elif v == '포장': pojang_col = c
        elif v == '전체': total_col = c

    data_map = {}
    for r in range(1, ws.max_row + 1):
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        c4 = ws.cell(row=r, column=4).value
        
        name = None
        for cand in [c3, c4, c2]:
            if cand and isinstance(cand, str) and ('(본)' in cand or '(기' in cand):
                name = cand.strip()
                break
        
        if name:
            n_val = ws.cell(row=r, column=novan_col).value
            g_val = ws.cell(row=r, column=gwedo_col).value
            p_val = ws.cell(row=r, column=pojang_col).value
            t_val = ws.cell(row=r, column=total_col).value

            nD = int(n_val) if isinstance(n_val, (int, float)) and n_val > 0 else 0
            gD = int(g_val) if isinstance(g_val, (int, float)) and g_val > 0 else 0
            pD = int(p_val) if isinstance(p_val, (int, float)) and p_val > 0 else 0
            tD = int(t_val) if isinstance(t_val, (int, float)) and t_val > 0 else (nD + gD + pD)

            data_map[name] = {
                'zone': zone_name,
                'novanDays': nD,
                'gwedoDays': gD,
                'pojangDays': pD,
                'totalDays': tD
            }
            print(f"[{zone_name}] {name} => 🟠노반:{nD}d | 🟢궤도:{gD}d | 🟦포장:{pD}d | ➔ 총:{tD}d")

    return data_map

d1 = extract_all_rows(file1, "1공구")
d2 = extract_all_rows(file2, "2공구")

all_parsed = {**d1, **d2}
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\chunwoo_exact_all.json', 'w', encoding='utf-8') as f:
    json.dump(all_parsed, f, ensure_ascii=False, indent=2)

print(f"\n✓ Successfully parsed exact durations for {len(all_parsed)} sections!")
