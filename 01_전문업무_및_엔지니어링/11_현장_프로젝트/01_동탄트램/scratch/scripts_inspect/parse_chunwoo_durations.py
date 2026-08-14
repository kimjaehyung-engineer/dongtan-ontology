import openpyxl, glob, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

durations = {}

def parse_excel(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        sec_name = ws.cell(row=r, column=4).value # D col: 구간명 (본)1-1..
        if sec_name and isinstance(sec_name, str) and ('(본)' in sec_name or '(기' in sec_name or '(' in sec_name):
            sec_name = sec_name.strip()
            # Find novan, gwedo, pojang columns
            # In 1공구 sheet: Col 100: Novan (일), Col 101: Gwedo (일), Col 102: Pojang (일), Col 103: Total
            # Let's inspect row for days values
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            nums = [v for v in row_vals if isinstance(v, (int, float)) and v > 0]
            
            # Look from the end for integer days
            # Usually: novan_days, gwedo_days, pojang_days, total_days
            # Let's scan from right side:
            valid_days = []
            for c in range(len(row_vals)-1, -1, -1):
                v = row_vals[c]
                if isinstance(v, (int, float)) and 1 <= v <= 200:
                    valid_days.append((c+1, v))
            
            # Print for verification
            print(f"[{zone_name}] {sec_name}: {valid_days[:10]}")

print("=== Parsing 1 Gonggu ===")
parse_excel(file1, "1공구")

print("\n=== Parsing 2 Gonggu ===")
parse_excel(file2, "2공구")
