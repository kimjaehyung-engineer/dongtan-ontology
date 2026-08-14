import sys
import os
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 1공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 2공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'

def read_excel_info(path, label):
    print(f"\n=================== {label} ===================")
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheet names:", wb.sheetnames)
    sheet = wb.active
    
    data = []
    for r in range(1, 40):
        row_vals = [sheet.cell(r, c).value for c in range(1, 20)]
        if any(v is not None for v in row_vals):
            data.append((r, row_vals))
            
    for r_idx, r_vals in data[:25]:
        # Filter out trailing Nones for printing
        trimmed = [str(v) if v is not None else "" for v in r_vals]
        print(f"Row {r_idx:2d}: {' | '.join(trimmed[:15])}")

read_excel_info(file1, "1공구 교차로 정보")
read_excel_info(file2, "2공구 교차로 정보")
