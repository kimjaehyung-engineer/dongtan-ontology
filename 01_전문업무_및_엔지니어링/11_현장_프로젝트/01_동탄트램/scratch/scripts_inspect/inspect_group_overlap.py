import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def inspect_group_summary(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    print(f"=== {zone_name} 전체 구간 행 데이터 파싱 ===")
    rows = []
    for r in range(1, ws.max_row + 1):
        sec_code = ws.cell(row=r, column=3).value or '' # C col: (본)1-1...
        sec_name = ws.cell(row=r, column=4).value or '' # D col
        grp_name = ws.cell(row=r, column=13).value or ws.cell(row=r, column=12).value or '' # M/L col: 시공구간 분할 (일반부지 1구간...)
        
        # Novan/Gwedo/Pojang days at col 95, 96, 97, 98
        n_days = ws.cell(row=r, column=95).value or 0
        g_days = ws.cell(row=r, column=96).value or 0
        p_days = ws.cell(row=r, column=97).value or 0
        t_days = ws.cell(row=r, column=98).value or 0

        # Critical path / Real duration summary cols (right side)
        # Col 99~112: 소요공기 정리 (노반, 궤도, 포장, 전체), 실 공사가능기간 등
        r_days = ws.cell(row=r, column=101).value or ws.cell(row=r, column=102).value or 0

        if sec_name and isinstance(sec_name, str) and ('(본)' in sec_name or '(기' in sec_name):
            print(f"Row {r:2d} | [{grp_name}] {sec_name.strip():<25} | 노반:{n_days}d 궤도:{g_days}d 포장:{p_days}d | 합계:{t_days}d | 실공기:{r_days}d")

inspect_group_summary(file1, "1공구")
print("\n")
inspect_group_summary(file2, "2공구")
