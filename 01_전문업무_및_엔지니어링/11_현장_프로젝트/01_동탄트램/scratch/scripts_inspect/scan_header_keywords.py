import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

wb = openpyxl.load_workbook(file1, data_only=True)
ws = wb.active

print("=== 1공구 헤더 내 '노반', '궤도', '포장', '합계' 키워드 셀 탐색 ===")
for r in range(1, 25):
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r, column=c).value or '').strip()
        if '노반' in val or '궤도' in val or '포장' in val or '합계' in val:
            print(f"Row {r:2d}, Col {c:3d}: '{val}'")
