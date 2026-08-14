import openpyxl, json, os, sys

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

def extract_exact_totals_from_cols(filepath, zone_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find exact col index for '노반합계', '궤도합계', '포장합계'
    novan_col, gwedo_col, pojang_col = 42, 69, 94
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value or '').replace('\n', '').replace(' ', '').strip()
            if val == '노반합계': novan_col = c
            elif val == '궤도합계': gwedo_col = c
            elif val == '포장합계': pojang_col = c

    print(f"[{zone_name}] Found exact header columns - 노반합계: Col {novan_col}, 궤도합계: Col {gwedo_col}, 포장합계: Col {pojang_col}")

    sec_totals = {}
    for r in range(1, ws.max_row + 1):
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        c4 = ws.cell(row=r, column=4).value
        
        name = None
        for cand in [c3, c4, c2]:
            if cand and isinstance(cand, str) and ('(본)' in cand or '(기' in cand):
                name = cand.strip()
                break

        if name:
            n_val = ws.cell(row=r, column=novan_col).value
            g_val = ws.cell(row=r, column=gwedo_col).value
            p_val = ws.cell(row=r, column=pojang_col).value

            nD = int(round(float(n_val))) if isinstance(n_val, (int, float)) and n_val > 0 else 0
            gD = int(round(float(g_val))) if isinstance(g_val, (int, float)) and g_val > 0 else 0
            pD = int(round(float(p_val))) if isinstance(p_val, (int, float)) and p_val > 0 else 0
            tD = nD + gD + pD

            sec_totals[name] = {
                'zone': zone_name,
                'novanDays': nD,
                'gwedoDays': gD,
                'pojangDays': pD,
                'totalDays': tD
            }
            print(f"  [{zone_name}] {name:<32} => 🟠노반합계:{nD}일 | 🟢궤도합계:{gD}일 | 🟦포장합계:{pD}일 | ➔ 개별총합:{tD}일")

    return sec_totals

print("=== 1공구 exact 노반합계/궤도합계/포장합계 데이터 추출 ===")
d1 = extract_exact_totals_from_cols(file1, "1공구")

print("\n=== 2공구 exact 노반합계/궤도합계/포장합계 데이터 추출 ===")
d2 = extract_exact_totals_from_cols(file2, "2공구")

merged_all = {**d1, **d2}
with open(r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\chunwoo_exact_header_totals.json', 'w', encoding='utf-8') as f:
    json.dump(merged_all, f, ensure_ascii=False, indent=2)

print(f"\n✓ Successfully extracted exact 노반합계, 궤도합계, 포장합계 for {len(merged_all)} sections!")
