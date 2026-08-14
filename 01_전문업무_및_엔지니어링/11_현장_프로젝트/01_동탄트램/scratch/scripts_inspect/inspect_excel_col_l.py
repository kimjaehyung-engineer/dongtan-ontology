import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

# 1공구 본선구간
path1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
path2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

for label, path in [("1공구", path1), ("2공구", path2)]:
    print(f"\n{'='*60}")
    print(f"=== {label} 본선구간 ===")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")
        # Print first 30 rows, columns A-N
        for row in range(1, min(ws.max_row + 1, 35)):
            vals = []
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                cell = ws[f'{col_letter}{row}']
                vals.append(str(cell.value) if cell.value is not None else '')
            print(f"  Row {row:3d}: {' | '.join(vals)}")
    wb.close()
