import sys
import os
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')

file1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 1공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'
file2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\05_기술제안 2공구 교차로구간 공기산출 근거_(주)천우씨엠_PST삭제.xlsx'

def get_all_rows(path, tool_label):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    rows_data = []
    
    # Header starts around row 5 in 1공구, row 8 in 2공구
    for r in range(1, 100):
        no_val = sheet.cell(r, 2).value
        # Check if no_val is numeric or valid row
        if isinstance(no_val, (int, float)) and no_val > 0:
            section = sheet.cell(r, 3).value
            name = sheet.cell(r, 4).value
            start_sta = sheet.cell(r, 9).value
            end_sta = sheet.cell(r, 10).value
            length = sheet.cell(r, 11).value
            method = sheet.cell(r, 12).value
            
            # Step / stage col
            if tool_label == '1공구':
                stage = sheet.cell(r, 13).value
                avg_len = sheet.cell(r, 14).value
            else: # 2공구
                stage_orig = sheet.cell(r, 13).value
                stage = sheet.cell(r, 14).value # 변경 단계
                avg_len = sheet.cell(r, 15).value
                
            rows_data.append({
                'tool': tool_label,
                'no': int(no_val),
                'section': str(section or ''),
                'name': str(name or ''),
                'start_sta': start_sta,
                'end_sta': end_sta,
                'length': length,
                'method': str(method or ''),
                'stage': stage,
                'avg_len': avg_len
            })
    return rows_data

data1 = get_all_rows(file1, '1공구')
data2 = get_all_rows(file2, '2공구')

print(f"Total 1공구 intersections: {len(data1)}")
for d in data1:
    print(f"1공구 #{d['no']:2d} | {d['section']:20s} | {d['name']:25s} | STA: {d['start_sta']}~{d['end_sta']} | 연장: {d['length']:.1f}m | 단계: {d['stage']}단계 | 평균작업연장: {d['avg_len']:.1f}m")

print(f"\nTotal 2공구 intersections: {len(data2)}")
for d in data2:
    avg_l = f"{d['avg_len']:.1f}m" if isinstance(d['avg_len'], (int, float)) else str(d['avg_len'])
    print(f"2공구 #{d['no']:2d} | {d['section']:20s} | {d['name']:25s} | STA: {d['start_sta']}~{d['end_sta']} | 연장: {d['length']:.1f}m | 단계: {d['stage']}단계 | 평균작업연장: {avg_l}")
