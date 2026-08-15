# -*- coding: utf-8 -*-
"""
동탄트램 콘크리트도상 시트 9000-5-13 ~ 9000-5-23 행 표준서/수행지침/체크리스트 하이퍼링크 수식 완결 연동 스크립트
"""

import os
import sys
import re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8")
TRACKBED_DIR = os.path.join(BASE_DIR, "5.콘크리트도상")

# 중복 폴더 정리 (폴더명 일원화)
duplicate_merges = [
    ("17_[TCL] 궤광 및 철근 조립", "17_[TCL] 궤광 및 철근조립"),
    ("18_[TCL] 거푸집설치", "18_[TCL] 거푸집 설치"),
    ("19_[TCL] 콘크리트타설및양생", "19_[TCL] 콘크리트 타설 및 양생"),
    ("20_[레일용접] 가스압접", "20_[레일용접] 가스 압접"),
    ("21_[레일용접] 테르밋용접", "21_[레일용접] 테르밋 용접")
]

for src_name, dst_name in duplicate_merges:
    src_p = os.path.join(TRACKBED_DIR, src_name)
    dst_p = os.path.join(TRACKBED_DIR, dst_name)
    if os.path.exists(src_p) and os.path.exists(dst_p):
        import shutil
        for sub in ['표준서', '수행지침', '체크리스트']:
            s_sub = os.path.join(src_p, sub)
            d_sub = os.path.join(dst_p, sub)
            if os.path.exists(s_sub):
                os.makedirs(d_sub, exist_ok=True)
                for fn in os.listdir(s_sub):
                    src_f = os.path.join(s_sub, fn)
                    dst_f = os.path.join(d_sub, fn)
                    if not os.path.exists(dst_f):
                        shutil.copy2(src_f, dst_f)
        shutil.rmtree(src_p)
        print(f"  중복 폴더 정리 완료: {src_name} ➔ {dst_name}")

# 더미 Task_24 폴더 제거
dummy_24 = os.path.join(TRACKBED_DIR, "24_Task_24")
if os.path.exists(dummy_24):
    import shutil
    shutil.rmtree(dummy_24)
    print("  더미 폴더 제거: 24_Task_24")

link_font_std = Font(name="맑은 고딕", size=9, bold=True, color="047857", underline="single")
link_font_guide = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")
link_font_chk = Font(name="맑은 고딕", size=9, bold=True, color="D97706", underline="single")

fill_link_std = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
fill_link_guide = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
fill_link_chk = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue

    print(f"\n=======================================================")
    print(f"콘크리트도상 하이퍼링크 수식 업데이트: {os.path.basename(file_path)}")
    wb = openpyxl.load_workbook(file_path)

    if "콘크리트도상" not in wb.sheetnames:
        continue

    ws = wb["콘크리트도상"]

    # 13개 세부 항목 및 1~23 전체 매핑
    for r in range(2, 25):
        idx = r - 1
        ws.cell(r, 4).value = f"9000-5-{idx}"
        task = ws.cell(r, 7).value
        if not task:
            continue

        # 해당 폴더 찾기
        matching_dirs = [d for d in os.listdir(TRACKBED_DIR) if d.startswith(f"{idx}_")]
        if not matching_dirs:
            print(f"❌ 폴더 없음: {idx}_ (Task: {task})")
            continue

        folder_name = matching_dirs[0]
        folder_path = os.path.join(TRACKBED_DIR, folder_name)

        std_dir = os.path.join(folder_path, "표준서")
        guide_dir = os.path.join(folder_path, "수행지침")
        chk_dir = os.path.join(folder_path, "체크리스트")

        def get_best_file(dir_path, kw):
            if not os.path.exists(dir_path): return ""
            files = [f for f in os.listdir(dir_path) if f.endswith(".html")]
            if not files: return ""
            for f in files:
                if not f.startswith(f"{idx}_") and kw in f:
                    return f
            for f in files:
                if kw in f:
                    return f
            return files[0]

        std_file = get_best_file(std_dir, "표준서")
        guide_file = get_best_file(guide_dir, "수행지침")
        chk_file = get_best_file(chk_dir, "체크리스트")

        std_rel = f"5.콘크리트도상/{folder_name}/표준서/{std_file}"
        guide_rel = f"5.콘크리트도상/{folder_name}/수행지침/{guide_file}"
        chk_rel = f"5.콘크리트도상/{folder_name}/체크리스트/{chk_file}"

        # N열: 표준서
        f_std = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{std_rel}", "표준서 열기")'
        c_std = ws.cell(r, 14, value=f_std)
        c_std.font = link_font_std; c_std.fill = fill_link_std; c_std.alignment = align_center

        # P열: 수행지침
        f_guide = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{guide_rel}", "수행지침서 열기")'
        c_guide = ws.cell(r, 16, value=f_guide)
        c_guide.font = link_font_guide; c_guide.fill = fill_link_guide; c_guide.alignment = align_center

        # R열: 체크리스트
        f_chk = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{chk_rel}", "체크리스트 열기")'
        c_chk = ws.cell(r, 18, value=f_chk)
        c_chk.font = link_font_chk; c_chk.fill = fill_link_chk; c_chk.alignment = align_center

    wb.save(file_path)
    print(f"  ✔ {os.path.basename(file_path)} 콘크리트도상 23개 전 행 하이퍼링크 수식 기입 완료!")

print(f"\n=======================================================")
print(f"콘크리트도상 9000-5-13 ~ 9000-5-23 행 연동 100% 완료!")
print(f"=======================================================")
