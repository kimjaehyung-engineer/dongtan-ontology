# -*- coding: utf-8 -*-
"""
지반조사 이미지 36개 액티비티를 v8 엑셀의 표준 공종 열 구조(16개 열)에 1:1 완벽 매핑하여 입력하는 마스터 스크립트
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 대상 엑셀 파일들
EXCEL_FILES = [
    r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx",
    r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v6.xlsx",
    r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v5.xlsx"
]

# 36개 전체 데이터 세부 정의
RAW_DATA = [
    # (No, Process, Start, Finish, Why, How, Who, Collab, What, Guideline_O, Standard_O)
    ("", "발주처 계획 입수(비공식)", "사전영업", "", "발주처 기본계획 사전 검토",
     "• 조사설계 용역 기본계획 도서 확보 여부 확인\n• 핵심이슈 (미시추 사유, 지층/암반등급 조정여부 등) 확보 여부 확인\n• 사업지 연접 지질/지반조사 자료 확보 여부 확인",
     "토목사업팀", "지반조사업체, 외부지질전문가 3인이상", "발주처 기본계획 보고서", "", ""),

    ("", "관심프로젝트 선정", "사전영업", "", "입수된 발주정보 평가를 통한 관심 Project 선정",
     "• 해당 프로젝트별 발주처/경쟁동향 / 사업성평가 수행\n• 기술형입찰 입찰참여 Guide Line으로 해당 프로젝트 평가",
     "토목사업팀", "포스코스마트엔지니어링팀", "관심사업 List", "", ""),

    ("1", "발주처 지반조사 모니터링", "사전영업", "", "발주처 지반조사 수립업체 파악",
     "• 발주처 지반조사업체 확인 후 당사 Pool List와 비교\n• 발주처 지반조사 업체 확인팀 Contact Point 확보",
     "포스코스마트엔지니어링팀", "토목사업팀", "발주처 기본계획 수행 List + Contact Point List", "", ""),

    ("2", "외부 지질전문가 Pool 관리", "매년 1분기", "", "지반조사 품질향상을 위한 제3자 검증 수행 지반조사 전문가 Pool 관리",
     "• Academic 전문가와 현장 전문가를 포함한 Pool 구성\n• 전문가 Pool의 학력/기술사 등 경력사항(조사 15년) 확인",
     "포스코스마트엔지니어링팀", "-", "외부지질전문가 Pool List", "", ""),

    ("3", "지반조사 Pool 관리", "매년 1분기", "", "발주처 계획 대응 및 개략 지반조사 결과 사전 예측을 위한 업체 관리",
     "• 공공기관별 입찰공고 모니터링 통한 지반조사업체 입찰참여 유치(입찰 공사 여부 확인)\n• 업체 Pool List 갱신 여부 확인\n• 지반조사업체 역량 및 기술수준 평가표 평가 여부 확인 (Pool 갱신 및 신규업체 등록)\n• 조사업체 강점, 장비보유 확인\n• 조사업체 Network 확보 여부 확인",
     "외주구매팀", "포스코스마트엔지니어링팀", "조사업체 POOL/필요시 Update", "", ""),

    ("", "기본계획 지반조사 자료 및 R/O 정보 입수", "D-210", "D-180", "발주처 기본계획 및 지반조사 결과 사전 획득",
     "• 기본계획 지반조사결과 및 실시설계 Contact을 통한 자료입수",
     "지반조사 협력업체", "설계 협력업체", "발주처 기본계획 보고서 및 지반조사 보고서(비공식)", "", ""),

    ("", "기본계획 지반조사 R/O 검토 (설계, 공정, 기획)", "D-180", "D-165", "발주처 기본계획 지반조사 문제점 검토",
     "• 시추조사 위치, 수량 적정성 검토\n• 지질이상 구간 상세 조사 여부 및 적정성 검토 확인\n• 미시추 구간 지반 특성 적정성 검토확인 (측점 검토, 위치 및 물량검토)\n• 인근 사업 지반조사자료 비교 통해 당사 지반특성 분석 결과 Cross-check 확인",
     "포스코스마트엔지니어링팀(설계), 토목수행팀(공정), 토목사업지원팀(기획)", "토목견적팀", "기본계획 지반조사 검토서(설계, 공정, 기획)", "", ""),

    ("4", "기본계획 지반조사 R/O 종합검토 (원가)", "D-165", "D-150", "발주처 지반조사 문제점 R/O 분석",
     "• 발주처 기본계획 지반조사 R/O별 Cost effect 분석",
     "토목견적팀", "포스코스마트엔지니어링팀(설계), 토목수행팀(공정), 토목사업지원팀(기획)", "기본계획 지반조사 개략 검토서", "", ""),

    ("5", "지반조사 업체 선정", "D-150", "D-135", "지반조사업체 Pool 내 최적의 업체 선정",
     "• 지반조사업체 Pool/실적 검토 및 공종별 적합성 정합 및 견적 본부 연계 검토",
     "일반구매팀", "포스코스마트엔지니어링팀", "지반조사 입찰 참여 Pool", "", ""),

    ("", "지반조사 업체 계약", "D-135", "D-120", "지반조사업체 용역 계약",
     "• 일반구매시스템을 이용한 용역계약",
     "일반구매팀", "포스코스마트엔지니어링팀", "지반조사 용역 계약서", "", ""),

    ("6", "지반조사 항목, 수량 및 위치검토", "D-135", "D-115", "프로젝트 정합에 부합하는 지반조사 항목 사전조사 수행, 위치 선정",
     "• 현장답사 수행\n• 공종별 실시설계에 준하여 수량 및 위치 검토",
     "포스코스마트엔지니어링팀", "지반조사업체", "지반조사 계획서", "", ""),

    ("7", "지반조사계획 적정성 심사의뢰", "D-115", "D-105", "지반조사 항목, 수량 및 위치 적정성 검토",
     "• 지반조사 계획서 검토 및 검토",
     "외부지질전문가 3인 이상", "포스코스마트엔지니어링팀", "지반조사계획 의견서", "", ""),

    ("", "지반조사 인허가", "D-135", "D-100", "지반조사 수행을 위한 인허가 승인",
     "• 산지일시사용 신고(사업계획서, 지형도, 산지일시사용예정실측도, 복구계획서, 복지봉투)\n• 굴착행위 신고(굴착행위신고서, 위치 또는 평면도, 원상복구계획서, 토지사용 승낙서)\n• 하천점용 허가(위치도, 이해관계인 동의서, 수리계산서, 하천점용등, 사업자등록증)",
     "지반조사업체", "포스코스마트엔지니어링팀", "인허가 서류", "", ""),

    ("", "지반조사 항목, 수량 및 위치 보완", "D-105", "D-100", "외부지질전문가 의견을 반영한 지반조사 계획 확정",
     "• 외부지질전문가 지반조사 의견 검토\n• 입찰설계 지반조사 계획 보완",
     "포스코스마트엔지니어링팀", "외부지질전문가 3인 이상, 지반조사업체", "지반조사 용역 계약서", "", ""),

    ("8", "지반조사 수행", "D-100", "D-50", "과업구간 지반여건 상세 파악",
     "• 물리탐사 수행(전기 비저항, 탄성파탐사 등)\n• 시추조사 및 현장 시험 수행(공내재하시험, 실내전단시험 등)\n• 실내시험 수행(요동시험, 입도시험 등)",
     "지반조사업체", "포스코스마트엔지니어링팀", "지반조사 보고서, 시추 Core 및 샘플", "", ""),

    ("9", "지반조사 수렴 결과 분석", "D-50", "D-40", "지반조사 결과 오류 및 지반정수의 적합성 검토",
     "• 입찰설계 지반조사 보고서 검토\n• 시추조사 위치 및 심도의 적정성 검토\n• 실내시험 결과 등급 적정성 검토\n• 실시설계시 추가지반조사 필요구간 검토",
     "외부지질전문가 3인 이상", "포스코스마트엔지니어링팀, 지반조사업체", "지반조사 결과 분석 의견서", "", ""),

    ("", "입찰설계 지반조사 R/O 검토 (설계, 원가, 공정, 기타)", "D-40", "D-30", "입찰설계 지반조사 결과에 따른 직분야 R/O 및 Cost impact 반영 사항 검토",
     "• 입찰설계 지반조사 결과 분석을 통한 원가 및 공정 영향 검토\n• 미시추 구간 평가 및 공정 Risk 반영 사항 검토\n• 지반조건을 고려한 발파공법성 여부 검토",
     "토목견적팀(원가), 토목수행팀(공정), 토목사업지원팀(기획)", "포스코스마트엔지니어링팀", "입찰설계 지반조사 검토서(설계, 원가, 공정, 기타)", "", ""),

    ("10", "지반조사 결과 R/O 종합검토", "D-40", "D-30", "원가 및 공정 검토 결과에 따른 입찰설계 R/O 최종 확정",
     "• 입찰설계 지반조사 R/O별 Cost effect 분석\n• 유사사업 L/L 분석, R/P 검토를 통한 실시설계시 고려사항 검토",
     "포스코스마트엔지니어링팀", "토목견적팀(원가), 토목수행팀(공정), 토목사업지원팀(기획)", "입찰설계 R/O 검토서", "", ""),

    ("", "REM 수립", "D-30", "D-20", "REM 차트 작성(회사의 Core Risk, T/C 공정검토 결과, 경쟁사 동향, 발주 경향) 및 보고",
     "• Risk Profiling/Merge 방안 검토\n• 공정 착공성 검토 및 T/C 검토\n• 설계차별화 및 Idea Frame 검토",
     "포스코스마트엔지니어링팀", "토목견적팀, 토목사업지원팀", "REM 회의록", "", ""),

    ("11", "추가지반조사 계획 수립", "D-30", "D-15", "R/O 분석 결과에 기반한 추가지반조사 항목 도출",
     "• 지반조사 수렴 결과 분석보고서 검토\n• 지반조사 결과 R/O 종합검토 결과 검토",
     "포스코스마트엔지니어링팀", "토목견적팀(원가), 토목수행팀(공정), 토목사업지원팀(기획)", "추가지반조사 List", "", ""),

    ("", "3D GPR탐사 장비 사양(차량형) 검토", "D-15", "D-7", "적정 추가지반조사 규명을 위한 장비 사양 검토",
     "• 3D GPR탐사 장비 사양(차량형)\n- 3D Radar(노르웨이)_200~3000Hz, 2.0m~3.0m : 하상도 높음\n- MIRA(스웨덴)_400~1300Hz, 1.0m~2.0m : 정밀한 입체 형상\n- StreamC(이탈리아)_200~600Hz, 1.5m~3.0m : 다양한 심도, 공동 탐지\n- VGPR(호주)_100~4000Hz, 2.0m~3.0m : 고속주행 데이터 획득\n- Road 3D(한국)_200~3000Hz, 2.0m~3.0m : 하상도 높음, 유용특화",
     "현장 공무팀", "현장 공사팀", "-", "", ""),

    ("", "ECRM 수립", "D-15", "D-7", "입찰설계 지반조사 결과에 따른 직분야 R/O 및 Cost impact 반영 사항 검토",
     "• 입찰설계 지반조사 결과 분석을 통한 원가 및 공정 영향 검토\n• 미시추 구간 평가 및 공정 Risk 반영 사항 검토\n• 지반조건을 고려한 발파공법성 여부 검토",
     "토목견적팀(원가)", "포스코스마트엔지니어링팀, 토목수행팀, 토목사업지원팀", "입찰설계 지반조사 검토서(설계, 원가, 공정, 기타)", "", ""),

    ("", "최종 Cost 검토 / 입찰", "D-7", "D-0", "입찰",
     "• 최종 R/O를 반영한 투찰내역서 작성",
     "토목견적팀", "-", "투찰내역서", "", ""),

    ("12", "입찰설계 현장 인수인계", "D+0", "D+30", "지반조사 포함 전체 R/O 승계자료 인수인계 및 추가지반조사 계획 공유",
     "• 최종 R/O 결과 및 입찰도서 인출",
     "현장소장", "토목수행팀", "인수인계서", "", ""),

    ("13", "추가지반조사 수행", "D+30", "D+100", "미시추구간 지반조사 수행을 통한 지반 Risk Hedge",
     "• 추가 지반조사 계획 및 시점 확인\n- 탄성파탐사/전기비저항탐사 등 물리탐사 추가조사 (현장 및 위치 확인)\n- 추가시추조사 수행(물리탐사, 시추조사)\n- 토질 및 암석 실내시험 수행 및 결과분석 확인\n- 공종별 핵심시설(교량, 접속도로, 토공, #200m 통과점, 터널 발파/심벽진입) 확인",
     "지반조사업체", "현장 공사팀/공무팀, 포스코스마트엔지니어링팀, 토목수행팀", "추가 지반조사 보고서", "○", "○"),

    ("14", "추가지반조사 결과 분석", "D+100", "D+110", "추가지반조사 결과에 기반한 최종 R/O 확정",
     "• 외부전문가 검토 통한 지반특성 최종 Cross-check 확인\n• 입찰설계 지반조사 대비 지반특성 변경구간 제원도 확인",
     "외부지질전문가 3인 이상", "현장 공사팀/공무팀, 포스코스마트엔지니어링팀, 토목수행팀", "추가 지반조사결과 종합보고서", "○", "○"),

    ("", "실시설계 지반조사 R/O 검토 (설계, 원가, 공정, 기타)", "D+110", "D+120", "실시설계 지반조사 결과에 따른 직분야 R/O 및 Cost impact 반영 사항 검토",
     "• 실시설계 추가지반조사 결과 분석을 통한 원가 및 공정 영향 검토\n• 미시추 구간 평가 및 공정 Risk 반영 사항 검토\n• 지반조건을 고려한 발파공법성 검토",
     "포스코스마트엔지니어링팀", "현장 공사팀/품질팀/공무팀, 토목견적팀(원가), 토목수행팀(공정), 토목사업지원팀(기획)", "실시설계 지반조사 검토서(설계, 원가, 공정, 기타)", "", ""),

    ("15", "실시설계 추가지반조사 결과 검토", "D+110", "D+120", "추가지반조사 결과에 기반한 최종 R/O 확정",
     "• 추가지반조사 결과 검토",
     "포스코스마트엔지니어링팀", "현장 공사팀/품질팀/공무팀, 토목견적팀(원가), 토목수행팀(공정), 토목사업지원팀(기획)", "실시설계 최종 R/O", "", ""),

    ("16", "최종 R/O 이행방안 수립", "D+120", "D+150", "최종 지반조사 결과에 따른 R/O 이행방안 수립",
     "• Core Risk 및 Hedge 방안 수립\n• Opportunity에 따른 본공 상가 추가 개선 방안 수립",
     "현장소장", "토목수행팀", "최종 R/O 이행계획서", "", ""),

    ("17", "외부지질전문가 업무평가", "-", "-", "참여인력 전문성 평가 및 특이사항 기록",
     "• 외부지질 전문가 업무평가서 작성",
     "포스코스마트엔지니어링팀", "현장 공사팀/공무팀", "외부지질 전문가 업무평가서", "", ""),

    ("18", "조사업체 PM 업무평가", "-", "-", "조사업체 PM 전문성 평가 및 특이사항 기록",
     "• 지반조사 PM 업무평가서 작성",
     "포스코스마트엔지니어링팀", "현장 공사팀/공무팀", "PM업무평가서", "", ""),

    ("19", "지반조사업체 업무평가", "-", "-", "조사업체 업무평가 및 특이사항 기록",
     "• 협력업체 업무평가서",
     "현장소장", "현장 공사팀/공무팀", "협력업체 업무평가서", "", ""),

    ("", "설계변경 LL 리포트 작성", "-", "-", "설계변경 사항 기록",
     "• 설계변경 사유 기록관리\n• 예상 R/O대비 설계변경 사항 비교 검토",
     "현장소장", "현장 공무팀", "LL리포트", "", ""),

    ("", "설계변경 LL 리포트 관리", "-", "-", "전체 현장 설계변경 사항 기록",
     "• 현장별 설계변경 LL 리포트 취합 관리",
     "토목수행팀", "현장소장", "LL리포트 List", "", ""),

    ("", "지반조사업체 정산계약", "-", "-", "지반조사업체 정산",
     "• 조사도 수량 Check 후 구매시스템 변경계약 수행\n• 변경계약 후 정산금 집행",
     "일반구매팀", "현장 공무팀", "정산계약서", "", ""),

    ("20", "매뉴얼 Update", "-", "-", "지반조사 매뉴얼 Update",
     "• 매뉴얼 Update 및 개정사항 사내공지",
     "혁신실장", "현장 공무팀", "지반조사 매뉴얼", "", "")
]

def build_ground_survey_sheet(excel_path):
    if not os.path.exists(excel_path):
        print(f"Skipping non-existent: {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    
    # 1. '지반조사' 시트 (표준 16개 열 체계)
    sheet_name = "지반조사"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    
    # 폰트 & 스타일
    font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="맑은 고딕", size=9, bold=False, color="000000")
    font_data_bold = Font(name="맑은 고딕", size=9, bold=True, color="000000")
    font_link = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")
    
    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # 1행 헤더 (다른 시트들과 100% 동일한 16개 열 체계)
    headers = [
        "L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "일정 (D-Day)",
        "작업단위 (Level 4 Task/Activity)", "주관", "협업부서", "목적", "방법",
        "산출물(결과)", "표준서 (Standard) 요약", "표준서 파일 (HTML)",
        "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
        "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)"
    ]
    
    ws.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # 2행부터 데이터 입력
    for row_idx, row in enumerate(RAW_DATA, 2):
        no_str, proc, start, finish, why, how, who, collab, what, guide_chk, std_chk = row
        
        # L4 코드 생성
        l4_code = f"9000-0-{row_idx-1}"
        
        # 일정 포맷팅
        if start and finish:
            d_day = f"{start} ~ {finish}"
        elif start:
            d_day = start
        else:
            d_day = "-"
            
        # 표준서/지침 요약 텍스트
        std_summary = f"{why} | {what}"
        guide_summary = how.replace("\n", " ")[:120] + "..." if len(how) > 120 else how.replace("\n", " ")
        chk_summary = f"1) {proc} 적정성을 확인하였는가? 2) {what} 산출물을 확보하였는가?"
        
        row_values = [
            "9000",                # L2 코드
            "9000-0",              # L3 코드
            "사전준비_지반조사",    # L3 대공종명
            l4_code,               # L4 코드
            d_day,                 # 일정 (D-Day)
            proc,                  # 작업단위 (Level 4 Task/Activity)
            who,                   # 주관
            collab,                # 협업부서
            why,                   # 목적
            how,                   # 방법
            what,                  # 산출물(결과)
            std_summary,           # 표준서 요약
            "👉 [더블클릭] 표준서 열기 📄", # 표준서 파일
            guide_summary,         # 수행지침 요약
            "👉 [더블클릭] 수행지침 열기 📄", # 수행지침 파일
            chk_summary,           # 체크리스트 요약
            "👉 [더블클릭] 체크리스트 열기 📄" # 체크리스트 파일
        ]
        
        ws.row_dimensions[row_idx].height = 45 if '\n' in how else 28
        row_fill = fill_highlight if guide_chk == "○" else (fill_zebra if row_idx % 2 == 1 else None)
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.font = font_data_bold if col_idx in [1, 2, 3, 4, 6, 11] else (font_link if col_idx in [13, 15, 17] else font_data)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
                
            # 정렬
            if col_idx in [1, 2, 3, 4, 5, 13, 15, 17]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [6, 7, 8, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
    # 컬럼 너비 설정
    col_widths = {
        'A': 10, 'B': 10, 'C': 16, 'D': 12, 'E': 16,
        'F': 28, 'G': 24, 'H': 26, 'I': 32, 'J': 55,
        'K': 32, 'L': 35, 'M': 22, 'N': 40, 'O': 22,
        'P': 40, 'Q': 22
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 2. 추가로 '지반조사_기술형(원본서식)' 시트도 함께 생성
    raw_sheet_name = "지반조사(원본서식)"
    if raw_sheet_name in wb.sheetnames:
        del wb[raw_sheet_name]
    ws_raw = wb.create_sheet(title=raw_sheet_name)
    
    font_title = Font(name="맑은 고딕", size=13, bold=True, color="000000")
    fill_header_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    ws_raw['A1'] = "■ 매뉴얼_공통_사전준비_지반조사(기술형)"
    ws_raw['A1'].font = font_title
    ws_raw.row_dimensions[1].height = 28
    
    raw_headers = [
        ("No.", "A2", "A3"), ("Process", "B2", "B3"), ("시점(When)", "C2", "D2"),
        ("목적(Why)", "E2", "E3"), ("Activity\n방법(How)", "F2", "F3"),
        ("주관부서(Who)", "G2", "G3"), ("협업부서", "H2", "H3"),
        ("산출물(What)", "I2", "I3"), ("수행지침", "J2", "J3"), ("표준서", "K2", "K3")
    ]
    for title, start_c, end_c in raw_headers:
        if start_c == end_c:
            c = ws_raw[start_c]
            c.value = title
            c.font = font_header
            c.fill = fill_header_blue
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
        else:
            ws_raw.merge_cells(f"{start_c}:{end_c}")
            c = ws_raw[start_c]
            c.value = title
            c.font = font_header
            c.fill = fill_header_blue
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            start_col, start_row = openpyxl.utils.coordinate_to_tuple(start_c)
            end_col, end_row = openpyxl.utils.coordinate_to_tuple(end_c)
            for r in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    ws_raw.cell(r, col).border = thin_border
                    ws_raw.cell(r, col).fill = fill_header_blue
                    
    ws_raw['C3'] = "Start"
    ws_raw['C3'].font = font_header
    ws_raw['C3'].fill = fill_header_blue
    ws_raw['C3'].alignment = Alignment(horizontal="center", vertical="center")
    ws_raw['C3'].border = thin_border
    
    ws_raw['D3'] = "Finish"
    ws_raw['D3'].font = font_header
    ws_raw['D3'].fill = fill_header_blue
    ws_raw['D3'].alignment = Alignment(horizontal="center", vertical="center")
    ws_raw['D3'].border = thin_border
    
    ws_raw.merge_cells("A4:K4")
    ws_raw['A4'] = "공통_사전준비_지반조사(기술형)"
    ws_raw['A4'].font = font_header
    ws_raw['A4'].fill = fill_header
    ws_raw['A4'].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_raw.row_dimensions[4].height = 20
    for col in range(1, 12):
        ws_raw.cell(4, col).border = thin_border
        ws_raw.cell(4, col).fill = fill_header
        
    for r_idx, r_data in enumerate(RAW_DATA, 5):
        ws_raw.row_dimensions[r_idx].height = 42 if '\n' in r_data[5] else 26
        for c_idx, val in enumerate(r_data, 1):
            cell = ws_raw.cell(r_idx, c_idx, val)
            cell.font = font_data_bold if c_idx in [1, 2, 9, 10, 11] else font_data
            cell.border = thin_border
            if c_idx in [1, 3, 4, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [2, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
    raw_col_widths = {'A': 6, 'B': 26, 'C': 12, 'D': 12, 'E': 32, 'F': 55, 'G': 24, 'H': 26, 'I': 30, 'J': 10, 'K': 10}
    for col_letter, width in raw_col_widths.items():
        ws_raw.column_dimensions[col_letter].width = width

    # 시트 순서 조정: '지반조사'를 '공정매뉴얼' 바로 뒤(3번째 위치)로 배치
    wb._sheets.remove(ws)
    wb._sheets.insert(2, ws)
    
    wb._sheets.remove(ws_raw)
    wb._sheets.insert(3, ws_raw)

    wb.save(excel_path)
    print(f"[SUCCESS] Updated {excel_path} -> Added '지반조사' and '지반조사(원본서식)' sheets with 36 rows!")

for f in EXCEL_FILES:
    build_ground_survey_sheet(f)

print("=== ALL TARGET EXCEL FILES UPDATED SUCCESSFULLY! ===")
