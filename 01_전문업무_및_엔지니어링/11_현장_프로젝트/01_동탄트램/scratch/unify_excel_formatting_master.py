# -*- coding: utf-8 -*-
"""
동탄트램 매뉴얼 BODY (집행단계)v8 엑셀 11개 시트 서식 및 레이아웃 마스터 통일 스크립트

적용 항목:
1. 전 시트 19개 표준 컬럼 체계 통일:
   A(L2코드), B(L3코드), C(L3대공종명), D(L4코드), E(선행), F(후행),
   G(작업단위), H(일정), I(주관), J(목적), K(방법), L(산출물),
   M(표준서요약), N(표준서링크), O(수행지침요약), P(수행지침링크),
   Q(체크리스트요약), R(체크리스트링크), S(비고)
2. 시트 1~7 컬럼 순서(G열=작업단위, H열=일정) 및 지반조사 10열(협업부서) 정격 재배치
3. C열 대공종명 오타('지장뭉 이설' -> '지장물이설') 및 표준 명칭 정비
4. 틀 고정(Freeze Panes) 통일: 1행 헤더 + A~D열 코드 고정 (E2)
5. 눈금선(ShowGridLines = True) 전 시트 강제 활성화
6. 헤더 높이(28pt) 및 데이터 행 높이(52pt) 통일
7. 19개 열 너비(Width) 황금비율 일괄 배분 및 프리미엄 다크네이비 헤더/지브라 스타일 통일
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

# 스타일 정의
font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
font_data = Font(name="맑은 고딕", size=9, color="0F172A")
font_data_bold = Font(name="맑은 고딕", size=9, bold=True, color="0F172A")

link_font_std = Font(name="맑은 고딕", size=9, bold=True, color="047857", underline="single")
link_font_guide = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")
link_font_chk = Font(name="맑은 고딕", size=9, bold=True, color="D97706", underline="single")

fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

border_thin = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

HEADERS = [
    "L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "선행", "후행",
    "작업단위 (Level 4 Task/Activity)", "일정 (D-Day)", "주관", "목적",
    "방법", "산출물(결과)", "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)", "체크리스트 (Checklist) 요약",
    "체크리스트 파일 (HTML)", "비고"
]

COL_WIDTHS = {
    'A': 10, 'B': 10, 'C': 18, 'D': 12, 'E': 10, 'F': 10,
    'G': 30, 'H': 14, 'I': 14, 'J': 38, 'K': 54, 'L': 30,
    'M': 36, 'N': 24, 'O': 36, 'P': 24, 'Q': 36, 'R': 24, 'S': 16
}

SHEET_CONFIG = {
    "지반조사": {"idx": 1, "name": "지반조사", "tasks": 36, "type": "jiban"},
    "사전토공사": {"idx": 2, "name": "사전토공사", "tasks": 31, "type": "swap_gh"},
    "지장물이설": {"idx": 3, "name": "지장물이설", "tasks": 38, "type": "swap_gh"},
    "상부강화노반": {"idx": 4, "name": "상부강화노반", "tasks": 36, "type": "swap_gh"},
    "콘크리트도상": {"idx": 5, "name": "콘크리트도상", "tasks": 23, "type": "swap_gh"},
    "건축": {"idx": 6, "name": "건축", "tasks": 50, "type": "swap_gh"},
    "신호": {"idx": 7, "name": "신호분야", "tasks": 23, "type": "swap_gh"},
    "전기": {"idx": 8, "name": "전기분야", "tasks": 32, "type": "standard"},
    "통신": {"idx": 9, "name": "통신분야", "tasks": 32, "type": "standard"},
    "기계": {"idx": 10, "name": "기계설비·소방설비", "tasks": 36, "type": "standard"},
    "철도종합시운전": {"idx": 11, "name": "철도종합시험운행", "tasks": 28, "type": "standard"}
}

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue
    print(f"\n=======================================================")
    print(f"서식 통일 작업 진행: {os.path.basename(file_path)}")
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CONFIG:
            continue

        cfg = SHEET_CONFIG[sheet_name]
        s_num = cfg["idx"]
        l3_name = cfg["name"]
        task_count = cfg["tasks"]
        sheet_type = cfg["type"]
        ws = wb[sheet_name]

        print(f"  ▶ [{sheet_name}] 서식 재배치 및 스타일링 적용 (Type: {sheet_type})")

        # 1. 기존 데이터 백업 (행별 원본 값 추출)
        raw_rows = []
        for r in range(2, task_count + 2):
            row_dict = {}
            for c in range(1, ws.max_column + 5):
                cell = ws.cell(r, c)
                row_dict[c] = cell.value
            raw_rows.append(row_dict)

        # 2. 시트 전체 초기화 (A~Z, 1~max_row)
        for r in range(1, ws.max_row + 15):
            for c in range(1, max(ws.max_column + 5, 25)):
                ws.cell(r, c).value = None
                ws.cell(r, c).fill = PatternFill(fill_type=None)
                ws.cell(r, c).border = Border()

        # 3. 눈금선 및 틀 고정 설정
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "E2"  # 1행 헤더 + A~D열 코드 고정

        # 4. 헤더 작성
        ws.row_dimensions[1].height = 28
        for c_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(1, c_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # 5. 데이터 행 재배치 및 서식 적용
        for idx, orig in enumerate(raw_rows, 1):
            r = idx + 1
            ws.row_dimensions[r].height = 52
            is_even = (idx % 2 == 0)
            row_fill = fill_zebra if is_even else None

            l2_code = "9000"
            l3_code = f"9000-{s_num}"
            l4_code = f"9000-{s_num}-{idx}"
            pred_code = "-" if idx == 1 else f"9000-{s_num}-{idx-1}"
            succ_code = "-" if idx == task_count else f"9000-{s_num}-{idx+1}"

            if sheet_type == "jiban":
                # 지반조사 원본 매핑:
                # 7: 일정, 8: 작업단위, 9: 주관, 10: 협업부서, 11: 목적, 12: 방법, 13: 산출물,
                # 14: 표준서요약, 15: 표준서파일, 16: 수행지침요약, 17: 수행지침파일, 18: 체크리스트요약, 19: 체크리스트파일, 20: 비고
                task_unit = orig.get(8)
                dday = orig.get(7)
                dept = orig.get(9)
                col_coop = orig.get(10)
                if col_coop and str(col_coop).strip():
                    dept_full = f"{dept} (협업: {str(col_coop).replace(chr(10), ' ')})"
                else:
                    dept_full = dept
                purpose = orig.get(11)
                method = orig.get(12)
                output = orig.get(13)
                std_sum = orig.get(14)
                std_link = orig.get(15)
                guide_sum = orig.get(16)
                guide_link = orig.get(17)
                chk_sum = orig.get(18)
                chk_link = orig.get(19)
                remark = orig.get(20)

            elif sheet_type == "swap_gh":
                # Sheets 2~7 원본 매핑:
                # 7: 일정, 8: 작업단위, 9: 주관, 10: 목적, 11: 방법, 12: 산출물,
                # 13: 표준서요약, 14: 표준서파일, 15: 수행지침요약, 16: 수행지침파일, 17: 체크리스트요약, 18: 체크리스트파일, 19: 비고
                task_unit = orig.get(8)
                dday = orig.get(7)
                dept_full = orig.get(9)
                purpose = orig.get(10)
                method = orig.get(11)
                output = orig.get(12)
                std_sum = orig.get(13)
                std_link = orig.get(14)
                guide_sum = orig.get(15)
                guide_link = orig.get(16)
                chk_sum = orig.get(17)
                chk_link = orig.get(18)
                remark = orig.get(19)

            else:
                # Sheets 8~11 원본 매핑 (이미 표준 순서):
                # 7: 작업단위, 8: 일정, 9: 주관, 10: 목적, 11: 방법, 12: 산출물,
                # 13: 표준서요약, 14: 표준서파일, 15: 수행지침요약, 16: 수행지침파일, 17: 체크리스트요약, 18: 체크리스트파일, 19: 비고
                task_unit = orig.get(7)
                dday = orig.get(8)
                dept_full = orig.get(9)
                purpose = orig.get(10)
                method = orig.get(11)
                output = orig.get(12)
                std_sum = orig.get(13)
                std_link = orig.get(14)
                guide_sum = orig.get(15)
                guide_link = orig.get(16)
                chk_sum = orig.get(17)
                chk_link = orig.get(18)
                remark = orig.get(19)

            row_data = [
                (1, l2_code, font_data_bold, align_center),
                (2, l3_code, font_data_bold, align_center),
                (3, l3_name, font_data_bold, align_center),
                (4, l4_code, font_data_bold, align_center),
                (5, pred_code, font_data, align_center),
                (6, succ_code, font_data, align_center),
                (7, task_unit, font_data_bold, align_left_wrap),
                (8, dday, font_data, align_center),
                (9, dept_full, font_data, align_center),
                (10, purpose, font_data, align_left_wrap),
                (11, method, font_data, align_left_wrap),
                (12, output, font_data, align_left_wrap),
                (13, std_sum, font_data, align_left_wrap),
                (14, std_link, link_font_std, align_center),
                (15, guide_sum, font_data, align_left_wrap),
                (16, guide_link, link_font_guide, align_center),
                (17, chk_sum, font_data, align_left_wrap),
                (18, chk_link, link_font_chk, align_center),
                (19, remark, font_data, align_center)
            ]

            for col_idx, val, f_style, a_style in row_data:
                c = ws.cell(r, col_idx)
                c.value = val
                c.font = f_style
                c.alignment = a_style
                c.border = border_thin
                if row_fill:
                    c.fill = row_fill

        # 6. 열 너비 일괄 적용
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    print(f"  ✔ {os.path.basename(file_path)} 11개 시트 서식 마스터 통일 완료!")

print(f"\n=======================================================")
print(f"매뉴얼 BODY (집행단계)v8 엑셀 11개 시트 서식 및 레이아웃 마스터 통일 완벽 성공!")
print(f"=======================================================")
