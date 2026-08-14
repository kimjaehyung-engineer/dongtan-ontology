# -*- coding: utf-8 -*-
"""
동탄도시철도(트램) 매뉴얼 BODY (집행단계)v8
[종합통계대시보드] 최상위 시트 생성 및 비즈니스 프리미엄 서식 구축 스크립트
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

# -------------------------------------------------------------
# 스타일 팔레트 (Executive Business Slate & Navy)
# -------------------------------------------------------------
font_title = Font(name="맑은 고딕", size=15, bold=True, color="FFFFFF")
font_subtitle = Font(name="맑은 고딕", size=9, color="94A3B8")
font_sec_hdr = Font(name="맑은 고딕", size=11, bold=True, color="0F172A")

font_tbl_hdr = Font(name="맑은 고딕", size=9.5, bold=True, color="FFFFFF")
font_tbl_data = Font(name="맑은 고딕", size=9, color="1E293B")
font_tbl_bold = Font(name="맑은 고딕", size=9, bold=True, color="0F172A")
font_tbl_total = Font(name="맑은 고딕", size=9.5, bold=True, color="0F172A")

font_kpi_label = Font(name="맑은 고딕", size=9, bold=True, color="475569")
font_kpi_val_blue = Font(name="맑은 고딕", size=16, bold=True, color="1E40AF")
font_kpi_val_slate = Font(name="맑은 고딕", size=16, bold=True, color="0F172A")
font_kpi_val_green = Font(name="맑은 고딕", size=16, bold=True, color="047857")
font_kpi_val_sky = Font(name="맑은 고딕", size=16, bold=True, color="0284C7")
font_kpi_val_amber = Font(name="맑은 고딕", size=16, bold=True, color="D97706")
font_kpi_sub = Font(name="맑은 고딕", size=8, color="64748B")

link_font_nav = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")

# 배경색 (Fills)
fill_banner = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
fill_tbl_hdr = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_tbl_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_tbl_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

fill_kpi_blue = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_kpi_slate = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_kpi_green = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
fill_kpi_sky = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
fill_kpi_amber = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")

# 테두리 (Borders)
border_thin = Border(
    left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_card_blue = Border(
    left=Side(style="medium", color="3B82F6"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_card_slate = Border(
    left=Side(style="medium", color="475569"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_card_green = Border(
    left=Side(style="medium", color="10B981"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_card_sky = Border(
    left=Side(style="medium", color="0EA5E9"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_card_amber = Border(
    left=Side(style="medium", color="F59E0B"), right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1")
)
border_total = Border(
    left=Side(style="thin", color="94A3B8"), right=Side(style="thin", color="94A3B8"),
    top=Side(style="medium", color="0F172A"), bottom=Side(style="double", color="0F172A")
)

# 정렬 (Alignments)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_title = Alignment(horizontal="left", vertical="center", indent=1)

# 11대 공종 기초 데이터
DISCIPLINES = [
    (1, "지반조사", "9000-1", "지반조사", 36, "토목사업팀"),
    (2, "사전토공사", "9000-2", "사전토공사", 31, "현장 공사팀"),
    (3, "지장물이설", "9000-3", "지장물이설", 38, "토목국내견적팀"),
    (4, "상부강화노반", "9000-4", "상부강화노반", 36, "현장 공사팀"),
    (5, "콘크리트도상", "9000-5", "콘크리트도상", 23, "현장소장"),
    (6, "건축", "9000-6", "건축", 50, "현장 공무팀"),
    (7, "신호", "9000-7", "신호분야", 23, "현장 공무팀"),
    (8, "전기", "9000-8", "전기분야", 32, "현장 공무팀"),
    (9, "통신", "9000-9", "통신분야", 32, "현장 공무팀"),
    (10, "기계", "9000-10", "기계설비·소방설비", 36, "기계/소방팀"),
    (11, "철도종합시운전", "9000-11", "철도종합시험운행", 28, "개통운영팀")
]

TOTAL_TASKS = sum(d[4] for d in DISCIPLINES)  # 375개

# 4대 D-Day 타임라인 단계별 데이터
TIMELINE_PHASES = [
    ("Stage 1 : 착수 전 사전준비 및 인허가", "D-300 ~ D-0", "설계적정성 검토, 인허가 취득, 인터페이스 Big Room 회의, 자재공급원 승인", 132, 35.2),
    ("Stage 2 : 현장 본시공 및 구조물 구축", "D+1 ~ D+100", "토공 굴착, 강화노반 슬래브, 레일 부설, 역사 골조 및 마감, 기계/전기 배관", 148, 39.5),
    ("Stage 3 : 시스템 연동 및 공종별 시험", "D+101 ~ D+180", "전철전력 가압, 통신/신호 단위시험, 소방 화재연동 실부하, 사전점검 수검", 67, 17.8),
    ("Stage 4 : 종합시운전 및 영업 개통", "D+181 ~ D+240", "속도별 시설물검증시험, 30일 영업 다이아 시운전, 운영사 인수인계 및 개통", 28, 7.5)
]

# 4대 주관 부문별 R&R 데이터
RR_SECTORS = [
    ("토목 / 궤도 / 지반 부문", "지반조사, 사전토공사, 지장물이설, 상부강화노반, 콘크리트도상", 164, 43.7, "현장 공사팀 / 토목사업팀"),
    ("건축 / 기계 / 소방 부문", "정거장 및 차량기지 건축 마감, 공조환기, 제연, 소방, 급배수", 86, 22.9, "건축공무팀 / 기계소방팀"),
    ("시스템 부문 (전기·신호·통신)", "수전/변전 급전, LTE-R 무선망, T-ATP/ATO 신호연동, SCADA", 87, 23.2, "시스템 공무/공사팀"),
    ("개통운영 / 안전품질 부문", "사전점검, 속도별 시설물검증, 영업시운전, TS 안전관리체계, 인수인계", 38, 10.1, "개통운영팀 / 안전품질팀")
]

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue

    print(f"\n=======================================================")
    print(f"통계 대시보드 시트 생성 중: {file_path}")
    wb = openpyxl.load_workbook(file_path)

    # 기존 대시보드 시트가 있으면 제거 후 재생성 (최신화)
    if "종합통계대시보드" in wb.sheetnames:
        wb.remove(wb["종합통계대시보드"])
    elif "대시보드" in wb.sheetnames:
        wb.remove(wb["대시보드"])

    # 최상위(Index 0)에 시트 생성
    ws = wb.create_sheet(title="종합통계대시보드", index=0)
    ws.views.sheetView[0].showGridLines = True

    # ---------------------------------------------------------
    # 1. 헤더 타이틀 배너 (Rows 1~2)
    # ---------------------------------------------------------
    ws.merge_cells("A1:J1")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    ws["A1"].value = "동탄도시철도(트램) 건설공사 집행단계 엔지니어링 매뉴얼 종합 통계 대시보드"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_banner
    ws["A1"].alignment = align_title

    ws["A2"].value = "■ 기준: 11개 공종 375개 액티비티 | 총 1,125개 매뉴얼(표준서·수행지침·체크리스트) 100% 디지털 연계 완료 | Ver 8.0"
    ws["A2"].font = font_subtitle
    ws["A2"].fill = fill_banner
    ws["A2"].alignment = align_title

    ws.row_dimensions[3].height = 10  # 빈 간격

    # ---------------------------------------------------------
    # 2. 마스터 5대 핵심 KPI 요약 카드 (Rows 4~6)
    # ---------------------------------------------------------
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 16

    kpi_cards = [
        ("A", "B", "총 관리 공종 수", "11개 공종", "(토목/궤도/시스템/시운전)", fill_kpi_blue, font_kpi_val_blue, border_card_blue),
        ("C", "D", "총 액티비티 (WBS)", "375개 작업", "(100% 코드화 및 체인연동)", fill_kpi_slate, font_kpi_val_slate, border_card_slate),
        ("E", "F", "표준서 (Standard)", "375 / 375", "(100% 기술기준 연계 완비)", fill_kpi_green, font_kpi_val_green, border_card_green),
        ("G", "H", "수행지침 (Guideline)", "375 / 375", "(100% 2D 도식·시뮬레이션)", fill_kpi_sky, font_kpi_val_sky, border_card_sky),
        ("I", "J", "체크리스트 (Checklist)", "375 / 375", "(100% 현장 검측 필증 연동)", fill_kpi_amber, font_kpi_val_amber, border_card_amber)
    ]

    for col1, col2, label, val, sub, fill_k, font_v, border_k in kpi_cards:
        c_range_1 = f"{col1}4:{col2}4"
        c_range_2 = f"{col1}5:{col2}5"
        c_range_3 = f"{col1}6:{col2}6"

        ws.merge_cells(c_range_1)
        ws.merge_cells(c_range_2)
        ws.merge_cells(c_range_3)

        c1 = ws[f"{col1}4"]
        c1.value = label; c1.font = font_kpi_label; c1.alignment = align_center

        c2 = ws[f"{col1}5"]
        c2.value = val; c2.font = font_v; c2.alignment = align_center

        c3 = ws[f"{col1}6"]
        c3.value = sub; c3.font = font_kpi_sub; c3.alignment = align_center

        for r_idx in range(4, 7):
            for col_l in [col1, col2]:
                cell = ws[f"{col_l}{r_idx}"]
                cell.fill = fill_k
                cell.border = border_k

    ws.row_dimensions[7].height = 14  # 간격

    # ---------------------------------------------------------
    # 3. [섹션 1] 11대 공종별 마스터 현황 및 시트 네비게이션 대장 (Rows 8~21)
    # ---------------------------------------------------------
    ws.merge_cells("A8:J8")
    ws["A8"].value = "1. 11대 공종별 마스터 현황 및 시트 네비게이션 대장"
    ws["A8"].font = font_sec_hdr
    ws.row_dimensions[8].height = 22

    headers_sec1 = ["순번", "공종명 (Discipline)", "L3 코드", "액티비티 수", "표준서", "수행지침", "체크리스트", "매뉴얼 총계", "비중 (%)", "시트 바로가기 (점프)"]
    ws.row_dimensions[9].height = 25

    for c_idx, h_text in enumerate(headers_sec1, 1):
        cell = ws.cell(9, c_idx, value=h_text)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_thin

    for idx, (s_idx, s_name, l3_code, l3_desc, tasks, dept) in enumerate(DISCIPLINES, 1):
        r = idx + 9
        ws.row_dimensions[r].height = 22
        is_even = (idx % 2 == 0)
        row_fill = fill_tbl_zebra if is_even else None
        share_pct = round((tasks / TOTAL_TASKS) * 100, 1)

        row_vals = [
            (1, idx, font_tbl_bold, align_center),
            (2, l3_desc if l3_desc else s_name, font_tbl_bold, align_left),
            (3, l3_code, font_tbl_bold, align_center),
            (4, tasks, font_tbl_bold, align_center),
            (5, tasks, font_tbl_data, align_center),
            (6, tasks, font_tbl_data, align_center),
            (7, tasks, font_tbl_data, align_center),
            (8, tasks * 3, font_tbl_bold, align_center),
            (9, f"{share_pct}%", font_tbl_data, align_center),
            (10, f'=HYPERLINK("#\'{s_name}\'!A1", "👉 [{s_name}] 시트 이동 ➔")', link_font_nav, align_center)
        ]

        for c_idx, val, f_style, a_style in row_vals:
            c = ws.cell(r, c_idx, value=val)
            c.font = f_style
            c.alignment = a_style
            c.border = border_thin
            if row_fill:
                c.fill = row_fill

    # 합계 행 (Row 21)
    r_total = len(DISCIPLINES) + 10
    ws.row_dimensions[r_total].height = 25
    total_vals = [
        (1, "합계", font_tbl_total, align_center),
        (2, "11개 공종 전체", font_tbl_total, align_center),
        (3, "-", font_tbl_total, align_center),
        (4, "=SUM(D10:D20)", font_tbl_total, align_center),
        (5, "=SUM(E10:E20)", font_tbl_total, align_center),
        (6, "=SUM(F10:F20)", font_tbl_total, align_center),
        (7, "=SUM(G10:G20)", font_tbl_total, align_center),
        (8, "=SUM(H10:H20)", font_tbl_total, align_center),
        (9, "100.0%", font_tbl_total, align_center),
        (10, "총 1,125개 매뉴얼 도서", font_tbl_total, align_center)
    ]

    for c_idx, val, f_style, a_style in total_vals:
        c = ws.cell(r_total, c_idx, value=val)
        c.font = f_style
        c.fill = fill_tbl_total
        c.alignment = a_style
        c.border = border_total

    ws.row_dimensions[r_total + 1].height = 14  # 간격

    # ---------------------------------------------------------
    # 4. [섹션 2] D-Day 일정 추진 단계별 (Timeline) 업무 분포 분석 (Rows 23~28)
    # ---------------------------------------------------------
    r_sec2_hdr = r_total + 2
    ws.merge_cells(f"A{r_sec2_hdr}:J{r_sec2_hdr}")
    ws[f"A{r_sec2_hdr}"].value = "2. D-Day 일정 추진 단계별(Timeline) 업무 분포 및 관리 목표"
    ws[f"A{r_sec2_hdr}"].font = font_sec_hdr
    ws.row_dimensions[r_sec2_hdr].height = 22

    headers_sec2 = ["단계 구분", "기간 (D-Day)", "주요 엔지니어링 관리 목표", "작업 수", "비중 (%)", "표준서", "지침서", "체크리스트", "주관 부서", "중점 리스크"]
    r_sec2_tbl_hdr = r_sec2_hdr + 1
    ws.row_dimensions[r_sec2_tbl_hdr].height = 25

    for c_idx, h_text in enumerate(headers_sec2, 1):
        cell = ws.cell(r_sec2_tbl_hdr, c_idx, value=h_text)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_thin

    for idx, (p_title, p_dday, p_desc, p_tasks, p_pct) in enumerate(TIMELINE_PHASES, 1):
        r = r_sec2_tbl_hdr + idx
        ws.row_dimensions[r].height = 24
        is_even = (idx % 2 == 0)
        row_fill = fill_tbl_zebra if is_even else None

        row_vals = [
            (1, p_title, font_tbl_bold, align_left),
            (2, p_dday, font_tbl_bold, align_center),
            (3, p_desc, font_tbl_data, align_left),
            (4, p_tasks, font_tbl_bold, align_center),
            (5, f"{p_pct}%", font_tbl_data, align_center),
            (6, p_tasks, font_tbl_data, align_center),
            (7, p_tasks, font_tbl_data, align_center),
            (8, p_tasks, font_tbl_data, align_center),
            (9, "공무/공사/시스템", font_tbl_data, align_center),
            (10, "공기지연·인허가", font_tbl_data, align_center)
        ]

        for c_idx, val, f_style, a_style in row_vals:
            c = ws.cell(r, c_idx, value=val)
            c.font = f_style
            c.alignment = a_style
            c.border = border_thin
            if row_fill:
                c.fill = row_fill

    r_sec2_end = r_sec2_tbl_hdr + len(TIMELINE_PHASES)
    ws.row_dimensions[r_sec2_end + 1].height = 14  # 간격

    # ---------------------------------------------------------
    # 5. [섹션 3] 4대 주관 부문별 업무 분장(R&R) 분포 현황 (Rows 29~34)
    # ---------------------------------------------------------
    r_sec3_hdr = r_sec2_end + 2
    ws.merge_cells(f"A{r_sec3_hdr}:J{r_sec3_hdr}")
    ws[f"A{r_sec3_hdr}"].value = "3. 4대 주관 부문별 업무 분장(R&R) 및 조직 협업 매트릭스"
    ws[f"A{r_sec3_hdr}"].font = font_sec_hdr
    ws.row_dimensions[r_sec3_hdr].height = 22

    headers_sec3 = ["주관 부문", "포함 공종 범위", "작업 수", "비중 (%)", "핵심 담당 조직", "주요 R&R 역할", "협업 인터페이스", "핵심 산출물", "모니터링 주기", "비고"]
    r_sec3_tbl_hdr = r_sec3_hdr + 1
    ws.row_dimensions[r_sec3_tbl_hdr].height = 25

    for c_idx, h_text in enumerate(headers_sec3, 1):
        cell = ws.cell(r_sec3_tbl_hdr, c_idx, value=h_text)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = align_center
        cell.border = border_thin

    for idx, (rr_title, rr_scope, rr_tasks, rr_pct, rr_org) in enumerate(RR_SECTORS, 1):
        r = r_sec3_tbl_hdr + idx
        ws.row_dimensions[r].height = 24
        is_even = (idx % 2 == 0)
        row_fill = fill_tbl_zebra if is_even else None

        row_vals = [
            (1, rr_title, font_tbl_bold, align_left),
            (2, rr_scope, font_tbl_data, align_left),
            (3, rr_tasks, font_tbl_bold, align_center),
            (4, f"{rr_pct}%", font_tbl_data, align_center),
            (5, rr_org, font_tbl_bold, align_center),
            (6, "시공계획 및 품질검측", font_tbl_data, align_center),
            (7, "주간 인터페이스 회의", font_tbl_data, align_center),
            (8, "검측서·시험성적서", font_tbl_data, align_center),
            (9, "일일 / 주간", font_tbl_data, align_center),
            (10, "무결점 시공 달성", font_tbl_data, align_center)
        ]

        for c_idx, val, f_style, a_style in row_vals:
            c = ws.cell(r, c_idx, value=val)
            c.font = f_style
            c.alignment = a_style
            c.border = border_thin
            if row_fill:
                c.fill = row_fill

    # ---------------------------------------------------------
    # 6. 열 너비(Width) 최적화
    # ---------------------------------------------------------
    col_widths_dash = {
        'A': 8,   # 순번 / 단계
        'B': 24,  # 공종명 / 기간
        'C': 12,  # L3 코드 / 주요목표
        'D': 14,  # 액티비티 수
        'E': 12,  # 표준서 / 비중
        'F': 12,  # 지침서
        'G': 12,  # 체크리스트
        'H': 14,  # 총계 / 산출물
        'I': 14,  # 비중 / 주기
        'J': 24   # 시트 바로가기 / 비고
    }

    for col_l, w in col_widths_dash.items():
        ws.column_dimensions[col_l].width = w

    wb.save(file_path)
    print(f"  ✔ {os.path.basename(file_path)} [종합통계대시보드] 시트 완벽 생성 및 1위치 배치 완료!")

print(f"\n=======================================================")
print(f"동탄트램 매뉴얼 BODY (집행단계)v8 [종합통계대시보드] 구축 100% 완료!")
print(f"=======================================================")
