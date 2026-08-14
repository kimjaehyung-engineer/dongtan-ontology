import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

# ============================================
# STEP 1: Extract construction sections from Excel
# ============================================
path1 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 1공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'
path2 = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\교차로 정보\03_기술제안 2공구 본선구간 공기산출 근거_(주)천우씨엠.xlsx'

sections = []

for tool_label, path in [("1공구", path1), ("2공구", path2)]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    current_section = None

    for row in range(12, ws.max_row + 1):
        col_l = ws[f'L{row}'].value
        col_b = ws[f'B{row}'].value
        col_c = ws[f'C{row}'].value
        col_g = ws[f'G{row}'].value
        col_h = ws[f'H{row}'].value
        col_i = ws[f'I{row}'].value
        col_k = ws[f'K{row}'].value
        col_m = ws[f'M{row}'].value
        col_n = ws[f'N{row}'].value

        if col_l and str(col_l).strip():
            if current_section:
                sections.append(current_section)
            section_name = str(col_l).strip().replace('\n', ' ')
            current_section = {
                'tool': tool_label,
                'section': section_name,
                'startSta': float(col_g) if col_g else 0.0,
                'endSta': None,
                'length': round(float(col_m), 1) if col_m else None,
                'duration': round(float(col_n), 1) if col_n else None,
                'segments': [],
                'landmark': str(col_k).strip() if col_k else None
            }

        if current_section and col_b is not None and col_c:
            seg_info = {
                'no': str(col_b),
                'name': str(col_c).strip(),
                'startSta': float(col_g) if col_g else None,
                'endSta': float(col_h) if col_h else None,
            }
            current_section['segments'].append(seg_info)
            if col_h:
                current_section['endSta'] = float(col_h)

    if current_section:
        sections.append(current_section)
    wb.close()

# Fix 2공구 1구간 startSta (was None/0)
for s in sections:
    if s['tool'] == '2공구' and s['startSta'] == 0.0 and s['segments']:
        first_sta = s['segments'][0].get('startSta')
        if first_sta is not None and first_sta > 0:
            s['startSta'] = first_sta
        else:
            s['startSta'] = 0.0

print(f"Extracted {len(sections)} sections")

# ============================================
# STEP 2: Build STA-to-stnPair lookup from intersectionData
# ============================================
html_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\08.메뉴얼 및 평면도\동탄트램 노선평면도\동탄트램_노선평면도V1.html'

with open(html_path, 'r', encoding='utf-8') as f:
    html = f.read()

m = re.search(r'const intersectionData = (\[.*?\]);', html, re.DOTALL)
idata = json.loads(m.group(1))

# Build sorted STA reference points per tool
# Each reference point: (sta, stnA, stnB, ratio)
def build_sta_refs(tool):
    items = [i for i in idata if i['tool'] == tool]
    refs = []
    for item in items:
        refs.append((item['startSta'], item['startStnA'], item['startStnB'], item['startRatio']))
        refs.append((item['endSta'], item['endStnA'], item['endStnB'], item['endRatio']))
    refs.sort(key=lambda x: x[0])
    # Deduplicate
    seen = set()
    unique = []
    for r in refs:
        key = r[0]
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique

refs_1 = build_sta_refs("1공구")
refs_2 = build_sta_refs("2공구")

def interpolate_sta(sta, refs):
    """Find the stnA/stnB/ratio for a given STA by interpolating between known reference points."""
    if not refs:
        return None
    
    # Clamp to range
    if sta <= refs[0][0]:
        return {'stnA': refs[0][1], 'stnB': refs[0][2], 'ratio': refs[0][3]}
    if sta >= refs[-1][0]:
        return {'stnA': refs[-1][1], 'stnB': refs[-1][2], 'ratio': refs[-1][3]}
    
    # Find the two bracketing references
    for i in range(len(refs) - 1):
        s1, a1, b1, r1 = refs[i]
        s2, a2, b2, r2 = refs[i + 1]
        if s1 <= sta <= s2:
            if a1 == a2 and b1 == b2:
                # Same segment pair, simple interpolation
                if s2 == s1:
                    t = 0
                else:
                    t = (sta - s1) / (s2 - s1)
                ratio = r1 + t * (r2 - r1)
                return {'stnA': a1, 'stnB': b1, 'ratio': round(ratio, 4)}
            else:
                # Different segment pairs - pick closest
                if (sta - s1) <= (s2 - sta):
                    return {'stnA': a1, 'stnB': b1, 'ratio': r1}
                else:
                    return {'stnA': a2, 'stnB': b2, 'ratio': r2}
    
    return {'stnA': refs[-1][1], 'stnB': refs[-1][2], 'ratio': refs[-1][3]}

# ============================================
# STEP 3: Compute stnA/stnB/ratio for each section start/end
# ============================================
cs_data = []
section_no = 0

for s in sections:
    section_no += 1
    refs = refs_1 if s['tool'] == '1공구' else refs_2
    
    start_pair = interpolate_sta(s['startSta'], refs)
    end_pair = interpolate_sta(s['endSta'], refs)
    
    # Midpoint STA for center position
    mid_sta = (s['startSta'] + s['endSta']) / 2
    mid_pair = interpolate_sta(mid_sta, refs)
    
    entry = {
        'no': section_no,
        'tool': s['tool'],
        'section': s['section'],
        'startSta': round(s['startSta'], 1),
        'endSta': round(s['endSta'], 1),
        'length': s['length'],
        'duration': s['duration'],
        'segCount': len(s['segments']),
        'startStnA': start_pair['stnA'],
        'startStnB': start_pair['stnB'],
        'startRatio': start_pair['ratio'],
        'endStnA': end_pair['stnA'],
        'endStnB': end_pair['stnB'],
        'endRatio': end_pair['ratio'],
        'midStnA': mid_pair['stnA'],
        'midStnB': mid_pair['stnB'],
        'midRatio': mid_pair['ratio'],
    }
    cs_data.append(entry)
    print(f"  [{s['tool']}] {s['section']}: STA {entry['startSta']}~{entry['endSta']} -> start({entry['startStnA']}/{entry['startStnB']}/{entry['startRatio']}) end({entry['endStnA']}/{entry['endStnB']}/{entry['endRatio']})")

# Save JS data
js_lines = json.dumps(cs_data, ensure_ascii=False, indent=2)

out_path = r'c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\11_현장_프로젝트\01_동탄트램\scratch\construction_sections_js.json'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(js_lines)

print(f"\nSaved {len(cs_data)} entries to construction_sections_js.json")
