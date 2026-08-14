import openpyxl
import json
import difflib

# Load target excel
file_path = 'c:/Users/sskjh/antigravity/01_전문업무_및_엔지니어링/동탄트램/06.체크리스트/07.최종/차량기지.xlsm'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['체크리스트(동탄트램전용)']

# Load v3 depot matches JSON
with open('scratch/v3_depot_matches.json', 'r', encoding='utf-8') as f:
    v3_matches = json.load(f)

# Helper function to find closest match in json by risk text
def find_closest_match(risk_text):
    if not risk_text:
        return None
    best_ratio = 0
    best_match = None
    for item in v3_matches:
        q_text = item.get('q') or ''
        # Compare similarity
        ratio = difflib.SequenceMatcher(None, risk_text, q_text).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = item
    if best_ratio > 0.7:  # high similarity threshold
        return best_match, best_ratio
    return None, 0

print("Matching Rows 5 to 38 with v3_depot_matches.json:")
matched_results = {}
for r in range(5, 39):
    code = sheet.cell(row=r, column=1).value
    lv2 = sheet.cell(row=r, column=5).value
    risk = sheet.cell(row=r, column=6).value
    
    match_item, ratio = find_closest_match(risk)
    if match_item:
        print(f"Row {r} ({code}) | Lv2: {lv2} -> Matched (ratio {ratio:.2f})")
        print(f"  Risk: {risk[:60]}...")
        print(f"  JSON q: {match_item['q'][:60]}...")
        print(f"  JSON c5 (Todo): {match_item.get('c5')}")
        print(f"  JSON c7 (Hedge): {match_item.get('c7')}")
        print(f"  JSON c9 (Ref): {match_item.get('c9')}")
        matched_results[r] = {
            "excel_code": code,
            "excel_lv2": lv2,
            "excel_risk": risk,
            "json_no": match_item.get('no'),
            "json_q": match_item.get('q'),
            "json_c5": match_item.get('c5'),
            "json_c7": match_item.get('c7'),
            "json_c9": match_item.get('c9')
        }
    else:
        print(f"Row {r} ({code}) | Lv2: {lv2} -> No close match in JSON")
        print(f"  Risk: {risk[:60]}...")
        matched_results[r] = {
            "excel_code": code,
            "excel_lv2": lv2,
            "excel_risk": risk,
            "json_no": None,
            "json_q": None,
            "json_c5": None,
            "json_c7": None,
            "json_c9": None
        }
    print("-" * 50)

# Save matching results to file for view
with open('scratch/excel_json_matching_analysis.txt', 'w', encoding='utf-8') as out_f:
    json.dump(matched_results, out_f, ensure_ascii=False, indent=2)
