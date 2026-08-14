# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

p_xlsm = r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsm"
wb = openpyxl.load_workbook(p_xlsm, data_only=False)
ws = wb["사전토공사"]

print("=== 최종 검증: Row 2 ~ 6 O, Q, S열 ===")
for r in range(2, 7):
    h = ws.cell(r, 8).value
    o = ws.cell(r, 15)
    q = ws.cell(r, 17)
    s = ws.cell(r, 19)
    print(f"Row {r} [{h}]:")
    print(f"  O열: val='{o.value}' | hl='{o.hyperlink.target if o.hyperlink else None}'")
    print(f"  Q열: val='{q.value}' | hl='{q.hyperlink.target if q.hyperlink else None}'")
    print(f"  S열: val='{s.value}' | hl='{s.hyperlink.target if s.hyperlink else None}'")
