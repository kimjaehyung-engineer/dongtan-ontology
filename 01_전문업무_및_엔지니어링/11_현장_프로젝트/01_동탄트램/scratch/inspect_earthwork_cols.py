# -*- coding: utf-8 -*-
import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

p_xlsm = r'08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsm'
wb = openpyxl.load_workbook(p_xlsm, data_only=False, keep_vba=True)
ws = wb['사전토공사']

print(f'Max col: {ws.max_column}, Max row: {ws.max_row}')
for c in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(c)
    val = ws.cell(1, c).value
    print(f'{col_letter}열 (Col {c}): {val}')

print('\n=== Row 2 ~ 6 샘플 데이터 ===')
for r in range(2, min(7, ws.max_row + 1)):
    f_val = ws.cell(r, 6).value # F열 (작업단위 등)
    l_val = ws.cell(r, 12).value
    n_val = ws.cell(r, 14).value
    o_val = ws.cell(r, 15).value
    p_val = ws.cell(r, 16).value
    q_val = ws.cell(r, 17).value
    s_val = ws.cell(r, 19).value
    print(f'Row {r}: F={f_val} | L={l_val} | N={n_val} | O={o_val} | P={p_val} | Q={q_val} | S={s_val}')
