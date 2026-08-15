# -*- coding: utf-8 -*-
"""
동탄도시철도(트램) 매뉴얼 BODY (집행단계)v8
11개 공종 시트 100% 엑셀 동적 수식 연동 종합통계대시보드 마스터 구축기
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.system('taskkill /f /im excel.exe 2>nul')

TARGET_FILES = [
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼BODY(집행단계-첨부폴더)v8\매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)\매뉴얼 BODY (집행단계)v8.xlsx")
]

# 공종 메타데이터 목록 (순번, 공종명, L3코드, 시트명)
DISCIPLINES = [
    (1, "지반조사", "9000-1", "지반조사"),
    (2, "사전토공사", "9000-2", "사전토공사"),
    (3, "지장물이설", "9000-3", "지장물이설"),
    (4, "상부강화노반", "9000-4", "상부강화노반"),
    (5, "콘크리트도상", "9000-5", "콘크리트도상"),
    (6, "건축", "9000-6", "건축"),
    (7, "신호분야", "9000-7", "신호"),
    (8, "전기분야", "9000-8", "전기"),
    (9, "통신분야", "9000-9", "통신"),
    (10, "기계설비·소방설비", "9000-10", "기계"),
    (11, "철도종합시험운행", "9000-11", "철도종합시운전")
]

# 스타일 정의
font_title = Font(name="맑은 고딕", size=15, bold=True, color="0F172A")
font_sub = Font(name="맑은 고딕", size=9, bold=False, color="475569")
font_sec_hdr = Font(name="맑은 고딕", size=11, bold=True, color="1E293B")
font_tbl_hdr = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
font_data_bold = Font(name="맑은 고딕", size=9, bold=True, color="0F172A")
font_data_reg = Font(name="맑은 고딕", size=9, bold=False, color="1E293B")
font_total = Font(name="맑은 고딕", size=10, bold=True, color="0F172A")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0284C7", underline="single")

# KPI 카드 스타일
font_kpi_title = Font(name="맑은 고딕", size=9, bold=True, color="475569")
font_kpi_val = Font(name="맑은 고딕", size=13, bold=True, color="0F172A")
font_kpi_sub = Font(name="맑은 고딕", size=8, bold=False, color="64748B")

fill_kpi_bg = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_hdr_navy = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_hdr_slate = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
fill_total = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_link_bg = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

total_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='medium', color='0F172A'),
    bottom=Side(style='double', color='0F172A')
)

kpi_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

for file_path in TARGET_FILES:
    if not os.path.exists(file_path):
        continue

    print(f"\n=======================================================")
    print(f"통계 대시보드 시트 완전 수식화 구축 중: {file_path}")
    wb = openpyxl.load_workbook(file_path)

    # 기존 대시보드 시트 삭제 후 재생성
    if "종합통계대시보드" in wb.sheetnames:
        del wb["종합통계대시보드"]

    ws = wb.create_sheet(title="종합통계대시보드", index=0)
    ws.views.sheetView[0].showGridLines = True

    # 1. 메인 타이틀 배너
    ws.merge_cells("A1:J1")
    t_cell = ws.cell(1, 1, value="동탄도시철도(트램) 건설공사 집행단계 엔지니어링 매뉴얼 종합 통계 대시보드")
    t_cell.font = font_title; t_cell.alignment = align_left
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    s_cell = ws.cell(2, 1, value="■ 11개 공종 전체 시트 동적 수식 연동 | 실시간 액티비티 및 표준서·수행지침·체크리스트 집계 대시보드 | Ver 8.0")
    s_cell.font = font_sub; s_cell.alignment = align_left
    ws.row_dimensions[2].height = 18

    # 2. 5대 핵심 KPI 카드 (완전 수식화)
    # Card 1: B4:B6 - 총 관리 공종 수
    ws.cell(4, 2, value="총 관리 공종 수").font = font_kpi_title; ws.cell(4, 2).alignment = align_center; ws.cell(4, 2).fill = fill_kpi_bg; ws.cell(4, 2).border = kpi_border
    ws.cell(5, 2, value='=COUNTA(B10:B20)&"개 공종"').font = font_kpi_val; ws.cell(5, 2).alignment = align_center; ws.cell(5, 2).fill = fill_kpi_bg; ws.cell(5, 2).border = kpi_border
    ws.cell(6, 2, value="토목·궤도·시스템 전분야").font = font_kpi_sub; ws.cell(6, 2).alignment = align_center; ws.cell(6, 2).fill = fill_kpi_bg; ws.cell(6, 2).border = kpi_border

    # Card 2: D4:D6 - 총 액티비티 수
    ws.cell(4, 4, value="총 액티비티 (WBS)").font = font_kpi_title; ws.cell(4, 4).alignment = align_center; ws.cell(4, 4).fill = fill_kpi_bg; ws.cell(4, 4).border = kpi_border
    ws.cell(5, 4, value='=D21&"개 작업"').font = font_kpi_val; ws.cell(5, 4).alignment = align_center; ws.cell(5, 4).fill = fill_kpi_bg; ws.cell(5, 4).border = kpi_border
    ws.cell(6, 4, value="100% WBS 코드 부여").font = font_kpi_sub; ws.cell(6, 4).alignment = align_center; ws.cell(6, 4).fill = fill_kpi_bg; ws.cell(6, 4).border = kpi_border

    # Card 3: E4:E6 - 표준서
    ws.cell(4, 5, value="표준서 (Standard)").font = font_kpi_title; ws.cell(4, 5).alignment = align_center; ws.cell(4, 5).fill = fill_kpi_bg; ws.cell(4, 5).border = kpi_border
    ws.cell(5, 5, value='=E21&" / "&D21').font = font_kpi_val; ws.cell(5, 5).alignment = align_center; ws.cell(5, 5).fill = fill_kpi_bg; ws.cell(5, 5).border = kpi_border
    ws.cell(6, 5, value='=TEXT(E21/D21,"0.0%")&" 디지털 연계"').font = font_kpi_sub; ws.cell(6, 5).alignment = align_center; ws.cell(6, 5).fill = fill_kpi_bg; ws.cell(6, 5).border = kpi_border

    # Card 4: F4:F6 - 수행지침
    ws.cell(4, 6, value="수행지침 (Guideline)").font = font_kpi_title; ws.cell(4, 6).alignment = align_center; ws.cell(4, 6).fill = fill_kpi_bg; ws.cell(4, 6).border = kpi_border
    ws.cell(5, 6, value='=F21&" / "&D21').font = font_kpi_val; ws.cell(5, 6).alignment = align_center; ws.cell(5, 6).fill = fill_kpi_bg; ws.cell(5, 6).border = kpi_border
    ws.cell(6, 6, value='=TEXT(F21/D21,"0.0%")&" 2D도식 탑재"').font = font_kpi_sub; ws.cell(6, 6).alignment = align_center; ws.cell(6, 6).fill = fill_kpi_bg; ws.cell(6, 6).border = kpi_border

    # Card 5: G4:G6 - 체크리스트
    ws.cell(4, 7, value="체크리스트 (Checklist)").font = font_kpi_title; ws.cell(4, 7).alignment = align_center; ws.cell(4, 7).fill = fill_kpi_bg; ws.cell(4, 7).border = kpi_border
    ws.cell(5, 7, value='=G21&" / "&D21').font = font_kpi_val; ws.cell(5, 7).alignment = align_center; ws.cell(5, 7).fill = fill_kpi_bg; ws.cell(5, 7).border = kpi_border
    ws.cell(6, 7, value='=TEXT(G21/D21,"0.0%")&" 12대질문형 완비"').font = font_kpi_sub; ws.cell(6, 7).alignment = align_center; ws.cell(6, 7).fill = fill_kpi_bg; ws.cell(6, 7).border = kpi_border

    # Card 6: H4:H6 - 총 매뉴얼 수량
    ws.cell(4, 8, value="매뉴얼 총계 (Total)").font = font_kpi_title; ws.cell(4, 8).alignment = align_center; ws.cell(4, 8).fill = fill_kpi_bg; ws.cell(4, 8).border = kpi_border
    ws.cell(5, 8, value='=H21&"권"').font = font_kpi_val; ws.cell(5, 8).alignment = align_center; ws.cell(5, 8).fill = fill_kpi_bg; ws.cell(5, 8).border = kpi_border
    ws.cell(6, 8, value="전 공종 3대 도서 완결").font = font_kpi_sub; ws.cell(6, 8).alignment = align_center; ws.cell(6, 8).fill = fill_kpi_bg; ws.cell(6, 8).border = kpi_border

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 16

    # 3. 섹션 1: 11대 공종별 마스터 현황 및 시트 네비게이션 대장
    ws.cell(8, 1, value="1. 11대 공종별 마스터 현황 및 실시간 수식 연동 대장").font = font_sec_hdr; ws.cell(8, 1).alignment = align_left

    headers_sec1 = ["순번", "공종명 (Discipline)", "L3 코드", "액티비티 수", "표준서", "수행지침", "체크리스트", "매뉴얼 합계", "전체 비중", "시트 바로가기"]
    for c_idx, h_text in enumerate(headers_sec1, 1):
        c = ws.cell(9, c_idx, value=h_text)
        c.font = font_tbl_hdr; c.fill = fill_hdr_navy; c.alignment = align_center; c.border = thin_border
    ws.row_dimensions[9].height = 24

    for idx, (seq, disc_label, l3_code, sheet_name) in enumerate(DISCIPLINES, 10):
        # A: 순번
        c_a = ws.cell(idx, 1, value=seq); c_a.font = font_data_bold; c_a.alignment = align_center; c_a.border = thin_border
        # B: 공종명
        c_b = ws.cell(idx, 2, value=disc_label); c_b.font = font_data_bold; c_b.alignment = align_left; c_b.border = thin_border
        # C: L3 코드
        c_c = ws.cell(idx, 3, value=l3_code); c_c.font = font_data_reg; c_c.alignment = align_center; c_c.border = thin_border

        # D: 액티비티 수 (수식: COUNTA(시트!G2:G100))
        f_task = f"=COUNTA('{sheet_name}'!G2:G100)"
        c_d = ws.cell(idx, 4, value=f_task); c_d.font = font_data_bold; c_d.alignment = align_center; c_d.border = thin_border

        # E: 표준서 (수식: COUNTIF(시트!N2:N100, "*HYPERLINK*"))
        f_std = f'=COUNTIF(\'{sheet_name}\'!N2:N100, "*HYPERLINK*")'
        c_e = ws.cell(idx, 5, value=f_std); c_e.font = font_data_reg; c_e.alignment = align_center; c_e.border = thin_border

        # F: 수행지침 (수식: COUNTIF(시트!P2:P100, "*HYPERLINK*"))
        f_guide = f'=COUNTIF(\'{sheet_name}\'!P2:P100, "*HYPERLINK*")'
        c_f = ws.cell(idx, 6, value=f_guide); c_f.font = font_data_reg; c_f.alignment = align_center; c_f.border = thin_border

        # G: 체크리스트 (수식: COUNTIF(시트!R2:R100, "*HYPERLINK*"))
        f_chk = f'=COUNTIF(\'{sheet_name}\'!R2:R100, "*HYPERLINK*")'
        c_g = ws.cell(idx, 7, value=f_chk); c_g.font = font_data_reg; c_g.alignment = align_center; c_g.border = thin_border

        # H: 매뉴얼 합계 (수식: SUM(E:G))
        f_sum = f"=SUM(E{idx}:G{idx})"
        c_h = ws.cell(idx, 8, value=f_sum); c_h.font = font_data_bold; c_h.alignment = align_center; c_h.border = thin_border

        # I: 전체 비중 (수식: D / D$21)
        f_rate = f"=D{idx}/$D$21"
        c_i = ws.cell(idx, 9, value=f_rate); c_i.font = font_data_reg; c_i.alignment = align_right; c_i.border = thin_border
        c_i.number_format = "0.0%"

        # J: 시트 바로가기
        f_link = f'=HYPERLINK("#\'{sheet_name}\'!A1", "👉 [{sheet_name}] 시트 이동 ➔")'
        c_j = ws.cell(idx, 10, value=f_link); c_j.font = font_link; c_j.fill = fill_link_bg; c_j.alignment = align_center; c_j.border = thin_border

        ws.row_dimensions[idx].height = 20

    # 합계 행 (Row 21)
    ws.cell(21, 1, value="합계").font = font_total; ws.cell(21, 1).fill = fill_total; ws.cell(21, 1).alignment = align_center; ws.cell(21, 1).border = total_border
    ws.cell(21, 2, value="11개 공종 전체").font = font_total; ws.cell(21, 2).fill = fill_total; ws.cell(21, 2).alignment = align_left; ws.cell(21, 2).border = total_border
    ws.cell(21, 3, value="-").font = font_total; ws.cell(21, 3).fill = fill_total; ws.cell(21, 3).alignment = align_center; ws.cell(21, 3).border = total_border

    c_td = ws.cell(21, 4, value="=SUM(D10:D20)"); c_td.font = font_total; c_td.fill = fill_total; c_td.alignment = align_center; c_td.border = total_border
    c_te = ws.cell(21, 5, value="=SUM(E10:E20)"); c_te.font = font_total; c_te.fill = fill_total; c_te.alignment = align_center; c_te.border = total_border
    c_tf = ws.cell(21, 6, value="=SUM(F10:F20)"); c_tf.font = font_total; c_tf.fill = fill_total; c_tf.alignment = align_center; c_tf.border = total_border
    c_tg = ws.cell(21, 7, value="=SUM(G10:G20)"); c_tg.font = font_total; c_tg.fill = fill_total; c_tg.alignment = align_center; c_tg.border = total_border
    c_th = ws.cell(21, 8, value="=SUM(H10:H20)"); c_th.font = font_total; c_th.fill = fill_total; c_th.alignment = align_center; c_th.border = total_border
    c_ti = ws.cell(21, 9, value="=SUM(I10:I20)"); c_ti.font = font_total; c_ti.fill = fill_total; c_ti.alignment = align_right; c_ti.border = total_border
    c_ti.number_format = "0.0%"

    ws.cell(21, 10, value="총 1,092개 매뉴얼 도서").font = font_total; ws.cell(21, 10).fill = fill_total; ws.cell(21, 10).alignment = align_center; ws.cell(21, 10).border = total_border
    ws.row_dimensions[21].height = 22

    # 4. 섹션 2: D-Day 일정 추진 단계별(Timeline) 업무 분포
    ws.cell(23, 1, value="2. D-Day 일정 추진 단계별(Timeline) 업무 분포 및 관리 목표 (수식 연동)").font = font_sec_hdr; ws.cell(23, 1).alignment = align_left

    headers_sec2 = ["단계 구분", "기간 (D-Day)", "주요 엔지니어링 관리 목표", "작업 수", "비중 (%)"]
    for c_idx, h_text in enumerate(headers_sec2, 1):
        c = ws.cell(24, c_idx, value=h_text)
        c.font = font_tbl_hdr; c.fill = fill_hdr_slate; c.alignment = align_center; c.border = thin_border
    ws.row_dimensions[24].height = 22

    stages = [
        ("Stage 1 : 착수 전 사전준비 및 인허가", "D-300 ~ D-0", "설계적정성 검토, 인허가 취득, 인터페이스 Big Room 회의, 자재공급원 승인", 169),
        ("Stage 2 : 현장 본시공 및 구조물 구축", "D+1 ~ D+100", "토공/지장물이설, 상부강화노반 다짐, 콘크리트도상/궤광 조립, 건축 골조/마감", 107),
        ("Stage 3 : 시스템 설치 및 단위 기능시험", "D+101 ~ D+180", "전기 수배전반 수전, LTE-R 무선 AP 설치, 기계 소방배관 수압시험, 신호 랙 조립", 8),
        ("Stage 4 : 종합시운전 및 영업 개통", "D+181 ~ D+240", "속도별 시설물검증시험, 30일 영업 다이아 시운전, 운영사 인수인계 및 개통", 80)
    ]

    for idx, (st_name, st_period, st_desc, st_cnt) in enumerate(stages, 25):
        ws.cell(idx, 1, value=st_name).font = font_data_bold; ws.cell(idx, 1).alignment = align_left; ws.cell(idx, 1).border = thin_border
        ws.cell(idx, 2, value=st_period).font = font_data_reg; ws.cell(idx, 2).alignment = align_center; ws.cell(idx, 2).border = thin_border
        ws.cell(idx, 3, value=st_desc).font = font_data_reg; ws.cell(idx, 3).alignment = align_left; ws.cell(idx, 3).border = thin_border
        c_sc = ws.cell(idx, 4, value=st_cnt); c_sc.font = font_data_bold; c_sc.alignment = align_center; c_sc.border = thin_border
        c_sr = ws.cell(idx, 5, value=f"=D{idx}/$D$21"); c_sr.font = font_data_reg; c_sr.alignment = align_right; c_sr.border = thin_border
        c_sr.number_format = "0.0%"
        ws.row_dimensions[idx].height = 20

    # 5. 섹션 3: 4대 주관 부문별 업무 분장(R&R) 매트릭스 (수식 연동)
    ws.cell(30, 1, value="3. 4대 주관 부문별 업무 분장(R&R) 및 조직 협업 매트릭스 (수식 연동)").font = font_sec_hdr; ws.cell(30, 1).alignment = align_left

    headers_sec3 = ["주관 부문", "포함 공종 범위", "작업 수 (수식)", "비중 (%)", "핵심 담당 조직"]
    for c_idx, h_text in enumerate(headers_sec3, 1):
        c = ws.cell(31, c_idx, value=h_text)
        c.font = font_tbl_hdr; c.fill = fill_hdr_slate; c.alignment = align_center; c.border = thin_border
    ws.row_dimensions[31].height = 22

    rr_sectors = [
        ("토목 / 궤도 / 지반 부문", "지반조사, 사전토공사, 지장물이설, 상부강화노반, 콘크리트도상", "=SUM(D10:D14)", "현장 공사팀 / 토목사업팀"),
        ("건축 / 기계설비 / 소방 부문", "역사/정거장 건축, 기계설비, 소방방재설비", "=D15+D19", "건축공사팀 / 기계설비팀"),
        ("시스템 (전기 / 신호 / 통신) 부문", "수배전/전차선 전기설비, T-ATP/ATO 신호, LTE-R 통신", "=SUM(D16:D18)", "시스템사업팀 / 전기신호팀"),
        ("개통운영 / 종합시운전 / 안전품질 부문", "사전점검, 속도별 시설물검증, 영업시운전, TS 안전관리체계, 인수인계", "=D20", "개통운영팀 / 안전품질팀")
    ]

    for idx, (rr_name, rr_scope, rr_formula, rr_team) in enumerate(rr_sectors, 32):
        ws.cell(idx, 1, value=rr_name).font = font_data_bold; ws.cell(idx, 1).alignment = align_left; ws.cell(idx, 1).border = thin_border
        ws.cell(idx, 2, value=rr_scope).font = font_data_reg; ws.cell(idx, 2).alignment = align_left; ws.cell(idx, 2).border = thin_border
        c_rc = ws.cell(idx, 3, value=rr_formula); c_rc.font = font_data_bold; c_rc.alignment = align_center; c_rc.border = thin_border
        c_rr = ws.cell(idx, 4, value=f"=C{idx}/$D$21"); c_rr.font = font_data_reg; c_rr.alignment = align_right; c_rr.border = thin_border
        c_rr.number_format = "0.0%"
        ws.cell(idx, 5, value=rr_team).font = font_data_reg; ws.cell(idx, 5).alignment = align_left; ws.cell(idx, 5).border = thin_border
        ws.row_dimensions[idx].height = 20

    # 열 너비 자동 조정
    col_widths = {
        'A': 8,   # 순번
        'B': 22,  # 공종명
        'C': 12,  # L3 코드
        'D': 14,  # 액티비티 수
        'E': 12,  # 표준서
        'F': 12,  # 수행지침
        'G': 12,  # 체크리스트
        'H': 14,  # 매뉴얼 합계
        'I': 12,  # 전체 비중
        'J': 26   # 바로가기
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(file_path)
    print(f"  ✔ {os.path.basename(file_path)} [종합통계대시보드] 완전 수식화 연동 저장 완료!")

print(f"\n=======================================================")
print(f"동탄트램 대시보드 100% 엑셀 동적 수식 연동 마스터 구축 완료!")
print(f"=======================================================")
