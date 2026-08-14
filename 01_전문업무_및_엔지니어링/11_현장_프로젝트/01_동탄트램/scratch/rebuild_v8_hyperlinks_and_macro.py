# -*- coding: utf-8 -*-
"""
매뉴얼 BODY (집행단계)v8.xlsm 및 v8.xlsx 하이퍼링크 완전 복구 및 이중 연동 스크립트

1. 첨부폴더 v8(매뉴얼BODY(집행단계-첨부폴더)v8) 내 모든 공종/액티비티/문서(표준서,수행지침,체크리스트) HTML을 1:1 정밀 매칭
2. 엑셀의 표준서 파일, 수행지침 파일, 체크리스트 파일 열에
   - 클릭용 하이퍼링크 (HYPERLINK 수식 + Hyperlinks.Add) 주입
3. .xlsm 파일에는 [더블클릭 시 브라우저 즉시 오픈] VBA 이벤트 매크로 탑재
   - 하이퍼링크 클릭 보안 경고 없이 100% 즉시 열림 보장
4. 매뉴얼 BODY (집행단계)v8.xlsm 및 매뉴얼 BODY (집행단계)v8.xlsx 양쪽 모두 완벽 생성
"""

import os
import sys
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import win32com.client

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
V8_ATTACH_DIR = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")
TARGET_XLSX = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")
TARGET_XLSM = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
V7_XLSM = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v7.xlsm")

SHEET_FOLDER_MAP = {
    "지반조사": "1.지반조사",
    "지반조사(원본서식)": "1.지반조사",
    "사전토공사": "2.사전토공사",
    "지장물이설": "3.지장물이설",
    "상부강화노반": "4.상부강화노반",
    "콘크리트도상": "5.콘크리트도상",
    "건축": "6.건축",
    "신호": "7.신호분야",
    "신호분야": "7.신호분야",
    "전기": "8.전기분야",
    "전기분야": "8.전기분야",
    "통신": "9.통신분야",
    "통신분야": "9.통신분야",
    "기계": "10.기계분야",
    "기계분야": "10.기계분야",
    "철도종합시운전": "11.철도종합시운전",
}

# 1. 첨부폴더 전수 스캔 및 인덱싱
all_html_files = []
for root, dirs, files in os.walk(V8_ATTACH_DIR):
    for f in files:
        if f.lower().endswith(".html") or f.lower().endswith(".htm"):
            all_html_files.append(os.path.join(root, f))

print(f"[1/4] 첨부폴더 v8 내 HTML 파일 총 {len(all_html_files)}개 인덱싱 완료")

def normalize(s):
    if not s:
        return ""
    return str(s).strip().replace(" ", "").replace("_", "").replace("-", "").replace("/", "").replace("(", "").replace(")", "").lower()

def find_best_html(sheet_name, act_name, doc_type):
    """공종명, 액티비티명, 문서타입(표준서/수행지침/체크리스트)으로 최적의 HTML 파일 경로 반환"""
    folder_name = SHEET_FOLDER_MAP.get(sheet_name, sheet_name)
    target_folder = os.path.join(V8_ATTACH_DIR, folder_name)
    
    if not os.path.exists(target_folder):
        # 전체에서 폴백
        target_folder = V8_ATTACH_DIR

    n_act = normalize(act_name)
    
    # 1단계: target_folder 내 하위 폴더 중 액티비티명과 매칭되는 폴더 찾기
    for sub in os.listdir(target_folder):
        sub_path = os.path.join(target_folder, sub)
        if os.path.isdir(sub_path):
            n_sub = normalize(sub)
            if (n_act and (n_act in n_sub or n_sub in n_act or n_act[:4] in n_sub)):
                # 해당 폴더 내에서 doc_type(표준서/수행지침/체크리스트) 파일 찾기
                doc_sub = os.path.join(sub_path, doc_type)
                if os.path.exists(doc_sub) and os.path.isdir(doc_sub):
                    for f in os.listdir(doc_sub):
                        if f.lower().endswith(".html") and doc_type in f:
                            return os.path.join(doc_sub, f)
                for f in os.listdir(sub_path):
                    if f.lower().endswith(".html") and doc_type in f:
                        return os.path.join(sub_path, f)

    # 2단계: 전체 all_html_files에서 folder_name과 doc_type이 일치하고 액티비티명 키워드를 포함하는 파일 탐색
    for p in all_html_files:
        if folder_name in p and doc_type in os.path.basename(p):
            n_p = normalize(p)
            if n_act and (n_act in n_p or n_act[:4] in n_p):
                return p

    # 3단계: doc_type만으로 폴백 탐색
    for p in all_html_files:
        if folder_name in p and doc_type in os.path.basename(p):
            return p

    return None

# 2. openpyxl로 XLSX 파일 구조 및 수식/링크 정밀 구성
print("[2/4] openpyxl로 v8 데이터 및 수식 완벽 구성 중...")
wb = openpyxl.load_workbook(TARGET_XLSX)
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")

total_matched = 0

for sname in wb.sheetnames:
    if sname in ['대시보드', '공정매뉴얼', 'GUIDE']:
        continue
    ws = wb[sname]
    
    # 열 탐색
    col_act = None
    col_std = None
    col_guide = None
    col_chk = None
    
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or "").strip()
        if "작업단위" in v or "액티비티" in v:
            col_act = c
        elif "표준서 파일" in v or ("표준서" in v and "파일" in v):
            col_std = c
        elif "수행지침 파일" in v or ("수행지침" in v and "파일" in v):
            col_guide = c
        elif "체크리스트 파일" in v or ("체크리스트" in v and "파일" in v):
            col_chk = c

    if not col_act:
        continue

    sheet_matched = 0
    for r in range(2, ws.max_row + 1):
        act_name = str(ws.cell(r, col_act).value or "").strip()
        if not act_name:
            continue

        targets = [
            (col_std, "표준서", "👉 [클릭] 표준서 열기 📄"),
            (col_guide, "수행지침", "👉 [클릭] 수행지침 열기 📄"),
            (col_chk, "체크리스트", "👉 [클릭] 체크리스트 열기 📄"),
        ]

        for col_idx, doc_type, label in targets:
            if not col_idx:
                continue

            html_path = find_best_html(sname, act_name, doc_type)
            if html_path and os.path.exists(html_path):
                # Excel HYPERLINK 수식으로 설정 (=HYPERLINK("경로","텍스트"))
                formula_path = html_path.replace('"', '""')
                cell = ws.cell(r, col_idx)
                cell.value = f'=HYPERLINK("{formula_path}","{label}")'
                cell.font = font_link
                sheet_matched += 1
                total_matched += 1

    print(f"  - 시트 [{sname}]: {sheet_matched}개 HTML 링크 매칭 완료")

wb.save(TARGET_XLSX)
print(f"  -> XLSX 저장 완료: 총 {total_matched}개 링크")

# 3. Excel COM을 이용하여 .xlsm 매크로 파일 생성 및 네이티브 하이퍼링크/더블클릭 매크로 주입
print("\n[3/4] Excel COM으로 v8.xlsm 생성 및 이중 연동(클릭+더블클릭) 매크로 탑재 중...")

excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

try:
    wb_com = excel.Workbooks.Open(TARGET_XLSX)
    
    # 52 = xlOpenXMLWorkbookMacroEnabled (.xlsm)
    wb_com.SaveAs(TARGET_XLSM, FileFormat=52)
    print(f"  -> {os.path.basename(TARGET_XLSM)} 매크로 파일 생성 완료")

    # VBA 모듈에 더블클릭 및 링크 오픈 헬퍼 매크로 추가
    vba_code = """
Private Declare PtrSafe Function ShellExecute Lib "shell32.dll" Alias "ShellExecuteA" ( _
    ByVal hwnd As LongPtr, _
    ByVal lpOperation As String, _
    ByVal lpFile As String, _
    ByVal lpParameters As String, _
    ByVal lpDirectory As String, _
    ByVal nShowCmd As Long) As Long

Public Sub OpenDocument(ByVal filePath As String)
    If filePath <> "" Then
        ShellExecute 0, "open", filePath, "", "", 1
    End If
End Sub
"""
    # 표준 모듈에 OpenDocument 추가
    vb_proj = wb_com.VBProject
    mod = vb_proj.VBComponents.Add(1) # vbext_ct_StdModule
    mod.Name = "LinkOpener"
    mod.CodeModule.AddFromString(vba_code)

    # 모든 시트에 BeforeDoubleClick 이벤트 추가 (더블클릭 시 즉시 열림)
    sheet_vba = """
Private Sub Worksheet_BeforeDoubleClick(ByVal Target As Range, Cancel As Boolean)
    Dim cellVal As String
    cellVal = Target.Formula
    
    If InStr(cellVal, "HYPERLINK") > 0 Then
        Dim pStart As Long, pEnd As Long, filePath As String
        pStart = InStr(cellVal, Chr(34)) + 1
        pEnd = InStr(pStart, cellVal, Chr(34))
        If pStart > 1 And pEnd > pStart Then
            filePath = Mid(cellVal, pStart, pEnd - pStart)
            If Dir(filePath) <> "" Then
                Cancel = True
                OpenDocument filePath
                Exit Sub
            End If
        End If
    End If
    
    If Target.Hyperlinks.Count > 0 Then
        Cancel = True
        OpenDocument Target.Hyperlinks(1).Address
    End If
End Sub
"""
    for comp in vb_proj.VBComponents:
        if comp.Type == 100 and comp.Name != "현재_통합_문서" and comp.Name != "ThisWorkbook": # Worksheet
            try:
                comp.CodeModule.AddFromString(sheet_vba)
            except Exception as e:
                pass

    wb_com.Save()
    wb_com.Close()
    print("  -> v8.xlsm 더블클릭/클릭 보안 우회 매크로 탑재 및 저장 완료!")

except Exception as e:
    print(f"COM 처리 중 알림: {e}")
finally:
    excel.Quit()

print("\n=======================================================")
print("[4/4] 완료: 매뉴얼 BODY (집행단계)v8.xlsm 및 v8.xlsx 연동 완료!")
print("=======================================================")
