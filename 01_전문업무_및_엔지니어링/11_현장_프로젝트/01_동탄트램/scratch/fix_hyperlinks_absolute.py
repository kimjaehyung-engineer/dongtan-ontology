# -*- coding: utf-8 -*-
"""
엑셀 v8/v6/v5 마스터 파일의 모든 하이퍼링크를
상대경로 -> file:/// 절대경로로 전환하여 더블클릭 즉시 실행되도록 복구
"""

import os
import openpyxl
from openpyxl.styles import Font

BASE_DIR = r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)"

EXCEL_FILES = [
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v6.xlsx"),
    os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v5.xlsx")
]

font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")

def convert_to_absolute(excel_path):
    if not os.path.exists(excel_path):
        print(f"Skipping: {excel_path}")
        return

    excel_abs = os.path.abspath(excel_path)
    excel_dir = os.path.dirname(excel_abs)

    print(f"\nProcessing: {excel_abs}")
    wb = openpyxl.load_workbook(excel_path)

    total = 0
    ok = 0
    broken = 0

    for sname in wb.sheetnames:
        if sname in ['대시보드', '공정매뉴얼', 'GUIDE']:
            continue
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if not cell.hyperlink:
                    continue
                target = cell.hyperlink.target
                if not target:
                    continue
                total += 1

                # 이미 절대경로면 건너뜀
                if target.startswith("file:///") or target.startswith("http"):
                    ok += 1
                    continue

                # 상대경로 -> 절대경로 변환
                abs_target = os.path.join(excel_dir, target)
                abs_target = os.path.normpath(abs_target)

                if os.path.exists(abs_target):
                    # file:/// URI 포맷으로 변환 (역슬래시 -> 슬래시)
                    uri = "file:///" + abs_target.replace("\\", "/")
                    cell.hyperlink = uri
                    cell.font = font_link
                    ok += 1
                else:
                    broken += 1
                    print(f"  BROKEN: {sname} | {cell.coordinate} | {abs_target}")

    wb.save(excel_path)
    print(f"  Done: total={total}, ok={ok}, broken={broken}")

for f in EXCEL_FILES:
    convert_to_absolute(f)

print("\n=== ALL DONE: Hyperlinks converted to file:/// absolute paths! ===")
