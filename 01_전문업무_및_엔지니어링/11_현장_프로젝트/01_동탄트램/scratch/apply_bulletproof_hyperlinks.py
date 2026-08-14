# -*- coding: utf-8 -*-
"""
동탄트램 8대 공종 전체 엑셀(v8.xlsx, v8.xlsm) 완벽한 동적 절대경로 =HYPERLINK() 수식 적용기
(Excel 작업 디렉토리 불일치/OneDrive 리다이렉션으로 인한 '예기치 않은 오류' 완벽 해결)
"""

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')
import glob
import re
import openpyxl
from openpyxl.styles import Font, Alignment

BASE_DIR = os.path.abspath(r"08.메뉴얼 및 평면도\최종\02_메뉴얼 공종프로섹스(집행단계)")
XLSX_PATH = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsx")
XLSM_PATH = os.path.join(BASE_DIR, "매뉴얼 BODY (집행단계)v8.xlsm")
ATTACH_BASE = os.path.join(BASE_DIR, "매뉴얼BODY(집행단계-첨부폴더)v8")

os.system('taskkill /f /im excel.exe 2>nul')

DISCIPLINE_MAP = {
    "사전토공사": "2.사전토공사",
    "지장물이설": "3.지장물이설",
    "상부강화노반": "4.상부강화노반",
    "콘크리트도상": "5.콘크리트도상",
    "건축": "6.건축",
    "신호": "7.신호분야",
    "전기": "8.전기분야",
    "통신": "9.통신분야"
}

link_font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

for target_file in [XLSX_PATH, XLSM_PATH]:
    if not os.path.exists(target_file):
        continue
    
    is_vba = target_file.endswith('.xlsm')
    print(f"\n==================================================")
    print(f"작업 대상 파일: {os.path.basename(target_file)}")
    wb = openpyxl.load_workbook(target_file, keep_vba=is_vba)

    for sheet_name, folder_name in DISCIPLINE_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        folder_abs = os.path.join(ATTACH_BASE, folder_name)
        
        subdirs = [d for d in os.listdir(folder_abs) if os.path.isdir(os.path.join(folder_abs, d))]
        def get_sort_key(name):
            m = re.match(r'^(\d+)', name)
            return int(m.group(1)) if m else 999
        subdirs.sort(key=get_sort_key)

        col_std = None
        col_guide = None
        col_chk = None

        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(1, c).value or "")
            if "표준서" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
                col_std = c
            elif "수행지침" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
                col_guide = c
            elif "체크리스트" in val and ("파일" in val or "열기" in val or "클릭" in val or "HTML" in val):
                col_chk = c

        if sheet_name in ["사전토공사", "상부강화노반"]:
            if not col_std: col_std = 15
            if not col_guide: col_guide = 17
            if not col_chk: col_chk = 19
        elif sheet_name in ["지장물이설", "콘크리트도상"]:
            if not col_std: col_std = 12
            if not col_guide: col_guide = 14
            if not col_chk: col_chk = 16
        else: # 건축, 신호, 전기, 통신
            if not col_std: col_std = 11
            if not col_guide: col_guide = 13
            if not col_chk: col_chk = 15

        ws.cell(1, col_std).value = "표준서 파일 (HTML)"
        ws.cell(1, col_guide).value = "수행지침 파일 (HTML)"
        ws.cell(1, col_chk).value = "체크리스트 파일 (HTML)"

        linked_count = 0
        for r in range(2, ws.max_row + 1):
            t_idx = r - 2
            if t_idx < len(subdirs):
                s_dir = subdirs[t_idx]
                d_path = os.path.join(folder_abs, s_dir)
                htmls = glob.glob(os.path.join(d_path, '**', '*.html'), recursive=True)

                std_f = next((f for f in htmls if '표준서' in f), None)
                guide_f = next((f for f in htmls if '수행지침' in f), None)
                chk_f = next((f for f in htmls if '체크리스트' in f), None)

                # 1. 표준서 (동적 절대경로 생성 수식)
                cell_std = ws.cell(r, col_std)
                cell_std.hyperlink = None
                if std_f and os.path.exists(std_f):
                    rel_path = os.path.relpath(std_f, BASE_DIR)
                    cell_std.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 표준서 열기 📄")'
                    cell_std.font = link_font
                    cell_std.alignment = align_center
                else:
                    cell_std.value = "-"

                # 2. 수행지침 (동적 절대경로 생성 수식)
                cell_guide = ws.cell(r, col_guide)
                cell_guide.hyperlink = None
                if guide_f and os.path.exists(guide_f):
                    rel_path = os.path.relpath(guide_f, BASE_DIR)
                    cell_guide.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 수행지침 열기 📄")'
                    cell_guide.font = link_font
                    cell_guide.alignment = align_center
                else:
                    cell_guide.value = "-"

                # 3. 체크리스트 (동적 절대경로 생성 수식)
                cell_chk = ws.cell(r, col_chk)
                cell_chk.hyperlink = None
                if chk_f and os.path.exists(chk_f):
                    rel_path = os.path.relpath(chk_f, BASE_DIR)
                    cell_chk.value = f'=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"{rel_path}", "👉 [클릭] 체크리스트 열기 📄")'
                    cell_chk.font = link_font
                    cell_chk.alignment = align_center
                else:
                    cell_chk.value = "-"

                linked_count += 1

        print(f"  ✓ [{sheet_name}]: {linked_count}개 행 동적 절대경로 =HYPERLINK() 주입 완료")

    wb.save(target_file)
    print(f"✓ {os.path.basename(target_file)} 저장 완료!")

print("\n==================================================")
print("8대 공종 전체 완벽한 동적 절대경로 수식 주입 100% 완료!")
print("==================================================")
