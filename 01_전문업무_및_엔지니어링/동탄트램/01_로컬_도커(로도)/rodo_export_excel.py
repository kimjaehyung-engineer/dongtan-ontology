import pandas as pd
from neo4j import GraphDatabase
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

# UTF-8 출력 강제 (윈도우 콘솔 한글 깨짐 방지)
try:
    sys.stdout.reconfigure(encoding='utf-8')
except AttributeError:
    pass

URI = "bolt://localhost:7687"
USER = ""
PASSWORD = ""

OUTPUT_FILE = r"C:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\동탄트램_기본설계_정합성_검토보고서.xlsx"

def export_compliance_report():
    print("=" * 70)
    print(" 📊 [RODO] 설계 정합성 검토 엑셀 보고서 자동 생성 엔진")
    print("=" * 70)
    print("1. 지식망(Memgraph)에서 설계 위배 데이터를 실시간 조회하는 중...")
    
    driver = GraphDatabase.driver(URI, auth=(USER, PASSWORD))
    
    query = """
    MATCH (fact:DesignFact)-[:VIOLATES]->(s:Requirement)
    MATCH (fact)-[:OCCURS_AT]->(loc:Location)
    MATCH (fact)-[:BELONGS_TO]->(disc:Discipline)
    MATCH (fact)-[ev:EVIDENCE_IN]->(doc:Document)
    RETURN disc.name AS `공종 분야`,
           loc.name AS `발생 위치`,
           COALESCE(s.metric, s.name) AS `검토 항목`,
           COALESCE(fact.issue_description, fact.description, fact.name) AS `상세 이슈 및 위배 내용`,
           s.limit_value AS `관리한계 기준치`,
           fact.actual_value AS `실제 기본설계치`,
           s.unit AS `단위`,
           doc.name AS `근거 서류명`,
           ev.page AS `근거 페이지`,
           COALESCE(doc.absolute_path, doc.file_path) AS `로컬 증빙 경로`
    ORDER BY `공종 분야`, `발생 위치`
    """
    
    with driver.session() as session:
        result = session.run(query)
        records = [dict(record) for record in result]
        
    driver.close()
    
    if not records:
        print("❌ [오류] 지식망에 등록된 리스크 데이터가 없어 엑셀을 생성할 수 없습니다.")
        return
        
    print(f" - 총 {len(records)}개의 리스크 데이터를 추출했습니다.")
    
    # Pandas DataFrame 변환
    df = pd.DataFrame(records)
    
    # 초과 차이 계산 및 상태 컬럼 추가
    df["초과 수치"] = df["실제 기본설계치"] - df["관리한계 기준치"]
    df["정합성 상태"] = "🔴 기준 초과 (위배)"
    
    # 컬럼 순서 재배치
    cols = [
        "공종 분야", "발생 위치", "정합성 상태", "검토 항목", 
        "관리한계 기준치", "실제 기본설계치", "초과 수치", "단위", 
        "상세 이슈 및 위배 내용", "근거 서류명", "근거 페이지", "로컬 증빙 경로"
    ]
    df = df[cols]
    
    print("2. 엑셀 파일 생성 및 비즈니스 스타일링 서식 적용 중...")
    
    # Excel Writer 기동 (openpyxl 엔진)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="정합성 검토 결과", index=False)
        
        workbook = writer.book
        worksheet = writer.sheets["정합성 검토 결과"]
        
        # 3. 디자인 서식 정의 (Sleek Dark Navy & Warm Red Accent)
        font_header = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep Navy Blue
        
        font_data = Font(name="맑은 고딕", size=10)
        font_alert = Font(name="맑은 고딕", size=10, bold=True, color="C00000") # Dark Red
        fill_alert = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Orange-Red
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        border_thin = Side(border_style="thin", color="D9D9D9")
        border_double = Side(border_style="double", color="333333")
        
        cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        bottom_total_border = Border(top=border_thin, bottom=border_double)
        
        # 행 높이 설정
        worksheet.row_dimensions[1].height = 28 # 헤더 행은 넉넉하게
        
        # 헤더 스타일 채우기
        for col_idx in range(1, len(cols) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            
        # 데이터 행 스타일 채우기 및 포맷팅
        max_row = worksheet.max_row
        for row in range(2, max_row + 1):
            worksheet.row_dimensions[row].height = 24 # 데이터 행
            
            # 값별 스타일링
            for col_idx in range(1, len(cols) + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.font = font_data
                cell.border = cell_border
                
                # 열별 맞춤 정렬 및 숫자 포맷
                col_name = cols[col_idx - 1]
                
                if col_name in ["공종 분야", "발생 위치", "정합성 상태", "단위", "근거 페이지"]:
                    cell.alignment = align_center
                elif col_name in ["관리한계 기준치", "실제 기본설계치", "초과 수치"]:
                    cell.alignment = align_right
                    cell.number_format = '#,##0.0'  # 소수점 첫째자리 정렬
                else:
                    cell.alignment = align_left
                    
                # 경고 셀 스타일링 (상태가 초과인 경우 강조)
                if col_name in ["정합성 상태", "실제 기본설계치", "초과 수치"]:
                    cell.fill = fill_alert
                    if col_name == "정합성 상태":
                        cell.font = font_alert
                        
        # 4. 열 넓이 자동 조정 (Auto-fit Columns)
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                val = str(cell.value or '')
                # 한글 문자 폭 가중치 조정
                val_len = sum(2 if ord(char) > 128 else 1 for char in val)
                if val_len > max_len:
                    max_len = val_len
            
            # 너무 길어지는 본문이나 경로는 상한선 설정하여 가독성 유지
            adjusted_width = min(max(max_len + 4, 11), 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
        # 그리드라인 보이기 설정 강제
        worksheet.views.sheetView[0].showGridLines = True
        
    print("=" * 70)
    print(f"🎉 [성공] 설계 정합성 검토 엑셀 보고서가 완성되었습니다!")
    print(f"📂 저장 경로: {OUTPUT_FILE}")
    print("=" * 70)

if __name__ == "__main__":
    export_compliance_report()
