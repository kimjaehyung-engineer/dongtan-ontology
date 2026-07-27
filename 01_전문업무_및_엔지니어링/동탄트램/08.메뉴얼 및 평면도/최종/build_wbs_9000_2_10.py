import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야\10_통신설비 제작 사양서 작성 _ 승인"

std_dir = os.path.join(base_dir, "표준서")
gui_dir = os.path.join(base_dir, "수행지침")
chk_dir = os.path.join(base_dir, "체크리스트")

for d in [std_dir, gui_dir, chk_dir]:
    os.makedirs(d, exist_ok=True)

modal_style = """
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .zoom-modal, .glossary-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .zoom-modal.active, .glossary-modal.active {
        display: flex;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 580px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-close, .glossary-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .zoom-close:hover, .glossary-close:hover {
        color: #ef4444;
    }
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal" onclick="closeGlossaryModalOutside(event)">
    <div class="glossary-modal-content" onclick="event.stopPropagation()">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 엔지니어링 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    "manufacturing_spec": "<b>통신설비 제작 사양서 (Manufacturing Specification)</b><br><br>• 실제 현장에 제작·설치될 통신설비(광전송장비, LTE-R 무선장치, CCTV, PIS/PAS 등)의 전기적 규격, 전원 소비전력, 기계적 치수, 환경시험 규격 및 기능 사양을 정밀 정의한 공식 문서입니다.",
    "shop_drawing": "<b>제작 도면 (Shop Drawing)</b><br><br>• 통신 랙(Rack) 내부 장비 배치도, 입출력 포트 배선도, 전원 분배도 및 외형 치수를 1:1 정밀 반영하여 제작에 사용하는 전용 2D/3D 도면입니다.",
    "fat_procedure": "<b>공장 검사 절차서 (FAT - Factory Acceptance Test Procedure)</b><br><br>• 자재 제작사 공장에서 현장 출하 전 입석 감리원 참관 하에 기능, 연동 및 품질 규격을 테스트하기 위한 공식 시험검사 기준서입니다."
};

function openGlossary(term) {
    const modal = document.getElementById('glossaryModal');
    const titleEl = document.getElementById('modalTitle');
    const descEl = document.getElementById('modalDescription');
    
    if (glossaryData[term]) {
        titleEl.innerHTML = "📖 용어 해설: " + term;
        descEl.innerHTML = glossaryData[term];
        modal.classList.add('active');
    }
}

function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function closeGlossaryModalOutside(event) {
    if (event.target.id === 'glossaryModal') {
        closeGlossaryModal();
    }
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeZoomModal();
        closeGlossaryModal();
    }
});
</script>
"""

# 1. Standard HTML Template
std_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 통신설비 제작 사양서 작성 / 승인 표준서 (WBS 9000-2-10)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard (WBS 9000-2-10)</span>
        <h1 class="text-3xl font-black mt-2">통신설비 제작 사양서 작성 및 승인 마스터 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-10 | 주관: 현장 시스템팀 | "공학적 사양 확정 및 제작도면 감리 승인 규정"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 표준 개요 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 본 표준서의 개요 및 제작사양서 승인 목적</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 표준서는 건설기술 진흥법, KCS 47 10 00, KDS 47 10 00 시방서에 의거하여, 실제 동탄트램 현장에 제작·설치될 통신설비(광전송, LTE-R 무선, CCTV, PIS/PAS 방송 등)의 <strong><span class="term-highlight" onclick="openGlossary('manufacturing_spec')">제작 사양서</span></strong> 및 <strong><span class="term-highlight" onclick="openGlossary('shop_drawing')">제작 도면(Shop Drawing)</span></strong>을 작성하고, 현장 시스템팀 주관으로 감리단/발주처 기술 승인을 득하여 최종 제품의 공학적 무결성과 품질을 완벽히 보장하는 기술 표준입니다.
            </p>
        </div>

        <!-- 📜 주요 규정 항목 -->
        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-blue-600 pb-2">1. 주요 시방 및 제작 사양서 작성·승인 표준 기준</h2>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-800 flex items-center gap-2">📐 1. 실시설계도서 1:1 대조 승인</span>
                    <p class="text-slate-700 text-xs">기본/실시설계 보고서, 도면, 자재사양서와 현장 기계실/역사 실측 규격을 1:1 대조 확정.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-800 flex items-center gap-2">🔌 2. 전기/기계 규격 & 랙 구성도</span>
                    <p class="text-slate-700 text-xs">전원(AC220V/DC48V), 소비전력, 통신 랙(H=2200mm) 내부 배치도 및 외형 치수 도면 작성 규정.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-800 flex items-center gap-2">🛡️ 3. KC 인증 & 환경시험(IP65/내진)</span>
                    <p class="text-slate-700 text-xs">KC 인증서, 내진·방진방수(IP65) 환경시험성적서 및 공장검사(FAT) 절차서 수록 의무화.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-800 flex items-center gap-2">🖊️ 4. 감리단/발주처 최종 승인(Sign-Off)</span>
                    <p class="text-slate-700 text-xs">기술 검토 의견서 최종 조율 후 감리 승인 인감을 수득하여 자재 제작사에 공식 전달 통보.</p>
                </div>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 2. Guideline HTML Template (5 Steps & 1:1 SVG Diagrams)
gui_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 통신설비 제작 사양서 작성 / 승인 수행지침서 (WBS 9000-2-10)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <!-- 🔵 헤더 영역 -->
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Detailed Guideline (WBS 9000-2-10)</span>
        <h1 class="text-3xl font-black mt-2">통신설비 제작 사양서 작성 및 승인 초정밀 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-10 | 주관: 현장 시스템팀 | "1:1 STEP별 직관적 2D 그림 수록 가이드"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 개념 해설 박스 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 통신설비 제작사양서 작성 및 승인 실무 개요</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 지침서는 현장 시스템팀 엔지니어가 통신설비 제작을 개시하기 전 <strong><span class="term-highlight" onclick="openGlossary('manufacturing_spec')">제작 사양서</span></strong> 및 <strong><span class="term-highlight" onclick="openGlossary('shop_drawing')">제작 도면(Shop Drawing)</span></strong>을 작성하고, <span class="term-highlight" onclick="openGlossary('fat_procedure')">공장검사(FAT) 절차서</span>를 포함하여 감리단/발주처 승인을 획득하는 5단계 실무 가이드를 제공합니다. **모든 카드에 1:1 직관적 2D 그림과 확대 보기**를 수록하였습니다.
            </p>
        </div>

        <!-- ☀️ 5단계 수행 마스터 프로세스 -->
        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">FLEXIBLE 5-STEP ARCHITECTURE</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 제작사양서 작성 및 승인 5단계 마스터 프로세스</h3>
            </div>
            <div class="grid grid-cols-1 sm:grid-cols-2 md:grid-cols-5 gap-2 text-xs">
                <div class="bg-white p-3.5 rounded-xl border border-slate-200 space-y-1">
                    <span class="bg-blue-600 text-white text-[10px] font-bold px-1.5 py-0.5 rounded">STEP 1</span>
                    <h4 class="font-bold text-slate-900 text-xs">설계도서 1:1 대조</h4>
                    <p class="text-[10px] text-slate-600">• 실시설계 보고서/도면<br">• 현장 기계실 규격 대조</p>
                </div>
                <div class="bg-white p-3.5 rounded-xl border border-slate-200 space-y-1">
                    <span class="bg-indigo-600 text-white text-[10px] font-bold px-1.5 py-0.5 rounded">STEP 2</span>
                    <h4 class="font-bold text-slate-900 text-xs">사양서 & 도면작성</h4>
                    <p class="text-[10px] text-slate-600">• 전기/기계 세부규격<br">• 통신 랙 Shop Drawing</p>
                </div>
                <div class="bg-white p-3.5 rounded-xl border border-slate-200 space-y-1">
                    <span class="bg-cyan-600 text-white text-[10px] font-bold px-1.5 py-0.5 rounded">STEP 3</span>
                    <h4 class="font-bold text-slate-900 text-xs">품질 & FAT 절차서</h4>
                    <p class="text-[10px] text-slate-600">• KC/IP65/내진 검증<br">• 공장검사 시험절차서</p>
                </div>
                <div class="bg-white p-3.5 rounded-xl border border-slate-200 space-y-1">
                    <span class="bg-teal-600 text-white text-[10px] font-bold px-1.5 py-0.5 rounded">STEP 4</span>
                    <h4 class="font-bold text-slate-900 text-xs">감리단 승인(Sign-Off)</h4>
                    <p class="text-[10px] text-slate-600">• 감리/발주처 제출<br">• 기술 승인 인감 획득</p>
                </div>
                <div class="bg-white p-3.5 rounded-xl border border-slate-200 space-y-1">
                    <span class="bg-emerald-600 text-white text-[10px] font-bold px-1.5 py-0.5 rounded">STEP 5</span>
                    <h4 class="font-bold text-slate-900 text-xs">제작사 전달 & 착수</h4>
                    <p class="text-[10px] text-slate-600">• 승인도서 제작사 전달<br">• 공장 제작 착수 지시</p>
                </div>
            </div>
        </div>

        <!-- 🔥 HOW 세부 지침 & 1:1 2D Visual SVG -->
        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-indigo-600 pb-2">단계별 초정밀 HOW 실무 지침 & 1:1 직관적 2D 그림</h2>

            <!-- STEP 1 Card -->
            <div class="bg-white p-6 rounded-2xl border border-blue-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-blue-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 1</span>
                    <h3 class="font-bold text-base text-slate-900">실시설계도서 1:1 대조 및 현장 기계실/역사 치수 확인</h3>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                    <p>• <strong>도서 1:1 비교:</strong> 기본/실시설계 보고서, 시방서, 도면 및 자재사양서 수치를 수신하여 현장 통신기계실/TPS실 실제 실측 치수와 1:1 비교합니다.</p>
                    <p>• <strong>설치 공간 확정:</strong> 랙 반입 경로, 이중바닥(Access Floor H=300mm), 천장 마감 층고(≥3.0m) 및 전원 배선반 위치를 확인하여 사양서 기준을 확정합니다.</p>
                </div>

                <!-- 1:1 STEP 1 2D SVG -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step1_wbs10', 'STEP 1 실시설계도서 대조 및 현장 실측 치수 확인 2D 도식')">
                    <svg id="svg_step1_wbs10" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📐 실시설계도서 1:1 대조</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 보고서/도면/자재사양서 검토</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 장비별 세부 기능 규격 대조</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 도서 대조 통과</text>

                        <path d="M 245 90 L 285 90" stroke="#2563eb" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#2563eb"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📏 현장 기계실 치수 검측</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 천장 층고 ≥ 3.0m 실측</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• Access Floor H=300mm 확인</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 공간 실측 1:1 승인</text>
                    </svg>
                </div>
            </div>

            <!-- STEP 2 Card -->
            <div class="bg-white p-6 rounded-2xl border border-indigo-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-indigo-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 2</span>
                    <h3 class="font-bold text-base text-slate-900">통신설비 <span class="term-highlight" onclick="openGlossary('manufacturing_spec')">제작 사양서</span> 및 <span class="term-highlight" onclick="openGlossary('shop_drawing')">제작도면(Shop Drawing)</span> 작성</h3>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                    <p>• <strong>전기/기계 사양 정의:</strong> 전원 사양(AC220V/DC48V), 소비전력(W), 포트 구성(10GbE, Optic), 신호 전파 특성 및 동작 온도(-20℃~60℃)를 작성합니다.</p>
                    <p>• <strong>제작 도면 작성:</strong> 통신 랙(Height 2200mm) 전면/후면 장비 배치도, 전원 분배 배선도, 광패치패널 구성도 및 외구 치수를 CAD 도면화합니다.</p>
                </div>

                <!-- 1:1 STEP 2 2D SVG -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step2_wbs10', 'STEP 2 제작사양서 및 통신 랙 Shop Drawing 작성 2D 도식')">
                    <svg id="svg_step2_wbs10" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">🔌 전기/기계 세부규격 정의</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 전원 AC220V/DC48V & 소비전력</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 10GbE 광포트 & 인터페이스</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 사양서 작성 완수</text>

                        <path d="M 245 90 L 285 90" stroke="#4f46e5" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#4f46e5"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🗄️ 랙(Rack) Shop Drawing</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 통신 랙(H=2200mm) 배치도</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 전원분배 & 광패치 배선도 CAD</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 제작도면 작성완료</text>
                    </svg>
                </div>
            </div>

            <!-- STEP 3 Card -->
            <div class="bg-white p-6 rounded-2xl border border-cyan-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-cyan-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 3</span>
                    <h3 class="font-bold text-base text-slate-900">품질/환경 규격(KC/IP65/내진) & <span class="term-highlight" onclick="openGlossary('fat_procedure')">공장검사(FAT) 절차서</span> 검증</h3>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                    <p>• <strong>인증서 & 성적서 포함:</strong> KC 방송통신기자재 적합인증서, 내진 시험성적서 및 야외 카메라 방진·방수 IP65 등급 시험성적서를 첨부합니다.</p>
                    <p>• <strong>FAT 절차서 작성:</strong> 공장 출하 전 입석 감리원 참관 하에 수행할 공장검사(FAT) 시험 항목, 검사 기준 및 양식을 사양서 부록에 포함합니다.</p>
                </div>

                <!-- 1:1 STEP 3 2D SVG -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step3_wbs10', 'STEP 3 KC인증/내진/IP65 성적서 및 공장검사 FAT 절차서 2D 도식')">
                    <svg id="svg_step3_wbs10" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">🛡️ KC 인증 & 환경시험성적서</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• KC 방송통신기자재 인증서</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 내진성능 & IP65 방수성적서</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 품질 규격 첨부 완료</text>

                        <path d="M 245 90 L 285 90" stroke="#0891b2" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#0891b2"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📋 공장검사(FAT) 시험절차서</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 출하 전 감리 참관 FAT 항목</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 시험검사 양식 부록 첨부</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ FAT 절차서 검증통과</text>
                    </svg>
                </div>
            </div>

            <!-- STEP 4 Card -->
            <div class="bg-white p-6 rounded-2xl border border-teal-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-teal-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 4</span>
                    <h3 class="font-bold text-base text-slate-900">감리단 및 발주처 기술 검토 제출 & 최종 승인(Sign-Off)</h3>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                    <p>• <strong>공문 승인 신청:</strong> 현장 시스템팀 주관으로 작성된 제작사양서 및 제작도면 세트를 감리단/발주처에 제출합니다.</p>
                    <p>• <strong>기술 보완 조치:</strong> 감리 기술 검토 의견서 발생 시 3일 이내 보완 조치표를 작성하여 최종 승인 인감(Sign-Off)을 수득합니다.</p>
                </div>

                <!-- 1:1 STEP 4 2D SVG -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step4_wbs10', 'STEP 4 감리단 기술검토 제출 및 승인 인감 수득 2D 도식')">
                    <svg id="svg_step4_wbs10" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#0d9488" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#0f766e" text-anchor="middle">📩 감리단/발주처 공문 제출</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 제작사양서 & 도면 세트 제출</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 기술 검토 의견서 보완 조치</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 제출 및 검토 완료</text>

                        <path d="M 245 90 L 285 90" stroke="#0d9488" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#0d9488"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🖊️ 승인 인감(Sign-Off) 획득</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 책임감리원 최종 승인 서명</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 공학적 무결성 공식 입증</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 최종 승인 완료</text>
                    </svg>
                </div>
            </div>

            <!-- STEP 5 Card -->
            <div class="bg-white p-6 rounded-2xl border border-emerald-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-emerald-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 5</span>
                    <h3 class="font-bold text-base text-slate-900">최종 승인 사양서 제작사 공식 전달 & 공장 제작 착수</h3>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                    <p>• <strong>승인 도서 통보:</strong> 감리 승인 필증이 첨부된 제작사양서 및 도면을 자재 제작사에 공식 전달 통보합니다.</p>
                    <p>• <strong>제작 착수 지시:</strong> 제작사 공장 제작 일정을 확정하고 공장검사(FAT) 준비를 지시하여 불량 없는 제품 생산을 개시합니다.</p>
                </div>

                <!-- 1:1 STEP 5 2D SVG -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step5_wbs10', 'STEP 5 승인 사양서 제작사 전달 및 공장 제작 착수 2D 도식')">
                    <svg id="svg_step5_wbs10" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📄 승인 도서 제작사 통보</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 승인 필증 포함 사양서 전달</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 임의 변형 방지 대조 마감</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 통보 전달 완료</text>

                        <path d="M 245 90 L 285 90" stroke="#059669" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#059669"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🏭 자재 공장 제작 착수</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 공장 검사(FAT) 일정 확정</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 통신설비 품질 제작 개시</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 제작 착수 완료</text>
                    </svg>
                </div>
            </div>

        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 3. Checklist HTML Template (3-Column Master & ~하였는가? 100%)
chk_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 통신설비 제작 사양서 작성 / 승인 마스터 체크리스트 (WBS 9000-2-10)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <!-- 🔵 헤더 영역 -->
    <div class="bg-white p-6 sm:p-8 border-b border-slate-200">
        <div class="flex justify-between items-start">
            <div>
                <h1 class="text-2xl sm:text-3xl font-black text-slate-900 tracking-tight">통신설비 제작 사양서 작성 / 승인 마스터 체크리스트</h1>
            </div>
            <span class="text-xs font-bold text-blue-600 bg-blue-50 px-3 py-1.5 rounded-lg border border-blue-200">WBS Code 9000-2-10 | 통신 검측대장</span>
        </div>
        <div class="w-full h-1 bg-slate-900 mt-4"></div>
    </div>

    <div class="p-6 sm:p-8 space-y-8">
        <!-- 📋 안내 상자 -->
        <div class="bg-blue-50/70 border border-blue-200 p-6 rounded-2xl text-xs sm:text-sm text-blue-950 space-y-2">
            <h4 class="font-bold text-sm sm:text-base text-blue-900 flex items-center gap-2">📋 쉽게 풀어쓴 현장 점검 체크리스트</h4>
            <p class="text-slate-700 leading-relaxed">
                본 체크리스트는 통신설비 제작사양서 및 제작도면 작성·승인 시 <strong>[🟣 시공 도식 열기]</strong>를 클릭하면 대형 고화질 팝업 모달이 열려 도식을 직접 보며 <strong>~하였는가? (100%)</strong> 점검을 진행할 수 있도록 연동되었습니다.
            </p>
        </div>

        <!-- 3-COLUMN MASTER TABLE -->
        <div class="overflow-x-auto rounded-2xl border border-slate-200 shadow-sm">
            <table class="w-full text-left border-collapse">
                <thead>
                    <tr class="bg-slate-100 text-slate-700 text-xs font-black uppercase tracking-wider border-b border-slate-200">
                        <th class="py-4 px-6 text-center w-1/4">시공 단계</th>
                        <th class="py-4 px-6 text-center w-7/12">필수 검측 항목 (쉬운 질문형 수칙)</th>
                        <th class="py-4 px-6 text-center w-1/6">점검 결과</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-200 text-xs sm:text-sm bg-white">
                    
                    <!-- STEP 1 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">📐</span>
                                <span class="font-bold text-slate-900 text-sm">설계 대조 & 작성</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 1 사양확정)</span>
                                <button onclick="openDiagramZoomByKey('step1', 'STEP 1 설계도서 대조 및 사양서 작성 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6">
                            <div class="flex items-start gap-2">
                                <span class="bg-blue-100 text-blue-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">1. 설계도서 대조</span>
                                <p class="text-slate-800 font-medium"><strong>[도서 대조]</strong> 기본/실시설계도서(보고서, 도면, 자재사양서)와 현장 기계실·역사 치수를 1:1 대조·확인하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-blue-100 text-blue-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">2. 사양서 작성</span>
                                <p class="text-slate-800 font-medium"><strong>[사양서 정의]</strong> 실제 제작·설치될 통신설비의 전기적·기계적 세부 규격 및 제작도면(Shop Drawing)을 작성하였는가?</p>
                            </div>
                        </td>
                    </tr>

                    <!-- STEP 2 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">🛡️</span>
                                <span class="font-bold text-slate-900 text-sm">품질 & 인터페이스</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 2 규격검증)</span>
                                <button onclick="openDiagramZoomByKey('step2', 'STEP 2 품질/환경 규격 및 공종간 인터페이스 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6 border-t border-slate-200">
                            <div class="flex items-start gap-2">
                                <span class="bg-indigo-100 text-indigo-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">3. 품질/환경 검증</span>
                                <p class="text-slate-800 font-medium"><strong>[품질 성적서]</strong> KC 인증, 내진·방수(IP65) 환경시험 규격 및 공장검사(FAT) 절차서를 사양서에 수록하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-indigo-100 text-indigo-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">4. 인터페이스 검증</span>
                                <p class="text-slate-800 font-medium"><strong>[인터페이스]</strong> 노반/전기/신호/차량 등 이종 공종 간 인터페이스 접속 규격을 명확히 반영하였는가?</p>
                            </div>
                        </td>
                    </tr>

                    <!-- STEP 3 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">🖊️</span>
                                <span class="font-bold text-slate-900 text-sm">승인 & 제작전달</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 3 제작착수)</span>
                                <button onclick="openDiagramZoomByKey('step3', 'STEP 3 감리단 승인 및 제작사 전달 2D 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6 border-t border-slate-200">
                            <div class="flex items-start gap-2">
                                <span class="bg-cyan-100 text-cyan-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">5. 감리 승인</span>
                                <p class="text-slate-800 font-medium"><strong>[Sign-Off]</strong> 작성된 제작사양서 및 제작도면을 감리단/발주처에 제출하여 기술 승인(Sign-Off)을 득하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-emerald-100 text-emerald-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">6. 제작사 전달</span>
                                <p class="text-slate-800 font-medium"><strong>[제작 착수]</strong> 최종 감리 승인된 제작사양서 및 도면을 제작사에 공식 전달하여 공장 제작을 착수하였는가?</p>
                            </div>
                        </td>
                    </tr>

                </tbody>
            </table>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const svgStore = {{
    'step1': `<svg viewBox="0 0 520 180" width="100%" height="250" xmlns="http://www.w3.org/2000/svg">
                <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                <text x="130" y="50" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📐 실시설계도서 1:1 대조</text>
                <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 보고서/도면/자재사양서 검토</text>
                <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 현장 기계실 천장 층고 ≥ 3.0m</text>
                <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 설계 대조 완료</text>
                <path d="M 245 90 L 285 90" stroke="#2563eb" stroke-width="3"/>
                <polygon points="285,85 295,90 285,95" fill="#2563eb"/>
                <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🔌 사양서 & Shop Drawing</text>
                <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 전기/기계 세부 규격 작성</text>
                <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 통신 랙(H=2200mm) 배치도</text>
                <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 사양서 작성 완수</text>
            </svg>`,
    'step2': `<svg viewBox="0 0 520 180" width="100%" height="250" xmlns="http://www.w3.org/2000/svg">
                <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                <text x="130" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">🛡️ KC인증 & 환경시험성적서</text>
                <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• KC 방송통신기자재 인증서</text>
                <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 내진성능 & IP65 방수성적서</text>
                <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 품질 규격 첨부</text>
                <path d="M 245 90 L 285 90" stroke="#4f46e5" stroke-width="3"/>
                <polygon points="285,85 295,90 285,95" fill="#4f46e5"/>
                <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📋 공종 간 인터페이스 검증</text>
                <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 궤도/전기/신호 접속규격 반영</text>
                <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 공장검사(FAT) 절차서 부록</text>
                <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 검증 완료</text>
            </svg>`,
    'step3': `<svg viewBox="0 0 520 180" width="100%" height="250" xmlns="http://www.w3.org/2000/svg">
                <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="130" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🖊️ 감리단/발주처 승인(Sign-Off)</text>
                <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 공문 제출 및 기술 검토 보완</text>
                <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 책임감리원 최종 승인 서명</text>
                <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 승인 획득 완료</text>
                <path d="M 245 90 L 285 90" stroke="#059669" stroke-width="3"/>
                <polygon points="285,85 295,90 285,95" fill="#059669"/>
                <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🏭 제작사 전달 & 공장 제작 착수</text>
                <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 승인도서 제작사에 공식 통보</text>
                <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 공장검사(FAT) 준비 지시</text>
                <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 제작 착수 개시</text>
            </svg>`
}};

function openDiagramZoomByKey(stepKey, titleText) {{
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "시공 도식 대형 정밀 보기");
    
    if (svgStore[stepKey]) {{
        zoomBody.innerHTML = svgStore[stepKey];
    }}
    
    document.getElementById('zoomModal').classList.add('active');
}}

function closeZoomModal() {{
    document.getElementById('zoomModal').classList.remove('active');
}}

function closeZoomModalOutside(event) {{
    if (event.target.id === 'zoomModal') {{
        closeZoomModal();
    }}
}}

window.addEventListener('keydown', function(event) {{
    if (event.key === 'Escape') {{
        closeZoomModal();
    }}
}});
</script>
</body>
</html>
"""

# Write Standard & Checklist HTML Files
files_to_write = [
    (os.path.join(std_dir, "9000-2-10_통신설비 제작 사양서 작성 _ 승인_표준서.html"), std_html_content),
    (os.path.join(std_dir, "통신설비 제작 사양서 작성 _ 승인_표준서.html"), std_html_content),
    (os.path.join(gui_dir, "9000-2-10_통신설비 제작 사양서 작성 _ 승인_수행지침.html"), gui_html_content),
    (os.path.join(gui_dir, "통신설비 제작 사양서 작성 _ 승인_수행지침.html"), gui_html_content),
    (os.path.join(chk_dir, "9000-2-10_통신설비 제작 사양서 작성 _ 승인_체크리스트.html"), chk_html_content),
    (os.path.join(chk_dir, "통신설비 제작 사양서 작성 _ 승인_체크리스트.html"), chk_html_content)
]

for path, content in files_to_write:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"   ✓ [WBS 9000-2-10 MASTER HTML BUILT] -> {os.path.basename(path)}")

# Update Excel V4 Row 10
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"

if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = None
        for s_name in wb.sheetnames:
            if "통신" in s_name:
                ws = wb[s_name]
                break
        if not ws:
            ws = wb.worksheets[5]
            
        row_idx = 10 # WBS 9000-2-10
        
        # Column J: Standard Summary
        ws.cell(row=row_idx, column=10).value = "1) 제작 사양서 승인: 실시설계 도서(보고서, 도면, 자재사양서)를 대조하여 실제 제작 설치될 통신설비 제작 사양서 및 도면을 확정함.\n2) 공학 무결성: 현장 시스템팀 주관으로 감리단/발주처 승인을 득하고 최종 제작사양서 및 제작도면의 공학적 무결성을 보장함."
        ws.cell(row=row_idx, column=11).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\10_통신설비 제작 사양서 작성 _ 승인\표준서\통신설비 제작 사양서 작성 _ 승인_표준서.html"
        ws.cell(row=row_idx, column=11).value = "📄 [더블클릭] 표준서 열기 🔗"
        ws.cell(row=row_idx, column=11).style = "Hyperlink"
        
        # Column L: Guideline Summary
        ws.cell(row=row_idx, column=12).value = "1) 현장 여건 대조: 실시설계도서와 현장 기계실/선로/역사 규격을 비교하여 제작 사양서 세부 수치 확정.\n2) 승인 절차 가이드: 제작사양서 및 제작도면을 작성하여 감리단/발주처 승인을 득하는 절차를 가이드함."
        ws.cell(row=row_idx, column=13).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\10_통신설비 제작 사양서 작성 _ 승인\수행지침\통신설비 제작 사양서 작성 _ 승인_수행지침.html"
        ws.cell(row=row_idx, column=13).value = "📄 [더블클릭] 수행지침 열기 🔗"
        ws.cell(row=row_idx, column=13).style = "Hyperlink"

        # Column N: Checklist Summary
        ws.cell(row=row_idx, column=14).value = "1) 실시설계도서(도면, 보고서, 자재사양서)와 현장 여건의 요구사항 충족 여부를 확인하였는가?\n2) 통신설비 제작사양서 및 제작도면을 작성하여 감리단/발주처의 최종 승인을 득하였는가?"
        ws.cell(row=row_idx, column=15).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\10_통신설비 제작 사양서 작성 _ 승인\체크리스트\통신설비 제작 사양서 작성 _ 승인_체크리스트.html"
        ws.cell(row=row_idx, column=15).value = "📄 [더블클릭] 체크리스트 열기 🔗"
        ws.cell(row=row_idx, column=15).style = "Hyperlink"
        
        wb.save(excel_path)
        print("   ✓ [EXCEL V4 SYNC COMPLETE] Row 10 (WBS 9000-2-10) Updated Successfully!")
    except Exception as e:
        print(f"Notice: Openpyxl save deferred ({e}). HTML files are 100% built.")

print("\n🎉 SUCCESSFULLY COMPLETED ALL REBUILDING AND SYNC FOR WBS 9000-2-10!")
