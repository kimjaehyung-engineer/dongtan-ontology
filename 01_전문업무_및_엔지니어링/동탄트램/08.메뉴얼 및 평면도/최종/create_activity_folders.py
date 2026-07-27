import openpyxl
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

def sanitize_folder_name(name):
    # Replace slashes, colons, invalid chars with _
    name = re.sub(r'[\/\\:\*\?"<>\|]', '_', name)
    # Clean up whitespace
    name = name.strip()
    return name

wb = openpyxl.load_workbook(excel_path, data_only=True)

target_sheets = ['사전토공사', '상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

created_summary = []

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found in workbook.")
        continue
    
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    
    # Find header row
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        if any(isinstance(val, str) and ('코드' in val or '작업단위' in val) for val in r if val):
            header_row_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
    
    l4_idx = None
    act_idx = None
    
    for i, h in enumerate(headers):
        if 'L4' in h:
            l4_idx = i
        elif '작업단위' in h or 'Activity' in h:
            act_idx = i
            
    discipline_dir = os.path.join(base_attach_dir, sheet_name)
    os.makedirs(discipline_dir, exist_ok=True)
    
    act_count = 0
    folder_count = 0
    
    for r in rows[header_row_idx+1:]:
        if all(v is None for v in r):
            continue
        l4_val = r[l4_idx] if l4_idx is not None and l4_idx < len(r) else None
        act_val = r[act_idx] if act_idx is not None and act_idx < len(r) else None
        
        if l4_val or act_val:
            act_count += 1
            act_str = str(act_val).strip() if act_val else f"Task_{act_count}"
            sanitized_act = sanitize_folder_name(act_str)
            folder_name = f"{act_count}_{sanitized_act}"
            
            act_folder_path = os.path.join(discipline_dir, folder_name)
            os.makedirs(act_folder_path, exist_ok=True)
            
            # Create subfolders
            for sub in ['표준서', '수행지침', '체크리스트']:
                sub_path = os.path.join(act_folder_path, sub)
                os.makedirs(sub_path, exist_ok=True)
                folder_count += 1

    msg = f"Sheet [{sheet_name}]: Created/Verified {act_count} activity folders and subfolders in '{discipline_dir}'"
    print(msg)
    created_summary.append(msg)

print("\n--- Folder Creation Complete ---")
