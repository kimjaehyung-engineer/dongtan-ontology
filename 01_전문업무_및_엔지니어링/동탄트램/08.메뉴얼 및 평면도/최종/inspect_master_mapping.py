import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

master_ws = wb['공정표에따른 매뉴얼']
print("Master sheet rows count:", master_ws.max_row)

header_row = 3
headers = [str(master_ws.cell(row=header_row, column=c).value or "").strip().replace('\n', ' ') for c in range(1, master_ws.max_column + 1)]

print(f"Master Sheet Headers (Row {header_row}):")
for i, h in enumerate(headers, 1):
    print(f"  Col {i}: '{h}'")

print("\nSample Rows (4~10):")
for r in range(4, 11):
    l3_name = master_ws.cell(row=r, column=3).value # Col C
    l4_code = master_ws.cell(row=r, column=4).value # Col D
    act_name = master_ws.cell(row=r, column=5).value # Col E
    print(f"Row {r:2d}: C(L3)='{l3_name}' | D(L4)='{l4_code}' | E(Act)='{act_name}'")
