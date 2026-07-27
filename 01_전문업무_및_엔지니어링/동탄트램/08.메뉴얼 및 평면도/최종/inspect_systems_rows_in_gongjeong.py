import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Source v1 excel path
v1_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)_최종공정매뉴얼완성본.xlsx"
wb = openpyxl.load_workbook(v1_excel_path)
ws = wb['공정매뉴얼']

header_row = 3

print(f"=== Inspecting Systems (신호, 통신, 전기) Rows in '공정매뉴얼' ===")

systems_rows = []

for r in range(header_row + 1, ws.max_row + 1):
    c_l3 = str(ws.cell(row=r, column=3).value or "").strip()
    d_l4 = str(ws.cell(row=r, column=4).value or "").strip()
    e_act = str(ws.cell(row=r, column=5).value or "").strip()
    p_disc = str(ws.cell(row=r, column=16).value or "").strip()
    
    text_combine = (c_l3 + " " + d_l4 + " " + e_act + " " + p_disc).lower()
    
    if any(k in text_combine for k in ['신호', '통신', '전기', '전차선', '급전', '변전소', 'lte', 'cctv', 'cbi', 'psd', '축차계수기', '광케이블']):
        systems_rows.append((r, c_l3, d_l4, e_act, p_disc))

print(f"Total Systems Rows Identified: {len(systems_rows)}")
for r, c_l3, d_l4, e_act, p_disc in systems_rows:
    print(f"  Row {r:2d} [Disc: {p_disc:10s}]: L4='{d_l4}' | L3='{c_l3}' | Act='{e_act}'")
