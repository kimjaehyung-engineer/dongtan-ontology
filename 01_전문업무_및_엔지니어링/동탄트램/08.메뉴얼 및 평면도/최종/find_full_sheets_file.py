import openpyxl
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

print("Checking sheets in all Excel files in destination folder:")

full_sheets_file = None

for f in os.listdir(base_dir):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        f_path = os.path.join(base_dir, f)
        try:
            wb = openpyxl.load_workbook(f_path, read_only=True)
            print(f" 📄 File: {f} | Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
            if len(wb.sheetnames) > 3 and not full_sheets_file:
                full_sheets_file = f_path
            wb.close()
        except Exception as e:
            print(f" 📄 File: {f} | Error: {e}")

print(f"\nBest Full Master File found: '{full_sheets_file}'")
