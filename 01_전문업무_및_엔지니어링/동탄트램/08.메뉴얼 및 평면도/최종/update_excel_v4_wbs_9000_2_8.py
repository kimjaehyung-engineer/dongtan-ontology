import os
import shutil
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

artifact_dir = r"C:\Users\sskjh\.gemini\antigravity\brain\887aacfa-3165-4be1-8e89-29f90e47a298"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토"

std_dir = os.path.join(base_dir, "표준서")
gui_dir = os.path.join(base_dir, "수행지침")
chk_dir = os.path.join(base_dir, "체크리스트")

for d in [std_dir, gui_dir, chk_dir]:
    os.makedirs(d, exist_ok=True)

# Copy image to gui_dir
orig_img = "fusion_splicer_calib_1785118781604.jpg"
target_img = "fusion_splicer_calib.jpg"

src_p = os.path.join(artifact_dir, orig_img)
dst_p = os.path.join(gui_dir, target_img)
if os.path.exists(src_p):
    shutil.copy(src_p, dst_p)
    print(f"   ✓ [IMAGE COPIED] {target_img} -> {gui_dir}")

# Update Excel V4 Row 8
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    
    # Find sheet for 통신분야
    target_sheet = None
    for s_name in wb.sheetnames:
        if "통신" in s_name:
            target_sheet = wb[s_name]
            break
            
    if target_sheet:
        row_idx = 8 # WBS 9000-2-8
        
        # Column J: Standard Summary
        target_sheet.cell(row=row_idx, column=10).value = "1) 투입 자원 사전 검토: 통신 공사 투입 인력, 광융착기/OTDR 측정장비, 자재 수급 계획 등 타당성/적합성 검토함\n2) 적합성 확보: 시스템업체/협력업체 및 감리단 주관으로 공정별 인력 및 장비 투입 제출서의 적합성을 최종 승인함"
        target_sheet.cell(row=row_idx, column=11).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\표준서\자재 인력 장비 등 투입 사전 검토_표준서.html"
        target_sheet.cell(row=row_idx, column=11).value = "📄 [더블클릭] 표준서 열기 🔗"
        target_sheet.cell(row=row_idx, column=11).style = "Hyperlink"
        
        # Column L: Guideline Summary
        target_sheet.cell(row=row_idx, column=12).value = "1) 자원 투입 수급: 공정별 숙련 통신공 투입, 광융착기/OTDR 시험 장비 검교정 상태 확인\n2) 안전/민원 대책: 현장 자재 야적장 확보, 도로 굴착시 교통통제 및 민원 대장 대책을 종합 검토함"
        target_sheet.cell(row=row_idx, column=13).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\수행지침\자재 인력 장비 등 투입 사전 검토_수행지침.html"
        target_sheet.cell(row=row_idx, column=13).value = "📄 [더블클릭] 수행지침 열기 🔗"
        target_sheet.cell(row=row_idx, column=13).style = "Hyperlink"

        # Column N: Checklist Summary
        target_sheet.cell(row=row_idx, column=14).value = "1) 공정별 숙련 인력 투입 계획 및 정밀 측정 장비(OTDR, 융착기) 검교정 상태를 확인하였는가?\n2) 자재 수급 계획, 야적장 확보 및 민원 대책을 포함한 투입 계획서를 작성하였는가?"
        target_sheet.cell(row=row_idx, column=15).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\체크리스트\자재 인력 장비 등 투입 사전 검토_체크리스트.html"
        target_sheet.cell(row=row_idx, column=15).value = "📄 [더블클릭] 체크리스트 열기 🔗"
        target_sheet.cell(row=row_idx, column=15).style = "Hyperlink"
        
        wb.save(excel_path)
        print("   ✓ [EXCEL V4 SYNC COMPLETE] Row 8 (WBS 9000-2-8) Updated Successfully!")

print("\n🎉 SUCCESSFULLY SYNCED EXCEL V4 FOR WBS 9000-2-8!")
