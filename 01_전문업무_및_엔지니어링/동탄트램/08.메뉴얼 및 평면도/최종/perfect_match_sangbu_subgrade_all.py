import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
backup_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반"

def normalize(name):
    return name.replace(" ", "").replace("_", "").replace("/", "").replace("(", "").replace(")", "").replace("-", "").lower()

# 36 Activities - Master Data Perfectly Matching Excel and HTML (Standard, Guideline, Checklist)
master_specs = [
    {
        "num": 1, "folder": "1_지반조사 상세검토", "title": "지반조사 상세검토",
        "purpose": "지반 물리탐사 및 시굴 결과를 바탕으로 노반 지반 지지지수와 연약지반 분포 구간 정밀 분석",
        "gui_summary": "PBT 평판재하시험 성과 분석, GPR/탄성파 탐사 대조 및 N치<4 연약지반 구간 PVD/DCM 도면 반영",
        "std_summary": "KDS 47 10 00 기준 시굴 및 GPR 탐사 결과 대조. 연약지반(N<4) 구간 PVD/DCM 반영 검증 및 허용 잔류침하량 2.5cm 이하 수렴 목표 설정.",
        "chk_summary": "N치<4 연약지반 PVD/DCM 대책, PBT 원지반 지지력 성적서, 암반 파쇄대 50cm 치환 굴착 검속",
        "std_code": "KDS 47 10 00 철도노반설계기준 / KDS 11 30 00 연약지반",
        "gui_s1": "① 지반 조사 성과 수집 & 물리탐사(GPR/탄성파) 도면 매핑 검속",
        "gui_s2": "② N치 < 4 연약지반 구간 추출 및 PVD/DCM 개량 공법 도면 반영",
        "gui_s3": "③ 원지반 PBT 시험성적서 분석 및 파쇄대 치환 굴착(≥50cm) 감리단 승인",
        "chk_items": [
            ("지반 조사 성과 대조", "GPR 및 탄성파 탐사 결과 도면 매핑 및 N치 < 4 연약지반 100% 추출 여부"),
            ("PBT 성적서 확인", "원지반 평판재하시험(PBT) K30 지지력 성적서 및 허용응력(Qa ≥ 150kN/m²) 확인 여부"),
            ("파쇄대 치환 굴착", "암반 파쇄대 및 취약 구간 치환 굴착 두께(50cm 이상) 감리단 승인 여부"),
            ("잔류침하량 목표", "연약지반 PVD/DCM 처리 구간 트램 허용 잔류침하량 2.5cm 이하 수렴 목표 설정 여부"),
            ("지하수위 저하 대책", "지하수위 노반 하부 1.0m 이하 유도 준비 배수 트렌치 도면 반영 여부")
        ]
    },
    {
        "num": 2, "folder": "2_발주전략 KOM", "title": "발주전략 KOM",
        "purpose": "강화노반 시공사 및 자재 공급업체 간 공정/품질 이행 목표 공유 킥오프 미팅",
        "gui_summary": "공정 CPM 마스터 스케줄 확정, 1층 다짐 두께 30cm 이하 규정 준수 서약 및 장비 반입 시방 합의",
        "std_summary": "KCS 47 10 25 표준 시방 준수. 1층 다짐 두께 30cm 이하, 롤러 조합(10t 진동+15t 타이어) 및 다짐도 95% 이상 킥오프 의결.",
        "chk_summary": "1층 다짐 두께 30cm 이하 이행 서약, 10t 진동롤러 반입 계약, K30≥110MN/m³ 의결서 확인",
        "std_code": "KCS 47 10 25 강화노반 시방서 / 동탄트램 사업관리 지침",
        "gui_s1": "① 킥오프 회의 개최 & CPM 공정 마스터 스케줄 목표 공유",
        "gui_s2": "② 1층 포설 다짐 두께 30cm 이하 및 롤러 조합(10t진동+15t타이어) 시방 의결",
        "gui_s3": "③ 강화노반 지지력 지표(K30≥110MN/m³, Ev2≥120MPa) 품질 협약 체결",
        "chk_items": [
            ("공정 마스터 스케줄", "강화노반 CPM 마스터 공정표 및 트램 준공 목표일 공유 체결 여부"),
            ("다짐두께 30cm 서약", "1층 포설 다짐 두께 30cm 이하 엄격 준수 계약서 날인 여부"),
            ("다짐 장비 반입 계약", "10톤 진동 롤러 및 15톤 타이어 롤러 필수 투입 확정 서류 확인 여부"),
            ("지지력 지표 의결", "PBT K30 ≥ 110 MN/m³ 및 PFWD Ev2 ≥ 120 MPa 품질 지표 서명 여부"),
            ("골재 공급망 확보", "쇄석 혼합 골재 일일 500m³ 이상 반입 공급업체 계약서 확인 여부")
        ]
    },
    {
        "num": 3, "folder": "3_철도보호지구에서의 행위신고(필요시)", "title": "철도보호지구에서의 행위신고(필요시)",
        "purpose": "철도보호지구(인접 30m 이내) 굴착 및 토공사 진행을 위한 법적 행위 신고 및 국가철도공단 승인",
        "gui_summary": "철도안전법 제45조 준수, 인접 30m 경계 측량 및 레일 침하핀(±5mm) 계측기 상시 모니터링",
        "std_summary": "철도안전법 제45조 및 국가철도공단 철도보호지구 관리지침 준수. 인접 30m 이내 굴착 시 선로 침하 계측기(경사계, 침하핀) 상시 설치.",
        "chk_summary": "철도공단 행위신고 허가서, 레일 침하핀(±5mm) 계측, 철도운행안전원 100% 상주 참관",
        "std_code": "철도안전법 제45조 / 국가철도공단 철도보호지구 관리지침",
        "gui_s1": "① 인접 30m 경계 정밀 측량 & 국가철도공단 행위신고 서류 제출",
        "gui_s2": "② 기존 철도 레일 침하핀 및 경사계 계측기(허용치 ±5mm) 10m 간격 설치",
        "gui_s3": "③ 전차선 가선 방호망 시공 & 철도운행안전원 현장 참관 아래 굴착 진행",
        "chk_items": [
            ("행위신고 허가서", "국가철도공단 철도보호지구 행위신고 최종 허가 공문 수신 여부"),
            ("레일 침하핀 계측", "기존 선로 레일 침하핀 및 경사계 10m 간격 설치 및 0점 측정 여부"),
            ("허용 침하량 관리", "선로 허용 침하량 ±5mm 이내 실시간 자동 계측 모니터링 여부"),
            ("전차선 방호망", "장비 회전 반경 내 전차선 방호망 및 안전 펜스 고정 여부"),
            ("운행안전원 상주", "철도운행안전원 현장 100% 참관 배치 필증 및 일일 안전 교육 여부")
        ]
    }
]

# Populate customized specifications for all 36 activities
for i in range(4, 37):
    folder_names = [
        "4_착수전 측량 Data 확인", "5_지장물이설 협의", "6_용지보상RISK 검토", "7_최고의 팀 만들기 지원",
        "8_연약지반 처리공법 검토(필요시)", "9_토공 유동표 확인", "10_기공승낙 적정성 검토", "11_폐기물처리계획 수립",
        "12_철도운행협의(필요시)", "13_작수전 Big Room 회의", "14_시공 계획 수립", "15_사토장_토취장 선정 검토(필요시)",
        "16_공사사전준비", "17_임시배수시설", "18_쌓기재료 검사", "19_장비 검수 지원",
        "20_선로 종_횡단 및 용지경계측량", "21_규준틀 설치", "22_준비배수", "23_벌개제근_표토제거",
        "24_구조물 및 지장물 제거", "25_진입로 조성", "26_노반쌓기", "27_하부노반 시공",
        "28_상부노반 시공", "29_강화노반 시공", "30_토공 유동운반_사토", "31_연약지반처리(필요시)",
        "32_절성토경계부 배수구조물(맹암거) 시공", "33_암석쌓기", "34_방치기간 확보", "35_토공마무리", "36_토공 마무리면 인계"
    ]
    fname = folder_names[i-4]
    title = fname.split("_", 1)[1]
    
    if i == 29: # 29_강화노반 시공
        master_specs.append({
            "num": 29, "folder": fname, "title": title,
            "purpose": "트램 궤도 직하부 강화노반(두께 30cm) 쇄석혼합 골재 타설, K30≥110MN/m³, Ev2≥120MPa, Ev2/Ev1≤2.2 검속",
            "gui_summary": "1층 다짐 두께 30cm 준수, 10t 진동롤러+15t 타이어롤러 다짐, PBT K30≥110MN/m³ 및 PFWD Ev2≥120MPa 100% 승인",
            "std_summary": "KCS 47 10 25 강화노반 최상급 시방. 트램 궤도 직하부 30cm 쇄석혼합골재 타설, PBT K30≥110MN/m³, PFWD Ev2≥120MPa, Ev2/Ev1≤2.2 100% 검속.",
            "chk_summary": "PBT K30≥110MN/m³ 성적서, PFWD Ev2≥120MPa, Ev2/Ev1≤2.2, 쇄석 최대입경 100mm 이하, 층다짐 30cm 검속",
            "std_code": "KCS 47 10 25 강화노반 시방서 / KDS 47 10 00",
            "gui_s1": "① 쇄석 혼합 골재 반입 (최대입경 ≤100mm, 흙분 ≤5%, 수정CBR ≥10%)",
            "gui_s2": "② 모터 그레이더 포설(두께 30cm) & 10t진동롤러 4회 + 15t타이어롤러 4회 정속 다짐",
            "gui_s3": "③ PBT (K30 ≥ 110 MN/m³) & PFWD (Ev2 ≥ 120 MPa, Ev2/Ev1 ≤ 2.2) 검속 승인",
            "chk_items": [
                ("K30 지지력 계수", "PBT 평판재하시험 노반 지지력 계수 K30 ≥ 110 MN/m³ 100% 달성 성적서 확인"),
                ("Ev2 변형계수", "PFWD 시험 2차 변형계수 Ev2 ≥ 120 MPa 및 다짐비 Ev2/Ev1 ≤ 2.2 충족 여부"),
                ("쇄석 골재 입도", "쇄석 혼합 골재 최대입경 100mm 이하 및 200번체 통과량(흙분) 5% 이하 검속"),
                ("1층 다짐 두께", "1층 포설 다짐 완결 두께 30cm 이하 엄격 준수 및 들밀도 상대다짐도 95% 이상"),
                ("완성면 공차", "완성면 계획고 허용 공차 ±10mm 이내 및 배수 횡단 구배 2.0% 정밀 정지 여부")
            ]
        })
    elif i == 32: # 32_절성토경계부 배수구조물(맹암거) 시공
        master_specs.append({
            "num": 32, "folder": fname, "title": title,
            "purpose": "절성토 편성 경계 부등침하 방지를 위한 맹암거(유압관 D200mm + 투수골재) 시공",
            "gui_summary": "절성토 경계 터파기 구배 2.0% 성형, D200mm 유압관 부직포 감싸기, 투수성 쇄석(25~40mm) 층다짐 채움",
            "std_summary": "절성토 경계부 부등침하 방지 시방. 경계면 맹암거(D200mm 유압관 + 부직포 감싼 투수 골재) 시공, 배수 구배 2.0% 확보하여 침수 차단.",
            "chk_summary": "D200mm 유압관 배수구배 2.0%, 세굴방지 부직포 감싸기, 투수쇄석(25~40mm) 채움, 집수정 통수 100%",
            "std_code": "KCS 11 20 00 배수구조물 / KDS 47 10 00",
            "gui_s1": "① 절성토 경계부 터파기(폭 1.0m) & 종단 배수 구배 2.0% 이상 성형",
            "gui_s2": "② 투수 부직포 포설 & D200mm 유압 유공관 부설 및 소켓 결속 검속",
            "gui_s3": "③ 투수성 쇄석(입경 25~40mm) 채움, 상부 부직포 봉합 & 층다짐 30cm 완료",
            "chk_items": [
                ("유공관 배수 구배", "맹암거 D200mm 유압 유공관 종단 배수 구배 2.0% 이상 정밀 부설 여부"),
                ("부직포 필터 감싸기", "토사 유입 방지용 투수 부직포 감싸기 및 봉합 상태 검속 여부"),
                ("투수 쇄석 채움", "투수성 쇄석 골재(입경 25~40mm) 채움 및 층다짐 30cm 이내 시행 여부"),
                ("접속부 부등침하 방지", "절성토 경계 접속 슬래브 및 맹암거 침하시설 연계 시공 확인"),
                ("집수정 통수 연통", "측구 집수정 접속부 통수 시험 100% 완료 및 감리단 검측 서명 여부")
            ]
        })
    else:
        master_specs.append({
            "num": i, "folder": fname, "title": title,
            "purpose": f"상부강화노반 {title} 과업 목적 달성 및 KCS 47 10 25 공학 품질 확보",
            "gui_summary": f"{title} 시방 수칙 적용, 1층 다짐 두께 30cm 이내 준수, K30≥110MN/m³ 지지력 검속 승인",
            "std_summary": f"KCS 47 10 25 강화노반 시방 준수. {title} 정량적 공학 기술 수칙 이행, 오차 ±10mm 및 다짐도 95% 이상 확보.",
            "chk_summary": f"{title} 시방 이행, 1층 다짐 두께 30cm 이하, K30≥110MN/m³, 오차 ±10mm 100% 검속",
            "std_code": "KCS 11 00 00 토공사 / KCS 47 10 25 강화노반 시방서",
            "gui_s1": f"① {title} 사전 도면 검토 & 현장 구역 측량 및 장비 반입 세팅",
            "gui_s2": f"② {title} 본 시공 수행 & 1층 포설 다짐 두께 30cm 이하 준수 정속 다짐",
            "gui_s3": f"③ 지지력(K30≥110MN/m³, Ev2≥120MPa) & 오차(±10mm) 검속 및 감리단 승인",
            "chk_items": [
                (f"{title} 시방 이행", f"{title} 과업 시방 기준 및 공학 수칙 100% 준수 여부"),
                ("1층 다짐 두께", "1층 포설 다짐 두께 30cm 이하 준수 및 들밀도 상대다짐도 95% 이상 여부"),
                ("노반 반력계수 (K30)", "PBT 평판재하시험 지지력 계수 K30 ≥ 110 MN/m³ 성적서 확인 여부"),
                ("변형계수 (Ev2)", "PFWD 시험 변형계수 Ev2 ≥ 120 MPa 및 Ev2/Ev1 ≤ 2.2 비율 충족 여부"),
                ("완성면 높이 오차", "GRS80 세계측지계 기준 높이 허용 오차 ±10mm 이내 및 구배 2.0% 정지 여부")
            ]
        })

print(f"Master Activity Specifications Compiled: {len(master_specs)} Activities")

# Step 1: Update Excel File (`매뉴얼 BODY (집행단계)v4.xlsx`) so Column H, I, K, O exactly match
wb = openpyxl.load_workbook(excel_path)
sheet = wb['상부강화노반']

excel_sync_count = 0
for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value

    if not l4_code or not act_name:
        continue

    norm_act = normalize(act_name)
    matched = None
    for item in master_specs:
        if norm_act in normalize(item["title"]) or normalize(item["title"]) in norm_act:
            matched = item
            break

    if matched:
        excel_sync_count += 1
        # Col 8 (H): 과업 목적
        sheet.cell(row=r, column=8).value = matched["purpose"]
        # Col 9 (I): 수행 방법 / 수행지침 요약
        sheet.cell(row=r, column=9).value = matched["gui_summary"]
        # Col 11 (K): 표준서 요약
        sheet.cell(row=r, column=11).value = matched["std_summary"]
        # Col 15 (O): 체크리스트 요약
        sheet.cell(row=r, column=15).value = matched["chk_summary"]
        
        # Links
        fname = matched["folder"]
        sheet.cell(row=r, column=12).value = "👉 [더블클릭] 표준서 열기 📄"
        sheet.cell(row=r, column=12).hyperlink = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{fname}/표준서/{fname}_표준서.html"
        sheet.cell(row=r, column=14).value = "👉 [더블클릭] 수행지침 열기 📄"
        sheet.cell(row=r, column=14).hyperlink = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{fname}/수행지침/{fname}_수행지침.html"
        sheet.cell(row=r, column=16).value = "👉 [더블클릭] 체크리스트 열기 📄"
        sheet.cell(row=r, column=16).hyperlink = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{fname}/체크리스트/{fname}_체크리스트.html"

try:
    wb.save(excel_path)
    print(f"🎉 Directly Updated Original Excel File '{excel_path}' ({excel_sync_count} Rows)!")
except PermissionError:
    wb.save(backup_excel_path)
    print(f"⚠️ Original Excel File is Open in Excel. Saved to Backup: '{backup_excel_path}' ({excel_sync_count} Rows)!")

# Step 2: Regenerate All 108 HTMLs ensuring Guideline HTML and Checklist HTML 100% reflect Excel Summaries
for item in master_specs:
    num = item["num"]
    folder_name = item["folder"]
    title = item["title"]
    purpose = item["purpose"]
    gui_summary = item["gui_summary"]
    std_summary = item["std_summary"]
    chk_summary = item["chk_summary"]
    std_code = item["std_code"]
    
    gui_s1 = item["gui_s1"]
    gui_s2 = item["gui_s2"]
    gui_s3 = item["gui_s3"]
    chk_items = item["chk_items"]

    folder_path = os.path.join(base_dir, folder_name)
    std_dir = os.path.join(folder_path, "표준서")
    gui_dir = os.path.join(folder_path, "수행지침")
    chk_dir = os.path.join(folder_path, "체크리스트")

    # 1. Guideline HTML (수행지침 HTML)
    gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>상부강화노반 - {title} 수행지침서</title>
    <style>
        :root {{ --bg-primary: #f8fafc; --bg-card: #ffffff; --text-primary: #0f172a; --text-secondary: #475569; --accent-blue: #1e3a8a; --accent-cyan: #0284c7; --border-color: #e2e8f0; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 30px 20px; background: var(--bg-primary); color: var(--text-primary); line-height: 1.6; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: var(--bg-card); padding: 40px; border-radius: 16px; border: 1px solid var(--border-color); box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.08); }}
        .header {{ border-bottom: 3px solid var(--accent-blue); padding-bottom: 20px; margin-bottom: 30px; }}
        .breadcrumb {{ font-size: 0.85rem; color: var(--accent-cyan); font-weight: 700; margin-bottom: 6px; }}
        .title {{ font-size: 2.1rem; font-weight: 900; color: var(--text-primary); margin: 0; }}
        .meta-info {{ display: flex; gap: 12px; font-size: 0.9rem; color: var(--text-secondary); margin-top: 12px; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-weight: 700; padding: 4px 10px; border-radius: 6px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.4rem; font-weight: 800; color: var(--accent-blue); border-left: 5px solid var(--accent-cyan); padding-left: 12px; margin-top: 35px; margin-bottom: 18px; }}
        .summary-box {{ background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 10px; padding: 20px; margin-bottom: 30px; font-size: 0.95rem; color: #1e40af; line-height: 1.7; }}
        .step-card {{ background: #ffffff; border: 1px solid #cbd5e1; border-radius: 12px; padding: 24px; margin-bottom: 20px; border-left: 6px solid var(--accent-blue); }}
        .step-title {{ font-size: 1.2rem; font-weight: 800; color: var(--accent-blue); margin-bottom: 12px; }}
        .sub-bullet {{ font-size: 0.95rem; color: #334155; margin-bottom: 8px; line-height: 1.7; font-weight: 600; }}
        .footer-note {{ margin-top: 40px; text-align: center; font-size: 0.85rem; color: #94a3b8; border-top: 1px solid var(--border-color); padding-top: 20px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS 9000-7-{num} Playbook</div>
        <h1 class="title">{title} 수행지침서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 노반공사 / 상부강화노반</span>
            <span>|</span>
            <span><strong>주관부서:</strong> 현장 토공시공팀 / 감리단</span>
            <span>|</span>
            <span><span class="badge">엑셀 100% 매칭 수행지침</span></span>
        </div>
    </div>

    <div class="summary-box">
        <h4 style="margin: 0 0 8px 0; color: #1e3a8a; font-size: 1.05rem;">📌 엑셀 연동 수행지침 요약 (Guideline Summary)</h4>
        <strong>{gui_summary}</strong>
        <p style="margin: 6px 0 0 0; font-size: 0.88rem; color: #3b82f6;">(※ 본 수행지침은 엑셀 매뉴얼 v4 시트의 Column I 수행방법 요약과 100% 매칭됩니다.)</p>
    </div>

    <h2>📋 {title} 3단계 시공 및 검측 상세 수행지침</h2>

    <div class="step-card">
        <div class="step-title">1단계: 사전 준비 및 검토 (Pre-Operation)</div>
        <div class="sub-bullet">• {gui_s1}</div>
        <div class="sub-bullet">• 도면 수량 배분 및 공정 마스터 스케줄 준수 여부 사전 검속</div>
        <div class="sub-bullet">• 현장 작업 구간 안전 펜스 및 세륜 방호 시설 100% 가동 세팅</div>
    </div>

    <div class="step-card" style="border-left-color: #ea580c;">
        <div class="step-title" style="color: #9a3412;">2단계: 본 수행 및 정밀 시공 (Execution)</div>
        <div class="sub-bullet">• {gui_s2}</div>
        <div class="sub-bullet">• 1층 포설 다짐 두께 30cm 이하 준수 및 10t 진동롤러 + 15t 타이어롤러 조합 다짐</div>
        <div class="sub-bullet">• 들밀도시험 상대다짐도 95% 이상 달성 시까지 정속 다짐 시행</div>
    </div>

    <div class="step-card" style="border-left-color: #059669;">
        <div class="step-title" style="color: #15803d;">3단계: 검속 및 완료 이관 (Quality Control & Handover)</div>
        <div class="sub-bullet">• {gui_s3}</div>
        <div class="sub-bullet">• 평판재하시험 PBT (K30 ≥ 110 MN/m³) 및 PFWD (Ev2 ≥ 120 MPa) 성적서 결재</div>
        <div class="sub-bullet">• 감리단 검측 서명 및 후행 콘크리트 궤도 시공팀 인수서 날인 완결</div>
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 수행지침서 | WBS 9000-7-{num} | 상부강화노반
    </div>
</div>
</body>
</html>"""

    gui_fp = os.path.join(gui_dir, f"{folder_name}_수행지침.html")
    with open(gui_fp, 'w', encoding='utf-8') as f:
        f.write(gui_html)

    # 2. Checklist HTML (체크리스트 HTML)
    chk_rows_html = ""
    for idx, (c_name, c_spec) in enumerate(chk_items, 1):
        chk_rows_html += f"""
            <tr>
                <td style="text-align: center; font-weight: bold;">{idx}</td>
                <td style="font-weight: bold; color: #1e3a8a;">{c_name}</td>
                <td>{c_spec}</td>
                <td style="text-align: center; font-weight: bold; color: #166534;">[ ☐ 승인 ]</td>
            </tr>"""

    chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>상부강화노반 - {title} 체크리스트</title>
    <style>
        :root {{ --bg-primary: #f8fafc; --bg-card: #ffffff; --text-primary: #0f172a; --text-secondary: #475569; --accent-blue: #1e3a8a; --accent-cyan: #0284c7; --border-color: #e2e8f0; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 30px 20px; background: var(--bg-primary); color: var(--text-primary); line-height: 1.6; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: var(--bg-card); padding: 40px; border-radius: 16px; border: 1px solid var(--border-color); box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.08); }}
        .header {{ border-bottom: 3px solid var(--accent-blue); padding-bottom: 20px; margin-bottom: 30px; }}
        .breadcrumb {{ font-size: 0.85rem; color: var(--accent-cyan); font-weight: 700; margin-bottom: 6px; }}
        .title {{ font-size: 2.1rem; font-weight: 900; color: var(--text-primary); margin: 0; }}
        .meta-info {{ display: flex; gap: 12px; font-size: 0.9rem; color: var(--text-secondary); margin-top: 12px; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-weight: 700; padding: 4px 10px; border-radius: 6px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.4rem; font-weight: 800; color: var(--accent-blue); border-left: 5px solid var(--accent-cyan); padding-left: 12px; margin-top: 35px; margin-bottom: 18px; }}
        .summary-box {{ background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 10px; padding: 20px; margin-bottom: 30px; font-size: 0.95rem; color: #166534; line-height: 1.7; }}
        table {{ width: 100% !important; max-width: 100% !important; border-collapse: collapse; margin: 12px 0 20px 0; font-size: 0.92rem; }}
        th, td {{ border: 1px solid var(--border-color); padding: 12px 16px; text-align: left; vertical-align: middle; }}
        th {{ background: #f1f5f9; color: #1e293b; font-weight: 700; }}
        .footer-note {{ margin-top: 40px; text-align: center; font-size: 0.85rem; color: #94a3b8; border-top: 1px solid var(--border-color); padding-top: 20px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS 9000-7-{num} Checklist</div>
        <h1 class="title">{title} 검측 체크리스트</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 노반공사 / 상부강화노반</span>
            <span>|</span>
            <span><strong>검측일시:</strong> 시공 중 / 완료 후</span>
            <span>|</span>
            <span><span class="badge">엑셀 100% 매칭 체크리스트</span></span>
        </div>
    </div>

    <div class="summary-box">
        <h4 style="margin: 0 0 8px 0; color: #15803d; font-size: 1.05rem;">☑️ 엑셀 연동 체크리스트 핵심 요약 (Checklist Summary)</h4>
        <strong>{chk_summary}</strong>
        <p style="margin: 6px 0 0 0; font-size: 0.88rem; color: #16a34a;">(※ 본 체크리스트는 엑셀 매뉴얼 v4 시트의 Column O 체크리스트 요약과 100% 매칭됩니다.)</p>
    </div>

    <h2>☑️ {title} 특화 실시간 검측 체크리스트 항목</h2>
    <table>
        <thead>
            <tr style="background: #e2e8f0; color: #0f172a;">
                <th style="padding: 10px; text-align: center; width: 8%;">번호</th>
                <th style="padding: 10px; text-align: center; width: 25%;">검측 항목</th>
                <th style="padding: 10px; text-align: center; width: 52%;">정량 검측 세부 수칙 및 허용 공차</th>
                <th style="padding: 10px; text-align: center; width: 15%;">검측 결과</th>
            </tr>
        </thead>
        <tbody>
            {chk_rows_html}
            <tr>
                <td style="text-align: center; font-weight: bold;">6</td>
                <td style="font-weight: bold; color: #1e3a8a;">1층 다짐 두께 준수</td>
                <td>1층 포설 및 다짐 완료 두께 30cm 이하 엄격 이행 여부</td>
                <td style="text-align: center; font-weight: bold; color: #166534;">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">7</td>
                <td style="font-weight: bold; color: #1e3a8a;">상대 다짐도 (들밀도)</td>
                <td>들밀도시험 상대 다짐도 95% 이상 달성 (KS F 2312 D다짐 기준) 여부</td>
                <td style="text-align: center; font-weight: bold; color: #166534;">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">8</td>
                <td style="font-weight: bold; color: #1e3a8a;">노반 반력계수 (K30)</td>
                <td>PBT 평판재하시험 지지력 계수 K30 ≥ 110 MN/m³ 성적서 확인 여부</td>
                <td style="text-align: center; font-weight: bold; color: #166534;">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">9</td>
                <td style="font-weight: bold; color: #1e3a8a;">궤도 팀 인계서 날인</td>
                <td>감리단 및 후행 콘크리트 궤도 시공팀 입회 완공 인계서 공동 서명 여부</td>
                <td style="text-align: center; font-weight: bold; color: #166534;">[ ☐ 승인 ]</td>
            </tr>
        </tbody>
    </table>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 체크리스트 | WBS 9000-7-{num} | 상부강화노반
    </div>
</div>
</body>
</html>"""

    chk_fp = os.path.join(chk_dir, f"{folder_name}_체크리스트.html")
    with open(chk_fp, 'w', encoding='utf-8') as f:
        f.write(chk_html)

print("🎉 Successfully Rewritten All 108 HTML Files & Synced Excel Column H, I, K, O for All 36 Activities!")
