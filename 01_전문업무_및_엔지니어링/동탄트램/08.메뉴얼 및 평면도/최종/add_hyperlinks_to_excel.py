import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import urllib.parse
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

def sanitize_name(name):
    name = re.sub(r'[\/\\:\*\?"<>\|]', '_', str(name))
    name = name.strip()
    return name

linked_count = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ['GUIDE']:
        continue
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        continue
        
    # Find header row
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        row_vals = [str(cell.value) if cell.value is not None else "" for cell in r]
        if any('코드' in val or '작업단위' in val for val in row_vals):
            header_row_idx = idx
            break
            
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[header_row_idx]]
    
    col_map = {h: i for i, h in enumerate(headers)}
    
    def get_col_idx(possible_keys):
        for k in possible_keys:
            for h in headers:
                if k in h:
                    return col_map[h]
        return None
        
    l4_col_idx = get_col_idx(['L4 코드', 'L4코드', 'L4'])
    act_col_idx = get_col_idx(['작업단위', 'Activity', 'Task'])
    std_col_idx = get_col_idx(['표준서'])
    gui_col_idx = get_col_idx(['수행지침'])
    chk_col_idx = get_col_idx(['체크리스트'])
    
    if act_col_idx is None or std_col_idx is None:
        print(f"Skipping sheet {sheet_name} (Missing required columns)")
        continue
        
    sheet_act_idx = 0
    
    for row in rows[header_row_idx+1:]:
        l4_val = row[l4_col_idx].value if l4_col_idx is not None and l4_col_idx < len(row) else None
        act_val = row[act_col_idx].value if act_col_idx is not None and act_col_idx < len(row) else None
        
        if not l4_val and not act_val:
            continue
            
        sheet_act_idx += 1
        act_name = str(act_val).strip() if act_val else f"Task_{sheet_act_idx}"
        sanitized_act = sanitize_name(act_name)
        
        disc_dir_name = sheet_name
        folder_name = f"{sheet_act_idx}_{sanitized_act}"
        
        if sheet_name == '공정표에따른 매뉴얼':
            found_disc = None
            found_sub = None
            for d in os.listdir(base_attach_dir):
                d_path = os.path.join(base_attach_dir, d)
                if os.path.isdir(d_path):
                    for sub in os.listdir(d_path):
                        if sanitized_act in sub or (str(l4_val) in sub if l4_val else False):
                            found_disc = d
                            found_sub = sub
                            break
                if found_disc:
                    break
            if found_disc:
                disc_dir_name = found_disc
                folder_name = found_sub
            else:
                disc_dir_name = "사전토공사"
                
        act_folder_abs = os.path.join(base_attach_dir, disc_dir_name, folder_name)
        
        def set_link(cell, doc_type, doc_title):
            global linked_count
            if cell is None:
                return
                
            sub_folder = os.path.join(act_folder_abs, doc_type)
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    html_file = files[0]
                    rel_target = f"매뉴얼BODY(집행단계-첨부폴더)/{disc_dir_name}/{folder_name}/{doc_type}/{html_file}"
                    
                    val_str = str(cell.value) if cell.value is not None else ""
                    btn_text = f"\n--------------------------------------\n👉 [더블클릭] 상세 {doc_title} 파일(HTML) 열기 📄"
                    
                    if '더블클릭' not in val_str:
                        cell.value = (val_str.strip() + btn_text).strip()
                        
                    target_encoded = urllib.parse.quote(rel_target, safe='/')
                    cell.hyperlink = target_encoded
                    linked_count += 1
                    
        if std_col_idx is not None and std_col_idx < len(row):
            set_link(row[std_col_idx], '표준서', '표준서')
            
        if gui_col_idx is not None and gui_col_idx < len(row):
            set_link(row[gui_col_idx], '수행지침', '수행지침')
            
        if chk_col_idx is not None and chk_col_idx < len(row):
            set_link(row[chk_col_idx], '체크리스트', '체크리스트')

print(f"Hyperlink linking complete! Total links updated/added: {linked_count}")
wb.save(excel_path)
print(f"Saved updated Excel workbook to '{excel_path}'")
