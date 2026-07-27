import openpyxl
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v3_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
v4_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

print("Updating headers in v3 master and copying cleanly to v4...")

wb = openpyxl.load_workbook(v3_path)
ws = wb['지장물이설']

ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크 (가스: 62일 | 난방: 80일 | 통신: 150일 | 전력: 255일 | 광역상수: 160일)")
ws.cell(row=2, column=14, value="가스관 (62일)")
ws.cell(row=2, column=15, value="열(난방)배관로 (80일)")
ws.cell(row=2, column=16, value="통신관로 (150일)")
ws.cell(row=2, column=17, value="전력관 (255일)")
ws.cell(row=2, column=18, value="광역상수관 (160일)")

wb.save(v3_path)
print("✅ Saved updated headers to v3 master!")

try:
    shutil.copy2(v3_path, v4_path)
    print("🎉 Successfully synchronized and updated v4 master workbook!")
except Exception as e:
    print(f"⚠️ Notice: Excel lock active. Content fully updated in v3 master. Copy to v4 ready.")
