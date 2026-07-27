import openpyxl
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

def sanitize_name(name):
    name = re.sub(r'[\/\\:\*\?"<>\|]', '_', str(name))
    name = name.strip()
    return name

def format_content(text, default_msg="● 해당사항 없음 또는 관련 설계기준서 준수"):
    if not text or not str(text).strip():
        return f"<p class='empty-text'>{default_msg}</p>"
    
    text = str(text).strip()
    lines = text.split('\n')
    
    html_out = []
    in_list = False
    
    for line in lines:
        line_s = line.strip()
        if not line_s:
            continue
        
        if line_s.startswith(('1)', '2)', '3)', '4)', '5)', '6)', '7)', '8)', '9)', '10)', '-', '●', '□', '👉')):
            if not in_list:
                html_out.append("<ul class='bullet-list'>")
                in_list = True
            
            clean_item = re.sub(r'^(1\)|2\)|3\)|4\)|5\)|6\)|7\)|8\)|9\)|10\)|-|●|□|👉)\s*', '', line_s)
            html_out.append(f"<li>{clean_item}</li>")
        else:
            if in_list:
                html_out.append("</ul>")
                in_list = False
            html_out.append(f"<p>{line_s}</p>")
            
    if in_list:
        html_out.append("</ul>")
        
    return "\n".join(html_out)

def generate_standard_html(discipline, l4_code, act_name, dept, supervisor, purpose, method, output, standard_text, criteria_text, advice_text):
    purp = purpose if purpose else '설계도서 준수 및 시공 품질 관리'
    meth = method if method else '표준 시공 지침 및 설계기준 준수'
    outp = output if output else '결과 보고서 및 검측 승인서'
    
    std_formatted = format_content(standard_text, "● 국가철도공단 설계지침 및 관련 법령 절대 기준 준수")
    crit_formatted = format_content(criteria_text, "● 관련 분야 설계기준서 준수")
    adv_formatted = format_content(advice_text, "● 협력사 자문 내용 없음 또는 해당사항 없음")

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{discipline} - {act_name} 기술 표준서</title>
    <style>
        :root {{
            --primary-color: #2563EB;
            --primary-light: #EFF6FF;
            --text-color: #1F2937;
            --bg-color: #F9FAFB;
            --card-bg: #FFFFFF;
            --border-color: #E5E7EB;
        }}
        body {{
            font-family: 'Inter', '맑은 고딕', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-color);
            margin: 0;
            padding: 40px 20px;
            line-height: 1.6;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            padding: 40px;
        }}
        .header {{
            border-bottom: 2px solid var(--primary-light);
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .breadcrumb {{
            font-size: 0.85rem;
            color: #6B7280;
            margin-bottom: 8px;
            text-transform: uppercase;
            font-weight: 600;
        }}
        .title {{
            font-size: 1.8rem;
            color: #111827;
            margin: 0;
            font-weight: 800;
        }}
        .meta-info {{
            display: flex;
            gap: 15px;
            margin-top: 12px;
            font-size: 0.9rem;
            color: #6B7280;
            flex-wrap: wrap;
        }}
        .badge {{
            background-color: var(--primary-light);
            color: var(--primary-color);
            padding: 4px 8px;
            border-radius: 6px;
            font-weight: 600;
            font-size: 0.8rem;
        }}
        h2 {{
            font-size: 1.25rem;
            color: #111827;
            border-left: 4px solid var(--primary-color);
            padding-left: 12px;
            margin-top: 35px;
            margin-bottom: 15px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
            margin-bottom: 25px;
            font-size: 0.9rem;
        }}
        th, td {{
            border: 1px solid var(--border-color);
            padding: 12px 16px;
        }}
        th {{
            background-color: var(--primary-light);
            color: #1E3A8A;
            font-weight: 700;
            width: 25%;
        }}
        .bullet-list {{
            list-style: none;
            padding-left: 0;
            margin-bottom: 25px;
        }}
        .bullet-list li {{
            position: relative;
            padding-left: 20px;
            margin-bottom: 10px;
            font-size: 0.95rem;
        }}
        .bullet-list li::before {{
            content: "•";
            position: absolute;
            left: 0;
            color: var(--primary-color);
            font-weight: 700;
        }}
        .footer-note {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid var(--border-color);
            font-size: 0.85rem;
            color: #9CA3AF;
            text-align: center;
        }}
        .empty-text {{
            color: #9CA3AF;
            font-style: italic;
        }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS {l4_code} Standard</div>
        <h1 class="title">{act_name} 기술 표준서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> {discipline} / {dept}</span>
            <span>|</span>
            <span><strong>주관부서:</strong> {supervisor}</span>
            <span>|</span>
            <span><span class="badge">설계 기술 표준 규격</span></span>
        </div>
    </div>

    <h2>1. 과업 개요 및 목적 (Overview & Scope)</h2>
    <table>
        <tbody>
            <tr>
                <th>과업 목적</th>
                <td>{purp}</td>
            </tr>
            <tr>
                <th>수행 방법</th>
                <td>{meth}</td>
            </tr>
            <tr>
                <th>주요 산출물</th>
                <td>{outp}</td>
            </tr>
        </tbody>
    </table>

    <h2>2. 정량적 기술 표준 (Technical Specifications)</h2>
    {std_formatted}

    <h2>3. 첨부서류 연계 상세 설계기준 (Design Criteria)</h2>
    {crit_formatted}

    <h2>4. 협력사 시공/공사관리 자문 (Subcontractor Advisory)</h2>
    {adv_formatted}

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 기술 표준서 | WBS {l4_code}
    </div>
</div>
</body>
</html>
"""

def generate_guideline_html(discipline, l4_code, act_name, dept, supervisor, purpose, method, output, guideline_text, risk_text):
    purp = purpose if purpose else '작업 절차 및 품질 관리 지침 이행'
    meth = method if method else '표준 시공 절차에 준하여 시행'
    outp = output if output else '시공 완료 보고서 및 품질 기록서'
    
    gui_formatted = format_content(guideline_text, "1) 사전 준비 단계: 설계 도서 및 기술 기준 검토를 완료한다.\n2) 본 시공 단계: 시공 표준 지침을 준수하여 관리한다.\n3) 검사 및 마감 단계: 품질 검측을 수행하고 기록을 보존한다.")
    risk_formatted = format_content(risk_text, "● 사전 시공성 검토 및 안전 관리 수칙 준수")

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{discipline} - {act_name} 작업 수행지침서</title>
    <style>
        :root {{
            --primary-color: #059669;
            --primary-light: #ECFDF5;
            --text-color: #1F2937;
            --bg-color: #F9FAFB;
            --card-bg: #FFFFFF;
            --border-color: #E5E7EB;
            --danger-color: #EF4444;
        }}
        body {{
            font-family: 'Inter', '맑은 고딕', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-color);
            margin: 0;
            padding: 40px 20px;
            line-height: 1.6;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            padding: 40px;
        }}
        .header {{
            border-bottom: 2px solid var(--primary-light);
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .breadcrumb {{
            font-size: 0.85rem;
            color: #6B7280;
            margin-bottom: 8px;
            text-transform: uppercase;
            font-weight: 600;
        }}
        .title {{
            font-size: 1.8rem;
            color: #111827;
            margin: 0;
            font-weight: 800;
        }}
        .meta-info {{
            display: flex;
            gap: 15px;
            margin-top: 12px;
            font-size: 0.9rem;
            color: #6B7280;
            flex-wrap: wrap;
        }}
        .badge {{
            background-color: var(--primary-light);
            color: var(--primary-color);
            padding: 4px 8px;
            border-radius: 6px;
            font-weight: 600;
            font-size: 0.8rem;
        }}
        h2 {{
            font-size: 1.25rem;
            color: #111827;
            border-left: 4px solid var(--primary-color);
            padding-left: 12px;
            margin-top: 35px;
            margin-bottom: 15px;
        }}
        .bullet-list {{
            list-style: none;
            padding-left: 0;
            margin-bottom: 25px;
        }}
        .bullet-list li {{
            position: relative;
            padding-left: 20px;
            margin-bottom: 10px;
            font-size: 0.95rem;
        }}
        .bullet-list li::before {{
            content: "✔";
            position: absolute;
            left: 0;
            color: var(--primary-color);
            font-weight: 700;
        }}
        .risk-box {{
            background-color: #FEF2F2;
            border-left: 4px solid var(--danger-color);
            padding: 16px 20px;
            border-radius: 0 8px 8px 0;
            margin: 25px 0;
            font-size: 0.9rem;
            color: #991B1B;
        }}
        .footer-note {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid var(--border-color);
            font-size: 0.85rem;
            color: #9CA3AF;
            text-align: center;
        }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS {l4_code} Guideline</div>
        <h1 class="title">{act_name} 작업 수행지침서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> {discipline} / {dept}</span>
            <span>|</span>
            <span><strong>주관부서:</strong> {supervisor}</span>
            <span>|</span>
            <span><span class="badge">현장 시공 관리 지침</span></span>
        </div>
    </div>

    <h2>1. 작업 개요 및 준비사항 (Preparation)</h2>
    <p><strong>주요 목적:</strong> {purp}</p>
    <p><strong>수행 방법:</strong> {meth}</p>
    <p><strong>최종 산출물:</strong> {outp}</p>

    <h2>2. 작업 절차 및 세부 지침 (Procedures)</h2>
    {gui_formatted}

    <h2>3. 하자 예방 및 위험요인 관리 (Risk Management)</h2>
    <div class="risk-box">
        <strong>[집행단계 리스크 검토사항]</strong><br>
        {risk_formatted}
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 수행지침서 | WBS {l4_code}
    </div>
</div>
</body>
</html>
"""

def generate_checklist_html(discipline, l4_code, act_name, dept, supervisor, checklist_text, risk_text, advice_text):
    chk_formatted = format_content(checklist_text, "☐ 설계도서 및 표준 시방 기준 준수 여부 확인\n☐ 품질 검측 성과표 작성 및 승인 여부 확인")
    risk_formatted = format_content(risk_text, "☐ 공종 간 인터페이스 및 주요 하자 리스크 사전 검토 이행 완료")
    adv_formatted = format_content(advice_text, "☐ 협력사 자문 내용 확인 완료")

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{discipline} - {act_name} 완료 검측 체크리스트</title>
    <style>
        :root {{
            --primary-color: #D97706;
            --primary-light: #FEF3C7;
            --text-color: #1F2937;
            --bg-color: #F9FAFB;
            --card-bg: #FFFFFF;
            --border-color: #E5E7EB;
        }}
        body {{
            font-family: 'Inter', '맑은 고딕', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-color);
            margin: 0;
            padding: 40px 20px;
            line-height: 1.6;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            padding: 40px;
        }}
        .header {{
            border-bottom: 2px solid var(--primary-light);
            padding-bottom: 20px;
            margin-bottom: 30px;
        }}
        .breadcrumb {{
            font-size: 0.85rem;
            color: #6B7280;
            margin-bottom: 8px;
            text-transform: uppercase;
            font-weight: 600;
        }}
        .title {{
            font-size: 1.8rem;
            color: #111827;
            margin: 0;
            font-weight: 800;
        }}
        .meta-info {{
            display: flex;
            gap: 15px;
            margin-top: 12px;
            font-size: 0.9rem;
            color: #6B7280;
            flex-wrap: wrap;
        }}
        .badge {{
            background-color: var(--primary-light);
            color: var(--primary-color);
            padding: 4px 8px;
            border-radius: 6px;
            font-weight: 600;
            font-size: 0.8rem;
        }}
        h2 {{
            font-size: 1.25rem;
            color: #111827;
            border-left: 4px solid var(--primary-color);
            padding-left: 12px;
            margin-top: 35px;
            margin-bottom: 15px;
        }}
        .bullet-list {{
            list-style: none;
            padding-left: 0;
            margin-bottom: 25px;
        }}
        .bullet-list li {{
            position: relative;
            padding-left: 28px;
            margin-bottom: 12px;
            font-size: 0.95rem;
        }}
        .bullet-list li::before {{
            content: "☐";
            position: absolute;
            left: 0;
            color: var(--primary-color);
            font-weight: 700;
            font-size: 1.1rem;
        }}
        .footer-note {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid var(--border-color);
            font-size: 0.85rem;
            color: #9CA3AF;
            text-align: center;
        }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS {l4_code} Checklist</div>
        <h1 class="title">{act_name} 완료 검측 체크리스트</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> {discipline} / {dept}</span>
            <span>|</span>
            <span><strong>주관부서:</strong> {supervisor}</span>
            <span>|</span>
            <span><span class="badge">완료 검측 표준</span></span>
        </div>
    </div>

    <h2>1. 완료 검측 세부 체크리스트 (Inspection Items)</h2>
    {chk_formatted}

    <h2>2. 집행단계 리스크 점검사항 (LLBS Risk Checklist)</h2>
    {risk_formatted}

    <h2>3. 협력사 공사관리 검측 확인사항 (Subcontractor Verification)</h2>
    {adv_formatted}

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 체크리스트 | WBS {l4_code}
    </div>
</div>
</body>
</html>
"""

wb = openpyxl.load_workbook(excel_path, data_only=True)
target_sheets = ['사전토공사', '상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

generated_count = 0
skipped_count = 0

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
        
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        if any(isinstance(val, str) and ('코드' in val or '작업단위' in val) for val in r if val):
            header_row_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
    h_map = {h: i for i, h in enumerate(headers)}
    
    def get_val(r, possible_keys):
        for k in possible_keys:
            for h in headers:
                if k in h:
                    idx = h_map[h]
                    if idx < len(r) and r[idx] is not None:
                        return str(r[idx]).strip()
        return ""

    discipline_dir = os.path.join(base_attach_dir, sheet_name)
    act_idx = 0
    
    for r in rows[header_row_idx+1:]:
        if all(v is None for v in r):
            continue
            
        l4_code = get_val(r, ['L4 코드', 'L4코드', 'L4'])
        act_name = get_val(r, ['작업단위', 'Activity', 'Task'])
        
        if not l4_code and not act_name:
            continue
            
        act_idx += 1
        sanitized_act = sanitize_name(act_name if act_name else f"Task_{act_idx}")
        folder_name = f"{act_idx}_{sanitized_act}"
        act_folder_path = os.path.join(discipline_dir, folder_name)
        
        dept = get_val(r, ['담당 분야', '담당분야', '분야'])
        supervisor = get_val(r, ['주관'])
        purpose = get_val(r, ['목적'])
        method = get_val(r, ['방법'])
        output = get_val(r, ['산출물'])
        standard_text = get_val(r, ['표준서'])
        guideline_text = get_val(r, ['수행지침'])
        checklist_text = get_val(r, ['체크리스트'])
        criteria_text = get_val(r, ['첨부서류', '설계기준'])
        risk_text = get_val(r, ['리스크'])
        advice_text = get_val(r, ['협력사', '자문'])
        
        # 1. 표준서 HTML
        std_dir = os.path.join(act_folder_path, '표준서')
        os.makedirs(std_dir, exist_ok=True)
        std_file = os.path.join(std_dir, f"{sanitized_act}_표준서.html")
        if not os.path.exists(std_file):
            with open(std_file, 'w', encoding='utf-8') as f:
                f.write(generate_standard_html(sheet_name, l4_code, act_name, dept, supervisor, purpose, method, output, standard_text, criteria_text, advice_text))
            generated_count += 1
        else:
            skipped_count += 1
            
        # 2. 수행지침 HTML
        gui_dir = os.path.join(act_folder_path, '수행지침')
        os.makedirs(gui_dir, exist_ok=True)
        gui_file = os.path.join(gui_dir, f"{sanitized_act}_수행지침.html")
        if not os.path.exists(gui_file):
            with open(gui_file, 'w', encoding='utf-8') as f:
                f.write(generate_guideline_html(sheet_name, l4_code, act_name, dept, supervisor, purpose, method, output, guideline_text, risk_text))
            generated_count += 1
        else:
            skipped_count += 1
            
        # 3. 체크리스트 HTML
        chk_dir = os.path.join(act_folder_path, '체크리스트')
        os.makedirs(chk_dir, exist_ok=True)
        chk_file = os.path.join(chk_dir, f"{sanitized_act}_체크리스트.html")
        if not os.path.exists(chk_file):
            with open(chk_file, 'w', encoding='utf-8') as f:
                f.write(generate_checklist_html(sheet_name, l4_code, act_name, dept, supervisor, checklist_text, risk_text, advice_text))
            generated_count += 1
        else:
            skipped_count += 1

print(f"Done! Generated: {generated_count} HTML files, Skipped (already existed): {skipped_count} HTML files.")
