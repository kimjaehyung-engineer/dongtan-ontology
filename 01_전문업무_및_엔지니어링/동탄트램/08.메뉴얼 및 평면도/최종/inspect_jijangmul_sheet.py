import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v2.xlsx"
wb = openpyxl.load_workbook(excel_path)

if '지장물이설' in wb.sheetnames:
    ws = wb['지장물이설']
    print(f"=== Sheet '지장물이설' Details ===")
    print("Max rows:", ws.max_row, "| Max cols:", ws.max_column)
    for r in range(1, min(15, ws.max_row + 1)):
        vals = [str(ws.cell(row=r, column=c).value or "").replace('\n', ' ')[:30] for c in range(1, min(10, ws.max_column + 1))]
        print(f"  Row {r:2d}: {vals}")
