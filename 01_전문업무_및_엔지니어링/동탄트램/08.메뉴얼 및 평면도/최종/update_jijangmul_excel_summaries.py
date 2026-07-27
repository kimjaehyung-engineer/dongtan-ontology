import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
source_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

print(f"Loading '{source_file}'...")
wb = openpyxl.load_workbook(source_file)
ws = wb['지장물이설']

# Find target row
target_row = None
for r in range(3, ws.max_row + 1):
    act_val = ws.cell(row=r, column=6).value
    if act_val and "지장물 이설 요청 (위수탁고)" in str(act_val):
        target_row = r
        break

if target_row:
    print(f"Found target activity at Row {target_row}")
    
    std_summary = "1) 위수탁기관별 관련 법령에 따른 이설 착수 요청 공문 정식 발송\n2) 발주기관 연계 이설 범위(1.5m 이격) 및 예산 분담(도급/위수탁) 명시"
    gui_summary = "1) GPR 줄따기 및 8대 관종별(도급m/위수탁OX) 이설 위치도 첨부 공문 발송\n2) 위수탁기관 14일 회신 관리, 무단수 Cut-over 및 PS 정산 수칙 준수"
    chk_summary = "1) 위수탁 공식 공문 발송 및 8대 관종 분류표 첨부 여부를 확인했는가?\n2) GPR 탐지 오차(±10cm) 및 기관 입회 하 이격거리(1.5m)를 검측했는가?"

    ws.cell(row=target_row, column=19, value=std_summary) # Std sum
    ws.cell(row=target_row, column=21, value=gui_summary) # Gui sum
    ws.cell(row=target_row, column=23, value=chk_summary) # Chk sum

    wb.save(target_file)
    print(f"🎉 Successfully saved updated workbook to '{target_file}'")
else:
    print("❌ Target row not found!")
