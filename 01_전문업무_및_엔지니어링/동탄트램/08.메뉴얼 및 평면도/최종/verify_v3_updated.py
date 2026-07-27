import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

v3_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3.xlsx"
wb = openpyxl.load_workbook(v3_path)

ws = wb['지장물이설']
print("=== Final Verification of Updated v3 '지장물이설' Sheet ===")
print("Max Rows:", ws.max_row, "| Max Cols:", ws.max_column)

print("\nSampling Rows (Row 4: 도급자분 이설업체 선정, Row 9: 인허가 절차 진행, Row 38: 후행공종 요구사항):")
for r in [4, 9, 38]:
    l4 = ws.cell(row=r, column=4).value
    act = ws.cell(row=r, column=6).value
    own = ws.cell(row=r, column=7).value
    std_sum = str(ws.cell(row=r, column=11).value).replace('\n', ' ')
    link12 = ws.cell(row=r, column=12).hyperlink is not None
    link14 = ws.cell(row=r, column=14).hyperlink is not None
    link16 = ws.cell(row=r, column=16).hyperlink is not None
    
    print(f"\n[Row {r:2d}] L4='{l4}' | Act='{act}' | Owner='{own}'")
    print(f"  - 표준서 요약: {std_sum[:75]}...")
    print(f"  - Links OK: Std={link12}, Gui={link14}, Chk={link16}")
