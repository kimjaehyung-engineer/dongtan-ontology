import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test"

# Cell A1: cell.hyperlink set on multi-line text
ws['A1'] = "Line 1 요약 내용\nLine 2 요약 내용\n👉 [더블클릭] 상세 표준서 파일(HTML) 열기 📄"
ws['A1'].hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\상부강화노반\1_지반조사 상세검토\표준서\지반조사 상세검토_표준서.html"

# Cell B1: formula HYPERLINK
rel_path = r"매뉴얼BODY(집행단계-첨부폴더)\상부강화노반\1_지반조사 상세검토\표준서\지반조사 상세검토_표준서.html"
ws['B1'] = f'=HYPERLINK("{rel_path}", "Line 1 요약 내용" & CHAR(10) & "👉 [더블클릭] 상세 표준서 파일(HTML) 열기 📄")'

wb.save("test_hyperlink_methods.xlsx")
print("Saved test_hyperlink_methods.xlsx")
