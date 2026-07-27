import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

target_folders = []
for f in os.listdir(base_dir):
    full_p = os.path.join(base_dir, f)
    if os.path.isdir(full_p):
        if f.startswith("4_") or f.startswith("5_") or f.startswith("6_"):
            target_folders.append(full_p)

print(f"Target Folders for Professional Term Refinement: {target_folders}")

# Replacements dictionary: (Unprofessional phrase -> Professional engineering phrase)
replacements = {
    "안테나 잘 터짐 합격증(TTA/KCA) 챙기기 & 관제·운영사 합격 도장: 무전기와 안테나가 잘 터진다는 정부 합격증(TTA/KCA 필증) 서류를 챙기고 관제·운영사·감리원 합동 서명 도장을 찍습니다.": "무선통신 성능검증서(TTA/KCA 필증) 확보 및 관제·운영사 최종 서명: 무선 장비 전파 성능 및 보안성이 검증된 국가 공인 필증(TTA/KCA)을 확보하고, 관제·운영사·감리원 합동 서명 회의록을 체결합니다.",
    "안테나 잘 터짐 합격증(TTA/KCA) 챙기기 & 관제·운영사 합격 도장": "무선통신 성능검증서(TTA/KCA 필증) 확보 및 관제·운영사 최종 서명",
    "안테나 잘 터짐 합격증(TTA/KCA) 챙기기 & 5개 회사가 모여 합격 도장 찍기": "무선통신 성능검증서(TTA/KCA 필증) 확보 및 제작사 5자 서명 체결",
    "안테나 잘 터짐 합격증(TTA/KCA) 챙기기": "무선통신 성능검증서(TTA/KCA 필증) 확보",
    "무전기와 안테나가 잘 터진다는 정부 합격증(TTA/KCA 필증) 서류를 챙기고": "무선 장비 전파 성능 및 보안성이 검증된 국가 공인 필증(TTA/KCA)을 확보하고",
    "무전기와 안테나가 잘 터지고 해킹당하지 않는다는 정부 합격증 서류(TTA/KCA)를 최종 챙기고": "무선 장비 전파 성능 및 보안성이 검증된 국가 공인 인증서(TTA/KCA)를 최종 확보하고",
    "무전기와 안테나가 잘 터지고 해킹당하지 않는다는 정부 합격증(TTA/KCA)을 챙기고": "무선 장비 전파 성능 및 보안성이 검증된 국가 공인 인증서(TTA/KCA)를 확보하고",
    "안테나 잘 터짐 인증서 (TTA)": "무선통신 전파성능 및 보안검증 (TTA)",
    "정부 무선 전파 합격증": "국가 공인 무선전파 적합필증 (TTA/KCA)",
    "정부 무선 합격증 (TTA/KCA)": "국가 공인 무선전파 적합필증 (TTA/KCA)",
    "정부 전파 합격증 (TTA/KCA) 챙기기": "국가 공인 무선전파 적합필증(TTA/KCA) 확보",
    "안테나 잘터짐 [TTA 인증]: 무전기 안테나가 잘 터지고 해킹 안 된다는 정부 인증서(TTA)를 챙겼는가?": "무선성능 검증 [TTA 인증]: 무선 장비 전파 성능 및 보안성이 검증된 국가 공인 TTA 성능검증서를 수신하였는가?",
    "안테나 잘터짐 [TTA 인증]: 무전기 안테나가 잘 터지고 해킹 안 된다는 정부 인증서(TTA)를 챙겼는가": "무선성능 검증 [TTA 인증]: 무선 장비 전파 성능 및 보안성이 검증된 국가 공인 TTA 성능검증서를 수신하였는가",
    "안테나 잘터짐": "무선성능 검증",
    "안테나 잘 터짐": "무선통신 성능 검증",
    "합격 도장 찍기": "최종 서명 체결",
    "합격 도장": "서명 날인"
}

# 1. Update HTML Files across 4_, 5_, 6_ folders
for folder in target_folders:
    for root, dirs, files in os.walk(folder):
        for file in files:
            if file.endswith('.html'):
                file_path = os.path.join(root, file)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                modified = False
                for old_text, new_text in replacements.items():
                    if old_text in content:
                        content = content.replace(old_text, new_text)
                        modified = True
                
                if modified:
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write(content)
                    print(f"   ✓ [REFINED PROFESSIONALLY] Updated: {file_path}")

# 2. Update Excel File (매뉴얼 BODY (집행단계)v4.xlsx)
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        if "통신분야" in wb.sheetnames:
            ws = wb["통신분야"]
            
            for r in range(4, 7):
                for col in [10, 12, 14]: # Cols J, L, N
                    cell_val = ws.cell(row=r, column=col).value
                    if cell_val and isinstance(cell_val, str):
                        for old_t, new_t in replacements.items():
                            if old_t in cell_val:
                                cell_val = cell_val.replace(old_t, new_t)
                        ws.cell(row=r, column=col, value=cell_val)
            
            wb.save(excel_path)
            print("   ✓ [EXCEL V4 REFINED] Successfully updated Rows 4, 5, 6 in 매뉴얼 BODY (집행단계)v4.xlsx with professional terms")
    except Exception as e:
        print(f"   ⚠️ Excel Refinement Note: {e}")

print("\n🎉 SUCCESSFULLY REFINED ALL WIRELESS TERMS INTO HIGHLY PROFESSIONAL ENGINEERING LANGUAGE!")
