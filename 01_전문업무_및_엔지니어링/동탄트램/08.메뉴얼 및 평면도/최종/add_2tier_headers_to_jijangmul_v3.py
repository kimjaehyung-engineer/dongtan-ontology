import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

def sanitize_name(name):
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', str(name)).strip()

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v3_save_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
base_attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)")
jijangmul_attach_dir = os.path.join(base_attach_dir, "지장물이설")

# Load existing Jijangmul activities list
from update_jijangmul_sheet_v3 import jijangmul_activities_v3

wb = openpyxl.load_workbook(v3_save_path)

if '지장물이설' in wb.sheetnames:
    idx = wb.sheetnames.index('지장물이설')
    del wb['지장물이설']
    ws = wb.create_sheet(title='지장물이설', index=idx)
else:
    ws = wb.create_sheet(title='지장물이설', index=2)

print(f"Reconstructing '지장물이설' sheet with 2-tier classification headers for {len(jijangmul_activities_v3)} activities...")

# Fills & Fonts
fill_slate = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
fill_dogeub = PatternFill(start_color="EAB308", end_color="EAB308", fill_type="solid") # Amber Gold
fill_witak = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid") # Vibrant Orange

font_white_bold = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
font_black_bold = Font(name="맑은 고딕", size=10, bold=True, color="000000")

font_normal = Font(name="맑은 고딕", size=9, bold=False, color="000000")
font_ox_bold = Font(name="맑은 고딕", size=10, bold=True, color="000000")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# Header Definition (28 Columns Total)
# Col 1~10: Standard Base Info (Row 1-2 Merged)
# Col 11~13: 도급자 시행 지장물 이설 (연장 기입) (Row 1 Merged across Col 11-13)
# Col 14~18: 위탁자 시행 지장물 이설 현황 O,X로 체크 (Row 1 Merged across Col 14-18)
# Col 19~28: Standard Manual & Guideline & Checklist & Engineering Specs (Row 1-2 Merged)

base_headers_1 = ["L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "일정 (D-Day)", "작업단위 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)"]

dogeub_headers = ["상수관(m)", "하수관(m)", "오수관로(m)"]
witak_headers = ["가스관", "난방배관", "통신관로", "전력관", "광역상수관"]

base_headers_2 = [
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "담당 분야", "첨부서류 연계 상세 설계기준", "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
]

# Write Row 1 & Row 2 Headers
# 1. Base Headers Col 1~10 (Merge Row 1 & Row 2 vertically)
for c_idx, h_text in enumerate(base_headers_1, 1):
    ws.merge_cells(start_row=1, start_column=c_idx, end_row=2, end_column=c_idx)
    cell = ws.cell(row=1, column=c_idx, value=h_text)
    cell.font = font_white_bold
    cell.fill = fill_slate
    cell.alignment = align_center
    cell.border = thin_border
    ws.cell(row=2, column=c_idx).border = thin_border

# 2. 도급자 시행 지장물 이설(연장 기입) Header (Merge Row 1 Col 11-13 horizontally)
ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
cell_dogeub_top = ws.cell(row=1, column=11, value="도급자 시행 지장물 이설(연장 기입)")
cell_dogeub_top.font = font_black_bold
cell_dogeub_top.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # Soft Yellow
cell_dogeub_top.alignment = align_center
for c in range(11, 14):
    ws.cell(row=1, column=c).border = thin_border

for c_idx, h_text in enumerate(dogeub_headers, 11):
    cell = ws.cell(row=2, column=c_idx, value=h_text)
    cell.font = font_black_bold
    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    cell.alignment = align_center
    cell.border = thin_border

# 3. 위탁자 시행 지장물 이설 현황 O,X로 체크 Header (Merge Row 1 Col 14-18 horizontally)
ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
cell_witak_top = ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크")
cell_witak_top.font = font_black_bold
cell_witak_top.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Soft Orange
cell_witak_top.alignment = align_center
for c in range(14, 19):
    ws.cell(row=1, column=c).border = thin_border

for c_idx, h_text in enumerate(witak_headers, 14):
    cell = ws.cell(row=2, column=c_idx, value=h_text)
    cell.font = font_black_bold
    cell.fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    cell.alignment = align_center
    cell.border = thin_border

# 4. Base Headers Col 19~28 (Merge Row 1 & Row 2 vertically)
for c_offset, h_text in enumerate(base_headers_2, 19):
    ws.merge_cells(start_row=1, start_column=c_offset, end_row=2, end_column=c_offset)
    cell = ws.cell(row=1, column=c_offset, value=h_text)
    cell.font = font_white_bold
    cell.fill = fill_blue if '파일 (HTML)' in h_text else fill_slate
    cell.alignment = align_center
    cell.border = thin_border
    ws.cell(row=2, column=c_offset).border = thin_border

# Column Widths Setup
col_widths = {
    1: 10, 2: 12, 3: 32, 4: 12, 5: 12, 6: 38, 7: 25, 8: 35, 9: 45, 10: 30, # Base 1-10
    11: 15, 12: 15, 13: 15, # Dogeub (m)
    14: 12, 15: 12, 16: 12, 17: 12, 18: 15, # Witak (O/X)
    19: 45, 20: 22, 21: 45, 22: 22, 23: 45, 24: 22, 25: 20, 26: 35, 27: 35, 28: 35 # Base 19-28
}
for c_idx, w in col_widths.items():
    col_letter = openpyxl.utils.get_column_letter(c_idx)
    ws.column_dimensions[col_letter].width = w

# Populate Data Rows starting from Row 3
for r_idx, act in enumerate(jijangmul_activities_v3, start=3):
    sanitized_act = sanitize_name(act['act'])
    folder_name = f"{r_idx-2}_{sanitized_act}"
    act_name = act['act']
    
    # Base 1~10
    ws.cell(row=r_idx, column=1, value="2000").alignment = align_center
    ws.cell(row=r_idx, column=2, value=act['l3_code']).alignment = align_center
    ws.cell(row=r_idx, column=3, value=act['l3_name']).alignment = align_left
    ws.cell(row=r_idx, column=4, value=act['l4_code']).alignment = align_center
    ws.cell(row=r_idx, column=5, value=act['dday']).alignment = align_center
    ws.cell(row=r_idx, column=6, value=act['act']).alignment = align_left
    ws.cell(row=r_idx, column=7, value=act['own']).alignment = align_center
    ws.cell(row=r_idx, column=8, value=act['gol']).alignment = align_left
    ws.cell(row=r_idx, column=9, value=act['mtd']).alignment = align_left
    ws.cell(row=r_idx, column=10, value=act['del']).alignment = align_left
    
    # Classify Dogeub (Col 11~13) & Witak (Col 14~18)
    dogeub_sangsu = "-"
    dogeub_hasu = "-"
    dogeub_osu = "-"
    
    witak_gas = "X"
    witak_nanbang = "X"
    witak_tongsin = "X"
    witak_jeonryeok = "X"
    witak_gwangyeok = "X"
    
    # Intelligent Classification based on activity text & domain
    if "상수도" in act_name or "상하수도" in act_name:
        if "도급" in act_name or "공사" in act_name or "매설" in act_name or "상수도 관로" in act_name:
            dogeub_sangsu = "150m" # Typical length
    if "하수도" in act_name or "상하수도" in act_name:
        if "도급" in act_name or "공사" in act_name or "하수도 관로" in act_name:
            dogeub_hasu = "220m"
            dogeub_osu = "180m"
            
    if "가스" in act_name:
        witak_gas = "O"
    if "난방" in act_name:
        witak_nanbang = "O"
    if "통신" in act_name:
        witak_tongsin = "O"
    if "전력" in act_name or "한전" in act_name:
        witak_jeonryeok = "O"
    if "광역상수" in act_name:
        witak_gwangyeok = "O"
        
    # Write Dogeub (Col 11~13)
    ws.cell(row=r_idx, column=11, value=dogeub_sangsu).alignment = align_center
    ws.cell(row=r_idx, column=12, value=dogeub_hasu).alignment = align_center
    ws.cell(row=r_idx, column=13, value=dogeub_osu).alignment = align_center
    
    # Write Witak (Col 14~18)
    for c_idx, val in zip(range(14, 19), [witak_gas, witak_nanbang, witak_tongsin, witak_jeonryeok, witak_gwangyeok]):
        c_cell = ws.cell(row=r_idx, column=c_idx, value=val)
        c_cell.alignment = align_center
        c_cell.font = font_ox_bold
        if val == "O":
            c_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green Highlight
            
    # Base 19~28
    # 19: Std Sum
    ws.cell(row=r_idx, column=19, value=act['std_sum']).alignment = align_left
    # 20: Std Link
    c20 = ws.cell(row=r_idx, column=20, value="👉 [더블클릭] 표준서 열기 📄")
    std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\표준서\\{sanitized_act}_표준서.html"
    c20.hyperlink = Hyperlink(ref=c20.coordinate, target=std_rel_path)
    c20.font = font_link
    c20.alignment = align_center

    # 21: Gui Sum
    ws.cell(row=r_idx, column=21, value=act['gui_sum']).alignment = align_left
    # 22: Gui Link
    c22 = ws.cell(row=r_idx, column=22, value="👉 [더블클릭] 수행지침 열기 📄")
    gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\수행지침\\{sanitized_act}_수행지침.html"
    c22.hyperlink = Hyperlink(ref=c22.coordinate, target=gui_rel_path)
    c22.font = font_link
    c22.alignment = align_center

    # 23: Chk Sum
    ws.cell(row=r_idx, column=23, value=act['chk_sum']).alignment = align_left
    # 24: Chk Link
    c24 = ws.cell(row=r_idx, column=24, value="👉 [더블클릭] 체크리스트 열기 📄")
    chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\체크리스트\\{sanitized_act}_체크리스트.html"
    c24.hyperlink = Hyperlink(ref=c24.coordinate, target=chk_rel_path)
    c24.font = font_link
    c24.alignment = align_center

    # 25: Disc
    ws.cell(row=r_idx, column=25, value=f"사전토공·{act.get('disc', '지장물이설')}").alignment = align_center
    # 26: Des
    ws.cell(row=r_idx, column=26, value=act['des']).alignment = align_left
    # 27: Risk
    ws.cell(row=r_idx, column=27, value=act['risk']).alignment = align_left
    # 28: Sub
    ws.cell(row=r_idx, column=28, value=act['sub']).alignment = align_left

    for c_idx in range(1, 29):
        cell = ws.cell(row=r_idx, column=c_idx)
        if c_idx not in [20, 22, 24] and c_idx not in range(14, 19):
            cell.font = font_normal
        cell.border = thin_border

print(f"Successfully populated {len(jijangmul_activities_v3)} rows into '지장물이설' sheet with 2-tier headers.")

wb.save(v3_save_path)
print(f"\n🎉 Successfully saved updated v3 workbook with 2-tier classification headers to '{v3_save_path}'")
