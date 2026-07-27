import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheet_name in ['상부강화노반', '콘크리트도상', '신호분야']:
    ws = wb[sheet_name]
    header_row = 1
    print(f"\n=== Verification for Sheet '{sheet_name}' (Row 2 Sample) ===")
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").replace('\n', ' ') for c in range(1, ws.max_column + 1)]
    
    std_sum_c = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '요약' in h), None)
    std_link_c = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '파일' in h), None)
    
    sum_cell = ws.cell(row=2, column=std_sum_c)
    link_cell = ws.cell(row=2, column=std_link_c)
    
    print(f"표준서 요약 셀 내용 (Lines={len(str(sum_cell.value).splitlines())}):")
    print(sum_cell.value)
    print(f"표준서 파일 링크 셀: Value='{link_cell.value}' | HasLink={link_cell.hyperlink is not None}")
