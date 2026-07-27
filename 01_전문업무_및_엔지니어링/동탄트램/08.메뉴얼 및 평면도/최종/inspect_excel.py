import openpyxl

wb = openpyxl.load_workbook("매뉴얼 BODY (집행단계).xlsx", read_only=True)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet: {name}")
    # Print first 5 rows
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows: {len(rows)}")
    for row in rows[:10]:
        print(row)
