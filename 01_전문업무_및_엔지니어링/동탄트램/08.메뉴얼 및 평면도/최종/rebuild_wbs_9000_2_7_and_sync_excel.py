import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

target_folder = None
for f in os.listdir(base_dir):
    if f.startswith("7_") or ("발주처" in f and "품질" in f):
        target_folder = os.path.join(base_dir, f)
        break

if not target_folder:
    print("❌ ERROR: Target folder for WBS 9000-2-7 not found!")
    sys.exit(1)

print(f"Target WBS 9000-2-7 Folder: {target_folder}")

zoom_modal_style = """
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .clickable-diagram::after {
        content: "🔍 클릭하여 대형 확대보기";
        position: absolute;
        bottom: 8px;
        right: 12px;
        background: rgba(15, 23, 42, 0.75);
        color: #ffffff;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        backdrop-filter: blur(4px);
        pointer-events: none;
        opacity: 0.85;
        transition: opacity 0.2s;
    }
    .clickable-diagram:hover::after {
        opacity: 1;
        background: rgba(2, 132, 199, 0.9);
    }
    .glossary-modal, .zoom-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .glossary-modal.active, .zoom-modal.active {
        display: flex;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 550px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-close, .zoom-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .glossary-close:hover, .zoom-close:hover {
        color: #ef4444;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal">
    <div class="glossary-modal-content">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 품질 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeZoomModal();
    }
});
</script>
"""

# 1. STANDARD HTML
std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 발주처 품질 요구사항 검토 표준서 (WBS 9000-2-7)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard (WBS 9000-2-7)</span>
        <h1 class="text-3xl font-black mt-2">발주처 품질 요구사항 검토 엔지니어링 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-7 | 주관: 현장 시스템팀 / 통신 감리단 | "발주처 ITP/ITC 및 광/무선 품질기준 검토"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 1. 개요 및 목적 -->
        <div class="space-y-4">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-blue-600 pb-2">1. 개요 및 엔지니어링 목적</h2>
            <div class="bg-slate-50 p-6 rounded-2xl border border-slate-200 text-sm text-slate-700 leading-relaxed space-y-3">
                <p><strong>💡 초직관 설명:</strong> 발주처가 요구하는 자재 및 시공 품질 검사 지침(ITP/ITC)을 사전 검토하여, 72-Core 광케이블 접속 손실(≤0.05dB) 및 LTE-R 무선망 수신 전계강도 등 법정 품질 기준에 미달하는 시공 결함을 방지하기 위한 엔지니어링 표준서입니다.</p>
                <div class="grid grid-cols-1 sm:grid-cols-2 gap-3 pt-2">
                    <div class="bg-white p-4 rounded-xl border border-slate-200 shadow-sm">
                        <span class="text-blue-600 font-bold text-xs">📋 발주처 ITP/ITC 품질 지침</span>
                        <p class="text-xs text-slate-800 font-bold mt-1">자재 및 시험검사 계획 사전 대조 및 승인</p>
                    </div>
                    <div class="bg-white p-4 rounded-xl border border-slate-200 shadow-sm">
                        <span class="text-emerald-600 font-bold text-xs">⚡ 광 접속 & 무선 품질</span>
                        <p class="text-xs text-slate-800 font-bold mt-1">광 손실 ≤0.05dB 및 무선국 준공 검증</p>
                    </div>
                </div>
            </div>
        </div>

        <!-- 🎨 2. 2D Visual SVG 기술 모식도 -->
        <div class="space-y-4">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-indigo-600 pb-2">2. ITP/ITC 품질검사 체계 2D 기술 모식도</h2>
            <div class="clickable-diagram bg-slate-50 p-6 rounded-2xl border border-slate-200 shadow-inner" onclick="openDiagramZoom('std_svg', '발주처 ITP/ITC 품질 검측 체계도')">
                <svg id="std_svg" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                    <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                    
                    <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                    <text x="130" y="50" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📋 발주처 품질검사 계획(ITP)</text>
                    <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• ITP/ITC 검측 절차서 사전 검토</text>
                    <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 시험 항목 및 시공 기준 대조</text>
                    <text x="35" y="122" font-size="11" font-weight="bold" fill="#1d4ed8">• 검측 서류 100% 준비 완료</text>

                    <path d="M 245 90 L 275 90" stroke="#2563eb" stroke-width="3"/>
                    <polygon points="275,85 285,90 275,95" fill="#2563eb"/>

                    <rect x="290" y="25" width="210" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                    <text x="395" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">⚡ 정량적 품질 검측 (ITC)</text>
                    <text x="305" y="78" font-size="11" font-weight="bold" fill="#334155">• 광 접속 손실: ≤ 0.05dB</text>
                    <text x="305" y="100" font-size="11" font-weight="bold" fill="#334155">• LTE-R 커버리지: 수신 전계 통과</text>
                    <text x="305" y="122" font-size="11" font-weight="black" fill="#047857">• ✔ 발주처 최종 서명 체결</text>
                </svg>
            </div>
            <div class="bg-blue-50 p-4 rounded-xl text-xs text-blue-900 border border-blue-200">
                <strong>📌 핵심 기준 요약:</strong> 발주처 ITP/ITC 품질 지침에 의거하여 광케이블 접속 손실 <strong>≤ 0.05dB</strong> 및 무선국 준공 검사 전파 적합 필증을 최종 승인합니다.
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 2. GUIDELINE HTML (Flexible 4-Step Architecture)
gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 발주처 품질 요구사항 검토 수행지침서 (WBS 9000-2-7)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Guideline (WBS 9000-2-7)</span>
        <h1 class="text-3xl font-black mt-2">발주처 품질 요구사항 검토 초직관 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-7 | 주관: 현장 시스템팀 / 통신 감리단 | "사안에 부합하는 4단계 유연 프로세스 수록"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 친절한 개요 해설 박스 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 한눈에 읽는 유연한 실무 가이드 (Flexible 4-Step)</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 지침서는 발주처 품질 검토 업무에 맞춘 <strong>"4단계 유연 프로세스(Flexible 4-Step Architecture)"</strong>로 구성되었습니다.
                아래의 <strong>Step별 1:1 직관적 2D Visual 그림</strong>을 보며 점검을 수행하십시오.
                모든 그림은 <strong><span class="term-highlight" onclick="openDiagramZoom('svg_step1', 'STEP 1 ITP/ITC 서류 대조 도식')">클릭하면 대형 팝업 모달로 크게 확장</span></strong>됩니다.
            </p>
        </div>

        <!-- ☀️ 4대 유연 핵심 프로세스 카드 -->
        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">FLEXIBLE PROCESS</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 발주처 품질 검토 4단계 유연 프로세스</h3>
            </div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📋</span> 1. ITP/ITC 계획서 사전 대조</span>
                    <p class="text-slate-700 text-xs">발주처 자재 및 시공 품질검사 계획서(ITP/ITC) 항목 및 시험 기준표 사전 검토.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>⚡</span> 2. 광케이블 접속 손실 0.05dB 측정</span>
                    <p class="text-slate-700 text-xs">72-Core 메인 광케이블 융착 접속 손실 0.05dB 이하 OTDR 측정 확인.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📶</span> 3. LTE-R 무선국 준공 검증</span>
                    <p class="text-slate-700 text-xs">LTE-R 안테나 수신 전계강도 및 무선국 준공검사 전파 품질 정밀 측정.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🖊️</span> 4. 품질 승인서(ITC) 최종 체결</span>
                    <p class="text-slate-700 text-xs">발주처·감리원·시스템팀 3자 서명 날인 및 최종 품질 관리대장 확정.</p>
                </div>
            </div>
        </div>

        <!-- 1. FLEXIBLE 4-STEP ARCHITECTURE -->
        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> 4단계 핵심 수행 마스터 프로세스 (Flexible Architecture)
            </h2>
            <div class="grid grid-cols-1 sm:grid-cols-2 md:grid-cols-4 gap-3">
                <div class="bg-blue-50 p-4 rounded-xl border border-blue-200 flex flex-col justify-between">
                    <span class="bg-blue-600 text-white text-[10px] font-black px-2 py-0.5 rounded w-fit">STEP 1</span>
                    <h4 class="font-bold text-slate-900 text-xs mt-2">ITP/ITC 사전대조</h4>
                    <p class="text-[11px] text-blue-900 mt-1 font-medium">• 품질계획서 대조<br">• 검측 항목 확정</p>
                </div>
                <div class="bg-indigo-50 p-4 rounded-xl border border-indigo-200 flex flex-col justify-between">
                    <span class="bg-indigo-600 text-white text-[10px] font-black px-2 py-0.5 rounded w-fit">STEP 2</span>
                    <h4 class="font-bold text-slate-900 text-xs mt-2">광 손실 측정</h4>
                    <p class="text-[11px] text-indigo-900 mt-1 font-medium">• 광 손실 ≤0.05dB<br">• OTDR 측정 판정</p>
                </div>
                <div class="bg-cyan-50 p-4 rounded-xl border border-cyan-200 flex flex-col justify-between">
                    <span class="bg-cyan-600 text-white text-[10px] font-black px-2 py-0.5 rounded w-fit">STEP 3</span>
                    <h4 class="font-bold text-slate-900 text-xs mt-2">LTE-R 무선검증</h4>
                    <p class="text-[11px] text-cyan-900 mt-1 font-medium">• 수신 전계 강도<br">• 무선국 준공검사</p>
                </div>
                <div class="bg-emerald-50 p-4 rounded-xl border border-emerald-200 flex flex-col justify-between">
                    <span class="bg-emerald-600 text-white text-[10px] font-black px-2 py-0.5 rounded w-fit">STEP 4</span>
                    <h4 class="font-bold text-slate-900 text-xs mt-2">품질승인 체결</h4>
                    <p class="text-[11px] text-emerald-900 mt-1 font-medium">• ITC 최종 서명<br">• 품질대장 확정</p>
                </div>
            </div>
        </div>

        <!-- 🔥 2. 초정밀 HOW 세부 실무 가이드 (STEP 1~4 1:1 직관적 2D VISUAL SVG) -->
        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-indigo-600 pb-2">
                <span class="text-indigo-600">2.</span> 유연 4단계 HOW 세부 실무 가이드 & 1:1 2D 그림
            </h2>

            <div class="space-y-8 text-sm">
                <!-- STEP 1 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-blue-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-blue-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 1</span>
                        <h3 class="font-bold text-base text-slate-900">발주처 품질검사 계획서(ITP/ITC) 항목 대조 & 시험 기준표 사전 검토</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 발주처 시방서 및 품질검사 지침(ITP/ITC) 항목을 사전 대조하여, 통신 자재 입고부터 통신선 포설·접속·시험까지의 검측 기준을 사전에 확정합니다.</p>
                    </div>
                    
                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step1', 'STEP 1 ITP/ITC 계획서 사전 대조 2D 시공 도식')">
                        <svg id="svg_step1" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                            <text x="130" y="48" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📋 ITP/ITC 품질검사 지침서</text>
                            <g transform="translate(35, 62)">
                                <rect x="0" y="0" width="190" height="25" fill="#e0f2fe" stroke="#2563eb" stroke-width="1.5" rx="3"/>
                                <text x="95" y="17" font-size="10" font-weight="bold" fill="#1d4ed8" text-anchor="middle">✔ 자재 검수 기준 100% 대조</text>

                                <rect x="0" y="32" width="190" height="25" fill="#e0f2fe" stroke="#2563eb" stroke-width="1.5" rx="3"/>
                                <text x="95" y="49" font-size="10" font-weight="bold" fill="#1d4ed8" text-anchor="middle">✔ 시공 검측 서류 100% 준비</text>
                            </g>

                            <path d="M 245 90 L 285 90" stroke="#2563eb" stroke-width="3"/>
                            <polygon points="285,85 295,90 285,95" fill="#2563eb"/>

                            <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="400" y="48" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📑 검측 시험 항목 확정</text>
                            <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 광 손실 ≤ 0.05dB 검사</text>
                            <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• LTE-R 전파 커버리지 검사</text>
                            <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 사전 승인 완료</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 2 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-indigo-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-indigo-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 2</span>
                        <h3 class="font-bold text-base text-slate-900">72-Core 메인 광케이블 접속 손실 0.05dB 이하 최상급 측정 시험</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 메인 72-Core 광케이블 접속 시 융착 접속기 및 OTDR 측정 장비로 1개 Core당 광 손실이 0.05dB 이하로 깨끗한지 정밀 측정합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step2', 'STEP 2 광케이블 접속 손실 0.05dB 측정 2D 시공 도식')">
                        <svg id="svg_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                            <text x="130" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">⚡ OTDR 광파워미터 측정</text>
                            <rect x="40" y="65" width="180" height="40" fill="#e0e7ff" stroke="#4f46e5" stroke-width="1.5" rx="4"/>
                            <text x="130" y="85" font-size="11" font-weight="black" fill="#3730a3" text-anchor="middle">접속 손실: 0.03 dB</text>
                            <text x="130" y="98" font-size="9" font-weight="bold" fill="#15803d" text-anchor="middle">(기준 0.05dB 이하 합격)</text>
                            <text x="130" y="135" font-size="10" font-weight="bold" fill="#334155" text-anchor="middle">72-Core 전수 파장 측정 통과</text>

                            <path d="M 245 90 L 285 90" stroke="#4f46e5" stroke-width="3"/>
                            <polygon points="285,85 295,90 285,95" fill="#4f46e5"/>

                            <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📊 광 성적서 승인</text>
                            <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 1Core ~ 72Core 무결점</text>
                            <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 광파장 1310/1550nm 합격</text>
                            <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ ITP 승인 필증 획득</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 3 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-cyan-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-cyan-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 3</span>
                        <h3 class="font-bold text-base text-slate-900">LTE-R 안테나 수신 전계강도 & 무선국 준공검사 전파 품질 측정</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> LTE-R 트램 전용 무선망 안테나 전파 수신 강도를 측정하여 음영 지역이 없는지 검사하고, 무선국 준공검사 전파 필증 서류를 준비합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step3', 'STEP 3 LTE-R 무선국 전파 품질 측정 2D 시공 도식')">
                        <svg id="svg_step3" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
                            <text x="130" y="50" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">📶 LTE-R 전파 전계 측정</text>
                            <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 수신 전계강도: ≥ -95dBm</text>
                            <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 음영지역 커버리지 100%</text>
                            <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 무선 커버리지 통과</text>

                            <path d="M 245 90 L 285 90" stroke="#0891b2" stroke-width="3"/>
                            <polygon points="285,85 295,90 285,95" fill="#0891b2"/>

                            <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📑 무선국 준공 검증</text>
                            <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 주파수 편차 적합</text>
                            <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 법정 합격 필증 확보</text>
                            <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 발주처 보고 완료</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 4 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-emerald-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-emerald-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 4</span>
                        <h3 class="font-bold text-base text-slate-900">발주처·감리원 품질 검측 승인서(ITC) 체결 & 품질 대장 확정</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 모든 품질 검측(ITP/ITC)이 통과되면 발주처, 감리원, 현장 시스템팀이 최종 검측 승인서(ITC)에 공동 서명하고 품질 대장을 확정합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step4', 'STEP 4 품질 검측 승인서(ITC) 최종 서명 체결 2D 시공 도식')">
                        <svg id="svg_step4" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="130" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🔴 ITP/ITC 합격 승인인</text>
                            <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 광/무선 품질 검측 통과</text>
                            <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 시방 규격 100% 만족</text>
                            <text x="130" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 품질 검측 승인 완료</text>

                            <rect x="260" y="25" width="240" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="380" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🖊️ 발주처·감리원 서명 체결</text>
                            <g transform="translate(295, 68)">
                                <circle cx="25" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="25" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">발주처</text>

                                <circle cx="85" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="85" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">감리원</text>

                                <circle cx="145" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="145" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">시스템</text>
                            </g>
                            <text x="380" y="122" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 품질 대장 최종 확정 완료</text>
                        </svg>
                    </div>
                </div>
            </div>
        </div>

        <!-- 3. 종합 2D VISUAL SVG DIAGRAM -->
        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-emerald-600 pb-2">
                <span class="text-emerald-600">3.</span> 종합 2D Visual 기술 도식 (Enriched 2D SVG)
            </h2>
            <div class="clickable-diagram bg-slate-50 p-5 rounded-2xl border border-slate-200 shadow-inner" onclick="openDiagramZoom('svg_r4', '[WBS 9000-2-7] 발주처 품질 요구사항 검토 초직관 종합 도식')">
                <svg id="svg_r4" viewBox="0 0 550 190" width="100%" height="190" xmlns="http://www.w3.org/2000/svg">
                    <rect x="0" y="0" width="550" height="190" fill="#f8fafc"/>
                    <rect x="25" y="15" width="230" height="145" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                    <text x="140" y="42" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📋 ITP/ITC & 광 접속 검사</text>
                    <text x="40" y="70" font-size="11" font-weight="bold" fill="#334155">• 발주처 품질 계획서(ITP) 대조</text>
                    <text x="40" y="93" font-size="11" font-weight="bold" fill="#334155">• 72-Core 광케이블 접속 손실</text>
                    <text x="40" y="116" font-size="11" font-weight="bold" fill="#1d4ed8">• 광 손실 ≤ 0.05dB 정밀 승인</text>
                    <text x="40" y="139" font-size="11" font-weight="bold" fill="#334155">• OTDR 및 광파워미터 통과</text>

                    <path d="M 265 85 L 295 85" stroke="#2563eb" stroke-width="3"/>
                    <polygon points="295,80 305,85 295,90" fill="#2563eb"/>

                    <rect x="310" y="15" width="215" height="145" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                    <text x="417" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📶 LTE-R & 서명 체결</text>
                    <text x="325" y="70" font-size="11" font-weight="bold" fill="#334155">• LTE-R 무선국 준공 검증 필증</text>
                    <text x="325" y="93" font-size="11" font-weight="bold" fill="#334155">• 무선 커버리지 100% 확보</text>
                    <text x="325" y="116" font-size="11" font-weight="bold" fill="#334155">• ITC 최종 검측 승인서 작성</text>
                    <text x="325" y="139" font-size="11" font-weight="bold" fill="#047857">• 발주처·감리원 3자 서명 날인</text>
                    <text x="275" y="175" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">WBS 9000-2-7 발주처 품질 요구사항 검토 100% 확정</text>
                </svg>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 3. CHECKLIST HTML (3-Column Master Template & Interrogative Phrasing `~하였는가?`)
chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 발주처 품질 요구사항 검토 체크리스트 (WBS 9000-2-7)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Checklist (WBS 9000-2-7)</span>
        <h1 class="text-3xl font-black mt-2">발주처 품질 요구사항 검토 현장 점검 체크리스트</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-7 | 주관: 현장 시스템팀 / 통신 감리단 | "종결 어미 ~하였는가? 100% 적용"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-8">
        <!-- 💡 현장 점검 안내 및 대형 팝업 모달 연동 버튼 -->
        <div class="bg-blue-50 border border-blue-200 p-5 rounded-xl flex flex-col sm:flex-row items-center justify-between gap-4">
            <div>
                <h4 class="font-bold text-slate-900 text-sm">🔍 ITP/ITC 품질검사 기술 도식 대형 확대보기</h4>
                <p class="text-xs text-slate-600 mt-1">시공 도식을 대형 모달 창으로 띄워 정밀 점검을 수행하십시오.</p>
            </div>
            <button onclick="openDiagramZoom('chk_svg', 'WBS 9000-2-7 품질 검측 정밀 도식')" class="bg-blue-600 hover:bg-blue-700 text-white font-bold text-xs px-4 py-2.5 rounded-lg shadow transition flex items-center gap-1.5 whitespace-nowrap">
                <span>🔍</span> 시공 도식 열기
            </button>
        </div>

        <!-- HIDDEN SVG FOR ZOOM MODAL -->
        <div class="hidden">
            <svg id="chk_svg" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                <text x="130" y="50" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📋 발주처 ITP/ITC 품질 계획</text>
                <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• ITP 검측 절차서 100% 대조</text>
                <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 자재/시공 시험기준 사전 승인</text>

                <rect x="270" y="25" width="230" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="385" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">⚡ 정량적 품질 검측 통과</text>
                <text x="285" y="78" font-size="11" font-weight="bold" fill="#334155">• 광 접속 손실: ≤ 0.05dB</text>
                <text x="285" y="100" font-size="11" font-weight="bold" fill="#334155">• LTE-R 무선국 준공 검사 완료</text>
                <text x="385" y="128" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 발주처·감리원 서명 체결</text>
            </svg>
        </div>

        <!-- 📝 3-COLUMN MASTER CHECKLIST TABLE -->
        <div class="overflow-x-auto rounded-xl border border-slate-200">
            <table class="w-full text-left text-sm text-slate-700">
                <thead class="bg-slate-900 text-white text-xs uppercase font-bold">
                    <tr>
                        <th class="p-3.5 text-center w-14">번호</th>
                        <th class="p-3.5">점검 항목 및 검측 세부 기준 (질문형 ~하였는가?)</th>
                        <th class="p-3.5 text-center w-24">점검 결과</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-200 bg-white">
                    <!-- 그룹 1: 사전 수립 및 ITP/ITC 절차 검토 -->
                    <tr class="bg-blue-50/50">
                        <td colspan="3" class="p-3 font-bold text-blue-900 text-xs">■ 1. 사전 수립 및 ITP/ITC 품질 계획서 대조</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">1</td>
                        <td class="p-3.5 font-medium">발주처 품질검사 계획서(ITP/ITC) 항목 및 시험검사 지침서를 사전에 확보하여 검토하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">2</td>
                        <td class="p-3.5 font-medium">통신 자재 입고 시 공장검과 품증 서류 및 발주처 승인 자재인지 대조하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">3</td>
                        <td class="p-3.5 font-medium">시공 검측 단계별 감리원 입회 시험 항목 및 ITP 양식을 정립하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>

                    <!-- 그룹 2: 72-Core 광케이블 & 무선망 품질 검측 -->
                    <tr class="bg-indigo-50/50">
                        <td colspan="3" class="p-3 font-bold text-indigo-900 text-xs">■ 2. 72-Core 광케이블 접속 & LTE-R 무선 전파 품질 검측</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">4</td>
                        <td class="p-3.5 font-medium">72-Core 메인 광케이블 융착 접속 시 접속 손실이 0.05dB 이하 기준을 만족하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">5</td>
                        <td class="p-3.5 font-medium">OTDR 및 광파워미터 측정 결과 리포트를 출력하여 파장별(1310/1550nm) 손실을 확인하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">6</td>
                        <td class="p-3.5 font-medium">LTE-R 안테나 및 기지국 수신 전계강도가 음영지역 없이 정량 기준을 만족하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">7</td>
                        <td class="p-3.5 font-medium">무선국 준공검사 및 전파법 관련 성능 검증 필증 서류를 확보하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>

                    <!-- 그룹 3: CCTV/PIS/PA 및 최종 ITC 승인 서명 -->
                    <tr class="bg-emerald-50/50">
                        <td colspan="3" class="p-3 font-bold text-emerald-900 text-xs">■ 3. CCTV/방송 품질 연동 & ITC 최종 승인 서명</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">8</td>
                        <td class="p-3.5 font-medium">4K IP CCTV 카메라 영상 스트리밍 화질 및 프레임 지연을 검측하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">9</td>
                        <td class="p-3.5 font-medium">승강장 안내전광판(PIS) 글자 표출 및 스피커(PA) 음압 품질을 검사하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">10</td>
                        <td class="p-3.5 font-medium">품질 검측 승인서(ITC) 양식에 발주처, 감리원, 현장 책임자 3자 서명을 체결하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">11</td>
                        <td class="p-3.5 font-medium">품질 검측 결함 사항에 대한 시정 조치 보고서를 작성하고 준공 대장에 반영하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                    <tr>
                        <td class="p-3.5 text-center font-bold text-xs text-slate-500">12</td>
                        <td class="p-3.5 font-medium">최종 발주처 품질 요구사항 검토 보고서를 준공 보존용 문서로 확정 등록하였는가?</td>
                        <td class="p-3.5 text-center font-bold text-emerald-600">☐ 적합</td>
                    </tr>
                </tbody>
            </table>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# Write HTML Files for WBS 9000-2-7
files_to_write = {
    os.path.join(target_folder, "표준서", "9000-2-7_발주처 품질 요구사항 검토_표준서.html"): std_html,
    os.path.join(target_folder, "표준서", "발주처 품질 요구사항 검토_표준서.html"): std_html,
    os.path.join(target_folder, "수행지침", "9000-2-7_발주처 품질 요구사항 검토_수행지침.html"): gui_html,
    os.path.join(target_folder, "수행지침", "발주처 품질 요구사항 검토_수행지침.html"): gui_html,
    os.path.join(target_folder, "체크리스트", "9000-2-7_발주처 품질 요구사항 검토_체크리스트.html"): chk_html,
    os.path.join(target_folder, "체크리스트", "발주처 품질 요구사항 검토_체크리스트.html"): chk_html
}

for fp, content in files_to_write.items():
    os.makedirs(os.path.dirname(fp), exist_ok=True)
    with open(fp, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"   ✓ [REBUILT WBS 9000-2-7 WITH EASY GUIDE & VISUAL SVG] -> {os.path.basename(fp)}")

# Update Excel File (매뉴얼 BODY (집행단계)v4.xlsx) Row 7 (WBS 9000-2-7)
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        if "통신분야" in wb.sheetnames:
            ws = wb["통신분야"]
            
            std_summary = "1) 품질 요구사항 검토: 발주처 자재/시공 품질검사계획(ITP/ITC) 및 72-Core 광케이블/LTE-R 전파법 관련 현장 품질 시험 항목을 전면 검토함.\n2) 품질 보장: 현장 시스템팀 주관으로 광 접속 손실(≤0.05dB) 및 무선 커버리지 측정 기준을 수립하여 품질검사의 완벽성을 확보함."
            gui_summary = "1) ITP/ITC 항목 검토: 발주처 자재/시공 품질검사(ITP/ITC) 항목 및 시험 기준 사전 대조\n2) 법정 품질 검사: 광케이블 OTDR 시험, LTE-R 주파수 편차 및 무선국 준공검사 기준을 수립하는 4단계 유연 품질 가이드."
            chk_summary = "1) 발주처 품질지침 적용: ITP/ITC 및 무선 설비별 품질 검사 항목을 정립 검토하였는가?\n2) 광 접속 손실(≤0.05dB) 및 LTE-R 무선 커버리지 품질 기준을 확정하였는가?"

            ws.cell(row=7, column=10, value=std_summary) # Col J
            ws.cell(row=7, column=12, value=gui_summary) # Col L
            ws.cell(row=7, column=14, value=chk_summary) # Col N

            wb.save(excel_path)
            print("   ✓ [EXCEL V4 ROW 7 SYNC COMPLETE] Successfully updated Row 7 (WBS 9000-2-7) Cols J, L, N in 매뉴얼 BODY (집행단계)v4.xlsx")
    except Exception as e:
        print(f"   ⚠️ Excel Sync Note: {e}")

print("\n🎉 SUCCESSFULLY REBUILT WBS 9000-2-7 WITH EASY GUIDE & HIGHLY VISUAL SVG DIAGRAMS!")
