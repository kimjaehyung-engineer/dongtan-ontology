import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

def sanitize_name(name):
    return re.sub(r'[\/\\:\*\?"<>\|]', '_', str(name)).strip()

def extract_clean_summary(html_path, doc_type):
    if not os.path.exists(html_path):
        return None
        
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')
        
    items = []
    
    for li in soup.find_all('li'):
        txt = li.get_text().strip()
        txt = re.sub(r'^\s*[\d\.\-☐☑📍📌]\s*', '', txt).strip()
        txt = re.sub(r'\[동탄트램 업무 매뉴얼 v1 연계\]\s*', '', txt)
        txt = re.sub(r'\[동탄트램 매뉴얼 v1\]\s*', '', txt)
        txt = re.sub(r'\[설계사 작성\]\s*', '', txt)
        txt = txt.replace('기술기준 연계', '기술기준').replace('설계기준 연계', '설계기준')
        
        if txt and len(txt) > 5 and txt not in items:
            items.append(txt)
            
    if doc_type == '표준서':
        items = sorted(items, key=lambda x: 0 if any(k in x for k in ['KCS', 'KDS', '≥', '≤', '±', '%', 'MPa', 'mm', 'SIL']) else 1)
    elif doc_type == '수행지침':
        items = sorted(items, key=lambda x: 0 if any(k in x for k in ['단계', '준비', '시공', '확인', '검측', '수칙']) else 1)
    elif doc_type == '체크리스트':
        items = [x for x in items if any(k in x for k in ['인가', '검측', '확인', '승인', '되었는가', '했는가', '자문'])]
        if not items:
            for li in soup.find_all('li'):
                txt = li.get_text().strip()
                txt = re.sub(r'^\s*[\d\.\-☐☑📍📌]\s*', '', txt).strip()
                if txt: items.append(txt)
                
    selected = items[:5]
    if not selected:
        return None
        
    bullets = [f"{i+1}) {item}" for i, item in enumerate(selected)]
    return "\n".join(bullets)

wb = openpyxl.load_workbook(excel_path)

header_fill_link = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
header_fill_sum = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate Dark
font_white = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

link_font = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

total_links_created = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    
    # 1. Determine header row
    if sheet_name == '공정표에따른 매뉴얼':
        header_row = 3
        has_dday = False
    elif sheet_name in ['사전토공사', '상부강화노반', '콘크리트도상']:
        header_row = 1
        has_dday = True
    else: # 건축, 신호분야, 통신분야, 전기분야
        header_row = 1
        has_dday = False
        
    act_col = 6 if has_dday else 5
    l4_col = 4
    
    rows_data = []
    max_r = ws.max_row
    
    for r in range(header_row + 1, max_r + 1):
        act_val = ws.cell(row=r, column=act_col).value
        l4_val = ws.cell(row=r, column=l4_col).value
        if not act_val and not l4_val: continue
        
        row_dict = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            # Avoid reading merged cell values incorrectly
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                row_dict[c] = cell.value
        rows_data.append(row_dict)

    print(f"\nProcessing '{sheet_name}' (Total Data Rows: {len(rows_data)})...")

    if has_dday:
        c_l2, c_l3, c_l3n, c_l4, c_dday, c_act, c_own, c_gol, c_mtd, c_del = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10
        c_std, c_gui, c_chk = 11, 13, 18
        c_disc, c_des, c_risk, c_sub = 16, 17, 19, 20
    elif sheet_name == '공정표에따른 매뉴얼':
        c_l2, c_l3, c_l3n, c_l4, c_act, c_own, c_gol, c_mtd, c_del = 1, 2, 3, 4, 5, 6, 7, 8, 9
        c_std, c_gui, c_chk = 10, 11, 12
        c_disc, c_des, c_risk, c_sub = 13, 14, 15, 16
        c_dday = None
    else:
        c_l2, c_l3, c_l3n, c_l4, c_act, c_own, c_gol, c_mtd, c_del = 1, 2, 3, 4, 5, 6, 7, 8, 9
        c_std, c_gui, c_chk = 10, 12, 17
        c_disc, c_des, c_risk, c_sub = 15, 16, 18, 19
        c_dday = None

    # Unmerge merged ranges if any in data area
    merged_ranges = list(ws.merged_cells.ranges)
    for mrange in merged_ranges:
        if mrange.min_row > header_row:
            ws.unmerge_cells(str(mrange))

    # Safely clear data rows
    for r in range(header_row + 1, max_r + 20):
        for c in range(1, ws.max_column + 10):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None
                cell.hyperlink = None

    # Define New Headers
    new_headers = []
    new_headers.extend(["L2 코드", "L3 코드", "L3 대공종명", "L4 코드"])
    if has_dday:
        new_headers.append("일정 (D-Day)")
    new_headers.extend([
        "작업단위 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)",
        "표준서 (Standard) 요약", "표준서 파일 (HTML)",
        "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
        "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
        "담당 분야", "첨부서류 연계 상세 설계기준", "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
    ])

    # Write Headers
    for c_idx, h_text in enumerate(new_headers, 1):
        cell = ws.cell(row=header_row, column=c_idx)
        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
            cell.value = h_text
            cell.font = font_white
            cell.alignment = align_center
            if '파일 (HTML)' in h_text:
                cell.fill = header_fill_link
            else:
                cell.fill = header_fill_sum

    # New column positions:
    n_act = 6 if has_dday else 5
    n_std_sum = 11 if has_dday else 10
    n_std_link = 12 if has_dday else 11
    n_gui_sum = 13 if has_dday else 12
    n_gui_link = 14 if has_dday else 13
    n_chk_sum = 15 if has_dday else 14
    n_chk_link = 16 if has_dday else 15
    n_disc = 17 if has_dday else 16
    n_des = 18 if has_dday else 17
    n_risk = 19 if has_dday else 18
    n_sub = 20 if has_dday else 19

    # Set Column Widths
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_std_sum)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_std_link)].width = 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_gui_sum)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_gui_link)].width = 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_chk_sum)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(n_chk_link)].width = 22

    sheet_act_count = 0
    disc_dir = sheet_name

    # Write Rows Data
    for r_idx, r_dict in enumerate(rows_data, start=header_row + 1):
        sheet_act_count += 1
        
        act_val = r_dict.get(c_act)
        l4_val = r_dict.get(c_l4)
        
        act_name = str(act_val).strip() if act_val else f"Task_{sheet_act_count}"
        sanitized_act = sanitize_name(act_name)
        folder_name = f"{sheet_act_count}_{sanitized_act}"
        
        if sheet_name == '공정표에따른 매뉴얼':
            found_disc = None
            found_sub = None
            for d in os.listdir(base_attach_dir):
                d_path = os.path.join(base_attach_dir, d)
                if os.path.isdir(d_path):
                    for sub in os.listdir(d_path):
                        if sanitized_act in sub or (str(l4_val) in sub if l4_val else False):
                            found_disc = d
                            found_sub = sub
                            break
                if found_disc: break
            if found_disc:
                disc_dir = found_disc
                folder_name = found_sub
            else:
                disc_dir = "사전토공사"

        act_dir_abs = os.path.join(base_attach_dir, disc_dir, folder_name)

        # Basic Columns
        ws.cell(row=r_idx, column=1, value=r_dict.get(c_l2)).alignment = align_center
        ws.cell(row=r_idx, column=2, value=r_dict.get(c_l3)).alignment = align_center
        ws.cell(row=r_idx, column=3, value=r_dict.get(c_l3n)).alignment = align_left
        ws.cell(row=r_idx, column=4, value=r_dict.get(c_l4)).alignment = align_center
        
        if has_dday:
            ws.cell(row=r_idx, column=5, value=r_dict.get(c_dday)).alignment = align_center
            
        ws.cell(row=r_idx, column=n_act, value=act_name).alignment = align_left
        ws.cell(row=r_idx, column=n_act + 1, value=r_dict.get(c_own)).alignment = align_center
        ws.cell(row=r_idx, column=n_act + 2, value=r_dict.get(c_gol)).alignment = align_left
        ws.cell(row=r_idx, column=n_act + 3, value=r_dict.get(c_mtd)).alignment = align_left
        ws.cell(row=r_idx, column=n_act + 4, value=r_dict.get(c_del)).alignment = align_left

        # Summaries and Links
        doc_tasks = [
            (c_std, n_std_sum, n_std_link, '표준서', '표준서'),
            (c_gui, n_gui_sum, n_gui_link, '수행지침', '수행지침'),
            (c_chk, n_chk_sum, n_chk_link, '체크리스트', '체크리스트')
        ]

        for orig_c, s_col, l_col, doc_type, doc_title in doc_tasks:
            orig_val = r_dict.get(orig_c)
            sub_folder = os.path.join(act_dir_abs, doc_type)
            
            clean_sum_text = None
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    clean_sum_text = extract_clean_summary(os.path.join(sub_folder, files[0]), doc_type)
                    
            if not clean_sum_text and orig_val:
                clean_sum_text = re.sub(r'--------------------------------------.*$', '', str(orig_val), flags=re.DOTALL).strip()
                
            # Summary Cell (NO HYPERLINK!)
            sum_cell = ws.cell(row=r_idx, column=s_col, value=clean_sum_text)
            sum_cell.hyperlink = None
            sum_cell.alignment = align_left
            sum_cell.border = thin_border
            
            # Link Cell (DEDICATED HYPERLINK!)
            link_cell = ws.cell(row=r_idx, column=l_col)
            link_cell.border = thin_border
            
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    raw_rel_target = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_dir}\\{folder_name}\\{doc_type}\\{files[0]}"
                    link_cell.value = f"👉 [더블클릭] {doc_title} 열기 📄"
                    link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=raw_rel_target)
                    link_cell.font = link_font
                    link_cell.alignment = align_center
                    total_links_created += 1

        # Extra metadata cols
        ws.cell(row=r_idx, column=n_disc, value=r_dict.get(c_disc)).alignment = align_center
        ws.cell(row=r_idx, column=n_des, value=r_dict.get(c_des)).alignment = align_left
        ws.cell(row=r_idx, column=n_risk, value=r_dict.get(c_risk)).alignment = align_left
        ws.cell(row=r_idx, column=n_sub, value=r_dict.get(c_sub)).alignment = align_left

        for c_idx in range(1, n_sub + 1):
            ws.cell(row=r_idx, column=c_idx).border = thin_border

print(f"\nPerfect layout reconstruction complete! Total {total_links_created} isolated hyperlinks created across dedicated columns.")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
