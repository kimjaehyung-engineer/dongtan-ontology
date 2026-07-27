import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

proc_ws = wb['공정매뉴얼']
print("Procurement/Process Manual Sheet Name:", proc_ws.title)
print("Max row:", proc_ws.max_row, "| Max col:", proc_ws.max_column)

header_row = 1
for r in range(1, 5):
    vals = [str(proc_ws.cell(row=r, column=c).value or "").strip() for c in range(1, proc_ws.max_column + 1)]
    if any('표준서' in v or '작업단위' in v for v in vals):
        header_row = r
        break

print(f"Header Row: {header_row}")
headers = [str(proc_ws.cell(row=header_row, column=c).value or "").replace('\n', ' ').strip() for c in range(1, proc_ws.max_column + 1)]
for idx, h in enumerate(headers, 1):
    print(f"  Col {idx}: '{h}'")

print("\nSampling first 10 rows in 공정매뉴얼:")
for r in range(header_row + 1, min(header_row + 15, proc_ws.max_row + 1)):
    c_val = proc_ws.cell(row=r, column=3).value # C col (L3)
    d_val = proc_ws.cell(row=r, column=4).value # D col (L4)
    e_val = proc_ws.cell(row=r, column=5).value # E col (Activity)
    f_val = proc_ws.cell(row=r, column=6).value # F col (Owner)
    print(f"Row {r:2d}: C(L3)='{c_val}' | D(L4)='{d_val}' | E(Act)='{e_val}' | F(Owner)='{f_val}'")
