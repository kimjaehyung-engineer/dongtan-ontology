import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

def keep_only_top_2_lines(val):
    if not val: return ""
    lines = str(val).splitlines()
    clean_items = []
    for l in lines:
        l_str = l.strip()
        # Remove any leading numbers or parenthesis
        l_str = re.sub(r'^\s*[\d\.\-\)\(☐☑📍📌]+\s*', '', l_str).strip()
        l_str = re.sub(r'^\s*[\d\.\-\)\(☐☑📍📌]+\s*', '', l_str).strip()
        if l_str and l_str not in clean_items:
            clean_items.append(l_str)
            
    # Take ONLY top 2 lines (1) and 2))
    selected = clean_items[:2]
    if not selected: return ""
    return "\n".join([f"{i+1}) {item}" for i, item in enumerate(selected)])

modified_std_gui_cells = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    header_row = 3 if sheet_name == '공정표에따른 매뉴얼' else 1
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    # Target ONLY '표준서' and '수행지침' summary columns (EXCLUDE '체크리스트' and '파일')
    target_cols = []
    for c_idx, h in enumerate(headers, 1):
        if ('표준서' in h or '수행지침' in h) and '파일' not in h:
            target_cols.append(c_idx)
            
    print(f"Sheet '{sheet_name}': Targeting columns {[headers[c-1] for c in target_cols]}")
    
    for r in range(header_row + 1, ws.max_row + 1):
        for c in target_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value:
                res = keep_only_top_2_lines(cell.value)
                if res and res != cell.value:
                    cell.value = res
                    modified_std_gui_cells += 1

print(f"\nSuccessfully condensed Standard and Guideline summaries to EXACTLY 2 lines across {modified_std_gui_cells} cells!")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
