import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb['공정표에따른 매뉴얼']

print("Row 1 in 공정표에따른 매뉴얼:", [c.value for c in ws[1] if c.value])
print("Row 2 in 공정표에따른 매뉴얼:", [c.value for c in ws[2] if c.value])
print("Row 3 in 공정표에따른 매뉴얼:", [c.value for c in ws[3] if c.value])
