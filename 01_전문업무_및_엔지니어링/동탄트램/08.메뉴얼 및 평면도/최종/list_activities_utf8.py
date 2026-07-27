import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path, data_only=True)

output_lines = []

output_lines.append("# Sheet and Activity Summary for Folder Structure Creation\n")

sheet_map = {}

for sheet in wb.sheetnames:
    if sheet in ['공정표에따른 매뉴얼', 'GUIDE']:
        continue
    
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        if any(isinstance(val, str) and ('코드' in val or '작업단위' in val) for val in r if val):
            header_row_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
    
    l4_idx = None
    act_idx = None
    
    for i, h in enumerate(headers):
        if 'L4' in h:
            l4_idx = i
        elif '작업단위' in h or 'Activity' in h:
            act_idx = i
            
    activities = []
    for r_idx, r in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):
        if all(v is None for v in r):
            continue
        l4_val = r[l4_idx] if l4_idx is not None and l4_idx < len(r) else None
        act_val = r[act_idx] if act_idx is not None and act_idx < len(r) else None
        
        if l4_val or act_val:
            l4_str = str(l4_val).strip() if l4_val else ""
            act_str = str(act_val).strip() if act_val else ""
            activities.append((l4_str, act_str))
            
    sheet_map[sheet] = activities
    output_lines.append(f"## Sheet: `{sheet}` (Total Activities: {len(activities)})")
    for idx, (l4, act) in enumerate(activities, start=1):
        output_lines.append(f"{idx}. [{l4}] {act}")
    output_lines.append("")

with open("activities_clean.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(output_lines))

print("Wrote activities_clean.txt")
