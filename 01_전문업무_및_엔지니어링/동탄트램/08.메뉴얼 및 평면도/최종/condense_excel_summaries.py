import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

def shorten_summary_to_2_3_lines(text):
    if not text:
        return ""
    
    text = str(text).strip()
    
    # Split by newlines or numbered patterns
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    
    clean_lines = []
    for line in lines:
        # Strip existing leading numbers or bullet characters
        clean = re.sub(r'^\s*[\d\.\-☐☑📍📌]\s*', '', line).strip()
        # Clean meta linkage phrases
        clean = re.sub(r'\[동탄트램 업무 매뉴얼 v1 연계\]\s*', '', clean)
        clean = re.sub(r'\[동탄트램 매뉴얼 v1\]\s*', '', clean)
        clean = re.sub(r'\[설계사 작성\]\s*', '', clean)
        clean = clean.replace('기술기준 연계', '기술기준').replace('설계기준 연계', '설계기준')
        
        if clean and clean not in clean_lines:
            clean_lines.append(clean)
            
    # Take at most 3 lines (2~3 lines)
    selected = clean_lines[:3]
    if not selected:
        return text[:100]
        
    formatted = [f"{i+1}) {line}" for i, line in enumerate(selected)]
    return "\n".join(formatted)

shortened_cells = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    header_row = 3 if sheet_name == '공정표에따른 매뉴얼' else 1
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    # Find summary column indices
    std_sum_col = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '요약' in h), None)
    gui_sum_col = next((c for c, h in enumerate(headers, 1) if '수행지침' in h and '요약' in h), None)
    chk_sum_col = next((c for c, h in enumerate(headers, 1) if '체크리스트' in h and '요약' in h), None)
    
    if not std_sum_col or not gui_sum_col or not chk_sum_col:
        # Fallback search if '요약' header text differs slightly
        std_sum_col = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '파일' not in h), None)
        gui_sum_col = next((c for c, h in enumerate(headers, 1) if '수행지침' in h and '파일' not in h), None)
        chk_sum_col = next((c for c, h in enumerate(headers, 1) if '체크리스트' in h and '파일' not in h), None)

    print(f"Sheet '{sheet_name}': std_sum={std_sum_col}, gui_sum={gui_sum_col}, chk_sum={chk_sum_col}")

    for r in range(header_row + 1, ws.max_row + 1):
        for col_idx in [std_sum_col, gui_sum_col, chk_sum_col]:
            if col_idx:
                cell = ws.cell(row=r, column=col_idx)
                if cell.value:
                    shortened = shorten_summary_to_2_3_lines(cell.value)
                    if shortened != cell.value:
                        cell.value = shortened
                        shortened_cells += 1

print(f"\nShortening complete! Total {shortened_cells} summary cells condensed into clean 2~3 lines.")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
