import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

for v_file in ["매뉴얼 BODY (집행단계)v3.xlsx", "매뉴얼 BODY (집행단계)v4.xlsx"]:
    v_path = os.path.join(base_dir, v_file)
    if os.path.exists(v_path):
        try:
            print(f"Updating 2-tier Excel headers with durations in '{v_file}'...")
            wb = openpyxl.load_workbook(v_path)
            if '지장물이설' in wb.sheetnames:
                ws = wb['지장물이설']
                
                # Row 1 sub-headers or Row 2 labels
                ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크 (가스: 62일 | 난방: 80일 | 통신: 150일 | 전력: 255일 | 광역상수: 160일)")
                
                ws.cell(row=2, column=14, value="가스관 (62일)")
                ws.cell(row=2, column=15, value="열(난방)배관로 (80일)")
                ws.cell(row=2, column=16, value="통신관로 (150일)")
                ws.cell(row=2, column=17, value="전력관 (255일)")
                ws.cell(row=2, column=18, value="광역상수관 (160일)")

                wb.save(v_path)
                print(f"✅ Successfully updated headers in '{v_file}'!")
        except Exception as e:
            print(f"⚠️ Could not save '{v_file}': {e}")
