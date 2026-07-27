import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반"

def normalize(name):
    return name.replace(" ", "").replace("_", "").replace("/", "").replace("(", "").replace(")", "").replace("-", "").lower()

master_activities = [
    (1, "1_지반조사 상세검토", "지반조사 상세검토", "지반 물리탐사 및 시굴 결과를 바탕으로 노반 지반 지지력과 연약지반 분포 구간을 정밀 분석", "PBT 평판재하시험 성과 분석 및 N치<4 연약지반 구간 도면 반영"),
    (2, "2_발주전략 KOM", "발주전략 KOM", "강화노반 시공사 및 자재 공급업체 간 공정/품질 이행 목표 공유 킥오프 미팅", "공정 CPM 마스터 스케줄 및 층다짐 30cm 이하 규정 합의"),
    (3, "3_철도보호지구에서의 행위신고(필요시)", "철도보호지구에서의 행위신고(필요시)", "철도보호지구(인접 30m 이내) 굴착 및 토공사 진행을 위한 법적 행위 신고 및 국가철도공단 승인", "철도안전법 제45조 준수 및 장비 작업 안전 가이드 수립"),
    (4, "4_착수전 측량 Data 확인", "착수전 측량 Data 확인", "GRS80 세계측지계 기준 선로 중심선, 중심고 및 횡단 현형 측량 성과 재검증", "허용 오차 ±10mm 이내 기준점 100m 간격 재설치"),
    (5, "5_지장물이설 협의", "지장물이설 협의", "강화노반 다짐 구간 내 매설 지하 지장물 이격거리(H≥1.5m) 및 공정 간섭 최종 조정", "5대 위탁기관 합동 3D BIM 간섭 검증 및 무단수 이설 방안 협의"),
    (6, "6_용지보상RISK 검토", "용지보상RISK 검토", "노반 시공 폭 및 임시 작업로 확보를 위한 용지 수용 경계 점검 및 미보상 부지 차단", "용지 수용 대장 및 미보상 사유지 진입 금지 가두리 표지 설치"),
    (7, "7_최고의 팀 만들기 지원", "최고의 팀 만들기 지원", "노반/토공 전문 엔지니어 및 다짐 장비 운전원 안전/품질 교육 및 최정예 팀 편성", "토공 기술사 및 PBT 시험 전문 검측원 참관 배치"),
    (8, "8_연약지반 처리공법 검토(필요시)", "연약지반 처리공법 검토(필요시)", "연약지반(N치 < 4) 구간 PVD, DCM, 성토재하 공법 적용성 및 허용 잔류침하량 검토", "허용 잔류침하량 2.5cm 이하 수렴 계측 계획 수립"),
    (9, "9_토공 유동표 확인", "토공 유동표 확인", "절토량 및 성토량 수량 배분 Mass-Curve 검증으로 토공 유동 운반거리 단축 및 사토 계획 수립", "토공 유동표 및 평균 운반거리 5km 이내 최적 경로 확정"),
    (10, "10_기공승낙 적정성 검토", "기공승낙 적정성 검토", "사유지 경계 부지 진입로 및 진입 진출로 토지소유자 기공승낙서 인감 및 범위 검속", "기공승낙서 100% 확보 및 인가 서류 관리대장 이관"),
    (11, "11_폐기물처리계획 수립", "폐기물처리계획 수립", "토공 굴착 중 발생하는 건설폐기물, 오염토양 분리수거 및 올바로시스템 신고 승인", "건설폐기물 재활용촉진법 및 올바로시스템 인허가 승인"),
    (12, "12_철도운행협의(필요시)", "철도운행협의(필요시)", "기존 철도 선로 인접 구역 차단작업 시간 확보 및 신호/전력 가선 방호 조치 협의", "야간 차단작업 시간(01:00~04:30) 확보 및 운행 안전원 배치"),
    (13, "13_작수전 Big Room 회의", "작수전 Big Room 회의", "노반, 궤도, 구조물, 위탁기관 간 3D BIM 간섭 검증 및 층다짐 밀도 일원화 빅룸 협의", "강화노반 K30≥110MN/m³, Ev2≥120MPa 및 LCLM 유동성채움재 합의"),
    (14, "14_시공 계획 수립", "시공 계획 수립", "KCS 11 00 00 토공사 시방 기준에 따른 다짐 장비 조합, 층다짐 두께(30cm이하) 계획 수립", "롤러 조합(진동10t+타이어15t) 및 시험성토 계획 수립"),
    (15, "15_사토장_토취장 선정 검토(필요시)", "사토장_토취장 선정 검토(필요시)", "강화노반용 적합 성토재(CBR≥10) 토취장 품질 시험 및 잔여토 사토장 반출 승인", "토취장 CBR 10% 이상, 흙분 5% 이하 시험성적서 첨부"),
    (16, "16_공사사전준비", "공사사전준비", "가설 사무소, 장비 세차장, 환경오염 방지시설(세륜기, 방진망) 설치 및 수직 측량 규준틀 준비", "세륜기 및 비산먼지 방지시설 100% 가동 준비"),
    (17, "17_임시배수시설", "임시배수시설", "강우 시 노반 유실 방지를 위한 가배수로, 집수정, 침사지 설치 (배수 구배 2.0% 이상)", "가배수로 및 침사지 용량 확보, 배수 경사 2.0% 유지"),
    (18, "18_쌓기재료 검사", "쌓기재료 검사", "강화노반 재료의 쇄석/입도조정 골재 품질 시험 (최대입경 100mm 이하, 흙분 5% 이하)", "입도분석 시험 및 수정CBR 10% 이상 검속 성적서"),
    (19, "19_장비 검수 지원", "장비 검수 지원", "진동 롤러(10톤 이상), 타이어 롤러, 모터 그레이더 등 다짐 장비 정비 상태 및 안전 검사", "장비 자체 안전검사 필증 구비 및 작업 안전 장치 검속"),
    (20, "20_선로 종_횡단 및 용지경계측량", "선로 종_횡단 및 용지경계측량", "트램 선로 중심선 기준 10m 간격 종횡단 측량 및 계획고 허용 오차(±10mm) 기준선 관리", "선로 중심선 좌표 GRS80 세계측지계 매핑 검속"),
    (21, "21_규준틀 설치", "규준틀 설치", "성토 비탈면 경계 및 노반 완성면 표고를 표시하는 경사 규준틀 50m 간격 설치", "수직 표고 허용오차 ±10mm 규준틀 견고 고정"),
    (22, "22_준비배수", "준비배수", "성토 바닥면 지하수위 저하 및 표면수 배출을 위한 암거, 맹암거 사전 배수 공조", "지하수위 노반 하부 1.0m 이하 유지 준비배수 완료"),
    (23, "23_벌개제근_표토제거", "벌개제근_표토제거", "시공 구역 내 유기물 함유 표토(두께 15~30cm) 및 수목 뿌리 완벽 제거 반출", "유기질 표토 100% 제거 및 외부 반출 처리"),
    (24, "24_구조물 및 지장물 제거", "구조물 및 지장물 제거", "노후 구조물, 기존 관로 및 콘크리트 전주 철거 후 소정의 재료로 되메우기 다짐", "지중 폐콘크리트 100% 적법 철거 및 사토 처리"),
    (25, "25_진입로 조성", "진입로 조성", "덤프트럭 및 다짐 롤러 통행용 가설 진입 도로(폭 W≥6.0m, 골재敷 200mm) 설치", "가설 통행로 지지력 확보 및 세륜 부지 조성"),
    (26, "26_노반쌓기", "노반쌓기", "KCS 11 20 00 토공 시방에 따른 1층 다짐 두께 30cm 이하 층다짐 시공 및 다짐도 95% 확보", "1층 다짐 두께 30cm 이하 준수 및 들밀도시험 95% 이상"),
    (27, "27_하부노반 시공", "하부노반 시공", "하부노반(두께 90cm) 입도 양호 골재 부설 및 변형계수 Ev2≥80MPa, Ev2/Ev1≤2.5 검속", "변형계수 Ev2≥80MPa 및 PBT 평판재하시험 승인"),
    (28, "28_상부노반 시공", "상부노반 시공", "상부노반(두께 30cm) 다짐 및 변형계수 Ev2≥100MPa, 노반반력계수 K30≥90MN/m³ 확보", "K30≥90MN/m³ 및 Ev2/Ev1≤2.2 다짐 검속"),
    (29, "29_강화노반 시공", "강화노반 시공", "트램 궤도 직하부 강화노반(두께 30cm) 쇄석혼합 골재 타설, K30≥110MN/m³, Ev2≥120MPa, Ev2/Ev1≤2.2 검속", "K30≥110MN/m³, Ev2≥120MPa, Ev2/Ev1≤2.2 100% 충족"),
    (30, "30_토공 유동운반_사토", "토공 유동운반_사토", "절토 잔여 흙 덤프트럭 사토장 운반 및 비산먼지 방지 덮개 개폐 이력 관리", "사토 반출 운반일지 및 지정 사토장 반입 확인서"),
    (31, "31_연약지반처리(필요시)", "연약지반처리(필요시)", "연약지반 구간 PVD 배수재 계측(간극수압, 침하판) 및 허용 잔류침하량 2.5cm 이하 확인", "침하판 계측 데이터 수렴(잔류침하량 ≤ 2.5cm) 승인"),
    (32, "32_절성토경계부 배수구조물(맹암거) 시공", "절성토경계부 배수구조물(맹암거) 시공", "절성토 편성 경계 부등침하 방지를 위한 맹암거(유압관 D200mm + 투수골재) 시공", "맹암거 투수 골재 부설 및 배수 구배 2.0% 확인"),
    (33, "33_암석쌓기", "암석쌓기", "발파 암석 재료 쌓기 시 최대 입경 300mm 이하 제한 및 공극 채움 쇄석 부설 다짐", "암석 최대 입경 300mm 이하 및 공극 쇄석 채움 다짐"),
    (34, "34_방치기간 확보", "방치기간 확보", "성토 완료 후 부등침하 수렴을 위한 계획 방치 기간(3~6개월) 계측 관리 및 침하판 측정", "침하 수렴 곡선 분석 및 잔류 침하 수렴 확인"),
    (35, "35_토공마무리", "토공마무리", "모터 그레이더를 이용한 횡단 구배 2.0% 정밀 정지 작업 및 종횡단 공차 ±10mm 검속", "횡단 구배 2.0% 및 완성면 높이 오차 ±10mm 검속"),
    (36, "36_토공 마무리면 인계", "토공 마무리면 인계", "후행 콘크리트 궤도 시공팀 및 감리단 입회 노반 인계인수서 서명 및 GIS 대장 이관", "노반 인계인수 서명서 및 K30≥110MN/m³ 성적서 인계")
]

# Generate High-Quality Standard, Guideline, Checklist HTMLs for each Activity
for num, folder_name, title, purpose, method in master_activities:
    folder_path = os.path.join(base_dir, folder_name)
    std_dir = os.path.join(folder_path, "표준서")
    gui_dir = os.path.join(folder_path, "수행지침")
    chk_dir = os.path.join(folder_path, "체크리스트")
    os.makedirs(std_dir, exist_ok=True)
    os.makedirs(gui_dir, exist_ok=True)
    os.makedirs(chk_dir, exist_ok=True)

    # 1. Standard HTML
    std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>상부강화노반 - {title} 기술 표준서</title>
    <style>
        :root {{ --bg-primary: #f8fafc; --bg-card: #ffffff; --text-primary: #0f172a; --text-secondary: #475569; --accent-blue: #1e3a8a; --accent-cyan: #0284c7; --border-color: #e2e8f0; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 30px 20px; background: var(--bg-primary); color: var(--text-primary); line-height: 1.6; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: var(--bg-card); padding: 40px; border-radius: 16px; border: 1px solid var(--border-color); box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.08); }}
        .header {{ border-bottom: 3px solid var(--accent-blue); padding-bottom: 20px; margin-bottom: 30px; }}
        .breadcrumb {{ font-size: 0.85rem; color: var(--accent-cyan); font-weight: 700; margin-bottom: 6px; }}
        .title {{ font-size: 2.1rem; font-weight: 900; color: var(--text-primary); margin: 0; }}
        .meta-info {{ display: flex; gap: 12px; font-size: 0.9rem; color: var(--text-secondary); margin-top: 12px; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-weight: 700; padding: 4px 10px; border-radius: 6px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.4rem; font-weight: 800; color: var(--accent-blue); border-left: 5px solid var(--accent-cyan); padding-left: 12px; margin-top: 35px; margin-bottom: 18px; }}
        table {{ width: 100% !important; max-width: 100% !important; border-collapse: collapse; margin: 12px 0 20px 0; font-size: 0.92rem; }}
        th, td {{ border: 1px solid var(--border-color); padding: 12px 16px; text-align: left; vertical-align: middle; }}
        th {{ background: #f1f5f9; color: #1e293b; font-weight: 700; }}
        .svg-container {{ background: #ffffff; border: 1px solid #cbd5e1; border-radius: 12px; padding: 20px; margin: 20px 0; text-align: center; }}
        .diagram-explanation {{ background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px; padding: 18px; margin-top: 15px; font-size: 0.9rem; color: #334155; text-align: left; }}
        .key-takeaway {{ background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 10px; padding: 16px; margin-top: 15px; color: #166534; font-size: 0.9rem; }}
        .footer-note {{ margin-top: 40px; text-align: center; font-size: 0.85rem; color: #94a3b8; border-top: 1px solid var(--border-color); padding-top: 20px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS 9000-7-{num} Standard</div>
        <h1 class="title">{title} 기술 표준서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 노반공사 / 상부강화노반</span>
            <span>|</span>
            <span><strong>주관부서:</strong> 현장 토공기술팀 / 감리단</span>
            <span>|</span>
            <span><span class="badge">강화노반 표준 규격</span></span>
        </div>
    </div>

    <h2>1. 과업 개요 및 목적 (Overview & Scope)</h2>
    <table>
        <tbody>
            <tr><th style="width: 20%;">과업 목적</th><td>{purpose}</td></tr>
            <tr><th>수행 방법</th><td>{method}</td></tr>
            <tr><th>핵심 품질 목표</th><td>노반 반력계수 K30 ≥ 110 MN/m³, 변형계수 Ev2 ≥ 120 MPa, Ev2/Ev1 ≤ 2.2, 층다짐도 ≥ 95% 100% 확보</td></tr>
            <tr><th>관련 시방 기준</th><td>KCS 11 00 00 토공사, KDS 47 10 00 철도노반설계기준, KCS 47 10 25 강화노반 시방서</td></tr>
        </tbody>
    </table>

    <h2>2. {title} 고유 정량 공학 시방 및 기술 수칙 표</h2>
    <div style="background: #f8fafc; border: 1px solid #cbd5e1; padding: 22px; border-radius: 12px; margin-bottom: 25px;">
        <h4 style="margin: 0 0 12px 0; color: #1e3a8a; font-size: 1.05rem;">📐 상부강화노반 정량적 공학 품질 수칙 및 허용 공차</h4>
        <table style="width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.9rem;">
            <thead>
                <tr style="background: #e2e8f0; color: #0f172a;">
                    <th style="padding: 10px; border: 1px solid #cbd5e1; text-align: center; width: 22%;">기술 검속 항목</th>
                    <th style="padding: 10px; border: 1px solid #cbd5e1; text-align: center; width: 28%;">관련 시방 및 검사 기준</th>
                    <th style="padding: 10px; border: 1px solid #cbd5e1; text-align: center; width: 50%;">핵심 정량 기술 수칙 및 허용 공차</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="font-weight: bold; text-align: center;">노반 지지지수 (K30)</td>
                    <td style="text-align: center;">KCS 47 10 25 / PBT 평판재하시험</td>
                    <td>• 평판재하시험 노반 지지력 계수 <strong>K30 ≥ 110 MN/m³</strong> 준수<br>• 트램 궤도 직하부 30cm 강화노반 필수 검속</td>
                </tr>
                <tr>
                    <td style="font-weight: bold; text-align: center;">변형계수 (Ev2)</td>
                    <td style="text-align: center;">DIN 18134 / PFWD 시험</td>
                    <td>• 2차 변형계수 <strong>Ev2 ≥ 120 MPa</strong> 이상 확보<br>• 다짐비 <strong>Ev2/Ev1 ≤ 2.2</strong> 만족 시 최종 승인</td>
                </tr>
                <tr>
                    <td style="font-weight: bold; text-align: center;">층다짐도 및 두께</td>
                    <td style="text-align: center;">KCS 11 20 00 / 들밀도시험</td>
                    <td>• 1층 다짐두께 <strong>30cm 이하</strong> 엄격 준수<br>• 상대 다짐도 <strong>95% 이상</strong> (KS F 2312 D다짐 기준)</td>
                </tr>
                <tr>
                    <td style="font-weight: bold; text-align: center;">형상 공차 & 구배</td>
                    <td style="text-align: center;">GRS80 세계측지계 측량</td>
                    <td>• 완성면 계획고 종횡단 허용 공차 <strong>±10mm 이내</strong><br>• 배수 횡단 구배 <strong>2.0%</strong> 정밀 정지 유지</td>
                </tr>
            </tbody>
        </table>
        <div style="margin-top: 15px; background: #eff6ff; padding: 12px 16px; border-radius: 8px; border: 1px solid #bfdbfe; font-size: 0.88rem; color: #1e40af;">
            <strong>📐 강화노반 기술 절대 수칙:</strong> 동탄트램 궤도 탈선 방지를 위해 <strong>K30 ≥ 110 MN/m³ 및 층다짐 30cm 이내</strong> 공학 기준을 100% 엄수합니다.
        </div>
    </div>

    <h2>3. {title} 핵심 프로세스 및 다짐 구조 모식도</h2>
    <div class="svg-container">
        <svg viewBox="0 0 900 340" width="100%" height="100%" xmlns="http://www.w3.org/2000/svg">
            <rect width="900" height="340" rx="12" fill="#f8fafc" stroke="#cbd5e1" stroke-width="1.5"/>
            <text x="450" y="32" text-anchor="middle" font-size="16" font-weight="bold" fill="#0f172a">상부강화노반 {title} 4단계 시공 및 지지력 검속 절차</text>

            <g transform="translate(30, 55)">
                <rect width="180" height="195" rx="8" fill="#ffffff" stroke="#2563eb" stroke-width="2"/>
                <rect width="180" height="38" rx="8" fill="#dbeafe"/>
                <text x="90" y="24" text-anchor="middle" font-size="13" font-weight="bold" fill="#1e40af">① 사전 측량 & 재료검사</text>
                <text x="14" y="65" font-size="11" font-weight="bold" fill="#0f172a">• GRS80 측량 기준선 설정</text>
                <text x="14" y="88" font-size="11" fill="#334155">• 쇄석/입도조정 골재 시험</text>
                <text x="14" y="111" font-size="11" fill="#334155">• 최대입경 100mm 이하</text>
                <text x="14" y="134" font-size="11" fill="#2563eb" font-weight="bold">• 적합 골재 100% 반입</text>
            </g>

            <text x="225" y="155" font-size="22" fill="#2563eb">➔</text>

            <g transform="translate(245, 55)">
                <rect width="190" height="195" rx="8" fill="#ffffff" stroke="#ea580c" stroke-width="2"/>
                <rect width="190" height="38" rx="8" fill="#ffedd5"/>
                <text x="95" y="24" text-anchor="middle" font-size="13" font-weight="bold" fill="#9a3412">② 층다짐 (두께 ≤30cm)</text>
                <text x="14" y="65" font-size="11" font-weight="bold" fill="#0f172a">• 모터 그레이더 평삭</text>
                <text x="14" y="88" font-size="11" fill="#334155">• 진동롤러(10t)+타이어롤러</text>
                <text x="14" y="111" font-size="11" fill="#334155">• 1층 다짐 두께 30cm 준수</text>
                <text x="14" y="134" font-size="11" fill="#ea580c" font-weight="bold">• 상대 다짐도 95% 달성</text>
            </g>

            <text x="450" y="155" font-size="22" fill="#ea580c">➔</text>

            <g transform="translate(470, 55)">
                <rect width="190" height="195" rx="8" fill="#ffffff" stroke="#059669" stroke-width="2"/>
                <rect width="190" height="35" rx="8" fill="#dcfce7"/>
                <text x="95" y="24" text-anchor="middle" font-size="13" font-weight="bold" fill="#15803d">③ 지지력 (PBT/PFWD)</text>
                <text x="14" y="65" font-size="11" font-weight="bold" fill="#0f172a">• K30 ≥ 110 MN/m³ 측정</text>
                <text x="14" y="88" font-size="11" fill="#334155">• Ev2 ≥ 120 MPa 검속</text>
                <text x="14" y="111" font-size="11" fill="#334155">• Ev2/Ev1 ≤ 2.2 비율 합격</text>
                <text x="14" y="134" font-size="11" fill="#059669" font-weight="bold">• 노반 반력 계속 100% 승인</text>
            </g>

            <text x="675" y="155" font-size="22" fill="#059669">➔</text>

            <g transform="translate(695, 55)">
                <rect width="175" height="195" rx="8" fill="#ffffff" stroke="#1e3a8a" stroke-width="2"/>
                <rect width="175" height="38" rx="8" fill="#e0e7ff"/>
                <text x="87" y="24" text-anchor="middle" font-size="13" font-weight="bold" fill="#1e1b4b">④ 마무리면 인계</text>
                <text x="14" y="65" font-size="11" font-weight="bold" fill="#0f172a">• 횡단 구배 2.0% 정밀 정지</text>
                <text x="14" y="88" font-size="11" fill="#334155">• 공차 ±10mm 검측</text>
                <text x="14" y="111" font-size="11" fill="#334155">• 후행 궤도팀 인계 서명</text>
                <text x="14" y="134" font-size="11" fill="#1e3a8a" font-weight="bold">• 노반 인계인수 완결</text>
            </g>

            <rect x="30" y="270" width="840" height="48" rx="8" fill="#1e3a8a"/>
            <text x="450" y="299" text-anchor="middle" font-size="13" font-weight="bold" fill="#ffffff">🚨 {title} 다짐 지지력(K30 ≥ 110 MN/m³) 미달 시 트램 궤도 부등침하 재난 전면 차단</text>
        </svg>
    </div>

    <div class="diagram-explanation">
        <h4 style="margin: 0 0 8px 0; color: #0f172a;">🔍 상부강화노반 엔지니어링 시공 및 품질 관리 해설</h4>
        <p style="margin: 0; line-height: 1.7;">본 과업은 KCS 47 10 25 기준에 따라 트램 궤도 직하부 상부강화노반의 다짐 밀도 95% 이상 및 지지력 K30 ≥ 110 MN/m³을 확보하여 궤도 인계인수를 완료하는 정밀 토공 절차입니다.</p>
    </div>

    <div class="key-takeaway">
        <strong>💡 핵심 요약:</strong> K30 ≥ 110 MN/m³, Ev2 ≥ 120 MPa, 1층 다짐 두께 30cm 이내 준수로 궤도 구조물 부등침하를 차단합니다!
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 기술 표준서 | WBS 9000-7-{num} | 상부강화노반
    </div>
</div>
</body>
</html>"""

    # Write Standard
    std_fp = os.path.join(std_dir, f"{folder_name}_표준서.html")
    with open(std_fp, 'w', encoding='utf-8') as f:
        f.write(std_html)

    # 2. Guideline HTML
    gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>상부강화노반 - {title} 수행지침서</title>
    <style>
        :root {{ --bg-primary: #f8fafc; --bg-card: #ffffff; --text-primary: #0f172a; --text-secondary: #475569; --accent-blue: #1e3a8a; --accent-cyan: #0284c7; --border-color: #e2e8f0; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 30px 20px; background: var(--bg-primary); color: var(--text-primary); line-height: 1.6; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: var(--bg-card); padding: 40px; border-radius: 16px; border: 1px solid var(--border-color); box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.08); }}
        .header {{ border-bottom: 3px solid var(--accent-blue); padding-bottom: 20px; margin-bottom: 30px; }}
        .breadcrumb {{ font-size: 0.85rem; color: var(--accent-cyan); font-weight: 700; margin-bottom: 6px; }}
        .title {{ font-size: 2.1rem; font-weight: 900; color: var(--text-primary); margin: 0; }}
        .meta-info {{ display: flex; gap: 12px; font-size: 0.9rem; color: var(--text-secondary); margin-top: 12px; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-weight: 700; padding: 4px 10px; border-radius: 6px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.4rem; font-weight: 800; color: var(--accent-blue); border-left: 5px solid var(--accent-cyan); padding-left: 12px; margin-top: 35px; margin-bottom: 18px; }}
        .step-card {{ background: #ffffff; border: 1px solid #cbd5e1; border-radius: 12px; padding: 24px; margin-bottom: 20px; border-left: 6px solid var(--accent-blue); }}
        .step-title {{ font-size: 1.2rem; font-weight: 800; color: var(--accent-blue); margin-bottom: 12px; display: flex; align-items: center; gap: 8px; }}
        .sub-bullet {{ margin-left: 20px; font-size: 0.93rem; color: #334155; margin-bottom: 6px; line-height: 1.7; }}
        .footer-note {{ margin-top: 40px; text-align: center; font-size: 0.85rem; color: #94a3b8; border-top: 1px solid var(--border-color); padding-top: 20px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS 9000-7-{num} Playbook</div>
        <h1 class="title">{title} 수행지침서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 노반공사 / 상부강화노반</span>
            <span>|</span>
            <span><strong>주관부서:</strong> 현장 토공시공팀 / 검측팀</span>
            <span>|</span>
            <span><span class="badge">3단계 시공 플레이북</span></span>
        </div>
    </div>

    <h2>📋 {title} 3단계 시공 및 검측 수행지침</h2>

    <div class="step-card">
        <div class="step-title">1단계: 사전 준비 및 품질 시험 (Pre-Construction)</div>
        <div class="sub-bullet">• <strong>골재재료 입도 및 CBR 시험:</strong> 강화노반 쇄석 혼합 재료의 최대 입경(100mm 이하) 및 수정 CBR(10% 이상) 성적서 100% 구비</div>
        <div class="sub-bullet">• <strong>측량 기준선 및 규준틀 설치:</strong> GRS80 측량 성과 기준 50m 간격 규준틀 견고 고정 및 수직 허용 오차(±10mm) 기준점 설치</div>
        <div class="sub-bullet">• <strong>준비배수 및 가배수로 구축:</strong> 강우 시 노반 침수 방지를 위한 가배수로 및 집수정 사전 배수구배(2.0% 이상) 설치</div>
    </div>

    <div class="step-card" style="border-left-color: #ea580c;">
        <div class="step-title" style="color: #9a3412;">2단계: 본 시공 및 층다짐 관리 (Construction Execution)</div>
        <div class="sub-bullet">• <strong>모터 그레이더 평삭 및 포설:</strong> 골재 재분리 방지 평삭 포설 및 1층 다짐 두께 30cm 이내 엄격 준수</div>
        <div class="sub-bullet">• <strong>롤러 조합 다짐 시공:</strong> 10톤 이상 진동 롤러(초기 3회) + 15톤 이상 타이어 롤러(마무리 4회) 조합 정속 다짐</div>
        <div class="sub-bullet">• <strong>최적 함수비(OMC) 관리:</strong> 현장 함수비 측정 후 OMC ±2% 범위 내 가수 또는 건조 작업 가동</div>
    </div>

    <div class="step-card" style="border-left-color: #059669;">
        <div class="step-title" style="color: #15803d;">3단계: 검속 및 노반 인계 (Quality Test & Handover)</div>
        <div class="sub-bullet">• <strong>PBT & PFWD 시험 시행:</strong> 평판재하시험 K30 ≥ 110 MN/m³, 2차 변형계수 Ev2 ≥ 120 MPa, Ev2/Ev1 ≤ 2.2 합격 확인</div>
        <div class="sub-bullet">• <strong>들밀도시험 검속:</strong> 상대 다짐도 95% 이상 들밀도시험 성적서 작성 및 100m 당 1회 검측</div>
        <div class="sub-bullet">• <strong>완성면 인계인수:</strong> 횡단 구배 2.0% 및 종횡단 오차 ±10mm 검속 후 감리단/궤도 시공팀 인수 서명 체결</div>
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 수행지침서 | WBS 9000-7-{num} | 상부강화노반
    </div>
</div>
</body>
</html>"""

    gui_fp = os.path.join(gui_dir, f"{folder_name}_수행지침.html")
    with open(gui_fp, 'w', encoding='utf-8') as f:
        f.write(gui_html)

    # 3. Checklist HTML
    chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>상부강화노반 - {title} 체크리스트</title>
    <style>
        :root {{ --bg-primary: #f8fafc; --bg-card: #ffffff; --text-primary: #0f172a; --text-secondary: #475569; --accent-blue: #1e3a8a; --accent-cyan: #0284c7; --border-color: #e2e8f0; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 30px 20px; background: var(--bg-primary); color: var(--text-primary); line-height: 1.6; }}
        .container {{ max-width: 1000px; margin: 0 auto; background: var(--bg-card); padding: 40px; border-radius: 16px; border: 1px solid var(--border-color); box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.08); }}
        .header {{ border-bottom: 3px solid var(--accent-blue); padding-bottom: 20px; margin-bottom: 30px; }}
        .breadcrumb {{ font-size: 0.85rem; color: var(--accent-cyan); font-weight: 700; margin-bottom: 6px; }}
        .title {{ font-size: 2.1rem; font-weight: 900; color: var(--text-primary); margin: 0; }}
        .meta-info {{ display: flex; gap: 12px; font-size: 0.9rem; color: var(--text-secondary); margin-top: 12px; }}
        .badge {{ background: #dbeafe; color: #1e40af; font-weight: 700; padding: 4px 10px; border-radius: 6px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.4rem; font-weight: 800; color: var(--accent-blue); border-left: 5px solid var(--accent-cyan); padding-left: 12px; margin-top: 35px; margin-bottom: 18px; }}
        table {{ width: 100% !important; max-width: 100% !important; border-collapse: collapse; margin: 12px 0 20px 0; font-size: 0.92rem; }}
        th, td {{ border: 1px solid var(--border-color); padding: 12px 16px; text-align: left; vertical-align: middle; }}
        th {{ background: #f1f5f9; color: #1e293b; font-weight: 700; }}
        .checkbox-cell {{ text-align: center; width: 80px; font-weight: bold; }}
        .footer-note {{ margin-top: 40px; text-align: center; font-size: 0.85rem; color: #94a3b8; border-top: 1px solid var(--border-color); padding-top: 20px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS 9000-7-{num} Checklist</div>
        <h1 class="title">{title} 검측 체크리스트</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 노반공사 / 상부강화노반</span>
            <span>|</span>
            <span><strong>검측일시:</strong> 시공 중 / 완료 후</span>
            <span>|</span>
            <span><span class="badge">강화노반 검측표</span></span>
        </div>
    </div>

    <h2>☑️ {title} 9대 핵심 실시간 O/X 검측 항목</h2>
    <table>
        <thead>
            <tr style="background: #e2e8f0; color: #0f172a;">
                <th style="padding: 10px; text-align: center; width: 8%;">번호</th>
                <th style="padding: 10px; text-align: center; width: 22%;">검측 항목</th>
                <th style="padding: 10px; text-align: center; width: 55%;">정량 검측 세부 수칙 및 허용 공차</th>
                <th style="padding: 10px; text-align: center; width: 15%;">검측 결과</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td style="text-align: center; font-weight: bold;">1</td>
                <td style="font-weight: bold;">골재 품질 시험 성적서</td>
                <td>최대입경 100mm 이하, 흙분 함유량 5% 이하, 수정 CBR 10% 이상 확인 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">2</td>
                <td style="font-weight: bold;">측량 오차 및 규준틀</td>
                <td>GRS80 세계측지계 기준 선로 중심선 및 규준틀 표고 허용오차 ±10mm 이내 고정 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">3</td>
                <td style="font-weight: bold;">1층 다짐 두께 준수</td>
                <td>1층 부설 및 다짐 두께 30cm 이하 이행 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">4</td>
                <td style="font-weight: bold;">상대 다짐도 (들밀도)</td>
                <td>들밀도시험 상대 다짐도 95% 이상 확보 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">5</td>
                <td style="font-weight: bold;">노반 반력계수 (K30)</td>
                <td>PBT 평판재하시험 K30 ≥ 110 MN/m³ 만족 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">6</td>
                <td style="font-weight: bold;">2차 변형계수 (Ev2)</td>
                <td>변형계수 Ev2 ≥ 120 MPa 및 다짐비 Ev2/Ev1 ≤ 2.2 비율 만족 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">7</td>
                <td style="font-weight: bold;">배수 횡단 구배</td>
                <td>노반 완성면 배수 횡단 구배 2.0% 정밀 정지 상태 유지 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">8</td>
                <td style="font-weight: bold;">임시 배수 및 준비배수</td>
                <td>강우 시 침수 방지 가배수로 및 집수정 작동 상태 정상 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
            <tr>
                <td style="text-align: center; font-weight: bold;">9</td>
                <td style="font-weight: bold;">노반 인계인수 서명</td>
                <td>감리단 및 후행 궤도 시공팀 입회 노반 완성면 인계서 날인 여부</td>
                <td class="checkbox-cell">[ ☐ 승인 ]</td>
            </tr>
        </tbody>
    </table>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 체크리스트 | WBS 9000-7-{num} | 상부강화노반
    </div>
</div>
</body>
</html>"""

    chk_fp = os.path.join(chk_dir, f"{folder_name}_체크리스트.html")
    with open(chk_fp, 'w', encoding='utf-8') as f:
        f.write(chk_html)

print("🎉 Successfully Generated 108 High-Quality HTML Files for 36 Sangbu Subgrade Activities!")

# Step 3: Update Excel Sheet '상부강화노반' Contents and Links
wb = openpyxl.load_workbook(excel_path)
sheet = wb['상부강화노반']

excel_updated = 0

for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value

    if not l4_code or not act_name:
        continue

    norm_act = normalize(act_name)
    
    matched_item = None
    for item in master_activities:
        if norm_act in normalize(item[2]) or normalize(item[2]) in norm_act:
            matched_item = item
            break
            
    if matched_item:
        num, folder_name, clean_title, purpose, method = matched_item
        excel_updated += 1
        
        # Col 8 (H): 과업 목적
        sheet.cell(row=r, column=8).value = purpose
        # Col 9 (I): 수행 방법
        sheet.cell(row=r, column=9).value = method
        
        # Relative HTML Links
        std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/표준서/{folder_name}_표준서.html"
        gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/수행지침/{folder_name}_수행지침.html"
        chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/체크리스트/{folder_name}_체크리스트.html"
        
        c_std = sheet.cell(row=r, column=12)
        c_gui = sheet.cell(row=r, column=14)
        c_chk = sheet.cell(row=r, column=16)
        
        c_std.value = "👉 [더블클릭] 표준서 열기 📄"
        c_std.hyperlink = std_rel_path
        
        c_gui.value = "👉 [더블클릭] 수행지침 열기 📄"
        c_gui.hyperlink = gui_rel_path
        
        c_chk.value = "👉 [더블클릭] 체크리스트 열기 📄"
        c_chk.hyperlink = chk_rel_path
        
        print(f"Row {r:02d} [{l4_code}] Activity: '{act_name}' ➔ Excel Updated & Linked to '{folder_name}'")

wb.save(excel_path)
print(f"\n🎉 Successfully Updated Excel Sheet '상부강화노반' with Matched Contents & Hyperlinks ({excel_updated} Rows)!")
