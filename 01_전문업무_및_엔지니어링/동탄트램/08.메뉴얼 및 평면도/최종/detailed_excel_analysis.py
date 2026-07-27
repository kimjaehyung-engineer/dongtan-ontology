import openpyxl
import pandas as pd
import sys
import os

# Set standard output to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

excel_path = "매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

report = []
report.append("# 매뉴얼 BODY (집행단계).xlsx 상세 분석 보고서\n")
report.append(f"- **분석 대상 파일**: `{excel_path}`")
report.append(f"- **파일 크기**: {os.path.getsize(excel_path) / 1024:.2f} KB")
report.append(f"- **시트 목록**: {wb.sheetnames}\n")

# Analyze each sheet
for name in wb.sheetnames:
    ws = wb[name]
    report.append(f"## 시트: `{name}`")
    report.append(f"- **전체 행 수**: {ws.max_row}")
    report.append(f"- **전체 열 수**: {ws.max_column}")
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        report.append("- 빈 시트입니다.\n")
        continue
    
    # Let's find where the header row is. Typically it's the first row that starts with 'L2' or has non-empty headers.
    header_idx = 0
    for idx, r in enumerate(rows[:5]):
        if any(isinstance(val, str) and '코드' in val for val in r if val):
            header_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(rows[header_idx])]
    report.append(f"- **헤더 시작 행**: {header_idx + 1}")
    report.append("- **열 정의 (Headers)**:")
    for h in headers:
        report.append(f"  - `{h}`")
    
    # Extract data rows
    data_rows = rows[header_idx + 1:]
    
    # Simple statistics
    total_data = len(data_rows)
    non_empty_rows = 0
    l4_codes = []
    l3_names = set()
    
    # Check key columns (e.g. L4 code, Guideline, Checklist, Standard)
    # Map headers to index
    h_map = {h: idx for idx, h in enumerate(headers)}
    
    for r in data_rows:
        # Check if row is completely empty
        if all(val is None for val in r):
            continue
        non_empty_rows += 1
        
        # Track codes
        l4_col = h_map.get("L4 코드") or h_map.get("L4코드") or h_map.get("Col_3")
        if l4_col is not None and r[l4_col] is not None:
            l4_codes.append(str(r[l4_col]))
            
        l3_name_col = h_map.get("L3 대공종명") or h_map.get("L3 대공종") or h_map.get("L3 명") or h_map.get("L3공종명") or h_map.get("Col_2")
        if l3_name_col is not None and r[l3_name_col] is not None:
            l3_names.add(str(r[l3_name_col]))
            
    report.append(f"- **유효 데이터 행 수 (비비어있는 행)**: {non_empty_rows} / {total_data}")
    if l4_codes:
        report.append(f"- **L4 코드 수**: {len(l4_codes)} (예: {l4_codes[:5]})")
    if l3_names:
        report.append(f"- **L3 대공종 목록 ({len(l3_names)}개)**:")
        for l3 in sorted(list(l3_names)):
            report.append(f"  - {l3}")
            
    # Sample data (first 3 rows of data)
    report.append("\n### 데이터 샘플 (최대 3개)")
    sample_count = 0
    for r in data_rows:
        if all(val is None for val in r):
            continue
        sample_count += 1
        report.append(f"#### 샘플 {sample_count}")
        for col_name, col_idx in h_map.items():
            val = r[col_idx]
            if val is not None:
                # Truncate long strings for readability
                val_str = str(val)
                if len(val_str) > 200:
                    val_str = val_str[:200] + "..."
                report.append(f"- **{col_name}**: {val_str}")
        if sample_count >= 3:
            break
    report.append("\n" + "="*50 + "\n")

# Write report to artifact directory
artifact_dir = r"C:\Users\sskjh\antigravity\brain\887aacfa-3165-4be1-8e89-29f90e47a298"
os.makedirs(artifact_dir, exist_ok=True)
report_file = os.path.join(artifact_dir, "excel_analysis_report.md")

with open(report_file, "w", encoding="utf-8") as f:
    f.write("\n".join(report))

print(f"Analysis complete. Report written to {report_file}")
