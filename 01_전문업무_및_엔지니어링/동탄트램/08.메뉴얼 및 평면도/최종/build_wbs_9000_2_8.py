import os
import shutil
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

artifact_dir = r"C:\Users\sskjh\.gemini\antigravity\brain\887aacfa-3165-4be1-8e89-29f90e47a298"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토"

std_dir = os.path.join(base_dir, "표준서")
gui_dir = os.path.join(base_dir, "수행지침")
chk_dir = os.path.join(base_dir, "체크리스트")

for d in [std_dir, gui_dir, chk_dir]:
    os.makedirs(d, exist_ok=True)

# Copy image to gui_dir
orig_img = "fusion_splicer_calib_1785118781604.jpg"
target_img = "fusion_splicer_calib.jpg"

src_p = os.path.join(artifact_dir, orig_img)
dst_p = os.path.join(gui_dir, target_img)
if os.path.exists(src_p):
    shutil.copy(src_p, dst_p)
    print(f"   ✓ [IMAGE COPIED] {target_img} -> {gui_dir}")

# Shared CSS & JS
modal_style = """
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .zoom-modal, .glossary-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .zoom-modal.active, .glossary-modal.active {
        display: flex;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 580px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-close, .glossary-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .zoom-close:hover, .glossary-close:hover {
        color: #ef4444;
    }
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal" onclick="closeGlossaryModalOutside(event)">
    <div class="glossary-modal-content" onclick="event.stopPropagation()">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 품질 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    "검교정": "<b>검교정 (Calibration)</b><br><br>• 정밀 시험 측정 장비(광융착기, OTDR 등)가 국가 표준 기준과 일치하는 정확한 오차범위 내에서 작동하는지 공인기관에서 <b>1년 마다 검사받고 승인받는 절차</b>입니다.",
    "숙련기술자": "<b>숙련 통신 기술자 (Professional Engineer)</b><br><br>• 정보통신공사업법에 의거한 고급/특급 기술자 자격증 보유자로, 72-Core 광케이블 접속 및 고난도 장비 세팅 실무 경력 5년 이상인 전문 인력입니다.",
    "야적장보관": "<b>자재 야적장 자원 관리</b><br><br>• 광케이블, CCTV, 전광판 등 고가 자재가 비, 습기, 도난으로부터 보호되도록 방수 천막, 방염 덮개 및 현장 도난방지 펜스를 설치하여 보관하는 수칙입니다."
};

function openGlossary(term) {
    const modal = document.getElementById('glossaryModal');
    const titleEl = document.getElementById('modalTitle');
    const descEl = document.getElementById('modalDescription');
    
    if (glossaryData[term]) {
        titleEl.innerHTML = "📖 용어 해설: " + term;
        descEl.innerHTML = glossaryData[term];
        modal.classList.add('active');
    }
}

function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function closeGlossaryModalOutside(event) {
    if (event.target.id === 'glossaryModal') {
        closeGlossaryModal();
    }
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeZoomModal();
        closeGlossaryModal();
    }
});
</script>
"""

# 1. Standard HTML Template
std_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 자재 / 인력 / 장비 등 투입 사전 검토 표준서 (WBS 9000-2-8)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard (WBS 9000-2-8)</span>
        <h1 class="text-3xl font-black mt-2">자재 / 인력 / 장비 등 투입 사전 검토 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-8 | 주관: 현장 시스템팀 / 협력업체 / 감리단 | "품질 기준 및 사전 승인 규정"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 표준 개요 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 본 표준서의 개요 및 적합성 승인 기준</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 표준서는 정보통신공사업법, KCS 47 10 00, KDS 47 10 00 시방서에 의거하여 동탄트램 통신공사 투입 인력의 자격 적합성, 광융착기 및 OTDR 등 핵심 정밀 측정 장비의 <span class="term-highlight" onclick="openGlossary('검교정')">검교정</span> 상태(1년 이내 유효), 그리고 현장 자재 야적장 보관 대책을 사전 심사·승인하는 필수 기술 규정입니다.
            </p>
        </div>

        <!-- 📜 주요 규정 항목 -->
        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-blue-600 pb-2">1. 주요 시방 및 자원 투입 품질 기준</h2>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 block">👤 숙련 기술자 배치 기준</span>
                    <p class="text-slate-700 text-xs">정보통신공사업법 자격 기준에 따른 고급/특급 기술자 배치 및 72-Core 광 접속 경력 확인.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 block">⚡ 정밀 측정 장비 검교정 규정</span>
                    <p class="text-slate-700 text-xs">광융착기 및 OTDR 시험 장비는 국가 공인 기관 검교정 유효기간(1년) 이내의 성적서 보유 의무화.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 block">📦 자재 야적장 보관 및 안전</span>
                    <p class="text-slate-700 text-xs">광케이블, CCTV, PIS전광판 습기 방지(방수 천막) 및 도난 예방 펜스 설치 규정 준수.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 block">🚧 교통통제 및 민원 방지 대책</span>
                    <p class="text-slate-700 text-xs">도로 굴착 시 안전 신호수 배치, 안내 표지판 설치 및 24시간 민원 처리 대장 운용.</p>
                </div>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 2. Guideline HTML Template
gui_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 자재 / 인력 / 장비 등 투입 사전 검토 수행지침서 (WBS 9000-2-8)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Guideline (WBS 9000-2-8)</span>
        <h1 class="text-3xl font-black mt-2">자재 / 인력 / 장비 등 투입 사전 검토 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-8 | 주관: 현장 시스템팀 / 협력업체 / 감리단 | "숙련 인력 & 장비 검교정 실무 가이드"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 개요 박스 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 한눈에 읽는 전문 실무 가이드 (Flexible 4-Step Architecture)</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 지침서는 공정별 통신 인력 수급, 광융착기/OTDR 시험 장비의 <span class="term-highlight" onclick="openGlossary('검교정')">검교정</span> 상태(1년 유효기간), 자재 <span class="term-highlight" onclick="openGlossary('야적장보관')">야적장 보관</span> 및 도로 굴착 교통통제 대책을 사전 심사·승인하는 4단계 실무 가이드입니다.
            </p>
        </div>

        <!-- ☀️ 4대 유연 핵심 프로세스 카드 -->
        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">FLEXIBLE PROCESS</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 자원 투입 사전 검토 4단계 유연 프로세스</h3>
            </div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📋</span> 1. 자원 투입 사전 계획</span>
                    <p class="text-slate-700 text-xs">공정별 통신 인력, 장비 수급 및 자재 수량 사전 정밀 대조.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>⚡</span> 2. 인력 숙련도 & 장비 검교정</span>
                    <p class="text-slate-700 text-xs">숙련 기술자 자격증 및 광융착기/OTDR 검교정 성적서(1년 이내) 검측.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📦</span> 3. 야적장 & 안전/민원 대책</span>
                    <p class="text-slate-700 text-xs">자재 야적장 방수/방범 및 도로 굴착시 교통통제/민원 처리 대책 수립.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🖊️</span> 4. 투입 승인서 최종 체결</span>
                    <p class="text-slate-700 text-xs">발주처·감리원 3자 서명 날인으로 자원 투입 계획서 최종 승인.</p>
                </div>
            </div>
        </div>

        <!-- 🔥 세부 실무 가이드 & 1:1 2D Visual SVG -->
        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 border-b-2 border-indigo-600 pb-2">엔지니어링 세부 실무 가이드 & 1:1 2D 그림</h2>
            
            <!-- STEP 2 Card with Image -->
            <div class="bg-white p-6 rounded-2xl border border-indigo-200 shadow-sm space-y-4">
                <div class="flex items-center gap-3">
                    <span class="bg-indigo-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 2</span>
                    <h3 class="font-bold text-base text-slate-900"><span class="term-highlight" onclick="openGlossary('숙련기술자')">숙련 기술자</span> 자격 및 정밀 장비 검교정 필증 확인</h3>
                </div>

                <!-- 📖 초심자 해설 박스 -->
                <div class="bg-indigo-50 border border-indigo-300 p-5 rounded-xl text-xs text-slate-800 space-y-3 shadow-inner">
                    <h4 class="font-black text-sm text-indigo-950 flex items-center gap-1.5">📖 [초심자 해설] 장비 검교정(Calibration)과 숙련 기술자 배치의 중요성</h4>
                    <div class="grid grid-cols-1 md:grid-cols-2 gap-3 bg-white p-4 rounded-lg border border-indigo-200">
                        <div>
                            <span class="font-bold text-indigo-900 block mb-1">⚡ 1. 장비 검교정(1년 주기)이란?</span>
                            <p class="text-slate-700 leading-relaxed">
                                광융착기나 OTDR 측정기가 국가 표준 오차 기준 내에서 정확히 작동하는지 공인 기관에서 검사받는 필수 필증입니다. 검교정이 안 된 장비로 측정 시 <b>잘못된 접속 손실 수치가 출력되어 정밀 하자가 발생</b>합니다.
                            </p>
                        </div>
                        <div>
                            <span class="font-bold text-indigo-900 block mb-1">👤 2. 숙련 기술자 배치의 목적</span>
                            <p class="text-slate-700 leading-relaxed">
                                72-Core 가느다란 유리선을 녹이는 작업은 숙련된 기술자만이 오차 없이 수행할 수 있으며, 불량 접속으로 인한 재작업 비용 및 공기 지연을 사전에 예방합니다.
                            </p>
                        </div>
                    </div>
                </div>

                <!-- 📸 실제 장비 사진 -->
                <div class="mt-4 bg-slate-50 p-3 rounded-xl border border-slate-200">
                    <span class="text-xs font-bold text-indigo-900 block mb-2">📸 실무 사진: 광융착기 및 OTDR 측정 장비 검교정 필증 실물 사진</span>
                    <img src="./fusion_splicer_calib.jpg" alt="광융착기 및 OTDR 검교정 장비 실물 사진" class="w-full h-52 object-cover rounded-lg border border-slate-300 shadow-sm cursor-pointer" onclick="openDiagramZoom('img_step2_tag', 'STEP 2 광융착기 및 OTDR 검교정 장비 현장 사진')">
                    <img id="img_step2_tag" src="./fusion_splicer_calib.jpg" alt="광융착기 검교정 장비" class="hidden">
                </div>

                <!-- 2D SVG Diagram -->
                <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step2', 'STEP 2 자원 투입 검교정 승인 2D 시공 도식')">
                    <svg id="svg_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                        <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                        <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                        <text x="130" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">👤 인력 & 장비 서류 검토</text>
                        <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 특급/고급 기술자 자격증</text>
                        <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 광융착기/OTDR 검교정 필증</text>
                        <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 유효기간 1년 이내 확인</text>

                        <path d="M 245 90 L 285 90" stroke="#4f46e5" stroke-width="3"/>
                        <polygon points="285,85 295,90 285,95" fill="#4f46e5"/>

                        <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                        <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📑 현장 투입 최종 승인</text>
                        <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 자재 야적장 보관 대책 통과</text>
                        <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 교통통제 신호수 배치 완료</text>
                        <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 자원 투입 서명 승인</text>
                    </svg>
                </div>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 3. Checklist HTML Template (3-Column Master)
chk_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 자재 / 인력 / 장비 등 투입 사전 검토 마스터 체크리스트 (WBS 9000-2-8)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <!-- 🔵 헤더 영역 -->
    <div class="bg-white p-6 sm:p-8 border-b border-slate-200">
        <div class="flex justify-between items-start">
            <div>
                <h1 class="text-2xl sm:text-3xl font-black text-slate-900 tracking-tight">자재 / 인력 / 장비 등 투입 사전 검토 마스터 체크리스트</h1>
            </div>
            <span class="text-xs font-bold text-blue-600 bg-blue-50 px-3 py-1.5 rounded-lg border border-blue-200">WBS Code 9000-2-8 | 통신 검측대장</span>
        </div>
        <div class="w-full h-1 bg-slate-900 mt-4"></div>
    </div>

    <div class="p-6 sm:p-8 space-y-8">
        <!-- 📋 안내 상자 -->
        <div class="bg-blue-50/70 border border-blue-200 p-6 rounded-2xl text-xs sm:text-sm text-blue-950 space-y-2">
            <h4 class="font-bold text-sm sm:text-base text-blue-900 flex items-center gap-2">📋 쉽게 풀어쓴 현장 점검 체크리스트</h4>
            <p class="text-slate-700 leading-relaxed">
                본 체크리스트는 통신 자원 투입 사전 검토 시 <strong>[🟣 시공 도식 열기]</strong>를 클릭하면 대형 고화질 팝업 모달이 열려 도식을 직접 보며 <strong>~하였는가? (100%)</strong> 점검을 진행할 수 있도록 연동되었습니다.
            </p>
        </div>

        <!-- 3-COLUMN MASTER TABLE -->
        <div class="overflow-x-auto rounded-2xl border border-slate-200 shadow-sm">
            <table class="w-full text-left border-collapse">
                <thead>
                    <tr class="bg-slate-100 text-slate-700 text-xs font-black uppercase tracking-wider border-b border-slate-200">
                        <th class="py-4 px-6 text-center w-1/4">시공 단계</th>
                        <th class="py-4 px-6 text-center w-7/12">필수 검측 항목 (쉬운 질문형 수칙)</th>
                        <th class="py-4 px-6 text-center w-1/6">점검 결과</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-200 text-xs sm:text-sm bg-white">
                    
                    <!-- STEP 1 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">📋</span>
                                <span class="font-bold text-slate-900 text-sm">사전 계획</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 1 자원 수급)</span>
                                <button onclick="openDiagramZoomByKey('step2', 'STEP 1 자원 투입 사전 대조 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6">
                            <div class="flex items-start gap-2">
                                <span class="bg-blue-100 text-blue-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">1. 투입 계획 대조</span>
                                <p class="text-slate-800 font-medium"><strong>[자원 계획]</strong> 공정별 통신 인력, 장비 수급 및 자재 수량을 사전 정밀 대조하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-blue-100 text-blue-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">2. 인력 숙련도</span>
                                <p class="text-slate-800 font-medium"><strong>[인력 자격]</strong> 통신 공종별 특급/고급 기술자 자격증 및 현장 배치 계획을 확인하였는가?</p>
                            </div>
                        </td>
                    </tr>

                    <!-- STEP 2 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">⚡</span>
                                <span class="font-bold text-slate-900 text-sm">장비 검교정</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 2 필증 확인)</span>
                                <button onclick="openDiagramZoomByKey('step2', 'STEP 2 장비 검교정 승인 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6 border-t border-slate-200">
                            <div class="flex items-start gap-2">
                                <span class="bg-indigo-100 text-indigo-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">3. 장비 검교정</span>
                                <p class="text-slate-800 font-medium"><strong>[검교정 필증]</strong> 광융착기 및 OTDR 시험 장비의 검교정 성적서(1년 유효기간)를 검측하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-indigo-100 text-indigo-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">4. 야적장 보관</span>
                                <p class="text-slate-800 font-medium"><strong>[자재 야적]</strong> 통신 자재 야적장의 방수, 방염 및 도난 방지 시설을 확인하였는가?</p>
                            </div>
                        </td>
                    </tr>

                    <!-- STEP 3 Row Group -->
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-6 px-6 text-center align-middle bg-slate-50/50 border-r border-slate-200" rowspan="2">
                            <div class="flex flex-col items-center gap-2">
                                <span class="text-xl">🚧</span>
                                <span class="font-bold text-slate-900 text-sm">안전/민원 대책</span>
                                <span class="text-xs text-slate-500 font-medium">(Step 3 투입 승인)</span>
                                <button onclick="openDiagramZoomByKey('step2', 'STEP 3 교통통제 및 투입 승인 도식')" class="mt-2 bg-indigo-600 hover:bg-indigo-700 text-white text-[11px] font-bold px-3 py-1.5 rounded-full shadow-sm transition-all">
                                    🟣 시공 도식 열기
                                </button>
                            </div>
                        </td>
                        <td class="py-4 px-6 border-t border-slate-200">
                            <div class="flex items-start gap-2">
                                <span class="bg-cyan-100 text-cyan-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">5. 안전/민원 대책</span>
                                <p class="text-slate-800 font-medium"><strong>[교통 통제]</strong> 도로 굴착 시 교통통제 신호수 배치 및 민원 예방 대책을 검토하였는가?</p>
                            </div>
                        </td>
                        <td class="py-4 px-6 text-center align-middle font-bold text-blue-600" rowspan="2">
                            ☐ 확인완료
                        </td>
                    </tr>
                    <tr class="hover:bg-slate-50 transition-colors">
                        <td class="py-4 px-6 border-t border-slate-100">
                            <div class="flex items-start gap-2">
                                <span class="bg-emerald-100 text-emerald-800 text-[11px] font-bold px-2 py-0.5 rounded shrink-0">6. 투입 승인 체결</span>
                                <p class="text-slate-800 font-medium"><strong>[최종 승인]</strong> 발주처·감리원 3자 서명 날인으로 자원 투입 계획서를 최종 승인하였는가?</p>
                            </div>
                        </td>
                    </tr>

                </tbody>
            </table>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const svgStore = {{
    'step2': `<svg viewBox="0 0 520 180" width="100%" height="250" xmlns="http://www.w3.org/2000/svg">
                <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                <text x="130" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">👤 인력 & 장비 서류 검토</text>
                <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 특급/고급 기술자 자격증</text>
                <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 광융착기/OTDR 검교정 필증</text>
                <text x="130" y="132" font-size="11" font-weight="black" fill="#15803d" text-anchor="middle">✔ 유효기간 1년 이내 확인</text>
                <path d="M 245 90 L 285 90" stroke="#4f46e5" stroke-width="3"/>
                <polygon points="285,85 295,90 285,95" fill="#4f46e5"/>
                <rect x="300" y="25" width="200" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                <text x="400" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📑 현장 투입 최종 승인</text>
                <text x="315" y="78" font-size="11" font-weight="bold" fill="#334155">• 자재 야적장 보관 대책 통과</text>
                <text x="315" y="100" font-size="11" font-weight="bold" fill="#334155">• 교통통제 신호수 배치 완료</text>
                <text x="400" y="132" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 자원 투입 서명 승인</text>
            </svg>`
}};

function openDiagramZoomByKey(stepKey, titleText) {{
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "시공 도식 대형 정밀 보기");
    
    if (svgStore[stepKey]) {{
        zoomBody.innerHTML = svgStore[stepKey];
    }}
    
    document.getElementById('zoomModal').classList.add('active');
}}

function closeZoomModal() {{
    document.getElementById('zoomModal').classList.remove('active');
}}

function closeZoomModalOutside(event) {{
    if (event.target.id === 'zoomModal') {{
        closeZoomModal();
    }}
}}

window.addEventListener('keydown', function(event) {{
    if (event.key === 'Escape') {{
        closeZoomModal();
    }}
}});
</script>
</body>
</html>
"""

# Write HTML Files
files_to_write = [
    (os.path.join(std_dir, "9000-2-8_자재 인력 장비 등 투입 사전 검토_표준서.html"), std_html_content),
    (os.path.join(std_dir, "자재 인력 장비 등 투입 사전 검토_표준서.html"), std_html_content),
    (os.path.join(gui_dir, "9000-2-8_자재 인력 장비 등 투입 사전 검토_수행지침.html"), gui_html_content),
    (os.path.join(gui_dir, "자재 인력 장비 등 투입 사전 검토_수행지침.html"), gui_html_content),
    (os.path.join(chk_dir, "9000-2-8_자재 인력 장비 등 투입 사전 검토_체크리스트.html"), chk_html_content),
    (os.path.join(chk_dir, "자재 인력 장비 등 투입 사전 검토_체크리스트.html"), chk_html_content)
]

for path, content in files_to_write:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"   ✓ [HTML MASTER BUILT] -> {os.path.basename(path)}")

# Update Excel V4 Row 8
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['통신분야']
    
    # Row 8 is WBS 9000-2-8
    row_idx = 8
    
    # Column J: Standard Summary
    ws.cell(row=row_idx, column=10).value = "1) 투입 자원 사전 검토: 통신 공사 투입 인력, 광융착기/OTDR 측정장비, 자재 수급 계획 등 타당성/적합성 검토함\n2) 적합성 확보: 시스템업체/협력업체 및 감리단 주관으로 공정별 인력 및 장비 투입 제출서의 적합성을 최종 승인함"
    ws.cell(row=row_idx, column=11).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\표준서\자재 인력 장비 등 투입 사전 검토_표준서.html"
    ws.cell(row=row_idx, column=11).value = "📄 [더블클릭] 표준서 열기 🔗"
    ws.cell(row=row_idx, column=11).style = "Hyperlink"
    
    # Column L: Guideline Summary
    ws.cell(row=row_idx, column=12).value = "1) 자원 투입 수급: 공정별 숙련 통신공 투입, 광융착기/OTDR 시험 장비 검교정 상태 확인\n2) 안전/민원 대책: 현장 자재 야적장 확보, 도로 굴착시 교통통제 및 민원 대장 대책을 종합 검토함"
    ws.cell(row=row_idx, column=13).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\수행지침\자재 인력 장비 등 투입 사전 검토_수행지침.html"
    ws.cell(row=row_idx, column=13).value = "📄 [더블클릭] 수행지침 열기 🔗"
    ws.cell(row=row_idx, column=13).style = "Hyperlink"

    # Column N: Checklist Summary
    ws.cell(row=row_idx, column=14).value = "1) 공정별 숙련 인력 투입 계획 및 정밀 측정 장비(OTDR, 융착기) 검교정 상태를 확인하였는가?\n2) 자재 수급 계획, 야적장 확보 및 민원 대책을 포함한 투입 계획서를 작성하였는가?"
    ws.cell(row=row_idx, column=15).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\체크리스트\자재 인력 장비 등 투입 사전 검토_체크리스트.html"
    ws.cell(row=row_idx, column=15).value = "📄 [더블클릭] 체크리스트 열기 🔗"
    ws.cell(row=row_idx, column=15).style = "Hyperlink"
    
    wb.save(excel_path)
    print("   ✓ [EXCEL V4 SYNC COMPLETE] Row 8 (WBS 9000-2-8) Updated Successfully!")

print("\n🎉 SUCCESSFULLY COMPLETED ALL REBUILDING AND SYNC FOR WBS 9000-2-8!")
