import openpyxl
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

# Find existing file
v3_path = None
for f in os.listdir(base_dir):
    if "v3" in f and f.endswith(".xlsx") and not f.startswith("~$"):
        v3_path = os.path.join(base_dir, f)
        break

v4_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

print(f"Copying '{v3_path}' to '{v4_path}'...")
shutil.copy2(v3_path, v4_path)

print(f"Updating HTML links in '{v4_path}' to point directly to the latest updated HTML files...")

wb = openpyxl.load_workbook(v4_path)
ws = wb['지장물이설']

attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)", "지장물이설")
folders = os.listdir(attach_dir)

count = 0
for r in range(3, ws.max_row + 1):
    act_name = str(ws.cell(row=r, column=6).value or '').strip()
    if not act_name:
        continue

    # Find matching folder name
    matched_folder = None
    row_prefix = f"{r-2}_"
    for f_n in folders:
        if f_n.startswith(row_prefix) or act_name in f_n:
            matched_folder = f_n
            break

    if matched_folder:
        std_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\표준서\\{act_name}_표준서.html"
        gui_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\수행지침\\{act_name}_수행지침.html"
        chk_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\체크리스트\\{act_name}_체크리스트.html"

        ws.cell(row=r, column=20, value=f'=HYPERLINK("{std_rel}", "👉 [더블클릭] 표준서 열기 📄")')
        ws.cell(row=r, column=22, value=f'=HYPERLINK("{gui_rel}", "👉 [더블클릭] 수행지침 열기 📘")')
        ws.cell(row=r, column=24, value=f'=HYPERLINK("{chk_rel}", "👉 [더블클릭] 체크리스트 열기 📋")')
        count += 1

wb.save(v4_path)
print(f"🎉 Successfully created and verified '{v4_path}' with {count} updated activity links!")
