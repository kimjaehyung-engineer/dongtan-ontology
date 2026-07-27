import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r_idx, row in enumerate(ws.iter_rows()):
        for c_idx, cell in enumerate(row):
            if cell.hyperlink:
                print(f"Sheet: {sheet_name}, Cell ({r_idx+1}, {c_idx+1})")
                print("  Cell Value:", str(cell.value)[-60:].replace('\n', ' '))
                print("  Hyperlink Target:", cell.hyperlink.target)
                print("  Hyperlink Location:", cell.hyperlink.location)
                print("  Hyperlink Display:", cell.hyperlink.display)
                break
        if cell.hyperlink:
            break
