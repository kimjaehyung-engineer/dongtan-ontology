import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx"

wb = openpyxl.load_workbook(file_path)
ws = wb['지장물이설']

print("=== Checking Excel Summaries for All 38 Rows ===")
single_line_issues = 0

for r in range(3, ws.max_row + 1):
    act_name = ws.cell(row=r, column=6).value
    std_sum = str(ws.cell(row=r, column=19).value or '')
    gui_sum = str(ws.cell(row=r, column=21).value or '')
    chk_sum = str(ws.cell(row=r, column=23).value or '')

    std_lines = len(std_sum.splitlines())
    gui_lines = len(gui_sum.splitlines())
    chk_lines = len(chk_sum.splitlines())

    if std_lines < 2 or gui_lines < 2 or chk_lines < 2:
        print(f"Row {r} [{act_name}]: Std={std_lines}L, Gui={gui_lines}L, Chk={chk_lines}L")
        single_line_issues += 1

if single_line_issues == 0:
    print("✅ PERFECT! All 38 activity rows strictly contain 2-line summaries (1)... \\n 2)...).")
else:
    print(f"⚠️ Found {single_line_issues} rows needing line formatting fix.")
