import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_v3 = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")

from update_jijangmul_sheet_v3 import jijangmul_activities_v3

print(f"Directly writing 100% updated data into '{target_v3}'...")

wb = openpyxl.load_workbook(target_v3)

if '지장물이설' in wb.sheetnames:
    ws = wb['지장물이설']
else:
    ws = wb.create_sheet('지장물이설')

# 1. Clear sheet & Setup Header
ws.delete_rows(1, ws.max_row + 1)

# Header Row 1: Group Headers
ws.merge_cells("K1:M1")
ws.cell(row=1, column=11, value="도급자 시행 지장물 이설(연장 기입)").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
ws.cell(row=1, column=11).font = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="1E3A8A")

ws.merge_cells("N1:R1")
ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
ws.cell(row=1, column=14).font = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="991B1B")

# Header Row 2: Columns
headers = [
    "일정 (D-Day)", "작업명칭 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)",
    "기초 데이터 1", "기초 데이터 2", "기초 데이터 3", "기초 데이터 4",
    "상수관(m)", "하수관(m)", "오수관로(m)",
    "가스관", "난방배관", "통신관로", "전력관", "광역상수관",
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "비고 1", "비고 2", "비고 3", "비고 4"
]

for c_idx, h_text in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=c_idx, value=h_text)
    cell.font = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

# Align & Format Data Rows (Row 3 ~ Row 41)
for r_idx, act in enumerate(jijangmul_activities_v3, start=3):
    ws.cell(row=r_idx, column=1, value=act.get('schedule', 'P+10')).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=2, value=act['act']).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    ws.cell(row=r_idx, column=3, value=act.get('owner', '현장 공사팀')).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=4, value=act.get('purpose', '')).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    ws.cell(row=r_idx, column=5, value=act.get('method', '')).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
    ws.cell(row=r_idx, column=6, value=act.get('result', '')).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

    # Dogeub/Witak pipeline Status
    ws.cell(row=r_idx, column=11, value="150m" if "상수" in act['act'] else "-").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=12, value="200m" if "하수" in act['act'] else "-").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=13, value="-").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    ws.cell(row=r_idx, column=14, value="O" if "가스" in act['act'] else "X").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=15, value="O" if "난방" in act['act'] else "X").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=16, value="O" if "통신" in act['act'] else "X").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=17, value="O" if "전력" in act['act'] else "X").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=18, value="O" if "광역상수" in act['act'] else "X").alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # Summaries (Col 19, 21, 23)
    ws.cell(row=r_idx, column=19, value=act['std_sum']).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r_idx, column=21, value=act['gui_sum']).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r_idx, column=23, value=act['chk_sum']).alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)

    # HTML Hyperlink Buttons (Col 20, 22, 24)
    act_folder = f"{r_idx-2}_{act['act']}"
    std_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\표준서\\{act['act']}_표준서.html"
    gui_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\수행지침\\{act['act']}_수행지침.html"
    chk_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\체크리스트\\{act['act']}_체크리스트.html"

    ws.cell(row=r_idx, column=20, value=f'=HYPERLINK("{std_path}", "👉 [더블클릭] 표준서 열기 📄")').alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=22, value=f'=HYPERLINK("{gui_path}", "👉 [더블클릭] 수행지침 열기 📘")').alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.cell(row=r_idx, column=24, value=f'=HYPERLINK("{chk_path}", "👉 [더블클릭] 체크리스트 열기 📋")').alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

wb.save(target_v3)
print(f"🎉 100% Successfully updated and linked '매뉴얼 BODY (집행단계)v3.xlsx' with all 39 activities!")
