import openpyxl
import os
import json
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반"

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['상부강화노반']

excel_data = []

# Column indices in '상부강화노반' sheet:
# Col 4 (D): Level 4 Task/Activity Code (e.g. 9000-7-1)
# Col 5 (E) or 6 (F): Activity Name (e.g. 지반조사 상세검토)
# Col 7 (G): 주관 (e.g. 현장 공사팀)
# Col 8 (H): 목적
# Col 9 (I): 방법 / 수행지침 요약
# Col 10 (J): 산출물 (결과)
# Col 11 (K): 표준서 요약
# Col 15 (O): 체크리스트 요약

for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value or sheet.cell(row=r, column=5).value
    if not l4_code or not act_name:
        continue
    
    dept = sheet.cell(row=r, column=7).value or "현장 공사팀"
    purpose = sheet.cell(row=r, column=8).value or ""
    method = sheet.cell(row=r, column=9).value or ""
    deliverable = sheet.cell(row=r, column=10).value or "보고서 및 성적서"
    std_sum = sheet.cell(row=r, column=11).value or ""
    gui_sum = sheet.cell(row=r, column=13).value or sheet.cell(row=r, column=9).value or "" # Col 13 or Col 9
    chk_sum = sheet.cell(row=r, column=15).value or ""

    excel_data.append({
        "row": r,
        "l4_code": str(l4_code).strip(),
        "act_name": str(act_name).strip(),
        "dept": str(dept).strip(),
        "purpose": str(purpose).strip(),
        "method": str(method).strip(),
        "deliverable": str(deliverable).strip(),
        "std_sum": str(std_sum).strip(),
        "gui_sum": str(gui_sum).strip(),
        "chk_sum": str(chk_sum).strip()
    })

print(f"Read {len(excel_data)} rows from '상부강화노반' sheet.")
for item in excel_data[:5]:
    print(item)
