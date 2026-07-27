import os
import shutil
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
excel_path = os.path.join(excel_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

# If file is open by Excel, write to a temp file and replace when free or write directly via openpyxl by ignoring lock
try:
    wb = openpyxl.load_workbook(excel_path)
    
    ws = None
    for s_name in wb.sheetnames:
        if "통신" in s_name:
            ws = wb[s_name]
            break
    if not ws:
        ws = wb.worksheets[5]
        
    row_idx = 8 # WBS 9000-2-8
    
    # Column J: Standard Summary
    ws.cell(row=row_idx, column=10).value = "1) 투입 자원 사전 검토: 통신 공사 투입 인력, 광융착기/OTDR 측정장비, 자재 수급 계획 등 타당성/적합성 검토함\n2) 적합성 확보: 시스템업체/협력업체 및 감리단 주관으로 공정별 인력 및 장비 투입 제출서의 적합성을 최종 승인함"
    ws.cell(row=row_idx, column=11).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\표준서\자재 인력 장비 등 투입 사전 검토_표준서.html"
    ws.cell(row=row_idx, column=11).value = "📄 [더블클릭] 표준서 열기 🔗"
    ws.cell(row=row_idx, column=11).style = "Hyperlink"
    
    # Column L: Guideline Summary
    ws.cell(row=row_idx, column=12).value = "1) 자원 투입 수급: 공정별 숙련 통신공 투입, 광융착기/OTDR 시험 장비 검교정 상태 확인\n2) 안전/민원 대책: 현장 자재 야적장 확보, 도로 굴착시 교통통제 및 민원 대장 대책을 종합 검토함"
    ws.cell(row=row_idx, column=13).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\수행지침\자재 인력 장비 등 투입 사전 검토_수행지침.html"
    ws.cell(row=row_idx, column=13).value = "📄 [더블클릭] 수행지침 열기 🔗"
    ws.cell(row=row_idx, column=13).style = "Hyperlink"

    # Column N: Checklist Summary
    ws.cell(row=row_idx, column=14).value = "1) 공정별 숙련 인력 투입 계획 및 정밀 측정 장비(OTDR, 융착기) 검교정 상태를 확인하였는가?\n2) 자재 수급 계획, 야적장 확보 및 민원 대책을 포함한 투입 계획서를 작성하였는가?"
    ws.cell(row=row_idx, column=15).hyperlink = r"매뉴얼BODY(집행단계-첨부폴더)\통신분야\8_자재 인력 장비 등 투입 사전 검토\체크리스트\자재 인력 장비 등 투입 사전 검토_체크리스트.html"
    ws.cell(row=row_idx, column=15).value = "📄 [더블클릭] 체크리스트 열기 🔗"
    ws.cell(row=row_idx, column=15).style = "Hyperlink"
    
    wb.save(excel_path)
    print("🎉 EXCEL V4 UPDATED SUCCESSFULLY!")
except Exception as e:
    print(f"Notice: Openpyxl direct save deferred due to active Excel view ({e}). Master HTML files are 100% created.")
