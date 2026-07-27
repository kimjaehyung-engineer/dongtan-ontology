import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"Sheet '{name}': {ws.max_column} cols, {ws.max_row} rows")
    if name != 'GUIDE':
        headers = [f"Col {i+1}: {repr(c.value)}" for i, c in enumerate(ws[1]) if c.value]
        print("  Headers:", headers)
