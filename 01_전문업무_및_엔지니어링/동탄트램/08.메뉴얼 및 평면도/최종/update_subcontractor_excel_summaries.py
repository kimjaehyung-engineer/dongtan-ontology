import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

print(f"Loading '{target_file}'...")
wb = openpyxl.load_workbook(target_file)
ws = wb['지장물이설']

# Find target row
target_row = None
for r in range(3, ws.max_row + 1):
    act_val = ws.cell(row=r, column=6).value
    if act_val and "도급자분 이설업체 선정(상/하수)" in str(act_val):
        target_row = r
        break

if target_row:
    print(f"Found target activity at Row {target_row}")
    
    std_summary = "1) 상하수도 전문면허 보유 업체 하도급 적격 심사 85점 이상 달성\n2) 설계변경 반영 물량 및 공사 수행 능력 100% 검증"
    gui_summary = "1) 하도급 계약 시 이설 수밀 시험 및 안전 시방 조항 명시\n2) 외주계약팀 승인 후 최종 이설 하도급 계약 체결"
    chk_summary = "1) 상하수도 이설 적격 심사(85점 이상) 및 하도급율(82% 이상)을 점검했는가?\n2) 이설 수밀시험/CCTV 특기시방 명시 및 현장대리인 배치를 확인했는가?"

    ws.cell(row=target_row, column=19, value=std_summary) # Std sum
    ws.cell(row=target_row, column=21, value=gui_summary) # Gui sum
    ws.cell(row=target_row, column=23, value=chk_summary) # Chk sum

    wb.save(target_file)
    print(f"🎉 Successfully saved updated workbook to '{target_file}'")
else:
    print("❌ Target row not found!")
