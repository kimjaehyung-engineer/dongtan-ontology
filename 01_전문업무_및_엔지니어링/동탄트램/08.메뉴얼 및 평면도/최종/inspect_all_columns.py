import openpyxl

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

activities = []
for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value  # Col F (or Col E/G)
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value
    if not act_name:
        act_name = sheet.cell(row=r, column=7).value
    
    # Get all non-empty cell values in row
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15) if sheet.cell(row=r, column=c).value is not None]
    activities.append((r-1, l4_code, row_vals))

for idx, code, vals in activities:
    print(f"{idx:02d}: {vals}")
