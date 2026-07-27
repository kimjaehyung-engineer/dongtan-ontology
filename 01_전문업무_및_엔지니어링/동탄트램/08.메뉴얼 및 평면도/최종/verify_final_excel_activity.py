import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx"

wb = openpyxl.load_workbook(file_path)
ws = wb['지장물이설']

print("=== Final Verification of Updated Activity (Row 4) ===")
act_name = ws.cell(row=4, column=6).value
std_sum = ws.cell(row=4, column=19).value
gui_sum = ws.cell(row=4, column=21).value
chk_sum = ws.cell(row=4, column=23).value

print(f"Activity: {act_name}")
print("\n[표준서 요약 (Col 19)]:\n", std_summary := str(std_sum))
print("\n[수행지침 요약 (Col 21)]:\n", gui_summary := str(gui_sum))
print("\n[체크리스트 요약 (Col 23)]:\n", chk_summary := str(chk_sum))

print("\nLine Count Test (Must be strictly 2 lines):")
print("Std lines:", len(std_summary.splitlines()))
print("Gui lines:", len(gui_summary.splitlines()))
print("Chk lines:", len(chk_summary.splitlines()))
