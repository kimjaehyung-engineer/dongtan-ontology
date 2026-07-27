import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)", "지장물이설")

v3_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
v4_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

print(f"Reading '{v3_file}' and updating HTML hyperlinks to latest files for '{v4_file}'...")

wb = openpyxl.load_workbook(v3_file)
ws = wb['지장물이설']

# Find all activity folders in attached dir
folders = os.listdir(attach_dir)
folder_map = {} # Maps activity index/name to actual folder name

for f_name in folders:
    f_path = os.path.join(attach_dir, f_name)
    if os.path.isdir(f_path):
        folder_map[f_name] = f_path

print(f"Found {len(folder_map)} activity folders in attachment directory.")

updated_links_count = 0

for r in range(3, ws.max_row + 1):
    act_name = str(ws.cell(row=r, column=6).value or '').strip()
    if not act_name:
        continue

    # Match corresponding folder
    matched_folder = None
    row_num_prefix = f"{r-2}_"
    
    for f_n in folder_map.keys():
        if f_n.startswith(row_num_prefix) or (act_name and act_name in f_n):
            matched_folder = f_n
            break

    if matched_folder:
        std_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\표준서\\{act_name}_표준서.html"
        gui_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\수행지침\\{act_name}_수행지침.html"
        chk_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\체크리스트\\{act_name}_체크리스트.html"

        # Update Col 20, 22, 24 HYPERLINK formulas
        ws.cell(row=r, column=20, value=f'=HYPERLINK("{std_path}", "👉 [더블클릭] 표준서 열기 📄")')
        ws.cell(row=r, column=22, value=f'=HYPERLINK("{gui_path}", "👉 [더블클릭] 수행지침 열기 📘")')
        ws.cell(row=r, column=24, value=f'=HYPERLINK("{chk_path}", "👉 [더블클릭] 체크리스트 열기 📋")')

        # Format link cells
        for col_idx in [20, 22, 24]:
            c = ws.cell(row=r, column=col_idx)
            c.font = openpyxl.styles.Font(name="맑은 고딕", size=9.5, bold=True, color="2563EB", underline="single")
            c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        updated_links_count += 1

wb.save(v4_file)
print(f"🎉 Successfully created and saved '{v4_file}' with latest HTML links for all {updated_links_count} activities!")
