import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

print(f"Loading '{target_file}'...")
wb = openpyxl.load_workbook(target_file)
ws = wb['지장물이설']

# Find target row
target_row = None
for r in range(3, ws.max_row + 1):
    act_val = ws.cell(row=r, column=6).value
    if act_val and "지장물 조사 (위탁기관 합동)" in str(act_val):
        target_row = r
        break

if target_row:
    print(f"Found target activity at Row {target_row}")
    
    std_summary = "1) 관리기관 지장물 도면과 실제 현장 맨홀/밸브 위치 100% 현장 대조\n2) 착공 전 측량 성과표 및 인접 주민 민원 우려 요소 사전 도출"
    gui_summary = "1) 유관기관 감독관 및 이설업체 기술자 현장 동행 조사\n2) 위치 일치 여부 현장 조사 보고서 및 사진대지 작성"
    chk_summary = "1) 위탁기관 합동 현장 조사 보고서 및 측량보고서를 확인했는가?\n2) 현장 맨홀 및 노출 관로 심도 일치 여부를 검측했는가?"

    ws.cell(row=target_row, column=19, value=std_summary) # Std sum
    ws.cell(row=target_row, column=21, value=gui_summary) # Gui sum
    ws.cell(row=target_row, column=23, value=chk_summary) # Chk sum

    wb.save(target_file)
    print(f"🎉 Successfully saved updated workbook to '{target_file}'")
else:
    print("❌ Target row not found!")
