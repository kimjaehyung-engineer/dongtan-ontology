import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

discipline_sheets = ['사전토공사', '상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

# Collect data dict from discipline sheets
# Key: (L3_name, act_name) or (act_name) or (L4_code)
discipline_data = {}

for sname in discipline_sheets:
    ws = wb[sname]
    header_row = 1
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip().replace('\n', ' ') for c in range(1, ws.max_column + 1)]
    col_map = {h: i for i, h in enumerate(headers, 1)}
    
    c_l3 = next((i for h, i in col_map.items() if 'L3' in h and '명' in h), 3)
    c_l4 = next((i for h, i in col_map.items() if 'L4' in h), 4)
    c_act = next((i for h, i in col_map.items() if '작업단위' in h or 'Activity' in h), 6)
    
    for r in range(header_row + 1, ws.max_row + 1):
        l3_val = str(ws.cell(row=r, column=c_l3).value or "").strip()
        l4_val = str(ws.cell(row=r, column=c_l4).value or "").strip()
        act_val = str(ws.cell(row=r, column=c_act).value or "").strip()
        
        if not act_val and not l4_val: continue
        
        # Store row cell values and hyperlinks
        row_cells = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            row_cells[c] = {
                'value': cell.value,
                'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                'header': headers[c-1] if c-1 < len(headers) else f"Col{c}"
            }
            
        # Register mapping keys
        if act_val:
            discipline_data[act_val] = row_cells
            if l3_val:
                discipline_data[(l3_val, act_val)] = row_cells
        if l4_val:
            discipline_data[l4_val] = row_cells

print(f"Total unique activity mapping keys indexed: {len(discipline_data)}")

# Now test mapping against '공정매뉴얼' sheet
proc_ws = wb['공정매뉴얼']
proc_header_row = 3

matched = 0
unmatched = 0

for r in range(proc_header_row + 1, proc_ws.max_row + 1):
    c_l3_val = str(proc_ws.cell(row=r, column=3).value or "").strip() # Col C (L3)
    d_l4_val = str(proc_ws.cell(row=r, column=4).value or "").strip() # Col D (L4)
    e_act_val = str(proc_ws.cell(row=r, column=5).value or "").strip() # Col E (Activity)
    
    if not e_act_val and not d_l4_val: continue
    
    match = discipline_data.get((c_l3_val, e_act_val)) or discipline_data.get(e_act_val) or discipline_data.get(d_l4_val)
    if match:
        matched += 1
    else:
        unmatched += 1
        print(f"Row {r:2d} UNMATCHED: C='{c_l3_val}' | D='{d_l4_val}' | E='{e_act_val}'")

print(f"\nMapping Check: Matched={matched}, Unmatched={unmatched}")
