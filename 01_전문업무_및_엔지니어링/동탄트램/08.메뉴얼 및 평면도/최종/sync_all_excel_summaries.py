import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

from update_jijangmul_sheet_v3 import jijangmul_activities_v3

print(f"Loading '{target_file}' to synchronize 2-line summaries for all {len(jijangmul_activities_v3)} activities...")
wb = openpyxl.load_workbook(target_file)
ws = wb['지장물이설']

updated_count = 0

for r_idx, act in enumerate(jijangmul_activities_v3, start=3):
    act_name = act['act']
    std_sum = act['std_sum']
    gui_sum = act['gui_sum']
    chk_sum = act['chk_sum']

    # Strict 2-line check
    ws.cell(row=r_idx, column=19, value=std_sum) # Col 19: Std sum
    ws.cell(row=r_idx, column=21, value=gui_sum) # Col 21: Gui sum
    ws.cell(row=r_idx, column=23, value=chk_sum) # Col 23: Chk sum
    updated_count += 1

wb.save(target_file)
print(f"🎉 Successfully synchronized 2-line summaries for all {updated_count} activities into '{target_file}'!")
