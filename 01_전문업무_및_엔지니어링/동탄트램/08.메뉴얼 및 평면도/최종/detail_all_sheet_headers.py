import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for name in wb.sheetnames:
    if name == 'GUIDE': continue
    ws = wb[name]
    print(f"=== Sheet: {name} ===")
    headers = [str(c.value).replace('\n', ' ') if c.value else "" for c in ws[1]]
    for idx, h in enumerate(headers, 1):
        print(f"  Col {idx}: {h}")
