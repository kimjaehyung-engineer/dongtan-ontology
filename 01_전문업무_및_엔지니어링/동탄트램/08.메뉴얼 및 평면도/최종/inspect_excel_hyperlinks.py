import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

print("Workbook loaded.")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Inspect cells with HYPERLINK or containing '더블클릭' or 'HTML'
    found_links = 0
    for r_idx, row in enumerate(ws.iter_rows()):
        for c_idx, cell in enumerate(row):
            val = str(cell.value) if cell.value is not None else ""
            hyperlink = cell.hyperlink.target if cell.hyperlink else None
            
            if '더블클릭' in val or 'HTML' in val or hyperlink:
                found_links += 1
                if found_links <= 5:
                    print(f"Cell ({r_idx+1}, {c_idx+1}) - Val: {val[:80]} | Hyperlink: {hyperlink}")
                    
    print(f"Total hyperlink/button cells found in {sheet_name}: {found_links}")
