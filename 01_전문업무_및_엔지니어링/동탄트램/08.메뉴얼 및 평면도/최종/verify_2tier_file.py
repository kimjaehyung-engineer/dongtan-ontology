import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

new_v3_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3_지장물분류반영.xlsx"
wb = openpyxl.load_workbook(new_v3_path)

ws = wb['지장물이설']
print("=== Final Verification of '매뉴얼 BODY (집행단계)v3_지장물분류반영.xlsx' ===")
print("Max Rows:", ws.max_row, "| Max Cols:", ws.max_column)

print("\nHeader Row 1 Sample:")
for c in range(1, 29):
    v = ws.cell(row=1, column=c).value
    if v:
        print(f"  Col {c:2d}: {v}")

print("\nHeader Row 2 Sample (Col 11~18):")
for c in range(11, 19):
    print(f"  Col {c:2d}: {ws.cell(row=2, column=c).value}")

print("\nData Row Sampling (Row 4, Row 17, Row 22, Row 24):")
for r in [4, 17, 22, 24]:
    act = ws.cell(row=4 if r==4 else r, column=6).value
    sangsu = ws.cell(row=r, column=11).value
    gas = ws.cell(row=r, column=14).value
    tongsin = ws.cell(row=r, column=16).value
    link_std = ws.cell(row=r, column=20).hyperlink is not None
    print(f"  Row {r:2d} ({act[:15]}...): 상수={sangsu}, 가스={gas}, 통신={tongsin} | StdLink={link_std}")
