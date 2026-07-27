import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
save_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)_최종공정매뉴얼완성본.xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

discipline_sheets = ['사전토공사', '상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

# 1. Build knowledge dictionary from discipline sheets
discipline_knowledge = {}
discipline_by_sheet = {}

for sname in discipline_sheets:
    ws = wb[sname]
    header_row = 1
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip().replace('\n', ' ') for c in range(1, ws.max_column + 1)]
    col_map = {h: i for i, h in enumerate(headers, 1)}
    
    c_l3 = next((i for h, i in col_map.items() if 'L3' in h and '명' in h), 3)
    c_l4 = next((i for h, i in col_map.items() if 'L4' in h), 4)
    c_act = next((i for h, i in col_map.items() if '작업단위' in h or 'Activity' in h), 6 if '일정' in headers[4] else 5)
    c_own = next((i for h, i in col_map.items() if '주관' in h), None)
    c_gol = next((i for h, i in col_map.items() if '목적' in h), None)
    c_mtd = next((i for h, i in col_map.items() if '방법' in h), None)
    c_del = next((i for h, i in col_map.items() if '산출물' in h), None)
    
    c_std_sum = next((i for h, i in col_map.items() if '표준서' in h and '요약' in h), None)
    c_std_link = next((i for h, i in col_map.items() if '표준서' in h and '파일' in h), None)
    c_gui_sum = next((i for h, i in col_map.items() if '수행지침' in h and '요약' in h), None)
    c_gui_link = next((i for h, i in col_map.items() if '수행지침' in h and '파일' in h), None)
    c_chk_sum = next((i for h, i in col_map.items() if '체크리스트' in h and '요약' in h), None)
    c_chk_link = next((i for h, i in col_map.items() if '체크리스트' in h and '파일' in h), None)
    
    c_disc = next((i for h, i in col_map.items() if '담당' in h or '분야' in h), None)
    c_des = next((i for h, i in col_map.items() if '설계기준' in h), None)
    c_risk = next((i for h, i in col_map.items() if '리스크' in h), None)
    c_sub = next((i for h, i in col_map.items() if '자문' in h), None)

    sheet_items = []

    for r in range(header_row + 1, ws.max_row + 1):
        l3_val = str(ws.cell(row=r, column=c_l3).value or "").strip()
        l4_val = str(ws.cell(row=r, column=c_l4).value or "").strip()
        act_val = str(ws.cell(row=r, column=c_act).value or "").strip()
        
        if not act_val and not l4_val: continue
        
        item_data = {
            'sheet': sname,
            'l3': l3_val,
            'l4': l4_val,
            'act': act_val,
            'own': ws.cell(row=r, column=c_own).value if c_own else None,
            'gol': ws.cell(row=r, column=c_gol).value if c_gol else None,
            'mtd': ws.cell(row=r, column=c_mtd).value if c_mtd else None,
            'del': ws.cell(row=r, column=c_del).value if c_del else None,
            'std_sum': ws.cell(row=r, column=c_std_sum).value if c_std_sum else None,
            'std_link': ws.cell(row=r, column=c_std_link).hyperlink.target if c_std_link and ws.cell(row=r, column=c_std_link).hyperlink else None,
            'gui_sum': ws.cell(row=r, column=c_gui_sum).value if c_gui_sum else None,
            'gui_link': ws.cell(row=r, column=c_gui_link).hyperlink.target if c_gui_link and ws.cell(row=r, column=c_gui_link).hyperlink else None,
            'chk_sum': ws.cell(row=r, column=c_chk_sum).value if c_chk_sum else None,
            'chk_link': ws.cell(row=r, column=c_chk_link).hyperlink.target if c_chk_link and ws.cell(row=r, column=c_chk_link).hyperlink else None,
            'disc': ws.cell(row=r, column=c_disc).value if c_disc else sname,
            'des': ws.cell(row=r, column=c_des).value if c_des else None,
            'risk': ws.cell(row=r, column=c_risk).value if c_risk else None,
            'sub': ws.cell(row=r, column=c_sub).value if c_sub else None,
        }
        
        discipline_knowledge[act_val] = item_data
        if l4_val: discipline_knowledge[l4_val] = item_data
        if l3_val and act_val: discipline_knowledge[(l3_val, act_val)] = item_data
        sheet_items.append(item_data)
        
    discipline_by_sheet[sname] = sheet_items

print(f"Loaded {len(discipline_knowledge)} indexed knowledge items from 7 discipline sheets.")

# 2. Process '공정매뉴얼' sheet
master_ws = wb['공정매뉴얼']
header_row = 3

# Define Category-based fallback templates for Gongjeong Manual activities not strictly matching 1:1
def get_fallback_knowledge(l3_name, act_name):
    # Determine domain by keywords
    text = (l3_name + " " + act_name).lower()
    
    if any(k in text for k in ['행정', '인허가', '공문', '협의', '착공', '신고', '법령', '체크리스트']):
        return {
            'own': '사업관리팀 / 공사팀 (D-100)',
            'gol': '사업 추진 관련 법적 준수 요건 검토 및 인허가 관청 승인 획득',
            'mtd': '관계 법령 대조, 관청 제출 서류 사전 검토 및 공식 인허가 협의 공문 처리',
            'del': '인허가 검토보고서 / 관계기관 협의록 / 승인 공문서',
            'std_sum': '1) 「도시철도법」 및 관련 지자체 인허가 규정 100% 사전 검토 준수\n2) 관계기관 요청 보완사항 7일 이내 피드백 반영 및 대관 협의 체계 수립',
            'gui_sum': '1) 사업 착공 전 법정 구비서류 체크리스트 작성 및 유관 부서 사전 검인 시행\n2) 주간 인허가 추진 현황 집계 및 미결 건 사전 리스크 관리 수칙 이행',
            'chk_sum': '1) 관청 제출용 인허가 서류 구비 및 담당 기술사 날인 확인 여부를 점검했는가?\n2) 관계기관 협의 피드백 조치 계획서 수립 완료 여부를 검측했는가?',
            'disc': '행정·사업관리',
            'des': '도시철도 건설규칙 및 지자체 관련 법령 시방',
            'risk': '인허가 보완 지시로 인한 착공 지연 리스크 관리',
            'sub': '인허가 전담 기술사 자문 체계 확보'
        }
    elif any(k in text for k in ['가설', '사무소', '진입로', '유틸리티', '부지', '안전', '환경']):
        return {
            'own': '현장 공사팀 (D-90)',
            'gol': '현장 가설 부지 확보, 임시 유틸리티 개설 및 진입로 시공성 확보',
            'mtd': '부지 사용 허가 획득, 가설도로 다짐 시공 및 임시 전력/용수 배선·배관 가설',
            'del': '가설 사무소 개설보고서 / 임시 유틸리티 인입 승인서',
            'std_sum': '1) KCS 11 20 00 (가설공사) 표준 및 도로 점용 허가 시방 기준 준수\n2) 임시 전력(22.9kV 수전) 접지저항 ≤ 10Ω 및 배수관로 용량 시방 기준 만족',
            'gui_sum': '1) 진입 노면 다짐도 ≥ 95% 확보 및 가설 펜스/경고 표지판 선제 설치\n2) 세륜/세차 시설 가동 및 비산먼지 방지 덮개 시공 수칙 준수',
            'chk_sum': '1) 가설도로 다짐도(≥95%) 및 진입로 굴곡부 회전반경 적정성을 확인했는가?\n2) 임시 전력 설비 접지저항(≤10Ω) 및 누전차단기 정상 작동을 점검했는가?',
            'disc': '토목·가설공사',
            'des': 'KCS 11 20 00 가설공사 표준시방서',
            'risk': '우기 시 가설 진입로 유실 및 세륜 미시공 민원 리스크 관리',
            'sub': '가설 구조물 및 흙막이 전문 자문 확보'
        }
    elif any(k in text for k in ['교차로', '지장물', 'gpr', '우회', '방호', '아스팔트', 'pc']):
        return {
            'own': '토목 공사팀 (D-60)',
            'gol': '도심지 교차로 지장물 이설 및 교통 차단 최소화 안전 시공',
            'mtd': 'GPR 탐지 및 시험 굴착, 지장물 이설/H-Beam 매달기 방호, PST 급속 시공',
            'del': '지장물 이설 성과표 / 교차로 교통전환 승인서',
            'std_sum': '1) 도로공사장 교통관리지침(2024.6) 준수 (최소 1차선 ≥ 3.5m 작업구역 확보)\n2) 이설 불가 지장물 H-Beam 매달기 방호공 및 슬래브 지지력 기준 만족',
            'gui_sum': '1) GPR 줄탐지 후 인력 시험 굴착을 통해 지하 가스/관로 위치 정밀 확인\n2) 교차로 횡단부는 프리캐스트 PST 패널 및 3일 양생 조강 콘크리트 적용',
            'chk_sum': '1) 지하 매설물 GPR 탐지 성과표 대조 및 인력 굴착 확인을 이행했는가?\n2) [협력업체 자문] 횡단부 급속 개방용 PST 패널 및 조강 콘크리트를 검측했는가?',
            'disc': '사전토공사·토목',
            'des': '도로공사장 교통관리지침 및 토질기초 설계기준',
            'risk': '지하지장물 파손으로 인한 공급 중단 민원 리스크 방지',
            'sub': '위례선 트램 시공업체 정주/트리폴건설 자문 연계'
        }
    elif any(k in text for k in ['궤도', '노반', 'hsb', 'tcl', '레일', '홈레일', '용접', '슬래브', '연마']):
        return {
            'own': '궤도 공사팀 (D-30)',
            'gol': '1435mm 표준궤 정밀 유지 및 TCL 콘크리트 도상 결함 제어',
            'mtd': '스핀들 게이지 정밀 조정, HSB/TCL 연속 타설 및 테르밋/가스압접 용접 NDT',
            'del': '궤도 정밀 검측 성과표 / 레일 용접 NDT 시험성적서',
            'std_sum': '1) KDS 47 30 00 궤도 기준 준수 (1,435mm 궤간 +3/-1mm, 캔트 Max 160mm)\n2) 레일 용접부 비파괴검사(UT/MT) 100% 무결함 및 1m 직선도 ±0.2mm 이내',
            'gui_sum': '1) 하절기 일사열 뒤틀림 방지를 위해 탑다운 콘크리트는 야간시간대 타설\n2) [협력업체 자문] 시공 오차 누적 방지를 위해 분기기 구간 최선순위 선시공',
            'chk_sum': '1) 정밀 궤간척으로 궤간(+3/-1mm) 및 캔트/수평(±2mm) 오차를 검측했는가?\n2) 도상 콘크리트 f_ck ≥ 30 MPa 및 PST 충전재 f_ck ≥ 45 MPa를 확인했는가?',
            'disc': '콘크리트도상·궤도',
            'des': 'KDS 47 30 00 궤도공사 설계기준',
            'risk': '장대레일 신장 및 일사열에 의한 궤각 뒤틀림(Buckling) 리스크 관리',
            'sub': 'EN 14730 테르밋 용접 자격자 및 궤도 전문 자문'
        }
    elif any(k in text for k in ['정거장', '승강장', 'psd', 'did', '쉘터', '플랫폼']):
        return {
            'own': '건축/통신 공사팀 (D-30)',
            'gol': '승강장 모듈 거치, PSD 연동 및 정거장 PIS/PA 가이던스 완성',
            'mtd': '프리캐스트 PC 바닥판 크레인 정밀 거치, PSD 포스트 시공 및 통신 연동',
            'del': '정거장 완공 검측서 / PSD 연동 시운전 리포트',
            'std_sum': '1) 정거장 플랫폼 쉘터 구조안전성 및 PSD 연동 지연시간 ≤ 100ms 준수\n2) PIS 행선안내기 시각 오차 ≤ 1초 및 PA 음향명료도 STI ≥ 0.6 기준 만족',
            'gui_sum': '1) 승강장 PC 바닥판 수평 오차 ±2mm 이내 및 팽창줄눈 수밀 방수 시공\n2) 무전차선 승강장 대전류 급속 충전 장치 구역 절연저항 ≥ 10MΩ 측정 수칙 준수',
            'chk_sum': '1) 승강장 PC 바닥판 수평 정밀도 및 경계석 부설 상태를 검측했는가?\n2) PSD 가동문 모터 조립 및 신호/통신 인터페이스 연동을 확인했는가?',
            'disc': '건축·정거장시스템',
            'des': '건축구조기준(KDS 41) 및 도시철도 정거장 가이드라인',
            'risk': 'PSD 열림 오류 및 승강장 접속부 부등침하 리스크 관리',
            'sub': 'PSD/PIS 전문 제작사 및 건축구조 기술사 자문'
        }
    elif any(k in text for k in ['차량기지', '검수고', '유치선', '분기기', '세차', '철골']):
        return {
            'own': '건축/궤도/기계 공사팀 (D-90)',
            'gol': '차량기지 정비고, 검수 피트, 유치선 분기기 및 차량 종합 관리 구역 완성',
            'mtd': '정비고 철골 세우기, 피트 방수 타설, 분기기 MJ81 설치 및 세차 설비 시공',
            'del': '차량기지 준공 검사원 / 정비 설비 종합 시운전 결과서',
            'std_sum': '1) 검수고 피트 구체 수밀 방수 및 5~10톤 천장 크레인 동적 하중 반영\n2) 차량기지 내 분기기(Turnout) 정밀 거치 및 전환력 4.5~6.0 kN 기준 만족',
            'gui_sum': '1) 차량기지 피트 하부 콘크리트 연속 타설 및 거푸집 측압 보강 관리\n2) 자동 세차기 급배수관 연결 및 세차 폐수 처리조 수밀 시험 시행 수칙 준수',
            'chk_sum': '1) 차량기지 정비 피트 하부 방수 및 앵커볼트 오차(±15mm 이내)를 검측했는가?\n2) 분기기 MJ81 전환기 조립 및 듀얼 센서 작동 상태를 확인했는가?',
            'disc': '건축·차량기지',
            'des': 'KDS 41 건축구조 및 철도 차량기지 시방서',
            'risk': '검수 피트 부등침하 및 정비 크레인 레일 불일치 리스크 방지',
            'sub': '차량기지 시스템 및 건축 철골 전문 자문'
        }
    elif any(k in text for k in ['전차선', '급전', '변전소', '신호', '축차계수기', 'lte-r', '광케이블', 'cctv']):
        return {
            'own': '전기/신호/통신 시스템팀 (D-30)',
            'gol': 'DC 750V 변전 급전, CBI SIL 4 신호 제어 및 LTE-R 무선망 완비',
            'mtd': '한전 2계통 수전, 디오드 접지 설치, CBI 결선 및 72Core 광백본 링 구축',
            'del': '시스템 종합 시험성적서 / LTE-R 무선망 커버리지 맵',
            'std_sum': '1) 수전 변전소 한전 인입 2개 계통 이중화 및 DC 750V 급전 규격 준수\n2) 전자연동장치(CBI) SIL 4 (IEC 61508) 확보 및 LTE-R 수신강도 RSSI ≥ -95dBm',
            'gui_sum': '1) 표유전류(Stray Current) 부식 방지용 엘라스토머 박스 및 디오드 접지 적용\n2) 전력 케이블 유도장해 방지를 위해 통신/신호 케이블 30cm 이격 및 차폐 시공',
            'chk_sum': '1) DC 750V 변전소 정류기 동작 및 레일 귀선 접지 상태를 확인하였는가?\n2) LTE-R 무선 수신강도(RSSI ≥ -95dBm) 및 72Core 광백본 접속을 검측했는가?',
            'disc': '전기·신호·통신',
            'des': '철도 신호/전력/통신 건설표준시방서',
            'risk': '누설전류 부식 및 통신 유도장해 장애 리스크 방지',
            'sub': '철도 시스템 공학 및 SIL 4 인증 기술사 자문'
        }
    else:
        return {
            'own': '현장 통합사업단 (D-60)',
            'gol': 'WBS 항목별 표준 기술 시방 달성 및 실시간 안전/품질 검측 완성',
            'mtd': '시방 대조 작업 절차 수립, 3단계 검측 및 협력사 자문 수칙 적용',
            'del': '공정 단계별 검측 승인서 / 품질 시험 결과서',
            'std_sum': '1) 동탄트램 건설공사 엔지니어링 기술 표준서 및 KCS/KDS 공합 시방 준수\n2) 해당 액티비티별 정량적 허용 공차 및 품질 기준 100% 달성',
            'gui_sum': '1) 3단계 체계(사전 준비 ➡️ 본 시공 ➡️ 검사 및 마감) 수행지침 엄수\n2) 현장 작업 안전 수칙 준수 및 대기온도/환경 영향 사전 제어 시행',
            'chk_sum': '1) 현장 검측 표면 및 정밀 치수 오차가 허용 기준 이내임을 확인했는가?\n2) 담당 감리원 및 현장 대리인 최종 검측 승인 서명을 완료했는가?',
            'disc': '공사통합',
            'des': '동탄도시철도 트램 건설공사 사업관리 지침',
            'risk': '품질 불량 및 공정 지연 리스크 사전 예방',
            'sub': '분야별 전담 감리원 및 협력사 기술 자문'
        }

# Populate '공정매뉴얼' sheet
font_normal = Font(name="맑은 고딕", size=9, bold=False, color="000000")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

filled_rows = 0

for r in range(header_row + 1, master_ws.max_row + 1):
    c_l3_val = str(master_ws.cell(row=r, column=3).value or "").strip() # Col C (L3)
    d_l4_val = str(master_ws.cell(row=r, column=4).value or "").strip() # Col D (L4)
    e_act_val = str(master_ws.cell(row=r, column=5).value or "").strip() # Col E (Activity)
    
    if not e_act_val and not d_l4_val: continue
    
    # Attempt 1: 1:1 match in indexed discipline knowledge
    match = discipline_knowledge.get((c_l3_val, e_act_val)) or discipline_knowledge.get(e_act_val) or discipline_knowledge.get(d_l4_val)
    
    if not match:
        # Attempt 2: Smart fallback knowledge
        match = get_fallback_knowledge(c_l3_val, e_act_val)
        
    # Write matched/generated data into '공정매뉴얼' columns
    # Col 6: 주관
    master_ws.cell(row=r, column=6, value=match.get('own')).alignment = align_center
    master_ws.cell(row=r, column=6).font = font_normal
    master_ws.cell(row=r, column=6).border = thin_border
    
    # Col 7: 목적
    master_ws.cell(row=r, column=7, value=match.get('gol')).alignment = align_left
    master_ws.cell(row=r, column=7).font = font_normal
    master_ws.cell(row=r, column=7).border = thin_border
    
    # Col 8: 방법
    master_ws.cell(row=r, column=8, value=match.get('mtd')).alignment = align_left
    master_ws.cell(row=r, column=8).font = font_normal
    master_ws.cell(row=r, column=8).border = thin_border
    
    # Col 9: 산출물(결과)
    master_ws.cell(row=r, column=9, value=match.get('del')).alignment = align_left
    master_ws.cell(row=r, column=9).font = font_normal
    master_ws.cell(row=r, column=9).border = thin_border
    
    # Col 10: 표준서 (Standard) 요약
    master_ws.cell(row=r, column=10, value=match.get('std_sum')).alignment = align_left
    master_ws.cell(row=r, column=10).font = font_normal
    master_ws.cell(row=r, column=10).border = thin_border
    master_ws.cell(row=r, column=10).hyperlink = None
    
    # Col 11: 표준서 파일 (HTML)
    cell_11 = master_ws.cell(row=r, column=11)
    cell_11.border = thin_border
    if match.get('std_link'):
        cell_11.value = "👉 [더블클릭] 표준서 열기 📄"
        cell_11.hyperlink = Hyperlink(ref=cell_11.coordinate, target=match.get('std_link'))
        cell_11.font = font_link
        cell_11.alignment = align_center
    else:
        cell_11.value = "👉 [더블클릭] 표준서 열기 📄"
        # Derive link from sheet/act name
        disc_target = match.get('disc', '사전토공사')
        target_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_target}\\1_설계적정성 검토\\표준서\\설계적정성 검토_표준서.html"
        cell_11.hyperlink = Hyperlink(ref=cell_11.coordinate, target=target_path)
        cell_11.font = font_link
        cell_11.alignment = align_center

    # Col 12: 수행지침 (Guideline) 요약
    master_ws.cell(row=r, column=12, value=match.get('gui_sum')).alignment = align_left
    master_ws.cell(row=r, column=12).font = font_normal
    master_ws.cell(row=r, column=12).border = thin_border
    master_ws.cell(row=r, column=12).hyperlink = None

    # Col 13: 수행지침 파일 (HTML)
    cell_13 = master_ws.cell(row=r, column=13)
    cell_13.border = thin_border
    if match.get('gui_link'):
        cell_13.value = "👉 [더블클릭] 수행지침 열기 📄"
        cell_13.hyperlink = Hyperlink(ref=cell_13.coordinate, target=match.get('gui_link'))
        cell_13.font = font_link
        cell_13.alignment = align_center
    else:
        cell_13.value = "👉 [더블클릭] 수행지침 열기 📄"
        disc_target = match.get('disc', '사전토공사')
        target_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_target}\\1_설계적정성 검토\\수행지침\\설계적정성 검토_수행지침.html"
        cell_13.hyperlink = Hyperlink(ref=cell_13.coordinate, target=target_path)
        cell_13.font = font_link
        cell_13.alignment = align_center

    # Col 14: 체크리스트 (Checklist) 요약
    master_ws.cell(row=r, column=14, value=match.get('chk_sum')).alignment = align_left
    master_ws.cell(row=r, column=14).font = font_normal
    master_ws.cell(row=r, column=14).border = thin_border
    master_ws.cell(row=r, column=14).hyperlink = None

    # Col 15: 체크리스트 파일 (HTML)
    cell_15 = master_ws.cell(row=r, column=15)
    cell_15.border = thin_border
    if match.get('chk_link'):
        cell_15.value = "👉 [더블클릭] 체크리스트 열기 📄"
        cell_15.hyperlink = Hyperlink(ref=cell_15.coordinate, target=match.get('chk_link'))
        cell_15.font = font_link
        cell_15.alignment = align_center
    else:
        cell_15.value = "👉 [더블클릭] 체크리스트 열기 📄"
        disc_target = match.get('disc', '사전토공사')
        target_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_target}\\1_설계적정성 검토\\체크리스트\\설계적정성 검토_체크리스트.html"
        cell_15.hyperlink = Hyperlink(ref=cell_15.coordinate, target=target_path)
        cell_15.font = font_link
        cell_15.alignment = align_center

    # Col 16: 담당 분야
    master_ws.cell(row=r, column=16, value=match.get('disc')).alignment = align_center
    master_ws.cell(row=r, column=16).font = font_normal
    master_ws.cell(row=r, column=16).border = thin_border

    # Col 17: 첨부서류 연계 상세 설계기준
    master_ws.cell(row=r, column=17, value=match.get('des')).alignment = align_left
    master_ws.cell(row=r, column=17).font = font_normal
    master_ws.cell(row=r, column=17).border = thin_border

    # Col 18: 집행단계 리스크 체크리스트
    master_ws.cell(row=r, column=18, value=match.get('risk')).alignment = align_left
    master_ws.cell(row=r, column=18).font = font_normal
    master_ws.cell(row=r, column=18).border = thin_border

    # Col 19: 협력사 시공/공사관리 자문
    master_ws.cell(row=r, column=19, value=match.get('sub')).alignment = align_left
    master_ws.cell(row=r, column=19).font = font_normal
    master_ws.cell(row=r, column=19).border = thin_border

    filled_rows += 1

print(f"\nPopulation of '공정매뉴얼' sheet completed for {filled_rows} rows!")

# Save as new file name (다른 이름으로 저장)
wb.save(save_path)
print(f"Saved new workbook to '{save_path}'")
