import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

target_folder = None
for f in os.listdir(base_dir):
    if f.startswith("5_") or ("제작사" in f and "인터페이스" in f):
        target_folder = os.path.join(base_dir, f)
        break

if not target_folder:
    print("❌ ERROR: Target folder for WBS 9000-2-5 not found!")
    sys.exit(1)

print(f"Target WBS 9000-2-5 Folder: {target_folder}")

zoom_modal_style = """
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .clickable-diagram::after {
        content: "🔍 클릭하여 대형 확대보기";
        position: absolute;
        bottom: 8px;
        right: 12px;
        background: rgba(15, 23, 42, 0.75);
        color: #ffffff;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        backdrop-filter: blur(4px);
        pointer-events: none;
        opacity: 0.85;
        transition: opacity 0.2s;
    }
    .clickable-diagram:hover::after {
        opacity: 1;
        background: rgba(2, 132, 199, 0.9);
    }
    .glossary-modal, .zoom-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .glossary-modal.active, .zoom-modal.active {
        display: flex;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 550px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-close, .zoom-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .glossary-close:hover, .zoom-close:hover {
        color: #ef4444;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal">
    <div class="glossary-modal-content">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 제작사 연동 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    'lter_du_rru': {
        title: '📡 LTE-R 무선 기지국 (DU/RRU) 연동',
        desc: 'LTE-R 주제어장치(EPC) 및 기지국(DU/RRU) 간 TTA 3GPP 표준 프로토콜 패킷 연동 시험과 72-Core/96-Core 광케이블망 핀 맵(Pin Map)을 1:1 대조·확정하는 수칙입니다.'
    },
    'pis_pa_cctv': {
        title: '📺 PIS/PA 방송 및 4K IP CCTV 0.5초 지연 연동',
        desc: '승강장 PIS 전광판, PA 음성 방송 모듈 및 4K IP CCTV 카메라 간 열차 도착 신호 수신 후 지연 시간 0.5초 이내 실시간 방송 및 영상 패킷 연동을 검증하는 수칙입니다.'
    },
    'psd_emergency_call': {
        title: '📻 PSD 비상통화버튼 3초 연동 시험',
        desc: '승강장 스크린도어(PSD) 함체에 설치된 비상통화버튼 조작 시 관제실 통신 콘솔로 3초 이내 즉시 착신 연동되는 회로 차폐 및 1:1 연동을 현장에서 시험하는 수칙입니다.'
    },
    'tta_kca_certified': {
        title: '📄 무선통신 호환성 인증 (TTA) & 전파 합격증 (KCA)',
        desc: '통신설비 제작사 간 무전기 및 안테나 호환성/해킹 방지 국가 공인 인증(TTA) 서류를 확보하고 전파법 최종 전파 합격증(KCA 필증) 서류를 5자 종합 관리대장에 체결하는 수칙입니다.'
    }
};

function openGlossary(termKey) {
    const data = glossaryData[termKey];
    if (!data) return;
    document.getElementById('modalTitle').innerText = data.title;
    document.getElementById('modalDescription').innerText = data.desc;
    document.getElementById('glossaryModal').classList.add('active');
}
function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeGlossaryModal();
        closeZoomModal();
    }
});
</script>
"""

# 1. STANDARD HTML
std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 통신설비 제작사 인터페이스 협의 표준서 (WBS 9000-2-5)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard (WBS 9000-2-5)</span>
        <h1 class="text-3xl font-black mt-2">통신설비 제작사 인터페이스 협의 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-5 | 주관: 현장 시스템팀 / 통신설비 각 제작사 협업업체</p>
    </div>
    
    <div class="p-8 space-y-8">
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-blue-950 mb-2">🎯 표준 목적 (Objective)</h3>
            <p class="text-slate-700 text-sm font-medium leading-relaxed">
                동탄트램 1단계 통신분야 설비(LTE-R 기지국, 96Core/72Core 광전송장치, PIS 승객안내전광판, PA 음성방송, 4K IP CCTV, PSD 스크린도어 비상통화) 제작사 간 기술 사양, RS-485/Ethernet 통신 프로토콜, 핀 맵(Pin Map) 및 전원 규격을 사전 대조·확정하고, 현장 시스템팀 주관으로 제작사 간 물리적·기능적 연동 무결성을 확보함을 목적으로 함.
            </p>
        </div>

        <div>
            <h3 class="text-lg font-bold text-slate-900 mb-4 border-b-2 border-blue-600 pb-2">📜 통신설비 제작사 인터페이스 정밀 시방 수칙 (Methodology)</h3>
            <ul class="space-y-3">
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 1</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>제작사 기술 프로토콜 확정:</strong> LTE-R 기지국(DU/RRU), 96Core/72Core 광전송장치, PIS/PA, 4K IP CCTV 제작사 간 RS-485/Ethernet 통신 프로토콜 및 커넥터 핀 맵(Pin Map) 사양을 1:1 확정함.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 2</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>제작사 전원 사양 및 무정전 UPS 결속:</strong> 각 제작사 장비 입력 전원(AC 220V / DC 48V) 및 통합 UPS 분전반 결선 사양을 대조하고, 무정전 백업 백플레인 규격을 확정함.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 3</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>PIS/PA 및 CCTV 지연 연동 시험 (≤0.5초):</strong> PIS 전광판 및 PA 방송 모듈 수신 지연 시간(0.5초 이내)과 4K CCTV 비디오 스트리밍 패킷 손실률(0.01% 이하)을 제작사 합동 검증함.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 4</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>PSD 비상통화 3초 연동 시험:</strong> 승강장 스크린도어(PSD) 함체 비상통화버튼 조작 시 관제실 통신 콘솔로 3초 이내 착신 연동되는 인터페이스 회로를 제작사 간 실측함.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 5</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>무선통신 호환성 인증 (TTA/KCA):</strong> 무선통신 장비 간 끊김 없이 잘 통하고 해킹당하지 않는 국가 공인 인증(TTA) 서류 및 전파법 최종 전파 합격증(KCA 필증) 서류를 확인·체결함.</span>
                </li>
            </ul>
        </div>

        <div class="bg-emerald-50 border border-emerald-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-emerald-950 mb-2">📦 증빙 산출물 (Deliverables)</h3>
            <p class="text-emerald-900 text-sm font-bold">제작사 인터페이스 협의 회의록, 제작사별 통신 프로토콜/핀 맵 승인서, TTA 보안성/상호운용성 인증서, KCA 전파 적합 필증</p>
        </div>
    </div>
</div>
</body>
</html>
"""

# 2. GUIDELINE HTML
gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 통신설비 제작사 인터페이스 협의 수행지침서 (WBS 9000-2-5)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Detailed Guideline (WBS 9000-2-5)</span>
        <h1 class="text-3xl font-black mt-2">통신설비 제작사 인터페이스 초정밀 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-5 | 주관: 현장 시스템팀 / 통신설비 각 제작사 | "Step별 1:1 2D Visual 그림 완벽 수록"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 친절한 개요 해설 박스 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 통신설비 제작사 인터페이스 현장 실무 개요</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                본 지침서는 현장 엔지니어가 작업 단계별로 <strong>어떻게(HOW) LTE-R 기지국, 96Core/72Core 광전송장치, PIS/PA 방송, 4K IP CCTV, PSD 비상통화 제작사 간 통신 프로토콜 및 전원 사양을 1:1 대조·검증</strong>해야 하는지 <strong>Step별 1:1 직관적 2D Visual 기술 도식(그림)</strong>을 함께 수록하였습니다.
                각 그림은 <strong><span class="term-highlight" onclick="openDiagramZoom('svg_step1', 'STEP 1 프로토콜 핀 맵 대조 도식')">클릭 시 대형 확대보기</span></strong>를 지원합니다.
            </p>
        </div>

        <!-- ☀️ 라이트 테마 특화 카드 섹션 -->
        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">SPECIAL FOCUS</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 통신설비 제작사 인터페이스 실무 검토 가이드</h3>
            </div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📡</span> 1. LTE-R & 광전송 제작사 연동</span>
                    <p class="text-slate-700 text-xs">LTE-R DU/RRU 기지국과 96Core/72Core 광전송장치 프로토콜 핀 맵 1:1 확정.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📺</span> 2. PIS/PA & 4K CCTV 지연 연동</span>
                    <p class="text-slate-700 text-xs">승강장 전광판, 음성방송 및 4K CCTV 패킷 전송 지연시간 0.5초 이내 검증.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📻</span> 3. PSD 비상통화버튼 3초 연동</span>
                    <p class="text-slate-700 text-xs">승강장 스크린도어 비상통화 조작 시 관제실 통신 콘솔 3초 착신 시험.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📄</span> 4. 무선 호환성 & 전파 합격증</span>
                    <p class="text-slate-700 text-xs">TTA 상호운용성/보안성 국가 인증서 및 KCA 최종 전파 합격 필증 체결.</p>
                </div>
            </div>
        </div>

        <!-- 1. FLEXIBLE 5-STEP ARCHITECTURE -->
        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> 5단계 수행 마스터 프로세스 (Flexible 5-Step Architecture)
            </h2>
            <div class="grid grid-cols-1 sm:grid-cols-2 md:grid-cols-5 gap-2">
                <div class="bg-blue-50 p-3 rounded-xl border border-blue-200 flex flex-col justify-between">
                    <span class="bg-blue-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 1</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">핀맵 대조</h4>
                    <p class="text-[10px] text-blue-900 mt-1 font-medium">• 통신 프로토콜<br">• RS-485/Ethernet</p>
                </div>
                <div class="bg-indigo-50 p-3 rounded-xl border border-indigo-200 flex flex-col justify-between">
                    <span class="bg-indigo-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 2</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">기지국 연동</h4>
                    <p class="text-[10px] text-indigo-900 mt-1 font-medium">• LTE-R DU/RRU<br">• 96Core 광배선</p>
                </div>
                <div class="bg-cyan-50 p-3 rounded-xl border border-cyan-200 flex flex-col justify-between">
                    <span class="bg-cyan-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 3</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">방송/CCTV</h4>
                    <p class="text-[10px] text-cyan-900 mt-1 font-medium">• PIS/PA 방송연동<br">• 지연시간 ≤ 0.5초</p>
                </div>
                <div class="bg-teal-50 p-3 rounded-xl border border-teal-200 flex flex-col justify-between">
                    <span class="bg-teal-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 4</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">PSD 비상통화</h4>
                    <p class="text-[10px] text-teal-900 mt-1 font-medium">• 3초 관제 착신<br">• 차폐 회로 검측</p>
                </div>
                <div class="bg-emerald-50 p-3 rounded-xl border border-emerald-200 flex flex-col justify-between">
                    <span class="bg-emerald-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 5</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">TTA/KCA 체결</h4>
                    <p class="text-[10px] text-emerald-900 mt-1 font-medium">• TTA 국가 인증<br">• KCA 전파 합격증</p>
                </div>
            </div>
        </div>

        <!-- 🔥 2. 초정밀 HOW 세부 실무 가이드 (STEP 1~5 1:1 2D Visual 그림 수록) -->
        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-indigo-600 pb-2">
                <span class="text-indigo-600">2.</span> 초정밀 HOW(어떻게 수행하는가) Step별 1:1 visual 실무 가이드
            </h2>

            <div class="space-y-8 text-sm">
                <!-- STEP 1 HOW + 1:1 Visual SVG -->
                <div class="bg-white p-6 rounded-2xl border border-blue-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-blue-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 1 HOW</span>
                        <h3 class="font-bold text-base text-slate-900">제작사 간 통신 프로토콜 & 커넥터 핀 맵 대조 방법 (HOW TO PREPARE)</h3>
                    </div>
                    <ul class="list-disc pl-5 text-slate-700 space-y-1.5 text-xs font-medium leading-relaxed">
                        <li><strong>제작사 프로토콜(RS-485/Ethernet) 핀 맵 대조:</strong> LTE-R, 광전송, PIS/PA, CCTV 제작사 도면을 수집하여 인터페이스 통신 포트 및 핀 맵(Pin Map) 1:1 결선 사양을 확인합니다.</li>
                        <li><strong>입력 전원(AC 220V/DC 48V) 검측:</strong> 각 제작사 장비의 동작 전원 사양을 대조하고 통합 UPS 분전반 결선 및 차단기 용량을 선제 확정합니다.</li>
                    </ul>
                    
                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step1', 'STEP 1 제작사 간 프로토콜 & 커넥터 핀 맵 대조 2D 시공 도식')">
                        <svg id="svg_step1" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- 제작사 A -->
                            <rect x="20" y="20" width="180" height="135" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                            <text x="110" y="45" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">🖥️ 제작사 A (LTE-R/광전송)</text>
                            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• 통신 규격: Ethernet / RJ45</text>
                            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 동작 전원: DC -48V / 15A</text>
                            <text x="35" y="121" font-size="11" font-weight="bold" fill="#2563eb">• Pin 1,2: TX / Pin 3,6: RX</text>

                            <!-- 화살표 핀맵 대조 -->
                            <path d="M 205 85 L 245 85" stroke="#2563eb" stroke-width="3"/>
                            <polygon points="245,80 255,85 245,90" fill="#2563eb"/>

                            <!-- 제작사 B -->
                            <rect x="260" y="20" width="240" height="135" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="380" y="45" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📻 제작사 B (PIS/PA/CCTV)</text>
                            <text x="275" y="75" font-size="11" font-weight="bold" fill="#334155">• 통신 규격: RS-485 / Terminal</text>
                            <text x="275" y="98" font-size="11" font-weight="bold" fill="#334155">• 동작 전원: AC 220V (통합 UPS)</text>
                            <text x="275" y="121" font-size="11" font-weight="bold" fill="#047857">• 핀 맵 1:1 대조 승인 완료</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 2 HOW + 1:1 Visual SVG -->
                <div class="bg-white p-6 rounded-2xl border border-indigo-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-indigo-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 2 HOW</span>
                        <h3 class="font-bold text-base text-slate-900">LTE-R DU/RRU 기지국 & 광전송 물리 연동 방법 (HOW TO MEASURE)</h3>
                    </div>
                    <ul class="list-disc pl-5 text-slate-700 space-y-1.5 text-xs font-medium leading-relaxed">
                        <li><strong>LTE-R 기지국 & 96Core 광배선 1:1 연결:</strong> LTE-R 주제어장치(EPC) 및 DU/RRU 기지국과 96Core/72Core 광전송장치 패널 간 광케이블 코어 1:1 접속 상태를 검측합니다.</li>
                        <li><strong>광 손실률(≤0.2dB/km) 광파워미터 시험:</strong> OTDR 및 광파워미터를 사용하여 제작사 접속 구간의 광 감쇄 손실률(0.2dB/km 이하)을 정밀 측정합니다.</li>
                    </ul>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step2', 'STEP 2 LTE-R 기지국 & 96Core 광전송 물리 연동 2D 시공 도식')">
                        <svg id="svg_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="20" width="220" height="135" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                            <text x="130" y="45" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">📡 LTE-R DU/RRU 기지국</text>
                            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• EPC 무선 주제어장치 수용</text>
                            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 3GPP 표준 패킷 전송 시험</text>
                            <text x="35" y="121" font-size="11" font-weight="bold" fill="#4f46e5">• 광 손실률 ≤ 0.2dB/km 합격</text>

                            <path d="M 245 85 L 275 85" stroke="#4f46e5" stroke-width="3"/>
                            <polygon points="275,80 285,85 275,90" fill="#4f46e5"/>

                            <rect x="290" y="20" width="210" height="135" fill="#ffffff" stroke="#0284c7" stroke-width="2" rx="8"/>
                            <text x="395" y="45" font-size="13" font-weight="black" fill="#0369a1" text-anchor="middle">🌐 96Core 광전송장치 (POTN)</text>
                            <text x="305" y="75" font-size="11" font-weight="bold" fill="#334155">• 상/하선 광케이블 코어 접속</text>
                            <text x="305" y="98" font-size="11" font-weight="bold" fill="#334155">• FDF 광분배함 패널 1:1 점검</text>
                            <text x="305" y="121" font-size="11" font-weight="bold" fill="#0284c7">• 물리적 무결성 100% 확보</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 3 HOW + 1:1 Visual SVG -->
                <div class="bg-white p-6 rounded-2xl border border-cyan-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-cyan-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 3 HOW</span>
                        <h3 class="font-bold text-base text-slate-900">PIS/PA 방송 & 4K IP CCTV 연동 방법 (HOW TO INSTALL & TEST)</h3>
                    </div>
                    <ul class="list-disc pl-5 text-slate-700 space-y-1.5 text-xs font-medium leading-relaxed">
                        <li><strong>PIS/PA 지연시간(≤0.5초) 실시간 시험:</strong> 승강장 PIS 전광판 표출 신호 및 PA 음성 방송 모듈 표출 시 관제 제어 신호와 방송 간 지연 시간(0.5초 이내)을 시뮬레이션 시험합니다.</li>
                        <li><strong>4K IP CCTV 비디오 패킷 검측:</strong> 4K IP CCTV 패킷 전송 손실률(0.01% 이하) 및 관제 모니터링 실시간 영상 전송 화질(4K UHD)을 제작사 합동 확인합니다.</li>
                    </ul>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step3', 'STEP 3 PIS/PA 방송 & 4K IP CCTV 연동 2D 시공 도식')">
                        <svg id="svg_step3" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="20" width="220" height="135" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
                            <text x="130" y="45" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">📺 PIS/PA 방송 모듈</text>
                            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• 승강장 PIS 전광판 표출</text>
                            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• PA 음성 안내 방송 출력</text>
                            <text x="35" y="121" font-size="11" font-weight="bold" fill="#0891b2">• 수신 지연시간 ≤ 0.5초 합격</text>

                            <path d="M 245 85 L 275 85" stroke="#0891b2" stroke-width="3"/>
                            <polygon points="275,80 285,85 275,90" fill="#0891b2"/>

                            <rect x="290" y="20" width="210" height="135" fill="#ffffff" stroke="#0284c7" stroke-width="2" rx="8"/>
                            <text x="395" y="45" font-size="13" font-weight="black" fill="#0369a1" text-anchor="middle">📹 4K IP CCTV 카메라</text>
                            <text x="305" y="75" font-size="11" font-weight="bold" fill="#334155">• 4K UHD 고화질 스트리밍</text>
                            <text x="305" y="98" font-size="11" font-weight="bold" fill="#334155">• 패킷 손실률 ≤ 0.01% 달성</text>
                            <text x="305" y="121" font-size="11" font-weight="bold" fill="#0284c7">• 관제 24시간 실시간 모니터링</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 4 HOW + 1:1 Visual SVG -->
                <div class="bg-white p-6 rounded-2xl border border-teal-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-teal-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 4 HOW</span>
                        <h3 class="font-bold text-base text-slate-900">PSD 비상통화버튼 3초 연동 시험 방법 (HOW TO CONNECT)</h3>
                    </div>
                    <ul class="list-disc pl-5 text-slate-700 space-y-1.5 text-xs font-medium leading-relaxed">
                        <li><strong>PSD 비상통화버튼 조작 시험:</strong> 정거장 PSD 스크린도어 표면에 설치된 비상통화버튼 조작 시 관제실 통신 콘솔로 3초 이내 즉시 착신 연동되는지 제작사 합동 시험합니다.</li>
                        <li><strong>통화 음질 및 잡음 차단 검측:</strong> 차폐 케이블 및 무정전 전원(UPS) 결속으로 통화 시 전력 유도 잡음이 발생하는지 데시벨(dB) 측정기로 현장 검측합니다.</li>
                    </ul>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step4', 'STEP 4 PSD 비상통화버튼 3초 착신 연동 2D 시공 도식')">
                        <svg id="svg_step4" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="20" width="220" height="135" fill="#ffffff" stroke="#0d9488" stroke-width="2" rx="8"/>
                            <text x="130" y="45" font-size="13" font-weight="black" fill="#0f766e" text-anchor="middle">📻 승강장 PSD 비상통화버튼</text>
                            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• PSD 스크린도어 함체 수용</text>
                            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 비상통화 조작 시 자동 발신</text>
                            <text x="35" y="121" font-size="11" font-weight="bold" fill="#0d9488">• 전력 잡음 차단 차폐 적용</text>

                            <path d="M 245 85 L 275 85" stroke="#0d9488" stroke-width="3"/>
                            <polygon points="275,80 285,85 275,90" fill="#0d9488"/>

                            <rect x="290" y="20" width="210" height="135" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="395" y="45" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🎧 관제실 통신 콘솔</text>
                            <text x="305" y="75" font-size="11" font-weight="bold" fill="#334155">• 3초 이내 관제 착신 통화</text>
                            <text x="305" y="98" font-size="11" font-weight="bold" fill="#334155">• 양방향 명확한 음질 통화</text>
                            <text x="305" y="121" font-size="11" font-weight="bold" fill="#047857">• 비상 연동 100% 합격</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 5 HOW + 1:1 Visual SVG -->
                <div class="bg-white p-6 rounded-2xl border border-emerald-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-emerald-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 5 HOW</span>
                        <h3 class="font-bold text-base text-slate-900">TTA/KCA 무선 인증 서류 & 5자 서명 체결 방법 (HOW TO SIGN-OFF)</h3>
                    </div>
                    <ul class="list-disc pl-5 text-slate-700 space-y-1.5 text-xs font-medium leading-relaxed">
                        <li><strong>무선통신 장비 호환성 & 전파 합격증 (TTA/KCA):</strong> 트램 안테나와 무전기 장비가 서로 통신이 잘 터지고 해킹당하지 않는지 국가 공인 기관 인증(TTA)을 받고, 전파법 기준에 맞는지 최종 전파 합격증(KCA 필증) 서류를 확인합니다.</li>
                        <li><strong>제작사 인터페이스 5자 서명 체결:</strong> 통신/전기/신호/차량/PSD 각 제작사 책임 엔지니어가 현장에 동시 입석하여 회의록 및 시스템 통합 대장에 서명 체결합니다.</li>
                    </ul>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step5', 'STEP 5 TTA/KCA 무선 인증서 & 5자 서명 체결 2D 시공 도식')">
                        <svg id="svg_step5" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="20" y="15" width="230" height="150" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="135" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📄 무선통신 호환성 & 전파 합격증</text>
                            <text x="35" y="70" font-size="11" font-weight="bold" fill="#334155">• 안테나·무전기 잘 터지는지 검사 (TTA)</text>
                            <text x="35" y="93" font-size="11" font-weight="bold" fill="#334155">• 무선 해킹 방지 인증서 완료 (TTA)</text>
                            <text x="35" y="116" font-size="11" font-weight="bold" fill="#334155">• 전파법 최종 전파 합격증 확보 (KCA)</text>
                            <text x="35" y="139" font-size="11" font-weight="bold" fill="#047857">• 법정 무선 승인 서류 100% 준비</text>

                            <rect x="270" y="15" width="230" height="150" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="385" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🤝 제작사 5자 서명 관리대장</text>
                            <text x="285" y="70" font-size="11" font-weight="bold" fill="#334155">• LTE-R / 광전송 제작사 날인</text>
                            <text x="285" y="93" font-size="11" font-weight="bold" fill="#334155">• PIS/PA/CCTV/PSD 제작사 날인</text>
                            <text x="285" y="116" font-size="11" font-weight="bold" fill="#334155">• 현장 시스템팀 감리원 최종 승인</text>
                            <text x="285" y="139" font-size="11" font-weight="bold" fill="#047857">• 1,2공구 종합대장 등재 완료</text>
                        </svg>
                    </div>
                </div>
            </div>
        </div>

        <!-- 3. 종합 2D VISUAL SVG DIAGRAM -->
        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-emerald-600 pb-2">
                <span class="text-emerald-600">3.</span> 종합 2D Visual 기술 도식 (Enriched 2D SVG)
            </h2>
            <div class="clickable-diagram bg-slate-50 p-5 rounded-2xl border border-slate-200 shadow-inner" onclick="openDiagramZoom('svg_r4', '[WBS 9000-2-5] 통신설비 제작사 인터페이스 2D Visual 도식')">
                <svg id="svg_r4" viewBox="0 0 550 190" width="100%" height="190" xmlns="http://www.w3.org/2000/svg">
                    <rect x="0" y="0" width="550" height="190" fill="#f8fafc"/>
                    <rect x="25" y="15" width="230" height="145" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                    <text x="140" y="42" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">📡 제작사 프로토콜 & 핀맵</text>
                    <text x="40" y="70" font-size="11" font-weight="bold" fill="#334155">• LTE-R/광전송/PIS 핀맵 1:1 대조</text>
                    <text x="40" y="93" font-size="11" font-weight="bold" fill="#334155">• RS-485 / Ethernet 규격 승인</text>
                    <text x="40" y="116" font-size="11" font-weight="bold" fill="#334155">• 광 손실률 ≤ 0.2dB/km 측정</text>
                    <text x="40" y="139" font-size="11" font-weight="bold" fill="#1d4ed8">• 물리적·기능적 무결성 검증</text>

                    <path d="M 265 85 L 295 85" stroke="#2563eb" stroke-width="3"/>
                    <polygon points="295,80 305,85 295,90" fill="#2563eb"/>

                    <rect x="310" y="15" width="215" height="145" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                    <text x="417" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📄 TTA/KCA & PSD 연동</text>
                    <text x="325" y="70" font-size="11" font-weight="bold" fill="#334155">• PIS/PA 방송 지연 ≤ 0.5초</text>
                    <text x="325" y="93" font-size="11" font-weight="bold" fill="#334155">• PSD 비상통화 3초 착신 연동</text>
                    <text x="325" y="116" font-size="11" font-weight="bold" fill="#334155">• 무선통신 TTA/KCA 필증 완료</text>
                    <text x="325" y="139" font-size="11" font-weight="bold" fill="#047857">• 제작사 5자 서명 체결 완료</text>
                    <text x="275" y="175" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">WBS 9000-2-5 통신설비 제작사 인터페이스 승인 완료</text>
                </svg>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 3. CHECKLIST HTML
chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>통신설비 제작사 인터페이스 협의 마스터 체크리스트 (WBS 9000-2-5)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; background-color: #f8fafc; }}
        .step-header {{ background-color: #ffffff; text-align: center; font-weight: 800; color: #1e3a8a; border-right: 1px solid #e2e8f0; }}
        .result-box {{ background-color: #ffffff; text-align: center; font-weight: 700; color: #2563eb; border-left: 1px solid #e2e8f0; }}
        {zoom_modal_style}
    </style>
</head>
<body class="p-6 sm:p-10 text-slate-800">
<div class="max-w-5xl mx-auto space-y-6">

    <!-- 대제목 & WBS 코드 -->
    <div class="flex justify-between items-end border-b-2 border-slate-900 pb-4">
        <div>
            <h1 class="text-3xl font-black text-slate-900 tracking-tight">통신설비 제작사 인터페이스 협의 마스터 체크리스트</h1>
        </div>
        <div class="text-right">
            <span class="text-blue-600 font-bold text-sm">WBS Code 9000-2-5 | 통신 검측대장</span>
        </div>
    </div>

    <!-- 📋 상단 안내 상자 (Notice Box) -->
    <div class="bg-blue-50/80 border border-blue-200 rounded-2xl p-6 shadow-sm space-y-2">
        <h3 class="text-base font-bold text-blue-950 flex items-center gap-2">
            <span>📋</span> 통신설비 제작사 인터페이스 12대 정밀 검측대장
        </h3>
        <p class="text-xs text-blue-900 leading-relaxed font-medium">
            본 체크리스트는 통신설비 제작사 간 프로토콜, 핀 맵, 광손실률, PIS/PA 및 PSD 비상통화 연동 점검 시 현장 엔지니어가 <strong>[🔍 시공 도식 열기]</strong>를 클릭하면 <strong>대형 고화질 팝업 모달</strong>이 열려 도식을 직접 보며 <strong>~하였는가? (100%)</strong> 점검을 진행할 수 있도록 연동되었습니다.
        </p>
    </div>

    <!-- 숨겨진 현장 점검용 Step 1~3 2D Visual SVG 소스 -->
    <div style="display:none;">
        <svg id="svg_chk_step1" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="20" width="220" height="140" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
            <text x="130" y="45" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">🖥️ 제작사 A (LTE-R/광전송)</text>
            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• RS-485 / Ethernet 핀 맵 대조</text>
            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 동작 전원 AC220V/DC48V 검측</text>
            <text x="35" y="121" font-size="11" font-weight="bold" fill="#1d4ed8">• 무정전 UPS 분전함 결선 완료</text>

            <rect x="270" y="20" width="230" height="140" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="385" y="45" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📻 제작사 B (PIS/PA/CCTV)</text>
            <text x="285" y="75" font-size="11" font-weight="bold" fill="#334155">• 커넥터 핀 맵 1:1 확정</text>
            <text x="285" y="98" font-size="11" font-weight="bold" fill="#334155">• 광 손실률 ≤ 0.2dB/km 측정</text>
            <text x="285" y="121" font-size="11" font-weight="bold" fill="#047857">• 제작사 인터페이스 도면 승인</text>
        </svg>

        <svg id="svg_chk_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="20" width="220" height="140" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
            <text x="130" y="45" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">📺 PIS/PA 방송 & 4K CCTV</text>
            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• PIS/PA 수신 지연 ≤ 0.5초</text>
            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 4K CCTV 손실률 ≤ 0.01%</text>
            <text x="35" y="121" font-size="11" font-weight="bold" fill="#0e7490">• 24시간 실시간 영상 전송 검증</text>

            <rect x="270" y="20" width="230" height="140" fill="#ffffff" stroke="#0d9488" stroke-width="2" rx="8"/>
            <text x="385" y="45" font-size="13" font-weight="black" fill="#0f766e" text-anchor="middle">🎧 PSD 비상통화 3초 연동</text>
            <text x="285" y="75" font-size="11" font-weight="bold" fill="#334155">• 스크린도어 버튼 3초 착신</text>
            <text x="285" y="98" font-size="11" font-weight="bold" fill="#334155">• 전력 유도 잡음 차단 회로</text>
            <text x="285" y="121" font-size="11" font-weight="bold" fill="#0f766e">• 관제실 음질 데시벨 합격</text>
        </svg>

        <svg id="svg_chk_step3" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="15" width="230" height="150" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="135" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📄 무선통신 호환성 & 전파 합격증</text>
            <text x="35" y="70" font-size="11" font-weight="bold" fill="#334155">• 안테나·무전기 잘 터지는지 검사 (TTA)</text>
            <text x="35" y="93" font-size="11" font-weight="bold" fill="#334155">• 무선 해킹 방지 인증서 완료 (TTA)</text>
            <text x="35" y="116" font-size="11" font-weight="bold" fill="#334155">• 전파법 최종 전파 합격증 확보 (KCA)</text>
            <text x="35" y="139" font-size="11" font-weight="bold" fill="#047857">• 법정 무선 승인 서류 100% 준비</text>

            <rect x="270" y="15" width="230" height="150" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="385" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🤝 제작사 5자 서명 관리대장</text>
            <text x="285" y="70" font-size="11" font-weight="bold" fill="#334155">• LTE-R / 광전송 제작사 날인</text>
            <text x="285" y="93" font-size="11" font-weight="bold" fill="#334155">• PIS/PA/CCTV/PSD 제작사 날인</text>
            <text x="285" y="116" font-size="11" font-weight="bold" fill="#334155">• 현장 시스템팀 감리원 최종 승인</text>
            <text x="285" y="139" font-size="11" font-weight="bold" fill="#047857">• 1,2공구 종합대장 등재 완료</text>
        </svg>
    </div>

    <!-- 3-Column 마스터 검측 테이블 -->
    <div class="bg-white border border-slate-200 rounded-2xl shadow-xl overflow-hidden">
        <table class="w-full border-collapse">
            <thead>
                <tr class="bg-slate-100 text-slate-700 text-sm font-extrabold border-b border-slate-200">
                    <th class="py-4 px-6 text-center w-48 border-r border-slate-200">시공 단계</th>
                    <th class="py-4 px-6 text-center">필수 검측 항목 (제작사 인터페이스 정밀 수칙)</th>
                    <th class="py-4 px-6 text-center w-36 border-l border-slate-200">점검 결과</th>
                </tr>
            </thead>
            <tbody class="divide-y divide-slate-200">
                <!-- STEP 1 Group -->
                <tr class="hover:bg-slate-50 transition-colors">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-amber-500 text-base">⚠️</span>
                            <div class="text-sm font-black text-slate-900">사전 준비</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 1 핀맵대조)</div>
                            <button onclick="openDiagramZoom('svg_chk_step1', 'Step 1 핀맵대조 시공 도식')" class="bg-blue-600 hover:bg-blue-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">1. 통신 프로토콜</span><strong class="text-slate-900">[RS-485/Ethernet]</strong> 통신설비 제작사 간 RS-485/Ethernet 통신 프로토콜 및 커넥터 핀 맵 사양을 확정<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">2. 전원 사양</span><strong class="text-slate-900">[통합 UPS]</strong> 각 제작사 장비 입력 전원(AC 220V/DC 48V) 및 통합 UPS 분전반 결선 사양을 확인<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">3. 광손실률</span><strong class="text-slate-900">[≤0.2dB/km]</strong> 96Core/72Core 광전송장치 상/하선 접속 구간 광 손실률(0.2dB/km 이하)을 측정<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">4. LTE-R 기지국</span><strong class="text-slate-900">[DU/RRU 수용]</strong> LTE-R 기지국과 무선 EPC 주제어장치 간 TTA 3GPP 패킷 전송을 검증<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>

                <!-- STEP 2 Group -->
                <tr class="hover:bg-slate-50 transition-colors border-t-2 border-slate-200">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-blue-500 text-base">📋</span>
                            <div class="text-sm font-black text-slate-900">8대 조건</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 2 방송/CCTV)</div>
                            <button onclick="openDiagramZoom('svg_chk_step2', 'Step 2 방송/CCTV 시공 도식')" class="bg-indigo-600 hover:bg-indigo-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">5. PIS/PA 지연</span><strong class="text-slate-900">[≤0.5초]</strong> 승강장 PIS 전광판 표출 및 PA 음성 방송 모듈 수신 지연 시간(0.5초 이내)을 시험<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">6. 4K CCTV</span><strong class="text-slate-900">[패킷 손실률]</strong> 4K IP CCTV 스트리밍 패킷 손실률(0.01% 이하) 및 관제 화질을 검측<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">7. PSD 비상통화</span><strong class="text-slate-900">[3초 착신]</strong> 승강장 PSD 스크린도어 비상통화버튼 조작 시 관제실 3초 착신 연동을 시험<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">8. 전력 잡음차단</span><strong class="text-slate-900">[차폐 회로]</strong> 통화 회로의 전력 유도 잡음 발생 여부를 데시벨(dB) 측정기로 검측<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>

                <!-- STEP 3 Group -->
                <tr class="hover:bg-slate-50 transition-colors border-t-2 border-slate-200">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-emerald-500 text-base">🤝</span>
                            <div class="text-sm font-black text-slate-900">마감 승인</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 3 결과 체결)</div>
                            <button onclick="openDiagramZoom('svg_chk_step3', 'Step 3 마감승인 시공 도식')" class="bg-emerald-600 hover:bg-emerald-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">9. 통신 호환성</span><strong class="text-slate-900">[TTA 인증]</strong> 무선통신 장비 간 끊김 없이 잘 통하고 해킹 안 되는 국가 공인 인증(TTA) 서류를 수신<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">10. 전파 합격증</span><strong class="text-slate-900">[KCA 필증]</strong> 전파법 기준에 맞는 최종 전파 합격증(KCA 필증) 수수료 반영 서류를 확인<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">11. 제작사 5자 서명</span><strong class="text-slate-900">[종합 대장]</strong> 제작사 책임 엔지니어 및 감리원 합동 서명 회의록을 확정<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">12. 결과서 등재</span><strong class="text-slate-900">[1,2공구 등재]</strong> 제작사 인터페이스 승인 서류를 동탄트램 시스템 통합 대장에 등재<strong class="text-blue-600 font-bold">하였는가?</strong></td>
                </tr>
            </tbody>
        </table>
    </div>

</div>
{common_js}
</body>
</html>
"""

sub_dirs = {
    "표준서": std_html,
    "수행지침": gui_html,
    "체크리스트": chk_html
}

for s_n, content in sub_dirs.items():
    sp = os.path.join(target_folder, s_n)
    if os.path.exists(sp):
        for fn in os.listdir(sp):
            if fn.endswith('.html'):
                with open(os.path.join(sp, fn), 'w', encoding='utf-8') as f_out:
                    f_out.write(content)
                print(f"   ✓ [UPGRADED WBS 9000-2-5] {s_n} -> {fn}")

# Excel Sync for Row 5 (WBS 9000-2-5)
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        if "통신분야" in wb.sheetnames:
            ws = wb["통신분야"]
            
            # Row 5 Summary Content
            std_summary = "1) 제작사 프로토콜 확정: LTE-R 기지국, 96Core/72Core 광전송장치, PIS/PA, 4K CCTV 제작사 간 RS-485/Ethernet 통신 프로토콜 및 커넥터 핀 맵 사양 1:1 확정\n2) 전원 및 인터페이스 연동: 장비 입력 전원(AC 220V/DC 48V) 및 통합 UPS 분전반 결선 검측, PIS/PA 방송 지연 ≤ 0.5초, PSD 비상통화 3초 착신 시험, TTA/KCA 무선 인증서 체결"
            gui_summary = "1) 프로토콜 & 핀맵 대조: 제작사 간 RS-485/Ethernet 커넥터 핀 맵 및 입력 전원 사양 1:1 대조\n2) 5단계 시공 가이드: ① 핀맵대조 ➔ ② 기지국연동(광손실률 ≤ 0.2dB/km) ➔ ③ 방송/CCTV(지연 ≤ 0.5초) ➔ ④ PSD 비상통화(3초 착신) ➔ ⑤ TTA/KCA 무선 인증서 5자 서명 체결 3단계 visual 가이드"
            chk_summary = "1) 통신설비 제작사 간 통신 프로토콜, 커넥터 핀 맵 및 입력 전원 사양을 확정하였는가?\n2) LTE-R 기지국, PIS/PA 방송 지연(≤0.5초), PSD 비상통화(3초 착신) 및 TTA/KCA 무선 인증 서류를 확인하고 5자 서명 회의록을 작성하였는가?"

            ws.cell(row=5, column=10, value=std_summary) # Col J
            ws.cell(row=5, column=12, value=gui_summary) # Col L
            ws.cell(row=5, column=14, value=chk_summary) # Col N

            wb.save(excel_path)
            print("   ✓ [EXCEL V4 SYNC COMPLETE] Successfully updated Row 5 (WBS 9000-2-5) in 매뉴얼 BODY (집행단계)v4.xlsx")
    except Exception as e:
        print(f"   ⚠️ Excel Sync Note: {e}")

print("\n🎉 SUCCESSFULLY COMPLETED ULTRA-DETAILED ENHANCEMENT & EXCEL SYNC FOR WBS 9000-2-5!")
