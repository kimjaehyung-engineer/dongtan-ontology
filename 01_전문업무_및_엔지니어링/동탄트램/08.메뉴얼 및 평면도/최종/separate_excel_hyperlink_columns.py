import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue for Link Cols
header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

link_font = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

target_sheets = ['공정표에따른 매뉴얼', '사전토공사', '상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

def sanitize_name(name):
    return re.sub(r'[\/\\:\*\?"<>\|]', '_', str(name)).strip()

processed_count = 0

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows: continue
    
    # Determine header row index (0-indexed)
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        row_vals = [str(c.value) if c.value is not None else "" for c in r]
        if any('표준서' in v for v in row_vals):
            header_row_idx = idx
            break
            
    header_row_num = header_row_idx + 1 # 1-based row number
    
    # Read headers
    headers = [str(ws.cell(row=header_row_num, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    # Find original target column indices (1-indexed)
    std_col = None
    gui_col = None
    chk_col = None
    act_col = None
    l4_col = None
    
    for c_idx, h in enumerate(headers, 1):
        if '표준서' in h and '파일' not in h and '링크' not in h: std_col = c_idx
        elif '수행지침' in h and '파일' not in h and '링크' not in h: gui_col = c_idx
        elif '체크리스트' in h and '파일' not in h and '링크' not in h: chk_col = c_idx
        elif '작업단위' in h or 'Activity' in h: act_col = c_idx
        elif 'L4' in h: l4_col = c_idx

    if not std_col or not gui_col or not chk_col or not act_col:
        print(f"Skipping {sheet_name}: missing required columns (std:{std_col}, gui:{gui_col}, chk:{chk_col}, act:{act_col})")
        continue

    print(f"Processing '{sheet_name}' (Header row: {header_row_num}, std_col: {std_col}, gui_col: {gui_col}, chk_col: {chk_col})...")

    # Insert link columns from right to left so indices don't shift during insertion
    # Insert Checklist Link col after chk_col
    ws.insert_cols(chk_col + 1)
    ws.cell(row=header_row_num, column=chk_col + 1, value="체크리스트 파일 (HTML)")
    ws.cell(row=header_row_num, column=chk_col + 1).fill = header_fill
    ws.cell(row=header_row_num, column=chk_col + 1).font = header_font
    ws.cell(row=header_row_num, column=chk_col + 1).alignment = align_center

    # Insert Guideline Link col after gui_col
    ws.insert_cols(gui_col + 1)
    ws.cell(row=header_row_num, column=gui_col + 1, value="수행지침 파일 (HTML)")
    ws.cell(row=header_row_num, column=gui_col + 1).fill = header_fill
    ws.cell(row=header_row_num, column=gui_col + 1).font = header_font
    ws.cell(row=header_row_num, column=gui_col + 1).alignment = align_center

    # Insert Standard Link col after std_col
    ws.insert_cols(std_col + 1)
    ws.cell(row=header_row_num, column=std_col + 1, value="표준서 파일 (HTML)")
    ws.cell(row=header_row_num, column=std_col + 1).fill = header_fill
    ws.cell(row=header_row_num, column=std_col + 1).font = header_font
    ws.cell(row=header_row_num, column=std_col + 1).alignment = align_center

    # Updated column indices after insertions:
    # std_col is original std_col
    # std_link_col = std_col + 1
    # gui_col = original gui_col + 1
    # gui_link_col = gui_col + 1
    # chk_col = original chk_col + 2
    # chk_link_col = chk_col + 1
    
    std_summary_col = std_col
    std_link_col = std_col + 1
    gui_summary_col = gui_col + 1
    gui_link_col = gui_col + 2
    chk_summary_col = chk_col + 2
    chk_link_col = chk_col + 3
    
    # Update header names for summary columns
    ws.cell(row=header_row_num, column=std_summary_col, value="표준서 (Standard) 요약")
    ws.cell(row=header_row_num, column=gui_summary_col, value="수행지침 (Guideline) 요약")
    ws.cell(row=header_row_num, column=chk_summary_col, value="체크리스트 (Checklist) 요약")

    # Set column widths
    ws.column_dimensions[openpyxl.utils.get_column_letter(std_summary_col)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(std_link_col)].width = 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(gui_summary_col)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(gui_link_col)].width = 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(chk_summary_col)].width = 45
    ws.column_dimensions[openpyxl.utils.get_column_letter(chk_link_col)].width = 22

    sheet_act_count = 0
    disc_dir = sheet_name

    for r_num in range(header_row_num + 1, ws.max_row + 1):
        act_val = ws.cell(row=r_num, column=act_col).value
        l4_val = ws.cell(row=r_num, column=l4_col).value if l4_col else None
        
        if not act_val and not l4_val: continue
        
        sheet_act_count += 1
        act_name = str(act_val).strip() if act_val else f"Task_{sheet_act_count}"
        sanitized_act = sanitize_name(act_name)
        folder_name = f"{sheet_act_count}_{sanitized_act}"
        
        if sheet_name == '공정표에따른 매뉴얼':
            found_disc = None
            found_sub = None
            for d in os.listdir(base_attach_dir):
                d_path = os.path.join(base_attach_dir, d)
                if os.path.isdir(d_path):
                    for sub in os.listdir(d_path):
                        if sanitized_act in sub or (str(l4_val) in sub if l4_val else False):
                            found_disc = d
                            found_sub = sub
                            break
                if found_disc: break
            if found_disc:
                disc_dir = found_disc
                folder_name = found_sub
            else:
                disc_dir = "사전토공사"

        act_dir_abs = os.path.join(base_attach_dir, disc_dir, folder_name)

        doc_types = [
            (std_summary_col, std_link_col, '표준서', '표준서'),
            (gui_summary_col, gui_link_col, '수행지침', '수행지침'),
            (chk_summary_col, chk_link_col, '체크리스트', '체크리스트')
        ]

        for sum_col, link_col, doc_type, doc_title in doc_types:
            sum_cell = ws.cell(row=r_num, column=sum_col)
            link_cell = ws.cell(row=r_num, column=link_col)
            
            # Clean summary cell value (remove hyperlink and button text)
            if sum_cell.value:
                val_str = str(sum_cell.value)
                cleaned_val = re.sub(r'--------------------------------------.*$', '', val_str, flags=re.DOTALL).strip()
                sum_cell.value = cleaned_val
                
            sum_cell.hyperlink = None # Strip hyperlink from summary cell!
            sum_cell.alignment = align_left

            # Build link for link_cell
            sub_folder = os.path.join(act_dir_abs, doc_type)
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    raw_rel_target = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_dir}\\{folder_name}\\{doc_type}\\{files[0]}"
                    link_cell.value = f"👉 [더블클릭] {doc_title} 열기 📄"
                    link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=raw_rel_target)
                    link_cell.font = link_font
                    link_cell.alignment = align_center
                    processed_count += 1

print(f"Separate link columns creation complete! Total {processed_count} links isolated into dedicated columns.")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
