import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
backup_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"

wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

# Target rows for test: 
# Usually, Row 2 (WBS 1_Site Survey Risk 검토), Row 3 (WBS 2_발주전략 KOM), etc.
# Column 15 (Col O) is Checklist Summary (체크리스트 요약)
# Let's inspect rows 2, 3, 4, 5 first to verify.

def clean_and_format_checklist(text):
    if not text:
        return ""
    
    # Split text by typical bullet characters or newlines
    # Matches bullet symbol like •, ·, *, or newline
    raw_sentences = re.split(r'[•·\n]', str(text))
    
    cleaned_sentences = []
    for s in raw_sentences:
        s = s.strip()
        if not s:
            continue
        
        # Remove checkbox character ☐
        s = s.replace("☐", "").replace("[ ]", "").strip()
        
        # Skip technical placeholder lines (e.g. details file hyperlink markers)
        if "상세 체크리스트 파일" in s or "더블클릭" in s or "---" in s:
            continue
            
        # Clean double spaces or duplicate bullet traces
        s = re.sub(r'\s+', ' ', s)
        
        if s:
            cleaned_sentences.append(s)
            
    # Join with standard bullet and newline
    formatted_text = "\n".join([f"• {s}" for s in cleaned_sentences])
    return formatted_text

print("--- BEFORE CORRECTION ---")
for r in [2, 3, 4]:
    l4_code = sheet.cell(row=r, column=4).value
    val = sheet.cell(row=r, column=15).value # Col O
    print(f"Row {r} [{l4_code}]:\n{val}\n")

# Apply test corrections
for r in [2, 3, 4]:
    original_val = sheet.cell(row=r, column=15).value
    corrected_val = clean_and_format_checklist(original_val)
    sheet.cell(row=r, column=15).value = corrected_val

print("--- AFTER CORRECTION ---")
for r in [2, 3, 4]:
    l4_code = sheet.cell(row=r, column=4).value
    val = sheet.cell(row=r, column=15).value
    print(f"Row {r} [{l4_code}]:\n{val}\n")

try:
    wb.save(excel_path)
    print(f"🎉 Saved to Original Excel '{excel_path}' successfully.")
except PermissionError:
    wb.save(backup_excel_path)
    print(f"⚠️ Original Excel is locked. Saved to Backup: '{backup_excel_path}'")
