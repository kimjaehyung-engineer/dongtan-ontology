import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v3_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")

wb = openpyxl.load_workbook(v3_path)
ws = wb['지장물이설']

print("=== Current '지장물이설' Sheet Columns in v3 ===")
print("Max Rows:", ws.max_row, "| Max Cols:", ws.max_column)
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
for i, h in enumerate(headers, 1):
    print(f"  Col {i:2d}: {h}")
