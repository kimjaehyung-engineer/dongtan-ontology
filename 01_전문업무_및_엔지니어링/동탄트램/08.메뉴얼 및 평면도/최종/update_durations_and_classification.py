import openpyxl
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)", "지장물이설")

# 1. Update HTML Files with Dogeub/Witak Classification & Standard Durations
table_html = """
    <!-- 도급자분 vs 위탁자분 8대 관종 정밀 분류 및 소요 공기 체계표 -->
    <h3>📋 지장물 이설 도급자 시행분 및 위탁자 시행분 정밀 분류 기준표</h3>
    <table>
        <thead>
            <tr>
                <th colspan="3" style="text-align: center; background: #dbeafe; color: #1e40af;">도급자 시행 지장물 이설 (연장 기입)</th>
                <th colspan="5" style="text-align: center; background: #ffedd5; color: #9a3412;">위탁자 시행 지장물 이설 현황 (O,X 체크 및 표준 소요 공기)</th>
            </tr>
            <tr>
                <th style="text-align: center;">상수관(m)</th>
                <th style="text-align: center;">하수관(m)</th>
                <th style="text-align: center;">오수관로(m)</th>
                <th style="text-align: center;">가스관 이설 (62일)</th>
                <th style="text-align: center;">열(난방)배관 이설 (80일)</th>
                <th style="text-align: center;">통신관로 이설 (150일)</th>
                <th style="text-align: center;">전력관 이설 (255일)</th>
                <th style="text-align: center;">광역상수관 이설 (160일)</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td>시공사 직접 토공 및 이설</td>
                <td>자연유하 구배 직접 시공</td>
                <td>오수 전용 관로 직접 시공</td>
                <td style="text-align: center; font-weight: bold; color: #ea580c;">삼천리/가스공사 (62일)</td>
                <td style="text-align: center; font-weight: bold; color: #ea580c;">지역난방공사 (80일)</td>
                <td style="text-align: center; font-weight: bold; color: #ea580c;">KT/LGU+/SKT (150일)</td>
                <td style="text-align: center; font-weight: bold; color: #ea580c;">한국전력공사 (255일)</td>
                <td style="text-align: center; font-weight: bold; color: #ea580c;">한국수자원공사 (160일)</td>
            </tr>
        </tbody>
    </table>
"""

print(f"Updating HTML files under '{target_dir}' with Dogeub/Witak durations...")
updated_html_count = 0

for root, dirs, files in os.walk(target_dir):
    for f in files:
        if f.endswith('.html') and '표준서' in f:
            f_path = os.path.join(root, f)
            with open(f_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            if "도급자 시행 지장물 이설" not in content:
                # Insert classification table before footer note or section 3
                if "<h2>3." in content:
                    content = content.replace("<h2>3.", table_html + "\n    <h2>3.")
                elif '<div class="footer-note">' in content:
                    content = content.replace('<div class="footer-note">', table_html + '\n    <div class="footer-note">')

                with open(f_path, 'w', encoding='utf-8') as file:
                    file.write(content)
                updated_html_count += 1

print(f"🎉 Updated {updated_html_count} Standard HTML files with 8-Pipeline Durations table!")

# 2. Update Excel Header with Durations (v3 & v4)
for v_file in ["매뉴얼 BODY (집행단계)v3.xlsx", "매뉴얼 BODY (집행단계)v4.xlsx"]:
    v_path = os.path.join(base_dir, v_file)
    if os.path.exists(v_path):
        print(f"Updating 2-tier Excel headers with durations in '{v_file}'...")
        wb = openpyxl.load_workbook(v_path)
        if '지장물이설' in wb.sheetnames:
            ws = wb['지장물이설']
            
            # Row 1 sub-headers or Row 2 labels
            ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크 (가스: 62일 | 난방: 80일 | 통신: 150일 | 전력: 255일 | 광역상수: 160일)")
            
            ws.cell(row=2, column=14, value="가스관 (62일)")
            ws.cell(row=2, column=15, value="열(난방)배관로 (80일)")
            ws.cell(row=2, column=16, value="통신관로 (150일)")
            ws.cell(row=2, column=17, value="전력관 (255일)")
            ws.cell(row=2, column=18, value="광역상수관 (160일)")

            wb.save(v_path)
            print(f"✅ Successfully updated headers in '{v_file}'!")

print("\n🎉 ALL HTML documents and Excel Workbooks successfully updated with Dogeub/Witak classifications & durations!")
