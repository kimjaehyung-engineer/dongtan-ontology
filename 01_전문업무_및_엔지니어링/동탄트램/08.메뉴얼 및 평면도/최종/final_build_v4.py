import openpyxl
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

v3_file = None
for f in os.listdir(base_dir):
    if f.endswith('.xlsx') and 'v3' in f and not f.startswith('~$'):
        v3_file = os.path.join(base_dir, f)
        break

v4_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

print(f"Source v3 file: {v3_file}")
print(f"Copying to v4 file: {v4_file}")

shutil.copy2(v3_file, v4_file)

wb = openpyxl.load_workbook(v4_file)
ws = wb['지장물이설']

attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)", "지장물이설")
folders = os.listdir(attach_dir)

count = 0
for r in range(3, ws.max_row + 1):
    act_name = str(ws.cell(row=r, column=6).value or '').strip()
    if not act_name:
        continue

    row_prefix = f"{r-2}_"
    matched_folder = None

    for f_n in folders:
        if f_n.startswith(row_prefix) or act_name in f_n:
            matched_folder = f_n
            break

    if matched_folder:
        std_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\표준서\\{act_name}_표준서.html"
        gui_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\수행지침\\{act_name}_수행지침.html"
        chk_rel = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{matched_folder}\\체크리스트\\{act_name}_체크리스트.html"

        ws.cell(row=r, column=20, value=f'=HYPERLINK("{std_rel}", "👉 [더블클릭] 표준서 열기 📄")')
        ws.cell(row=r, column=22, value=f'=HYPERLINK("{gui_rel}", "👉 [더블클릭] 수행지침 열기 📘")')
        ws.cell(row=r, column=24, value=f'=HYPERLINK("{chk_rel}", "👉 [더블클릭] 체크리스트 열기 📋")')

        for col_idx in [20, 22, 24]:
            c = ws.cell(row=r, column=col_idx)
            c.font = openpyxl.styles.Font(name="맑은 고딕", size=9.5, bold=True, color="2563EB", underline="single")
            c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        count += 1

wb.save(v4_file)
print(f"🎉 SUCCESS! Created '매뉴얼 BODY (집행단계)v4.xlsx' with {count} updated activity links!")
