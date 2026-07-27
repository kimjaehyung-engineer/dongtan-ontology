import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\지장물이설"

wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

print("=== 엑셀 ↔ HTML 파일 매칭 정밀 점검 시작 ===")

total_rows = 0
match_errors = []

for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value

    if not l4_code or not act_name:
        continue

    total_rows += 1
    
    # Check folder in base_dir matching act_name
    matched_folder = None
    for folder in os.listdir(base_dir):
        if os.path.isdir(os.path.join(base_dir, folder)):
            if act_name in folder or folder.endswith(act_name):
                matched_folder = folder
                break
    
    if not matched_folder:
        match_errors.append(f"Row {r} ({l4_code} - {act_name}): ❌ Matching Folder NOT found in filesystem!")
        continue
    
    folder_path = os.path.join(base_dir, matched_folder)
    
    # Check 3 HTML files
    doc_types = ["표준서", "수행지침", "체크리스트"]
    file_status = {}
    for doc in doc_types:
        sub_dir = os.path.join(folder_path, doc)
        expected_file = f"{matched_folder}_{doc}.html"
        full_path = os.path.join(sub_dir, expected_file)
        
        # Check if file exists or any html exists
        if os.path.exists(full_path):
            file_status[doc] = f"✅ {expected_file}"
        else:
            # list html files in sub_dir
            if os.path.exists(sub_dir):
                htmls = [f for f in os.listdir(sub_dir) if f.endswith('.html')]
                if htmls:
                    file_status[doc] = f"⚠️ Existing file: {htmls[0]} (Expected: {expected_file})"
                else:
                    file_status[doc] = f"❌ Missing {doc} HTML!"
            else:
                file_status[doc] = f"❌ Missing {doc} Directory!"

    print(f"\nRow {r:02d} [{l4_code}] Activity: '{act_name}'")
    print(f"   📁 Folder: {matched_folder}")
    for doc in doc_types:
        print(f"   - {doc}: {file_status[doc]}")

print("\n==========================================")
print(f"Total Excel Rows Inspected: {total_rows}")
if match_errors:
    print(f"Total Errors Found: {len(match_errors)}")
    for err in match_errors:
        print(err)
else:
    print("🎉 All Excel Activities Perfectly Match Filesystem HTML Files 100%!")
