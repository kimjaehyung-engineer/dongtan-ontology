import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    
    # Find header row
    header_row_idx = 1
    for r in range(1, 6):
        row_vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, ws.max_column + 1)]
        if any('표준서' in v for v in row_vals):
            header_row_idx = r
            break
            
    print(f"\nSheet '{sheet_name}' Header Row = {header_row_idx}:")
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=header_row_idx, column=c).value or "").replace('\n', ' ')
        if v:
            print(f"  Col {c}: {v}")
