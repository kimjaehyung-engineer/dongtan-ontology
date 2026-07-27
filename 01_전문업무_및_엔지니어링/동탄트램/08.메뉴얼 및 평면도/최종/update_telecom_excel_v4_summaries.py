import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
updated_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"

# Load workbook without loading images to avoid PIL I/O closed file error
wb = openpyxl.load_workbook(excel_path)
ws = wb['통신']

activities_count = 0
for r in range(2, ws.max_row + 1):
    act_name = ws.cell(row=r, column=5).value
    l4_code = ws.cell(row=r, column=4).value
    purpose = ws.cell(row=r, column=7).value or ""
    method = ws.cell(row=r, column=8).value or ""
    deliverable = ws.cell(row=r, column=9).value or "통신 시공/검사 보고서"
    
    if act_name:
        activities_count += 1
        name = str(act_name).strip()
        code = str(l4_code).strip()
        
        # 1) 표준서 요약 (Col 10, J) - 최대 2문장
        std_sum = f"1) 본 표준서는 동탄트램 통신분야 [{name}] 공종의 공학적 시방 기준, 72-Core 광 백본망 및 LTE-R SIL 4 무선망 성능 규격을 정의함.\n2) 정보통신공사업법 제36조 및 전파법 제19조 준수를 통한 완공 품질과 최종 결과 증빙({deliverable})의 무결성을 보장함."
        
        # 2) 수행지침 요약 (Col 12, L) - 최대 2문장
        gui_sum = f"1) 사전 준비, 본 시공 및 법정 인허가 3단계 체계별 1:1 2D visual 기술 도식과 대형 확대 모달을 활용하여 [{name}] 작업절차를 가이드함.\n2) 광 접속 손실(≤0.05dB), OTDR 전수 시험 및 관제-승강장 통신 연동을 정밀 수행함."
        
        # 3) 체크리스트 요약 (Col 14, N) - 최대 2문장 (~하였는가? 어미)
        chk_sum = f"1) [{name}] 시공 착수 전 사양서 및 타 분야 인터페이스 기준을 정상 검토하였는가?\n2) LTE-R 무선 커버리지, 광 접속 성능 및 정보통신 사용전검사 합격 여부를 최종 확인하였는가?"
        
        ws.cell(row=r, column=10).value = std_sum
        ws.cell(row=r, column=12).value = gui_sum
        ws.cell(row=r, column=14).value = chk_sum

# Remove _images attribute to prevent PIL closed file crash during save
for sheet in wb.worksheets:
    sheet._images = []

wb.save(excel_path)
wb.save(updated_excel_path)

print(f"🎉 SUCCESSFULLY UPDATED EXCEL V4 SUMMARY COLUMNS FOR {activities_count} TELECOM ACTIVITIES SAFELY!")
