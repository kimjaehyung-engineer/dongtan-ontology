import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v2.xlsx"

if not os.path.exists(excel_path):
    excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)_최종공정매뉴얼완성본.xlsx"

wb = openpyxl.load_workbook(excel_path)

print(f"Loaded Excel: {excel_path}")
print("Sheet names:", wb.sheetnames)

ref_ws = wb['사전토공사']
print("\n=== Reference Sheet '사전토공사' Layout ===")
print("Max cols:", ref_ws.max_column)
headers = [str(ref_ws.cell(row=1, column=c).value or "").replace('\n', ' ') for c in range(1, ref_ws.max_column + 1)]
for i, h in enumerate(headers, 1):
    print(f"  Col {i:2d}: '{h}'")
