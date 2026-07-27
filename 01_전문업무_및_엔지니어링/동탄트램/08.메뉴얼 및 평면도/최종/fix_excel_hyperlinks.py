import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

target_sheets = ['상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야', '사전토공사', '공정표에따른 매뉴얼']

fixed_count = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ['GUIDE']:
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows: continue
    
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(headers)}
    
    std_idx = next((i for h, i in col_map.items() if '표준서' in h), None)
    gui_idx = next((i for h, i in col_map.items() if '수행지침' in h), None)
    chk_idx = next((i for h, i in col_map.items() if '체크리스트' in h), None)
    act_idx = next((i for h, i in col_map.items() if '작업단위' in h or 'Activity' in h), None)
    
    if act_idx is None: continue
    
    sheet_act_count = 0
    disc_dir = sheet_name
    
    for row in rows[1:]:
        act_val = row[act_idx].value
        if not act_val: continue
        
        sheet_act_count += 1
        act_name = str(act_val).strip()
        sanitized_act = act_name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_').strip()
        
        folder_name = f"{sheet_act_count}_{sanitized_act}"
        
        if sheet_name == '공정표에따른 매뉴얼':
            found_disc = None
            found_sub = None
            for d in os.listdir(base_attach_dir):
                d_path = os.path.join(base_attach_dir, d)
                if os.path.isdir(d_path):
                    for sub in os.listdir(d_path):
                        if sanitized_act in sub:
                            found_disc = d
                            found_sub = sub
                            break
                if found_disc: break
            if found_disc:
                disc_dir = found_disc
                folder_name = found_sub
            else:
                disc_dir = "사전토공사"
                
        act_dir_abs = os.path.join(base_attach_dir, disc_dir, folder_name)
        
        tasks = [
            (std_idx, '표준서'),
            (gui_idx, '수행지침'),
            (chk_idx, '체크리스트')
        ]
        
        for c_idx, doc_type in tasks:
            if c_idx is None or c_idx >= len(row): continue
            cell = row[c_idx]
            sub_folder = os.path.join(act_dir_abs, doc_type)
            
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    # Unencoded raw relative path for Windows Excel!
                    raw_rel_target = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_dir}\\{folder_name}\\{doc_type}\\{files[0]}"
                    
                    # Create openpyxl Hyperlink with raw unencoded target
                    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=raw_rel_target)
                    fixed_count += 1

print(f"Fixed {fixed_count} hyperlinks with unencoded Windows paths!")
wb.save(excel_path)
print(f"Saved to '{excel_path}'")
