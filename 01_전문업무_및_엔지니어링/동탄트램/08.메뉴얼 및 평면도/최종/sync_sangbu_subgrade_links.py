import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
backup_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반"

wb = openpyxl.load_workbook(excel_path)
sheet = wb['상부강화노반']

# List actual folders on disk
folders_on_disk = [f for f in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, f))]

def find_matched_folder(seq_num):
    prefix = f"{seq_num}_"
    for f in folders_on_disk:
        if f.startswith(prefix):
            return f
    return None

updated_links_count = 0

for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    if not l4_code:
        continue
    
    # Extract sequence number from level 4 code (e.g. 9000-7-1 -> 1, 9000-7-29 -> 29)
    try:
        seq_num = int(str(l4_code).split("-")[-1])
    except ValueError:
        continue

    folder_name = find_matched_folder(seq_num)
    if not folder_name:
        print(f"⚠️ Row {r}: No folder on disk starting with '{seq_num}_'")
        continue

    # Create relative paths
    std_rel = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/표준서/{folder_name}_표준서.html"
    gui_rel = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/수행지침/{folder_name}_수행지침.html"
    chk_rel = f"매뉴얼BODY(집행단계-첨부폴더)/상부강화노반/{folder_name}/체크리스트/{folder_name}_체크리스트.html"

    # Cell coordinates
    # Col 12 (L): 표준서 파일 (HTML)
    # Col 14 (N): 수행지침 파일 (HTML)
    # Col 16 (P): 체크리스트 파일 (HTML)
    c_std = sheet.cell(row=r, column=12)
    c_gui = sheet.cell(row=r, column=14)
    c_chk = sheet.cell(row=r, column=16)

    c_std.value = "👉 [더블클릭] 표준서 열기 📄"
    c_std.hyperlink = std_rel
    
    c_gui.value = "👉 [더블클릭] 수행지침 열기 📄"
    c_gui.hyperlink = gui_rel
    
    c_chk.value = "👉 [더블클릭] 체크리스트 열기 📄"
    c_chk.hyperlink = chk_rel

    updated_links_count += 1
    print(f"Row {r:02d} [{l4_code}] ➔ Links Synced to Folder: '{folder_name}'")

try:
    wb.save(excel_path)
    print(f"🎉 Successfully Saved Links to Original Excel '{excel_path}' ({updated_links_count} Rows)!")
except PermissionError:
    wb.save(backup_excel_path)
    print(f"⚠️ Original Excel File is Open. Saved to Backup: '{backup_excel_path}' ({updated_links_count} Rows)!")
