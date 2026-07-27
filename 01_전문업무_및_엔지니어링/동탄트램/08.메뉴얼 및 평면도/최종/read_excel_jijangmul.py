import openpyxl
import os

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"

wb = openpyxl.load_workbook(excel_path)
print("Sheet Names:", wb.sheetnames)

sheet = wb['지장물이설']
excel_activities = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row and len(row) > 1 and row[1]:
        wbs = row[0]
        act_name = str(row[1]).strip()
        excel_activities.append((wbs, act_name))

print(f"Total Activities in Excel ('지장물이설'): {len(excel_activities)}")
for idx, (wbs, name) in enumerate(excel_activities, 1):
    print(f"Excel {idx}: WBS={wbs} | Name={name}")
