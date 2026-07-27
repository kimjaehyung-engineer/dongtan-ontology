import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\지장물이설"

wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

# Helper to normalize activity name matching
def normalize(name):
    return name.replace(" ", "").replace("_", "").replace("/", "").replace("(", "").replace(")", "").replace("-", "").lower()

existing_folders = [f for f in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, f))]

print("Starting Excel ↔ HTML File Path Synchronization...")

matched_count = 0

for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value

    if not l4_code or not act_name:
        continue

    norm_act = normalize(act_name)
    
    # Find matching folder in base_dir
    target_folder = None
    for folder in existing_folders:
        norm_folder = normalize(folder)
        if norm_act in norm_folder or norm_folder in norm_act:
            target_folder = folder
            break
            
    if target_folder:
        matched_count += 1
        folder_path = os.path.join(base_dir, target_folder)
        
        # Verify and Update Hyperlink or cell text in Excel
        # Col 12 (L), Col 14 (N), Col 16 (P) contain std, gui, chk links or text
        std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/지장물이설/{target_folder}/표준서/{target_folder}_표준서.html"
        gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/지장물이설/{target_folder}/수행지침/{target_folder}_수행지침.html"
        chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)/지장물이설/{target_folder}/체크리스트/{target_folder}_체크리스트.html"
        
        # We set hyperlink and value
        c_std = sheet.cell(row=r, column=12)
        c_gui = sheet.cell(row=r, column=14)
        c_chk = sheet.cell(row=r, column=16)
        
        c_std.hyperlink = std_rel_path
        c_gui.hyperlink = gui_rel_path
        c_chk.hyperlink = chk_rel_path
        
        print(f"Row {r:02d} [{l4_code}] Activity: '{act_name}' ➔ Matched Folder: '{target_folder}'")
        print(f"   - Std Path: {std_rel_path}")

wb.save(excel_path)
print(f"\n🎉 Successfully Synchronized {matched_count} Activities in Excel '지장물이설' Sheet with HTML files!")
