import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

def clean_lines_perfectly(val):
    if not val: return ""
    lines = str(val).splitlines()
    clean_items = []
    for l in lines:
        l_str = l.strip()
        # Remove any leading digits, parens, bullets
        l_str = re.sub(r'^\s*[\d\.\-\)\(☐☑📍📌]+\s*', '', l_str).strip()
        l_str = re.sub(r'^\s*[\d\.\-\)\(☐☑📍📌]+\s*', '', l_str).strip()
        if l_str and l_str not in clean_items:
            clean_items.append(l_str)
    
    selected = clean_items[:3]
    if not selected: return ""
    return "\n".join([f"{i+1}) {item}" for i, item in enumerate(selected)])

cleaned = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    header_row = 3 if sheet_name == '공정표에따른 매뉴얼' else 1
    
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            header_val = str(ws.cell(row=header_row, column=c).value or "")
            if ('표준서' in header_val or '수행지침' in header_val or '체크리스트' in header_val) and '파일' not in header_val:
                cell = ws.cell(row=r, column=c)
                if cell.value:
                    res = clean_lines_perfectly(cell.value)
                    if res and res != cell.value:
                        cell.value = res
                        cleaned += 1

print(f"Perfect cleanup applied to {cleaned} cells!")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
