import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

v2_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v2.xlsx"
wb = openpyxl.load_workbook(v2_excel_path)
ws = wb['공정매뉴얼']

header_row = 3

print("=== Verification of v2 Workbook '공정매뉴얼' Sheet ===")
print("Total rows:", ws.max_row, "| Total cols:", ws.max_column)

print("\nSampling Updated Systems Rows (Row 45, 48, 49):")
for r in [45, 48, 49]:
    l4 = ws.cell(row=r, column=4).value
    act = ws.cell(row=r, column=5).value
    disc = ws.cell(row=r, column=16).value
    std_sum = str(ws.cell(row=r, column=10).value).replace('\n', ' ')
    chk_sum = str(ws.cell(row=r, column=14).value).replace('\n', ' ')
    link11 = ws.cell(row=r, column=11).hyperlink is not None
    link13 = ws.cell(row=r, column=13).hyperlink is not None
    link15 = ws.cell(row=r, column=15).hyperlink is not None
    
    print(f"\n[Row {r}] L4='{l4}' | Act='{act}' | Disc='{disc}'")
    print(f"  - 표준서 요약: {std_sum[:80]}...")
    print(f"  - 체크리스트 요약: {chk_sum[:80]}...")
    print(f"  - Links OK: Std={link11}, Gui={link13}, Chk={link15}")
