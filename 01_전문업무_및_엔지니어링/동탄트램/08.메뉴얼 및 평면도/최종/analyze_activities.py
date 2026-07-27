import openpyxl
import os
import re

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheet Names:", wb.sheetnames)

for sheet in wb.sheetnames:
    if sheet in ['공정표에따른 매뉴얼', 'GUIDE']:
        continue
    
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    
    # Find header row
    header_row_idx = 0
    for idx, r in enumerate(rows[:5]):
        if any(isinstance(val, str) and ('코드' in val or '작업단위' in val) for val in r if val):
            header_row_idx = idx
            break
            
    headers = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
    
    # Find column indices for L4 code and Activity name
    l4_idx = None
    act_idx = None
    
    for i, h in enumerate(headers):
        if 'L4' in h:
            l4_idx = i
        elif '작업단위' in h or 'Activity' in h:
            act_idx = i
            
    print(f"\n--- Sheet: {sheet} (Header Row: {header_row_idx+1}) ---")
    print(f"L4 col idx: {l4_idx}, Act col idx: {act_idx}")
    
    activities = []
    for r_idx, r in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):
        if all(v is None for v in r):
            continue
        l4_val = r[l4_idx] if l4_idx is not None and l4_idx < len(r) else None
        act_val = r[act_idx] if act_idx is not None and act_idx < len(r) else None
        
        if l4_val or act_val:
            activities.append((r_idx, str(l4_val).strip() if l4_val else "", str(act_val).strip() if act_val else ""))
            
    print(f"Total activities found: {len(activities)}")
    for item in activities[:5]:
        print("  ", item)
