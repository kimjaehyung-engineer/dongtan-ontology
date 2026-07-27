import openpyxl
from bs4 import BeautifulSoup
import urllib.parse
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

target_sheets = ['상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

def extract_summary_from_html(html_path, doc_type):
    if not os.path.exists(html_path):
        return None
        
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')
        
    lines = []
    
    if doc_type == '표준서':
        # Look for spec lists or table or bullet lists
        spec_box = soup.find('div', class_='spec-box')
        if spec_box:
            for li in spec_box.find_all('li'):
                lines.append(li.get_text().strip())
        
        # Also check designer criteria or subcontractor advisory if available
        advisory_box = soup.find('div', class_='subcontractor-advisory-box')
        if advisory_box:
            for li in advisory_box.find_all('li'):
                lines.append("📌 " + li.get_text().strip())
                
        if not lines:
            # Fallback to general list items
            for li in soup.find_all('li')[:5]:
                lines.append(li.get_text().strip())
                
    elif doc_type == '수행지침':
        # Look for step sections or bullet list
        step_boxes = soup.find_all('div', class_='step-box')
        if step_boxes:
            for s in step_boxes:
                title = s.find(['h3', 'h4', 'strong'])
                title_text = title.get_text().strip() if title else ""
                desc = s.get_text().replace(title_text, "").strip() if title else s.get_text().strip()
                desc_short = desc.split('\n')[0][:100]
                lines.append(f"{title_text}: {desc_short}")
        else:
            for li in soup.find_all('li')[:5]:
                lines.append(li.get_text().strip())
                
    elif doc_type == '체크리스트':
        # Look for checklist items (☐ or li)
        chk_items = soup.find_all('li')
        for li in chk_items:
            txt = li.get_text().strip()
            if '☐' in txt or '확인' in txt or '인가' in txt or '검측' in txt:
                lines.append(txt)
        if not lines:
            for li in chk_items[:6]:
                lines.append(li.get_text().strip())

    # Format into numbered 1), 2), 3) list
    formatted_bullets = []
    for i, line in enumerate(lines[:6]):
        clean_line = re.sub(r'^\d+[\.\)]\s*', '', line) # remove existing numbers
        clean_line = re.sub(r'^[☐☑📍📌]\s*', '', clean_line).strip()
        if clean_line:
            formatted_bullets.append(f"{i+1}) {clean_line}")
            
    if not formatted_bullets:
        return "1) 시방 및 정량 기술 스펙 준수\n2) 안전/품질 검측 기준 및 협력사 자문 수칙 적용"
        
    return "\n".join(formatted_bullets)

updated_cells = 0

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows:
        continue
        
    header_row = rows[0]
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_row]
    col_map = {h: i for i, h in enumerate(headers)}
    
    std_idx = None
    gui_idx = None
    chk_idx = None
    act_idx = None
    l4_idx = None
    
    for h, i in col_map.items():
        if '표준서' in h: std_idx = i
        elif '수행지침' in h: gui_idx = i
        elif '체크리스트' in h: chk_idx = i
        elif '작업단위' in h or 'Activity' in h: act_idx = i
        elif 'L4' in h: l4_idx = i
        
    if act_idx is None:
        continue
        
    sheet_act_count = 0
    disc_dir = sheet_name
    
    for r_idx, row in enumerate(rows[1:], start=2):
        act_val = row[act_idx].value
        if not act_val:
            continue
            
        sheet_act_count += 1
        act_name = str(act_val).strip()
        sanitized_act = re.sub(r'[\/\\:\*\?"<>\|]', '_', act_name).strip()
        
        folder_name = f"{sheet_act_count}_{sanitized_act}"
        act_dir_abs = os.path.join(base_attach_dir, disc_dir, folder_name)
        
        # Process 표준서, 수행지침, 체크리스트
        tasks = [
            (std_idx, '표준서', '표준서'),
            (gui_idx, '수행지침', '수행지침'),
            (chk_idx, '체크리스트', '체크리스트')
        ]
        
        for c_idx, doc_type, doc_title in tasks:
            if c_idx is None or c_idx >= len(row):
                continue
                
            cell = row[c_idx]
            sub_folder = os.path.join(act_dir_abs, doc_type)
            
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    html_path = os.path.join(sub_folder, files[0])
                    summary_text = extract_summary_from_html(html_path, doc_type)
                    
                    if summary_text:
                        rel_target = f"매뉴얼BODY(집행단계-첨부폴더)/{disc_dir}/{folder_name}/{doc_type}/{files[0]}"
                        target_encoded = urllib.parse.quote(rel_target, safe='/')
                        
                        btn_text = f"\n--------------------------------------\n👉 [더블클릭] 상세 {doc_title} 파일(HTML) 열기 📄"
                        cell.value = (summary_text + btn_text).strip()
                        cell.hyperlink = target_encoded
                        updated_cells += 1

print(f"Populating Excel summaries complete!")
print(f"  - Total Excel cells updated with rich summaries: {updated_cells}")

wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
