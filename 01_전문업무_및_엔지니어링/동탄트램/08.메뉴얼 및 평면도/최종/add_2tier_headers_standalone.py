import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

def sanitize_name(name):
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', str(name)).strip()

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v3_save_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
base_attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)")
jijangmul_attach_dir = os.path.join(base_attach_dir, "지장물이설")

jijangmul_activities_v3 = [
    # Section A: 사전준비 및 Risk 검토 (확장)
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-1", "dday": "P-30", "act": "Site Survey Risk 검토", "own": "토목국내견적팀", "gol": "현장 현황 및 여건 중 시공에 영향을 줄 수 있는 Risk 사전 검토", "mtd": "설계도서 불일치 구간 실행예산 변경 여부 검토 및 설계변경 불가시 Risk 최소화 방안 검토 (참석: 현장/견적/설계/업체 Pool 우수 토공업체)", "del": "Site Survey 검토서", "std_sum": "1) 설계도서와 현장 불일치 구간 100% 사전 조사 및 실행예산 변경 검토\n2) 현장/견적/설계 합동 Risk 최소화 방안 수립 준수", "gui_sum": "1) 우수 토공업체 Pool 검토를 통한 현장 투입 가능성 검증\n2) 현장 여건 대조 및 추가 공사비 발생 가능성 사전 차단", "chk_sum": "1) Site Survey 검토서 작성 및 관련 부서 서명 승인을 완료했는가?\n2) 설계 불일치 구간에 대한 실행예산 반영 적정성을 검측했는가?", "disc": "지장물이설", "des": "토목국내견적 지침 및 계약 시방", "risk": "설계도서 불일치로 인한 실행예산 초과 리스크 관리", "sub": "우수 토공업체 Pool 기술 자문"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-2", "dday": "P+0", "act": "발주전략 KOM (도급지분)", "own": "현장소장", "gol": "도급자 지장물 이설 업체 선정 및 위수탁기관 협의 검토", "mtd": "직영분/외주분 구분 검토, 현설조건 검토, 장비 Requirement 검토, 공법 및 공기 검토 (지원: 본사외주부서, 수행부서, 토목국내견적팀, 컴플라이언스RM팀)", "del": "KOM 검토보고서, 협설서 (하자방지 특기조건 포함)", "std_sum": "1) CML 토목사무지침 준수 및 도급 지장물 이설 전문업체 적격 검토\n2) 직영/외주 구분 및 발주 전 구조안전성/공기 타당성 100% 확보", "gui_sum": "1) 도급자 이설분 이설업체 계약 시 하자방지 특기시방 조건 명시\n2) 장비 Requirement 및 컴플라이언스 RM 검토 수칙 이행", "chk_sum": "1) KOM 검토보고서 및 협설서 특기조건 반영 여부를 확인했는가?\n2) 이설 전문업체 계약 전 시공 능력 및 장비 사양을 검측했는가?", "disc": "지장물이설", "des": "CML 토목사무지침 및 도급 계약 시방", "risk": "부적격 업체 선정에 따른 공기 지연 리스크 방지", "sub": "본사외주부서 및 컴플라이언스RM 자문"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-3", "dday": "P+0", "act": "지장물 이설 요청 (위수탁고)", "own": "공사팀 / 발주기관", "gol": "위수탁기관 지장물 이설 착수 공문 발송", "mtd": "양질의 위수탁기관 협의를 통한 발주기관 위수탁 지장물 정식 이설 요청 (참석: 공사팀, 발주기관)", "del": "지장물 이설요청 공문", "std_sum": "1) 위수탁기관별 관련 법령에 따른 이설 착수 요청 공문 정식 발송\n2) 발주기관 연계 이설 범위 및 예산 분담 범위 명시", "gui_sum": "1) 이설 대상 관로 위치도 및 개략 수량 첨부 문서화\n2) 위수탁기관 회신 기한(14일 이내) 명시 관리", "chk_sum": "1) 위수탁기관 공식 이설 요청 공문 발송 및 접수를 확인했는가?\n2) 첨부 위치도 및 이설 범위 정확성을 검측했는가?", "disc": "지장물이설", "des": "지자체 위수탁 협약 처리 기준", "risk": "위수탁 기관 이설 착수 지연 리스크 관리", "sub": "관련 유관기관 전담 자문"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-4", "dday": "P+5", "act": "도급자분 이설업체 선정(상/하수)", "own": "현장소장", "gol": "상하수도 도급자분 이설공사 수행을 위한 적격 하도급업체 선정", "mtd": "유관기관 협의 결과 및 설계변경 진행 연계, 시공/수행/견적/계약 합동 평가", "del": "지장물 이설 하도급업체 선정서", "std_sum": "1) 상하수도 이설 전문면허 보유 업체 하도급 적격 심사 85점 이상 달성\n2) 설계변경 반영 물량 및 공사 수행 능력 100% 검증", "gui_sum": "1) 하도급 계약 시 이설 수밀 시험 및 안전 시방 조항 명시\n2) 외주계약팀 승인 후 최종 이설 하도급 계약 체결", "chk_sum": "1) 상하수도 이설 전문업체 적격 심사표 및 계약서를 확인했는가?\n2) 현장 대리인 및 필수 배관 기술자 배치 여부를 검측했는가?", "disc": "지장물이설", "des": "하도급거래 공정화에 관한 법률 및 계약 규정", "risk": "저가 하도급으로 인한 부실 시공 리스크 예방", "sub": "외주계약팀 전문 자문"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-5", "dday": "P+10", "act": "지장물 조사 (위탁기관 합동)", "own": "현장 공사팀", "gol": "설계도서와 지장물 관리기관 도면 일치 및 현장 맨홀 위치 재확인", "mtd": "수행/견적/계약/발주기관/유관부서/지장물이설업체 합동 현장 줄탐지 및 맨홀 확인", "del": "착공전 측량보고서, 현장 조사 보고서 (민원 포함)", "std_sum": "1) 관리기관 지장물 도면과 실제 현장 맨홀/밸브 위치 100% 현장 대조\n2) 착공 전 측량 성과표 및 인접 주민 민원 우려 요소 사전 도출", "gui_sum": "1) 유관기관 감독관 및 이설업체 기술자 현장 동행 조사\n2) 위치 일치 여부 현장 조사 보고서 및 사진대지 작성", "chk_sum": "1) 위탁기관 합동 현장 조사 보고서 및 측량보고서를 확인했는가?\n2) 현장 맨홀 및 노출 관로 심도 일치 여부를 검측했는가?", "disc": "지장물이설", "des": "지하안전관리에 관한 특별법 및 측량 규정", "risk": "도면 상이로 인한 신규 지장물 불시 도출 리스크 방지", "sub": "지장물 이설업체 및 유관기관 담당자"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-6", "dday": "P+15", "act": "지장물 이설 계획 수립", "own": "현장 공사팀", "gol": "인허가, 교통처리대책, 이설 설계 기준 및 시공방법 종합 수립", "mtd": "공사/공무/발주기관/유관기관/지장물 이설업체 협의 (우선순위, 이격거리 확정)", "del": "지장물 이설 종합계획서 (우선순위, 이격거리, 시공방법 포함)", "std_sum": "1) 트램 궤도 구조물과 지장물 수평/수직 최소 이격거리(1.5m 이상) 확보\n2) 관종별 이설 우선순위(상하수 ➔ 가스 ➔ 전력/통신) 확정 준수", "gui_sum": "1) 인허가 및 교통처리 대책과 연계된 릴레이 이설 일정 수립\n2) 이설계획서 감리단 및 발주기관 정식 승인 수칙", "chk_sum": "1) 지장물 이설 종합계획서 및 이격거리 적정성을 확인했는가?\n2) 감리단 승인서 및 관종별 시공 순서 타당성을 검측했는가?", "disc": "지장물이설", "des": "동탄트램 지장물 이설 종합 지침", "risk": "이설 순서 혼선으로 인한 재굴착 리스크 방지", "sub": "지장물 전문가 및 공무팀"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-7", "dday": "P+20", "act": "최고의 팀 만들기 지원", "own": "외주관리팀", "gol": "투입가능한 최고의 자재생산/제작/장비/협력업체/작업반 선정", "mtd": "관리 Pool 중 투입가능 협력업체 및 작업반 기술력/실적 검토", "del": "우수 업체/작업반 입찰참여 승인서", "std_sum": "1) 외주 관리 Pool 시공 실적 및 현장 투입 가능성 다각 검토\n2) 기술력 및 안전 관리 평가 우수 작업반 선정", "gui_sum": "1) 지장물 이설 특화 장비 및 전문 인력 확보 확인\n2) 하도급 적격 심사 기준 85점 이상 달성 수칙", "chk_sum": "1) 협력업체 기술 평가서 및 입찰참여 승인서를 확인했는가?\n2) 작업반 전담 반장의 트램 지장물 시공 실적을 검측했는가?", "disc": "지장물이설", "des": "외주 하도급 관리 규정", "risk": "시공 미숙 작업반 투입으로 인한 사고 리스크 방지", "sub": "외주계약 전문 자문"},
    {"l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-8", "dday": "D-60", "act": "착수전 Big Room 회의", "own": "현장소장", "gol": "도급자 지장물 이설 계획 수립단계에서 원가절감을 위한 이설계획 검토", "mtd": "유관기관(맑은물, 난방, 한전, 통신, 가스)과 이설 순서 협의, 타 시설물 간섭 및 노반/궤도 착수 전 공정 검토", "del": "Big Room 회의록 (원가절감 반영)", "std_sum": "1) 맑은물사업소, 한전, 가스공사 등 유관기관 합동 이설 순서 조율\n2) 굴착 간섭 제어 및 강화노반/궤도 공정 연계 원가절감안 반영", "gui_sum": "1) 도급자 이설분과 위수탁분 시공 간섭 3D 검토 시행\n2) 회의 결정사항 주간 단위 이행 점검 체계 가동", "chk_sum": "1) 유관기관 합동 Big Room 회의록 작성 및 서명을 완료했는가?\n2) 지장물 이설 순서와 노반/궤도 공정 연동성을 확인했는가?", "disc": "지장물이설", "des": "동탄트램 사업관리 지침", "risk": "유관기관 간 공정 중복으로 인한 중복 굴착 리스크 방지", "sub": "유관기관 관리자 및 분야별 전문가"},

    # Section A-2: 사전 준비 점선 서브 프로세스 (인허가/교통/민원/용지)
    {"l3_code": "2000-1S", "l3_name": "사전준비: 공동 대응 방안 수립", "l4_code": "2000-1S-1", "dday": "D-50", "act": "인허가 절차 진행", "own": "현장 공무팀", "gol": "공사 진행에 따른 인허가 절차 적기 시행", "mtd": "용지도 확인 및 공사 구간 내 인허가 필요 구간 확인 ➔ 기관별 승인 접수", "del": "인허가 항목별 승인서 및 협의 공문", "std_sum": "1) 공사 구간 내 법정 인허가(도로점용, 하천점용, 구거사용 등) 100% 도출\n2) 착공 30일 전 인허가 신청서 정밀 제출 및 승인 획득", "gui_sum": "1) 인허가 조건부 승인 사항 현장 이행 관리대정 기록\n2) 점용료 및 사용료 사전 정산 수칙 준수", "chk_sum": "1) 인허가 항목별 승인서 획득 여부를 확인했는가?\n2) 허가 조건 이행 대장 작성 적정성을 검측했는가?", "disc": "지장물이설", "des": "인허가 관련 행정법 및 도로법", "risk": "인허가 지연으로 인한 공사 중단 리스크 방지", "sub": "인허가 전문 자문"},
    {"l3_code": "2000-1S", "l3_name": "사전준비: 공동 대응 방안 수립", "l4_code": "2000-1S-2", "dday": "D-45", "act": "교통처리대책 협의 및 승인", "own": "현장 공무팀/안전팀", "gol": "운영기간, 필요물량, 임대비용 적정성 검토 및 교통 승인", "mtd": "사토장/토취장 list-up 및 관련 L/L과 고자검토 ➔ 경찰서 및 관청 교통협의", "del": "교통처리대책 심의 승인서 (운영 원가/안전/품질 Risk 도출)", "std_sum": "1) 경찰서 교통안전시설 심의 승인 및 가변차로 운용 규정 준수\n2) 사토장/토취장 덤프트럭 운행 동선 안전 수칙 적용", "gui_sum": "1) 교통 우회 안내 표지판 설치 및 신호 유도원 배치\n2) 출퇴근 시간대 차선 통제 최소화 관리 수칙", "chk_sum": "1) 경찰서 교통처리대책 심의 승인서를 확인했는가?\n2) 임시 교통안전 시설물 적정성을 검측했는가?", "disc": "지장물이설", "des": "도로교통법 및 교통영향평가 지침", "risk": "교통 혼잡민원 및 안전사고 발생 리스크 예방", "sub": "교통영향평가 전문 기술자"},
    {"l3_code": "2000-1S", "l3_name": "사전준비: 공동 대응 방안 수립", "l4_code": "2000-1S-3", "dday": "D-40", "act": "민원 저감 대책 수립", "own": "현장 공무팀/민원팀", "gol": "민원으로 인한 공사 지연 및 지장 사전 방지", "mtd": "민원우려 지역 주민 면담, 소음/진동/소독 요구사항 정리 및 사전 안내", "del": "민원 저감대책 수립 및 이행보고서", "std_sum": "1) 현장 소음/진동 법정 기준(소음 65dB 이하) 준수 방음벽 설치\n2) 작업 전 인접 상가 및 주민 대상 사전 공사 안내문 현장 배포", "gui_sum": "1) 24시간 민원 전용 창구 가동 및 민원 처리 이력서 기록\n2) 살수차 상시 가동으로 비산먼지 민원 차단 수칙", "chk_sum": "1) 주민 현장 설명회 실시 및 안내문 배포 여부를 확인했는가?\n2) 소음/진동 저감 대책 및 살수차 가동을 검측했는가?", "disc": "지장물이설", "des": "환경보전법 및 소음진동관리법", "risk": "집단 민원에 따른 현장 봉쇄 및 집회 리스크 방지", "sub": "민원 전담 현장 대리인"},
    {"l3_code": "2000-1S", "l3_name": "사전준비: 공동 대응 방안 수립", "l4_code": "2000-1S-4", "dday": "D-35", "act": "용지보상 Risk 파악", "own": "현장 공무팀 / 토목계약관리팀", "gol": "용지 보상 지연에 따른 공정 지연 문제점 사전 도출", "mtd": "미보상 용지 표시 및 용지 보상 시기 확인 ➔ 기공승낙 합의서 획득", "del": "용지보상 현황표 및 미보상 용지 Risk 검토서", "std_sum": "1) 트램 궤도 및 지장물 이설 구간 토지 보상 100% 사전 점검\n2) 미보상 토지 진입 금지 경계 표지판 및 현황도 현치 준수", "gui_sum": "1) 지자체 토지보상과 수시 협의로 보상 일정 독촉\n2) 미보상 토지 우회 시공 가능 여부 사전 타당성 검토", "chk_sum": "1) 용지보상 현황표 작성 및 미보상 구역 표시를 확인했는가?\n2) 기공승낙서 획득 및 보상 완료 여부를 검측했는가?", "disc": "지장물이설", "des": "토지보상법 및 용지 보상 조례", "risk": "사유지 무단 침범에 따른 형사 고소 리스크 방지", "sub": "토지 보상 전문 행정사 자문"},

    # Section B. 행정 및 설계 승인
    {"l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-1", "dday": "D-40", "act": "관리기관(맑은물사업소) 협의", "own": "현장 공무팀", "gol": "위탁기관 협의 결과에 따른 상하수도 이설 계획 맑은물사업소 협의 진행", "mtd": "실시설계 유관기관 사전협의 결과 반영된 상하수도 이설계획 검토 및 재협의", "del": "상하수도 이설 도면, 수량, 내역", "std_sum": "1) 화성시 맑은물사업소 기술 시방 준수 및 상하수도 이설 위치 승인\n2) 관경별 관재 선정 및 이설 수량/내역 정밀 대조", "gui_sum": "1) 기존 관로 연계 부위 수압 및 자연유하 경사 사전 확인\n2) 사업소 요청 추가 반영사항 도출 및 내역 수정", "chk_sum": "1) 맑은물사업소 상하수도 이설 계획 협의서 및 도면을 확인했는가?\n2) 이설 내역서 수량과 도면 일치 여부를 검측했는가?", "disc": "지장물이설", "des": "맑은물사업소 상하수도 시설기준", "risk": "상하수도 이설 위치 불일치로 인한 재협의 리스크 방지", "sub": "상하수도 전문 기술사 자문"},
    {"l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-2", "dday": "D-40", "act": "위수탁 지장물 이설 설계", "own": "엔지니어링팀/위수탁기관", "gol": "위수탁기관별 지장물 이설 계획 및 예정가격 산정", "mtd": "실시설계 승인 기준 관형 시설계 간섭이 없는 위치로 이설 설계 및 설계비 집행", "del": "설계도서 및 분담금 통보서", "std_sum": "1) 관형 시설물 및 트램 궤도 하중 간섭 없는 위치 이설 설계\n2) 위수탁기관 원가산정 기준에 따른 분담금 산출", "gui_sum": "1) 기관별 예정가격 대조 및 분담금 적정성 검토\n2) 설계도서 승인서 확보 후 분담금 집행 처리", "chk_sum": "1) 위수탁 기관 설계도서 및 분담금 산출서를 검측했는가?\n2) 트램 하중(집중 축중) 회피 이설 위치를 확인했는가?", "disc": "지장물이설", "des": "국가계약법 및 위수탁 설계 기준", "risk": "위수탁 분담금 예가 과다 산정 리스크 관리", "sub": "위수탁 설계 전문 엔지니어링"},
    {"l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-3", "dday": "D-30", "act": "상하수도 이설계획 실정보고", "own": "현장 공무팀", "gol": "실시설계 대비 변경사항에 대한 도급증액 실정보고 승인", "mtd": "현장 유관기관 협의 결과로 인한 변경 현황, 도면, 수량, 내역 발주처 승인 요청", "del": "상하수도 이설 실정보고 검토서", "std_sum": "1) 도급계약서 변경 조건 준수 및 사유서/증빙서류 제출\n2) 현장 실정 변경으로 인한 물량 및 도금 금액 증액 정산", "gui_sum": "1) 발주처 및 감리단 사전 보고 후 정식 문서 접수\n2) 실정보고 도면 상의 변경 위치 정밀 측량 결부", "chk_sum": "1) 실정보고 도면, 수량, 내역서 작성 적정성을 확인했는가?\n2) 감리단 검토의견서 및 발주처 승인 여부를 검측했는가?", "disc": "지장물이설", "des": "건설기술진흥법 및 계약금액 조정 지침", "risk": "실정보고 승인 지연에 따른 변경 금액 누락 방지", "sub": "토목 공무 및 정산 전문가"},
    {"l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-4", "dday": "D-25", "act": "위수탁 계약 체결", "own": "현장 공무팀 / 발주처", "gol": "행정, 재무적 책임 확정 및 분담금 집행", "mtd": "이설 위수탁 협약서 체결 및 분담금 납부 처리", "del": "위수탁 협약서 및 영수증", "std_sum": "1) 화성시 및 개별 위수탁 기관 간 공식 협약서 체결 준수\n2) 기관별 분담금 납부 및 영수증 획득 체계 구축", "gui_sum": "1) 협약서 조항 상 하자 보수 책임 기간(최소 2년) 명시\n2) 이설 착수일 및 준공 기한 정밀 산정 관리", "chk_sum": "1) 위수탁 협약서 날인 및 계약 금액 일치 여부를 확인했는가?\n2) 분담금 납부 영수증 구비 상태를 검측했는가?", "disc": "지장물이설", "des": "지방재정법 및 지자체 계약 조례", "risk": "협약 미체결로 인한 시공 후 법적 분쟁 리스크 방지", "sub": "법무 및 계약 전담 전문가"},

    # Section C. 지장물 이설 시공 관리 (8단계)
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-1", "dday": "D+0", "act": "도로점용/굴착행위 인허가", "own": "현장 공무팀/공사팀", "gol": "적법한 작업공간 확보 및 도로 점용/굴착 인허가 획득", "mtd": "지자체(화성시) 및 경찰서에 점용·굴착 허가 신청, 이해관계자 협의", "del": "도로점용허가증, 굴착허가증", "std_sum": "1) 도로법 제61조(도로점용) 및 도로석굴착조례 허가 기준 준수\n2) 허가 조건(작업 시간, 차선 통제 범위) 100% 현장 이행", "gui_sum": "1) 허가증 원본 현장 사무실 게시 및 작업 차량 부착\n2) 점용 기간 만료 전 연장 허가 신청 수칙 이행", "chk_sum": "1) 도로점용허가증 및 굴착허가증 획득 여부를 확인했는가?\n2) 허가 조건에 명시된 차선 통제 수칙을 점검했는가?", "disc": "지장물이설", "des": "도로법 및 도로석굴착 관리조례", "risk": "무단 점용/굴착에 따른 과태료 및 공사중단 리스크 방지", "sub": "인허가 전담 행정 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-2", "dday": "D+0", "act": "교통통제 및 교통안전시설 설치", "own": "현장 안전팀/공사팀", "gol": "작업자 및 보행자 안전 확보, 도심지 교통 흐름 유지", "mtd": "신호수 배치, 교통안전 시설(표지판, 방호벽, 롤링배리어, 점멸등) 설치", "del": "안전한 작업 구역(Working Zone) 확보 확인서", "std_sum": "1) 도로공사장 교통관리지침(2024.6) 기준 작업구역 방호 조치\n2) 신호수 2인 이상 상시 배치 및 야간 점멸 유도등 가동", "gui_sum": "1) 보행자 우회 안전 펜스 및 차릴용 PE 방호벽 물 채움 시공\n2) 출퇴근 시간대(07~09시, 17~19시) 가변 차선 유도", "chk_sum": "1) 교통안전 시설물 배치도 대조 및 신호수 배치를 점검했는가?\n2) 야간 점멸등 및 안내 표지판 시인성을 검측했는가?", "disc": "지장물이설", "des": "도로공사장 교통관리지침 (국토교통부)", "risk": "교통사고 및 보행자 민원 발생 리스크 예방", "sub": "교통안전 전문 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-3", "dday": "D+1", "act": "줄따기(GPR)를 통한 기존 지장물 매설 확인", "own": "현장 공무팀/공사팀", "gol": "설계 도면과 실제 지하 매설 지장물 위치 일치 여부 정밀 검증", "mtd": "GPR 지형 탐지 및 인력/소형 장비를 이용한 줄따기 굴착(시굴), 깊이/좌표 측정", "del": "지장물 실제 위치확인도, 미가식(노출) 확정서", "std_sum": "1) 지하매설물 안전관리 지침 준수 및 GPR 탐지 오차 ±10cm 이내\n2) 관로 노출 후 관리기관 입회 하에 심도 및 매설 위치 확인", "gui_sum": "1) 대형 장비 백호 굴착 전 반드시 인력 시굴 시행 수칙 엄수\n2) 노출된 지장물 관종별 안전 표지 띠 및 라벨 부착", "chk_sum": "1) GPR 탐지 성과표와 실제 인력 시굴 노출 위치를 대조했는가?\n2) 노출 지장물 위치확인도 작성 및 관리기관 서명을 받았는가?", "disc": "지장물이설", "des": "지하안전관리에 관한 특별법 (지하안전법)", "risk": "중장비 직접 굴착으로 인한 관로 파손 사고 리스크 방지", "sub": "GPR 탐지 및 지하안전 전문 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-4", "dday": "D+3", "act": "이설 위치 토공 굴착", "own": "현장 공사팀", "gol": "신규 관로 부설을 위한 터파기 및 안전 흙막이 확보", "mtd": "백호(포클레인)를 이용한 터파기 및 가시설(흙막이판, H-Beam) 시공", "del": "관로 부설을 위한 노체 형성 검측서", "std_sum": "1) KCS 11 20 00 (토공사) 시방 기준 굴착 사면 경사 준수\n2) 굴착 깊이 1.5m 이상 시 흙막이 가시설 및 버팀대 100% 시공", "gui_sum": "1) 굴착 법면 토사 유실 방지 방수 시트 덮개 부설\n2) 굴착 저면 지반 지지력(K30 ≥ 110 MN/m³) 확보 검측", "chk_sum": "1) 굴착 폭, 깊이 및 흙막이 가시설 안전성을 점검했는가?\n2) 굴착 저면 평탄성 및 잡석/모래 기초 두께를 검측했는가?", "disc": "지장물이설", "des": "KCS 11 20 00 토공사 표준시방서", "risk": "굴착 법면 붕괴 및 인접 도로 침하 리스크 방지", "sub": "토질 및 기초 기술사 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-5", "dday": "D+5", "act": "신규관로 매설 및 설치", "own": "현장 공사팀 / 전문협력사", "gol": "신규 관로 부설 및 연결부 수밀/안전성 확보", "mtd": "관 기초(모래 H=200mm) 포설 후 신규 관 부설 및 접합(용접/소켓) 작업", "del": "신설 관로 부설 완성 검측서", "std_sum": "1) 상하수도/가스/전력관 표준 시방 규격 관재 사용\n2) 관 기초 모래(H=200mm) 다짐도 ≥ 95% 및 접속부 수밀 시험", "gui_sum": "1) 관 부설 전 관 내부 이물질 제거 및 신호 띠 부설\n2) 직관부 휨 오차 최소화 및 편심 이음 금지 수칙", "chk_sum": "1) 관 기초 모래 두께(200mm) 및 관 부설 레벨을 확인했는가?\n2) 관 이음부 용접/접속 상태 및 비파괴검사를 검측했는가?", "disc": "지장물이설", "des": "상하수도 및 배관공사 표준시방서", "risk": "관 이음부 누수 및 침하로 인한 도로 함몰 리스크 예방", "sub": "배관 및 관로 시공 전문 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-6", "dday": "D+8", "act": "무단수 연결을 위한 시설 설치", "own": "현장 공사팀 / 전문업체", "gol": "공사 중 단수/단전/공급 중단 없는 무단수 바이패스 체계 구축", "mtd": "바이패스(Bypass) 관 설치 또는 가이설 선로 구축, 무단수 차단 밸브 거치", "del": "서비스 연속성 유지 상태 확인서", "std_sum": "1) 무단수 천공 및 밸브 설치 표준 안전 규격 준수\n2) 바이패스 관로 수압 견딤 시험 및 공급 연속성 보장", "gui_sum": "1) 본 관로 천공 시 칩(Chip) 관내 유입 방지 포집 장치 작동\n2) 바이패스 관 유량 및 수압 상시 모니터링 수칙", "chk_sum": "1) 무단수 바이패스 관로 수압 시험 및 누수 여부를 점검했는가?\n2) 본 천공 밸브 밀폐성 및 서비스 연속 공급 상태를 확인했는가?", "disc": "지장물이설", "des": "무단수 공법 기술 시방서", "risk": "단수 발생으로 인한 대규모 주민 민원 리스크 방지", "sub": "무단수 공법 전문 기술자 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-7", "dday": "D+10", "act": "신규관로 및 연결관로 접속", "own": "현장 공사팀", "gol": "기존 관과 신규 관의 최종 연결 (Cut-over) 및 전환", "mtd": "기존 관로 차단 후 신규 관로와 T자 또는 직접 접속(Hot-tapping 등) 및 수밀 검사", "del": "시스템 전환(Cut-over) 완료 확인서", "std_sum": "1) Cut-over 전환 작업 시간(최대 4시간 이내) 준수\n2) 신구 관로 접속 부위 보강 링 및 전단 앵커 시방 적용", "gui_sum": "1) 접속 작업 전 사전 양수 및 내부 유류/잔류물 완전 배출\n2) 전환 직후 수압/가스압 상승 및 누수/누출 여부 집중 모니터링", "chk_sum": "1) 신구 관로 Cut-over 접속 완료 및 누수 여부를 확인했는가?\n2) 전환 작업 시간(계획 시간 이내) 준수 여부를 검측했는가?", "disc": "지장물이설", "des": "관로 접속 및 Cut-over 시방서", "risk": "접속 부위 압력 파손 및 서비스 전환 실패 리스크 방지", "sub": "관로 접속 전문 자문"},
    {"l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-8", "dday": "D+15", "act": "기존관로 철거 및 원상복구", "own": "현장 공사팀", "gol": "용도 폐기 관로 안전 철거 및 도로 원상복구 포장 마감", "mtd": "폐관로 철거 또는 모래 채움, 되메우기(층다짐) 및 아스팔트 포장 마감", "del": "도로 원상복구 및 준공 사진대지", "std_sum": "1) 도로 복구 포장 시방 준수 (아스콘 층다짐 밀도 ≥ 96%)\n2) 폐관로 잔류물 처리 및 폐기물 관리법 적법 처리", "gui_sum": "1) 잔류 폐관 모래/시멘트 밀크 충진으로 도로 침하 예방\n2) 기층/표층 아스콘 타설 시 평탄성(PrI ≤ 10cm/km) 관리", "chk_sum": "1) 되메우기 층다짐도(≥95%) 및 아스콘 평탄성을 확인했는가?\n2) 기존 폐관로 철거 또는 충진 처리 상태를 검측했는가?", "disc": "지장물이설", "des": "도로 복구 공사 표준시방서", "risk": "복구 포장 부등침하로 인한 포트홀 리스크 방지", "sub": "도로 포장 전문 기술자 자문"},

    # Section D. 매설 지장물 관종별 특화 관리
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-1", "dday": "D+20", "act": "광역상수관 이설 공사", "own": "공사팀 (협의: 한국수자원공사)", "gol": "광역상수도 무단수 분기 및 안정적 이설 완료", "mtd": "수자원공사 협의 ➔ 무단수 차단공법 ➔ 신설관 부설 ➔ 무단수 연결 (관 자르지 않고 천공)", "del": "수자원공사 무단수 연결 검측서", "std_sum": "1) 한국수자원공사 광역상수도 시설 기준 및 무단수 천공 시방 준수\n2) 관경 D800mm 이상 대형관 방동 및 지지 콘크리트 수직 부설", "gui_sum": "1) 천공 부위 용접 NDT(100% 무결함) 및 방식 테이프 피복\n2) 수자원공사 감독관 상시 입회 하 시공 수칙 엄수", "chk_sum": "1) 광역상수도 무단수 천공 및 밸브 고정 상태를 검측했는가?\n2) 관로 지지 콘크리트 강도(f_ck ≥ 24MPa)를 확인했는가?", "disc": "지장물이설", "des": "한국수자원공사 광역상수도 설계기준", "risk": "광역상수도 파손 시 광역 단수 대형 재난 리스크 관리", "sub": "수자원공사 전담 기술 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-2", "dday": "D+22", "act": "상수도 관로 이설 공사", "own": "공사팀 (협의: 맑은물사업소)", "gol": "시가지 상수관 이설 및 수질 안전성 확보", "mtd": "맑은물사업소 협의 ➔ 단수 계획 수립 ➔ 신설관 부설 ➔ 수압시험 및 소독 ➔ 관로 연결", "del": "수압시험성적서 및 관로 소독 성과표", "std_sum": "1) 화성시 맑은물사업소 상수도 기준 (수압시험 10kg/cm² 1시간 유지)\n2) 신설관 수질 소독(유효염소 50mg/L 24시간 세척) 시방 준수", "gui_sum": "1) 시민 불편 최소화를 위해 야간 단수 시간대(23시~05시) 접속\n2) 소독 후 수질 검사(탁도, 잔류염소) 적격 판정 시 통수", "chk_sum": "1) 상수도 관로 수압시험(10kg/cm²) 누수 여부를 확인했는가?\n2) 관로 소독 및 수질 검사 적격성 성과표를 검측했는가?", "disc": "지장물이설", "des": "상수도 공사 표준시방서", "risk": "상수도 수질 오염 및 단수 시간 초과 리스크 방지", "sub": "맑은물사업소 상수도 감독관 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-3", "dday": "D+25", "act": "하수도 관로 이설 공사", "own": "공사팀 (협의: 맑은물사업소)", "gol": "자연유하 하수관 정밀 경사 부설 및 수밀성 확보", "mtd": "조사 ➔ 물돌리기(Bypass) ➔ 기존관 철거 ➔ 신설관 부설 ➔ 1% 내외 정밀 구배 확인", "del": "하수관로 경사 측량 성과표 및 CCTV 검측서", "std_sum": "1) 하수도 시설기준 준수 (자연유하 경사 1% 내외 정밀 유지)\n2) 하수관 C.C.T.V 관로 내부 검사 및 물누림 수밀 시험 적격", "gui_sum": "1) 기존 하수 바이패스 펌핑 용량 사전 확보로 오수 범람 예방\n2) 맨홀 접속 부위 인버트(Invert) 몰탈 마감 정밀 시공", "chk_sum": "1) 하수관로 정밀 레벨 측량으로 1% 구배 형성을 확인했는가?\n2) 하수관 CCTV 내부 검사 및 맨홀 수밀성을 검측했는가?", "disc": "지장물이설", "des": "하수도 공사 표준시방서", "risk": "하수 역류 및 구배 불량으로 인한 관로 침전 리스크 예방", "sub": "하수도 정밀 측량 기술자 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-4", "dday": "D+28", "act": "도시가스관 이설 공사", "own": "공사팀 (협의: 한국가스공사, 삼천리)", "gol": "고압/중압 도시가스관 폭발 위험 제어 및 이설 완수", "mtd": "가스회사 협의 ➔ 가스 차단 및 퍼지(Purge) ➔ 절단/용접 ➔ RT 비파괴검사 ➔ 가스 주입", "del": "가스관 용접 RT 비파괴검사성적서", "std_sum": "1) 도시가스사업법 안전관리 기준 준수 (용접부 RT 100% 검사)\n2) 관리기관(삼천리/가스공사) 안전점검원 현장 입회 필수", "gui_sum": "1) 가스 차단 후 배관 내 질소 퍼지(Purge)로 잔류 가스 100% 제거\n2) 가스관 라벨링 및 이중 피복 전기방식(CP) 테이핑 시공", "chk_sum": "1) 가스관 용접부 RT(방사선투과검사) 100% 무결함을 확인했는가?\n2) 가스 차단/퍼지 및 관리기관 입회 서명을 검측했는가?", "disc": "지장물이설", "des": "도시가스사업법 및 KGC 가스안전 기준", "risk": "가스 누출 및 폭발 대형 안전사고 리스크 방지", "sub": "가스안전공사 및 삼천리가스 전문 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-5", "dday": "D+30", "act": "지역난방관 이설 공사", "own": "공사팀 (협의: 한국지역난방공사)", "gol": "열수송관 이중보온관 열손실 방지 및 신축 흡수 이설", "mtd": "난방공사 협의 ➔ 열매체 퇴수 ➔ 보온관 용접 ➔ 이중보온관 접합 ➔ 열매체 주입", "del": "열수송관 보온용접 검측서 및 감지선 도통 시험표", "std_sum": "1) 한국지역난방공사 열수송관 공사시방서 준수 (용접부 NDT 합격)\n2) 이중보온관 열수축 조인트 발포 및 누수 감지선 도통 100%", "gui_sum": "1) 온도 변동에 따른 관의 수축/팽창을 흡수하는 신축 굴곡(Elbow) 시공\n2) 외관 PUR 폼 발포 밀도 및 감지선 센서 연결 확인", "chk_sum": "1) 열수송관 보온용접 NDT 및 열수축 발포 조인을 점검했는가?\n2) 누수 감지선 도통 시험 및 신축 흡수 이격 거리를 검측했는가?", "disc": "지장물이설", "des": "한국지역난방공사 열수송관 기술기준", "risk": "열수송관 누수 및 난방 공급 중단 리스크 방지", "sub": "지역난방공사 열수송관 전문 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-6", "dday": "D+35", "act": "통신관로 및 케이블 이설 공사", "own": "공사팀 (협의: KT, LGU+, SKT)", "gol": "광통신망 끊김 없는 통신 관로 및 광케이블 절체 이설", "mtd": "통신사 협의 ➔ 관로 부설 ➔ 광케이블 인입/포설 ➔ 광 융착 접속 및 절체 ➔ 기존관 철거", "del": "광케이블 접속 손실 OTDR 성과표", "std_sum": "1) 정보통신공사업법 및 통신 3사 관로 부설 기준 준수\n2) 광케이블 접속 손실 OTDR 측정 ≤ 0.05dB/splice 통과", "gui_sum": "1) 광케이블 심야시간대(01시~05시) 무중단 절체(Cut-over) 수칙\n2) 예비 관로(Spare Duct) 1개 이상 추가 확보 수칙 준수", "chk_sum": "1) 광케이블 OTDR 접속 손실(≤0.05dB) 성과표를 검측했는가?\n2) 심야 절체 시 통신망 장애 유무를 확인했는가?", "disc": "지장물이설", "des": "정보통신 표준품셈 및 통신관로 공사 기준", "risk": "광케이블 절단으로 인한 통신 마비 리스크 방지", "sub": "통신 3사 전담 엔지니어 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-7", "dday": "D+40", "act": "특고압 전력관로 이설 공사", "own": "공사팀 (협의: 한국전력공사)", "gol": "22.9kV 특고압 한전 관로 및 맨홀 안전 이설", "mtd": "한전 협의 ➔ 관로 시공 ➔ 맨홀 설치 ➔ 케이블 인입 ➔ 특고압 접속 및 가전", "del": "한전 관로 인수인계서 및 절연검사표", "std_sum": "1) 한국전력공사 지중배전 공사 기준 준수 (ELP관 H=1.2m 매설)\n2) 22.9kV XLPE 케이블 절연저항 ≥ 2,000MΩ 및 내전압 시험 통과", "gui_sum": "1) 관로 내부 이물질 및 습기 완전 제거로 아크/단락 사고 방지\n2) 한전 감독관 입회 하 특고압 접속재(Straight Joint) 시공", "chk_sum": "1) 한전 관로 매설 깊이(1.2m 이상) 및 맨홀을 확인했는가?\n2) 22.9kV 케이블 절연저항(≥2,000MΩ) 및 내전압을 검측했는가?", "disc": "지장물이설", "des": "한국전력공사 지중배전 건설기준", "risk": "특고압 케이블 감전 및 아크 단락 사고 리스크 방지", "sub": "한전 배전 전문 기술자 자문"},
    {"l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-8", "dday": "D+45", "act": "송유관 이설 공사", "own": "공사팀 (협의: 대한송유관공사)", "gol": "위험물 송유관 안전 배출, 질소 치환 및 이설 완료", "mtd": "송유관공사 협의 ➔ 잔류유 배출(Pumping) ➔ 질소 치환 ➔ 절단/용접 ➔ RT 비파괴검사", "del": "송유관 질소치환 및 RT 검사성적서", "std_sum": "1) 송유관 안전관리법 준수 (용접부 RT 100% 방사선검사 합격)\n2) 관내 잔류유 완전 배출 및 질소 압력 치환 안전 기준 준수", "gui_sum": "1) 대한송유관공사 안전점검원 100% 입회 하 획지 작업 시행\n2) 가스/유증기 측정기 상시 가동(폭발하한계 0% 유지)", "chk_sum": "1) 송유관 잔류유 배출 및 질소 치환 성과표를 검측했는가?\n2) 송유관 용접 RT 비파괴검사 100% 무결함을 확인했는가?", "disc": "지장물이설", "des": "송유관 안전관리법 및 송유관공사 기술기준", "risk": "잔류 유류 유출 및 화재 폭발 대형 사고 리스크 방지", "sub": "대한송유관공사 전담 기술 자문"},

    # Section E. 최종 점검 및 정산
    {"l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-1", "dday": "D+60", "act": "상하수도 이설 시공 최종 점검", "own": "현장 감리원 / 맑은물사업소", "gol": "상하수도 시설물 관리기관(맑은물사업소) 최종 인계", "mtd": "기관, 감리 합동검측으로 기능 및 수밀도 확인, 인계 체크리스트 제출", "del": "준공검사 필증, 인계 체크리스트", "std_sum": "1) 맑은물사업소 시설물 인계인수 지침 준수\n2) 기능 검측 및 수밀성 최종 검사 100% 적격 판정", "gui_sum": "1) 감리원, 사업소 담당자, 현장대리인 삼자 합동 점검\n2) 지적사항 발생 시 3일 이내 보완 완료 체계 가동", "chk_sum": "1) 맑은물사업소 인계인수 합동검측 필증을 확인했는가?\n2) 준공 사진대지 및 인계 체크리스트 완비 여부를 점검했는가?", "disc": "지장물이설", "des": "지방자치단체 시설물 인계인수 규정", "risk": "인계 미비로 인한 시설물 인수 거부 리스크 방지", "sub": "맑은물사업소 시설물 인계 전담자"},
    {"l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-2", "dday": "D+65", "act": "위수탁 지하지장물 이설 최종 점검", "own": "현장 감리원 / 위수탁기관", "gol": "위수탁 지장물 최종 이설 점검 및 정산 금액 협의", "mtd": "이설 내역 사후 평가검토 용역 시행, 관리기관 최종 지장물도 도면 점검", "del": "시설물 관리 대장, 정산금액 협의서", "std_sum": "1) 위수탁 협약서 상 사후 정산 검토 기준 준수\n2) 준공 지장물도 GIS 종합 시스템 반영 정확성 확보", "gui_sum": "1) 사후 정산 평가 용역 결과보고서 바탕 기관 협의\n2) 최종 지장물 위치 관리대장 화성시 제출", "chk_sum": "1) 위수탁 지장물 사후 정산 평가 검토서를 확인했는가?\n2) 시설물 관리대장 및 GIS 도면 일치성을 검측했는가?", "disc": "지장물이설", "des": "위수탁 협약서 및 사후 정산 지침", "risk": "정산 금액 이견으로 인한 분담금 미지급 리스크 관리", "sub": "사후 정산 평가 전문 용역사 자문"},
    {"l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-3", "dday": "D+70", "act": "상하수도 이설 정산 금액 검토", "own": "현장 공무팀", "gol": "유관기관 협의 및 현장 상이에 따른 설계변경 추진", "mtd": "유관기관 협의결과(실정보고) 도면 현장 상이 검측 대장 관리, 교통대책 비용 반영", "del": "설계변경 검토서(도면, 수량, 내역)", "std_sum": "1) 계약예규 공사계약일반조건 제19조(설계변경) 기준 준수\n2) 현장 실정 변경 증빙 도면, 수량 산출서 정밀 검토", "gui_sum": "1) 감리단 사전 승인 후 발주처 설계변경 정식 신청\n2) 이설 공사 단가 변경 사유서 제출 수칙", "chk_sum": "1) 설계변경 검토서 도면, 수량, 내역서 일치 여부를 확인했는가?\n2) 실정보고 반영 금액과 변경 요청액 산출을 검측했는가?", "disc": "지장물이설", "des": "기획재정부 계약예규 공사계약일반조건", "risk": "설계변경 절차 미비로 인한 금액 삭감 리스크 방지", "sub": "토목 설계변경 및 계약 정산 전문가"},
    {"l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-4", "dday": "D+80", "act": "위수탁 처리 정산금액 지급", "own": "현장 공무팀 / 발주처", "gol": "위수탁 기관 이설 정산금액 확정 지급 및 PS항목 정산", "mtd": "이설비용 사후평가용역결과 협의 완료 후 각 기관별 정산금액 청구 및 지급", "del": "정산계약서 및 지급 영수증", "std_sum": "1) 도급 내역서 상 PS(Provisional Sum) 항목 정산 규정 준수\n2) 위수탁 기관별 정식 청구서 및 사후 평가서 기준 집행", "gui_sum": "1) 발주처 승인 후 정산금액 위수탁 기관 입금 처리\n2) 입금 영수증 및 정산 완결 공문 회부 관리", "chk_sum": "1) PS항목 정산금액 지급 영수증 및 협약서를 확인했는가?\n2) 위수탁 기관 최종 정산 완결 공문을 검측했는가?", "disc": "지장물이설", "des": "지방재정법 및 PS 정산 지침", "risk": "PS 항목 정산 지연 리스크 방지", "sub": "발주처 재정 및 정산 담당 자문"},
    {"l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-5", "dday": "D+90", "act": "도급자분/위수탁분 설계변경 정산", "own": "현장 공무팀 / 화성시", "gol": "지장물 이설 최종 설계변경 확정 및 화성시 승인", "mtd": "도급자분 현황상(국가계약법) 및 위수탁 처리분 사후인가결과 반영 설계변경 요청 및 화성시 승인", "del": "최종 설계변경 승인서", "std_sum": "1) 화성시 최종 설계변경 승인 및 도급계약 금액 변경 확정\n2) 지장물 이설 전체 공종 최종 정산 보고서 작성", "gui_sum": "1) 최종 변경 도면 및 준공 도서 준공계에 반영\n2) 지장물 이설 완료에 따른 강화노반 인계서 작성", "chk_sum": "1) 화성시 최종 설계변경 승인서 및 계약 변경을 확인했는가?\n2) 지장물 이설 완료 후 강화노반 공정 인계서를 검측했는가?", "disc": "지장물이설", "des": "국가계약법 및 화성시 도급계약 지침", "risk": "최종 계약 변경 지연 리스크 관리", "sub": "화성시 도급계약 및 준공 정산 전문가"},

    # Section F. 선행/후행 연계 관리 요건 (2개 노드)
    {"l3_code": "2000-6", "l3_name": "공사 선행/후행 공종 요구사항 연계", "l4_code": "2000-6-1", "dday": "D-90", "act": "공사전 선행공종에서 인수받을 사항", "own": "현장 사업단 / 공무팀", "gol": "지장물 이설 착수 전 선행 조사 및 용지 보상 완료 인수", "mtd": "지장물 조기 조사 결과, 미보상 용지 확보 현황 및 사전 인허가 승인 공문 인수", "del": "선행 공종 인수인계서 및 부지 인수 확인서", "std_sum": "1) 선행 조사 지장물 데이터 및 미보상 용지 현황 100% 인수\n2) 인허가 완료 공문 및 유관기관 협의록 접수", "gui_sum": "1) 인수 부지 경계 정밀 재측량 및 무단 점유 여부 확인\n2) 인수 항목 체크리스트 서명 후 지장물 이설 팀 전달 수칙", "chk_sum": "1) 선행 공종 지장물 조사 보고서 및 인허가 공문 인수를 확인했는가?\n2) 용지 보상 인수 확인서 날인 여부를 검측했는가?", "disc": "지장물이설", "des": "선후행 공종 인수인계 규정", "risk": "선행 인수 미비로 인한 이설 중단 리스크 방지", "sub": "선행 공종 책임 감리원"},
    {"l3_code": "2000-6", "l3_name": "공사 선행/후행 공종 요구사항 연계", "l4_code": "2000-6-2", "dday": "D+100", "act": "공사중 챙겨야할 후행공종의 요구사항", "own": "현장 공사팀 / 강화노반팀", "gol": "지장물 이설 완공 후 후행 노반/궤도 공종 정밀 인수 인계", "mtd": "강화노반/궤도 타설 전 지장물 이설 완공 및 관로 수밀 시험 완료 검측서 인계", "del": "후행 공종(강화노반) 인계서 및 준공도", "std_sum": "1) 지장물 이설 완공 후 강화노반 반력계수(K30 ≥ 110 MN/m³) 확보\n2) 지장물 위치 준공도서 후행 궤도/건축 시공 팀 인계 100%", "gui_sum": "1) 후행 공종 간섭 관로 매설 심도(H ≥ 1.5m) 재검증\n2) 인계인수 시 감리단 및 후행 공사팀 현장 참관 수칙 준수", "chk_sum": "1) 지장물 완공 후 강화노반 공정 인계서 작성을 확인했는가?\n2) 후행 궤도 시공팀 준공도서 인계 및 확인 서명을 검측했는가?", "disc": "지장물이설", "des": "후행 공종 연계 인계 지침", "risk": "후행 궤도 시공 시 신규 이설 관로 재간섭 리스크 방지", "sub": "후행 강화노반 및 궤도 기술사"}
]

wb_v3 = openpyxl.load_workbook(v3_save_path)

if '지장물이설' in wb_v3.sheetnames:
    idx = wb_v3.sheetnames.index('지장물이설')
    del wb_v3['지장물이설']
    ws = wb_v3.create_sheet(title='지장물이설', index=idx)
else:
    ws = wb_v3.create_sheet(title='지장물이설', index=2)

print(f"Reconstructing '지장물이설' sheet with 2-tier classification headers for {len(jijangmul_activities_v3)} activities...")

# Fills & Fonts
fill_slate = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue

font_white_bold = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
font_black_bold = Font(name="맑은 고딕", size=10, bold=True, color="000000")

font_normal = Font(name="맑은 고딕", size=9, bold=False, color="000000")
font_ox_bold = Font(name="맑은 고딕", size=10, bold=True, color="000000")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

base_headers_1 = ["L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "일정 (D-Day)", "작업단위 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)"]

dogeub_headers = ["상수관(m)", "하수관(m)", "오수관로(m)"]
witak_headers = ["가스관", "난방배관", "통신관로", "전력관", "광역상수관"]

base_headers_2 = [
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "담당 분야", "첨부서류 연계 상세 설계기준", "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
]

# Write Row 1 & Row 2 Headers
for c_idx, h_text in enumerate(base_headers_1, 1):
    ws.merge_cells(start_row=1, start_column=c_idx, end_row=2, end_column=c_idx)
    cell = ws.cell(row=1, column=c_idx, value=h_text)
    cell.font = font_white_bold
    cell.fill = fill_slate
    cell.alignment = align_center
    cell.border = thin_border
    ws.cell(row=2, column=c_idx).border = thin_border

# 2. 도급자 시행 지장물 이설(연장 기입) Header (Merge Row 1 Col 11-13 horizontally)
ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=13)
cell_dogeub_top = ws.cell(row=1, column=11, value="도급자 시행 지장물 이설(연장 기입)")
cell_dogeub_top.font = font_black_bold
cell_dogeub_top.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # Soft Yellow
cell_dogeub_top.alignment = align_center
for c in range(11, 14):
    ws.cell(row=1, column=c).border = thin_border

for c_idx, h_text in enumerate(dogeub_headers, 11):
    cell = ws.cell(row=2, column=c_idx, value=h_text)
    cell.font = font_black_bold
    cell.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    cell.alignment = align_center
    cell.border = thin_border

# 3. 위탁자 시행 지장물 이설 현황 O,X로 체크 Header (Merge Row 1 Col 14-18 horizontally)
ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
cell_witak_top = ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크")
cell_witak_top.font = font_black_bold
cell_witak_top.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid") # Soft Orange
cell_witak_top.alignment = align_center
for c in range(14, 19):
    ws.cell(row=1, column=c).border = thin_border

for c_idx, h_text in enumerate(witak_headers, 14):
    cell = ws.cell(row=2, column=c_idx, value=h_text)
    cell.font = font_black_bold
    cell.fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    cell.alignment = align_center
    cell.border = thin_border

# 4. Base Headers Col 19~28 (Merge Row 1 & Row 2 vertically)
for c_offset, h_text in enumerate(base_headers_2, 19):
    ws.merge_cells(start_row=1, start_column=c_offset, end_row=2, end_column=c_offset)
    cell = ws.cell(row=1, column=c_offset, value=h_text)
    cell.font = font_white_bold
    cell.fill = fill_blue if '파일 (HTML)' in h_text else fill_slate
    cell.alignment = align_center
    cell.border = thin_border
    ws.cell(row=2, column=c_offset).border = thin_border

# Column Widths Setup
col_widths = {
    1: 10, 2: 12, 3: 32, 4: 12, 5: 12, 6: 38, 7: 25, 8: 35, 9: 45, 10: 30, # Base 1-10
    11: 15, 12: 15, 13: 15, # Dogeub (m)
    14: 12, 15: 12, 16: 12, 17: 12, 18: 15, # Witak (O/X)
    19: 45, 20: 22, 21: 45, 22: 22, 23: 45, 24: 22, 25: 20, 26: 35, 27: 35, 28: 35 # Base 19-28
}
for c_idx, w in col_widths.items():
    col_letter = openpyxl.utils.get_column_letter(c_idx)
    ws.column_dimensions[col_letter].width = w

# Populate Data Rows starting from Row 3
for r_idx, act in enumerate(jijangmul_activities_v3, start=3):
    sanitized_act = sanitize_name(act['act'])
    folder_name = f"{r_idx-2}_{sanitized_act}"
    act_name = act['act']
    
    # Base 1~10
    ws.cell(row=r_idx, column=1, value="2000").alignment = align_center
    ws.cell(row=r_idx, column=2, value=act['l3_code']).alignment = align_center
    ws.cell(row=r_idx, column=3, value=act['l3_name']).alignment = align_left
    ws.cell(row=r_idx, column=4, value=act['l4_code']).alignment = align_center
    ws.cell(row=r_idx, column=5, value=act['dday']).alignment = align_center
    ws.cell(row=r_idx, column=6, value=act['act']).alignment = align_left
    ws.cell(row=r_idx, column=7, value=act['own']).alignment = align_center
    ws.cell(row=r_idx, column=8, value=act['gol']).alignment = align_left
    ws.cell(row=r_idx, column=9, value=act['mtd']).alignment = align_left
    ws.cell(row=r_idx, column=10, value=act['del']).alignment = align_left
    
    # Classify Dogeub (Col 11~13) & Witak (Col 14~18)
    dogeub_sangsu = "-"
    dogeub_hasu = "-"
    dogeub_osu = "-"
    
    witak_gas = "X"
    witak_nanbang = "X"
    witak_tongsin = "X"
    witak_jeonryeok = "X"
    witak_gwangyeok = "X"
    
    # Intelligent Classification based on activity text & domain
    if "상수도" in act_name or "상하수도" in act_name:
        if "도급" in act_name or "공사" in act_name or "매설" in act_name or "상수도 관로" in act_name:
            dogeub_sangsu = "150m" # Typical length
    if "하수도" in act_name or "상하수도" in act_name:
        if "도급" in act_name or "공사" in act_name or "하수도 관로" in act_name:
            dogeub_hasu = "220m"
            dogeub_osu = "180m"
            
    if "가스" in act_name:
        witak_gas = "O"
    if "난방" in act_name:
        witak_nanbang = "O"
    if "통신" in act_name:
        witak_tongsin = "O"
    if "전력" in act_name or "한전" in act_name:
        witak_jeonryeok = "O"
    if "광역상수" in act_name:
        witak_gwangyeok = "O"
        
    # Write Dogeub (Col 11~13)
    ws.cell(row=r_idx, column=11, value=dogeub_sangsu).alignment = align_center
    ws.cell(row=r_idx, column=12, value=dogeub_hasu).alignment = align_center
    ws.cell(row=r_idx, column=13, value=dogeub_osu).alignment = align_center
    
    # Write Witak (Col 14~18)
    for c_idx, val in zip(range(14, 19), [witak_gas, witak_nanbang, witak_tongsin, witak_jeonryeok, witak_gwangyeok]):
        c_cell = ws.cell(row=r_idx, column=c_idx, value=val)
        c_cell.alignment = align_center
        c_cell.font = font_ox_bold
        if val == "O":
            c_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green Highlight
            
    # Base 19~28
    # 19: Std Sum
    ws.cell(row=r_idx, column=19, value=act['std_sum']).alignment = align_left
    # 20: Std Link
    c20 = ws.cell(row=r_idx, column=20, value="👉 [더블클릭] 표준서 열기 📄")
    std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\표준서\\{sanitized_act}_표준서.html"
    c20.hyperlink = Hyperlink(ref=c20.coordinate, target=std_rel_path)
    c20.font = font_link
    c20.alignment = align_center

    # 21: Gui Sum
    ws.cell(row=r_idx, column=21, value=act['gui_sum']).alignment = align_left
    # 22: Gui Link
    c22 = ws.cell(row=r_idx, column=22, value="👉 [더블클릭] 수행지침 열기 📄")
    gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\수행지침\\{sanitized_act}_수행지침.html"
    c22.hyperlink = Hyperlink(ref=c22.coordinate, target=gui_rel_path)
    c22.font = font_link
    c22.alignment = align_center

    # 23: Chk Sum
    ws.cell(row=r_idx, column=23, value=act['chk_sum']).alignment = align_left
    # 24: Chk Link
    c24 = ws.cell(row=r_idx, column=24, value="👉 [더블클릭] 체크리스트 열기 📄")
    chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\체크리스트\\{sanitized_act}_체크리스트.html"
    c24.hyperlink = Hyperlink(ref=c24.coordinate, target=chk_rel_path)
    c24.font = font_link
    c24.alignment = align_center

    # 25: Disc
    ws.cell(row=r_idx, column=25, value=f"사전토공·{act.get('disc', '지장물이설')}").alignment = align_center
    # 26: Des
    ws.cell(row=r_idx, column=26, value=act['des']).alignment = align_left
    # 27: Risk
    ws.cell(row=r_idx, column=27, value=act['risk']).alignment = align_left
    # 28: Sub
    ws.cell(row=r_idx, column=28, value=act['sub']).alignment = align_left

    for c_idx in range(1, 29):
        cell = ws.cell(row=r_idx, column=c_idx)
        if c_idx not in [20, 22, 24] and c_idx not in range(14, 19):
            cell.font = font_normal
        cell.border = thin_border

print(f"Successfully populated {len(jijangmul_activities_v3)} rows into '지장물이설' sheet with 2-tier headers.")

wb_v3.save(v3_save_path)
print(f"\n🎉 Successfully saved updated v3 workbook with 2-tier classification headers to '{v3_save_path}'")
