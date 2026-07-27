import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

print(f"Loading '{target_file}'...")
wb = openpyxl.load_workbook(target_file)
ws = wb['지장물이설']

# Find target row
target_row = None
for r in range(3, ws.max_row + 1):
    act_val = ws.cell(row=r, column=6).value
    if act_val and "지장물 이설 계획 수립" in str(act_val):
        target_row = r
        break

if target_row:
    print(f"Found target activity at Row {target_row}")
    
    std_summary = "1) 트램 궤도 구조물과 지장물 수평/수직 최소 이격거리(1.5m 이상) 확보\n2) 관종별 이설 우선순위(상하수 ➔ 가스 ➔ 전력/통신) 확정 준수"
    gui_summary = "1) 인허가 및 교통처리 대책과 연계된 릴레이 이설 일정 수립\n2) 이설계획서 감리단 및 발주기관 정식 승인 수칙"
    chk_summary = "1) 궤도 최소 이격거리(1.5m 이상) 및 관종별 릴레이 이설 순서를 검측했는가?\n2) 지장물 이설 종합계획서 감리단 및 발주처 공식 승인을 확인했는가?"

    ws.cell(row=target_row, column=19, value=std_summary) # Std sum
    ws.cell(row=target_row, column=21, value=gui_summary) # Gui sum
    ws.cell(row=target_row, column=23, value=chk_summary) # Chk sum

    wb.save(target_file)
    print(f"🎉 Successfully saved updated workbook to '{target_file}'")
else:
    print("❌ Target row not found!")
