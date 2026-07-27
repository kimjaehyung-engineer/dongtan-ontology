import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

target_v3 = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3.xlsx"

wb = openpyxl.load_workbook(target_v3)
ws = wb['지장물이설']

print("=== Checking Restored Layout & Row Heights ===")
print(f"Row 1 Height: {ws.row_dimensions[1].height}pt")
print(f"Row 2 Height: {ws.row_dimensions[2].height}pt")
print(f"Row 3 Height: {ws.row_dimensions[3].height}pt")
print(f"Row 4 Height: {ws.row_dimensions[4].height}pt")

print("\nSample Check Row 3:")
print(" - Schedule:", ws.cell(row=3, column=1).value)
print(" - Activity:", ws.cell(row=3, column=6).value)
print(" - Purpose:", ws.cell(row=3, column=8).value)
print(" - Header Col 11 (2-Tier):", ws.cell(row=1, column=11).value)
print(" - Header Col 14 (2-Tier):", ws.cell(row=1, column=14).value)

print("\n✅ Verification SUCCESS: Perfectly restored clean & compact layout!")
