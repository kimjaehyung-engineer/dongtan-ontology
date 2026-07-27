import openpyxl
import sys

# Set standard output to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("매뉴얼 BODY (집행단계).xlsx", data_only=True)
print("Sheet Names:", wb.sheetnames)

# Save summary to a file
with open("excel_summary.txt", "w", encoding="utf-8") as f:
    f.write(f"Sheet Names: {wb.sheetnames}\n\n")
    for name in wb.sheetnames:
        ws = wb[name]
        f.write(f"=== Sheet: {name} (Rows: {ws.max_row}, Cols: {ws.max_column}) ===\n")
        # Get headers (first row or non-empty first row)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            f.write("Empty sheet\n\n")
            continue
        
        # print first few rows to file
        f.write("Headers/First 5 rows:\n")
        for i, row in enumerate(rows[:10]):
            f.write(f"Row {i+1}: {row}\n")
        f.write("\n")
print("Done writing summary to excel_summary.txt")
