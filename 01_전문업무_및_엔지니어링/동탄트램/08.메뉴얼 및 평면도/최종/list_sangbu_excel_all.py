import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\상부강화노반"

wb = openpyxl.load_workbook(excel_path)
sheet = wb['상부강화노반']

print("=== 엑셀 '상부강화노반' 36개 전체 액티비티 목록 ===")
excel_acts = []
for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if not act_name:
        act_name = sheet.cell(row=r, column=5).value
    if l4_code and act_name:
        excel_acts.append((r-1, l4_code, act_name))
        print(f"{r-1:02d}. Row {r:02d} | L4 Code: {l4_code} | Name: {act_name}")

print(f"Total Master Activities: {len(excel_acts)}")
