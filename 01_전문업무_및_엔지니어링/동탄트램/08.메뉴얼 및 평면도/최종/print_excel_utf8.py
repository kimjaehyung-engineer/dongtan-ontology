import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

print("=== 엑셀 '지장물이설' 시트 38개 액티비티 표준 순서 ===")
for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1) if sheet.cell(row=r, column=c).value is not None]
    if l4_code:
        print(f"Row {r-1:02d} | L4 Code: {l4_code} | Values: {row_vals}")
