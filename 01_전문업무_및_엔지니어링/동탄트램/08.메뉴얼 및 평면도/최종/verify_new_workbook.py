import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)_최종공정매뉴얼완성본.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb['공정매뉴얼']

header_row = 3

print(f"=== Verification of '공정매뉴얼' Sheet in New Workbook ===")
print("Total rows:", ws.max_row, "| Total cols:", ws.max_column)

print("\nHeaders (Row 3):")
for c in range(1, ws.max_column + 1):
    print(f"  Col {c:2d}: {ws.cell(row=header_row, column=c).value}")

print("\nSample Row 4 (1st Data Row):")
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=4, column=c).value
    has_link = ws.cell(row=4, column=c).hyperlink is not None
    val_snippet = str(val).replace('\n', ' ')[:50] if val else "None"
    print(f"  Col {c:2d} [{ws.cell(row=header_row, column=c).value}]: HasLink={has_link} | Val='{val_snippet}'")

print("\nSample Row 25 (Middle Data Row):")
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=25, column=c).value
    has_link = ws.cell(row=25, column=c).hyperlink is not None
    val_snippet = str(val).replace('\n', ' ')[:50] if val else "None"
    print(f"  Col {c:2d} [{ws.cell(row=header_row, column=c).value}]: HasLink={has_link} | Val='{val_snippet}'")
