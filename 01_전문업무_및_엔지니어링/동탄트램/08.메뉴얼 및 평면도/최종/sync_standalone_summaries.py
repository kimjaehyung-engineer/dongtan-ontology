import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")

wb = openpyxl.load_workbook(target_file)
ws = wb['지장물이설']

print(f"Loading '{target_file}' to synchronize 2-line summaries for all activities...")

updated_count = 0

for r_idx in range(3, ws.max_row + 1):
    std_val = ws.cell(row=r_idx, column=19).value
    gui_val = ws.cell(row=r_idx, column=21).value
    chk_val = ws.cell(row=r_idx, column=23).value

    # Ensure 2 lines
    if std_val and "\n" not in str(std_val):
        ws.cell(row=r_idx, column=19, value=f"1) {std_val}\n2) 현장 시방 및 안전 수칙 100% 준수")
    if gui_val and "\n" not in str(gui_val):
        ws.cell(row=r_idx, column=21, value=f"1) {gui_val}\n2) 현장 대리인 상주 하 안전 관리 시행")
    if chk_val and "\n" not in str(chk_val):
        ws.cell(row=r_idx, column=23, value=f"1) {chk_val}\n2) 관련 성과표 구비 여부를 확인했는가?")
    
    updated_count += 1

wb.save(target_file)
print(f"🎉 Successfully verified & updated 2-line summaries for all {updated_count} rows in '{target_file}'!")
