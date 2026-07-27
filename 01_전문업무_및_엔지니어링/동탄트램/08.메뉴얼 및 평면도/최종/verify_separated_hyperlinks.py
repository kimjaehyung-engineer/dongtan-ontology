import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    header_row = 3 if sheet_name == '공정표에따른 매뉴얼' else 1
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").replace('\n', ' ') for c in range(1, ws.max_column + 1)]
    print(f"\n=== Sheet '{sheet_name}' (Total Cols: {len(headers)}) ===")
    for c_idx, h in enumerate(headers, 1):
        if any(k in h for k in ['표준서', '수행지침', '체크리스트']):
            r2_cell = ws.cell(row=header_row + 1, column=c_idx)
            has_link = r2_cell.hyperlink is not None
            val_snippet = str(r2_cell.value)[:40].replace('\n', ' ') if r2_cell.value else "Empty"
            print(f"  Col {c_idx} [{h}]: HasHyperlink={has_link} | SampleVal='{val_snippet}'")
