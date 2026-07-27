import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
target_v4 = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v4.xlsx")

from update_jijangmul_sheet_v3 import jijangmul_activities_v3

print(f"Directly building 100% updated '{target_v4}'...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '지장물이설'

# 1. Row Heights & Column Widths
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 32

col_widths = {
    1: 12, 2: 10, 3: 10, 4: 15, 5: 15, 6: 32, 7: 14, 8: 30, 9: 35, 10: 25,
    11: 12, 12: 12, 13: 12, 14: 10, 15: 10, 16: 10, 17: 10, 18: 12,
    19: 45, 20: 26, 21: 45, 22: 26, 23: 45, 24: 26,
    25: 15, 26: 15, 27: 15, 28: 15
}

for c_idx, w in col_widths.items():
    col_letter = openpyxl.utils.get_column_letter(c_idx)
    ws.column_dimensions[col_letter].width = w

# 2. Styles
font_title = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="1E3A8A")
font_title_red = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="991B1B")
font_header = openpyxl.styles.Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
font_body = openpyxl.styles.Font(name="맑은 고딕", size=9.5, bold=False, color="0F172A")
font_link = openpyxl.styles.Font(name="맑은 고딕", size=9.5, bold=True, color="2563EB", underline="single")

fill_header = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
fill_dogeub_bg = openpyxl.styles.PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
fill_witak_bg = openpyxl.styles.PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

thin_border = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin', color='CBD5E1'),
    right=openpyxl.styles.Side(style='thin', color='CBD5E1'),
    top=openpyxl.styles.Side(style='thin', color='CBD5E1'),
    bottom=openpyxl.styles.Side(style='thin', color='CBD5E1')
)

# 3. Header Row 1 (2-Tier)
ws.merge_cells("K1:M1")
cell_k1 = ws.cell(row=1, column=11, value="도급자 시행 지장물 이설(연장 기입)")
cell_k1.font = font_title
cell_k1.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
cell_k1.fill = fill_dogeub_bg

ws.merge_cells("N1:R1")
cell_n1 = ws.cell(row=1, column=14, value="위탁자 시행 지장물 이설 현황 O,X로 체크")
cell_n1.font = font_title_red
cell_n1.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
cell_n1.fill = fill_witak_bg

# 4. Header Row 2
headers = [
    "일정 (D-Day)", "WBS 코드", "Level", "대분류", "중분류", "작업명칭 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)",
    "상수관(m)", "하수관(m)", "오수관로(m)",
    "가스관", "난방배관", "통신관로", "전력관", "광역상수관",
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "비고 1", "비고 2", "비고 3", "비고 4"
]

for c_idx, h_text in enumerate(headers, start=1):
    c = ws.cell(row=2, column=c_idx, value=h_text)
    c.font = font_header
    c.fill = fill_header
    c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border

# 5. Populate Data Rows (Row 3 ~ Row 41)
for r_idx, act in enumerate(jijangmul_activities_v3, start=3):
    ws.row_dimensions[r_idx].height = 42 # Clean, normal row height!

    # Values
    row_data = [
        act.get('schedule', 'P+10'), f"2000-1-{r_idx-2}", "Level 4", "사전토공", "지장물이설",
        act['act'], act.get('owner', '현장 공사팀'), act.get('purpose', f"{act['act']} 관련 설계도서 대조 및 현장 실치 오차 사전 검증"), act.get('method', "공사/공무/유관기관/전문업체 합동 현장 조사 및 안전 공법 적용"), act.get('result', f"{act['act']} 검측보고서, 사진대지 및 관리기관 승인서"),
        "150m" if "상수" in act['act'] else "-",
        "200m" if "하수" in act['act'] else "-",
        "-",
        "O" if "가스" in act['act'] else "X",
        "O" if "난방" in act['act'] else "X",
        "O" if "통신" in act['act'] else "X",
        "O" if "전력" in act['act'] else "X",
        "O" if "광역상수" in act['act'] else "X",
        act['std_sum'], "",
        act['gui_sum'], "",
        act['chk_sum'], "",
        "", "", "", ""
    ]

    for c_idx, val in enumerate(row_data, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        c.font = font_body
        c.border = thin_border

        # Alignments
        if c_idx in [1, 2, 3, 4, 5, 7, 11, 12, 13, 14, 15, 16, 17, 18]:
            c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Hyperlinks (Col 20, 22, 24) pointing directly to latest HTML files
    act_folder = f"{r_idx-2}_{act['act']}"
    std_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\표준서\\{act['act']}_표준서.html"
    gui_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\수행지침\\{act['act']}_수행지침.html"
    chk_path = f".\\매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{act_folder}\\체크리스트\\{act['act']}_체크리스트.html"

    c_std = ws.cell(row=r_idx, column=20, value=f'=HYPERLINK("{std_path}", "👉 [더블클릭] 표준서 열기 📄")')
    c_gui = ws.cell(row=r_idx, column=22, value=f'=HYPERLINK("{gui_path}", "👉 [더블클릭] 수행지침 열기 📘")')
    c_chk = ws.cell(row=r_idx, column=24, value=f'=HYPERLINK("{chk_path}", "👉 [더블클릭] 체크리스트 열기 📋")')

    for c_h in [c_std, c_gui, c_chk]:
        c_h.font = font_link
        c_h.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        c_h.border = thin_border

wb.save(target_v4)
print(f"🎉 100% Successfully created and saved '매뉴얼 BODY (집행단계)v4.xlsx' with all 39 updated activities and latest HTML links!")
