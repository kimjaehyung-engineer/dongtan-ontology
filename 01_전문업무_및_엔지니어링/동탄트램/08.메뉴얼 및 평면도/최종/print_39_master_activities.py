import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb['지장물이설']

print("=== 엑셀 '지장물이설' 시트 39개 액티비티 마스터 목록 ===")
master_list = []
for r in range(2, sheet.max_row + 1):
    l4_code = sheet.cell(row=r, column=4).value
    act_name = sheet.cell(row=r, column=6).value
    if l4_code and act_name:
        idx = r - 1
        master_list.append((idx, l4_code, act_name))
        print(f"{idx:02d}. L4 Code: {l4_code} | Activity Name: {act_name}")

print(f"Total Master Activities: {len(master_list)}")
