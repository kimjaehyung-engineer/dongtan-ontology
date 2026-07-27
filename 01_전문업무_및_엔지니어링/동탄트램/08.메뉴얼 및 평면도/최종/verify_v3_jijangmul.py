import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

v3_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3.xlsx"
wb = openpyxl.load_workbook(v3_path)

print("=== Verification of v3 Workbook '지장물이설' Sheet ===")
print("Sheet Names:", wb.sheetnames)

ws = wb['지장물이설']
print("Max Rows:", ws.max_row, "| Max Cols:", ws.max_column)

print("\nHeader Row (Row 1):")
for c in range(1, ws.max_column + 1):
    print(f"  Col {c:2d}: '{ws.cell(row=1, column=c).value}'")

print("\nSampling Rows (Row 2, Row 15, Row 25):")
for r in [2, 15, 25]:
    l4 = ws.cell(row=r, column=4).value
    act = ws.cell(row=r, column=6).value
    own = ws.cell(row=r, column=7).value
    std_sum = str(ws.cell(row=r, column=11).value).replace('\n', ' ')
    link12 = ws.cell(row=r, column=12).hyperlink is not None
    link14 = ws.cell(row=r, column=14).hyperlink is not None
    link16 = ws.cell(row=r, column=16).hyperlink is not None
    
    print(f"\n[Row {r:2d}] L4='{l4}' | Act='{act}' | Owner='{own}'")
    print(f"  - 표준서 요약: {std_sum[:70]}...")
    print(f"  - Links OK: Std={link12}, Gui={link14}, Chk={link16}")
