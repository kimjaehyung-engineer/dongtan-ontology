import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

def sanitize_name(name):
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', str(name)).strip()

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v2_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v2.xlsx")
if not os.path.exists(v2_path):
    v2_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계).xlsx")

v3_save_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")
base_attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)")
jijangmul_attach_dir = os.path.join(base_attach_dir, "지장물이설")

os.makedirs(jijangmul_attach_dir, exist_ok=True)

# 29 Detailed Jijangmul Activities extracted from images
jijangmul_activities = [
    # Section A: 사전준비 및 Risk 검토
    {
        "l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-1", "dday": "P-30",
        "act": "Site Survey Risk 검토", "own": "토목국내견적팀",
        "gol": "현장 현황 및 여건 중 시공에 영향을 줄 수 있는 Risk 사전 검토",
        "mtd": "설계도서 불일치 구간 실행예산 변경 여부 검토 및 설계변경 불가시 Risk 최소화 방안 검토",
        "del": "Site Survey 검토서",
        "std_sum": "1) 설계도서와 현장 불일치 구간 100% 사전 조사 및 실행예산 변경 검토\n2) 현장/견적/설계 합동 Risk 최소화 방안 수립 준수",
        "gui_sum": "1) 우수 토공업체 Pool 검토를 통한 현장 투입 가능성 검증\n2) 현장 여건 대조 및 추가 공사비 발생 가능성 사전 차단",
        "chk_sum": "1) Site Survey 검토서 작성 및 관련 부서 서명 승인을 완료했는가?\n2) 설계 불일치 구간에 대한 실행예산 반영 적정성을 검측했는가?",
        "disc": "지장물이설", "des": "토목국내견적 지침 및 계약 시방", "risk": "설계도서 불일치로 인한 실행예산 초과 리스크 관리", "sub": "우수 토공업체 Pool 기술 자문"
    },
    {
        "l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-2", "dday": "P+0",
        "act": "발주전적 KOM (도급지분)", "own": "현장소장",
        "gol": "도급자 지장물 이설 업체 선정 및 위수탁기관 협의 검토",
        "mtd": "착공전/발주전 구조, 공법, 조건, 장비 Requirement 검토 및 공기 검토 (CML 토목사무지침)",
        "del": "KOM 검토보고서 (변동사항 지침 특기조건 포함)",
        "std_sum": "1) CML 토목사무지침 준수 및 도급 지장물 이설 전문업체 적격 검토\n2) 발주 전 구조안전성 및 공기 타당성 100% 확보",
        "gui_sum": "1) 도급자 이설분 이설업체 계약 시 특기시방 조건 명시\n2) 장비 사양 및 현장 진입 여건 사전 확정 수칙",
        "chk_sum": "1) KOM 검토보고서 및 변동사항 특기조건 반영 여부를 확인했는가?\n2) 이설 전문업체 계약 전 시공 능력 및 장비 사양을 검측했는가?",
        "disc": "지장물이설", "des": "CML 토목사무지침 및 도급 계약 시방", "risk": "부적격 업체 선정에 따른 공기 지연 리스크 방지", "sub": "외주관리팀 전문 자문"
    },
    {
        "l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-3", "dday": "P+0",
        "act": "지장물 이설 요청 (위수탁고)", "own": "공무팀/사업팀",
        "gol": "위수탁기관 지장물 이설 착수 공문 발송",
        "mtd": "양질의 위수탁기관 협의를 통한 위수탁 지장물 이설 공식 요청",
        "del": "지장물 이설요청 공문",
        "std_sum": "1) 위수탁기관별 관련 법령에 따른 이설 착수 요청 공문 정식 발송\n2) 협의 지연 예방을 위한 기관별 전담 창구 일원화",
        "gui_sum": "1) 이설 대상 관로 위치도 및 개략 수량 첨부 문서화\n2) 위수탁기관 회신 기한(14일 이내) 명시 관리",
        "chk_sum": "1) 위수탁기관 공식 이설 요청 공문 발송 및 접수를 확인했는가?\n2) 첨부 위치도 및 이설 범위 정확성을 검측했는가?",
        "disc": "지장물이설", "des": "지자체 위수탁 협약 처리 기준", "risk": "위수탁 기관 이설 착수 지연 리스크 관리", "sub": "관련 유관기관 전담 자문"
    },
    {
        "l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-4", "dday": "P+10",
        "act": "최고의 팀 만들기 지원", "own": "외주관리팀",
        "gol": "투입가능한 최고의 자재생산/제작/장비/협력업체/작업반 선정",
        "mtd": "관리 Pool 중 투입가능 협력업체 및 작업반 기술력/실적 검토",
        "del": "우수 업체/작업반 입찰참여 승인서",
        "std_sum": "1) 외주 관리 Pool 시공 실적 및 현장 투입 가능성 다각 검토\n2) 기술력 및 안전 관리 평가 우수 작업반 선정",
        "gui_sum": "1) 지장물 이설 특화 장비 및 전문 인력 확보 확인\n2) 하도급 적격 심사 기준 85점 이상 달성 수칙",
        "chk_sum": "1) 협력업체 기술 평가서 및 입찰참여 승인서를 확인했는가?\n2) 작업반 전담 반장의 트램 지장물 시공 실적을 검측했는가?",
        "disc": "지장물이설", "des": "외주 하도급 관리 규정", "risk": "시공 미숙 작업반 투입으로 인한 사고 리스크 방지", "sub": "외주계약 전문 자문"
    },
    {
        "l3_code": "2000-1", "l3_name": "공통 관리 Key Point & 사전 준비", "l4_code": "2000-1-5", "dday": "D-60",
        "act": "착수전 Big Room 회의", "own": "현장소장",
        "gol": "도급자 지장물 이설 계획 수립단계에서 원가절감을 위한 이설계획 검토",
        "mtd": "유관기관(맑은물, 난방, 한전, 통신, 가스)과 이설 순서 협의, 타 시설물 간섭 및 노반/궤도 착수 전 공정 검토",
        "del": "Big Room 회의록 (원가절감 반영)",
        "std_sum": "1) 맑은물사업소, 한전, 가스공사 등 유관기관 합동 이설 순서 조율\n2) 굴착 간섭 제어 및 강화노반/궤도 공정 연계 원가절감안 반영",
        "gui_sum": "1) 도급자 이설분과 위수탁분 시공 간섭 3D 검토 시행\n2) 회의 결정사항 주간 단위 이행 점검 체계 가동",
        "chk_sum": "1) 유관기관 합동 Big Room 회의록 작성 및 서명을 완료했는가?\n2) 지장물 이설 순서와 노반/궤도 공정 연동성을 확인했는가?",
        "disc": "지장물이설", "des": "동탄트램 사업관리 지침", "risk": "유관기관 간 공정 중복으로 인한 중복 굴착 리스크 방지", "sub": "유관기관 관리자 및 분야별 전문가"
    },

    # Section B. 행정 및 설계 승인
    {
        "l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-1", "dday": "D-40",
        "act": "관리기관(맑은물사업소) 협의", "own": "현장 공무팀",
        "gol": "위탁기관 협의 결과에 따른 상하수도 이설 계획 맑은물사업소 협의 진행",
        "mtd": "실시설계 유관기관 사전협의 결과 반영된 상하수도 이설계획 검토 및 재협의",
        "del": "상하수도 이설 도면, 수량, 내역",
        "std_sum": "1) 화성시 맑은물사업소 기술 시방 준수 및 상하수도 이설 위치 승인\n2) 관경별 관재 선정 및 이설 수량/내역 정밀 대조",
        "gui_sum": "1) 기존 관로 연계 부위 수압 및 자연유하 경사 사전 확인\n2) 사업소 요청 추가 반영사항 도출 및 내역 수정",
        "chk_sum": "1) 맑은물사업소 상하수도 이설 계획 협의서 및 도면을 확인했는가?\n2) 이설 내역서 수량과 도면 일치 여부를 검측했는가?",
        "disc": "지장물이설", "des": "맑은물사업소 상하수도 시설기준", "risk": "상하수도 이설 위치 불일치로 인한 재협의 리스크 방지", "sub": "상하수도 전문 기술사 자문"
    },
    {
        "l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-2", "dday": "D-40",
        "act": "위수탁 지장물 이설 설계", "own": "엔지니어링팀/위수탁기관",
        "gol": "위수탁기관별 지장물 이설 계획 및 예정가격 산정",
        "mtd": "실시설계 승인 기준 관형 시설계 간섭이 없는 위치로 이설 설계 및 설계비 집행",
        "del": "설계도서 및 분담금 통보서",
        "std_sum": "1) 관형 시설물 및 트램 궤도 하중 간섭 없는 위치 이설 설계\n2) 위수탁기관 원가산정 기준에 따른 분담금 산출",
        "gui_sum": "1) 기관별 예정가격 대조 및 분담금 적정성 검토\n2) 설계도서 승인서 확보 후 분담금 집행 처리",
        "chk_sum": "1) 위수탁 기관 설계도서 및 분담금 산출서를 검측했는가?\n2) 트램 하중(집중 축중) 회피 이설 위치를 확인했는가?",
        "disc": "지장물이설", "des": "국가계약법 및 위수탁 설계 기준", "risk": "위수탁 분담금 예가 과다 산정 리스크 관리", "sub": "위수탁 설계 전문 엔지니어링"
    },
    {
        "l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-3", "dday": "D-30",
        "act": "상하수도 이설계획 실정보고", "own": "현장 공무팀",
        "gol": "실시설계 대비 변경사항에 대한 도급증액 실정보고 승인",
        "mtd": "현장 유관기관 협의 결과로 인한 변경 현황, 도면, 수량, 내역 발주처 승인 요청",
        "del": "상하수도 이설 실정보고 검토서",
        "std_sum": "1) 도급계약서 변경 조건 준수 및 사유서/증빙서류 제출\n2) 현장 실정 변경으로 인한 물량 및 도금 금액 증액 정산",
        "gui_sum": "1) 발주처 및 감리단 사전 보고 후 정식 문서 접수\n2) 실정보고 도면 상의 변경 위치 정밀 측량 결부",
        "chk_sum": "1) 실정보고 도면, 수량, 내역서 작성 적정성을 확인했는가?\n2) 감리단 검토의견서 및 발주처 승인 여부를 검측했는가?",
        "disc": "지장물이설", "des": "건설기술진흥법 및 계약금액 조정 지침", "risk": "실정보고 승인 지연에 따른 변경 금액 누락 방지", "sub": "토목 공무 및 정산 전문가"
    },
    {
        "l3_code": "2000-2", "l3_name": "도급자/위수탁분 설계 및 행정", "l4_code": "2000-2-4", "dday": "D-25",
        "act": "위수탁 계약 체결", "own": "현장 공무팀 / 발주처",
        "gol": "행정, 재무적 책임 확정 및 분담금 집행",
        "mtd": "이설 위수탁 협약서 체결 및 분담금 납부 처리",
        "del": "위수탁 협약서 및 영수증",
        "std_sum": "1) 화성시 및 개별 위수탁 기관 간 공식 협약서 체결 준수\n2) 기관별 분담금 납부 및 영수증 획득 체계 구축",
        "gui_sum": "1) 협약서 조항 상 하자 보수 책임 기간(최소 2년) 명시\n2) 이설 착수일 및 준공 기한 정밀 산정 관리",
        "chk_sum": "1) 위수탁 협약서 날인 및 계약 금액 일치 여부를 확인했는가?\n2) 분담금 납부 영수증 구비 상태를 검측했는가?",
        "disc": "지장물이설", "des": "지방재정법 및 지자체 계약 조례", "risk": "협약 미체결로 인한 시공 후 법적 분쟁 리스크 방지", "sub": "법무 및 계약 전담 전문가"
    },

    # Section C. 지장물 이설 시공 관리 (8단계)
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-1", "dday": "D+0",
        "act": "도로점용/굴착행위 인허가", "own": "현장 공무팀/공사팀",
        "gol": "적법한 작업공간 확보 및 도로 점용/굴착 인허가 획득",
        "mtd": "지자체(화성시) 및 경찰서에 점용·굴착 허가 신청, 이해관계자 협의",
        "del": "도로점용허가증, 굴착허가증",
        "std_sum": "1) 도로법 제61조(도로점용) 및 도로석굴착조례 허가 기준 준수\n2) 허가 조건(작업 시간, 차선 통제 범위) 100% 현장 이행",
        "gui_sum": "1) 허가증 원본 현장 사무실 게시 및 작업 차량 부착\n2) 점용 기간 만료 전 연장 허가 신청 수칙 이행",
        "chk_sum": "1) 도로점용허가증 및 굴착허가증 획득 여부를 확인했는가?\n2) 허가 조건에 명시된 차선 통제 수칙을 점검했는가?",
        "disc": "지장물이설", "des": "도로법 및 도로석굴착 관리조례", "risk": "무단 점용/굴착에 따른 과태료 및 공사중단 리스크 방지", "sub": "인허가 전담 행정 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-2", "dday": "D+0",
        "act": "교통통제 및 교통안전시설 설치", "own": "현장 안전팀/공사팀",
        "gol": "작업자 및 보행자 안전 확보, 도심지 교통 흐름 유지",
        "mtd": "신호수 배치, 교통안전 시설(표지판, 방호벽, 롤링배리어, 점멸등) 설치",
        "del": "안전한 작업 구역(Working Zone) 확보 확인서",
        "std_sum": "1) 도로공사장 교통관리지침(2024.6) 기준 작업구역 방호 조치\n2) 신호수 2인 이상 상시 배치 및 야간 점멸 유도등 가동",
        "gui_sum": "1) 보행자 우회 안전 펜스 및 차릴용 PE 방호벽 물 채움 시공\n2) 출퇴근 시간대(07~09시, 17~19시) 가변 차선 유도",
        "chk_sum": "1) 교통안전 시설물 배치도 대조 및 신호수 배치를 점검했는가?\n2) 야간 점멸등 및 안내 표지판 시인성을 검측했는가?",
        "disc": "지장물이설", "des": "도로공사장 교통관리지침 (국토교통부)", "risk": "교통사고 및 보행자 민원 발생 리스크 예방", "sub": "교통안전 전문 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-3", "dday": "D+1",
        "act": "줄따기(GPR)를 통한 기존 지장물 매설 확인", "own": "현장 공무팀/공사팀",
        "gol": "설계 도면과 실제 지하 매설 지장물 위치 일치 여부 정밀 검증",
        "mtd": "GPR 지형 탐지 및 인력/소형 장비를 이용한 줄따기 굴착(시굴), 깊이/좌표 측정",
        "del": "지장물 실제 위치확인도, 미가식(노출) 확정서",
        "std_sum": "1) 지하매설물 안전관리 지침 준수 및 GPR 탐지 오차 ±10cm 이내\n2) 관로 노출 후 관리기관 입회 하에 심도 및 매설 위치 확인",
        "gui_sum": "1) 대형 장비 백호 굴착 전 반드시 인력 시굴 시행 수칙 엄수\n2) 노출된 지장물 관종별 안전 표지 띠 및 라벨 부착",
        "chk_sum": "1) GPR 탐지 성과표와 실제 인력 시굴 노출 위치를 대조했는가?\n2) 노출 지장물 위치확인도 작성 및 관리기관 서명을 받았는가?",
        "disc": "지장물이설", "des": "지하안전관리에 관한 특별법 (지하안전법)", "risk": "중장비 직접 굴착으로 인한 관로 파손 사고 리스크 방지", "sub": "GPR 탐지 및 지하안전 전문 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-4", "dday": "D+3",
        "act": "이설 위치 토공 굴착", "own": "현장 공사팀",
        "gol": "신규 관로 부설을 위한 터파기 및 안전 흙막이 확보",
        "mtd": "백호(포클레인)를 이용한 터파기 및 가시설(흙막이판, H-Beam) 시공",
        "del": "관로 부설을 위한 노체 형성 검측서",
        "std_sum": "1) KCS 11 20 00 (토공사) 시방 기준 굴착 사면 경사 준수\n2) 굴착 깊이 1.5m 이상 시 흙막이 가시설 및 버팀대 100% 시공",
        "gui_sum": "1) 굴착 법면 토사 유실 방지 방수 시트 덮개 부설\n2) 굴착 저면 지반 지지력(K30 ≥ 110 MN/m³) 확보 검측",
        "chk_sum": "1) 굴착 폭, 깊이 및 흙막이 가시설 안전성을 점검했는가?\n2) 굴착 저면 평탄성 및 잡석/모래 기초 두께를 검측했는가?",
        "disc": "지장물이설", "des": "KCS 11 20 00 토공사 표준시방서", "risk": "굴착 법면 붕괴 및 인접 도로 침하 리스크 방지", "sub": "토질 및 기초 기술사 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-5", "dday": "D+5",
        "act": "신규관로 매설 및 설치", "own": "현장 공사팀 / 전문협력사",
        "gol": "신규 관로 부설 및 연결부 수밀/안전성 확보",
        "mtd": "관 기초(모래 H=200mm) 포설 후 신규 관 부설 및 접합(용접/소켓) 작업",
        "del": "신설 관로 부설 완성 검측서",
        "std_sum": "1) 상하수도/가스/전력관 표준 시방 규격 관재 사용\n2) 관 기초 모래(H=200mm) 다짐도 ≥ 95% 및 접속부 수밀 시험",
        "gui_sum": "1) 관 부설 전 관 내부 이물질 제거 및 신호 띠 부설\n2) 직관부 휨 오차 최소화 및 편심 이음 금지 수칙",
        "chk_sum": "1) 관 기초 모래 두께(200mm) 및 관 부설 레벨을 확인했는가?\n2) 관 이음부 용접/접속 상태 및 비파괴검사를 검측했는가?",
        "disc": "지장물이설", "des": "상하수도 및 배관공사 표준시방서", "risk": "관 이음부 누수 및 침하로 인한 도로 함몰 리스크 예방", "sub": "배관 및 관로 시공 전문 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-6", "dday": "D+8",
        "act": "무단수 연결을 위한 시설 설치", "own": "현장 공사팀 / 전문업체",
        "gol": "공사 중 단수/단전/공급 중단 없는 무단수 바이패스 체계 구축",
        "mtd": "바이패스(Bypass) 관 설치 또는 가이설 선로 구축, 무단수 차단 밸브 거치",
        "del": "서비스 연속성 유지 상태 확인서",
        "std_sum": "1) 무단수 천공 및 밸브 설치 표준 안전 규격 준수\n2) 바이패스 관로 수압 견딤 시험 및 공급 연속성 보장",
        "gui_sum": "1) 본 관로 천공 시 칩(Chip) 관내 유입 방지 포집 장치 작동\n2) 바이패스 관 유량 및 수압 상시 모니터링 수칙",
        "chk_sum": "1) 무단수 바이패스 관로 수압 시험 및 누수 여부를 점검했는가?\n2) 본 천공 밸브 밀폐성 및 서비스 연속 공급 상태를 확인했는가?",
        "disc": "지장물이설", "des": "무단수 공법 기술 시방서", "risk": "단수 발생으로 인한 대규모 주민 민원 리스크 방지", "sub": "무단수 공법 전문 기술자 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-7", "dday": "D+10",
        "act": "신규관로 및 연결관로 접속", "own": "현장 공사팀",
        "gol": "기존 관과 신규 관의 최종 연결 (Cut-over) 및 전환",
        "mtd": "기존 관로 차단 후 신규 관로와 T자 또는 직접 접속(Hot-tapping 등) 및 수밀 검사",
        "del": "시스템 전환(Cut-over) 완료 확인서",
        "std_sum": "1) Cut-over 전환 작업 시간(최대 4시간 이내) 준수\n2) 신구 관로 접속 부위 보강 링 및 전단 앵커 시방 적용",
        "gui_sum": "1) 접속 작업 전 사전 양수 및 내부 유류/잔류물 완전 배출\n2) 전환 직후 수압/가스압 상승 및 누수/누출 여부 집중 모니터링",
        "chk_sum": "1) 신구 관로 Cut-over 접속 완료 및 누수 여부를 확인했는가?\n2) 전환 작업 시간(계획 시간 이내) 준수 여부를 검측했는가?",
        "disc": "지장물이설", "des": "관로 접속 및 Cut-over 시방서", "risk": "접속 부위 압력 파손 및 서비스 전환 실패 리스크 방지", "sub": "관로 접속 전문 자문"
    },
    {
        "l3_code": "2000-3", "l3_name": "지장물 이설 시공 관리", "l4_code": "2000-3-8", "dday": "D+15",
        "act": "기존관로 철거 및 원상복구", "own": "현장 공사팀",
        "gol": "용도 폐기 관로 안전 철거 및 도로 원상복구 포장 마감",
        "mtd": "폐관로 철거 또는 모래 채움, 되메우기(층다짐) 및 아스팔트 포장 마감",
        "del": "도로 원상복구 및 준공 사진대지",
        "std_sum": "1) 도로 복구 포장 시방 준수 (아스콘 층다짐 밀도 ≥ 96%)\n2) 폐관로 잔류물 처리 및 폐기물 관리법 적법 처리",
        "gui_sum": "1) 잔류 폐관 모래/시멘트 밀크 충진으로 도로 침하 예방\n2) 기층/표층 아스콘 타설 시 평탄성(PrI ≤ 10cm/km) 관리",
        "chk_sum": "1) 되메우기 층다짐도(≥95%) 및 아스콘 평탄성을 확인했는가?\n2) 기존 폐관로 철거 또는 충진 처리 상태를 검측했는가?",
        "disc": "지장물이설", "des": "도로 복구 공사 표준시방서", "risk": "복구 포장 부등침하로 인한 포트홀 리스크 방지", "sub": "도로 포장 전문 기술자 자문"
    },

    # Section D. 매설 지장물 관종별 특화 관리
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-1", "dday": "D+20",
        "act": "광역상수관 이설 공사", "own": "공사팀 (협의: 한국수자원공사)",
        "gol": "광역상수도 무단수 분기 및 안정적 이설 완료",
        "mtd": "수자원공사 협의 ➔ 무단수 차단공법 ➔ 신설관 부설 ➔ 무단수 연결 (관 자르지 않고 천공)",
        "del": "수자원공사 무단수 연결 검측서",
        "std_sum": "1) 한국수자원공사 광역상수도 시설 기준 및 무단수 천공 시방 준수\n2) 관경 D800mm 이상 대형관 방동 및 지지 콘크리트 수직 부설",
        "gui_sum": "1) 천공 부위 용접 NDT(100% 무결함) 및 방식 테이프 피복\n2) 수자원공사 감독관 상시 입회 하 시공 수칙 엄수",
        "chk_sum": "1) 광역상수도 무단수 천공 및 밸브 고정 상태를 검측했는가?\n2) 관로 지지 콘크리트 강도(f_ck ≥ 24MPa)를 확인했는가?",
        "disc": "지장물이설", "des": "한국수자원공사 광역상수도 설계기준", "risk": "광역상수도 파손 시 광역 단수 대형 재난 리스크 관리", "sub": "수자원공사 전담 기술 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-2", "dday": "D+22",
        "act": "상수도 관로 이설 공사", "own": "공사팀 (협의: 맑은물사업소)",
        "gol": "시가지 상수관 이설 및 수질 안전성 확보",
        "mtd": "맑은물사업소 협의 ➔ 단수 계획 수립 ➔ 신설관 부설 ➔ 수압시험 및 소독 ➔ 관로 연결",
        "del": "수압시험성적서 및 관로 소독 성과표",
        "std_sum": "1) 화성시 맑은물사업소 상수도 기준 (수압시험 10kg/cm² 1시간 유지)\n2) 신설관 수질 소독(유효염소 50mg/L 24시간 세척) 시방 준수",
        "gui_sum": "1) 시민 불편 최소화를 위해 야간 단수 시간대(23시~05시) 접속\n2) 소독 후 수질 검사(탁도, 잔류염소) 적격 판정 시 통수",
        "chk_sum": "1) 상수도 관로 수압시험(10kg/cm²) 누수 여부를 확인했는가?\n2) 관로 소독 및 수질 검사 적격성 성과표를 검측했는가?",
        "disc": "지장물이설", "des": "상수도 공사 표준시방서", "risk": "상수도 수질 오염 및 단수 시간 초과 리스크 방지", "sub": "맑은물사업소 상수도 감독관 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-3", "dday": "D+25",
        "act": "하수도 관로 이설 공사", "own": "공사팀 (협의: 맑은물사업소)",
        "gol": "자연유하 하수관 정밀 경사 부설 및 수밀성 확보",
        "mtd": "조사 ➔ 물돌리기(Bypass) ➔ 기존관 철거 ➔ 신설관 부설 ➔ 1% 내외 정밀 구배 확인",
        "del": "하수관로 경사 측량 성과표 및 CCTV 검측서",
        "std_sum": "1) 하수도 시설기준 준수 (자연유하 경사 1% 내외 정밀 유지)\n2) 하수관 C.C.T.V 관로 내부 검사 및 물누림 수밀 시험 적격",
        "gui_sum": "1) 기존 하수 바이패스 펌핑 용량 사전 확보로 오수 범람 예방\n2) 맨홀 접속 부위 인버트(Invert) 몰탈 마감 정밀 시공",
        "chk_sum": "1) 하수관로 정밀 레벨 측량으로 1% 구배 형성을 확인했는가?\n2) 하수관 CCTV 내부 검사 및 맨홀 수밀성을 검측했는가?",
        "disc": "지장물이설", "des": "하수도 공사 표준시방서", "risk": "하수 역류 및 구배 불량으로 인한 관로 침전 리스크 예방", "sub": "하수도 정밀 측량 기술자 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-4", "dday": "D+28",
        "act": "도시가스관 이설 공사", "own": "공사팀 (협의: 한국가스공사, 삼천리)",
        "gol": "고압/중압 도시가스관 폭발 위험 제어 및 이설 완수",
        "mtd": "가스회사 협의 ➔ 가스 차단 및 퍼지(Purge) ➔ 절단/용접 ➔ RT 비파괴검사 ➔ 가스 주입",
        "del": "가스관 용접 RT 비파괴검사성적서",
        "std_sum": "1) 도시가스사업법 안전관리 기준 준수 (용접부 RT 100% 검사)\n2) 관리기관(삼천리/가스공사) 안전점검원 현장 입회 필수",
        "gui_sum": "1) 가스 차단 후 배관 내 질소 퍼지(Purge)로 잔류 가스 100% 제거\n2) 가스관 라벨링 및 이중 피복 전기방식(CP) 테이핑 시공",
        "chk_sum": "1) 가스관 용접부 RT(방사선투과검사) 100% 무결함을 확인했는가?\n2) 가스 차단/퍼지 및 관리기관 입회 서명을 검측했는가?",
        "disc": "지장물이설", "des": "도시가스사업법 및 KGC 가스안전 기준", "risk": "가스 누출 및 폭발 대형 안전사고 리스크 방지", "sub": "가스안전공사 및 삼천리가스 전문 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-5", "dday": "D+30",
        "act": "지역난방관 이설 공사", "own": "공사팀 (협의: 한국지역난방공사)",
        "gol": "열수송관 이중보온관 열손실 방지 및 신축 흡수 이설",
        "mtd": "난방공사 협의 ➔ 열매체 퇴수 ➔ 보온관 용접 ➔ 이중보온관 접합 ➔ 열매체 주입",
        "del": "열수송관 보온용접 검측서 및 감지선 도통 시험표",
        "std_sum": "1) 한국지역난방공사 열수송관 공사시방서 준수 (용접부 NDT 합격)\n2) 이중보온관 열수축 조인트 발포 및 누수 감지선 도통 100%",
        "gui_sum": "1) 온도 변동에 따른 관의 수축/팽창을 흡수하는 신축 굴곡(Elbow) 시공\n2) 외관 PUR 폼 발포 밀도 및 감지선 센서 연결 확인",
        "chk_sum": "1) 열수송관 보온용접 NDT 및 열수축 발포 조인을 점검했는가?\n2) 누수 감지선 도통 시험 및 신축 흡수 이격 거리를 검측했는가?",
        "disc": "지장물이설", "des": "한국지역난방공사 열수송관 기술기준", "risk": "열수송관 누수 및 난방 공급 중단 리스크 방지", "sub": "지역난방공사 열수송관 전문 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-6", "dday": "D+35",
        "act": "통신관로 및 케이블 이설 공사", "own": "공사팀 (협의: KT, LGU+, SKT)",
        "gol": "광통신망 끊김 없는 통신 관로 및 광케이블 절체 이설",
        "mtd": "통신사 협의 ➔ 관로 부설 ➔ 광케이블 인입/포설 ➔ 광 융착 접속 및 절체 ➔ 기존관 철거",
        "del": "광케이블 접속 손실 OTDR 성과표",
        "std_sum": "1) 정보통신공사업법 및 통신 3사 관로 부설 기준 준수\n2) 광케이블 접속 손실 OTDR 측정 ≤ 0.05dB/splice 통과",
        "gui_sum": "1) 광케이블 심야시간대(01시~05시) 무중단 절체(Cut-over) 수칙\n2) 예비 관로(Spare Duct) 1개 이상 추가 확보 수칙 준수",
        "chk_sum": "1) 광케이블 OTDR 접속 손실(≤0.05dB) 성과표를 검측했는가?\n2) 심야 절체 시 통신망 장애 유무를 확인했는가?",
        "disc": "지장물이설", "des": "정보통신 표준품셈 및 통신관로 공사 기준", "risk": "광케이블 절단으로 인한 통신 마비 리스크 방지", "sub": "통신 3사 전담 엔지니어 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-7", "dday": "D+40",
        "act": "특고압 전력관로 이설 공사", "own": "공사팀 (협의: 한국전력공사)",
        "gol": "22.9kV 특고압 한전 관로 및 맨홀 안전 이설",
        "mtd": "한전 협의 ➔ 관로 시공 ➔ 맨홀 설치 ➔ 케이블 인입 ➔ 특고압 접속 및 가전",
        "del": "한전 관로 인수인계서 및 절연검사표",
        "std_sum": "1) 한국전력공사 지중배전 공사 기준 준수 (ELP관 H=1.2m 매설)\n2) 22.9kV XLPE 케이블 절연저항 ≥ 2,000MΩ 및 내전압 시험 통과",
        "gui_sum": "1) 관로 내부 이물질 및 습기 완전 제거로 아크/단락 사고 방지\n2) 한전 감독관 입회 하 특고압 접속재(Straight Joint) 시공",
        "chk_sum": "1) 한전 관로 매설 깊이(1.2m 이상) 및 맨홀을 확인했는가?\n2) 22.9kV 케이블 절연저항(≥2,000MΩ) 및 내전압을 검측했는가?",
        "disc": "지장물이설", "des": "한국전력공사 지중배전 건설기준", "risk": "특고압 케이블 감전 및 아크 단락 사고 리스크 방지", "sub": "한전 배전 전문 기술자 자문"
    },
    {
        "l3_code": "2000-4", "l3_name": "주요 지장물 이설 소유자 및 관리특성", "l4_code": "2000-4-8", "dday": "D+45",
        "act": "송유관 이설 공사", "own": "공사팀 (협의: 대한송유관공사)",
        "gol": "위험물 송유관 안전 배출, 질소 치환 및 이설 완료",
        "mtd": "송유관공사 협의 ➔ 잔류유 배출(Pumping) ➔ 질소 치환 ➔ 절단/용접 ➔ RT 비파괴검사",
        "del": "송유관 질소치환 및 RT 검사성적서",
        "std_sum": "1) 송유관 안전관리법 준수 (용접부 RT 100% 방사선검사 합격)\n2) 관내 잔류유 완전 배출 및 질소 압력 치환 안전 기준 준수",
        "gui_sum": "1) 대한송유관공사 안전점검원 100% 입회 하 획지 작업 시행\n2) 가스/유증기 측정기 상시 가동(폭발하한계 0% 유지)",
        "chk_sum": "1) 송유관 잔류유 배출 및 질소 치환 성과표를 검측했는가?\n2) 송유관 용접 RT 비파괴검사 100% 무결함을 확인했는가?",
        "disc": "지장물이설", "des": "송유관 안전관리법 및 송유관공사 기술기준", "risk": "잔류 유류 유출 및 화재 폭발 대형 사고 리스크 방지", "sub": "대한송유관공사 전담 기술 자문"
    },

    # Section E. 최종 점검 및 정산
    {
        "l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-1", "dday": "D+60",
        "act": "상하수도 이설 시공 최종 점검", "own": "현장 감리원 / 맑은물사업소",
        "gol": "상하수도 시설물 관리기관(맑은물사업소) 최종 인계",
        "mtd": "기관, 감리 합동검측으로 기능 및 수밀도 확인, 인계 체크리스트 제출",
        "del": "준공검사 필증, 인계 체크리스트",
        "std_sum": "1) 맑은물사업소 시설물 인계인수 지침 준수\n2) 기능 검측 및 수밀성 최종 검사 100% 적격 판정",
        "gui_sum": "1) 감리원, 사업소 담당자, 현장대리인 삼자 합동 점검\n2) 지적사항 발생 시 3일 이내 보완 완료 체계 가동",
        "chk_sum": "1) 맑은물사업소 인계인수 합동검측 필증을 확인했는가?\n2) 준공 사진대지 및 인계 체크리스트 완비 여부를 점검했는가?",
        "disc": "지장물이설", "des": "지방자치단체 시설물 인계인수 규정", "risk": "인계 미비로 인한 시설물 인수 거부 리스크 방지", "sub": "맑은물사업소 시설물 인계 전담자"
    },
    {
        "l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-2", "dday": "D+65",
        "act": "위수탁 지하지장물 이설 최종 점검", "own": "현장 감리원 / 위수탁기관",
        "gol": "위수탁 지장물 최종 이설 점검 및 정산 금액 협의",
        "mtd": "이설 내역 사후 평가검토 용역 시행, 관리기관 최종 지장물도 도면 점검",
        "del": "시설물 관리 대장, 정산금액 협의서",
        "std_sum": "1) 위수탁 협약서 상 사후 정산 검토 기준 준수\n2) 준공 지장물도 GIS 종합 시스템 반영 정확성 확보",
        "gui_sum": "1) 사후 정산 평가 용역 결과보고서 바탕 기관 협의\n2) 최종 지장물 위치 관리대장 화성시 제출",
        "chk_sum": "1) 위수탁 지장물 사후 정산 평가 검토서를 확인했는가?\n2) 시설물 관리대장 및 GIS 도면 일치성을 검측했는가?",
        "disc": "지장물이설", "des": "위수탁 협약서 및 사후 정산 지침", "risk": "정산 금액 이견으로 인한 분담금 미지급 리스크 관리", "sub": "사후 정산 평가 전문 용역사 자문"
    },
    {
        "l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-3", "dday": "D+70",
        "act": "상하수도 이설 정산 금액 검토", "own": "현장 공무팀",
        "gol": "유관기관 협의 및 현장 상이에 따른 설계변경 추진",
        "mtd": "유관기관 협의결과(실정보고) 도면 현장 상이 검측 대장 관리, 교통대책 비용 반영",
        "del": "설계변경 검토서(도면, 수량, 내역)",
        "std_sum": "1) 계약예규 공사계약일반조건 제19조(설계변경) 기준 준수\n2) 현장 실정 변경 증빙 도면, 수량 산출서 정밀 검토",
        "gui_sum": "1) 감리단 사전 승인 후 발주처 설계변경 정식 신청\n2) 이설 공사 단가 변경 사유서 제출 수칙",
        "chk_sum": "1) 설계변경 검토서 도면, 수량, 내역서 일치 여부를 확인했는가?\n2) 실정보고 반영 금액과 변경 요청액 산출을 검측했는가?",
        "disc": "지장물이설", "des": "기획재정부 계약예규 공사계약일반조건", "risk": "설계변경 절차 미비로 인한 금액 삭감 리스크 방지", "sub": "토목 설계변경 및 계약 정산 전문가"
    },
    {
        "l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-4", "dday": "D+80",
        "act": "위수탁 처리 정산금액 지급", "own": "현장 공무팀 / 발주처",
        "gol": "위수탁 기관 이설 정산금액 확정 지급 및 PS항목 정산",
        "mtd": "이설비용 사후평가용역결과 협의 완료 후 각 기관별 정산금액 청구 및 지급",
        "del": "정산계약서 및 지급 영수증",
        "std_sum": "1) 도급 내역서 상 PS(Provisional Sum) 항목 정산 규정 준수\n2) 위수탁 기관별 정식 청구서 및 사후 평가서 기준 집행",
        "gui_sum": "1) 발주처 승인 후 정산금액 위수탁 기관 입금 처리\n2) 입금 영수증 및 정산 완결 공문 회부 관리",
        "chk_sum": "1) PS항목 정산금액 지급 영수증 및 협약서를 확인했는가?\n2) 위수탁 기관 최종 정산 완결 공문을 검측했는가?",
        "disc": "지장물이설", "des": "지방재정법 및 PS 정산 지침", "risk": "PS 항목 정산 지연 리스크 방지", "sub": "발주처 재정 및 정산 담당 자문"
    },
    {
        "l3_code": "2000-5", "l3_name": "최종 점검, 정산 및 설계변경", "l4_code": "2000-5-5", "dday": "D+90",
        "act": "도급자분/위수탁분 설계변경 정산", "own": "현장 공무팀 / 화성시",
        "gol": "지장물 이설 최종 설계변경 확정 및 화성시 승인",
        "mtd": "도급자분 현황상(국가계약법) 및 위수탁 처리분 사후인가결과 반영 설계변경 요청 및 화성시 승인",
        "del": "최종 설계변경 승인서",
        "std_sum": "1) 화성시 최종 설계변경 승인 및 도급계약 금액 변경 확정\n2) 지장물 이설 전체 공종 최종 정산 보고서 작성",
        "gui_sum": "1) 최종 변경 도면 및 준공 도서 준공계에 반영\n2) 지장물 이설 완료에 따른 강화노반 인계서 작성",
        "chk_sum": "1) 화성시 최종 설계변경 승인서 및 계약 변경을 확인했는가?\n2) 지장물 이설 완료 후 강화노반 공정 인계서를 검측했는가?",
        "disc": "지장물이설", "des": "국가계약법 및 화성시 도급계약 지침", "risk": "최종 계약 변경 지연 리스크 관리", "sub": "화성시 도급계약 및 준공 정산 전문가"
    }
]

print(f"Defined {len(jijangmul_activities)} detailed Jijangmul engineering activities from uploaded images.")

# Step 2. Create HTML files for each Jijangmul activity
print("\nCreating HTML files for Jijangmul activities...")

html_created = 0

for idx, act in enumerate(jijangmul_activities, start=1):
    sanitized_act = sanitize_name(act['act'])
    folder_name = f"{idx}_{sanitized_act}"
    act_dir = os.path.join(jijangmul_attach_dir, folder_name)
    
    std_dir = os.path.join(act_dir, "표준서")
    gui_dir = os.path.join(act_dir, "수행지침")
    chk_dir = os.path.join(act_dir, "체크리스트")
    
    os.makedirs(std_dir, exist_ok=True)
    os.makedirs(gui_dir, exist_ok=True)
    os.makedirs(chk_dir, exist_ok=True)
    
    # 1. Standard HTML
    std_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <title>지장물이설 - {act['act']} 기술 표준서</title>
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; margin: 0; padding: 20px; background: #F8FAFC; color: #1E293B; }}
        .container {{ max-width: 900px; margin: 0 auto; background: #FFF; padding: 30px; border-radius: 12px; border: 1px solid #E2E8F0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); }}
        .header {{ border-bottom: 2px solid #2563EB; padding-bottom: 15px; margin-bottom: 20px; }}
        .breadcrumb {{ font-size: 0.85rem; color: #2563EB; font-weight: bold; margin-bottom: 5px; }}
        .title {{ font-size: 1.8rem; font-weight: 900; color: #0F172A; margin: 0; }}
        .meta-info {{ font-size: 0.9rem; color: #64748B; margin-top: 8px; }}
        .badge {{ background: #DBEAFE; color: #1D4ED8; font-weight: bold; padding: 3px 8px; border-radius: 4px; font-size: 0.8rem; }}
        h2 {{ font-size: 1.3rem; font-weight: 700; color: #1E3A8A; border-left: 4px solid #2563EB; padding-left: 10px; margin-top: 30px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.95rem; }}
        th, td {{ border: 1px solid #E2E8F0; padding: 10px 14px; text-align: left; }}
        th {{ background: #F1F5F9; color: #334155; font-weight: bold; width: 25%; }}
        .spec-box {{ background: #EFF6FF; border: 1px solid #BFDBFE; padding: 15px; border-radius: 8px; margin-top: 10px; }}
        .designer-box {{ background: #FFFBEB; border: 1px solid #FCD34D; padding: 15px; border-radius: 8px; margin-top: 15px; }}
        .footer-note {{ margin-top: 30px; text-align: center; font-size: 0.85rem; color: #94A3B8; border-top: 1px solid #E2E8F0; padding-top: 15px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <div class="breadcrumb">Dongtan Tram WBS {act['l4_code']} Standard</div>
        <h1 class="title">{act['act']} 기술 표준서</h1>
        <div class="meta-info">
            <span><strong>공종/분야:</strong> 지장물이설 / {act['disc']}</span>
            <span>|</span>
            <span><strong>주관부서:</strong> {act['own']} ({act['dday']})</span>
            <span>|</span>
            <span><span class="badge">공통 기술 표준</span></span>
        </div>
    </div>

    <h2>1. 과업 개요 및 목적 (Overview & Scope)</h2>
    <table>
        <tbody>
            <tr>
                <th>과업 목적</th>
                <td>{act['gol']}</td>
            </tr>
            <tr>
                <th>수행 방법</th>
                <td>{act['mtd']}</td>
            </tr>
            <tr>
                <th>주요 산출물</th>
                <td>{act['del']}</td>
            </tr>
            <tr>
                <th>관련 기술 시방</th>
                <td>{act['des']}</td>
            </tr>
        </tbody>
    </table>

    <h2>2. 정량적 공학 절대 기준 (Technical Specifications)</h2>
    <div class="spec-box">
        <ul style="margin: 0; padding-left: 20px; line-height: 1.8;">
            <li>{act['std_sum'].replace(chr(10), '</li><li>')}</li>
        </ul>
    </div>

    <h2>3. 프로세스맵 리스크 및 협력사 자문 수칙</h2>
    <div class="designer-box">
        <p style="margin: 0 0 8px 0; font-weight: bold; color: #B45309;">📍 리스크 관리 및 협력사 이행 지침:</p>
        <ul style="margin: 0; padding-left: 20px; line-height: 1.7; color: #78350F;">
            <li><strong>리스크 관리:</strong> {act['risk']}</li>
            <li><strong>협력사 자문:</strong> {act['sub']}</li>
        </ul>
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 기술 표준서 | WBS {act['l4_code']} | 사전토공·지장물이설
    </div>
</div>
</body>
</html>"""

    with open(os.path.join(std_dir, f"{sanitized_act}_표준서.html"), 'w', encoding='utf-8') as hf:
        hf.write(std_html_content)
        
    # 2. Guideline HTML
    gui_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <title>지장물이설 - {act['act']} 현장 수행지침</title>
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; margin: 0; padding: 20px; background: #F0FDF4; color: #166534; }}
        .container {{ max-width: 900px; margin: 0 auto; background: #FFF; padding: 30px; border-radius: 12px; border: 1px solid #BBF7D0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); }}
        .header {{ border-bottom: 2px solid #16A34A; padding-bottom: 15px; margin-bottom: 20px; }}
        .title {{ font-size: 1.8rem; font-weight: 900; color: #14532D; margin: 0; }}
        .step-box {{ background: #F0FDF4; border-left: 4px solid #16A34A; padding: 15px; margin-bottom: 15px; border-radius: 4px; }}
        .step-title {{ font-weight: bold; color: #15803D; margin-bottom: 5px; }}
        .footer-note {{ margin-top: 30px; text-align: center; font-size: 0.85rem; color: #86EFAC; border-top: 1px solid #DCFCE7; padding-top: 15px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1 class="title">{act['act']} 현장 수행지침 (Playbook)</h1>
        <p style="color: #4AD66D; margin-top: 5px;">WBS {act['l4_code']} | 주관: {act['own']} ({act['dday']})</p>
    </div>

    <h2 style="color: #15803D;">📌 3단계 현장 수행 가이드</h2>
    <div class="step-box">
        <div class="step-title">① 사전 준비 단계 (Preparation)</div>
        <p style="margin: 0; font-size: 0.95rem;">{act['mtd']}</p>
    </div>
    <div class="step-box">
        <div class="step-title">② 본 시공 수행 단계 (Execution)</div>
        <p style="margin: 0; font-size: 0.95rem;">{act['gui_sum'].replace(chr(10), '<br>')}</p>
    </div>
    <div class="step-box">
        <div class="step-title">③ 검사 및 마무리 단계 (Sign-off)</div>
        <p style="margin: 0; font-size: 0.95rem;">산출물 ({act['del']}) 작성 및 {act['sub']} 수칙 확인 완료.</p>
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 수행지침 | WBS {act['l4_code']}
    </div>
</div>
</body>
</html>"""

    with open(os.path.join(gui_dir, f"{sanitized_act}_수행지침.html"), 'w', encoding='utf-8') as hf:
        hf.write(gui_html_content)

    # 3. Checklist HTML
    chk_html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <title>지장물이설 - {act['act']} 검측 체크리스트</title>
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; margin: 0; padding: 20px; background: #FFFBEB; color: #78350F; }}
        .container {{ max-width: 900px; margin: 0 auto; background: #FFF; padding: 30px; border-radius: 12px; border: 1px solid #FDE68A; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); }}
        .header {{ border-bottom: 2px solid #D97706; padding-bottom: 15px; margin-bottom: 20px; }}
        .title {{ font-size: 1.8rem; font-weight: 900; color: #78350F; margin: 0; }}
        .chk-item {{ background: #FEF3C7; padding: 12px 15px; margin-bottom: 10px; border-radius: 6px; display: flex; align-items: center; gap: 10px; }}
        .chk-item input {{ width: 18px; height: 18px; }}
        .footer-note {{ margin-top: 30px; text-align: center; font-size: 0.85rem; color: #D97706; border-top: 1px solid #FEF3C7; padding-top: 15px; }}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1 class="title">{act['act']} 검측 체크리스트</h1>
        <p style="color: #B45309; margin-top: 5px;">WBS {act['l4_code']} | 실시간 O/X 검측 도구</p>
    </div>

    <h2 style="color: #B45309;">☐ 실시간 검측 필수 항목</h2>
    <div class="chk-item">
        <input type="checkbox">
        <span>{act['chk_sum'].splitlines()[0] if act['chk_sum'] else act['act'] + " 기준을 확인했는가?"}</span>
    </div>
    <div class="chk-item">
        <input type="checkbox">
        <span>{act['chk_sum'].splitlines()[1] if len(act['chk_sum'].splitlines())>1 else "관련 산출물(" + act['del'] + ") 작성을 완료하였는가?"}</span>
    </div>
    <div class="chk-item">
        <input type="checkbox">
        <span>[협력사 자문] {act['sub']} 및 리스크 조치를 확인했는가?</span>
    </div>

    <div class="footer-note">
        동탄도시철도(트램) 건설공사 엔지니어링 체크리스트 | WBS {act['l4_code']}
    </div>
</div>
</body>
</html>"""

    with open(os.path.join(chk_dir, f"{sanitized_act}_체크리스트.html"), 'w', encoding='utf-8') as hf:
        hf.write(chk_html_content)
        
    html_created += 3

print(f"Successfully generated {html_created} HTML files across {len(jijangmul_activities)} Jijangmul activity folders.")

# Step 3. Write into '지장물이설' sheet in Excel v3
print("\nReconstructing '지장물이설' sheet in v3 Excel...")

wb_v2 = openpyxl.load_workbook(v2_path)

if '지장물이설' in wb_v2.sheetnames:
    # Rename existing sheet to avoid confusion or clear it
    idx = wb_v2.sheetnames.index('지장물이설')
    del wb_v2['지장물이설']
    ws = wb_v2.create_sheet(title='지장물이설', index=idx)
else:
    ws = wb_v2.create_sheet(title='지장물이설', index=2)

# Styles
header_fill_sum = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate Dark
header_fill_link = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
font_white = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

font_normal = Font(name="맑은 고딕", size=9, bold=False, color="000000")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# 20 Columns layout 준용 (사전토공사 시트 형식 100% 준용)
headers_20 = [
    "L2 코드", "L3 코드", "L3 대공종명", "L4 코드", "일정 (D-Day)",
    "작업단위 (Level 4 Task/Activity)", "주관", "목적", "방법", "산출물(결과)",
    "표준서 (Standard) 요약", "표준서 파일 (HTML)",
    "수행지침 (Guideline) 요약", "수행지침 파일 (HTML)",
    "체크리스트 (Checklist) 요약", "체크리스트 파일 (HTML)",
    "담당 분야", "첨부서류 연계 상세 설계기준", "집행단계 리스크 체크리스트", "협력사 시공/공사관리 자문"
]

header_row = 1

# Write Headers
for c_idx, h_text in enumerate(headers_20, 1):
    cell = ws.cell(row=header_row, column=c_idx, value=h_text)
    cell.font = font_white
    cell.alignment = align_center
    if '파일 (HTML)' in h_text:
        cell.fill = header_fill_link
    else:
        cell.fill = header_fill_sum

# Set Column Widths
ws.column_dimensions['A'].width = 10 # L2
ws.column_dimensions['B'].width = 12 # L3
ws.column_dimensions['C'].width = 30 # L3 Name
ws.column_dimensions['D'].width = 12 # L4
ws.column_dimensions['E'].width = 12 # D-Day
ws.column_dimensions['F'].width = 38 # Activity
ws.column_dimensions['G'].width = 25 # Owner
ws.column_dimensions['H'].width = 35 # Goal
ws.column_dimensions['I'].width = 45 # Method
ws.column_dimensions['J'].width = 30 # Deliverable
ws.column_dimensions['K'].width = 45 # Std Sum
ws.column_dimensions['L'].width = 22 # Std Link
ws.column_dimensions['M'].width = 45 # Gui Sum
ws.column_dimensions['N'].width = 22 # Gui Link
ws.column_dimensions['O'].width = 45 # Chk Sum
ws.column_dimensions['P'].width = 22 # Chk Link
ws.column_dimensions['Q'].width = 20 # Disc
ws.column_dimensions['R'].width = 35 # Des
ws.column_dimensions['S'].width = 35 # Risk
ws.column_dimensions['T'].width = 35 # Sub

# Populate Data Rows
for r_idx, act in enumerate(jijangmul_activities, start=2):
    sanitized_act = sanitize_name(act['act'])
    folder_name = f"{r_idx-1}_{sanitized_act}"
    
    ws.cell(row=r_idx, column=1, value="2000").alignment = align_center
    ws.cell(row=r_idx, column=2, value=act['l3_code']).alignment = align_center
    ws.cell(row=r_idx, column=3, value=act['l3_name']).alignment = align_left
    ws.cell(row=r_idx, column=4, value=act['l4_code']).alignment = align_center
    ws.cell(row=r_idx, column=5, value=act['dday']).alignment = align_center
    ws.cell(row=r_idx, column=6, value=act['act']).alignment = align_left
    ws.cell(row=r_idx, column=7, value=act['own']).alignment = align_center
    ws.cell(row=r_idx, column=8, value=act['gol']).alignment = align_left
    ws.cell(row=r_idx, column=9, value=act['mtd']).alignment = align_left
    ws.cell(row=r_idx, column=10, value=act['del']).alignment = align_left
    
    # 11: Std Sum
    ws.cell(row=r_idx, column=11, value=act['std_sum']).alignment = align_left
    # 12: Std Link
    c12 = ws.cell(row=r_idx, column=12, value="👉 [더블클릭] 표준서 열기 📄")
    std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\표준서\\{sanitized_act}_표준서.html"
    c12.hyperlink = Hyperlink(ref=c12.coordinate, target=std_rel_path)
    c12.font = font_link
    c12.alignment = align_center
    
    # 13: Gui Sum
    ws.cell(row=r_idx, column=13, value=act['gui_sum']).alignment = align_left
    # 14: Gui Link
    c14 = ws.cell(row=r_idx, column=14, value="👉 [더블클릭] 수행지침 열기 📄")
    gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\수행지침\\{sanitized_act}_수행지침.html"
    c14.hyperlink = Hyperlink(ref=c14.coordinate, target=gui_rel_path)
    c14.font = font_link
    c14.alignment = align_center
    
    # 15: Chk Sum
    ws.cell(row=r_idx, column=15, value=act['chk_sum']).alignment = align_left
    # 16: Chk Link
    c16 = ws.cell(row=r_idx, column=16, value="👉 [더블클릭] 체크리스트 열기 📄")
    chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\지장물이설\\{folder_name}\\체크리스트\\{sanitized_act}_체크리스트.html"
    c16.hyperlink = Hyperlink(ref=c16.coordinate, target=chk_rel_path)
    c16.font = font_link
    c16.alignment = align_center
    
    # 17: Disc
    ws.cell(row=r_idx, column=17, value=f"사전토공·{act['disc']}").alignment = align_center
    # 18: Des
    ws.cell(row=r_idx, column=18, value=act['des']).alignment = align_left
    # 19: Risk
    ws.cell(row=r_idx, column=19, value=act['risk']).alignment = align_left
    # 20: Sub
    ws.cell(row=r_idx, column=20, value=act['sub']).alignment = align_left
    
    for c_idx in range(1, 21):
        cell = ws.cell(row=r_idx, column=c_idx)
        if c_idx not in [12, 14, 16]:
            cell.font = font_normal
        cell.border = thin_border

print(f"Successfully populated {len(jijangmul_activities)} rows into '지장물이설' sheet.")

# Step 4. Save workbook as v3 (다른 이름으로 저장 - v3)
wb_v2.save(v3_save_path)
print(f"\n🎉 Successfully saved updated workbook as v3 to '{v3_save_path}'")
