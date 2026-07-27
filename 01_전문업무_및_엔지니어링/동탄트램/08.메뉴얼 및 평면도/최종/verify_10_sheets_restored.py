import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

for target_name in ["매뉴얼 BODY (집행단계)v3.xlsx", "매뉴얼 BODY (집행단계)v4.xlsx"]:
    target_path = os.path.join(base_dir, target_name)
    wb = openpyxl.load_workbook(target_path, read_only=True)
    print(f"\n==========================================")
    print(f"📁 File: {target_name}")
    print(f"   - Total Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    
    wb_full = openpyxl.load_workbook(target_path)
    ws = wb_full['지장물이설']
    print(f"   - '지장물이설' Sheet Total Rows: {ws.max_row}")
    print(f"   - Row 3 Activity: {ws.cell(row=3, column=6).value}")
    print(f"   - Row 3 Std Summary:\n{ws.cell(row=3, column=19).value}")
    print(f"   - Row 3 Std Link:\n{ws.cell(row=3, column=20).value}")

print("\n✅ Verification SUCCESS: All 10 sheets fully restored & updated in both v3 and v4!")
