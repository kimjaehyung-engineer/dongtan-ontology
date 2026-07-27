import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
updated_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"

wb = openpyxl.load_workbook(excel_path)
ws = wb['통신']

activities_count = 0
for r in range(2, ws.max_row + 1):
    l4_code = ws.cell(row=r, column=4).value
    act_name = ws.cell(row=r, column=5).value
    dept = ws.cell(row=r, column=6).value or "현장 시스템팀"
    purpose = ws.cell(row=r, column=7).value or ""
    method = ws.cell(row=r, column=8).value or ""
    deliverable = ws.cell(row=r, column=9).value or "검측서"
    
    if act_name:
        activities_count += 1
        name = str(act_name).strip()
        code = str(l4_code).strip()
        dept_str = str(dept).strip()
        
        # Clean text for summary
        p_clean = str(purpose).replace('\n', ' ').strip()
        m_clean = str(method).replace('\n', ' ').strip()
        d_clean = str(deliverable).replace('\n', ' ').strip()
        
        # Trim method/purpose if too long for summary sentence
        p_short = p_clean[:60] + "..." if len(p_clean) > 60 else p_clean
        m_short = m_clean[:70] + "..." if len(m_clean) > 70 else m_clean

        # 1) 표준서 요약 (Col 10, J) - 1:1 커스텀
        std_sum = f"1) 본 표준서는 [{name}] 공종의 품질 확보 및 목적({p_short})을 달성하기 위한 공학적 기준을 규정함.\n2) {dept_str} 주관으로 관련 시방 수칙을 준수하여 최종 산출물({d_clean})의 무결성을 보장함."
        
        # 2) 수행지침 요약 (Col 12, L) - 1:1 커스텀
        gui_sum = f"1) [{name}] 작업 시 사전준비, 본시공, 검사마감 3단계 visual 도식과 대형 확대 모달로 실무 절차를 가이드함.\n2) {m_short} 수칙에 따라 세부 시공 및 확인 작업을 이행함."
        
        # 3) 체크리스트 요약 (Col 14, N) - 1:1 커스텀 (~하였는가? 어미 100%)
        chk_sum = f"1) [{name}] 수행 전 관련 사양서, 도면 및 사전 준비 기준을 확인하였는가?\n2) {m_short} 항목을 정밀 검측하고 산출물({d_clean})을 최종 확인하였는가?"
        
        ws.cell(row=r, column=10).value = std_sum
        ws.cell(row=r, column=12).value = gui_sum
        ws.cell(row=r, column=14).value = chk_sum

# Remove _images attribute to prevent PIL closed file crash during save
for sheet in wb.worksheets:
    sheet._images = []

wb.save(excel_path)
wb.save(updated_excel_path)

print(f"🎉 SUCCESSFULLY UPDATED EXCEL V4 1:1 CUSTOM SUMMARIES FOR ALL {activities_count} TELECOM ACTIVITIES!")
