import openpyxl
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
source_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_최종업그레이드.xlsx")
target_file = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3.xlsx")

print(f"Syncing upgraded '지장물이설' sheet from '{source_file}' into '{target_file}'...")

try:
    # 1. Load source workbook
    wb_src = openpyxl.load_workbook(source_file)
    ws_src = wb_src['지장물이설']

    # 2. Load target workbook
    wb_tgt = openpyxl.load_workbook(target_file)
    
    # Check if '지장물이설' sheet exists in target, if so remove and copy fresh
    if '지장물이설' in wb_tgt.sheetnames:
        idx = wb_tgt.sheetnames.index('지장물이설')
        wb_tgt.remove(wb_tgt['지장물이설'])
        ws_tgt = wb_tgt.create_sheet('지장물이설', idx)
    else:
        ws_tgt = wb_tgt.create_sheet('지장물이설')

    # 3. Copy cells, formatting, merges, and values
    for row in ws_src.iter_rows():
        for cell in row:
            tgt_cell = ws_tgt.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                tgt_cell.font = openpyxl.styles.Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                tgt_cell.fill = openpyxl.styles.PatternFill(fill_type=cell.fill.fill_type, start_color=cell.fill.start_color, end_color=cell.fill.end_color)
                tgt_cell.border = openpyxl.styles.Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=cell.border.bottom)
                tgt_cell.alignment = openpyxl.styles.Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text)
                tgt_cell.number_format = cell.number_format

    # Copy column dimensions
    for col in ws_src.column_dimensions:
        ws_tgt.column_dimensions[col].width = ws_src.column_dimensions[col].width

    # Copy row dimensions
    for r in ws_src.row_dimensions:
        ws_tgt.row_dimensions[r].height = ws_src.row_dimensions[r].height

    # Copy merged cells
    for merged_range in ws_src.merged_cells.ranges:
        ws_tgt.merge_cells(str(merged_range))

    wb_tgt.save(target_file)
    print(f"🎉 Successfully synchronized '지장물이설' sheet directly into '{target_file}'!")

except PermissionError:
    print(f"⚠️ Notice: '{target_file}' is currently locked by MS Excel.")
    backup_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v3_연동완료.xlsx")
    shutil.copy(source_file, backup_path)
    print(f"✅ Created updated copy as '{backup_path}'. Please save/close Excel to sync '{target_file}'.")
