import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"
v1_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v1.xlsx")
if not os.path.exists(v1_path):
    v1_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)_최종공정매뉴얼완성본.xlsx")
if not os.path.exists(v1_path):
    v1_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계).xlsx")

v2_save_path = os.path.join(base_dir, "매뉴얼 BODY (집행단계)v2.xlsx")
base_attach_dir = os.path.join(base_dir, "매뉴얼BODY(집행단계-첨부폴더)")

print(f"Loading base workbook from: {v1_path}")
wb = openpyxl.load_workbook(v1_path)

# System specific highly detailed specs (Gongjeong Manual mapping dictionary)
system_specs_db = {
    # Signal Discipline
    "신호": {
        "own": "신호 시스템팀 (D-30)",
        "gol": "전자연동장치(CBI) SIL 4 안전성 확보 및 교차로 트램 우선신호 연동",
        "mtd": "CBI 응답검측, 분기기 축차계수기 듀얼센서 설치, LTE-R 단말-도로교통 제어기 연동",
        "del": "SIL 4 안전인증서 / 우선신호 응답시험성적서 / 축차계수기 설치검측서",
        "std_sum": "1) 전자연동장치(CBI) SIL 4 (IEC 61508/62425) 100% 안전 무결성 확보\n2) 교차로 PPC 트램 우선신호 LTE-R 제어기 연동 응답시간 ≤ 100ms 준수",
        "gui_sum": "1) 도상 타설 시 바이브레이터가 신호 루프 센서(Loop Coil)에 직접 타격 없도록 5cm 이격\n2) 축차계수기 듀얼센서 토크(45~50Nm) 및 선로전환기 밀착유격(≤1.5mm) 연동",
        "chk_sum": "1) 전자연동장치(CBI) SIL 4 무결성 및 루프 코일 매설 파손 방지 조치를 검측했는가?\n2) 교차로 PPC 우선신호 응답시간(≤100ms) 및 TOD 녹색 현시 연동을 확인했는가?",
        "disc": "신호분야",
        "des": "IEC 61508 / IEC 62425 SIL 4 철도 신호 안전 규격",
        "risk": "신호 센서 파손으로 인한 열차 감지 오류 및 교차로 신호 지연 리스크 관리",
        "sub": "위례선 트램 시공사 정주/트리폴건설 자문 및 신호 전문 자격자 전담 배치"
    },
    
    # Telecom Discipline
    "통신": {
        "own": "통신 시스템팀 (D-30)",
        "gol": "700MHz LTE-R 음영지역 제로화 및 72-Core 광백본 이중화 링 구축",
        "mtd": "OTDR 광접속 손실 검측, 4K IP CCTV/NVR 24시간 이중화, PIS/PA/PSD 통신 연동",
        "del": "LTE-R 무선망 측정 성과표 / 광백본 OTDR 시험성적서 / PSD 연동 보고서",
        "std_sum": "1) 700MHz LTE-R 전 구간 무선 수신강도 RSSI ≥ -95dBm, 핸드오버 성공률 ≥ 99.5%\n2) 72-Core 싱글모드 광케이블(ITU-T G.652D) OTDR 접속손실 ≤ 0.05dB/splice",
        "gui_sum": "1) 지하지장물 이설 시 내역서상 PS(Provisional Sum) 항목 지정 및 수칙 준수\n2) 전력 케이블 유도장해 방지를 위해 통신 케이블 30cm 이격 및 차폐 트레이 시공",
        "chk_sum": "1) 700MHz LTE-R 전 구간 수신강도(RSSI ≥ -95dBm) 및 72Core 광백본을 검측했는가?\n2) PIS 시각 오차(≤1초), PA STI(≥0.6), PSD 통신 지연(≤100ms)을 확인했는가?",
        "disc": "통신분야",
        "des": "700MHz TTA LTE-R 표준 규격 및 광통신 설비 기준",
        "risk": "통신 음영지역 발생으로 인한 승강장 행선안내 및 4K CCTV 영상 끊김 리스크 방지",
        "sub": "광통신 융착 전문 기술자 및 통신 시스템 시공 자문"
    },
    
    # Electrical Power Discipline
    "전기": {
        "own": "전력 시스템팀 (D-30)",
        "gol": "DC 750V 12-Pulse 급전, 수전 변전소 이중화 및 표유전류 부식 방지",
        "mtd": "한전 2계통 수전, 무전차선 승강장 DC 750V/1000A 급속충전, 디오드 접지 부설",
        "del": "변전소 수전 시험성적서 / 표유전류 전압 측정보고서 / ESS 급속충전 결과서",
        "std_sum": "1) 수전 변전소 한전 인입 선로 2개 계통 독립 이중화 및 DC 750V 정류기 규격 준수\n2) 표유전류(Stray Current) 상수도관 부식 방지용 디오드 접지(Polarization Cell) 부설",
        "gui_sum": "1) 무전차선 승강장 정차 30초 내 80% 급속충전 오버헤드 펜타그래프 인입 및 제어\n2) 레일 주변 엘라스토머 박스(Elastomer Box) 절연재 부설 및 절연저항 ≥ 10MΩ 측정",
        "chk_sum": "1) DC 750V 수전 변전소 정류기 동작 및 2개 계통 한전 인입 이중화를 확인했는가?\n2) 표유전류 방지 디오드 접지 및 레일-대지 전압(≤120V)을 검측하였는가?",
        "disc": "전기분야",
        "des": "EN 50122-2 표유전류 부식 방지 규격 및 전기설비기술기준",
        "risk": "누설전류에 의한 상수도관 부식 및 변전 정전 발생 리스크 관리",
        "sub": "22.9kV 특고압 기술사 및 무가선 급속충전 전문 자문"
    }
}

font_normal = Font(name="맑은 고딕", size=9, bold=False, color="000000")
font_link = Font(name="맑은 고딕", size=9, bold=True, color="0000FF", underline="single")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

master_ws = wb['공정매뉴얼']
header_row = 3

reupdated_systems_count = 0

for r in range(header_row + 1, master_ws.max_row + 1):
    c_l3 = str(master_ws.cell(row=r, column=3).value or "").strip()
    d_l4 = str(master_ws.cell(row=r, column=4).value or "").strip()
    e_act = str(master_ws.cell(row=r, column=5).value or "").strip()
    p_disc = str(master_ws.cell(row=r, column=16).value or "").strip()
    
    text_combine = (c_l3 + " " + d_l4 + " " + e_act + " " + p_disc).lower()
    
    # Check if this row belongs to Signal, Telecom, or Electrical Systems
    matched_domain = None
    if any(k in text_combine for k in ['신호', 'cbi', '축차계수기', '선로전환기', '우선신호']):
        matched_domain = "신호"
    elif any(k in text_combine for k in ['통신', 'lte', 'cctv', 'psd', '광케이블', '광백본']):
        matched_domain = "통신"
    elif any(k in text_combine for k in ['전기', '전차선', '급전', '변전소', 'dc 750v', '누설전류', '표유전류', '충전', '오버헤드']):
        matched_domain = "전기"
        
    if matched_domain:
        spec = system_specs_db[matched_domain]
        disc_folder = f"{matched_domain}분야"
        
        # Determine specific folder target
        target_act_folder = "1_설계적정성 검토"
        for act_f in os.listdir(os.path.join(base_attach_dir, disc_folder)):
            sanitized_e = e_act.replace('/', '_').replace('\\', '_').strip()
            if sanitized_e in act_f or d_l4 in act_f:
                target_act_folder = act_f
                break
                
        # 1. Col 6: 주관
        master_ws.cell(row=r, column=6, value=spec['own']).alignment = align_center
        master_ws.cell(row=r, column=6).font = font_normal
        master_ws.cell(row=r, column=6).border = thin_border
        
        # 2. Col 7: 목적
        master_ws.cell(row=r, column=7, value=spec['gol']).alignment = align_left
        master_ws.cell(row=r, column=7).font = font_normal
        master_ws.cell(row=r, column=7).border = thin_border
        
        # 3. Col 8: 방법
        master_ws.cell(row=r, column=8, value=spec['mtd']).alignment = align_left
        master_ws.cell(row=r, column=8).font = font_normal
        master_ws.cell(row=r, column=8).border = thin_border
        
        # 4. Col 9: 산출물(결과)
        master_ws.cell(row=r, column=9, value=spec['del']).alignment = align_left
        master_ws.cell(row=r, column=9).font = font_normal
        master_ws.cell(row=r, column=9).border = thin_border
        
        # 5. Col 10: 표준서 요약 (2줄)
        master_ws.cell(row=r, column=10, value=spec['std_sum']).alignment = align_left
        master_ws.cell(row=r, column=10).font = font_normal
        master_ws.cell(row=r, column=10).border = thin_border
        master_ws.cell(row=r, column=10).hyperlink = None
        
        # 6. Col 11: 표준서 파일 (HTML) - Dedicated Link
        c11 = master_ws.cell(row=r, column=11, value="👉 [더블클릭] 표준서 열기 📄")
        std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\표준서\\설계적정성 검토_표준서.html"
        # Check if actual file exists
        if os.path.exists(os.path.join(base_attach_dir, disc_folder, target_act_folder, "표준서")):
            files = [f for f in os.listdir(os.path.join(base_attach_dir, disc_folder, target_act_folder, "표준서")) if f.endswith('.html')]
            if files:
                std_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\표준서\\{files[0]}"
        c11.hyperlink = Hyperlink(ref=c11.coordinate, target=std_rel_path)
        c11.font = font_link
        c11.alignment = align_center
        c11.border = thin_border

        # 7. Col 12: 수행지침 요약 (2줄)
        master_ws.cell(row=r, column=12, value=spec['gui_sum']).alignment = align_left
        master_ws.cell(row=r, column=12).font = font_normal
        master_ws.cell(row=r, column=12).border = thin_border
        master_ws.cell(row=r, column=12).hyperlink = None
        
        # 8. Col 13: 수행지침 파일 (HTML) - Dedicated Link
        c13 = master_ws.cell(row=r, column=13, value="👉 [더블클릭] 수행지침 열기 📄")
        gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\수행지침\\설계적정성 검토_수행지침.html"
        if os.path.exists(os.path.join(base_attach_dir, disc_folder, target_act_folder, "수행지침")):
            files = [f for f in os.listdir(os.path.join(base_attach_dir, disc_folder, target_act_folder, "수행지침")) if f.endswith('.html')]
            if files:
                gui_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\수행지침\\{files[0]}"
        c13.hyperlink = Hyperlink(ref=c13.coordinate, target=gui_rel_path)
        c13.font = font_link
        c13.alignment = align_center
        c13.border = thin_border

        # 9. Col 14: 체크리스트 요약 (2줄)
        master_ws.cell(row=r, column=14, value=spec['chk_sum']).alignment = align_left
        master_ws.cell(row=r, column=14).font = font_normal
        master_ws.cell(row=r, column=14).border = thin_border
        master_ws.cell(row=r, column=14).hyperlink = None

        # 10. Col 15: 체크리스트 파일 (HTML) - Dedicated Link
        c15 = master_ws.cell(row=r, column=15, value="👉 [더블클릭] 체크리스트 열기 📄")
        chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\체크리스트\\설계적정성 검토_체크리스트.html"
        if os.path.exists(os.path.join(base_attach_dir, disc_folder, target_act_folder, "체크리스트")):
            files = [f for f in os.listdir(os.path.join(base_attach_dir, disc_folder, target_act_folder, "체크리스트")) if f.endswith('.html')]
            if files:
                chk_rel_path = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_folder}\\{target_act_folder}\\체크리스트\\{files[0]}"
        c15.hyperlink = Hyperlink(ref=c15.coordinate, target=chk_rel_path)
        c15.font = font_link
        c15.alignment = align_center
        c15.border = thin_border

        # 11. Col 16: 담당 분야
        master_ws.cell(row=r, column=16, value=spec['disc']).alignment = align_center
        master_ws.cell(row=r, column=16).font = font_normal
        master_ws.cell(row=r, column=16).border = thin_border

        # 12. Col 17: 첨부서류 연계 상세 설계기준
        master_ws.cell(row=r, column=17, value=spec['des']).alignment = align_left
        master_ws.cell(row=r, column=17).font = font_normal
        master_ws.cell(row=r, column=17).border = thin_border

        # 13. Col 18: 집행단계 리스크 체크리스트
        master_ws.cell(row=r, column=18, value=spec['risk']).alignment = align_left
        master_ws.cell(row=r, column=18).font = font_normal
        master_ws.cell(row=r, column=18).border = thin_border

        # 14. Col 19: 협력사 시공/공사관리 자문
        master_ws.cell(row=r, column=19, value=spec['sub']).alignment = align_left
        master_ws.cell(row=r, column=19).font = font_normal
        master_ws.cell(row=r, column=19).border = thin_border

        reupdated_systems_count += 1

print(f"\nSystems re-population complete for {reupdated_systems_count} rows!")

# Save as v2 (다른 이름으로 저장 - v2)
wb.save(v2_save_path)
print(f"Successfully saved new v2 workbook to '{v2_save_path}'")
