import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

def clean_and_format_2_3_lines(text):
    if not text: return ""
    text = str(text).strip()
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    
    clean_lines = []
    for l in lines:
        # Strip all existing numbers, parenthesis, bullets
        clean = re.sub(r'^\s*[\d\.\-☐☑📍📌\)\(\]]+\s*', '', l).strip()
        clean = re.sub(r'^\s*\d+[\.\)]\s*', '', clean).strip()
        clean = re.sub(r'\[동탄트램 업무 매뉴얼 v1 연계\]\s*', '', clean)
        clean = re.sub(r'\[동탄트램 매뉴얼 v1\]\s*', '', clean)
        clean = re.sub(r'\[설계사 작성\]\s*', '', clean)
        clean = clean.replace('기술기준 연계', '기술기준').replace('설계기준 연계', '설계기준')
        
        if clean and clean not in clean_lines:
            clean_lines.append(clean)
            
    selected = clean_lines[:3] # Maximum 3 lines (2~3 lines)
    if not selected: return text[:100]
    
    return "\n".join([f"{i+1}) {line}" for i, line in enumerate(selected)])

cleaned_cells = 0

for sheet_name in wb.sheetnames:
    if sheet_name == 'GUIDE': continue
    ws = wb[sheet_name]
    header_row = 3 if sheet_name == '공정표에따른 매뉴얼' else 1
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    sum_cols = [c for c, h in enumerate(headers, 1) if ('표준서' in h or '수행지침' in h or '체크리스트' in h) and '요약' in h]
    if not sum_cols:
        sum_cols = [c for c, h in enumerate(headers, 1) if ('표준서' in h or '수행지침' in h or '체크리스트' in h) and '파일' not in h]
        
    for r in range(header_row + 1, ws.max_row + 1):
        for c in sum_cols:
            cell = ws.cell(row=r, column=c)
            if cell.value:
                formatted = clean_and_format_2_3_lines(cell.value)
                if formatted != cell.value:
                    cell.value = formatted
                    cleaned_cells += 1

print(f"Numbering cleanup complete for {cleaned_cells} summary cells!")
wb.save(excel_path)
print(f"Saved updated Excel file to '{excel_path}'")
