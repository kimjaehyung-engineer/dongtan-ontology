import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

target_folder = None
for f in os.listdir(base_dir):
    if f.startswith("6_") or ("관제" in f and "운영사" in f):
        target_folder = os.path.join(base_dir, f)
        break

if not target_folder:
    print("❌ ERROR: Target folder for WBS 9000-2-6 not found!")
    sys.exit(1)

print(f"Target WBS 9000-2-6 Folder: {target_folder}")

zoom_modal_style = """
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .clickable-diagram::after {
        content: "🔍 클릭하여 대형 확대보기";
        position: absolute;
        bottom: 8px;
        right: 12px;
        background: rgba(15, 23, 42, 0.75);
        color: #ffffff;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        backdrop-filter: blur(4px);
        pointer-events: none;
        opacity: 0.85;
        transition: opacity 0.2s;
    }
    .clickable-diagram:hover::after {
        opacity: 1;
        background: rgba(2, 132, 199, 0.9);
    }
    .glossary-modal, .zoom-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .glossary-modal.active, .zoom-modal.active {
        display: flex;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 550px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-close, .zoom-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .glossary-close:hover, .zoom-close:hover {
        color: #ef4444;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal">
    <div class="glossary-modal-content">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 관제 인터페이스 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    'occ_console_display': {
        title: '🖥️ 종합관제실(OCC) 모니터 화면 배치 & 4K CCTV 65개소 명당 자리 정하기',
        desc: '동탄 종합관제실(OCC)과 수원·화성 도로교통관제센터 모니터에 트램 운행 상황이 한눈에 들어오도록 화면을 맞추고, 교차로 4K CCTV 카메라 65개소가 사각지대 없이 보이도록 위치를 배치하는 가이드입니다.'
    },
    'city_traffic_control': {
        title: '🚦 수원·화성시 도로교통 관제센터와 신호등 신호 연동하기',
        desc: '수원시 및 화성시 도로교통관제센터와 트램 우대 신호(신호등을 푸른불로 바꿔주는 장치) 신호를 통신선으로 연결하여 막힘없이 연동되는지 대조하는 수칙입니다.'
    },
    'station_pis_pa_05s': {
        title: '📺 승강장 전광판, 스피커 안내방송, 관제실 도착 정보 동시에 0.5초 만에 띄우기',
        desc: '트램이 역에 들어올 때 승강장 전광판의 이번 열차 도착 글자, 스피커 안내방송, 관제실 모니터 운행 정보가 0.5초 만에 동시에 작동하는지 시험하는 지침입니다.'
    },
    'psd_bell_driver_facility': {
        title: '🚨 승강장 안전문 비상 벨 3초 관제 연결 & 기관사 편의시설 위치 확인',
        desc: '승강장 안전문(PSD) 비상 벨 조작 시 3초 안에 관제실로 연결되는지 확인하고, 트램 운전실 및 승강장에 무전기/안내판이 일하기 좋은 위치에 붙었는지 확인하는 수칙입니다.'
    }
};

function openGlossary(termKey) {
    const data = glossaryData[termKey];
    if (!data) return;
    document.getElementById('modalTitle').innerText = data.title;
    document.getElementById('modalDescription').innerText = data.desc;
    document.getElementById('glossaryModal').classList.add('active');
}
function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeGlossaryModal();
        closeZoomModal();
    }
});
</script>
"""

# 1. STANDARD HTML (EASY GUIDANCE)
std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 관제 및 운영사 인터페이스 협의 표준서 (WBS 9000-2-6 초직관 개편)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard (WBS 9000-2-6)</span>
        <h1 class="text-3xl font-black mt-2">관제 및 운영사 인터페이스 협의 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-6 | 주관: 현장 시스템팀 / 통신 협업업체</p>
    </div>
    
    <div class="p-8 space-y-8">
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-blue-950 mb-2">🎯 쉽게 설명한 표준 목적 (Objective)</h3>
            <p class="text-slate-700 text-sm font-medium leading-relaxed">
                동탄 종합관제센터(OCC), 수원·화성 도로교통관제센터 및 종합운영사의 요구사항을 사전 조사하여, <strong>관제실 모니터 화면 배치, 교차로 4K CCTV 65개소 사각지대 제로 배치, 트램 우대 신호등 연동, 승강장 전광판·스피커·CCTV 0.5초 동시 띄우기 및 비상 벨 3초 관제 연결 요구사항을 100% 확정</strong>하는 실무 표준입니다.
            </p>
        </div>

        <div>
            <h3 class="text-lg font-bold text-slate-900 mb-4 border-b-2 border-blue-600 pb-2">📜 쉽게 풀어서 쓴 5대 관제·운영사 표준 수칙</h3>
            <ul class="space-y-3">
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 1</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>종합관제실(OCC) 모니터 배치 & 4K CCTV 65개소 명당 자리 배치:</strong> 동탄 종합관제실(OCC) 모니터 화면 배치를 확정하고, 교차로 12m 강관주 4K CCTV 65개소 위치를 사각지대 없이 명당으로 결정합니다.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 2</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>수원·화성시 도로교통 관제센터와 신호등 신호 연동:</strong> 수원시 및 화성시 도로교통관제센터와 트램 우대 신호(신호등을 푸른불로 바꿔주는 장치) 신호를 통신선으로 연동 대조합니다.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 3</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>승강장 전광판·스피커·관제 화면 0.5초 동시 띄우기:</strong> 트램 진입 시 승강장 전광판 글자, 스피커 안내방송, 관제실 모니터 운행 정보가 0.5초 만에 동시에 표출되는지 시험합니다.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 4</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>승강장 비상 벨 3초 관제 연결 & 기관사 편의시설 배치:</strong> 승강장 안전문(PSD) 비상 벨 조작 시 3초 안에 관제실 연결을 검측하고, 트램 운전실 무전기/안내판 명당 배치를 확정합니다.</span>
                </li>
                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙 5</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed"><strong>안테나 잘 터짐 합격증(TTA/KCA) 챙기기 & 관제·운영사 합격 도장:</strong> 무전기와 안테나가 잘 터진다는 정부 합격증(TTA/KCA 필증) 서류를 챙기고 관제·운영사·감리원 합동 서명 도장을 찍습니다.</span>
                </li>
            </ul>
        </div>

        <div class="bg-emerald-50 border border-emerald-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-emerald-950 mb-2">📦 증빙 서류 (Deliverables)</h3>
            <p class="text-emerald-900 text-sm font-bold">관제 및 운영사 인터페이스 회의록, 4K CCTV 65개소 위치 승인도, TTA/KCA 전파 합격증, 관제·운영사 서명 관리대장</p>
        </div>
    </div>
</div>
</body>
</html>
"""

# 2. GUIDELINE HTML (ULTRA-EASY & VISUAL SVG)
gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - 관제 및 운영사 인터페이스 초직관 수행지침서 (WBS 9000-2-6)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Detailed Guideline (WBS 9000-2-6)</span>
        <h1 class="text-3xl font-black mt-2">관제 및 운영사 인터페이스 초직관 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: 9000-2-6 | 주관: 현장 시스템팀 / 통신 협업업체 | "Step별 1:1 2D Visual 그림 완벽 수록"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <!-- 💡 친절한 개요 해설 박스 -->
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-3">
            <h4 class="font-bold text-base flex items-center gap-2">💡 한눈에 읽는 쉬운 실무 가이드</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                어려운 전문 용어를 모두 배제하고 <strong>"종합관제실(OCC), 시교통관제센터 및 운영사가 일하기 가장 편한 형태로 어떻게 통신선과 모니터, 카메라를 연결하는지"</strong> 쉬운 텍스트와 <strong>Step별 1:1 직관적 2D Visual 그림</strong>으로 가이드합니다.
                모든 그림은 <strong><span class="term-highlight" onclick="openDiagramZoom('svg_step1', 'STEP 1 관제실 모니터 배치 도식')">클릭하면 커다란 모달 창으로 크게 연동</span></strong>됩니다.
            </p>
        </div>

        <!-- ☀️ 5대 핵심 실무 요약 카드 -->
        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">EASY SUMMARY</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 관제 및 운영사 협의 한눈에 보기</h3>
            </div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🖥️</span> 1. OCC 모니터 & 4K 카메라 65개소</span>
                    <p class="text-slate-700 text-xs">관제실 모니터 화면을 배치하고 4K 카메라 65개소 사각지대 제로 자리 확정.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🚦</span> 2. 수원·화성 시교통관제 신호 연동</span>
                    <p class="text-slate-700 text-xs">수원/화성 도로교통관제센터와 트램 우대 신호등 신호선을 통신선으로 연결.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📺</span> 3. 전광판·스피커·관제 0.5초 동시 작동</span>
                    <p class="text-slate-700 text-xs">트램 진입 시 전광판, 스피커, 관제 화면이 0.5초 만에 동시에 켜지는지 확인.</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🚨</span> 4. 비상 벨 3초 관제 연결 & 기관사 편의</span>
                    <p class="text-slate-700 text-xs">승강장 비상 벨 3초 관제 연결 및 트램 운전실 무전기/안내판 명당 배치.</p>
                </div>
            </div>
        </div>

        <!-- 1. FLEXIBLE 5-STEP ARCHITECTURE -->
        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> 5단계 수행 마스터 프로세스
            </h2>
            <div class="grid grid-cols-1 sm:grid-cols-2 md:grid-cols-5 gap-2">
                <div class="bg-blue-50 p-3 rounded-xl border border-blue-200 flex flex-col justify-between">
                    <span class="bg-blue-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 1</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">OCC 모니터배치</h4>
                    <p class="text-[10px] text-blue-900 mt-1 font-medium">• 관제 화면 배치<br">• 4K 카메라 65개소</p>
                </div>
                <div class="bg-indigo-50 p-3 rounded-xl border border-indigo-200 flex flex-col justify-between">
                    <span class="bg-indigo-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 2</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">시교통 신호연동</h4>
                    <p class="text-[10px] text-indigo-900 mt-1 font-medium">• 수원/화성 관제센터<br">• 트램 우대 신호등</p>
                </div>
                <div class="bg-cyan-50 p-3 rounded-xl border border-cyan-200 flex flex-col justify-between">
                    <span class="bg-cyan-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 3</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">0.5초 3중연동</h4>
                    <p class="text-[10px] text-cyan-900 mt-1 font-medium">• 전광판+스피커<br">• 관제실 운행모니터</p>
                </div>
                <div class="bg-teal-50 p-3 rounded-xl border border-teal-200 flex flex-col justify-between">
                    <span class="bg-teal-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 4</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">3초 비상통화</h4>
                    <p class="text-[10px] text-teal-900 mt-1 font-medium">• 비상 벨 3초 연결<br">• 기관사 무전기 배치</p>
                </div>
                <div class="bg-emerald-50 p-3 rounded-xl border border-emerald-200 flex flex-col justify-between">
                    <span class="bg-emerald-600 text-white text-[9px] font-black px-1.5 py-0.5 rounded w-fit">STEP 5</span>
                    <h4 class="font-bold text-slate-900 text-[11px] mt-1">합격증&관제도장</h4>
                    <p class="text-[10px] text-emerald-900 mt-1 font-medium">• 정부 전파 합격증<br">• 관제/운영사 도장</p>
                </div>
            </div>
        </div>

        <!-- 🔥 2. 초정밀 HOW 세부 실무 가이드 (STEP 1~5 1:1 직관적 2D VISUAL SVG) -->
        <div class="space-y-8">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-indigo-600 pb-2">
                <span class="text-indigo-600">2.</span> 쉬운 설명 & Step별 1:1 직관적 2D 그림 지침
            </h2>

            <div class="space-y-8 text-sm">
                <!-- STEP 1 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-blue-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-blue-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 1</span>
                        <h3 class="font-bold text-base text-slate-900">종합관제실(OCC) 모니터 화면 배치 & 4K CCTV 카메라 65개소 명당 자리 정하기</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 동탄 종합관제실(OCC)과 수원·화성 도로교통관제센터 모니터에 트램 운행 상황이 한눈에 들어오도록 화면 위치를 맞추고, 교차로 4K CCTV 카메라 65개소가 사각지대 없이 잘 보이도록 위치를 정합니다.</p>
                    </div>
                    
                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step1', 'STEP 1 OCC 관제실 모니터 배치 & 4K 카메라 65개소 2D 시공 도식')">
                        <svg id="svg_step1" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- OCC 관제 대형 모니터 -->
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                            <text x="130" y="48" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">🖥️ 동탄 OCC 종합관제실 모니터</text>
                            <g transform="translate(35, 62)">
                                <rect x="0" y="0" width="55" height="35" fill="#e0f2fe" stroke="#2563eb" stroke-width="1.5" rx="3"/>
                                <text x="27" y="22" font-size="9" font-weight="black" fill="#1d4ed8" text-anchor="middle">트램 1호</text>

                                <rect x="65" y="0" width="55" height="35" fill="#e0f2fe" stroke="#2563eb" stroke-width="1.5" rx="3"/>
                                <text x="92" y="22" font-size="9" font-weight="black" fill="#1d4ed8" text-anchor="middle">트램 2호</text>

                                <rect x="130" y="0" width="55" height="35" fill="#e0f2fe" stroke="#2563eb" stroke-width="1.5" rx="3"/>
                                <text x="157" y="22" font-size="9" font-weight="black" fill="#1d4ed8" text-anchor="middle">전력/신호</text>

                                <rect x="0" y="42" width="185" height="20" fill="#dcfce7" stroke="#15803d" stroke-width="1.5" rx="3"/>
                                <text x="92" y="56" font-size="10" font-weight="black" fill="#15803d" text-anchor="middle">종합 운행 상황판 100% 연결 완료</text>
                            </g>

                            <path d="M 245 90 L 285 90" stroke="#2563eb" stroke-width="3"/>
                            <polygon points="285,85 295,90 285,95" fill="#2563eb"/>

                            <!-- 교차로 4K CCTV 65개소 -->
                            <rect x="305" y="25" width="195" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="402" y="48" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📹 4K CCTV 65개소 배치</text>
                            <circle cx="402" cy="85" r="22" fill="#d1fae5" stroke="#059669" stroke-width="2"/>
                            <text x="402" y="90" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">65개소</text>
                            <text x="402" y="125" font-size="10" font-weight="bold" fill="#334155" text-anchor="middle">12m Pole 사각지대 제로 배치</text>
                            <text x="402" y="142" font-size="10" font-weight="bold" fill="#047857" text-anchor="middle">✔ 관제실 화질 승인 완료</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 2 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-indigo-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-indigo-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 2</span>
                        <h3 class="font-bold text-base text-slate-900">수원·화성시 도로교통 관제센터와 신호등 신호 연동 방법</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 수원시 및 화성시 도로교통관제센터와 트램 우대 신호(신호등을 트램이 올 때 푸른불로 바꿔주는 장치) 신호를 통신선으로 연결하여 신호가 막힘없이 연동되는지 대조합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step2', 'STEP 2 수원/화성 시교통관제센터 신호 연동 2D 시공 도식')">
                        <svg id="svg_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- 도로 신호등 -->
                            <rect x="20" y="25" width="140" height="130" fill="#ffffff" stroke="#4f46e5" stroke-width="2" rx="8"/>
                            <text x="90" y="50" font-size="13" font-weight="black" fill="#3730a3" text-anchor="middle">🚦 트램 우대 신호등</text>
                            <g transform="translate(45, 68)">
                                <circle cx="15" cy="15" r="10" fill="#ef4444"/>
                                <circle cx="45" cy="15" r="10" fill="#f59e0b"/>
                                <circle cx="75" cy="15" r="12" fill="#10b981" stroke="#047857" stroke-width="2"/>
                            </g>
                            <text x="90" y="120" font-size="10" font-weight="bold" fill="#15803d" text-anchor="middle">트램 진입 시 푸른불 켜짐</text>
                            <text x="90" y="140" font-size="10" font-weight="bold" fill="#334155" text-anchor="middle">교차로 제어기(PPC) 연동</text>

                            <!-- 통신선 연동 화살표 -->
                            <path d="M 160 90 L 250 90" stroke="#4f46e5" stroke-width="3"/>
                            <polygon points="250,85 260,90 250,95" fill="#4f46e5"/>
                            <text x="205" y="80" font-size="10" font-weight="black" fill="#4f46e5" text-anchor="middle">광통신선 전송</text>

                            <!-- 수원/화성 시교통관제센터 -->
                            <rect x="265" y="25" width="235" height="130" fill="#ffffff" stroke="#0284c7" stroke-width="2" rx="8"/>
                            <text x="382" y="50" font-size="13" font-weight="black" fill="#0369a1" text-anchor="middle">🏢 수원/화성 시교통관제센터</text>
                            <rect x="285" y="68" width="195" height="32" fill="#e0f2fe" stroke="#0284c7" stroke-width="1.5" rx="4"/>
                            <text x="382" y="88" font-size="11" font-weight="black" fill="#0369a1" text-anchor="middle">🚦 트램 우선신호 제어 100% 연동</text>
                            <text x="382" y="122" font-size="11" font-weight="bold" fill="#15803d" text-anchor="middle">✔ 신호 지연 없는 실시간 제어 승인</text>
                            <text x="382" y="142" font-size="10" font-weight="bold" fill="#64748b" text-anchor="middle">교통 제어 패킷 정상 수신 확인</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 3 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-cyan-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-cyan-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 3</span>
                        <h3 class="font-bold text-base text-slate-900">승강장 전광판, 스피커 안내방송, 관제실 도착 정보 동시에 0.5초 만에 띄우기</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 트램이 역에 들어올 때 승강장 전광판의 '이번 열차 도착' 글자, 스피커의 안내방송, 관제실 모니터의 운행 정보가 0.5초 만에 동시에 작동하는지 꼼꼼히 확인합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step3', 'STEP 3 전광판·스피커·관제실 0.5초 동시 띄우기 2D 시공 도식')">
                        <svg id="svg_step3" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- 트램 진입 신호 -->
                            <rect x="20" y="25" width="120" height="130" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
                            <text x="80" y="55" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">🚃 트램 역 진입</text>
                            <circle cx="80" cy="95" r="22" fill="#cffafe" stroke="#0891b2" stroke-width="2"/>
                            <text x="80" y="100" font-size="11" font-weight="black" fill="#0891b2" text-anchor="middle">⏱️ 0.5초</text>
                            <text x="80" y="140" font-size="10" font-weight="bold" fill="#047857" text-anchor="middle">동시 신호 발생</text>

                            <!-- 동시 작동 3종 세트 -->
                            <g transform="translate(160, 25)">
                                <rect x="0" y="0" width="340" height="38" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="6"/>
                                <text x="170" y="24" font-size="12" font-weight="black" fill="#1d4ed8" text-anchor="middle">📺 승강장 PIS 전광판: "이번 열차 잠시 후 도착" (0.5초 표출)</text>

                                <rect x="0" y="46" width="340" height="38" fill="#ffffff" stroke="#059669" stroke-width="2" rx="6"/>
                                <text x="170" y="70" font-size="12" font-weight="black" fill="#047857" text-anchor="middle">🔊 스피커 PA 방송: "열차가 들어옵니다" (0.5초 안내)</text>

                                <rect x="0" y="92" width="340" height="38" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="6"/>
                                <text x="170" y="116" font-size="12" font-weight="black" fill="#0e7490" text-anchor="middle">🖥️ 관제실 운행 모니터: 위치 맵 100% 동시 갱신 완료</text>
                            </g>
                        </svg>
                    </div>
                </div>

                <!-- STEP 4 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-teal-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-teal-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 4</span>
                        <h3 class="font-bold text-base text-slate-900">승강장 안전문 비상 벨 3초 관제 연결 & 기관사 편의시설 위치 확인 방법</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 승강장 안전문(PSD)에 붙은 빨간 비상 벨을 누르면 3초 안에 관제실로 연결되는지 확인하고, 트램 운전실 및 승강장에 무전기와 비상 안내판이 일하기 좋은 위치에 붙었는지 확인합니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step4', 'STEP 4 비상 벨 3초 관제 연결 & 기관사 무전기 배치 2D 시공 도식')">
                        <svg id="svg_step4" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- 승강장 비상 벨 -->
                            <rect x="20" y="25" width="150" height="130" fill="#ffffff" stroke="#ef4444" stroke-width="2" rx="8"/>
                            <text x="95" y="50" font-size="13" font-weight="black" fill="#b91c1c" text-anchor="middle">🚨 승강장 비상 벨</text>
                            <circle cx="95" cy="90" r="22" fill="#fee2e2" stroke="#ef4444" stroke-width="3"/>
                            <text x="95" y="95" font-size="11" font-weight="black" fill="#b91c1c" text-anchor="middle">PRESS</text>
                            <text x="95" y="135" font-size="10" font-weight="bold" fill="#334155" text-anchor="middle">비상 버튼 꾹 누름</text>

                            <!-- 3초 신호 이동 화살표 -->
                            <path d="M 170 90 L 260 90" stroke="#0d9488" stroke-width="4"/>
                            <polygon points="260,85 270,90 260,95" fill="#0d9488"/>
                            <text x="215" y="80" font-size="11" font-weight="black" fill="#0f766e" text-anchor="middle">⚡ 3초 이내 연결</text>

                            <!-- 관제실 & 트램 운전실 배치 -->
                            <rect x="280" y="25" width="220" height="130" fill="#ffffff" stroke="#0d9488" stroke-width="2" rx="8"/>
                            <text x="390" y="50" font-size="13" font-weight="black" fill="#0f766e" text-anchor="middle">🕹️ 트램 운전실 무전기 명당</text>
                            <rect x="300" y="68" width="180" height="32" fill="#ccfbf1" stroke="#0d9488" stroke-width="1.5" rx="4"/>
                            <text x="390" y="88" font-size="11" font-weight="black" fill="#0f766e" text-anchor="middle">📻 무전기/안내판 손닿는 배치</text>
                            <text x="390" y="122" font-size="11" font-weight="bold" fill="#15803d" text-anchor="middle">✔ 기관사 편의시설 승인 완료</text>
                            <text x="390" y="142" font-size="10" font-weight="bold" fill="#64748b" text-anchor="middle">조작 편의성 100% 합격</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 5 HOW + VISUAL SVG -->
                <div class="bg-white p-6 rounded-2xl border border-emerald-200 shadow-sm space-y-4">
                    <div class="flex items-center gap-3">
                        <span class="bg-emerald-600 text-white font-bold text-xs px-2.5 py-1 rounded">STEP 5</span>
                        <h3 class="font-bold text-base text-slate-900">안테나 잘 터짐 정부 합격증(TTA/KCA) 챙기기 & 관제·운영사 담당자 합격 도장 찍기</h3>
                    </div>
                    <div class="bg-slate-50 p-4 rounded-xl text-xs text-slate-700 space-y-2 leading-relaxed">
                        <p><strong>💡 쉬운 실무 설명:</strong> 무전기와 안테나가 잘 터지고 해킹당하지 않는다는 정부 합격증(TTA/KCA)을 최종 챙기고, 현장 감리원과 종합관제실·운영사 담당자가 모여 합격 서명 도장을 찍습니다.</p>
                    </div>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl border border-slate-200 mt-3" onclick="openDiagramZoom('svg_step5', 'STEP 5 정부 전파 합격증 & 관제/운영사 날인 2D 시공 도식')">
                        <svg id="svg_step5" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <!-- 정부 합격증 -->
                            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="130" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📑 정부 무선 전파 합격증</text>
                            <text x="35" y="78" font-size="11" font-weight="bold" fill="#334155">• 안테나 잘 터짐 인증서 (TTA)</text>
                            <text x="35" y="100" font-size="11" font-weight="bold" fill="#334155">• 무선 해킹 방지 인증 완료 (TTA)</text>
                            <text x="35" y="122" font-size="11" font-weight="bold" fill="#047857">• 최종 전파 적합 합격증 (KCA)</text>
                            <text x="35" y="142" font-size="10" font-weight="bold" fill="#64748b">• 법정 승인 서류 100% 준비</text>

                            <!-- 관제/운영사 서명 도장 -->
                            <rect x="260" y="25" width="240" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                            <text x="380" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🖊️ 관제·운영사·감리원 합격 도장</text>
                            <g transform="translate(295, 68)">
                                <circle cx="25" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="25" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">관제</text>

                                <circle cx="85" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="85" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">운영사</text>

                                <circle cx="145" cy="18" r="16" fill="#fee2e2" stroke="#dc2626" stroke-width="1.5"/>
                                <text x="145" y="22" font-size="10" font-weight="black" fill="#dc2626" text-anchor="middle">감리원</text>
                            </g>
                            <text x="380" y="122" font-size="11" font-weight="black" fill="#047857" text-anchor="middle">✔ 종합 관리대장 최종 체결 완료</text>
                            <text x="380" y="142" font-size="10" font-weight="bold" fill="#64748b" text-anchor="middle">1,2공구 연동 완결</text>
                        </svg>
                    </div>
                </div>
            </div>
        </div>

        <!-- 3. 종합 2D VISUAL SVG DIAGRAM -->
        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-emerald-600 pb-2">
                <span class="text-emerald-600">3.</span> 종합 2D Visual 기술 도식 (Enriched 2D SVG)
            </h2>
            <div class="clickable-diagram bg-slate-50 p-5 rounded-2xl border border-slate-200 shadow-inner" onclick="openDiagramZoom('svg_r4', '[WBS 9000-2-6] 관제 및 운영사 인터페이스 초직관 종합 도식')">
                <svg id="svg_r4" viewBox="0 0 550 190" width="100%" height="190" xmlns="http://www.w3.org/2000/svg">
                    <rect x="0" y="0" width="550" height="190" fill="#f8fafc"/>
                    <rect x="25" y="15" width="230" height="145" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
                    <text x="140" y="42" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">🖥️ OCC 모니터 & 시교통 신호</text>
                    <text x="40" y="70" font-size="11" font-weight="bold" fill="#334155">• OCC 관제실 모니터 화면 배치</text>
                    <text x="40" y="93" font-size="11" font-weight="bold" fill="#334155">• 4K CCTV 65개소 사각지대 제로</text>
                    <text x="40" y="116" font-size="11" font-weight="bold" fill="#334155">• 수원/화성 시교통관제 신호 연동</text>
                    <text x="40" y="139" font-size="11" font-weight="bold" fill="#1d4ed8">• 트램 우대 신호등 푸른불 제어</text>

                    <path d="M 265 85 L 295 85" stroke="#2563eb" stroke-width="3"/>
                    <polygon points="295,80 305,85 295,90" fill="#2563eb"/>

                    <rect x="310" y="15" width="215" height="145" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                    <text x="417" y="42" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📺 0.5초 연동 & 관제 도장</text>
                    <text x="325" y="70" font-size="11" font-weight="bold" fill="#334155">• 전광판+스피커+관제 0.5초 동시</text>
                    <text x="325" y="93" font-size="11" font-weight="bold" fill="#334155">• 승강장 비상 벨 3초 관제 소리</text>
                    <text x="325" y="116" font-size="11" font-weight="bold" fill="#334155">• 정부 전파 합격증(TTA/KCA) 챙기기</text>
                    <text x="325" y="139" font-size="11" font-weight="bold" fill="#047857">• 관제·운영사·감리원 합격 도장</text>
                    <text x="275" y="175" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">WBS 9000-2-6 관제 및 운영사 인터페이스 100% 확정</text>
                </svg>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

# 3. CHECKLIST HTML (ULTRA-EASY QUESTIONS)
chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>관제 및 운영사 인터페이스 협의 마스터 체크리스트 (WBS 9000-2-6 초직관 개편)</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; background-color: #f8fafc; }}
        .step-header {{ background-color: #ffffff; text-align: center; font-weight: 800; color: #1e3a8a; border-right: 1px solid #e2e8f0; }}
        .result-box {{ background-color: #ffffff; text-align: center; font-weight: 700; color: #2563eb; border-left: 1px solid #e2e8f0; }}
        {zoom_modal_style}
    </style>
</head>
<body class="p-6 sm:p-10 text-slate-800">
<div class="max-w-5xl mx-auto space-y-6">

    <!-- 대제목 & WBS 코드 -->
    <div class="flex justify-between items-end border-b-2 border-slate-900 pb-4">
        <div>
            <h1 class="text-3xl font-black text-slate-900 tracking-tight">관제 및 운영사 인터페이스 협의 마스터 체크리스트</h1>
        </div>
        <div class="text-right">
            <span class="text-blue-600 font-bold text-sm">WBS Code 9000-2-6 | 통신 검측대장</span>
        </div>
    </div>

    <!-- 📋 상단 안내 상자 (Notice Box) -->
    <div class="bg-blue-50/80 border border-blue-200 rounded-2xl p-6 shadow-sm space-y-2">
        <h3 class="text-base font-bold text-blue-950 flex items-center gap-2">
            <span>📋</span> 쉽게 풀어쓴 12대 현장 점검 체크리스트
        </h3>
        <p class="text-xs text-blue-900 leading-relaxed font-medium">
            본 체크리스트는 동탄 종합관제실(OCC) 모니터 배치, 4K 카메라 65개소 명당 자리, 수원/화성 시교통관제 신호 연동, 전광판·스피커 0.5초 동시 작동 및 비상 벨 3초 관제 연결 점검 시 현장 엔지니어가 <strong>[🔍 시공 도식 열기]</strong>를 클릭하면 <strong>대형 고화질 팝업 모달</strong>이 열려 도식을 직접 보며 <strong>~하였는가? (100%)</strong> 점검을 진행할 수 있도록 연동되었습니다.
        </p>
    </div>

    <!-- 숨겨진 현장 점검용 Step 1~3 2D Visual SVG 소스 -->
    <div style="display:none;">
        <svg id="svg_chk_step1" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="8"/>
            <text x="130" y="50" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">🖥️ OCC 관제실 모니터 배치</text>
            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• 종합관제실 모니터 화면배치 완결</text>
            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 4K CCTV 65개소 사각지대 제로</text>
            <text x="35" y="121" font-size="11" font-weight="bold" fill="#1d4ed8">• 화면 표출 규격 100% 승인</text>

            <rect x="270" y="25" width="230" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="385" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🚦 수원/화성 시교통관제 연동</text>
            <text x="285" y="75" font-size="11" font-weight="bold" fill="#334155">• 수원·화성시 관제센터 신호선 연결</text>
            <text x="285" y="98" font-size="11" font-weight="bold" fill="#334155">• 트램 우대 신호등 푸른불 제어</text>
            <text x="285" y="121" font-size="11" font-weight="bold" fill="#047857">• 교차로 신호 지연 0초 달성</text>
        </svg>

        <svg id="svg_chk_step2" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#0891b2" stroke-width="2" rx="8"/>
            <text x="130" y="50" font-size="13" font-weight="black" fill="#0e7490" text-anchor="middle">📺 0.5초 동시 작동 시험</text>
            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• 승강장 전광판 + 스피커 0.5초 방송</text>
            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 관제실 운행 모니터 동시 갱신</text>
            <text x="35" y="121" font-size="11" font-weight="bold" fill="#0e7490">• 3중 0.5초 동시 작동 합격</text>

            <rect x="270" y="25" width="230" height="130" fill="#ffffff" stroke="#0d9488" stroke-width="2" rx="8"/>
            <text x="385" y="50" font-size="13" font-weight="black" fill="#0f766e" text-anchor="middle">🚨 비상 벨 3초 관제 소리 연결</text>
            <text x="285" y="75" font-size="11" font-weight="bold" fill="#334155">• 승강장 비상 벨 3초 관제 연결</text>
            <text x="285" y="98" font-size="11" font-weight="bold" fill="#334155">• 기관사 운전실 무전기 손닿는 배치</text>
            <text x="285" y="121" font-size="11" font-weight="bold" fill="#0f766e">• 기관사 편의시설 100% 승인</text>
        </svg>

        <svg id="svg_chk_step3" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
            <rect x="20" y="25" width="220" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="130" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📄 정부 무선 합격증 (TTA/KCA)</text>
            <text x="35" y="75" font-size="11" font-weight="bold" fill="#334155">• 무전기 잘 터짐 인증서 (TTA)</text>
            <text x="35" y="98" font-size="11" font-weight="bold" fill="#334155">• 무선 해킹 방지 인증서 완료 (TTA)</text>
            <text x="35" y="121" font-size="11" font-weight="bold" fill="#047857">• 전파법 최종 전파 합격증 확보 (KCA)</text>

            <rect x="270" y="25" width="230" height="130" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
            <text x="385" y="50" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">🖊️ 관제·운영사 합격 도장</text>
            <text x="285" y="75" font-size="11" font-weight="bold" fill="#334155">• 관제/운영사/감리원 담당자 날인</text>
            <text x="285" y="98" font-size="11" font-weight="bold" fill="#334155">• 회의록 종합 대장 등재 완료</text>
            <text x="285" y="121" font-size="11" font-weight="bold" fill="#047857">• 1,2공구 인터페이스 완결</text>
        </svg>
    </div>

    <!-- 3-Column 마스터 검측 테이블 -->
    <div class="bg-white border border-slate-200 rounded-2xl shadow-xl overflow-hidden">
        <table class="w-full border-collapse">
            <thead>
                <tr class="bg-slate-100 text-slate-700 text-sm font-extrabold border-b border-slate-200">
                    <th class="py-4 px-6 text-center w-48 border-r border-slate-200">시공 단계</th>
                    <th class="py-4 px-6 text-center">필수 검측 항목 (쉬운 질문형 12대 수칙)</th>
                    <th class="py-4 px-6 text-center w-36 border-l border-slate-200">점검 결과</th>
                </tr>
            </thead>
            <tbody class="divide-y divide-slate-200">
                <!-- STEP 1 Group -->
                <tr class="hover:bg-slate-50 transition-colors">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-amber-500 text-base">⚠️</span>
                            <div class="text-sm font-black text-slate-900">사전 준비</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 1 모니터배치)</div>
                            <button onclick="openDiagramZoom('svg_chk_step1', 'Step 1 모니터배치 시공 도식')" class="bg-blue-600 hover:bg-blue-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">1. OCC 화면배치</span><strong class="text-slate-900">[화면 맞추기]</strong> 동탄 종합관제실(OCC) 모니터에 트램 운행 상황이 한눈에 들어오도록 화면을 배치하였는가?</td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">2. 4K CCTV 65개</span><strong class="text-slate-900">[명당 자리]</strong> 교차로 4K CCTV 카메라 65개소가 사각지대 없이 보이도록 명당 자리를 배치하였는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">3. 시교통 관제</span><strong class="text-slate-900">[수원/화성 연동]</strong> 수원시 및 화성시 도로교통관제센터와 통신선을 연결하여 신호 연동을 대조하였는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">4. 트램 우대신호</span><strong class="text-slate-900">[푸른불 제어]</strong> 트램 진입 시 신호등이 푸른불로 바뀌는 교차로 제어기(PPC) 연동을 확인하였는가?</td>
                </tr>

                <!-- STEP 2 Group -->
                <tr class="hover:bg-slate-50 transition-colors border-t-2 border-slate-200">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-blue-500 text-base">📋</span>
                            <div class="text-sm font-black text-slate-900">8대 조건</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 2 동시3중연동)</div>
                            <button onclick="openDiagramZoom('svg_chk_step2', 'Step 2 동시3중연동 시공 도식')" class="bg-indigo-600 hover:bg-indigo-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">5. 전광판 0.5초</span><strong class="text-slate-900">[글자 표출]</strong> 트램 진입 시 승강장 전광판 글자가 0.5초 만에 작동하는지 확인하였는가?</td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">6. 스피커 안내</span><strong class="text-slate-900">[동시 방송]</strong> 트램 진입 시 스피커 안내방송이 0.5초 만에 동시에 울리는지 시험하였는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">7. 관제 화면동시</span><strong class="text-slate-900">[위치 갱신]</strong> 트램 진입 시 관제실 모니터 운행 위치가 0.5초 만에 동시에 갱신되는지 확인하였는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">8. 비상 벨 3초</span><strong class="text-slate-900">[관제실 소리]</strong> 승강장 비상 벨 조작 시 3초 안에 관제실로 연결되어 울리는지 시험하였는가?</td>
                </tr>

                <!-- STEP 3 Group -->
                <tr class="hover:bg-slate-50 transition-colors border-t-2 border-slate-200">
                    <td rowspan="4" class="step-header p-6 w-48 align-middle bg-slate-50/50 border-r border-slate-200">
                        <div class="space-y-2">
                            <span class="text-emerald-500 text-base">🤝</span>
                            <div class="text-sm font-black text-slate-900">마감 승인</div>
                            <div class="text-xs text-blue-600 font-bold">(Step 3 결과 체결)</div>
                            <button onclick="openDiagramZoom('svg_chk_step3', 'Step 3 마감승인 시공 도식')" class="bg-emerald-600 hover:bg-emerald-700 text-white text-[10px] font-bold px-2 py-1 rounded shadow transition mt-1">🔍 시공 도식 열기</button>
                        </div>
                    </td>
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">9. 기관사 편의</span><strong class="text-slate-900">[무전기 배치]</strong> 트램 운전실 무전기와 승강장 안내판이 기관사 손닿는 명당에 배치되었는가?</td>
                    <td rowspan="4" class="result-box p-6 w-36 align-middle bg-slate-50/30 text-blue-600 font-bold text-sm">☐ 확인완료</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">10. 안테나 잘터짐</span><strong class="text-slate-900">[TTA 인증]</strong> 무전기와 안테나가 잘 터지고 해킹 안 된다는 정부 인증서(TTA)를 챙겼는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">11. 전파 합격증</span><strong class="text-slate-900">[KCA 필증]</strong> 전파법 기준에 맞는 최종 전파 합격증(KCA 필증) 서류를 확인하였는가?</td>
                </tr>
                <tr class="hover:bg-slate-50 transition-colors">
                    <td class="p-4 border-b border-slate-200 text-sm font-medium text-slate-800 leading-relaxed"><span class="inline-block bg-blue-100 text-blue-800 font-bold px-2 py-0.5 rounded text-xs mr-2">12. 관제·운영 도장</span><strong class="text-slate-900">[합격 날인]</strong> 현장 감리원, 종합관제실 및 운영사 담당자가 모여 합격 도장을 찍었는가?</td>
                </tr>
            </tbody>
        </table>
    </div>

</div>
{common_js}
</body>
</html>
"""

sub_dirs = {
    "표준서": std_html,
    "수행지침": gui_html,
    "체크리스트": chk_html
}

for s_n, content in sub_dirs.items():
    sp = os.path.join(target_folder, s_n)
    if os.path.exists(sp):
        for fn in os.listdir(sp):
            if fn.endswith('.html'):
                with open(os.path.join(sp, fn), 'w', encoding='utf-8') as f_out:
                    f_out.write(content)
                print(f"   ✓ [REBUILT WBS 9000-2-6 WITH EASY GUIDE & VISUAL SVG] {s_n} -> {fn}")

# Excel Sync for Row 6 (WBS 9000-2-6) with EASY TERMS
excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
if os.path.exists(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        if "통신분야" in wb.sheetnames:
            ws = wb["통신분야"]
            
            # Easy Summary Content for Row 6
            std_summary = "1) 종합관제실(OCC) 및 시교통관제 연동: 동탄 OCC 관제실 모니터 화면 배치 확정 및 교차로 4K CCTV 65개소 사각지대 제로 명당 배치\n2) 0.5초 방송/관제 연동 & 3초 비상 벨: 수원/화성 시교통관제 트램 우대 신호등 연동, 승강장 전광판·스피커·관제 0.5초 동시 띄우기, 비상 벨 3초 관제 소리 연결, 정부 합격증(TTA/KCA) 챙기기 & 관제·운영사 도장 찍기"
            gui_summary = "1) OCC 모니터 배치 & 4K 카메라 65개소: 관제실 모니터 화면 배치 및 4K 카메라 65개소 사각지대 제로 명당 배치\n2) 5단계 초직관 시공 가이드: ① OCC 모니터배치 ➔ ② 수원/화성 시교통 신호연동 ➔ ③ 전광판·스피커 0.5초 동시 띄우기 ➔ ④ 비상 벨 3초 관제 소리 연결 ➔ ⑤ 정부 전파 합격증(TTA/KCA) 챙기기 & 관제·운영사 도장 찍기 3단계 visual 가이드"
            chk_summary = "1) 동탄 OCC 관제실 모니터 화면을 배치하고 교차로 4K CCTV 카메라 65개소를 사각지대 없이 명당 자리에 배치하였는가?\n2) 수원/화성 시교통관제 신호 연동, 승강장 전광판·스피커 0.5초 동시 띄우기, 비상 벨 3초 관제 연결 및 정부 합격증(TTA/KCA)을 확인하고 관제·운영사 서명 도장을 찍었는가?"

            ws.cell(row=6, column=10, value=std_summary) # Col J
            ws.cell(row=6, column=12, value=gui_summary) # Col L
            ws.cell(row=6, column=14, value=chk_summary) # Col N

            wb.save(excel_path)
            print("   ✓ [EXCEL V4 SYNC COMPLETE] Successfully updated Row 6 (WBS 9000-2-6) with EASY TERMS in 매뉴얼 BODY (집행단계)v4.xlsx")
    except Exception as e:
        print(f"   ⚠️ Excel Sync Note: {e}")

print("\n🎉 SUCCESSFULLY REBUILT WBS 9000-2-6 WITH EASY GUIDE & HIGHLY VISUAL SVG DIAGRAMS!")
