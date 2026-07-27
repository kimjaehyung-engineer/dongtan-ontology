import os
import sys
import openpyxl
import shutil

sys.stdout.reconfigure(encoding='utf-8')

possible_paths = [
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx",
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\매뉴얼 BODY (집행단계)v4.xlsx",
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\매뉴얼 BODY (집행단계).xlsx"
]

excel_path = None
for p in possible_paths:
    if os.path.exists(p):
        excel_path = p
        break

if not excel_path:
    parent_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램"
    for fname in os.listdir(parent_dir):
        if fname.endswith('.xlsx') and '집행단계' in fname:
            excel_path = os.path.join(parent_dir, fname)
            break

print(f"Target Excel File: {excel_path}")

wb = openpyxl.load_workbook(excel_path)

# Find telecom sheet
target_sheet_name = None
for sname in wb.sheetnames:
    if '통신' in sname:
        target_sheet_name = sname
        break

if not target_sheet_name:
    target_sheet_name = wb.sheetnames[0]

sheet = wb[target_sheet_name]
print(f"Target Sheet: {target_sheet_name}")

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"
all_folders = os.listdir(base_dir)

updated_rows = 0

for r in range(2, 34):
    l4_code = sheet.cell(row=r, column=4).value or ""
    act_name = sheet.cell(row=r, column=5).value or ""
    ju_gwan = sheet.cell(row=r, column=6).value or ""
    purpose = sheet.cell(row=r, column=7).value or ""
    method = sheet.cell(row=r, column=8).value or ""
    deliverable = sheet.cell(row=r, column=9).value or ""

    if not act_name:
        continue

    act_clean = str(act_name).strip()
    idx = r - 1

    method_lines = [m.strip() for m in str(method).split('\n') if m.strip()]
    first_method = method_lines[0] if method_lines else f"{act_clean} 시방 수칙"
    first_method_clean = first_method.lstrip('0123456789.-) ').strip()
    
    if not first_method_clean.endswith('하였는가?'):
        if first_method_clean.endswith('함') or first_method_clean.endswith('임') or first_method_clean.endswith('다'):
            first_method_clean = first_method_clean[:-1] + '하였는가?'
        elif first_method_clean.endswith('확인') or first_method_clean.endswith('검토') or first_method_clean.endswith('시공') or first_method_clean.endswith('설치'):
            first_method_clean = first_method_clean + '하였는가?'
        else:
            first_method_clean = first_method_clean + ' 적정 여부를 확인하였는가?'

    std_summary = f"1) {act_clean} 시방 기준 및 {str(purpose)[:30]} 준수 수칙을 확인한다.\n2) {str(deliverable)[:30]} 산출물의 무결성을 검증하고 서명 체결한다."
    gui_summary = f"1) 사전 준비부터 시공/시험 및 마감 검사까지 2D Visual 절차에 따라 수행한다.\n2) 타 공종 인터페이스 및 8대 현장조건을 반영하여 미결사항을 즉시 해결한다."
    chk_summary = f"1) {first_method_clean}\n2) 시방서 요구조건 및 {str(deliverable)[:25]} 증빙 제출 여부를 점검하였는가?"

    sheet.cell(row=r, column=10).value = std_summary
    sheet.cell(row=r, column=12).value = gui_summary
    sheet.cell(row=r, column=14).value = chk_summary

    updated_rows += 1
    print(f"Row {r:02d} [{idx}] {act_clean:<30} -> J, L, N Columns Synced!")

# Try saving to main file, or output file if locked
out_path = excel_path.replace(".xlsx", "_통신연동완료.xlsx")
try:
    wb.save(excel_path)
    print(f"\n🎉 SUCCESSFULLY SAVED DIRECTLY TO ORIGINAL EXCEL V4 FILE:\n{excel_path}")
except PermissionError:
    wb.save(out_path)
    print(f"\n⚠️ ORIGINAL FILE IS CURRENTLY OPEN IN EXCEL. SAVED TO UPDATED FILE:\n{out_path}")
    print("💡 Please close the excel file and rename/overwrite if needed.")
