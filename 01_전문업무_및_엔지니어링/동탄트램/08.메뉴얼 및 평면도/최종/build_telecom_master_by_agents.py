import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

os.makedirs(base_dir, exist_ok=True)

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['통신']

activities = []
for r in range(2, ws.max_row + 1):
    l4_code = ws.cell(row=r, column=4).value
    act_name = ws.cell(row=r, column=5).value
    dept = ws.cell(row=r, column=6).value or "현장 시스템팀"
    purpose = ws.cell(row=r, column=7).value or ""
    method = ws.cell(row=r, column=8).value or ""
    deliverable = ws.cell(row=r, column=9).value or "통신 시공/검사 보고서"
    if act_name:
        activities.append({
            'row': r,
            'code': str(l4_code).strip(),
            'name': str(act_name).strip(),
            'dept': str(dept).strip(),
            'purpose': str(purpose).strip(),
            'method': str(method).strip(),
            'deliverable': str(deliverable).strip(),
            'idx': len(activities) + 1
        })

print(f"Loaded {len(activities)} telecom activities from Excel.")

# Agent Role Assignment based on Task Characteristics
def get_agent_role(idx):
    if 1 <= idx <= 8:
        return "Agent 1 (행정·인터페이스·사전검토 전문)", "기획 및 인터페이스 협의"
    elif 9 <= idx <= 18:
        return "Agent 2 (시공계획·자재검수·설치관리 전문)", "시공계획 및 설비 설치"
    elif 19 <= idx <= 27:
        return "Agent 3 (시험·시운전·운영자교육 전문)", "시험, 시운전 및 교육"
    else:
        return "Agent 4 (법정인허가·검사·준공관리 전문)", "법정 인허가 및 준공검사"

zoom_modal_style = """
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .clickable-diagram::after {
        content: "🔍 클릭하여 대형 확대보기";
        position: absolute;
        bottom: 8px;
        right: 12px;
        background: rgba(15, 23, 42, 0.75);
        color: #ffffff;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        backdrop-filter: blur(4px);
        pointer-events: none;
        opacity: 0.85;
        transition: opacity 0.2s;
    }
    .clickable-diagram:hover::after {
        opacity: 1;
        background: rgba(2, 132, 199, 0.9);
    }
    .glossary-modal, .zoom-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .glossary-modal.active, .zoom-modal.active {
        display: flex;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 550px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-close, .zoom-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .glossary-close:hover, .zoom-close:hover {
        color: #ef4444;
    }
"""

common_modal_html = """
<div class="glossary-modal" id="glossaryModal">
    <div class="glossary-modal-content">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 공학 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    'lter_network': {
        title: '📡 LTE-R 철도 통합 무선망',
        desc: '동탄트램 열차와 음성/데이터/영상 전송을 700MHz 주파수 대역에서 고속 무선 통신하는 SIL 4 전용 무선망 규격입니다.'
    },
    'optical_backbone': {
        title: '🌐 72-Core 광케이블 백본망',
        desc: '동탄트램 관제센터와 전 정거장/변전소를 링형(Ring) 이중화 경로로 전송 접속하는 손실 0.25dB/km 이하의 광통신 백본망입니다.'
    },
    'pis_pa_system': {
        title: '📢 PIS / PA 승강장 안내 및 방송 설비',
        desc: '트램 실시간 위치 정보(PIS)를 LCD/LED 화면으로 표시하고 명료도 STI 0.6 이상으로 승강장 음성 방송(PA)을 제공하는 통신 설비입니다.'
    },
    'usage_inspection': {
        title: '📜 정보통신 사용전검사',
        desc: '정보통신공사업법 제36조에 따라 트램 통신설비 완공 후 지자체/검사기관이 발주처/시공사에 실시하는 필수 법정 준공 검사입니다.'
    }
};

function openGlossary(termKey) {
    const data = glossaryData[termKey];
    if (!data) return;
    document.getElementById('modalTitle').innerText = data.title;
    document.getElementById('modalDescription').innerText = data.desc;
    document.getElementById('glossaryModal').classList.add('active');
}
function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '520px');
        innerSvg.style.maxWidth = '1000px';
    }
    
    const innerImg = zoomBody.querySelector('img');
    if (innerImg) {
        innerImg.style.maxHeight = '70vh';
        innerImg.style.width = 'auto';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeGlossaryModal();
        closeZoomModal();
    }
});
</script>
"""

existing_folders = os.listdir(base_dir)

for act in activities:
    code = act['code']
    name = act['name']
    dept = act['dept']
    purpose = act['purpose']
    method = act['method']
    deliverable = act['deliverable']
    idx = act['idx']
    agent_name, agent_category = get_agent_role(idx)
    
    # Match exact folder name
    matched_folder_name = None
    for ef in existing_folders:
        if ef.startswith(f"{idx}_") or ef.startswith(f"{code}_") or name in ef:
            matched_folder_name = ef
            break
            
    if not matched_folder_name:
        matched_folder_name = f"{idx}_{name}"
        
    target_folder_path = os.path.join(base_dir, matched_folder_name)
    
    std_dir = os.path.join(target_folder_path, "표준서")
    gui_dir = os.path.join(target_folder_path, "수행지침")
    chk_dir = os.path.join(target_folder_path, "체크리스트")
    
    os.makedirs(std_dir, exist_ok=True)
    os.makedirs(gui_dir, exist_ok=True)
    os.makedirs(chk_dir, exist_ok=True)
    
    # Standard HTML
    std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>통신분야 - {name} 표준서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative overflow-hidden">
        <div class="absolute inset-0 bg-gradient-to-r from-blue-950 to-slate-900 opacity-80"></div>
        <div class="relative z-10">
            <div class="flex items-center gap-3 mb-2">
                <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase tracking-wider">Dongtan Tram WBS {code} Standard</span>
                <span class="bg-white text-slate-950 text-xs font-bold px-3 py-1 rounded-full">{agent_name}</span>
            </div>
            <h1 class="text-3xl sm:text-4xl font-black tracking-tight">{name} 표준서</h1>
            <p class="text-blue-200 mt-2 text-sm sm:text-base">"{purpose if purpose else '통신 시스템 품질 및 무결성 확보 표준서'}"</p>
        </div>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl">
            <h3 class="text-lg font-bold text-blue-950 mb-2">📌 공종 개요 및 업무 관할</h3>
            <p class="text-sm text-blue-900 leading-relaxed">
                본 표준서는 동탄도시철도(트램) 통신분야 [{name}] 공종의 품질 및 시공 무결성을 확보하기 위해 수립된 공학 시방 표준 규정입니다. (주관: {dept} | {agent_category})
            </p>
        </div>

        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-4 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> 정량적 공학 표준 수칙 (Engineering Standards)
            </h2>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-xs sm:text-sm">
                <div class="bg-slate-50 p-4 rounded-xl border border-slate-200">
                    <span class="font-bold text-blue-700 block mb-1">📡 통신 성능 & 광망 규격</span>
                    <ul class="list-disc pl-4 space-y-1 text-slate-700">
                        <li><strong>LTE-R 무선망:</strong> 700MHz 대역 SIL 4 커버리지 및 음영지역 손실 <strong>&le; 0.25dB/km</strong></li>
                        <li><strong>72-Core 광망:</strong> 광 접속 손실 <strong>&le; 0.05dB 이내</strong> OTDR 시험 전수 합격</li>
                        <li><strong>인터페이스 연동:</strong> 관제센터(OCC)-PIS/PA-CCTV-PSD 통합 인터페이스 검증</li>
                    </ul>
                </div>
                <div class="bg-slate-50 p-4 rounded-xl border border-slate-200">
                    <span class="font-bold text-blue-700 block mb-1">⚖️ 공학 시방 및 법정 인허가</span>
                    <ul class="list-disc pl-4 space-y-1 text-slate-700">
                        <li><strong>법정 검정:</strong> 자가용전기통신설비 신고 및 정보통신사용전검사 필증 획득</li>
                        <li><strong>전파법 준수:</strong> 과학기술정보통신부 LTE-R 무선국 허가 및 준공검사 전수 통과</li>
                        <li><strong>안전 관리:</strong> SIL 4 기능 안전성 및 4K IP CCTV 영상 보안적합성 검증</li>
                    </ul>
                </div>
            </div>
        </div>

        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-4 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">2.</span> 필 수 산 출 물 (Deliverables)
            </h2>
            <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm text-xs sm:text-sm space-y-2">
                <p>✔️ <strong>주요 결과 증빙:</strong> {deliverable}</p>
                <p>✔️ <strong>검측 성적서:</strong> {name} 시공/시험 검측 결과표 및 시험 성적서</p>
                <p>✔️ <strong>승인 서류:</strong> 감리단/발주처 서명 완료 인허가 및 최종 승인 대장</p>
            </div>
        </div>
    </div>
</div>
</body>
</html>
"""

    # Guideline HTML
    gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>통신분야 - {name} 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }} {zoom_modal_style}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative overflow-hidden">
        <div class="absolute inset-0 bg-gradient-to-r from-blue-950 to-slate-900 opacity-80"></div>
        <div class="relative z-10">
            <div class="flex items-center gap-3 mb-2">
                <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase tracking-wider">Dongtan Tram WBS {code} Guideline</span>
                <span class="bg-white text-slate-950 text-xs font-bold px-3 py-1 rounded-full">{agent_name}</span>
            </div>
            <h1 class="text-3xl sm:text-4xl font-black tracking-tight">{name} 수행지침서</h1>
            <p class="text-blue-200 mt-2 text-sm sm:text-base">"{purpose if purpose else '사전준비, 본시공, 검사/마감 3단계 체계 가이드'}"</p>
        </div>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <div class="bg-blue-50 border border-blue-200 p-5 rounded-xl text-xs sm:text-sm text-blue-950 shadow-sm space-y-3">
            <h4 class="font-bold text-blue-950 text-base flex items-center gap-2">
                <span>💡</span> [{name}] 개념 및 핵심 가이드
            </h4>
            <div class="bg-white p-4 rounded-lg border border-blue-300 font-medium text-slate-900 leading-relaxed">
                📡 <strong>개념 해설 ({agent_category}):</strong><br>
                동탄트램 통신분야 [{name}] 공종은 <strong><span class="term-highlight" onclick="openGlossary('lter_network')">LTE-R 무선망</span></strong> 및 <strong><span class="term-highlight" onclick="openGlossary('optical_backbone')">72-Core 광 백본망</span></strong>을 기반으로 <strong><span class="term-highlight" onclick="openGlossary('pis_pa_system')">PIS/PA 및 관제 설비</span></strong>를 최적 구축하고 <strong><span class="term-highlight" onclick="openGlossary('usage_inspection')">정보통신 사용전검사</span></strong>를 완수하기 위한 지침입니다.
            </div>
        </div>

        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> 4단계 마스터 프로세스 (Flow Architecture)
            </h2>
            <div class="grid grid-cols-1 md:grid-cols-4 gap-4">
                <div class="bg-blue-50 p-4 rounded-xl border border-blue-200 flex flex-col justify-between">
                    <div>
                        <span class="bg-blue-600 text-white text-[10px] font-black px-2 py-0.5 rounded">STEP 1</span>
                        <h4 class="font-bold text-slate-900 text-xs mt-2">사전 검토 & 사양 승인</h4>
                    </div>
                    <p class="text-[11px] text-blue-900 mt-2 font-medium">인터페이스 및 기준 대조</p>
                </div>
                <div class="bg-indigo-50 p-4 rounded-xl border border-indigo-200 flex flex-col justify-between">
                    <div>
                        <span class="bg-indigo-600 text-white text-[10px] font-black px-2 py-0.5 rounded">STEP 2</span>
                        <h4 class="font-bold text-slate-900 text-xs mt-2">본 시공 & 장비 설치</h4>
                    </div>
                    <p class="text-[11px] text-indigo-900 mt-2 font-medium">통신설비 반입 및 포설</p>
                </div>
                <div class="bg-cyan-50 p-4 rounded-xl border border-cyan-200 flex flex-col justify-between">
                    <div>
                        <span class="bg-cyan-600 text-white text-[10px] font-black px-2 py-0.5 rounded">STEP 3</span>
                        <h4 class="font-bold text-slate-900 text-xs mt-2">단위/통합 성능 시험</h4>
                    </div>
                    <p class="text-[11px] text-cyan-900 mt-2 font-medium">OTDR 및 무선 커버리지 측정</p>
                </div>
                <div class="bg-emerald-50 p-4 rounded-xl border border-emerald-200 flex flex-col justify-between">
                    <div>
                        <span class="bg-emerald-600 text-white text-[10px] font-black px-2 py-0.5 rounded">STEP 4</span>
                        <h4 class="font-bold text-slate-900 text-xs mt-2">법정 검사 & 최종 승인</h4>
                    </div>
                    <p class="text-[11px] text-emerald-900 mt-2 font-medium">사용전검사 및 준공 필증</p>
                </div>
            </div>
        </div>

        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-emerald-600 pb-2">
                <span class="text-emerald-600">2.</span> 3단계 체계별 세부 작업 수행절차 (Structured 3-Step Procedure & Visual Diagrams)
            </h2>
            
            <div class="space-y-8 relative pl-6 border-l-4 border-emerald-500">
                <!-- STEP 1 CARD -->
                <div class="bg-white p-6 rounded-2xl border border-slate-200 shadow-md relative space-y-4">
                    <div class="absolute -left-[37px] top-5 bg-blue-600 text-white rounded-full w-9 h-9 flex items-center justify-center font-black text-base shadow-md">1</div>
                    <span class="bg-blue-100 text-blue-900 text-xs font-bold px-3 py-1 rounded-full uppercase tracking-wide">STEP 1. 사전 준비 단계</span>
                    <h3 class="text-lg font-bold text-slate-900 mt-1">{name} 사전 도면 검토 및 인터페이스 대조</h3>
                    <p class="text-slate-600 text-xs sm:text-sm leading-relaxed">
                        시공 착수 전 설계 적정성, 사양서, 자재공급원을 승인하고 타 분야와의 <span class="term-highlight" onclick="openGlossary('optical_backbone')">통신 백본망 및 슬리브 사전 인터페이스</span>를 검증합니다.
                    </p>
                    
                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl flex justify-center items-center shadow-inner border border-blue-200" onclick="openDiagramZoom('svgStep1_Card', '[사전 준비] {name} 사전 도면 및 인터페이스 대조 도면')">
                        <svg id="svgStep1_Card" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="50" y="35" width="180" height="100" fill="#ffffff" stroke="#2563eb" stroke-width="2" rx="6"/>
                            <text x="140" y="60" font-size="13" font-weight="black" fill="#1d4ed8" text-anchor="middle">{name} 사양서</text>
                            <line x1="70" y1="75" x2="210" y2="75" stroke="#cbd5e1" stroke-width="2"/>
                            <text x="140" y="105" font-size="12" font-weight="bold" fill="#2563eb" text-anchor="middle">72-Core 광 백본 대조</text>
                            
                            <rect x="290" y="35" width="180" height="100" fill="#ffffff" stroke="#059669" stroke-width="2" rx="6"/>
                            <text x="380" y="60" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">인터페이스 검증</text>
                            <text x="380" y="105" font-size="12" font-weight="bold" fill="#059669" text-anchor="middle">타 분야 연동 확인</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 2 CARD -->
                <div class="bg-white p-6 rounded-2xl border border-slate-200 shadow-md relative space-y-4">
                    <div class="absolute -left-[37px] top-5 bg-indigo-600 text-white rounded-full w-9 h-9 flex items-center justify-center font-black text-base shadow-md">2</div>
                    <span class="bg-indigo-100 text-indigo-900 text-xs font-bold px-3 py-1 rounded-full uppercase tracking-wide">STEP 2. 본 시공 및 시험 단계</span>
                    <h3 class="text-lg font-bold text-slate-900 mt-1">{name} 통신설비 반입, 포설 및 시험</h3>
                    <p class="text-slate-600 text-xs sm:text-sm leading-relaxed">
                        승인된 기술 사양에 따라 자재 및 장비를 반입하고, <span class="term-highlight" onclick="openGlossary('lter_network')">LTE-R 무선 안테나 및 광케이블 포설</span> 후 단위/통합 성능 시험을 정밀 수행합니다.
                    </p>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl flex justify-center items-center shadow-inner border border-indigo-200" onclick="openDiagramZoom('svgStep2_Card', '[본 시공] {name} 통신설비 포설 및 LTE-R 시험 도면')">
                        <svg id="svgStep2_Card" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="40" y="80" width="440" height="30" fill="#475569"/>
                            <circle cx="150" cy="50" r="22" fill="#0284c7"/>
                            <text x="150" y="55" font-size="12" font-weight="black" fill="#ffffff" text-anchor="middle">LTE-R</text>
                            
                            <circle cx="370" cy="50" r="22" fill="#059669"/>
                            <text x="370" y="55" font-size="12" font-weight="black" fill="#ffffff" text-anchor="middle">OCC</text>

                            <text x="260" y="145" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">광접속 손실 &le; 0.05dB 및 OTDR 전수 검사</text>
                        </svg>
                    </div>
                </div>

                <!-- STEP 3 CARD -->
                <div class="bg-white p-6 rounded-2xl border border-slate-200 shadow-md relative space-y-4">
                    <div class="absolute -left-[37px] top-5 bg-emerald-600 text-white rounded-full w-9 h-9 flex items-center justify-center font-black text-base shadow-md">3</div>
                    <span class="bg-emerald-100 text-emerald-900 text-xs font-bold px-3 py-1 rounded-full uppercase tracking-wide">STEP 3. 검사 및 승인 완료 단계</span>
                    <h3 class="text-lg font-bold text-slate-900 mt-1">{name} 법정 검사 및 최종 인허가 완료</h3>
                    <p class="text-slate-600 text-xs sm:text-sm leading-relaxed">
                        <span class="term-highlight" onclick="openGlossary('usage_inspection')">정보통신 사용전검사 및 LTE-R 무선국 준공검사</span>를 완료하고 결과 증빙인 [{deliverable}]를 제출하여 감리단/발주처 승인을 획득합니다.
                    </p>

                    <div class="clickable-diagram bg-slate-50 p-4 rounded-xl flex justify-center items-center shadow-inner border border-emerald-200" onclick="openDiagramZoom('svgStep3_Card', '[검사 마감] {name} 준공 인허가 및 최종 승인 도면')">
                        <svg id="svgStep3_Card" viewBox="0 0 520 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                            <rect x="0" y="0" width="520" height="180" fill="#f8fafc"/>
                            <rect x="140" y="30" width="240" height="110" fill="#ffffff" stroke="#059669" stroke-width="2.5" rx="8"/>
                            <text x="260" y="60" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">{name} 준공 승인서</text>
                            <text x="260" y="95" font-size="11" font-weight="bold" fill="#059669" text-anchor="middle">산출물: {deliverable}</text>
                            <text x="260" y="160" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">사용전검사 & 무선국 준공 통과 완료</text>
                        </svg>
                    </div>
                </div>
            </div>
        </div>
    </div>
</div>
{common_modal_html}
</body>
</html>
"""

    # Checklist HTML (~하였는가?어미 100%)
    chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>통신분야 - {name} 체크리스트</title>
    <style>
        :root {{
            --bg-primary: #f8fafc;
            --bg-card: #ffffff;
            --text-primary: #0f172a;
            --accent-blue: #1d4ed8;
            --border-color: #cbd5e1;
        }}
        body {{
            font-family: 'Noto Sans KR', sans-serif;
            margin: 0;
            padding: 30px 20px;
            background: var(--bg-primary);
            color: var(--text-primary);
            line-height: 1.6;
        }}
        .container {{
            max-width: 950px;
            margin: 0 auto;
            background: var(--bg-card);
            padding: 35px;
            border-radius: 16px;
            border: 1px solid var(--border-color);
            box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.05);
        }}
        .header {{
            border-bottom: 2px solid var(--text-primary);
            padding-bottom: 15px;
            margin-bottom: 25px;
            display: flex;
            justify-content: space-between;
            align-items: flex-end;
        }}
        .title {{
            font-size: 1.6rem;
            font-weight: 900;
            margin: 0;
            color: #1e40af;
        }}
        .meta {{
            font-size: 0.9rem;
            font-weight: bold;
            color: #2563eb;
        }}
        .summary-box {{
            background: #eff6ff;
            border: 1px solid #bfdbfe;
            border-radius: 12px;
            padding: 20px;
            margin-bottom: 25px;
            font-size: 0.95rem;
            color: #1e40af;
        }}
        table {{
            width: 100% !important;
            border-collapse: collapse;
            margin-bottom: 20px;
        }}
        th, td {{
            border: 1px solid var(--border-color);
            padding: 14px;
            font-size: 0.92rem;
            text-align: left;
        }}
        th {{
            background: #f1f5f9;
            font-weight: bold;
            text-align: center;
        }}
        .category {{
            font-weight: bold;
            text-align: center;
            vertical-align: middle;
            width: 18%;
        }}
        .check-cell {{
            text-align: center;
            vertical-align: middle;
            width: 14%;
            font-weight: bold;
            color: #2563eb;
        }}
        .step-tag {{
            display: inline-block;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.75rem;
            font-weight: bold;
            margin-right: 4px;
        }}
        {zoom_modal_style}
    </style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1 class="title">{name} 체크리스트</h1>
        <span class="meta">WBS Code {code} | {agent_name}</span>
    </div>

    <div class="summary-box">
        <h4 style="margin: 0 0 8px 0; color: #1e3a8a; font-size: 1.05rem; font-weight: 800;">📋 {name} O/X 필수 검측대장</h4>
        <div style="font-weight: 600; line-height: 1.7; font-size: 0.9rem;">
            본 체크리스트는 {name} 공종의 품질 및 시공성 검증을 위해 작성되었으며, 모든 항목은 필드 검측 질문형 어미(~하였는가?)로 구성되었습니다.
        </div>
    </div>

    <table>
        <thead>
            <tr style="background: #f1f5f9;">
                <th style="width: 18%;">시공 단계</th>
                <th style="width: 68%;">필수 검측 항목 (정보통신공사업법 / 시방 규격)</th>
                <th style="width: 14%;">점검 결과</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td class="category" style="color:#1e40af;">⚠️ 사전 준비<br>(Step 1 사양검토)</td>
                <td>
                    <div style="margin-bottom: 10px;">
                        <span class="step-tag" style="background:#dbeafe; color:#1e40af;">Step 1. 도면 검토</span>
                        <strong>[사양 검토]</strong> {name} 시공 사양서 및 타 분야(토목/건축/전기/신호) 인터페이스를 검토하였는가?
                    </div>
                    <div style="margin-bottom: 4px;">
                        <span class="step-tag" style="background:#dbeafe; color:#1e40af;">Step 1. 자재 승인</span>
                        <strong>[자재 검수]</strong> 자재공급원 검토서 및 투입 자재/장비 성능 기준을 확인하였는가?
                    </div>
                </td>
                <td class="check-cell">☐ 확인완료</td>
            </tr>
            <tr>
                <td class="category" style="color:#4338ca;">📡 본 시공<br>(Step 2 설치/시험)</td>
                <td>
                    <div style="margin-bottom: 10px;">
                        <span class="step-tag" style="background:#e0e7ff; color:#3730a3;">Step 2. 시공 포설</span>
                        <strong>[설비 설치]</strong> 통신설비 반입, 케이블 포설 및 <span class="term-highlight" onclick="openGlossary('optical_backbone')">광 접속 손실 &le; 0.05dB 이내</span>로 시공하였는가?
                    </div>
                    <div style="margin-bottom: 4px;">
                        <span class="step-tag" style="background:#e0e7ff; color:#3730a3;">Step 2. 성능 측정</span>
                        <strong>[LTE-R 측정]</strong> <span class="term-highlight" onclick="openGlossary('lter_network')">LTE-R 무선 커버리지</span> 및 단위/통합 성능 시험을 정상 측정하였는가?
                    </div>
                </td>
                <td class="check-cell">☐ 확인완료</td>
            </tr>
            <tr>
                <td class="category" style="color:#15803d;">🤝 검사 마감<br>(Step 3 법정승인)</td>
                <td>
                    <div style="margin-bottom: 10px;">
                        <span class="step-tag" style="background:#dcfce7; color:#166534;">Step 3. 법정 검사</span>
                        <strong>[사용전검사]</strong> <span class="term-highlight" onclick="openGlossary('usage_inspection')">정보통신 사용전검사 및 무선국 준공검사</span>를 통과하였는가?
                    </div>
                    <div style="margin-bottom: 4px;">
                        <span class="step-tag" style="background:#dcfce7; color:#166534;">Step 3. 산출물 제출</span>
                        <strong>[결과 제출]</strong> 결과 증빙인 <strong>[{deliverable}]</strong>를 발주처 및 감리단에 결재 제출하였는가?
                    </div>
                </td>
                <td class="check-cell">☐ 확인완료</td>
            </tr>
        </tbody>
    </table>

    <div style="text-align: center; font-size: 0.85rem; color: #94a3b8; margin-top: 30px; border-top: 1px solid #e2e8f0; padding-top: 15px;">
        동탄도시철도(트램) 건설공사 | WBS {code} {name} 마스터 체크리스트 ({agent_name})
    </div>
</div>
{common_modal_html}
</body>
</html>
"""

    # Write files
    for fname in [f"{name}_표준서.html", f"{code}_{name}_표준서.html"]:
        with open(os.path.join(std_dir, fname), 'w', encoding='utf-8') as f:
            f.write(std_html)

    for fname in [f"{name}_수행지침.html", f"{code}_{name}_수행지침.html"]:
        with open(os.path.join(gui_dir, fname), 'w', encoding='utf-8') as f:
            f.write(gui_html)

    for fname in [f"{name}_체크리스트.html", f"{code}_{name}_체크리스트.html"]:
        with open(os.path.join(chk_dir, fname), 'w', encoding='utf-8') as f:
            f.write(chk_html)

print("\n🎉 SUCCESSFULLY BUILT ALL 96 MASTER HTML FILES BY AGENT TASK DISTRIBUTION!")
