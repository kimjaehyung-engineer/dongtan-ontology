import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# Find excel file in workspace
possible_paths = [
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\매뉴얼 BODY (집행단계)v4.xlsx",
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\매뉴얼 BODY (집행단계).xlsx",
    r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
]

excel_path = None
for p in possible_paths:
    if os.path.exists(p):
        excel_path = p
        break

if not excel_path:
    parent_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램"
    for fname in os.listdir(parent_dir):
        if fname.endswith('.xlsx') and '집행단계' in fname:
            excel_path = os.path.join(parent_dir, fname)
            break

print(f"Found Excel File: {excel_path}")

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)\통신분야"

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Robust sheet selection by scanning sheet names
target_sheet_name = None
for sname in wb.sheetnames:
    if '통신' in sname:
        target_sheet_name = sname
        break

if not target_sheet_name:
    target_sheet_name = wb.sheetnames[0]

sheet = wb[target_sheet_name]
print(f"Selected Worksheet: {target_sheet_name}")

print("=== STARTING AUTOMATED SEQUENTIAL ENHANCEMENT FROM ROW 4 TO ROW 33 (ROW 4 START BY USER DIRECTIVE) ===")

zoom_modal_style = """
    .term-highlight {
        color: #0284c7 !important;
        font-weight: 700 !important;
        border-bottom: 2px dashed #0284c7 !important;
        cursor: pointer !important;
        transition: all 0.2s ease !important;
        padding: 0 2px !important;
    }
    .term-highlight:hover {
        background: #e0f2fe !important;
        color: #0369a1 !important;
        border-radius: 4px !important;
    }
    .clickable-diagram {
        cursor: zoom-in !important;
        transition: all 0.25s ease !important;
        position: relative !important;
    }
    .clickable-diagram:hover {
        transform: scale(1.015) !important;
        box-shadow: 0 12px 25px -5px rgba(0, 0, 0, 0.15) !important;
    }
    .clickable-diagram::after {
        content: "🔍 클릭하여 대형 확대보기";
        position: absolute;
        bottom: 8px;
        right: 12px;
        background: rgba(15, 23, 42, 0.75);
        color: #ffffff;
        font-size: 11px;
        font-weight: 700;
        padding: 4px 10px;
        border-radius: 20px;
        backdrop-filter: blur(4px);
        pointer-events: none;
        opacity: 0.85;
        transition: opacity 0.2s;
    }
    .clickable-diagram:hover::after {
        opacity: 1;
        background: rgba(2, 132, 199, 0.9);
    }
    .glossary-modal, .zoom-modal {
        display: none;
        position: fixed;
        z-index: 9999;
        left: 0;
        top: 0;
        width: 100%;
        height: 100%;
        overflow: auto;
        background-color: rgba(15, 23, 42, 0.75);
        backdrop-filter: blur(6px);
        align-items: center;
        justify-content: center;
    }
    .glossary-modal.active, .zoom-modal.active {
        display: flex;
    }
    .glossary-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 24px;
        border: 1px solid #e2e8f0;
        width: 90%;
        max-width: 550px;
        border-radius: 16px;
        box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1);
        position: relative;
        text-align: left;
    }
    .zoom-modal-content {
        background-color: #ffffff;
        margin: auto;
        padding: 28px;
        border: 1px solid #cbd5e1;
        width: 95%;
        max-width: 1100px;
        max-height: 90vh;
        border-radius: 20px;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25);
        position: relative;
        overflow-y: auto;
        text-align: center;
    }
    .glossary-close, .zoom-close {
        color: #64748b;
        position: absolute;
        right: 20px;
        top: 16px;
        font-size: 32px;
        font-weight: bold;
        cursor: pointer;
        transition: color 0.2s;
    }
    .glossary-close:hover, .zoom-close:hover {
        color: #ef4444;
    }
"""

common_js = """
<div class="glossary-modal" id="glossaryModal">
    <div class="glossary-modal-content">
        <span class="glossary-close" onclick="closeGlossaryModal()">&times;</span>
        <h3 id="modalTitle" style="font-size: 1.25rem; font-weight: 800; color: #1e3a8a; margin-top: 0; margin-bottom: 12px; border-bottom: 2px solid #e2e8f0; padding-bottom: 8px;">용어 및 기술 해설</h3>
        <div class="modal-body">
            <p id="modalDescription" style="font-size: 0.95rem; color: #334155; line-height: 1.7; margin: 0; word-break: keep-all;"></p>
        </div>
    </div>
</div>

<div class="zoom-modal" id="zoomModal" onclick="closeZoomModalOutside(event)">
    <div class="zoom-modal-content" onclick="event.stopPropagation()">
        <span class="zoom-close" onclick="closeZoomModal()">&times;</span>
        <h3 id="zoomTitle" style="font-size: 1.35rem; font-weight: 900; color: #0f172a; margin-top: 0; margin-bottom: 16px; border-bottom: 2px solid #38bdf8; padding-bottom: 10px; text-align: left;">🔍 도식 대형 고화질 정밀 보기</h3>
        <div id="zoomBody" class="bg-slate-50 p-6 rounded-xl border border-slate-200 shadow-inner flex justify-center items-center overflow-auto min-h-[400px]">
        </div>
        <div style="margin-top: 14px; text-align: right; font-size: 0.85rem; font-weight: 700; color: #64748b;">
            💡 팁: ESC 키를 누르시거나 닫기(×) 버튼을 누르면 이전 화면으로 복귀합니다.
        </div>
    </div>
</div>

<script>
const glossaryData = {
    'lter_spec': {
        title: '📡 LTE-R 무선 통신 규격',
        desc: '동탄트램 700MHz 대역 전자기파 환경에서 열차 차상 및 지상 간 SIL 4 등급 통신 수신 레벨 ≥ -95dBm 확보 및 음영 Zero화를 달성하는 트램 전용 무선망 규격입니다.'
    },
    'optical_spec': {
        title: '🌐 72-Core 광 전송 백본망',
        desc: '차량기지 OCC 통합관제센터 및 정거장 18개소를 ring 이중화 구조로 연결하는 72-Core 싱글모드 광전송 케이블 망입니다.'
    },
    'checklist_spec': {
        title: '✅ 검측 무결성 검증 수칙',
        desc: '현장 감리단 및 감리원이 시공 무결성을 확보하기 위해 모든 검측 체크리스트 질문항목을 "~하였는가?" 질문형으로 검측 및 판정하는 기준입니다.'
    }
};

function openGlossary(termKey) {
    const data = glossaryData[termKey] || { title: '📌 기술 해설', desc: '동탄트램 통신공사 표준시방 기준에 준하여 검측을 수행합니다.' };
    document.getElementById('modalTitle').innerText = data.title;
    document.getElementById('modalDescription').innerText = data.desc;
    document.getElementById('glossaryModal').classList.add('active');
}
function closeGlossaryModal() {
    document.getElementById('glossaryModal').classList.remove('active');
}

function openDiagramZoom(elementId, titleText) {
    const srcEl = document.getElementById(elementId);
    if (!srcEl) return;
    
    const zoomBody = document.getElementById('zoomBody');
    document.getElementById('zoomTitle').innerText = "🔍 " + (titleText || "도식 대형 정밀 보기");
    
    zoomBody.innerHTML = srcEl.outerHTML;
    
    const innerSvg = zoomBody.querySelector('svg');
    if (innerSvg) {
        innerSvg.setAttribute('width', '100%');
        innerSvg.setAttribute('height', '550px');
        innerSvg.style.maxWidth = '1050px';
    }
    
    document.getElementById('zoomModal').classList.add('active');
}

function closeZoomModal() {
    document.getElementById('zoomModal').classList.remove('active');
}

function closeZoomModalOutside(event) {
    if (event.target.id === 'zoomModal') {
        closeZoomModal();
    }
}

window.addEventListener('keydown', function(event) {
    if (event.key === 'Escape') {
        closeGlossaryModal();
        closeZoomModal();
    }
});
</script>
"""

all_folders = os.listdir(base_dir)
processed_count = 0

# START FROM ROW 4 BY USER DIRECTIVE
for r in range(4, 34):
    l4_code = sheet.cell(row=r, column=4).value or "" # D열 L4 Code
    act_name = sheet.cell(row=r, column=5).value or "" # E열 Activity
    ju_gwan = sheet.cell(row=r, column=6).value or "" # F열 주관
    purpose = sheet.cell(row=r, column=7).value or "" # G열 목적
    method = sheet.cell(row=r, column=8).value or "" # H열 방법
    deliverable = sheet.cell(row=r, column=9).value or "" # I열 산출물

    if not act_name:
        continue

    act_clean = str(act_name).strip()
    idx = r - 1

    # Find matching directory
    target_folder = None
    for f in all_folders:
        if f.startswith(f"{idx}_") or f.endswith(act_clean) or act_clean in f:
            target_folder = os.path.join(base_dir, f)
            break
    
    if not target_folder:
        for f in all_folders:
            if f.startswith(f"{idx}_"):
                target_folder = os.path.join(base_dir, f)
                break

    if not target_folder:
        print(f"⚠️ Warning: Could not find folder for Row {r} ({act_clean})")
        continue

    print(f"🔄 Processing Row {r}/{33}: [{idx}] {act_clean} -> {os.path.basename(target_folder)}")

    method_lines = [m.strip() for m in str(method).split('\n') if m.strip()]
    if not method_lines:
        method_lines = [f"{act_clean} 관련 현장 시공 및 시방 수칙 준수여부 검측"]

    checklist_items = []
    for m_line in method_lines:
        line_clean = m_line.lstrip('0123456789.-) ').strip()
        if not line_clean.endswith('하였는가?'):
            if line_clean.endswith('함') or line_clean.endswith('임') or line_clean.endswith('다'):
                line_clean = line_clean[:-1] + '하였는가?'
            elif line_clean.endswith('확인') or line_clean.endswith('검토') or line_clean.endswith('시공') or line_clean.endswith('설치'):
                line_clean = line_clean + '하였는가?'
            else:
                line_clean = line_clean + ' 적정 여부를 확인하였는가?'
        checklist_items.append(line_clean)

    extra_checks = [
        "입찰안내서 및 시방서 요구조건과의 1:1 일치 여부를 검측하였는가?",
        "KRSA 철도표준자재 및 KS/KC 공인 시험성적서 첨부 여부를 확인하였는가?",
        "타 공종(토목/건축/전기/신호/차량) 8대 인터페이스 상호 간섭 여부를 점검하였는가?",
        "안전관리 계획 및 감리원 실시간 공무 서명 체결 여부를 확인하였는가?"
    ]
    for ec in extra_checks:
        if len(checklist_items) < 12:
            checklist_items.append(ec)

    step_count = max(3, min(5, len(method_lines)))

    # STANDARD HTML
    std_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - {act_clean} 표준서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Standard</span>
        <h1 class="text-3xl font-black mt-2">{act_clean} 표준서</h1>
        <p class="text-blue-200 text-sm mt-1">L4 Code: {l4_code} | 주관: {ju_gwan}</p>
    </div>
    
    <div class="p-8 space-y-8">
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-blue-950 mb-2">🎯 표준 목적 (Objective)</h3>
            <p class="text-slate-700 text-sm font-medium leading-relaxed">{purpose}</p>
        </div>

        <div>
            <h3 class="text-lg font-bold text-slate-900 mb-4 border-b-2 border-blue-600 pb-2">📜 시공 및 품질 검토 시방 수칙 (Methodology)</h3>
            <ul class="space-y-3">
"""
    for ml in method_lines:
        std_html += f"""                <li class="bg-slate-50 p-4 rounded-xl border border-slate-200 flex items-start gap-3">
                    <span class="bg-blue-600 text-white text-xs font-bold px-2.5 py-1 rounded mt-0.5">수칙</span>
                    <span class="text-slate-800 text-sm font-medium leading-relaxed">{ml}</span>
                </li>\n"""
    std_html += f"""            </ul>
        </div>

        <div class="bg-emerald-50 border border-emerald-200 p-6 rounded-xl">
            <h3 class="text-base font-bold text-emerald-950 mb-2">📦 증빙 산출물 (Deliverables)</h3>
            <p class="text-emerald-900 text-sm font-bold">{deliverable}</p>
        </div>
    </div>
</div>
</body>
</html>
"""

    # GUIDELINE HTML
    gui_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - {act_clean} 수행지침서</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>
        body {{ font-family: 'Noto Sans KR', sans-serif; }}
        {zoom_modal_style}
    </style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8 relative">
        <span class="bg-blue-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Guideline</span>
        <h1 class="text-3xl font-black mt-2">{act_clean} 유연 {step_count}단계 수행지침서</h1>
        <p class="text-blue-200 text-sm mt-1">"주관: {ju_gwan} | 현장 실무 맞춤형 2D Visual 마스터 수행 매뉴얼"</p>
    </div>
    
    <div class="p-6 sm:p-10 space-y-10">
        <div class="bg-blue-50 border border-blue-200 p-6 rounded-2xl text-sm text-blue-950 space-y-2">
            <h4 class="font-bold text-base flex items-center gap-2">💡 {act_clean} 실무 핵심 지침</h4>
            <p class="bg-white p-4 rounded-xl border border-blue-300 font-medium text-slate-900 leading-relaxed">
                {purpose} 본 지침서는 공종 특성에 준하여 <strong>{step_count}단계(Step 1~{step_count}) 마스터 프로세스</strong>로 구체화되었습니다.
            </p>
        </div>

        <div class="bg-blue-50/70 border border-blue-200 p-7 rounded-2xl shadow-md space-y-6">
            <div class="border-b border-blue-200 pb-4">
                <span class="bg-blue-600 text-white text-xs font-black px-3 py-1 rounded-full uppercase">SPECIAL FOCUS</span>
                <h3 class="text-xl font-black text-blue-950 mt-2">📋 {act_clean} 핵심 시공 및 검측 요약</h3>
            </div>
            <div class="grid grid-cols-1 md:grid-cols-2 gap-4 text-sm">
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>🎯</span> 주요 목적 및 시방</span>
                    <p class="text-slate-700 leading-relaxed text-xs">{purpose}</p>
                </div>
                <div class="bg-white p-5 rounded-xl border border-slate-200 shadow-sm space-y-2">
                    <span class="font-bold text-blue-700 flex items-center gap-2"><span>📦</span> 최종 증빙 산출물</span>
                    <p class="text-slate-700 leading-relaxed text-xs font-bold">{deliverable}</p>
                </div>
            </div>
        </div>

        <div>
            <h2 class="text-xl font-bold text-slate-900 mb-6 flex items-center gap-2 border-b-2 border-blue-600 pb-2">
                <span class="text-blue-600">1.</span> {step_count}단계 수행 마스터 프로세스 (Flexible {step_count}-Step Architecture)
            </h2>
            <div class="grid grid-cols-1 sm:grid-cols-2 md:grid-cols-{step_count} gap-3">
"""
    for s_idx in range(step_count):
        step_title = method_lines[s_idx] if s_idx < len(method_lines) else f"단계 {s_idx+1} 마감 및 결재"
        gui_html += f"""                <div class="bg-blue-50 p-4 rounded-xl border border-blue-200 flex flex-col justify-between">
                    <span class="bg-blue-600 text-white text-[10px] font-black px-2 py-0.5 rounded w-fit">STEP {s_idx+1}</span>
                    <h4 class="font-bold text-slate-900 text-xs mt-2">{step_title[:18]}...</h4>
                    <p class="text-[11px] text-blue-900 mt-1 font-medium">• 정밀 시공 및 검측</p>
                </div>\n"""
    gui_html += f"""            </div>
        </div>

        <div class="space-y-6">
            <h2 class="text-xl font-bold text-slate-900 mb-4 border-b-2 border-emerald-600 pb-2">
                <span class="text-emerald-600">2.</span> 2D Visual 기술 도식 (Enriched 2D SVG)
            </h2>
            <div class="clickable-diagram bg-slate-50 p-5 rounded-2xl border border-slate-200 shadow-inner" onclick="openDiagramZoom('svg_{idx}', '[{act_clean}] 시공 및 검측 2D visual 도식')">
                <svg id="svg_{idx}" viewBox="0 0 550 180" width="100%" height="180" xmlns="http://www.w3.org/2000/svg">
                    <rect x="0" y="0" width="550" height="180" fill="#f8fafc"/>
                    <rect x="30" y="20" width="220" height="120" fill="#ffffff" stroke="#0284c7" stroke-width="2" rx="8"/>
                    <text x="140" y="45" font-size="13" font-weight="black" fill="#0369a1" text-anchor="middle">📡 {act_clean[:14]}</text>
                    <text x="50" y="75" font-size="11" font-weight="bold" fill="#334155">• 시방 기준 100% 준수</text>
                    <text x="50" y="98" font-size="11" font-weight="bold" fill="#334155">• 인터페이스 검증 완료</text>
                    
                    <path d="M 260 80 L 290 80" stroke="#0284c7" stroke-width="3"/>
                    <polygon points="290,75 300,80 290,85" fill="#0284c7"/>
                    
                    <rect x="305" y="20" width="215" height="120" fill="#ffffff" stroke="#059669" stroke-width="2" rx="8"/>
                    <text x="412" y="45" font-size="13" font-weight="black" fill="#047857" text-anchor="middle">📋 {deliverable[:14]}</text>
                    <text x="325" y="75" font-size="11" font-weight="bold" fill="#334155">• 검측 무결성 서명 체결</text>
                    <text x="325" y="98" font-size="11" font-weight="bold" fill="#334155">• 최종 승인보고서 작성</text>
                    <text x="275" y="162" font-size="13" font-weight="black" fill="#0f172a" text-anchor="middle">{act_clean} 표준 시공 & 산출물 무결성 확보</text>
                </svg>
            </div>
        </div>
    </div>
</div>
{common_js}
</body>
</html>
"""

    # CHECKLIST HTML (~하였는가? 100%)
    chk_html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>동탄트램 통신분야 - {act_clean} 체크리스트</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
    <style>body {{ font-family: 'Noto Sans KR', sans-serif; }}</style>
</head>
<body class="bg-slate-50 text-slate-800 antialiased p-6 sm:p-10">
<div class="max-w-5xl mx-auto bg-white shadow-2xl rounded-2xl border border-slate-200 overflow-hidden">
    <div class="bg-slate-900 text-white p-8">
        <span class="bg-emerald-600 text-white text-xs font-bold px-3 py-1 rounded-full uppercase">Dongtan Tram Checklist</span>
        <h1 class="text-3xl font-black mt-2">{act_clean} 체크리스트</h1>
        <p class="text-emerald-200 text-sm mt-1">L4 Code: {l4_code} | 주관: {ju_gwan}</p>
    </div>
    
    <div class="p-8 space-y-6">
        <div class="bg-slate-100 p-4 rounded-xl border border-slate-300 flex justify-between items-center text-xs font-bold">
            <span>공종: 통신분야</span>
            <span>작업단위: {act_clean}</span>
            <span>산출물: {deliverable}</span>
        </div>

        <table class="w-full border-collapse border border-slate-300 text-sm text-left">
            <thead>
                <tr class="bg-slate-800 text-white text-xs">
                    <th class="border border-slate-300 p-3 text-center w-12">NO</th>
                    <th class="border border-slate-300 p-3 text-center">검측 및 점검 항목 statement (질문형 종결어미)</th>
                    <th class="border border-slate-300 p-3 text-center w-20">판정</th>
                </tr>
            </thead>
            <tbody>
"""
    for idx_c, item_text in enumerate(checklist_items, 1):
        chk_html += f"""                <tr class="hover:bg-slate-50">
                    <td class="border border-slate-300 p-3 text-center font-bold">{idx_c}</td>
                    <td class="border border-slate-300 p-3 font-medium text-slate-900">{item_text}</td>
                    <td class="border border-slate-300 p-3 text-center font-bold text-emerald-600">☐ 적합</td>
                </tr>\n"""
    chk_html += """            </tbody>
        </table>
    </div>
</div>
</body>
</html>
"""

    sub_map = {
        "표준서": std_html,
        "수행지침": gui_html,
        "체크리스트": chk_html
    }

    for sub_name, html_content in sub_map.items():
        sub_dir = os.path.join(target_folder, sub_name)
        if os.path.exists(sub_dir):
            for file_in_sub in os.listdir(sub_dir):
                if file_in_sub.endswith('.html'):
                    full_file_path = os.path.join(sub_dir, file_in_sub)
                    with open(full_file_path, 'w', encoding='utf-8') as out_f:
                        out_f.write(html_content)
                    print(f"   ✓ [OVERWROTE] {sub_name} -> {file_in_sub}")

    processed_count += 1

print(f"\n🎉 SUCCESSFULLY COMPLETED ALL TELECOM ACTIVITIES FROM ROW 4 TO ROW 33 ({processed_count} ACTIVITIES PROCESSSED)!")
