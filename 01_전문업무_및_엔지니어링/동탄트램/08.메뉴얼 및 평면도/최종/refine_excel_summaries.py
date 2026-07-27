import openpyxl
from bs4 import BeautifulSoup
import urllib.parse
import os
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
base_attach_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

wb = openpyxl.load_workbook(excel_path)

target_sheets = ['상부강화노반', '콘크리트도상', '건축', '신호분야', '통신분야', '전기분야']

def extract_rich_summary(html_path, doc_type):
    if not os.path.exists(html_path):
        return None
        
    with open(html_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')
        
    items = []
    
    for li in soup.find_all('li'):
        txt = li.get_text().strip()
        txt = re.sub(r'^\s*[\d\.\-☐☑📍📌]\s*', '', txt).strip()
        # Clean meta phrases
        txt = re.sub(r'\[동탄트램 업무 매뉴얼 v1 연계\]\s*', '', txt)
        txt = re.sub(r'\[동탄트램 매뉴얼 v1\]\s*', '', txt)
        txt = re.sub(r'\[설계사 작성\]\s*', '', txt)
        txt = txt.replace('기술기준 연계', '기술기준').replace('설계기준 연계', '설계기준')
        
        if txt and len(txt) > 5 and txt not in items:
            items.append(txt)
            
    if doc_type == '표준서':
        items = sorted(items, key=lambda x: 0 if any(k in x for k in ['KCS', 'KDS', '≥', '≤', '±', '%', 'MPa', 'mm', 'SIL']) else 1)
    elif doc_type == '수행지침':
        items = sorted(items, key=lambda x: 0 if any(k in x for k in ['단계', '준비', '시공', '확인', '검측', '수칙']) else 1)
    elif doc_type == '체크리스트':
        items = [x for x in items if any(k in x for k in ['인가', '검측', '확인', '승인', '되었는가', '했는가', '자문'])]
        if not items:
            for li in soup.find_all('li'):
                txt = li.get_text().strip()
                txt = re.sub(r'\[동탄트램 매뉴얼 v1\]\s*', '', txt)
                txt = re.sub(r'\[설계사 작성\]\s*', '', txt)
                if txt: items.append(txt)
                
    selected = items[:5]
    if not selected:
        return None
        
    bullets = [f"{i+1}) {item}" for i, item in enumerate(selected)]
    return "\n".join(bullets)

updated_count = 0

for sheet_name in target_sheets:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if not rows: continue
    
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    col_map = {h: i for i, h in enumerate(headers)}
    
    std_idx = next((i for h, i in col_map.items() if '표준서' in h), None)
    gui_idx = next((i for h, i in col_map.items() if '수행지침' in h), None)
    chk_idx = next((i for h, i in col_map.items() if '체크리스트' in h), None)
    act_idx = next((i for h, i in col_map.items() if '작업단위' in h or 'Activity' in h), None)
    
    if act_idx is None: continue
    
    sheet_act_count = 0
    disc_dir = sheet_name
    
    for row in rows[1:]:
        act_val = row[act_idx].value
        if not act_val: continue
        
        sheet_act_count += 1
        act_name = str(act_val).strip()
        sanitized_act = re.sub(r'[\/\\:\*\?"<>\|]', '_', act_name).strip()
        folder_name = f"{sheet_act_count}_{sanitized_act}"
        act_dir_abs = os.path.join(base_attach_dir, disc_dir, folder_name)
        
        tasks = [
            (std_idx, '표준서', '표준서'),
            (gui_idx, '수행지침', '수행지침'),
            (chk_idx, '체크리스트', '체크리스트')
        ]
        
        for c_idx, doc_type, doc_title in tasks:
            if c_idx is None or c_idx >= len(row): continue
            cell = row[c_idx]
            sub_folder = os.path.join(act_dir_abs, doc_type)
            
            if os.path.exists(sub_folder):
                files = [f for f in os.listdir(sub_folder) if f.endswith('.html')]
                if files:
                    html_path = os.path.join(sub_folder, files[0])
                    summary_text = extract_rich_summary(html_path, doc_type)
                    
                    if summary_text:
                        raw_rel_target = f"매뉴얼BODY(집행단계-첨부폴더)\\{disc_dir}\\{folder_name}\\{doc_type}\\{files[0]}"
                        btn_text = f"\n--------------------------------------\n👉 [더블클릭] 상세 {doc_title} 파일(HTML) 열기 📄"
                        cell.value = (summary_text + btn_text).strip()
                        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(ref=cell.coordinate, target=raw_rel_target)
                        updated_count += 1

print(f"Refinement complete! Total cells updated: {updated_count}")
wb.save(excel_path)
print(f"Saved Excel file to '{excel_path}'")
