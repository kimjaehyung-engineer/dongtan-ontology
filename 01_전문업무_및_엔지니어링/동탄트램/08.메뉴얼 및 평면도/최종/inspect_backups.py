import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종"

print("Checking available Excel files and their '지장물이설' sheet structures:")

for f in os.listdir(base_dir):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        f_path = os.path.join(base_dir, f)
        try:
            wb = openpyxl.load_workbook(f_path, data_only=True)
            if '지장물이설' in wb.sheetnames:
                ws = wb['지장물이설']
                row3_act = ws.cell(row=3, column=6).value or ws.cell(row=3, column=2).value
                print(f"\n📁 File: {f}")
                print(f"   - Max Row: {ws.max_row}, Max Col: {ws.max_column}")
                print(f"   - Row 3 Activity: {row3_act}")
                print(f"   - Row 3 Height: {ws.row_dimensions[3].height}")
        except Exception as e:
            print(f"📁 File: {f} (Error: {e})")
