import os
import re
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4.xlsx"
backup_excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v4_updated.xlsx"
base_dir = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼BODY(집행단계-첨부폴더)"

targets = [
    {
        "sheet": "사전토공사",
        "wbs": "9000-5-2",
        "folder": "2_발주전략 KOM",
        "row": 3
    },
    {
        "sheet": "상부강화노반",
        "wbs": "9000-7-2",
        "folder": "2_발주전략 KOM",
        "row": 3
    },
    {
        "sheet": "콘크리트도상",
        "wbs": "9000-6-4",
        "folder": "4_발주전략 KOM",
        "row": 5
    }
]

def extract_standard_summary_from_html(file_path):
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        html = f.read()
    # Find text inside <tr><th>표준서 (Standard) 요약</th><td>...</td></tr>
    match = re.search(r'표준서\s*\(Standard\)\s*요약</th>\s*<td>(.*?)</td>', html, re.DOTALL)
    if match:
        text = re.sub(r'<[^>]*>', '', match.group(1)).strip()
        text = re.sub(r'\s+', ' ', text)
        return text
    return None

def extract_checklist_summary_from_html(file_path):
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        html = f.read()
    # Find text inside <div class="summary-box"> ... </div>
    match = re.search(r'<div class="summary-box">.*?<div[^>]*>(.*?)</div>', html, re.DOTALL)
    if match:
        text = re.sub(r'<[^>]*>', '', match.group(1)).strip()
        text = re.sub(r'\s+', ' ', text)
        return text
    return None

def extract_guideline_summary_from_html(file_path):
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        html = f.read()
    
    # Extract guideline bullet points to make a clean summary
    # Or find if there's any summary structure
    # We can join first bullets of each step or take key texts
    bullets = re.findall(r'<li><strong>(.*?)</strong>(.*?)</li>', html)
    if bullets:
        summary_lines = []
        for title, desc in bullets[:3]: # Take first 3 key bullets
            clean_title = re.sub(r'<[^>]*>', '', title).replace(":", "").strip()
            clean_desc = re.sub(r'<[^>]*>', '', desc).strip()
            summary_lines.append(f"• {clean_title}: {clean_desc}")
        return "\n".join(summary_lines)
    return None

# Load Excel
wb = openpyxl.load_workbook(excel_path)

excel_updated = 0
for t in targets:
    sheet_name = t["sheet"]
    folder = t["folder"]
    row_num = t["row"]
    wbs = t["wbs"]
    
    # HTML File paths
    std_file = os.path.join(base_dir, sheet_name, folder, "표준서", f"발주전략 KOM_표준서.html" if sheet_name != "콘크리트도상" else f"4_발주전략 KOM_표준서.html")
    chk_file = os.path.join(base_dir, sheet_name, folder, "체크리스트", f"발주전략 KOM_체크리스트.html" if sheet_name != "콘크리트도상" else f"4_발주전략 KOM_체크리스트.html")
    gui_file = os.path.join(base_dir, sheet_name, folder, "수행지침", f"발주전략 KOM_수행지침.html" if sheet_name != "콘크리트도상" else f"4_발주전략 KOM_수행지침.html")
    
    # Extract values
    std_sum = extract_standard_summary_from_html(std_file)
    chk_sum = extract_checklist_summary_from_html(chk_file)
    gui_sum = extract_guideline_summary_from_html(gui_file)
    
    sheet = wb[sheet_name]
    
    # Verify WBS code
    excel_wbs = sheet.cell(row=row_num, column=4).value
    if excel_wbs != wbs:
        print(f"⚠️ WBS code mismatch on {sheet_name} Row {row_num}: Excel has '{excel_wbs}', expected '{wbs}'")
        # Find the correct row
        for r in range(1, sheet.max_row+1):
            if sheet.cell(row=r, column=4).value == wbs:
                row_num = r
                break
                
    if std_sum:
        sheet.cell(row=row_num, column=11).value = std_sum # Col K: 표준서 요약
        print(f"Updated {sheet_name} Row {row_num} Std Summary ➔ '{std_sum[:40]}...'")
    if gui_sum:
        sheet.cell(row=row_num, column=13).value = gui_sum # Col M: 수행지침 요약
        print(f"Updated {sheet_name} Row {row_num} Guideline Summary ➔ '{gui_sum[:40]}...'")
    if chk_sum:
        sheet.cell(row=row_num, column=15).value = chk_sum # Col O: 체크리스트 요약
        print(f"Updated {sheet_name} Row {row_num} Checklist Summary ➔ '{chk_sum[:40]}...'")
        
    excel_updated += 1

try:
    wb.save(excel_path)
    print(f"\n🎉 Successfully saved and aligned Excel from original HTMLs!")
except PermissionError:
    wb.save(backup_excel_path)
    print(f"\n⚠️ Excel locked. Saved to Backup: '{backup_excel_path}'")
