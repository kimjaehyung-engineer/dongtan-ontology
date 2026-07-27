import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

excel_path = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계).xlsx"
wb = openpyxl.load_workbook(excel_path)

for sheet_name in ['상부강화노반', '콘크리트도상', '신호분야']:
    ws = wb[sheet_name]
    header_row = 1
    print(f"\n=== Verification for Sheet '{sheet_name}' (Row 2 Sample) ===")
    
    headers = [str(ws.cell(row=header_row, column=c).value or "").replace('\n', ' ') for c in range(1, ws.max_column + 1)]
    
    std_sum_c = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '요약' in h), None)
    gui_sum_c = next((c for c, h in enumerate(headers, 1) if '수행지침' in h and '요약' in h), None)
    chk_sum_c = next((c for c, h in enumerate(headers, 1) if '체크리스트' in h and '요약' in h), None)
    std_link_c = next((c for c, h in enumerate(headers, 1) if '표준서' in h and '파일' in h), None)
    
    print("--- [표준서 요약] (Lines=" + str(len(str(ws.cell(row=2, column=std_sum_c).value).splitlines())) + ") ---")
    print(ws.cell(row=2, column=std_sum_c).value)
    
    print("--- [수행지침 요약] (Lines=" + str(len(str(ws.cell(row=2, column=gui_sum_c).value).splitlines())) + ") ---")
    print(ws.cell(row=2, column=gui_sum_c).value)
    
    print("--- [체크리스트 요약] (Lines=" + str(len(str(ws.cell(row=2, column=chk_sum_c).value).splitlines())) + ") ---")
    print(ws.cell(row=2, column=chk_sum_c).value)
    
    print(f"--- [표준서 파일 링크] HasLink={ws.cell(row=2, column=std_link_c).hyperlink is not None} ---")
