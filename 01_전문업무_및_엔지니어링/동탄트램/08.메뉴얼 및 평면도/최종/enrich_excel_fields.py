import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

target_v3 = r"c:\Users\sskjh\antigravity\01_전문업무_및_엔지니어링\동탄트램\08.메뉴얼 및 평면도\최종\매뉴얼 BODY (집행단계)v3.xlsx"

wb = openpyxl.load_workbook(target_v3)
ws = wb['지장물이설']

print("Enriching purpose, method, result text for all rows...")

for r in range(3, ws.max_row + 1):
    act_name = ws.cell(row=r, column=6).value or ''
    
    # Purpose
    if not ws.cell(row=r, column=8).value:
        ws.cell(row=r, column=8, value=f"{act_name} 관련 설계도서 대조 및 현장 실치 오차 사전 검증")
    # Method
    if not ws.cell(row=r, column=9).value:
        ws.cell(row=r, column=9, value=f"공사/공무/유관기관/전문업체 합동 현장 조사 및 안전 공법 적용")
    # Result
    if not ws.cell(row=r, column=10).value:
        ws.cell(row=r, column=10, value=f"{act_name} 검측보고서, 사진대지 및 관리기관 승인서")

wb.save(target_v3)
print("🎉 Purpose, Method, Result fields successfully enriched for all 39 rows!")
